"""ScoreMax V6.2.8.1 Power House V3 academic-review workbook compatibility tests."""
from __future__ import annotations

import io
import json
import os
import sys
import tempfile
from pathlib import Path

ROOT=Path(__file__).resolve().parent
from smoke_tests_v5_5 import install_framework_stubs

HEADERS=[
    'index','Review No','Priority','Review Requirement','Question ID','Family ID','Relationship',
    'Canonical Content ID','Mastery','Cognitive Demand','Question Type','Question / Task','Stimulus / Context',
    'Statements / Options','Key Answer','Explanation / Marking Rubric','Independent-AI Status',
    'AI Finding / Correction Rationale','Reviewer 1 Name','Reviewer 1 Decision','Scientific Accuracy',
    'Answer / Rubric Accuracy','Clarity / Ambiguity','Option / Distractor Quality','Mastery Decision',
    'Suggested Mastery','Exact Correction / Comments','Reviewer 1 Date','Reviewer 2 Required','Reviewer 2 Name',
    'Reviewer 2 Decision','Reviewer 2 Comments','Reviewer 2 Date','Adjudication Required',
    'Senior Adjudicator / Notes','Final Academic Decision','Agreed Final Wording / Version','Final Decision Date'
]


def row_values(sheet_no:int,index:int,p0_limit:int):
    qid=f'PH-CH1-B{sheet_no}-{index:04d}'
    priority='P0_PRESENTATION_REFERENCE' if index<=p0_limit else ('P1_CORRECTED_OR_DISPUTED' if index%17==0 else 'P2_GOVERNED_DUAL_REVIEW' if index%3==0 else 'P3_STANDARD_REVIEW')
    dual=index==1 or index%3==0
    requirement='DUAL_REVIEW_REQUIRED' if dual else 'STANDARD_FIRST_REVIEW'
    qtype='SHORT_CONSTRUCTED_RESPONSE' if index%25==0 else 'MATCHING_SET' if index%20==0 else 'STANDARD_MCQ'
    stimulus=f'Scenario for sheet {sheet_no}, row {index}.' if index%4==0 else ''
    if qtype=='STANDARD_MCQ':
        content='Statements:\n\n\nOptions:\nA. Correct option\nB. Distractor one\nC. Distractor two\nD. Distractor three'
        key='A'
        explanation='The configured key is A.'
    elif qtype=='MATCHING_SET':
        content='Left items:\nL1. Term one\nL2. Term two\n\nRight options:\nR1. Meaning one\nR2. Meaning two\nR3. Extra distractor'
        key='{"L1":"R1","L2":"R2"}'
        explanation='Each term is linked to its corresponding meaning.'
    else:
        content=''
        key=''
        explanation='Analytic Rubric: 1 mark for the correct concept and 1 mark for a justified comparison.'
    values=[
        index-1,index,priority,requirement,qid,f'FG-{sheet_no}-{index:04d}','NEW_SEED',qid,
        'EXAM_READY' if index%2 else 'ADVANCED','APPLICATION',qtype,
        f'Question task {sheet_no}-{index}',stimulus,content,key,explanation,'NO_ROW_LEVEL_AI_FINDING','',
        '','PENDING','PENDING','PENDING','PENDING','PENDING','KEEP','','','',
        'YES' if dual else 'CONDITIONAL_IF_CHANGED','','PENDING','','','NO','','PENDING','',''
    ]
    assert len(values)==len(HEADERS)
    return values


def build_v3_workbook() -> bytes:
    from openpyxl import Workbook
    wb=Workbook()
    start=wb.active
    start.title='Start Here'
    start.append(['Chapter 1 Academic Review — Post Independent-AI Assurance'])
    start.append(['Use the three batch sheets for review.'])
    mastery=wb.create_sheet('Mastery & Decisions'); mastery.append(['Mastery Classification and Academic Review Standards'])
    sources=wb.create_sheet('Sources & Limitations'); sources.append(['Source / Evidence','Role'])
    specs=[('Batch 1 Review',1199,65),('Batch 2 Review',1168,64),('Batch 3 Review',1199,64)]
    for sheet_no,(name,count,p0) in enumerate(specs,1):
        ws=wb.create_sheet(name)
        ws.append(HEADERS)
        for i in range(1,count+1):
            ws.append(row_values(sheet_no,i,p0))
    out=io.BytesIO(); wb.save(out); wb.close(); return out.getvalue()


def main():
    install_framework_stubs()
    temp=Path(tempfile.mkdtemp(prefix='scoremax_v6281_'))
    os.environ['SCOREMAX_DB']=str(temp/'scoremax.db')
    os.environ['SCOREMAX_ENV']='local'
    os.environ['SCOREMAX_SECRET']='v6.2.8.1-regression-secret'
    os.environ['SCOREMAX_BOOTSTRAP_ADMIN_PASSWORD']='V6281-Admin-Password'
    sys.path.insert(0,str(ROOT))
    import app
    from werkzeug.security import generate_password_hash
    ri=app.reviewer_import
    rw=app.reviewer_workspace
    checks=[]
    def ok(name,condition=True):
        if not condition: raise AssertionError(name)
        checks.append(name); print('PASS:',name)

    app.init(); c=app.db()
    raw=build_v3_workbook()
    rows=rw.parse_upload('PH_CH1_Final_Academic_Review_Workbook_Post_AI_Assurance_v3_0.xlsx',raw)
    counts=ri.sheet_counts(rows)
    ok('instructional and reference worksheets are ignored',len(rows)==3566 and [x['sheet'] for x in counts]==['Batch 1 Review','Batch 2 Review','Batch 3 Review'])
    ok('all three Power House review sheet counts are preserved',[x['rows'] for x in counts]==[1199,1168,1199])
    columns=ri.headers(rows); mapping=ri.suggest_mapping(columns)
    ok('Power House V3 headers map without manual configuration',mapping['question_text']=='Question / Task' and mapping['correct_answer']=='Key Answer' and mapping['explanation']=='Explanation / Marking Rubric' and mapping['options_text']=='Statements / Options')
    ok('Power House V3 workbook profile is recognised',ri.detect_profile(columns,rows)=='POWER_HOUSE_ACADEMIC_REVIEW_V3')
    normalized,errors=ri.validate_preview(rows,mapping,chapter='Chapter 1',topic='Biodiversity')
    ok('all 3566 records import including rubric-only constructed responses',len(normalized)==3566 and not errors)
    ok('combined option text becomes reviewer choices',len(normalized[0]['options'])==4 and normalized[0]['options'][0]['id']=='A')
    matching=next(x for x in normalized if x['question_type']=='MATCHING_SET')
    ok('matching and statement content is preserved for the reviewer','Left items:' in matching['review_content'] and not matching['options'])
    constructed=next(x for x in normalized if x['question_type']=='SHORT_CONSTRUCTED_RESPONSE')
    ok('marking rubric safely supplies a missing separate key','Analytic Rubric:' in constructed['correct_answer'])
    ok('stimulus/context is preserved inside the minimal review snapshot',any(x['stimulus_context'].startswith('Scenario') for x in normalized))

    staged=ri.preview_import(c,rows,title='Chapter 1 V3 Academic Review',filename='PH_CH1_Final_Academic_Review_Workbook_Post_AI_Assurance_v3_0.xlsx',chapter='Chapter 1',topic='Biodiversity',created_by=1)
    confirmed=ri.confirm_import(c,staged['id'],mapping,actor_user_id=1)
    ok('3566 questions create 36 batches without crossing source-sheet boundaries',confirmed['valid_rows']==3566 and confirmed['batch_count']==36 and confirmed['invalid_rows']==0)
    batch_rows=c.execute('SELECT source_sheet,source_part_number,source_part_count,question_count FROM reviewer_batches WHERE import_id=? ORDER BY batch_number',(staged['id'],)).fetchall()
    ok('each original review sheet creates twelve auditable parts',len(batch_rows)==36 and all(int(x['source_part_count'])==12 for x in batch_rows))
    ok('source sheet part sizes are 100 with the correct final remainder',[(x['source_sheet'],x['question_count']) for x in batch_rows if int(x['source_part_number'])==12]==[('Batch 1 Review',99),('Batch 2 Review',68),('Batch 3 Review',99)])
    crossing=c.execute('''SELECT rb.id,COUNT(DISTINCT rq.source_sheet) n FROM reviewer_batches rb JOIN reviewer_questions rq ON rq.batch_id=rb.id WHERE rb.import_id=? GROUP BY rb.id HAVING n>1''',(staged['id'],)).fetchall()
    ok('no ScoreMax batch mixes questions from different workbook sheets',not crossing)
    stored=c.execute("SELECT * FROM reviewer_questions WHERE question_type='STANDARD_MCQ' ORDER BY id LIMIT 1").fetchone()
    ok('reviewer snapshot stores source lineage and required content only',stored['source_sheet']=='Batch 1 Review' and stored['source_row']==2 and len(json.loads(stored['options_json']))==4)

    reviewer_id=c.execute("""INSERT INTO users(system_user_id,role,full_name,email,username,password_hash,account_status)
      VALUES('REV-628100','reviewer','V3 Reviewer','v3reviewer@example.com','v3reviewer',?,'active')""",(generate_password_hash('Reviewer-Password'),)).lastrowid
    c.commit()
    first_batch=batch_rows[0]
    batch_id=c.execute('SELECT id FROM reviewer_batches WHERE import_id=? ORDER BY batch_number LIMIT 1',(staged['id'],)).fetchone()['id']
    assignment=rw.create_assignment(c,batch_id=batch_id,reviewer_user_id=reviewer_id,created_by=1)
    item=c.execute('SELECT id FROM reviewer_assignment_items WHERE assignment_id=? ORDER BY display_order LIMIT 1',(assignment['assignment_id'],)).fetchone()
    result=rw.submit_decision(c,item_id=item['id'],reviewer_user_id=reviewer_id,decision='ACCEPT_UNCHANGED',mastery_suitability='SUITABLE')
    ok('pre-governed dual-review items still require a second reviewer after unchanged acceptance',result['outcome_status']=='SECOND_REVIEW_REQUIRED')
    template=(ROOT/'templates/reviewer_item.html').read_text()
    ok('reviewer page displays context and structured response content without wider architecture','item.stimulus_context' in template and 'item.review_content' in template and 'CONFIGURED ANSWER / RUBRIC' in template)
    admin_template=(ROOT/'templates/admin_reviewer_workspace.html').read_text()
    ok('Admin preview reports recognised Power House sheets and counts','POWER_HOUSE_ACADEMIC_REVIEW_V3' in admin_template and 'pending_import.sheet_counts' in admin_template)
    importer_source=(ROOT/'reviewer_import_engine.py').read_text()
    ok('misleading map-question-and-answer failure has been removed','Map the Question and Correct answer columns first' not in importer_source)
    ok('release health marker is V6.2.8.1',app.healthz()[0]['version']=='6.2.8.1')
    c.close()
    print(f'\nV6.2.8.1 POWER HOUSE V3 IMPORT CHECKS PASSED: {len(checks)}')


if __name__=='__main__':
    main()
