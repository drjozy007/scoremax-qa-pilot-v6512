from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, jsonify, abort
import sqlite3, random, csv, io, os, json, secrets, time, re, hashlib, smtplib, ssl, logging, shutil
from email.message import EmailMessage
from pathlib import Path
from datetime import datetime, timedelta
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook, load_workbook
from written_response_engine import validate_assessment_package, package_checksum, package_signature, mark_written_response
from academic_messaging_engine import (clean_text, parse_list, profile_completeness, detect_message_policy,
    validate_teacher_listing, teacher_match_score)
from PIL import Image, ImageStat, ImageFilter
from blueprint_engine import (calculate_checksum, calculate_signature, canonical_json, compare_blueprints,
    confidence_label, normalize_blueprint_payload, normalize_difficulty, normalize_mix, rigor_mix,
    allocate_counts, validate_blueprint_payload)
from pilot_readiness_engine import (validate_prompt_pack, parse_generation_output, generation_transport,
    payload_checksum as pilot_payload_checksum, safe_filename, sqlite_backup, feedback_route, readiness_status)
from student_experience_engine import (MATRIC_COMMON_SUBJECTS, PATHWAY_CATALOGUE, is_matric_level,
    pathway as pathway_definition, workload_range, workload_fit, minutes_per_study_day)
from daily_spark_engine import WORD_LIBRARY, age_from_dob, choose_word, word_payload, academic_payload
import mastery_lab_engine as mastery_lab
import reviewer_workspace_engine as reviewer_workspace
import reviewer_import_engine as reviewer_import
import commercial_access_engine as commercial_access
import universal_mastery_engine as universal_mastery
import scoremax_integration_v1 as integration_v1
import qa_synthetic_learner as qa_synthetic

BASE=Path(__file__).parent
DB=Path(os.environ.get('SCOREMAX_DB', BASE/'scoremax_v4.db'))
SCOREMAX_ENV=os.environ.get('SCOREMAX_ENV','local').strip().lower() or 'local'
app=Flask(__name__)
_secret=os.environ.get('SCOREMAX_SECRET','').strip()
if not _secret:
    if SCOREMAX_ENV=='production':
        raise RuntimeError('SCOREMAX_SECRET must be set in production.')
    _secret=secrets.token_hex(32)
    print('[ScoreMax V6.5.10] SCOREMAX_SECRET not set; using an ephemeral local-session secret for this run.')
app.secret_key=_secret
app.config['MAX_CONTENT_LENGTH']=int(os.environ.get('SCOREMAX_MAX_UPLOAD_MB','40'))*1024*1024
_cookie_secure=os.environ.get('SCOREMAX_COOKIE_SECURE','0')=='1' or SCOREMAX_ENV=='production'
app.config.update(SESSION_COOKIE_HTTPONLY=True,SESSION_COOKIE_SAMESITE='Lax',SESSION_COOKIE_SECURE=_cookie_secure)
logging.basicConfig(level=getattr(logging, os.environ.get('SCOREMAX_LOG_LEVEL','INFO').upper(), logging.INFO))
if SCOREMAX_ENV=='production':
    _smtp_host=os.environ.get('SCOREMAX_SMTP_HOST','').strip()
    _smtp_from=os.environ.get('SCOREMAX_SMTP_FROM','').strip()
    if not _smtp_host or not _smtp_from:
        raise RuntimeError('SCOREMAX_SMTP_HOST and SCOREMAX_SMTP_FROM must be set in production so password reset can function.')
COMMERCIAL_GATES_ENABLED=os.environ.get('SCOREMAX_ENFORCE_PAYWALL','0').strip()=='1'
INTERNAL_FULL_ACCESS_ENABLED=os.environ.get('SCOREMAX_INTERNAL_FULL_ACCESS','0').strip()=='1'
_RATE_BUCKETS={}
LIVE_MARKABLE_TYPES={'single_choice','true_false','fill_blank','multiple_select','numerical'}
SCOREMAX_RELEASE_VERSION='6.5.12'
SCOREMAX_BUILD_NAME='ScoreMax Synthetic Learner Isolation Rectification Candidate'
STUDENT_PROGRAMME_CHOICES=[
    {'code':'fsc1','label':'FSc 1','value':'FSc Part 1'},
    {'code':'fsc2','label':'FSc 2','value':'FSc Part 2'},
    {'code':'mdcat','label':'MDCAT','value':'MDCAT'},
]
STUDENT_PROGRAMME_BY_CODE={x['code']:x for x in STUDENT_PROGRAMME_CHOICES}


def db():
    c=sqlite3.connect(DB, timeout=5)
    c.row_factory=sqlite3.Row
    c.execute('PRAGMA foreign_keys=ON')
    c.execute('PRAGMA journal_mode=WAL')
    c.execute('PRAGMA busy_timeout=5000')
    return c

def _csrf_token():
    token=session.get('_csrf_token')
    if not token:
        token=secrets.token_urlsafe(24)
        session['_csrf_token']=token
    return token

app.jinja_env.globals['csrf_token']=_csrf_token
app.jinja_env.globals['parse_list']=parse_list
# Academic review is now owned by Power House; historical ScoreMax reviewer code is retained only for rollback/audit compatibility.
app.jinja_env.globals['reviewer_workspace_forward_enabled']=False

def _session_version(value, missing=-1):
    """Return a valid integer session version without treating zero as missing."""
    if value is None:
        return missing
    try:
        return int(value)
    except (TypeError, ValueError):
        return missing

@app.before_request
def _v54_security_gate():
    # Machine-to-machine integration endpoints use scoped bearer + HMAC authentication,
    # not browser sessions/CSRF. Every integration route performs its own service-auth gate.
    if request.path.startswith('/api/integration/v1/'):
        return None
    _csrf_token()
    if request.method in {'POST','PUT','PATCH','DELETE'}:
        expected=session.get('_csrf_token','')
        supplied=request.form.get('_csrf_token','') or request.headers.get('X-CSRF-Token','')
        if not expected or not supplied or not secrets.compare_digest(str(expected),str(supplied)):
            abort(400, description='Invalid or missing form security token.')
    # Structural fence: a forgotten route-level check cannot accidentally make a private page anonymous.
    public_endpoints={'index','login','register','forgot_password','reset_password','about_page','how_it_works','faq_page','updates_page','contact_page','healthz','static','knowledge_home','knowledge_article','connect_page','sustainability_page','teacher_of_year_page','reviewer_invite'}
    if session.get('user_id'):
        c=db(); u=c.execute("SELECT account_status,COALESCE(session_version,0) session_version FROM users WHERE id=?",(session.get('user_id'),)).fetchone(); c.close()
        expected_version=_session_version(u['session_version'],0) if u else -1
        presented_version=_session_version(session.get('session_version'),-1)
        if not u or (u['account_status'] and u['account_status']!='active') or presented_version!=expected_version:
            session.clear()
            if request.path.startswith('/api/'):
                return jsonify({'error':'session_invalid'}),401
            flash('Please log in again to continue.','info')
            return redirect(url_for('login'))
    if session.get('role')=='reviewer':
        reviewer_endpoints={'reviewer_home','reviewer_continue','reviewer_assignment','reviewer_item','reviewer_timer','reviewer_reveal','reviewer_submit','logout','static'}
        if request.endpoint and request.endpoint not in reviewer_endpoints:
            return redirect(url_for('reviewer_home'))
    if request.endpoint and request.endpoint not in public_endpoints and not session.get('user_id'):
        if request.path.startswith('/api/'):
            return jsonify({'error':'authentication_required'}),401
        return redirect(url_for('login'))


@app.before_request
def _qa_synthetic_learner_route_fence():
    """P0 isolation fence for authenticated synthetic learner identities.

    qa_student accounts authenticate through the normal ScoreMax login, but can never
    traverse live learner/product routes. This prevents accidental writes to attempts,
    mastery, referrals, plans, teacher/parent evidence or Growth events.
    """
    if session.get('role')!='qa_student':
        return None
    allowed={
        'qa_synthetic_home','qa_synthetic_start','qa_synthetic_session','qa_synthetic_result',
        'qa_synthetic_visual_capture','qa_synthetic_visual_judgement','logout','static'
    }
    if request.endpoint in allowed:
        return None
    return redirect(url_for('qa_synthetic_home'))


_INTEGRATION_LAST_TICK=0.0

@app.before_request
def _integration_housekeeping_tick():
    # Synthetic learner traffic is qualification-only and must not mutate unrelated
    # integration/product state as an incidental side effect of a QA request.
    if session.get('role')=='qa_student':
        return None
    # Cheap bounded housekeeping only; no cross-system call is allowed to block a learner request.
    global _INTEGRATION_LAST_TICK
    now=time.time()
    if now-_INTEGRATION_LAST_TICK<30:
        return None
    _INTEGRATION_LAST_TICK=now
    try:
        c=db()
        if c.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name='integration_outbox'").fetchone():
            integration_v1.activate_due_releases(c)
            c.commit()
        c.close()
    except Exception:
        # Integration housekeeping must fail isolated; learner/product transactions continue.
        app.logger.exception('Non-blocking integration housekeeping failed')
    return None

def rate_limit(key, limit=8, window_seconds=300):
    now=time.time()
    bucket=[t for t in _RATE_BUCKETS.get(key,[]) if now-t<window_seconds]
    if len(bucket)>=limit:
        _RATE_BUCKETS[key]=bucket
        return False
    bucket.append(now); _RATE_BUCKETS[key]=bucket
    return True


def mask_email(value):
    email=str(value or '').strip()
    if '@' not in email: return 'registered address'
    local,domain=email.split('@',1)
    visible=(local[:1] or '*')+'***'
    return f'{visible}@{domain}'

def init():
    c=db()
    c.executescript('''
    CREATE TABLE IF NOT EXISTS institutions(
      id INTEGER PRIMARY KEY, institution_code TEXT UNIQUE, name TEXT NOT NULL,
      province TEXT, division TEXT, district TEXT, board TEXT, institution_type TEXT DEFAULT 'College', active INTEGER DEFAULT 1);
    CREATE TABLE IF NOT EXISTS users(
      id INTEGER PRIMARY KEY, system_user_id TEXT UNIQUE, role TEXT, full_name TEXT, father_name TEXT,
      dob TEXT, mobile TEXT, email TEXT, username TEXT UNIQUE, password_hash TEXT,
      province TEXT, division TEXT, district TEXT, board TEXT, academic_level TEXT,
      primary_institution_id INTEGER, subjects TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS classrooms(
      id INTEGER PRIMARY KEY, teacher_id INTEGER, institution_id INTEGER, name TEXT, level TEXT, subject TEXT, join_code TEXT UNIQUE);
    CREATE TABLE IF NOT EXISTS classroom_students(
      classroom_id INTEGER, student_id INTEGER, roll_no TEXT, PRIMARY KEY(classroom_id,student_id));
    CREATE TABLE IF NOT EXISTS curriculum(
      id INTEGER PRIMARY KEY, programme TEXT, subject TEXT, chapter TEXT, topic TEXT, subtopic TEXT, active INTEGER DEFAULT 1);
    CREATE TABLE IF NOT EXISTS questions(
      id INTEGER PRIMARY KEY, question_id TEXT UNIQUE, family_id TEXT, variant TEXT, programme TEXT,
      subject TEXT, chapter TEXT, topic TEXT, subtopic TEXT, qtype TEXT, level TEXT, question TEXT,
      option_a TEXT, option_b TEXT, option_c TEXT, option_d TEXT, answer TEXT, explanation TEXT, status TEXT);
    CREATE TABLE IF NOT EXISTS assessment_sessions(
      id INTEGER PRIMARY KEY, student_id INTEGER, mode TEXT DEFAULT 'practice', duration_minutes INTEGER,
      started_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, expires_at TEXT, status TEXT DEFAULT 'in_progress',
      current_index INTEGER DEFAULT 0, question_ids TEXT NOT NULL, flagged_ids TEXT DEFAULT '',
      saved_answers TEXT DEFAULT '{}', meta_json TEXT DEFAULT '{}');
    CREATE TABLE IF NOT EXISTS attempts(
      id INTEGER PRIMARY KEY, student_id INTEGER, scope TEXT, programme TEXT, subject TEXT, chapters TEXT,
      topic TEXT, subtopic TEXT, level TEXT, score REAL, correct_count INTEGER, total_count INTEGER,
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS attempt_answers(
      id INTEGER PRIMARY KEY, attempt_id INTEGER, question_db_id INTEGER, selected_answer TEXT, is_correct INTEGER);
    ''')
    identity_indexes=(
        ('email','idx_users_unique_email'),
        ('username','idx_users_unique_username_ci'),
        ('system_user_id','idx_users_unique_system_user_id_ci'),
    )
    for column,index_name in identity_indexes:
        duplicate=c.execute(f"SELECT 1 FROM users WHERE COALESCE({column},'')<>'' GROUP BY lower({column}) HAVING COUNT(*)>1 LIMIT 1").fetchone()
        if not duplicate:
            c.execute(f"CREATE UNIQUE INDEX IF NOT EXISTS {index_name} ON users(lower({column})) WHERE {column} IS NOT NULL AND {column}<>''")
        else:
            print(f'[ScoreMax V6.5.10] Existing duplicate {column} values detected; resolve them before enforcing its case-insensitive unique index.')

    if not c.execute("SELECT 1 FROM users WHERE username='admin'").fetchone():
        bootstrap=os.environ.get('SCOREMAX_BOOTSTRAP_ADMIN_PASSWORD','').strip() or secrets.token_urlsafe(10)
        c.execute("INSERT INTO users(system_user_id,role,full_name,username,password_hash) VALUES('ADM-000001','admin','Platform Admin','admin',?)",(generate_password_hash(bootstrap),))
        print(f'[ScoreMax V6.5.10] One-time bootstrap admin created: admin / {bootstrap}')
    else:
        existing_admin=c.execute("SELECT id,password_hash FROM users WHERE username='admin' AND role='admin'").fetchone()
        if existing_admin and existing_admin['password_hash'] and check_password_hash(existing_admin['password_hash'],'admin123'):
            bootstrap=os.environ.get('SCOREMAX_BOOTSTRAP_ADMIN_PASSWORD','').strip() or secrets.token_urlsafe(10)
            c.execute("UPDATE users SET password_hash=? WHERE id=?",(generate_password_hash(bootstrap),existing_admin['id']))
            print(f'[ScoreMax V6.5.10] Legacy demo admin password rotated. New one-time local password: admin / {bootstrap}')
    if c.execute('SELECT COUNT(*) n FROM questions').fetchone()['n']==0:
        seed=[
        ('BIO001','OSM001','A','FSc Part 1','Biology','Cell Biology','Cell Transport','Osmosis','MCQ','Foundation','Osmosis is the movement of:','Water through a selectively permeable membrane','Solute using ATP','Proteins through ribosomes','Gases against a gradient','A','Osmosis is passive movement of water.'),
        ('BIO002','OSM001','B','FSc Part 1','Biology','Cell Biology','Cell Transport','Osmosis','True/False','Foundation','Osmosis requires ATP.','True','False','','','B','Osmosis is passive.'),
        ('BIO003','OSM002','A','FSc Part 1','Biology','Cell Biology','Cell Transport','Osmosis','MCQ','Exam Ready','A plant cell in concentrated salt solution will:','Gain water','Lose water','Remain turgid','Divide','B','Water leaves the cell by osmosis.'),
        ('BIO004','OSM003','A','FSc Part 1','Biology','Cell Biology','Cell Transport','Osmosis','MCQ','Distinction','A cell at -500 kPa is placed in solution at -200 kPa. Water moves:','Into the cell','Out of the cell','No movement','Only actively','A','Water moves from higher to lower water potential.'),
        ('BIO005','ACT001','A','FSc Part 1','Biology','Cell Biology','Cell Transport','Active Transport','MCQ','Elite','A membrane pump moving ions against a gradient directly requires:','ATP hydrolysis','Osmosis','Simple diffusion','Passive transport','A','Active transport requires energy.')]
        c.executemany('''INSERT INTO questions(question_id,family_id,variant,programme,subject,chapter,topic,subtopic,qtype,level,question,option_a,option_b,option_c,option_d,answer,explanation,status)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'Approved')''',seed)
        c.execute("INSERT INTO curriculum(programme,subject,chapter,topic,subtopic) SELECT DISTINCT programme,subject,chapter,topic,subtopic FROM questions")
    migrate_v5(c)
    migrate_v6(c)
    migrate_v6_1(c)
    migrate_v6_2(c)
    migrate_v6_2_3(c)
    migrate_v6_2_5(c)
    mastery_lab.init_mastery_lab_schema(c)
    qa_synthetic.init_schema(c)
    reviewer_workspace.init_reviewer_schema(c)
    reviewer_import.init_schema(c)
    commercial_access.init_schema(c)
    universal_mastery.init_schema(c)
    init_chapter_catalogue_schema(c)
    backfill_chapter_catalogue(c)
    migrate_v6_4(c)
    integration_v1.init_schema(c)
    integration_v1.activate_due_releases(c)
    # V6.3 universal mastery is feature-flagged. Legacy mastery remains authoritative
    # until the internal pilot acceptance gate is deliberately opened.
    if os.environ.get('SCOREMAX_UNIVERSAL_MASTERY','0').strip()=='1':
        universal_mastery.set_feature_flag(c,'universal_mastery_runtime',True,mode='PILOT')
    c.commit(); c.close()



def table_columns(c, table):
    return {r['name'] for r in c.execute(f"PRAGMA table_info({table})").fetchall()}

def ensure_column(c, table, name, definition):
    if name not in table_columns(c, table):
        c.execute(f"ALTER TABLE {table} ADD COLUMN {name} {definition}")


def migrate_v6_4(c):
    """V6.4.0 live-pilot UX/operations migration. Idempotent and additive."""
    ensure_column(c,'users','active_programme',"TEXT DEFAULT ''")
    c.execute("UPDATE users SET active_programme=academic_level WHERE role='student' AND COALESCE(active_programme,'')='' AND COALESCE(academic_level,'')<>''")

    # Referral attribution is immutable after a valid owner is recorded. One teacher override level only.
    ensure_column(c,'referral_attributions','attribution_kind',"TEXT DEFAULT 'DIRECT'")
    ensure_column(c,'referral_attributions','locked_at',"TEXT DEFAULT ''")
    ensure_column(c,'referral_rewards','rule_version','INTEGER DEFAULT 1')
    ensure_column(c,'referral_rewards','override_referrer_user_id','INTEGER')
    ensure_column(c,'referral_rewards','override_reward_rate','REAL DEFAULT 0')
    ensure_column(c,'referral_rewards','override_reward_amount_minor','INTEGER DEFAULT 0')
    ensure_column(c,'referral_rewards','override_rule_version','INTEGER DEFAULT 0')
    ensure_column(c,'referral_rewards','override_status',"TEXT DEFAULT ''")
    ensure_column(c,'referral_rewards','override_available_at',"TEXT DEFAULT ''")
    ensure_column(c,'referral_rewards','override_paid_at',"TEXT DEFAULT ''")
    ensure_column(c,'referral_rewards','override_reversed_at',"TEXT DEFAULT ''")
    ensure_column(c,'referral_rewards','override_notes',"TEXT DEFAULT ''")
    ensure_column(c,'referral_programs','programme_version','INTEGER DEFAULT 1')
    c.execute("INSERT OR IGNORE INTO referral_programs(role_group,reward_type,reward_rate,hold_days,active,programme_version) VALUES('teacher_direct','commission',0.00,14,1,1)")
    c.execute("INSERT OR IGNORE INTO referral_programs(role_group,reward_type,reward_rate,hold_days,active,programme_version) VALUES('teacher_override','commission',0.00,14,1,1)")

    # Emergency direct intake is a governed wrapper around the existing importer, not a second importer.
    ensure_column(c,'content_import_batches','intake_mode',"TEXT DEFAULT 'STANDARD'")
    ensure_column(c,'content_import_batches','release_status',"TEXT DEFAULT 'NOT_RELEASED'")
    ensure_column(c,'content_import_batches','release_attested_at',"TEXT DEFAULT ''")
    ensure_column(c,'content_import_batches','release_attested_by','INTEGER')
    ensure_column(c,'content_import_batches','release_note',"TEXT DEFAULT ''")
    ensure_column(c,'content_import_batches','released_count','INTEGER DEFAULT 0')
    ensure_column(c,'content_import_batches','released_at',"TEXT DEFAULT ''")


# V6.3.2 — governed chapter identity for learner-facing chapter names.
# The raw questions.chapter value remains the source/filter key. Display metadata is
# deliberately separate so improving a title never changes question identity or mastery scope.
_CHAPTER_PREFIX_RE=re.compile(r'^\s*(?:chapter|ch\.?)[\s_-]*([0-9]+[A-Za-z]?)\s*(?:[\-–—:.]\s*)?(.*?)\s*$',re.I)
_CHAPTER_NUMERIC_TITLE_RE=re.compile(r'^\s*([0-9]+[A-Za-z]?)\s*[.\-–—:]\s*(.+?)\s*$')
_CHAPTER_NUMBER_ONLY_RE=re.compile(r'^\s*([0-9]+[A-Za-z]?)\s*$')

def parse_chapter_identity(raw_chapter):
    """Extract only chapter identity explicitly present in the source label.

    No chapter name or number is guessed from curriculum knowledge. A title-only
    source such as ``Cell Biology`` stays title-only; ``Chapter 3`` stays
    number-only until governed metadata supplies the missing name.
    """
    raw=str(raw_chapter or '').strip()
    if not raw:
        return {'chapter_number':'','chapter_name':'','display_label':'','identity_status':'MISSING'}
    m=_CHAPTER_PREFIX_RE.match(raw)
    if m:
        number=(m.group(1) or '').strip(); name=(m.group(2) or '').strip(' \t-–—:.')
        display=f'Chapter {number} — {name}' if name else f'Chapter {number}'
        return {'chapter_number':number,'chapter_name':name,'display_label':display,
                'identity_status':'COMPLETE_SOURCE_LABEL' if name else 'NUMBER_ONLY_SOURCE_LABEL'}
    m=_CHAPTER_NUMERIC_TITLE_RE.match(raw)
    if m:
        number=(m.group(1) or '').strip(); name=(m.group(2) or '').strip()
        return {'chapter_number':number,'chapter_name':name,'display_label':f'Chapter {number} — {name}',
                'identity_status':'COMPLETE_SOURCE_LABEL'}
    m=_CHAPTER_NUMBER_ONLY_RE.match(raw)
    if m:
        number=m.group(1).strip()
        return {'chapter_number':number,'chapter_name':'','display_label':f'Chapter {number}',
                'identity_status':'NUMBER_ONLY_SOURCE_LABEL'}
    return {'chapter_number':'','chapter_name':raw,'display_label':raw,'identity_status':'TITLE_ONLY_SOURCE_LABEL'}


def _chapter_display_label(chapter_number='', chapter_name='', raw_chapter=''):
    number=str(chapter_number or '').strip(); name=str(chapter_name or '').strip(); raw=str(raw_chapter or '').strip()
    if number and name: return f'Chapter {number} — {name}'
    if name: return name
    if number: return f'Chapter {number}'
    return raw


def init_chapter_catalogue_schema(c):
    c.executescript("""
    CREATE TABLE IF NOT EXISTS chapter_catalogue(
      id INTEGER PRIMARY KEY,
      programme TEXT NOT NULL DEFAULT '',
      subject TEXT NOT NULL DEFAULT '',
      source_chapter TEXT NOT NULL,
      chapter_number TEXT NOT NULL DEFAULT '',
      chapter_name TEXT NOT NULL DEFAULT '',
      display_label TEXT NOT NULL DEFAULT '',
      identity_status TEXT NOT NULL DEFAULT 'UNRESOLVED',
      metadata_source TEXT NOT NULL DEFAULT 'SOURCE_LABEL',
      review_status TEXT NOT NULL DEFAULT 'Derived',
      active INTEGER NOT NULL DEFAULT 1,
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      UNIQUE(programme,subject,source_chapter));
    CREATE INDEX IF NOT EXISTS idx_chapter_catalogue_scope
      ON chapter_catalogue(programme,subject,active,source_chapter);
    """)


def upsert_chapter_catalogue(c, programme, subject, source_chapter, chapter_number='', chapter_name='',
                             metadata_source='SOURCE_LABEL', review_status='Derived'):
    """Persist chapter presentation metadata without changing the raw chapter key.

    Explicit governed/import metadata wins over deterministic parsing. Derived
    backfill never overwrites an existing governed row.
    """
    programme=str(programme or '').strip(); subject=str(subject or '').strip(); source=str(source_chapter or '').strip()
    if not source:
        return None
    parsed=parse_chapter_identity(source)
    number=str(chapter_number or '').strip() or parsed['chapter_number']
    name=str(chapter_name or '').strip() or parsed['chapter_name']
    display=_chapter_display_label(number,name,source)
    explicit=bool(str(chapter_number or '').strip() or str(chapter_name or '').strip())
    source_kind=str(metadata_source or ('GOVERNED_IMPORT' if explicit else 'SOURCE_LABEL')).strip()
    status='COMPLETE_GOVERNED_METADATA' if explicit and number and name else parsed['identity_status']
    existing=c.execute("""SELECT * FROM chapter_catalogue WHERE programme=? AND subject=? AND source_chapter=?""",
                       (programme,subject,source)).fetchone()
    if existing:
        governed=(existing['metadata_source'] or '').upper() in {'POWER_HOUSE','GOVERNED_IMPORT','ADMIN_GOVERNED'}
        if governed and not explicit:
            return existing['id']
        c.execute("""UPDATE chapter_catalogue SET chapter_number=?,chapter_name=?,display_label=?,identity_status=?,
          metadata_source=?,review_status=?,active=1,updated_at=CURRENT_TIMESTAMP WHERE id=?""",
          (number,name,display,status,source_kind,review_status,existing['id']))
        return existing['id']
    cur=c.execute("""INSERT INTO chapter_catalogue(programme,subject,source_chapter,chapter_number,chapter_name,
      display_label,identity_status,metadata_source,review_status,active) VALUES(?,?,?,?,?,?,?,?,?,1)""",
      (programme,subject,source,number,name,display,status,source_kind,review_status))
    return cur.lastrowid


def backfill_chapter_catalogue(c):
    rows=c.execute("""SELECT DISTINCT COALESCE(programme,'') programme,COALESCE(subject,'') subject,chapter
      FROM questions WHERE COALESCE(chapter,'')<>''""").fetchall()
    for row in rows:
        upsert_chapter_catalogue(c,row['programme'],row['subject'],row['chapter'])
    rows=c.execute("""SELECT DISTINCT COALESCE(programme,'') programme,COALESCE(subject,'') subject,chapter
      FROM curriculum WHERE COALESCE(chapter,'')<>''""").fetchall()
    for row in rows:
        upsert_chapter_catalogue(c,row['programme'],row['subject'],row['chapter'])


def chapter_identity(c, programme, subject, source_chapter):
    """Return the canonical learner-facing label while preserving raw source identity."""
    programme=str(programme or '').strip(); subject=str(subject or '').strip(); source=str(source_chapter or '').strip()
    row=c.execute("""SELECT * FROM chapter_catalogue WHERE active=1 AND programme=? AND lower(subject)=lower(?)
      AND source_chapter=? ORDER BY CASE WHEN metadata_source IN ('POWER_HOUSE','GOVERNED_IMPORT','ADMIN_GOVERNED') THEN 0 ELSE 1 END,id DESC LIMIT 1""",
      (programme,subject,source)).fetchone()
    if not row:
        row=c.execute("""SELECT * FROM chapter_catalogue WHERE active=1 AND lower(subject)=lower(?) AND source_chapter=?
          ORDER BY CASE WHEN programme=? THEN 0 WHEN programme='' THEN 1 ELSE 2 END,
                   CASE WHEN metadata_source IN ('POWER_HOUSE','GOVERNED_IMPORT','ADMIN_GOVERNED') THEN 0 ELSE 1 END,id DESC LIMIT 1""",
          (subject,source,programme)).fetchone()
    if row:
        return {'chapter':source,'chapter_number':row['chapter_number'] or '',
                'chapter_name':row['chapter_name'] or '',
                'display_label':row['display_label'] or source,
                'identity_status':row['identity_status'] or 'UNRESOLVED',
                'metadata_source':row['metadata_source'] or 'SOURCE_LABEL'}
    parsed=parse_chapter_identity(source)
    return {'chapter':source,**parsed,'metadata_source':'SOURCE_LABEL_RUNTIME'}

def canonical_family_key(family_id, country='', qualification='', exam_board='', curriculum_version='', programme='', subject=''):
    """Create a stable cross-market family key without assuming family_id is globally unique."""
    parts=[country,qualification,exam_board,curriculum_version,programme,subject,family_id]
    payload='|'.join(str(x or '').strip().lower() for x in parts)
    return 'FAM-'+hashlib.sha256(payload.encode('utf-8')).hexdigest()[:24]

def live_question_clause(alias='q'):
    """Single authoritative publication gate for every student-facing question pool.

    A question is live only when the question itself is Approved + active AND its
    governed question family is Approved + active. Imported Draft rows therefore
    cannot leak into Practice, tests, mastery, challenges, teacher assignments or
    Exam Centre builders merely because a spreadsheet contained Status=Approved.
    """
    p=f"{alias}." if alias else ''
    demo_fence='' if SCOREMAX_ENV!='production' else f" AND COALESCE({p}is_demo,0)=0"
    return (f"COALESCE({p}status,'Draft')='Approved' AND "
            f"COALESCE({p}review_status,'Draft')='Approved' AND COALESCE({p}active,0)=1 AND "
            f"COALESCE({p}family_key,'')<>'' AND EXISTS(SELECT 1 FROM question_families qf "
            f"WHERE qf.family_key={p}family_key AND qf.review_status='Approved' AND COALESCE(qf.active,0)=1){demo_fence}")

def effective_mastery_status(record, today=None):
    """Resolve current mastery validity from dates, not only the stored status label."""
    if not record:
        return ''
    status=record['status'] if hasattr(record,'keys') else record.get('status','')
    due_raw=record['verification_due_at'] if hasattr(record,'keys') else record.get('verification_due_at','')
    due=_parse_date(due_raw)
    today=today or datetime.now().date()
    if status in ('Verified','Elite Candidate') and due and due<today:
        return 'Verification Due'
    return status or ''

def record_mastery_history(c, record, event_type, new_level=None, new_status=None, attempt_id=None,
                           assessment_session_id=None, note='', metadata=None):
    """Append-only mastery governance history. `record` is the state before the event."""
    if not record:
        return
    c.execute("""INSERT INTO mastery_history(
      mastery_record_id,student_id,scope_type,scope_key,event_type,previous_level,new_level,
      previous_status,new_status,attempt_id,assessment_session_id,note,metadata_json)
      VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",
      (record['id'],record['student_id'],record['scope_type'],record['scope_key'],event_type,
       record['mastery_level'] or '',new_level if new_level is not None else (record['mastery_level'] or ''),
       record['status'] or '',new_status if new_status is not None else (record['status'] or ''),
       attempt_id,assessment_session_id,note,json.dumps(metadata or {})))

def upsert_question_family(c, row, review_status='Draft', active=0):
    """Create/update family governance from a question-like mapping and return family_key."""
    family_id=str(row.get('family_id') or row.get('Family ID') or '').strip()
    if not family_id:
        return ''
    country=str(row.get('country') or row.get('Country') or 'Pakistan').strip() or 'Pakistan'
    qualification=str(row.get('qualification') or row.get('Qualification') or row.get('programme') or row.get('Programme') or '').strip()
    exam_board=str(row.get('exam_board') or row.get('Exam Board') or '').strip()
    curriculum_version=str(row.get('curriculum_version') or row.get('Curriculum Version') or '').strip()
    programme=str(row.get('programme') or row.get('Programme') or '').strip()
    subject=str(row.get('subject') or row.get('Subject') or '').strip()
    family_key=canonical_family_key(family_id,country,qualification,exam_board,curriculum_version,programme,subject)
    construct=str(row.get('construct_signature') or row.get('Family Construct') or row.get('concept') or row.get('Concept') or '').strip()
    invariants=str(row.get('invariants_json') or row.get('Family Invariants') or '').strip()
    if invariants and not invariants.startswith(('[','{')):
        invariants=json.dumps([x.strip() for x in invariants.split('|') if x.strip()])
    if not invariants:
        invariants='[]'
    c.execute("""INSERT INTO question_families(
      family_key,family_id,country,qualification,exam_board,curriculum_version,programme,subject,
      learning_outcome,concept,construct_signature,invariants_json,review_status,active,source_type,updated_at)
      VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,CURRENT_TIMESTAMP)
      ON CONFLICT(family_key) DO UPDATE SET
        learning_outcome=CASE WHEN excluded.learning_outcome<>'' THEN excluded.learning_outcome ELSE question_families.learning_outcome END,
        concept=CASE WHEN excluded.concept<>'' THEN excluded.concept ELSE question_families.concept END,
        construct_signature=CASE WHEN excluded.construct_signature<>'' THEN excluded.construct_signature ELSE question_families.construct_signature END,
        invariants_json=CASE WHEN excluded.invariants_json<>'[]' THEN excluded.invariants_json ELSE question_families.invariants_json END,
        source_type=CASE WHEN excluded.source_type<>'' THEN excluded.source_type ELSE question_families.source_type END,
        updated_at=CURRENT_TIMESTAMP""",
      (family_key,family_id,country,qualification,exam_board,curriculum_version,programme,subject,
       str(row.get('learning_outcome') or row.get('Learning Outcome') or '').strip(),
       str(row.get('concept') or row.get('Concept') or '').strip(),construct,invariants,
       review_status,1 if active else 0,str(row.get('source_type') or row.get('Source Type') or 'ScoreMax Original').strip()))
    return family_key

def ensure_curriculum_mapping_for_question(c, question_db_id, row):
    """Map imported content into the curriculum tree without hard-coding one market."""
    levels=[
      ('country',str(row.get('Country') or row.get('country') or 'Pakistan').strip()),
      ('qualification',str(row.get('Qualification') or row.get('qualification') or row.get('Programme') or row.get('programme') or '').strip()),
      ('board',str(row.get('Exam Board') or row.get('exam_board') or '').strip()),
      ('curriculum',str(row.get('Curriculum Version') or row.get('curriculum_version') or '').strip()),
      ('subject',str(row.get('Subject') or row.get('subject') or '').strip()),
      ('chapter',str(row.get('Chapter') or row.get('chapter') or '').strip()),
      ('learning_outcome',str(row.get('Learning Outcome') or row.get('learning_outcome') or '').strip()),
      ('concept',str(row.get('Concept') or row.get('concept') or '').strip())]
    parent_id=None; primary_id=None
    country=levels[0][1]; version=levels[3][1]
    for node_type,name in levels:
        if not name:
            continue
        existing=c.execute("""SELECT id FROM curriculum_nodes WHERE
          ((parent_id IS NULL AND ? IS NULL) OR parent_id=?) AND node_type=? AND name=? AND version=?""",
          (parent_id,parent_id,node_type,name,version)).fetchone()
        if existing:
            node_id=existing['id']
        else:
            cur=c.execute("""INSERT INTO curriculum_nodes(parent_id,node_type,code,name,country,version,active)
              VALUES(?,?,?,?,?,?,1)""",(parent_id,node_type,'',name,country,version))
            node_id=cur.lastrowid
        parent_id=node_id; primary_id=node_id
    if primary_id:
        c.execute("INSERT OR IGNORE INTO question_curriculum_map(question_id,curriculum_node_id,relation_type) VALUES(?,?,'primary')",
                  (question_db_id,primary_id))
    return primary_id

def migrate_v5(c):
    """Backward-compatible V5 schema additions. Existing V4 rows remain valid."""
    question_columns = {
        'country': "TEXT DEFAULT 'Pakistan'",
        'qualification': "TEXT DEFAULT ''",
        'exam_board': "TEXT DEFAULT ''",
        'curriculum_version': "TEXT DEFAULT ''",
        'learning_outcome': "TEXT DEFAULT ''",
        'concept': "TEXT DEFAULT ''",
        'difficulty': "TEXT DEFAULT ''",
        'cognitive_skill': "TEXT DEFAULT ''",
        'command_word': "TEXT DEFAULT ''",
        'marks': "REAL DEFAULT 1",
        'estimated_time_seconds': "INTEGER DEFAULT 60",
        'stimulus_type': "TEXT DEFAULT ''",
        'stimulus_data': "TEXT DEFAULT '{}'",
        'answer_config': "TEXT DEFAULT '{}'",
        'marking_config': "TEXT DEFAULT '{}'",
        'feedback_config': "TEXT DEFAULT '{}'",
        'misconception_tags': "TEXT DEFAULT '[]'",
        'prerequisite_tags': "TEXT DEFAULT '[]'",
        'question_version': "INTEGER DEFAULT 1",
        'review_status': "TEXT DEFAULT 'Approved'",
        'reviewer': "TEXT DEFAULT ''",
        'reviewed_at': "TEXT DEFAULT ''",
        'source_type': "TEXT DEFAULT 'ScoreMax Original'",
        'secure_bank': "INTEGER DEFAULT 0",
        'language': "TEXT DEFAULT 'en'",
        'translation_group': "TEXT DEFAULT ''",
        'family_key': "TEXT DEFAULT ''",
        'active': "INTEGER DEFAULT 1"
    }
    for name, definition in question_columns.items():
        ensure_column(c, 'questions', name, definition)

    session_columns = {
        'confidence_json': "TEXT DEFAULT '{}'",
        'response_times_json': "TEXT DEFAULT '{}'",
        # V6.3 post-Claude hardening: persist the immutable result link for idempotent submit replay.
        'submitted_attempt_id': "INTEGER"
    }
    for name, definition in session_columns.items():
        ensure_column(c, 'assessment_sessions', name, definition)

    answer_columns = {
        'marks_awarded': "REAL DEFAULT 0",
        'confidence': "TEXT DEFAULT ''",
        'response_time_seconds': "INTEGER DEFAULT 0",
        'question_version': "INTEGER DEFAULT 1",
        'misconception_triggered': "TEXT DEFAULT ''",
        'created_at': "TIMESTAMP DEFAULT CURRENT_TIMESTAMP"
    }
    for name, definition in answer_columns.items():
        ensure_column(c, 'attempt_answers', name, definition)

    attempt_columns = {
        'assessment_kind': "TEXT DEFAULT 'standard'",
        'recovery_parent_attempt_id': "INTEGER",
        'recovery_focus_type': "TEXT DEFAULT ''",
        'recovery_focus_name': "TEXT DEFAULT ''",
        'exam_paper_id': "INTEGER",
        'exam_paper_code': "TEXT DEFAULT ''",
        'exam_title': "TEXT DEFAULT ''",
        'guided_mode': "INTEGER DEFAULT 0",
        # One assessment session may create at most one scored attempt.
        'assessment_session_id': "INTEGER"
    }
    for name, definition in attempt_columns.items():
        ensure_column(c, 'attempts', name, definition)
    # Database-level idempotency backs up the route-level atomic claim, including multi-process retries.
    duplicate_session_attempt=c.execute("SELECT assessment_session_id FROM attempts WHERE assessment_session_id IS NOT NULL GROUP BY assessment_session_id HAVING COUNT(*)>1 LIMIT 1").fetchone()
    if not duplicate_session_attempt:
        c.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_attempts_unique_assessment_session ON attempts(assessment_session_id) WHERE assessment_session_id IS NOT NULL")
    else:
        logging.error('Duplicate assessment_session_id values exist in attempts; unique idempotency index was not created.')

    c.executescript("""
    CREATE TABLE IF NOT EXISTS question_families(
      family_key TEXT PRIMARY KEY,
      family_id TEXT NOT NULL,
      country TEXT DEFAULT 'Pakistan',
      qualification TEXT DEFAULT '',
      exam_board TEXT DEFAULT '',
      curriculum_version TEXT DEFAULT '',
      programme TEXT DEFAULT '',
      subject TEXT DEFAULT '',
      learning_outcome TEXT DEFAULT '',
      concept TEXT DEFAULT '',
      construct_signature TEXT DEFAULT '',
      invariants_json TEXT DEFAULT '[]',
      review_status TEXT DEFAULT 'Draft',
      active INTEGER DEFAULT 0,
      source_type TEXT DEFAULT 'ScoreMax Original',
      reviewer_id INTEGER,
      reviewed_at TEXT DEFAULT '',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    CREATE INDEX IF NOT EXISTS idx_question_families_lookup ON question_families(family_id,subject,review_status,active);
    CREATE INDEX IF NOT EXISTS idx_questions_family_key ON questions(family_key);
    CREATE INDEX IF NOT EXISTS idx_questions_live_scope ON questions(status,review_status,active,programme,subject,chapter,level);
    CREATE TABLE IF NOT EXISTS curriculum_nodes(
      id INTEGER PRIMARY KEY, parent_id INTEGER, node_type TEXT NOT NULL, code TEXT, name TEXT NOT NULL,
      country TEXT DEFAULT '', version TEXT DEFAULT '', active INTEGER DEFAULT 1,
      UNIQUE(parent_id,node_type,name,version));
    CREATE TABLE IF NOT EXISTS question_curriculum_map(
      question_id INTEGER NOT NULL, curriculum_node_id INTEGER NOT NULL, relation_type TEXT DEFAULT 'primary',
      PRIMARY KEY(question_id,curriculum_node_id,relation_type));
    CREATE INDEX IF NOT EXISTS idx_question_curriculum_question ON question_curriculum_map(question_id);
    CREATE INDEX IF NOT EXISTS idx_question_curriculum_node ON question_curriculum_map(curriculum_node_id);
    CREATE TABLE IF NOT EXISTS question_versions(
      id INTEGER PRIMARY KEY, question_id INTEGER NOT NULL, version INTEGER NOT NULL, snapshot_json TEXT NOT NULL,
      change_note TEXT DEFAULT '', created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      UNIQUE(question_id,version));
    CREATE TABLE IF NOT EXISTS question_review_events(
      id INTEGER PRIMARY KEY, question_id INTEGER NOT NULL, action TEXT NOT NULL, reviewer_id INTEGER, reason_code TEXT,
      note TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    """)

    # Backfill family governance for trusted pre-V5.4.2 content. New imports are
    # deliberately handled later as Draft + inactive and never inherit spreadsheet status.
    for q in c.execute("""SELECT id,family_id,country,qualification,exam_board,curriculum_version,programme,subject,
      learning_outcome,concept,source_type,status,review_status,active,family_key FROM questions""").fetchall():
        if not (q['family_id'] or '').strip():
            continue
        family_key=(q['family_key'] or '').strip() or canonical_family_key(
            q['family_id'],q['country'],q['qualification'],q['exam_board'],q['curriculum_version'],q['programme'],q['subject'])
        c.execute("UPDATE questions SET family_key=? WHERE id=?",(family_key,q['id']))
        live=(q['status']=='Approved' and q['review_status']=='Approved' and int(q['active'] or 0)==1)
        c.execute("""INSERT OR IGNORE INTO question_families(
          family_key,family_id,country,qualification,exam_board,curriculum_version,programme,subject,
          learning_outcome,concept,construct_signature,invariants_json,review_status,active,source_type)
          VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
          (family_key,q['family_id'],q['country'] or 'Pakistan',q['qualification'] or '',q['exam_board'] or '',q['curriculum_version'] or '',
           q['programme'] or '',q['subject'] or '',q['learning_outcome'] or '',q['concept'] or '',q['concept'] or '','[]',
           'Approved' if live else 'Draft',1 if live else 0,q['source_type'] or 'ScoreMax Original'))

    c.executescript("""
    CREATE TABLE IF NOT EXISTS assignments(
      id INTEGER PRIMARY KEY, teacher_id INTEGER NOT NULL, classroom_id INTEGER,
      title TEXT NOT NULL, subject TEXT DEFAULT '', chapter TEXT DEFAULT '', topic TEXT DEFAULT '',
      subtopic TEXT DEFAULT '', level TEXT DEFAULT '', assessment_mode TEXT DEFAULT 'exam',
      due_at TEXT DEFAULT '', status TEXT DEFAULT 'active', created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS assignment_students(
      assignment_id INTEGER NOT NULL, student_id INTEGER NOT NULL, attempt_id INTEGER,
      status TEXT DEFAULT 'assigned', completed_at TEXT DEFAULT '',
      PRIMARY KEY(assignment_id,student_id));
    """)

    assignment_columns = {
        'assignment_kind': "TEXT DEFAULT 'standard'",
        'focus_type': "TEXT DEFAULT ''",
        'focus_name': "TEXT DEFAULT ''",
        'question_count': "INTEGER DEFAULT 8",
        'duration_minutes': "INTEGER",
        'created_from_diagnostic': "INTEGER DEFAULT 0",
        'baseline_accuracy': "REAL"
    }
    for name, definition in assignment_columns.items():
        ensure_column(c, 'assignments', name, definition)

    # V5 Checkpoint 9 — Exam Centre, official past-paper simulations and ScoreMax mocks.
    c.executescript("""
    CREATE TABLE IF NOT EXISTS exam_blueprints(
      id INTEGER PRIMARY KEY, code TEXT UNIQUE NOT NULL, title TEXT NOT NULL,
      country TEXT DEFAULT 'Pakistan', qualification TEXT DEFAULT '', exam_board TEXT DEFAULT '',
      programme TEXT DEFAULT '', subject TEXT NOT NULL, curriculum_version TEXT DEFAULT '',
      paper_name TEXT DEFAULT '', duration_minutes INTEGER, total_marks REAL,
      section_config_json TEXT DEFAULT '[]', coverage_json TEXT DEFAULT '{}',
      question_type_mix_json TEXT DEFAULT '{}', difficulty_mix_json TEXT DEFAULT '{}',
      active INTEGER DEFAULT 1, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS exam_papers(
      id INTEGER PRIMARY KEY, blueprint_id INTEGER, code TEXT UNIQUE NOT NULL, title TEXT NOT NULL,
      paper_kind TEXT DEFAULT 'scoremax_mock', official_year TEXT DEFAULT '',
      source_label TEXT DEFAULT '', source_url TEXT DEFAULT '', reproduction_status TEXT DEFAULT 'scoremax_original',
      duration_minutes INTEGER, total_marks REAL, instructions TEXT DEFAULT '',
      premium_required INTEGER DEFAULT 0, active INTEGER DEFAULT 1,
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS exam_paper_questions(
      paper_id INTEGER NOT NULL, question_id INTEGER NOT NULL, position INTEGER NOT NULL,
      section_label TEXT DEFAULT '', display_number TEXT DEFAULT '', marks REAL,
      PRIMARY KEY(paper_id,question_id), UNIQUE(paper_id,position));
    CREATE TABLE IF NOT EXISTS student_exam_dates(
      id INTEGER PRIMARY KEY, student_id INTEGER NOT NULL, subject TEXT NOT NULL,
      exam_name TEXT DEFAULT '', exam_date TEXT NOT NULL, active INTEGER DEFAULT 1,
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    """)

    # V5.1 — Student Experience Refinement + Learning Capsules.
    v51_user_columns = {
        'goal_type': "TEXT DEFAULT ''",
        'goal_name': "TEXT DEFAULT ''",
        'target_percentage': "REAL",
        'preferred_view': "TEXT DEFAULT 'dashboard'"
    }
    for name, definition in v51_user_columns.items():
        ensure_column(c, 'users', name, definition)

    v51_question_columns = {
        'concept_id': "TEXT DEFAULT ''",
        'capsule_id': "TEXT DEFAULT ''",
        'misconception_id': "TEXT DEFAULT ''"
    }
    for name, definition in v51_question_columns.items():
        ensure_column(c, 'questions', name, definition)

    c.executescript("""
    CREATE TABLE IF NOT EXISTS learning_capsules(
      id INTEGER PRIMARY KEY,
      capsule_id TEXT UNIQUE NOT NULL,
      country TEXT DEFAULT 'Pakistan',
      programme TEXT DEFAULT '',
      subject TEXT NOT NULL,
      chapter TEXT DEFAULT '',
      topic TEXT DEFAULT '',
      subtopic TEXT DEFAULT '',
      learning_outcome TEXT DEFAULT '',
      concept_id TEXT DEFAULT '',
      concept TEXT NOT NULL,
      simple_explanation TEXT NOT NULL,
      remember_this TEXT DEFAULT '',
      formula_rule TEXT DEFAULT '',
      worked_example TEXT DEFAULT '',
      common_mistake TEXT DEFAULT '',
      memory_tip TEXT DEFAULT '',
      visual_brief TEXT DEFAULT '',
      video_url TEXT DEFAULT '',
      quick_check_question_ids TEXT DEFAULT '[]',
      recovery_question_ids TEXT DEFAULT '[]',
      review_status TEXT DEFAULT 'Draft',
      reviewer TEXT DEFAULT '',
      version INTEGER DEFAULT 1,
      active INTEGER DEFAULT 1,
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS misconception_library(
      id INTEGER PRIMARY KEY,
      misconception_id TEXT UNIQUE NOT NULL,
      concept_id TEXT DEFAULT '',
      subject TEXT DEFAULT '',
      title TEXT NOT NULL,
      student_explanation TEXT DEFAULT '',
      corrective_hint TEXT DEFAULT '',
      active INTEGER DEFAULT 1);
    CREATE TABLE IF NOT EXISTS student_external_results(
      id INTEGER PRIMARY KEY,
      student_id INTEGER NOT NULL,
      exam_type TEXT NOT NULL,
      subject TEXT DEFAULT '',
      board_authority TEXT DEFAULT '',
      exam_year INTEGER,
      marks_obtained REAL,
      total_marks REAL,
      percentage REAL,
      grade TEXT DEFAULT '',
      verification_status TEXT DEFAULT 'self_reported',
      scoremax_level_snapshot TEXT DEFAULT '',
      scoremax_score_snapshot REAL,
      recorded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS level_snapshots(
      id INTEGER PRIMARY KEY,
      student_id INTEGER NOT NULL,
      programme TEXT DEFAULT '',
      subject TEXT NOT NULL,
      level_name TEXT NOT NULL,
      readiness_score REAL,
      evidence_count INTEGER DEFAULT 0,
      reason TEXT DEFAULT '',
      snapshot_type TEXT DEFAULT 'routine',
      related_exam_type TEXT DEFAULT '',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS content_updates(
      id INTEGER PRIMARY KEY,
      category TEXT NOT NULL,
      title TEXT NOT NULL,
      summary TEXT DEFAULT '',
      source_name TEXT DEFAULT '',
      source_url TEXT DEFAULT '',
      published_date TEXT DEFAULT '',
      important_date TEXT DEFAULT '',
      audience TEXT DEFAULT 'all',
      status TEXT DEFAULT 'published',
      verified INTEGER DEFAULT 0,
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS contact_messages(
      id INTEGER PRIMARY KEY,
      name TEXT NOT NULL,
      email TEXT DEFAULT '',
      mobile TEXT DEFAULT '',
      enquiry_type TEXT DEFAULT 'general',
      message TEXT NOT NULL,
      status TEXT DEFAULT 'new',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    """)

    # Seed a few learning capsules for the built-in demonstration concepts.
    capsule_seed = [
      ('LC-BIO-OSM','FSc Part 1','Biology','Cell Biology','Cell Transport','Osmosis','BIO-OSM','Osmosis',
       'Osmosis is the movement of water through a selectively permeable membrane from a region with more available water to a region with less available water.',
       'Water moves; the membrane controls what can pass.','','A plant cell placed in a concentrated salt solution loses water and becomes less turgid.',
       'A common mistake is saying that osmosis moves solute. Osmosis specifically describes water movement.',
       'Think: OSMOSIS = WATER moving across a membrane.','Simple membrane diagram showing water moving across a selectively permeable membrane.'),
      ('LC-BIO-ACT','FSc Part 1','Biology','Cell Biology','Cell Transport','Active Transport','BIO-ACT','Active Transport',
       'Active transport moves substances across a cell membrane against their concentration gradient. Because this movement goes against the natural direction of diffusion, the cell needs energy.',
       'Against the gradient means energy is needed.','ATP provides energy','A membrane pump can move ions from a lower concentration to a higher concentration by using energy from ATP.',
       'Do not confuse active transport with diffusion. Diffusion does not require ATP.',
       'Against the gradient = active = energy.','Membrane pump moving ions from low concentration to high concentration.')
    ]
    for row in capsule_seed:
        c.execute("""INSERT OR IGNORE INTO learning_capsules(
          capsule_id,programme,subject,chapter,topic,subtopic,concept_id,concept,simple_explanation,
          remember_this,formula_rule,worked_example,common_mistake,memory_tip,visual_brief,review_status)
          VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'Approved')""", row)

    c.execute("""UPDATE questions SET concept_id='BIO-OSM',concept='Osmosis',capsule_id='LC-BIO-OSM'
      WHERE subject='Biology' AND subtopic='Osmosis' AND COALESCE(capsule_id,'')=''""")
    c.execute("""UPDATE questions SET concept_id='BIO-ACT',concept='Active Transport',capsule_id='LC-BIO-ACT'
      WHERE subject='Biology' AND subtopic='Active Transport' AND COALESCE(capsule_id,'')=''""")

    # V5.1 Refined — Study Plans, richer achievements, challenge types and admin account controls.
    for name, definition in {
        'study_plan_pathway': "TEXT DEFAULT ''",
        'study_plan_source': "TEXT DEFAULT ''",
        'study_plan_active': "INTEGER DEFAULT 0",
        'account_status': "TEXT DEFAULT 'active'",
        'help_tips_enabled': "INTEGER DEFAULT 1",
        'weekly_email_enabled': "INTEGER DEFAULT 1",
        'parent_weekly_email_enabled': "INTEGER DEFAULT 0",
        'parent_link_code': "TEXT DEFAULT ''",
        'session_version': "INTEGER DEFAULT 0"
    }.items():
        ensure_column(c,'users',name,definition)

    for u in c.execute("SELECT id FROM users WHERE role='student' AND COALESCE(parent_link_code,'')=''").fetchall():
        c.execute("UPDATE users SET parent_link_code=? WHERE id=?",('P-'+secrets.token_urlsafe(12).upper(),u['id']))

    # Replace legacy predictable parent-link codes from earlier prototypes with high-entropy codes.
    for u in c.execute("SELECT id,parent_link_code FROM users WHERE role='student'").fetchall():
        legacy=(u['parent_link_code'] or '').strip()
        if not legacy or re.fullmatch(r'P\d{4,12}',legacy):
            c.execute("UPDATE users SET parent_link_code=? WHERE id=?",('P-'+secrets.token_urlsafe(12).upper(),u['id']))

    # Challenge columns are added after the base challenges table is created below.

    c.executescript("""
    CREATE TABLE IF NOT EXISTS study_plans(
      id INTEGER PRIMARY KEY,
      student_id INTEGER NOT NULL,
      pathway TEXT NOT NULL DEFAULT 'Core',
      source TEXT NOT NULL DEFAULT 'scoremax',
      title TEXT DEFAULT '',
      target_exam TEXT DEFAULT '',
      target_date TEXT DEFAULT '',
      target_percentage REAL,
      status TEXT DEFAULT 'active',
      assigned_by_user_id INTEGER,
      notes TEXT DEFAULT '',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS study_plan_activities(
      id INTEGER PRIMARY KEY,
      plan_id INTEGER NOT NULL,
      student_id INTEGER NOT NULL,
      activity_date TEXT DEFAULT '',
      subject TEXT DEFAULT '',
      chapter TEXT DEFAULT '',
      topic TEXT DEFAULT '',
      activity_type TEXT DEFAULT 'practice',
      title TEXT NOT NULL,
      target_score REAL,
      status TEXT DEFAULT 'planned',
      source_reason TEXT DEFAULT '',
      completed_at TEXT,
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS external_result_subjects(
      id INTEGER PRIMARY KEY,
      external_result_id INTEGER NOT NULL,
      student_id INTEGER NOT NULL,
      subject TEXT NOT NULL,
      marks_obtained REAL,
      total_marks REAL,
      percentage REAL,
      grade TEXT DEFAULT '',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    """)

    for name, definition in {
        'readiness_date': "TEXT DEFAULT ''",
        'exam_mode_days': "INTEGER DEFAULT 30",
        'days_per_week': "INTEGER DEFAULT 6",
        'minutes_per_day': "INTEGER DEFAULT 120",
        'starting_coverage': "REAL DEFAULT 0",
        'starting_accuracy': "REAL",
        'readiness_target': "REAL",
        'pace_status': "TEXT DEFAULT ''",
        'last_rebalanced_at': "TEXT DEFAULT ''",
        'custom_settings_json': "TEXT DEFAULT '{}'"
    }.items():
        ensure_column(c,'study_plans',name,definition)

    for name, definition in {
        'week_no': "INTEGER",
        'phase': "TEXT DEFAULT ''",
        'priority': "INTEGER DEFAULT 3",
        'estimated_minutes': "INTEGER DEFAULT 30",
        'required_mastery': "REAL",
        'mandatory': "INTEGER DEFAULT 1"
    }.items():
        ensure_column(c,'study_plan_activities',name,definition)

    c.executescript("""
    CREATE TABLE IF NOT EXISTS parent_student_links(
      id INTEGER PRIMARY KEY,
      parent_user_id INTEGER NOT NULL,
      student_user_id INTEGER NOT NULL,
      relationship TEXT DEFAULT 'Parent',
      status TEXT DEFAULT 'active',
      share_plan INTEGER DEFAULT 1,
      share_progress INTEGER DEFAULT 1,
      share_results INTEGER DEFAULT 1,
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      UNIQUE(parent_user_id,student_user_id));

    CREATE TABLE IF NOT EXISTS weekly_progress_reports(
      id INTEGER PRIMARY KEY,
      student_id INTEGER NOT NULL,
      week_start TEXT NOT NULL,
      week_end TEXT NOT NULL,
      plan_id INTEGER,
      plan_completion REAL DEFAULT 0,
      tests_completed INTEGER DEFAULT 0,
      avg_score REAL,
      status_label TEXT DEFAULT '',
      summary_text TEXT DEFAULT '',
      priorities_text TEXT DEFAULT '',
      student_email_enabled INTEGER DEFAULT 1,
      parent_email_enabled INTEGER DEFAULT 0,
      generated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      UNIQUE(student_id,week_start));
    """)

    # V5 Checkpoint 12 — referral rewards, wallet ledger and integration controls.
    referral_user_columns = {
        'own_referral_code': "TEXT DEFAULT ''",
        'wallet_balance_minor': "INTEGER DEFAULT 0"
    }
    for name, definition in referral_user_columns.items():
        ensure_column(c, 'users', name, definition)

    c.executescript("""
    CREATE UNIQUE INDEX IF NOT EXISTS idx_users_own_referral_code
      ON users(own_referral_code) WHERE own_referral_code IS NOT NULL AND own_referral_code<>'';
    CREATE TABLE IF NOT EXISTS referral_programs(
      role_group TEXT PRIMARY KEY, reward_type TEXT NOT NULL, reward_rate REAL DEFAULT 0.10,
      hold_days INTEGER DEFAULT 14, active INTEGER DEFAULT 1,
      updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS referral_rewards(
      id INTEGER PRIMARY KEY, payment_transaction_id INTEGER UNIQUE NOT NULL,
      referrer_user_id INTEGER NOT NULL, referred_user_id INTEGER NOT NULL,
      reward_type TEXT DEFAULT 'wallet_credit', currency TEXT DEFAULT 'PKR',
      qualifying_amount_minor INTEGER DEFAULT 0, reward_rate REAL DEFAULT 0.10,
      reward_amount_minor INTEGER DEFAULT 0, status TEXT DEFAULT 'pending',
      available_at TEXT, paid_at TEXT, reversed_at TEXT, notes TEXT DEFAULT '',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS wallet_transactions(
      id INTEGER PRIMARY KEY, user_id INTEGER NOT NULL, referral_reward_id INTEGER,
      currency TEXT DEFAULT 'PKR', amount_minor INTEGER NOT NULL,
      transaction_type TEXT DEFAULT 'referral_reward', status TEXT DEFAULT 'posted',
      description TEXT DEFAULT '', created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS wallet_balances(
      user_id INTEGER NOT NULL, currency TEXT NOT NULL, balance_minor INTEGER DEFAULT 0,
      updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      PRIMARY KEY(user_id,currency));
    """)

    # V5 Checkpoint 11 — Institution / College operating layer.
    institution_columns = {
        'parent_institution_id': "INTEGER",
        'campus_name': "TEXT DEFAULT ''",
        'contact_name': "TEXT DEFAULT ''",
        'contact_email': "TEXT DEFAULT ''"
    }
    for name, definition in institution_columns.items():
        ensure_column(c, 'institutions', name, definition)

    c.execute("INSERT OR IGNORE INTO referral_programs(role_group,reward_type,reward_rate,hold_days,active) VALUES('student','wallet_credit',0.10,14,1)")
    c.execute("INSERT OR IGNORE INTO referral_programs(role_group,reward_type,reward_rate,hold_days,active) VALUES('partner','commission',0.10,14,1)")

    for u in c.execute("SELECT id,full_name,own_referral_code FROM users WHERE COALESCE(own_referral_code,'')=''").fetchall():
        stem=''.join(ch for ch in (u['full_name'] or 'USER').upper() if ch.isalnum())[:5] or 'USER'
        c.execute("UPDATE users SET own_referral_code=? WHERE id=?",(f"{stem}{int(u['id']):05d}",u['id']))

    c.executescript("""
    CREATE TABLE IF NOT EXISTS institution_staff(
      institution_id INTEGER NOT NULL, user_id INTEGER NOT NULL,
      institution_role TEXT DEFAULT 'viewer', active INTEGER DEFAULT 1,
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      PRIMARY KEY(institution_id,user_id));
    CREATE TABLE IF NOT EXISTS institution_targets(
      id INTEGER PRIMARY KEY, institution_id INTEGER NOT NULL, subject TEXT DEFAULT '',
      target_type TEXT DEFAULT 'accuracy', target_value REAL DEFAULT 70,
      starts_at TEXT, ends_at TEXT, active INTEGER DEFAULT 1,
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    """)

    # V5 Checkpoint 10 — voluntary rankings and monthly ScoreMax Challenges.
    ranking_columns = {
        'ranking_opt_in': "INTEGER DEFAULT 0",
        'ranking_display_name': "TEXT DEFAULT ''"
    }
    for name, definition in ranking_columns.items():
        ensure_column(c, 'users', name, definition)

    c.executescript("""
    CREATE TABLE IF NOT EXISTS challenges(
      id INTEGER PRIMARY KEY, code TEXT UNIQUE NOT NULL, title TEXT NOT NULL,
      country TEXT DEFAULT 'Pakistan', subject TEXT NOT NULL, qualification TEXT DEFAULT '',
      exam_board TEXT DEFAULT '', challenge_month TEXT NOT NULL, description TEXT DEFAULT '',
      duration_minutes INTEGER DEFAULT 30, question_count INTEGER DEFAULT 20,
      opens_at TEXT NOT NULL, closes_at TEXT NOT NULL, premium_required INTEGER DEFAULT 1,
      ranking_enabled INTEGER DEFAULT 1, exact_rank_min_score REAL DEFAULT 80,
      max_attempts INTEGER DEFAULT 1, status TEXT DEFAULT 'draft',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS challenge_questions(
      challenge_id INTEGER NOT NULL, question_id INTEGER NOT NULL, position INTEGER NOT NULL,
      PRIMARY KEY(challenge_id,question_id), UNIQUE(challenge_id,position));
    CREATE TABLE IF NOT EXISTS challenge_entries(
      id INTEGER PRIMARY KEY, challenge_id INTEGER NOT NULL, student_id INTEGER NOT NULL,
      assessment_session_id INTEGER, attempt_id INTEGER, status TEXT DEFAULT 'registered',
      started_at TEXT, completed_at TEXT, score REAL, correct_count INTEGER, total_count INTEGER,
      elapsed_seconds INTEGER DEFAULT 0, ranking_consent INTEGER DEFAULT 0,
      ranking_band TEXT DEFAULT '', exact_rank INTEGER, percentile REAL,
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      UNIQUE(challenge_id,student_id));
    CREATE TABLE IF NOT EXISTS ranking_snapshots(
      id INTEGER PRIMARY KEY, challenge_id INTEGER NOT NULL, generated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      participant_count INTEGER DEFAULT 0, snapshot_json TEXT DEFAULT '[]');
    """)

    for name, definition in {
        'challenge_type': "TEXT DEFAULT 'subject'",
        'chapter': "TEXT DEFAULT ''",
        'topic': "TEXT DEFAULT ''",
        'created_by_user_id': "INTEGER",
        'official': "INTEGER DEFAULT 1"
    }.items():
        ensure_column(c,'challenges',name,definition)

    # V5 Checkpoint 8 — commercial access, subscriptions and payment tracking.
    user_commercial_columns = {
        'referral_source': "TEXT DEFAULT ''",
        'referral_code': "TEXT DEFAULT ''"
    }
    for name, definition in user_commercial_columns.items():
        ensure_column(c, 'users', name, definition)

    c.executescript("""
    CREATE TABLE IF NOT EXISTS plans(
      id INTEGER PRIMARY KEY, code TEXT UNIQUE NOT NULL, name TEXT NOT NULL, audience TEXT NOT NULL,
      billing_period TEXT DEFAULT 'none', currency TEXT DEFAULT 'PKR', price_minor INTEGER,
      entitlements_json TEXT DEFAULT '{}', active INTEGER DEFAULT 1, sort_order INTEGER DEFAULT 0,
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS subscriptions(
      id INTEGER PRIMARY KEY, user_id INTEGER, institution_id INTEGER, plan_id INTEGER NOT NULL,
      status TEXT DEFAULT 'active', source TEXT DEFAULT 'manual', provider TEXT DEFAULT '',
      provider_subscription_ref TEXT DEFAULT '', starts_at TEXT NOT NULL, ends_at TEXT,
      renews_at TEXT, auto_renew INTEGER DEFAULT 0, trial_ends_at TEXT,
      cancelled_at TEXT, cancellation_reason TEXT DEFAULT '', created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS payment_transactions(
      id INTEGER PRIMARY KEY, user_id INTEGER, institution_id INTEGER, subscription_id INTEGER,
      plan_id INTEGER, provider TEXT DEFAULT 'manual', provider_transaction_ref TEXT DEFAULT '',
      currency TEXT DEFAULT 'PKR', gross_amount_minor INTEGER DEFAULT 0, discount_minor INTEGER DEFAULT 0,
      tax_minor INTEGER DEFAULT 0, net_amount_minor INTEGER DEFAULT 0, status TEXT DEFAULT 'pending',
      payment_method TEXT DEFAULT '', paid_at TEXT, failure_reason TEXT DEFAULT '',
      refund_amount_minor INTEGER DEFAULT 0, refund_status TEXT DEFAULT '',
      referral_source TEXT DEFAULT '', promo_code TEXT DEFAULT '', notes TEXT DEFAULT '',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS promo_codes(
      id INTEGER PRIMARY KEY, code TEXT UNIQUE NOT NULL, discount_type TEXT DEFAULT 'percent',
      discount_value INTEGER DEFAULT 0, starts_at TEXT, expires_at TEXT, usage_limit INTEGER,
      times_used INTEGER DEFAULT 0, active INTEGER DEFAULT 1, campaign TEXT DEFAULT '',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS promo_redemptions(
      id INTEGER PRIMARY KEY, promo_code_id INTEGER NOT NULL, user_id INTEGER,
      payment_transaction_id INTEGER, redeemed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS referral_attributions(
      id INTEGER PRIMARY KEY, user_id INTEGER UNIQUE NOT NULL, referral_source TEXT DEFAULT '',
      referral_code TEXT DEFAULT '', referrer_type TEXT DEFAULT '', referrer_id INTEGER,
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS institution_licenses(
      id INTEGER PRIMARY KEY, institution_id INTEGER NOT NULL, plan_id INTEGER NOT NULL,
      seat_count INTEGER DEFAULT 0, starts_at TEXT NOT NULL, ends_at TEXT, status TEXT DEFAULT 'active',
      invoice_ref TEXT DEFAULT '', payment_status TEXT DEFAULT 'pending', notes TEXT DEFAULT '',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS institution_license_users(
      institution_license_id INTEGER NOT NULL, user_id INTEGER NOT NULL,
      allocated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      PRIMARY KEY(institution_license_id,user_id));
    CREATE TABLE IF NOT EXISTS commission_events(
      id INTEGER PRIMARY KEY, payment_transaction_id INTEGER NOT NULL, beneficiary_type TEXT DEFAULT '',
      beneficiary_id INTEGER, commission_rate REAL DEFAULT 0, commission_amount_minor INTEGER DEFAULT 0,
      status TEXT DEFAULT 'pending', paid_at TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    """)

    free_entitlements = {
        'diagnostic': True, 'basic_results': True, 'topic_practice': True,
        'advanced_diagnostics': False, 'recovery_engine': False,
        'all_mocks': False, 'monthly_challenges': False, 'ranking_eligible': False,
        'premium_reports': False
    }
    premium_entitlements = {
        'diagnostic': True, 'basic_results': True, 'topic_practice': True,
        'advanced_diagnostics': True, 'recovery_engine': True,
        'all_mocks': True, 'monthly_challenges': True, 'ranking_eligible': True,
        'premium_reports': True
    }
    institution_entitlements = dict(premium_entitlements)
    seeded_plans=[
        ('free_student','Free Student','student','none','PKR',0,free_entitlements,10),
        ('premium_monthly','Premium Monthly','student','monthly','PKR',None,premium_entitlements,20),
        ('premium_annual','Premium Annual','student','annual','PKR',None,premium_entitlements,30),
        ('institution_student','Institution Student','student','institution','PKR',None,institution_entitlements,40),
        ('free_teacher','Free Teacher','teacher','none','PKR',0,{'classes':True,'assignments':True,'basic_class_analytics':True},50),
        ('institution_plan','School / College Plan','institution','annual','PKR',None,{'institution_dashboard':True,'student_seats':True,'teacher_analytics':True},60)
    ]
    for code,name,audience,billing,currency,price,entitlements,sort_order in seeded_plans:
        c.execute("""INSERT OR IGNORE INTO plans(code,name,audience,billing_period,currency,price_minor,entitlements_json,active,sort_order)
                     VALUES(?,?,?,?,?,?,?,?,?)""",
                  (code,name,audience,billing,currency,price,json.dumps(entitlements),1,sort_order))

    # V5.3 — evidence-backed planning, recall, misconception state, notifications and versioned admissions rules.
    for name, definition in {
        'evidence_status': "TEXT DEFAULT 'scheduled'",
        'evidence_type': "TEXT DEFAULT ''",
        'linked_attempt_id': "INTEGER",
        'verified_score': "REAL",
        'outcome_status': "TEXT DEFAULT ''",
        'last_evidence_at': "TEXT DEFAULT ''",
        'concept_key': "TEXT DEFAULT ''"
    }.items():
        ensure_column(c,'study_plan_activities',name,definition)

    c.executescript("""
    CREATE TABLE IF NOT EXISTS recall_items(
      id INTEGER PRIMARY KEY,
      student_id INTEGER NOT NULL,
      concept_key TEXT NOT NULL,
      subject TEXT DEFAULT '', chapter TEXT DEFAULT '', topic TEXT DEFAULT '', area_name TEXT DEFAULT '',
      interval_days INTEGER DEFAULT 7,
      successful_recalls INTEGER DEFAULT 0,
      last_score REAL,
      last_attempt_at TEXT DEFAULT '',
      next_due_date TEXT DEFAULT '',
      status TEXT DEFAULT 'scheduled',
      updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      UNIQUE(student_id,concept_key));

    CREATE TABLE IF NOT EXISTS student_learning_states(
      id INTEGER PRIMARY KEY,
      student_id INTEGER NOT NULL,
      area_key TEXT NOT NULL,
      subject TEXT DEFAULT '', chapter TEXT DEFAULT '', topic TEXT DEFAULT '', area_name TEXT DEFAULT '',
      evidence_count INTEGER DEFAULT 0,
      correct_count INTEGER DEFAULT 0,
      accuracy REAL DEFAULT 0,
      status TEXT DEFAULT 'Emerging Concern',
      last_attempt_id INTEGER,
      last_evidence_at TEXT DEFAULT '',
      recovered_at TEXT DEFAULT '',
      UNIQUE(student_id,area_key));

    CREATE TABLE IF NOT EXISTS student_misconceptions(
      id INTEGER PRIMARY KEY,
      student_id INTEGER NOT NULL,
      misconception_key TEXT NOT NULL,
      subject TEXT DEFAULT '', area_name TEXT DEFAULT '',
      evidence_count INTEGER DEFAULT 0,
      confident_wrong_count INTEGER DEFAULT 0,
      status TEXT DEFAULT 'Emerging Concern',
      last_attempt_id INTEGER,
      last_seen_at TEXT DEFAULT '',
      recovered_at TEXT DEFAULT '',
      UNIQUE(student_id,misconception_key));

    CREATE TABLE IF NOT EXISTS user_notifications(
      id INTEGER PRIMARY KEY,
      user_id INTEGER NOT NULL,
      notification_type TEXT DEFAULT 'info',
      title TEXT NOT NULL,
      message TEXT DEFAULT '',
      read_at TEXT DEFAULT '',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS admission_formulas(
      id INTEGER PRIMARY KEY,
      route_code TEXT NOT NULL,
      country TEXT NOT NULL,
      route_name TEXT NOT NULL,
      admission_year INTEGER,
      version_label TEXT NOT NULL,
      components_json TEXT NOT NULL,
      source_url TEXT DEFAULT '',
      source_label TEXT DEFAULT '',
      verified_at TEXT DEFAULT '',
      effective_from TEXT DEFAULT '',
      effective_to TEXT DEFAULT '',
      active INTEGER DEFAULT 1,
      UNIQUE(route_code,admission_year,version_label));

    CREATE TABLE IF NOT EXISTS student_admission_targets(
      id INTEGER PRIMARY KEY,
      student_id INTEGER NOT NULL,
      route_code TEXT NOT NULL,
      institution_name TEXT DEFAULT '',
      programme_name TEXT DEFAULT '',
      target_aggregate REAL,
      admission_year INTEGER,
      active INTEGER DEFAULT 1,
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    """)

    # Current PM&DC public merit weighting is stored as a versioned rule instead of hard-coded UI JavaScript.
    c.execute("""INSERT OR IGNORE INTO admission_formulas(
      route_code,country,route_name,admission_year,version_label,components_json,source_url,source_label,verified_at,effective_from,active)
      VALUES(?,?,?,?,?,?,?,?,?,?,1)""",
      ('PK_MEDICAL','Pakistan','Medical / Dental',2026,'PMDC-2026-FAQ',
       json.dumps([{'key':'matric','label':'SSC / Matric / Equivalent','weight':0.10},
                   {'key':'fsc','label':'F.Sc / HSSC / Equivalent','weight':0.40},
                   {'key':'mdcat','label':'MDCAT','weight':0.50}]),
       'https://pmdc.pk/Home/FAQs','PM&DC FAQs','2026-07-26','2026-01-01'))

    c.execute("UPDATE questions SET qualification=programme WHERE COALESCE(qualification,'')='' ")
    c.execute("UPDATE questions SET review_status=COALESCE(NULLIF(status,''),'Approved') WHERE COALESCE(review_status,'')='' OR review_status='Approved'")
    c.execute("UPDATE questions SET difficulty=level WHERE COALESCE(difficulty,'')='' ")

    rows=c.execute("SELECT id,qtype,option_a,option_b,option_c,option_d,answer,answer_config,marking_config,marks FROM questions").fetchall()
    for r in rows:
        if r['answer_config'] not in (None,'','{}') and r['marking_config'] not in (None,'','{}'):
            continue
        qtype=(r['qtype'] or 'MCQ').strip()
        if qtype=='True/False':
            options=[{'id':'A','text':r['option_a'] or 'True'},{'id':'B','text':r['option_b'] or 'False'}]
            answer_config={'options':options}
            marking_config={'correct_option_ids':[r['answer'] or ''], 'marks': float(r['marks'] or 1)}
        elif qtype=='Fill Blank':
            accepted=[x for x in [r['option_a'],r['answer']] if x]
            answer_config={'accepted_answers':list(dict.fromkeys(accepted)), 'case_sensitive':False, 'trim_spaces':True}
            marking_config={'marks': float(r['marks'] or 1)}
        else:
            opts=[]
            for code,key in [('A','option_a'),('B','option_b'),('C','option_c'),('D','option_d')]:
                if r[key]:
                    opts.append({'id':code,'text':r[key]})
            answer_config={'options':opts}
            marking_config={'correct_option_ids':[r['answer'] or ''], 'marks': float(r['marks'] or 1)}
        c.execute("UPDATE questions SET answer_config=?, marking_config=? WHERE id=?",(json.dumps(answer_config),json.dumps(marking_config),r['id']))




    # V5.4 — Mastery & Progression architecture.
    for name, definition in {
        'access_override_code': "TEXT DEFAULT ''",
        'login_provider': "TEXT DEFAULT 'password'",
        'profile_completed': "INTEGER DEFAULT 0",
        'last_login_at': "TEXT DEFAULT ''"
    }.items():
        ensure_column(c,'users',name,definition)

    for name, definition in {
        'is_demo': "INTEGER DEFAULT 0",
        'response_count': "INTEGER DEFAULT 0",
        'facility_value': "REAL",
        'discrimination_value': "REAL",
        'calibration_status': "TEXT DEFAULT 'PROVISIONAL'",
        'calibrated_at': "TEXT DEFAULT ''"
    }.items():
        ensure_column(c,'questions',name,definition)

    for name, definition in {
        'access_rank': "INTEGER DEFAULT 0",
        'mastery_ceiling': "TEXT DEFAULT 'Foundation'"
    }.items():
        ensure_column(c,'plans',name,definition)
    for name, definition in {
        'access_rank': "INTEGER DEFAULT 0",
        'mastery_ceiling': "TEXT DEFAULT 'Foundation'"
    }.items():
        ensure_column(c,'study_plans',name,definition)

    for name, definition in {
        'scope_type': "TEXT DEFAULT 'Subject'",
        'programme': "TEXT DEFAULT ''",
        'required_mastery_level': "TEXT DEFAULT ''",
        'scope_label': "TEXT DEFAULT ''"
    }.items():
        ensure_column(c,'challenges',name,definition)

    c.executescript("""
    CREATE TABLE IF NOT EXISTS password_reset_tokens(
      id INTEGER PRIMARY KEY,
      user_id INTEGER NOT NULL,
      token_hash TEXT UNIQUE NOT NULL,
      expires_at TEXT NOT NULL,
      used_at TEXT DEFAULT '',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS mastery_policies(
      id INTEGER PRIMARY KEY,
      mastery_level TEXT UNIQUE NOT NULL,
      level_rank INTEGER NOT NULL,
      min_forms INTEGER DEFAULT 1,
      min_questions INTEGER DEFAULT 10,
      min_accuracy REAL DEFAULT 70,
      verification_days INTEGER DEFAULT 90,
      external_percentile_target REAL,
      target_band_pct REAL DEFAULT 0.25,
      unseen_family_pct REAL DEFAULT 0.60,
      active INTEGER DEFAULT 1,
      updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS mastery_form_results(
      id INTEGER PRIMARY KEY,
      student_id INTEGER NOT NULL,
      assessment_session_id INTEGER,
      attempt_id INTEGER,
      scope_type TEXT NOT NULL,
      scope_key TEXT NOT NULL,
      programme TEXT DEFAULT '',
      subject TEXT DEFAULT '',
      chapter TEXT DEFAULT '',
      target_level TEXT NOT NULL,
      score REAL DEFAULT 0,
      question_count INTEGER DEFAULT 0,
      passed INTEGER DEFAULT 0,
      demo_only INTEGER DEFAULT 0,
      breadth_ok INTEGER DEFAULT 0,
      unseen_family_ratio REAL DEFAULT 0,
      family_ids_json TEXT DEFAULT '[]',
      policy_snapshot_json TEXT DEFAULT '{}',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS mastery_records(
      id INTEGER PRIMARY KEY,
      student_id INTEGER NOT NULL,
      scope_type TEXT NOT NULL,
      scope_key TEXT NOT NULL,
      programme TEXT DEFAULT '',
      subject TEXT DEFAULT '',
      chapter TEXT DEFAULT '',
      mastery_level TEXT NOT NULL DEFAULT 'Foundation',
      status TEXT NOT NULL DEFAULT 'Verified',
      verified_at TEXT DEFAULT '',
      verification_due_at TEXT DEFAULT '',
      best_accuracy REAL DEFAULT 0,
      forms_passed INTEGER DEFAULT 0,
      questions_total INTEGER DEFAULT 0,
      failed_reconfirmations INTEGER DEFAULT 0,
      source TEXT DEFAULT 'ScoreMax Mastery',
      updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      UNIQUE(student_id,scope_type,scope_key));

    CREATE TABLE IF NOT EXISTS mastery_history(
      id INTEGER PRIMARY KEY,
      mastery_record_id INTEGER NOT NULL,
      student_id INTEGER NOT NULL,
      scope_type TEXT NOT NULL,
      scope_key TEXT NOT NULL,
      event_type TEXT NOT NULL,
      previous_level TEXT DEFAULT '',
      new_level TEXT DEFAULT '',
      previous_status TEXT DEFAULT '',
      new_status TEXT DEFAULT '',
      attempt_id INTEGER,
      assessment_session_id INTEGER,
      note TEXT DEFAULT '',
      metadata_json TEXT DEFAULT '{}',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS access_change_history(
      id INTEGER PRIMARY KEY,
      student_id INTEGER NOT NULL,
      actor_user_id INTEGER,
      source TEXT NOT NULL DEFAULT 'admin_override',
      previous_access_code TEXT DEFAULT '',
      new_access_code TEXT DEFAULT '',
      note TEXT DEFAULT '',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE INDEX IF NOT EXISTS idx_mastery_forms_student_scope ON mastery_form_results(student_id,scope_type,scope_key,target_level);
    CREATE INDEX IF NOT EXISTS idx_mastery_records_student ON mastery_records(student_id,status);
    CREATE INDEX IF NOT EXISTS idx_mastery_history_student ON mastery_history(student_id,created_at);
    CREATE INDEX IF NOT EXISTS idx_access_history_student ON access_change_history(student_id,created_at);
    CREATE INDEX IF NOT EXISTS idx_reset_tokens_user ON password_reset_tokens(user_id,expires_at);
    CREATE INDEX IF NOT EXISTS idx_attempts_student_created ON attempts(student_id,created_at);
    CREATE INDEX IF NOT EXISTS idx_attempt_answers_attempt_question ON attempt_answers(attempt_id,question_db_id);
    """)

    policies=[
      ('Foundation',0,1,10,65,120,None,0.20,0.50),
      ('Exam Ready',1,1,12,72,90,None,0.25,0.50),
      ('Advanced',2,2,15,76,75,None,0.30,0.60),
      ('Distinction',3,2,18,82,60,None,0.35,0.65),
      ('Expert',4,2,22,86,45,90,0.40,0.70),
      ('Elite',5,3,25,90,30,99,0.45,0.75)
    ]
    for row in policies:
        c.execute("""INSERT OR IGNORE INTO mastery_policies(
          mastery_level,level_rank,min_forms,min_questions,min_accuracy,verification_days,
          external_percentile_target,target_band_pct,unseen_family_pct) VALUES(?,?,?,?,?,?,?,?,?)""",row)

    # Four cumulative student access levels. Legacy Premium remains a Full Access entitlement.
    access_entitlements={
      'free_access': ('Free Access',0,'Foundation', {'diagnostic':True,'basic_results':True,'topic_practice':True,'live_study_plan':False}),
      'level_1_access': ('Level 1 Access',1,'Exam Ready', {'diagnostic':True,'basic_results':True,'topic_practice':True,'advanced_diagnostics':True,'live_study_plan':True}),
      'level_2_access': ('Level 2 Access',2,'Distinction', {'diagnostic':True,'basic_results':True,'topic_practice':True,'advanced_diagnostics':True,'recovery_engine':True,'live_study_plan':True,'all_mocks':True}),
      'full_access': ('Full Access',3,'Elite', {'diagnostic':True,'basic_results':True,'topic_practice':True,'advanced_diagnostics':True,'recovery_engine':True,'live_study_plan':True,'all_mocks':True,'monthly_challenges':True,'ranking_eligible':True,'premium_reports':True})
    }
    for code,(name,rank,ceiling,ent) in access_entitlements.items():
        c.execute("""INSERT OR IGNORE INTO plans(code,name,audience,billing_period,currency,price_minor,entitlements_json,active,sort_order,access_rank,mastery_ceiling)
          VALUES(?,?, 'student','none','PKR',NULL,?,1,?,?,?)""",(code,name,json.dumps(ent),100+rank,rank,ceiling))
        c.execute("UPDATE plans SET name=?,entitlements_json=?,active=1,access_rank=?,mastery_ceiling=? WHERE code=?",
                  (name,json.dumps(ent),rank,ceiling,code))
    c.execute("UPDATE plans SET active=0 WHERE code IN ('free_student','premium_monthly','premium_annual','institution_student')")
    c.execute("UPDATE plans SET access_rank=3,mastery_ceiling='Elite' WHERE code IN ('premium_monthly','premium_annual','institution_student')")
    c.execute("UPDATE plans SET access_rank=0,mastery_ceiling='Foundation' WHERE code='free_student'")

    # V5.3 plan-pathway names conflicted with mastery names; preserve evidence but rename future route labels.
    c.execute("UPDATE study_plans SET pathway='Stretch',title=replace(title,'Advanced','Stretch') WHERE pathway='Advanced'")
    c.execute("UPDATE study_plans SET pathway='Peak',title=replace(title,'Elite','Peak') WHERE pathway='Elite'")
    c.execute("UPDATE users SET study_plan_pathway='Stretch' WHERE study_plan_pathway='Advanced'")
    c.execute("UPDATE users SET study_plan_pathway='Peak' WHERE study_plan_pathway='Elite'")

    # Clearly-labelled software-test content. Demo items can exercise journeys but can never award production mastery.
    built_in_demo_ids=('BIO001','BIO002','BIO003','BIO004','BIO005')
    c.execute("UPDATE questions SET is_demo=1,calibration_status='DEMO' WHERE question_id IN (?,?,?,?,?)",built_in_demo_ids)
    demo_levels=['Foundation','Exam Ready','Advanced','Distinction','Expert','Elite']
    demo_subjects=['Biology','Chemistry','Physics']
    for subject in demo_subjects:
        prefix=subject[:3].upper()
        for i in range(1,31):
            level=demo_levels[(i-1)//5]
            chapter=f'Demo Chapter {(((i-1)//5)%3)+1}'
            qid=f'DEMO-{prefix}-{i:03d}'
            fam=f'DEMO-{prefix}-FAM-{i:03d}'
            c.execute("""INSERT OR IGNORE INTO questions(
              question_id,family_id,variant,programme,subject,chapter,topic,subtopic,qtype,level,question,
              option_a,option_b,option_c,option_d,answer,explanation,status,country,qualification,difficulty,
              review_status,source_type,active,is_demo,calibration_status)
              VALUES(?,?, 'A','FSc Part 1',?,?,?,?, 'MCQ',?,?, 'A','B','C','D','A',?,'Approved','Pakistan','FSc Part 1',?,'Approved','ScoreMax Demo',1,1,'DEMO')""",
              (qid,fam,subject,chapter,f'{level} workflow',f'{level} workflow',level,
               f'[DEMO] {subject} software-test item {i}. Select option A to test the ScoreMax journey.',
               'Demo-only item; not academic content.',level))
            if not c.execute("SELECT 1 FROM curriculum WHERE programme='FSc Part 1' AND subject=? AND chapter=? AND topic=? LIMIT 1",
                             (subject,chapter,f'{level} workflow')).fetchone():
                c.execute("INSERT INTO curriculum(programme,subject,chapter,topic,subtopic) VALUES('FSc Part 1',?,?,?,?)",
                          (subject,chapter,f'{level} workflow',f'{level} workflow'))

    # Demo rows are seeded after the first family backfill. Run a second idempotent
    # governance pass so local V5.4.2 testing keeps the full demo journeys available.
    for q in c.execute("""SELECT id,family_id,country,qualification,exam_board,curriculum_version,programme,subject,
      learning_outcome,concept,source_type,status,review_status,active,family_key FROM questions
      WHERE COALESCE(family_id,'')<>''""").fetchall():
        family_key=(q['family_key'] or '').strip() or canonical_family_key(
            q['family_id'],q['country'],q['qualification'],q['exam_board'],q['curriculum_version'],q['programme'],q['subject'])
        if not (q['family_key'] or '').strip():
            c.execute("UPDATE questions SET family_key=? WHERE id=?",(family_key,q['id']))
        live=(q['status']=='Approved' and q['review_status']=='Approved' and int(q['active'] or 0)==1)
        c.execute("""INSERT OR IGNORE INTO question_families(
          family_key,family_id,country,qualification,exam_board,curriculum_version,programme,subject,
          learning_outcome,concept,construct_signature,invariants_json,review_status,active,source_type)
          VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
          (family_key,q['family_id'],q['country'] or 'Pakistan',q['qualification'] or '',q['exam_board'] or '',q['curriculum_version'] or '',
           q['programme'] or '',q['subject'] or '',q['learning_outcome'] or '',q['concept'] or '',q['concept'] or '','[]',
           'Approved' if live else 'Draft',1 if live else 0,q['source_type'] or 'ScoreMax Original'))


    # V5.5 — Versioned Assessment Blueprint Integration and calibration-policy foundation.
    # Power House owns the authoritative blueprint. ScoreMax stores immutable snapshots
    # and applies them to authentic mocks, projections, planning and bank-readiness checks.
    c.executescript("""
    CREATE TABLE IF NOT EXISTS assessment_frameworks(
      id INTEGER PRIMARY KEY,
      powerhouse_framework_id TEXT NOT NULL UNIQUE,
      name TEXT NOT NULL,
      country TEXT DEFAULT '',
      authority TEXT DEFAULT '',
      active INTEGER DEFAULT 1,
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS assessment_framework_versions(
      id INTEGER PRIMARY KEY,
      framework_id INTEGER NOT NULL,
      powerhouse_framework_version_id TEXT NOT NULL,
      version_name TEXT NOT NULL,
      effective_from TEXT DEFAULT '',
      effective_to TEXT DEFAULT '',
      status TEXT DEFAULT 'ACTIVE',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      UNIQUE(framework_id,powerhouse_framework_version_id));

    CREATE TABLE IF NOT EXISTS assessment_blueprints(
      id INTEGER PRIMARY KEY,
      powerhouse_blueprint_id TEXT NOT NULL,
      framework_id INTEGER NOT NULL,
      framework_version_id INTEGER NOT NULL,
      blueprint_version TEXT NOT NULL,
      source_status TEXT NOT NULL,
      local_status TEXT NOT NULL DEFAULT 'IMPORTED',
      authority TEXT DEFAULT '',
      source_reference TEXT DEFAULT '',
      governance_note TEXT DEFAULT '',
      total_questions INTEGER NOT NULL,
      duration_minutes INTEGER,
      difficulty_distribution_json TEXT DEFAULT '{}',
      activation_date TEXT DEFAULT '',
      superseded_date TEXT DEFAULT '',
      source_created_at TEXT DEFAULT '',
      source_approved_at TEXT DEFAULT '',
      source_approved_by TEXT DEFAULT '',
      source_policy_version TEXT DEFAULT '',
      imported_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      imported_by INTEGER,
      payload_checksum TEXT NOT NULL,
      signature_status TEXT DEFAULT 'NOT_CONFIGURED',
      sync_status TEXT DEFAULT 'IMPORTED',
      validation_report_json TEXT DEFAULT '{}',
      immutable_payload_json TEXT NOT NULL,
      activated_at TEXT DEFAULT '',
      activated_by INTEGER,
      superseded_by_blueprint_id INTEGER,
      UNIQUE(powerhouse_blueprint_id,blueprint_version));

    CREATE TABLE IF NOT EXISTS assessment_blueprint_sections(
      id INTEGER PRIMARY KEY,
      blueprint_id INTEGER NOT NULL,
      section_order INTEGER NOT NULL,
      section_code TEXT DEFAULT '',
      section_title TEXT DEFAULT '',
      subject TEXT NOT NULL,
      question_count INTEGER NOT NULL,
      weight_percent REAL NOT NULL,
      duration_minutes INTEGER,
      difficulty_distribution_json TEXT DEFAULT '{}',
      rules_json TEXT DEFAULT '{}',
      UNIQUE(blueprint_id,subject));

    CREATE TABLE IF NOT EXISTS assessment_blueprint_sync_events(
      id INTEGER PRIMARY KEY,
      blueprint_id INTEGER,
      action TEXT NOT NULL,
      sync_status TEXT DEFAULT '',
      checksum TEXT DEFAULT '',
      message TEXT DEFAULT '',
      actor_user_id INTEGER,
      metadata_json TEXT DEFAULT '{}',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS assessment_blueprint_audit(
      id INTEGER PRIMARY KEY,
      blueprint_id INTEGER,
      action TEXT NOT NULL,
      actor_user_id INTEGER,
      previous_status TEXT DEFAULT '',
      new_status TEXT DEFAULT '',
      reason TEXT DEFAULT '',
      impact_snapshot_json TEXT DEFAULT '{}',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS assessment_assembly_policies(
      id INTEGER PRIMARY KEY,
      policy_code TEXT NOT NULL UNIQUE,
      policy_version TEXT NOT NULL,
      scope_type TEXT DEFAULT 'global',
      scope_key TEXT DEFAULT '',
      framework_version_id INTEGER,
      blueprint_id INTEGER,
      name TEXT NOT NULL,
      rigor_score INTEGER NOT NULL DEFAULT 50,
      mastery_standard_score INTEGER NOT NULL DEFAULT 50,
      selection_config_json TEXT DEFAULT '{}',
      evidence_config_json TEXT DEFAULT '{}',
      source TEXT DEFAULT 'ScoreMax Academic Governance',
      status TEXT DEFAULT 'DRAFT',
      effective_from TEXT DEFAULT '',
      superseded_at TEXT DEFAULT '',
      created_by INTEGER,
      approved_by INTEGER,
      approved_at TEXT DEFAULT '',
      reason TEXT DEFAULT '',
      preview_json TEXT DEFAULT '{}',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      UNIQUE(scope_type,scope_key,policy_version));

    CREATE TABLE IF NOT EXISTS assessment_policy_audit(
      id INTEGER PRIMARY KEY,
      policy_id INTEGER,
      action TEXT NOT NULL,
      actor_user_id INTEGER,
      previous_status TEXT DEFAULT '',
      new_status TEXT DEFAULT '',
      reason TEXT DEFAULT '',
      snapshot_json TEXT DEFAULT '{}',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS content_requirement_requests(
      id INTEGER PRIMARY KEY,
      request_code TEXT NOT NULL UNIQUE,
      framework_version_id INTEGER,
      blueprint_id INTEGER,
      subject TEXT NOT NULL,
      chapter TEXT DEFAULT '',
      learning_outcome TEXT DEFAULT '',
      mastery_level TEXT DEFAULT '',
      difficulty TEXT DEFAULT '',
      assets_required INTEGER NOT NULL DEFAULT 0,
      families_required INTEGER NOT NULL DEFAULT 0,
      intended_use TEXT DEFAULT 'authentic_mock',
      priority TEXT DEFAULT 'High',
      deadline TEXT DEFAULT '',
      reason TEXT DEFAULT '',
      existing_bank_evidence_json TEXT DEFAULT '{}',
      status TEXT DEFAULT 'OPEN',
      created_by INTEGER,
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS student_blueprint_projections(
      id INTEGER PRIMARY KEY,
      student_id INTEGER NOT NULL,
      blueprint_id INTEGER NOT NULL,
      blueprint_version TEXT NOT NULL,
      framework_version TEXT DEFAULT '',
      projection_json TEXT NOT NULL,
      confidence_label TEXT DEFAULT 'Low',
      evidence_sufficiency REAL DEFAULT 0,
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE INDEX IF NOT EXISTS idx_assessment_blueprints_status ON assessment_blueprints(framework_version_id,local_status,activation_date);
    CREATE INDEX IF NOT EXISTS idx_blueprint_sections_blueprint ON assessment_blueprint_sections(blueprint_id,section_order);
    CREATE INDEX IF NOT EXISTS idx_blueprint_sync_blueprint ON assessment_blueprint_sync_events(blueprint_id,created_at);
    CREATE INDEX IF NOT EXISTS idx_blueprint_audit_blueprint ON assessment_blueprint_audit(blueprint_id,created_at);
    CREATE INDEX IF NOT EXISTS idx_assembly_policy_active ON assessment_assembly_policies(status,scope_type,scope_key);
    CREATE INDEX IF NOT EXISTS idx_policy_audit_policy ON assessment_policy_audit(policy_id,created_at);
    CREATE INDEX IF NOT EXISTS idx_content_requests_status ON content_requirement_requests(status,blueprint_id,subject);
    CREATE INDEX IF NOT EXISTS idx_student_blueprint_projection ON student_blueprint_projections(student_id,blueprint_id,created_at);
    """)

    for name, definition in {
        'rights_status': "TEXT DEFAULT 'ScoreMax Original'",
        'scoremax_ready': "INTEGER DEFAULT 1",
        'assessment_purpose': "TEXT DEFAULT 'practice|test|mock|mastery'",
        'empirical_difficulty': "REAL",
        'difficulty_source': "TEXT DEFAULT 'authoring'"
    }.items():
        ensure_column(c,'questions',name,definition)

    blueprint_pinned_columns={
        'assessment_blueprint_id': "INTEGER",
        'blueprint_source_id': "TEXT DEFAULT ''",
        'blueprint_version': "TEXT DEFAULT ''",
        'framework_version': "TEXT DEFAULT ''",
        'blueprint_snapshot_json': "TEXT DEFAULT '{}'",
        'assembly_policy_id': "INTEGER",
        'assembly_policy_version': "TEXT DEFAULT ''"
    }
    for table in ('assessment_sessions','attempts'):
        for name, definition in blueprint_pinned_columns.items():
            ensure_column(c,table,name,definition)
    ensure_column(c,'mastery_form_results','assembly_policy_id','INTEGER')
    ensure_column(c,'mastery_form_results','assembly_policy_version',"TEXT DEFAULT ''")
    ensure_column(c,'mastery_form_results','effective_policy_json',"TEXT DEFAULT '{}'")

    for name, definition in {
        **blueprint_pinned_columns,
        'authenticity_status': "TEXT DEFAULT 'LEGACY_UNPINNED'",
        'preflight_json': "TEXT DEFAULT '{}'",
        'generated_at': "TEXT DEFAULT ''"
    }.items():
        ensure_column(c,'exam_papers',name,definition)

    for name, definition in {
        'assessment_blueprint_id': "INTEGER",
        'blueprint_version': "TEXT DEFAULT ''",
        'framework_version': "TEXT DEFAULT ''",
        'subject_priority_snapshot_json': "TEXT DEFAULT '{}'"
    }.items():
        ensure_column(c,'study_plans',name,definition)

    c.execute("""INSERT OR IGNORE INTO assessment_assembly_policies(
      policy_code,policy_version,scope_type,scope_key,name,rigor_score,mastery_standard_score,
      selection_config_json,evidence_config_json,status,source,reason,approved_at)
      VALUES('SMX-RIGOR-BASELINE','1','global','','ScoreMax Baseline Rigor',50,50,
      '{"unseen_family_ratio":0.70,"duplicate_family_limit":1,"calibrated_preference":true}',
      '{"historical_results_immutable":true,"policy_tightening_action":"Verification Due"}',
      'ACTIVE','ScoreMax Academic Governance','Baseline versioned policy',CURRENT_TIMESTAMP)""")

    # Existing papers pre-date authoritative blueprint pinning. Never fabricate that
    # they followed a later Power House blueprint.
    c.execute("""UPDATE exam_papers SET authenticity_status='LEGACY_UNPINNED'
      WHERE COALESCE(assessment_blueprint_id,0)=0 AND COALESCE(authenticity_status,'')=''""")


# ---------------------------------------------------------------------------
# V6.0 — Written Response Intelligence
# ---------------------------------------------------------------------------

def migrate_v6(c):
    """Idempotent V6 schema. Power House packages are immutable; student evidence is separate."""
    ensure_column(c,'users','written_pilot_enabled','INTEGER DEFAULT 0')
    c.executescript("""
    CREATE TABLE IF NOT EXISTS written_feature_controls(
      feature_code TEXT PRIMARY KEY,
      name TEXT NOT NULL,
      state TEXT NOT NULL DEFAULT 'HIDDEN',
      required_access_code TEXT DEFAULT 'full_access',
      available_from TEXT DEFAULT '',
      available_to TEXT DEFAULT '',
      configuration_json TEXT DEFAULT '{}',
      updated_by INTEGER,
      updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS written_assessment_packages(
      id INTEGER PRIMARY KEY,
      assessment_package_id TEXT NOT NULL,
      assessment_package_version TEXT NOT NULL,
      framework_id TEXT NOT NULL,
      framework_version_id TEXT NOT NULL,
      blueprint_snapshot_id TEXT DEFAULT '',
      subject_id TEXT NOT NULL,
      chapter_id TEXT NOT NULL,
      academic_approval_status TEXT NOT NULL,
      local_status TEXT NOT NULL DEFAULT 'IMPORTED',
      approved_at TEXT DEFAULT '',
      source_approved_by TEXT DEFAULT '',
      export_checksum TEXT NOT NULL,
      immutable_payload_json TEXT NOT NULL,
      imported_by INTEGER,
      imported_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      activated_by INTEGER,
      activated_at TEXT DEFAULT '',
      UNIQUE(assessment_package_id,assessment_package_version));

    CREATE TABLE IF NOT EXISTS written_questions(
      id INTEGER PRIMARY KEY,
      package_id INTEGER NOT NULL,
      question_source_id TEXT NOT NULL,
      question_family_id TEXT NOT NULL,
      variant_id TEXT DEFAULT '',
      question_type TEXT NOT NULL,
      question_text TEXT NOT NULL,
      command_verb TEXT NOT NULL,
      maximum_marks REAL NOT NULL,
      estimated_time INTEGER,
      difficulty TEXT DEFAULT '',
      cognitive_demand TEXT DEFAULT '',
      mastery_level TEXT DEFAULT '',
      purpose TEXT DEFAULT 'practice',
      question_json TEXT NOT NULL,
      active INTEGER DEFAULT 1,
      UNIQUE(package_id,question_source_id));

    CREATE TABLE IF NOT EXISTS written_attempts(
      id INTEGER PRIMARY KEY,
      student_id INTEGER NOT NULL,
      written_question_id INTEGER NOT NULL,
      package_id INTEGER NOT NULL,
      parent_attempt_id INTEGER,
      attempt_mode TEXT NOT NULL DEFAULT 'practice',
      entry_method TEXT NOT NULL DEFAULT 'typed',
      evidence_type TEXT NOT NULL DEFAULT 'independent_production',
      support_level TEXT DEFAULT 'independent',
      novelty_status TEXT DEFAULT 'seen_family',
      status TEXT NOT NULL DEFAULT 'DRAFT',
      current_mark REAL,
      maximum_mark REAL,
      marking_confidence REAL,
      result_state TEXT DEFAULT '',
      package_version TEXT NOT NULL,
      rubric_version TEXT DEFAULT '',
      mastery_policy_version TEXT DEFAULT '',
      rigor_policy_version TEXT DEFAULT '',
      grader_a_version TEXT DEFAULT '',
      grader_b_version TEXT DEFAULT '',
      reconciliation_policy_version TEXT DEFAULT '',
      original_submitted_at TEXT DEFAULT '',
      completed_at TEXT DEFAULT '',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS written_answer_versions(
      id INTEGER PRIMARY KEY,
      attempt_id INTEGER NOT NULL,
      version_no INTEGER NOT NULL,
      version_type TEXT NOT NULL,
      answer_text TEXT DEFAULT '',
      original_ocr_text TEXT DEFAULT '',
      confirmed_transcript TEXT DEFAULT '',
      correction_log_json TEXT DEFAULT '[]',
      word_count INTEGER DEFAULT 0,
      is_frozen INTEGER DEFAULT 0,
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      UNIQUE(attempt_id,version_no));

    CREATE TABLE IF NOT EXISTS written_marking_runs(
      id INTEGER PRIMARY KEY,
      attempt_id INTEGER NOT NULL,
      answer_version_id INTEGER NOT NULL,
      proposed_mark REAL NOT NULL,
      maximum_mark REAL NOT NULL,
      percentage REAL NOT NULL,
      confidence REAL NOT NULL,
      result_state TEXT NOT NULL,
      command_verb_met INTEGER DEFAULT 1,
      grader_a_json TEXT DEFAULT '{}',
      grader_b_json TEXT DEFAULT '{}',
      reconciliation_json TEXT DEFAULT '{}',
      feedback_json TEXT DEFAULT '[]',
      contradictions_json TEXT DEFAULT '[]',
      misconceptions_json TEXT DEFAULT '[]',
      validation_boundary TEXT DEFAULT '',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS written_mark_point_results(
      id INTEGER PRIMARY KEY,
      marking_run_id INTEGER NOT NULL,
      point_id TEXT NOT NULL,
      description TEXT NOT NULL,
      available_marks REAL NOT NULL,
      awarded_marks REAL NOT NULL,
      status TEXT NOT NULL,
      evidence_json TEXT DEFAULT '[]',
      improvement_instruction TEXT DEFAULT '',
      grader_a_status TEXT DEFAULT '',
      grader_b_status TEXT DEFAULT '');

    CREATE TABLE IF NOT EXISTS written_mastery_evidence(
      id INTEGER PRIMARY KEY,
      student_id INTEGER NOT NULL,
      attempt_id INTEGER NOT NULL,
      package_id INTEGER NOT NULL,
      framework_id TEXT DEFAULT '',
      framework_version_id TEXT DEFAULT '',
      subject_id TEXT DEFAULT '',
      chapter_id TEXT DEFAULT '',
      learning_outcome_ids_json TEXT DEFAULT '[]',
      concept_ids_json TEXT DEFAULT '[]',
      proposition_ids_json TEXT DEFAULT '[]',
      command_verb TEXT DEFAULT '',
      cognitive_demand TEXT DEFAULT '',
      evidence_level TEXT NOT NULL,
      support_level TEXT DEFAULT '',
      novelty_status TEXT DEFAULT '',
      score_percentage REAL,
      confidence REAL,
      evidence_status TEXT DEFAULT 'PROVISIONAL',
      evidence_json TEXT DEFAULT '{}',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS written_recovery_tasks(
      id INTEGER PRIMARY KEY,
      student_id INTEGER NOT NULL,
      attempt_id INTEGER NOT NULL,
      task_type TEXT NOT NULL,
      title TEXT NOT NULL,
      reason TEXT DEFAULT '',
      approved_activity_id TEXT DEFAULT '',
      status TEXT DEFAULT 'PLANNED',
      scheduled_for TEXT DEFAULT '',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      completed_at TEXT DEFAULT '');

    CREATE TABLE IF NOT EXISTS written_upload_pages(
      id INTEGER PRIMARY KEY,
      attempt_id INTEGER NOT NULL,
      page_no INTEGER NOT NULL,
      storage_path TEXT NOT NULL,
      original_filename TEXT DEFAULT '',
      mime_type TEXT DEFAULT '',
      width INTEGER DEFAULT 0,
      height INTEGER DEFAULT 0,
      quality_score REAL DEFAULT 0,
      quality_status TEXT DEFAULT '',
      quality_json TEXT DEFAULT '{}',
      original_file_hash TEXT DEFAULT '',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      UNIQUE(attempt_id,page_no));

    CREATE TABLE IF NOT EXISTS written_processing_jobs(
      id INTEGER PRIMARY KEY,
      attempt_id INTEGER NOT NULL,
      job_type TEXT NOT NULL,
      state TEXT NOT NULL,
      provider TEXT DEFAULT '',
      provider_version TEXT DEFAULT '',
      retry_count INTEGER DEFAULT 0,
      idempotency_key TEXT NOT NULL UNIQUE,
      input_json TEXT DEFAULT '{}',
      output_json TEXT DEFAULT '{}',
      error_message TEXT DEFAULT '',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS written_exemplar_candidates(
      id INTEGER PRIMARY KEY,
      attempt_id INTEGER NOT NULL UNIQUE,
      student_id INTEGER NOT NULL,
      answer_version_id INTEGER NOT NULL,
      eligibility_reason TEXT NOT NULL,
      academic_status TEXT DEFAULT 'PENDING_REVIEW',
      academic_reviewer_id INTEGER,
      academic_reviewed_at TEXT DEFAULT '',
      academic_note TEXT DEFAULT '',
      consent_status TEXT DEFAULT 'PENDING',
      exemplar_type TEXT DEFAULT 'INDEPENDENT_FULL_MARK',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS written_exemplar_consents(
      id INTEGER PRIMARY KEY,
      candidate_id INTEGER NOT NULL,
      student_id INTEGER NOT NULL,
      consent_status TEXT NOT NULL,
      attribution_preference TEXT DEFAULT 'ANONYMOUS',
      guardian_confirmation INTEGER DEFAULT 0,
      consent_text_version TEXT NOT NULL,
      consented_at TEXT DEFAULT '',
      withdrawn_at TEXT DEFAULT '',
      UNIQUE(candidate_id,student_id));

    CREATE TABLE IF NOT EXISTS written_exemplars(
      id INTEGER PRIMARY KEY,
      candidate_id INTEGER NOT NULL UNIQUE,
      written_question_id INTEGER NOT NULL,
      package_id INTEGER NOT NULL,
      answer_version_id INTEGER NOT NULL,
      exemplar_type TEXT NOT NULL,
      display_name TEXT DEFAULT 'Anonymous student',
      publication_status TEXT DEFAULT 'APPROVED_HIDDEN',
      approved_by INTEGER NOT NULL,
      approved_at TEXT NOT NULL,
      published_at TEXT DEFAULT '',
      package_version TEXT NOT NULL,
      rubric_version TEXT DEFAULT '',
      marking_policy_version TEXT DEFAULT '',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS written_usage_ledger(
      id INTEGER PRIMARY KEY,
      student_id INTEGER NOT NULL,
      attempt_id INTEGER,
      operation TEXT NOT NULL,
      provider TEXT DEFAULT 'local',
      units REAL DEFAULT 1,
      estimated_cost_minor INTEGER DEFAULT 0,
      metadata_json TEXT DEFAULT '{}',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE INDEX IF NOT EXISTS idx_written_questions_package ON written_questions(package_id,active);
    CREATE INDEX IF NOT EXISTS idx_written_attempts_student ON written_attempts(student_id,created_at);
    CREATE INDEX IF NOT EXISTS idx_written_marking_attempt ON written_marking_runs(attempt_id,created_at);
    CREATE INDEX IF NOT EXISTS idx_written_evidence_student ON written_mastery_evidence(student_id,subject_id,chapter_id);
    """)
    controls=[
      ('written_response_engine','Written Response Intelligence','PILOT','full_access',{'typed_answers':True,'routine_human_marking':False}),
      ('written_handwriting','Handwriting and OCR','PILOT','full_access',{'provider_required_for_live':True,'local_admin_simulation':True}),
      ('written_build_answer','Build the Answer','PILOT','level_2_access',{'powerhouse_approved_scaffolds_only':True}),
      ('written_exemplar_library','Approved Student Exemplar Library','HIDDEN','full_access',{'perfect_score_only':True,'academic_approval':True,'explicit_opt_in':True})]
    for code,name,state,access,config in controls:
        c.execute("""INSERT OR IGNORE INTO written_feature_controls(feature_code,name,state,required_access_code,configuration_json)
          VALUES(?,?,?,?,?)""",(code,name,state,access,json.dumps(config)))



# ---------------------------------------------------------------------------
# V6.1 — Teacher Discovery & Academic Messages
# ---------------------------------------------------------------------------

def migrate_v6_1(c):
    """Idempotent marketplace/messaging schema with privacy and safeguarding boundaries."""
    ensure_column(c,'users','teacher_marketplace_pilot_enabled','INTEGER DEFAULT 0')
    ensure_column(c,'users','academic_messages_pilot_enabled','INTEGER DEFAULT 0')
    c.executescript("""
    CREATE TABLE IF NOT EXISTS community_feature_controls(
      feature_code TEXT PRIMARY KEY,
      name TEXT NOT NULL,
      state TEXT NOT NULL DEFAULT 'HIDDEN',
      configuration_json TEXT DEFAULT '{}',
      updated_by INTEGER,
      updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS community_user_agreements(
      id INTEGER PRIMARY KEY,
      user_id INTEGER NOT NULL,
      agreement_code TEXT NOT NULL,
      agreement_version TEXT NOT NULL,
      status TEXT NOT NULL DEFAULT 'ACCEPTED',
      accepted_at TEXT DEFAULT '',
      revoked_at TEXT DEFAULT '',
      ip_evidence TEXT DEFAULT '',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      UNIQUE(user_id,agreement_code,agreement_version));

    CREATE TABLE IF NOT EXISTS teacher_profiles(
      id INTEGER PRIMARY KEY,
      teacher_id INTEGER NOT NULL UNIQUE,
      headline TEXT DEFAULT '',
      bio TEXT DEFAULT '',
      subjects_json TEXT DEFAULT '[]',
      frameworks_json TEXT DEFAULT '[]',
      qualifications_text TEXT DEFAULT '',
      experience_years INTEGER DEFAULT 0,
      languages_json TEXT DEFAULT '[]',
      delivery_modes_json TEXT DEFAULT '[]',
      platforms_json TEXT DEFAULT '[]',
      location_text TEXT DEFAULT '',
      price_from_minor INTEGER DEFAULT 0,
      currency TEXT DEFAULT 'PKR',
      availability_text TEXT DEFAULT '',
      response_expectation_hours INTEGER DEFAULT 24,
      office_hours TEXT DEFAULT '',
      intro_video_url TEXT DEFAULT '',
      allow_one_to_one INTEGER DEFAULT 1,
      allow_groups INTEGER DEFAULT 1,
      profile_status TEXT DEFAULT 'DRAFT',
      identity_verification_status TEXT DEFAULT 'UNVERIFIED',
      qualification_verification_status TEXT DEFAULT 'UNVERIFIED',
      experience_verification_status TEXT DEFAULT 'UNVERIFIED',
      moderation_note TEXT DEFAULT '',
      moderated_by INTEGER,
      moderated_at TEXT DEFAULT '',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS teacher_verification_events(
      id INTEGER PRIMARY KEY,
      teacher_id INTEGER NOT NULL,
      verification_type TEXT NOT NULL,
      previous_status TEXT DEFAULT '',
      new_status TEXT NOT NULL,
      reviewer_id INTEGER NOT NULL,
      evidence_note TEXT DEFAULT '',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS teacher_service_listings(
      id INTEGER PRIMARY KEY,
      teacher_id INTEGER NOT NULL,
      profile_id INTEGER NOT NULL,
      service_type TEXT NOT NULL,
      title TEXT NOT NULL,
      description TEXT DEFAULT '',
      subject TEXT NOT NULL,
      framework TEXT DEFAULT '',
      chapter_scope TEXT DEFAULT '',
      delivery_mode TEXT DEFAULT 'ONLINE',
      platform_options_json TEXT DEFAULT '[]',
      price_minor INTEGER DEFAULT 0,
      currency TEXT DEFAULT 'PKR',
      pricing_unit TEXT DEFAULT 'PER_SESSION',
      capacity INTEGER DEFAULT 1,
      availability_text TEXT DEFAULT '',
      listing_status TEXT DEFAULT 'DRAFT',
      academic_endorsement_status TEXT DEFAULT 'NOT_SCOREMAX_ENDORSED',
      moderation_note TEXT DEFAULT '',
      moderated_by INTEGER,
      moderated_at TEXT DEFAULT '',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS teacher_enquiries(
      id INTEGER PRIMARY KEY,
      student_id INTEGER NOT NULL,
      teacher_id INTEGER NOT NULL,
      listing_id INTEGER,
      subject TEXT DEFAULT '',
      framework TEXT DEFAULT '',
      chapter TEXT DEFAULT '',
      support_need TEXT NOT NULL,
      preferred_mode TEXT DEFAULT '',
      initial_message TEXT DEFAULT '',
      status TEXT DEFAULT 'PENDING',
      academic_context_json TEXT DEFAULT '{}',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      responded_at TEXT DEFAULT '',
      accepted_at TEXT DEFAULT '',
      closed_at TEXT DEFAULT '');

    CREATE TABLE IF NOT EXISTS academic_groups(
      id INTEGER PRIMARY KEY,
      teacher_id INTEGER NOT NULL,
      listing_id INTEGER,
      name TEXT NOT NULL,
      subject TEXT NOT NULL,
      framework TEXT DEFAULT '',
      description TEXT DEFAULT '',
      group_type TEXT DEFAULT 'TUITION',
      posting_policy TEXT DEFAULT 'TEACHER_ONLY',
      join_policy TEXT DEFAULT 'REQUEST',
      max_members INTEGER DEFAULT 30,
      status TEXT DEFAULT 'DRAFT',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS academic_group_members(
      group_id INTEGER NOT NULL,
      user_id INTEGER NOT NULL,
      member_role TEXT NOT NULL,
      status TEXT DEFAULT 'PENDING',
      joined_at TEXT DEFAULT '',
      approved_by INTEGER,
      PRIMARY KEY(group_id,user_id));

    CREATE TABLE IF NOT EXISTS academic_conversations(
      id INTEGER PRIMARY KEY,
      conversation_code TEXT NOT NULL UNIQUE,
      conversation_type TEXT NOT NULL,
      teacher_id INTEGER NOT NULL,
      student_id INTEGER,
      enquiry_id INTEGER,
      group_id INTEGER,
      status TEXT DEFAULT 'ACTIVE',
      purpose TEXT DEFAULT '',
      academic_context_json TEXT DEFAULT '{}',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      archived_at TEXT DEFAULT '');

    CREATE TABLE IF NOT EXISTS academic_conversation_members(
      conversation_id INTEGER NOT NULL,
      user_id INTEGER NOT NULL,
      member_role TEXT NOT NULL,
      status TEXT DEFAULT 'ACTIVE',
      joined_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      last_read_message_id INTEGER,
      notification_preference TEXT DEFAULT 'ALL',
      PRIMARY KEY(conversation_id,user_id));

    CREATE TABLE IF NOT EXISTS academic_messages(
      id INTEGER PRIMARY KEY,
      conversation_id INTEGER NOT NULL,
      sender_id INTEGER NOT NULL,
      message_type TEXT DEFAULT 'TEXT',
      body TEXT NOT NULL,
      attachment_json TEXT DEFAULT '{}',
      moderation_status TEXT DEFAULT 'VISIBLE',
      policy_flags_json TEXT DEFAULT '[]',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      removed_at TEXT DEFAULT '',
      removed_by INTEGER,
      removal_reason TEXT DEFAULT '');

    CREATE TABLE IF NOT EXISTS academic_message_reports(
      id INTEGER PRIMARY KEY,
      reporter_id INTEGER NOT NULL,
      conversation_id INTEGER NOT NULL,
      message_id INTEGER,
      reported_user_id INTEGER,
      category TEXT NOT NULL,
      detail TEXT DEFAULT '',
      status TEXT DEFAULT 'OPEN',
      resolution TEXT DEFAULT '',
      resolved_by INTEGER,
      resolved_at TEXT DEFAULT '',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS academic_user_blocks(
      blocker_id INTEGER NOT NULL,
      blocked_id INTEGER NOT NULL,
      active INTEGER DEFAULT 1,
      reason TEXT DEFAULT '',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      PRIMARY KEY(blocker_id,blocked_id));

    CREATE TABLE IF NOT EXISTS academic_guardian_consents(
      id INTEGER PRIMARY KEY,
      student_id INTEGER NOT NULL,
      parent_user_id INTEGER NOT NULL,
      consent_scope TEXT NOT NULL DEFAULT 'TEACHER_DISCOVERY_AND_MESSAGES',
      status TEXT NOT NULL DEFAULT 'APPROVED',
      consent_text_version TEXT NOT NULL DEFAULT 'V6.1-GUARDIAN-1',
      granted_at TEXT DEFAULT '',
      revoked_at TEXT DEFAULT '',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      UNIQUE(student_id,parent_user_id,consent_scope));

    CREATE TABLE IF NOT EXISTS teacher_engagements(
      id INTEGER PRIMARY KEY,
      enquiry_id INTEGER,
      conversation_id INTEGER NOT NULL,
      student_id INTEGER NOT NULL,
      teacher_id INTEGER NOT NULL,
      service_type TEXT DEFAULT 'ONE_TO_ONE',
      teacher_confirmed INTEGER DEFAULT 0,
      student_confirmed INTEGER DEFAULT 0,
      status TEXT DEFAULT 'ACTIVE',
      session_date TEXT DEFAULT '',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      completed_at TEXT DEFAULT '',
      UNIQUE(conversation_id,student_id,teacher_id));

    CREATE TABLE IF NOT EXISTS teacher_reviews(
      id INTEGER PRIMARY KEY,
      engagement_id INTEGER NOT NULL UNIQUE,
      student_id INTEGER NOT NULL,
      teacher_id INTEGER NOT NULL,
      rating INTEGER NOT NULL,
      review_text TEXT DEFAULT '',
      moderation_status TEXT DEFAULT 'PENDING',
      policy_flags_json TEXT DEFAULT '[]',
      moderated_by INTEGER,
      moderated_at TEXT DEFAULT '',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE INDEX IF NOT EXISTS idx_teacher_profiles_status ON teacher_profiles(profile_status,teacher_id);
    CREATE INDEX IF NOT EXISTS idx_teacher_listings_search ON teacher_service_listings(listing_status,subject,framework,service_type);
    CREATE INDEX IF NOT EXISTS idx_teacher_enquiries_teacher ON teacher_enquiries(teacher_id,status,created_at);
    CREATE INDEX IF NOT EXISTS idx_teacher_enquiries_student ON teacher_enquiries(student_id,status,created_at);
    CREATE INDEX IF NOT EXISTS idx_academic_members_user ON academic_conversation_members(user_id,status,conversation_id);
    CREATE INDEX IF NOT EXISTS idx_academic_messages_conversation ON academic_messages(conversation_id,id);
    CREATE INDEX IF NOT EXISTS idx_academic_reports_status ON academic_message_reports(status,created_at);
    CREATE INDEX IF NOT EXISTS idx_teacher_reviews_teacher ON teacher_reviews(teacher_id,moderation_status,created_at);
    """)
    ensure_column(c,'teacher_enquiries','guardian_consent_status',"TEXT DEFAULT 'NOT_REQUIRED'")
    ensure_column(c,'teacher_enquiries','guardian_parent_user_id','INTEGER')
    controls=[
      ('teacher_discovery','Teacher Discovery','PILOT',{'authenticated_only':True,'professional_boundaries':True}),
      ('academic_messages','Academic Messages','PILOT',{'accepted_relationship_required':True,'phone_numbers_hidden':True,'student_student_dm':False}),
      ('teacher_group_channels','Teacher-led Group Channels','PILOT',{'teacher_owned':True,'student_student_dm':False}),
      ('student_direct_messages','Student Direct Messages','HIDDEN',{'launch_blocked':True,'reason':'Safeguarding and moderation must be proven first'})]
    for code,name,state,config in controls:
        c.execute("""INSERT OR IGNORE INTO community_feature_controls(feature_code,name,state,configuration_json)
          VALUES(?,?,?,?)""",(code,name,state,json.dumps(config,sort_keys=True)))



def migrate_v6_2(c):
    """Idempotent V6.2 pilot-readiness, persistent content intake and Knowledge Hub schema."""
    ensure_column(c,'users','is_demo_account','INTEGER DEFAULT 0')
    ensure_column(c,'questions','source_import_batch_id','INTEGER')
    ensure_column(c,'questions','content_environment',"TEXT DEFAULT 'CANDIDATE'")
    ensure_column(c,'question_families','source_import_batch_id','INTEGER')
    ensure_column(c,'written_assessment_packages','source_import_batch_id','INTEGER')
    c.executescript("""
    CREATE TABLE IF NOT EXISTS pilot_feature_controls(
      feature_code TEXT PRIMARY KEY,
      name TEXT NOT NULL,
      state TEXT NOT NULL DEFAULT 'PILOT',
      configuration_json TEXT DEFAULT '{}',
      updated_by INTEGER,
      updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS powerhouse_prompt_packs(
      id INTEGER PRIMARY KEY,
      prompt_pack_id TEXT NOT NULL,
      prompt_pack_version TEXT NOT NULL,
      source_status TEXT NOT NULL,
      local_status TEXT NOT NULL DEFAULT 'IMPORTED',
      framework TEXT DEFAULT '',
      framework_version TEXT DEFAULT '',
      subject TEXT DEFAULT '',
      chapter TEXT DEFAULT '',
      learning_outcome_ids_json TEXT DEFAULT '[]',
      source_evidence_ids_json TEXT DEFAULT '[]',
      prompt_text TEXT NOT NULL,
      expected_output_schema_json TEXT DEFAULT '{}',
      immutable_payload_json TEXT NOT NULL,
      payload_checksum TEXT NOT NULL,
      imported_by INTEGER,
      imported_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      copied_count INTEGER DEFAULT 0,
      last_copied_at TEXT DEFAULT '',
      UNIQUE(prompt_pack_id,prompt_pack_version));

    CREATE TABLE IF NOT EXISTS powerhouse_generation_batches(
      id INTEGER PRIMARY KEY,
      prompt_pack_db_id INTEGER NOT NULL,
      provider TEXT NOT NULL,
      model TEXT DEFAULT '',
      provider_run_id TEXT DEFAULT '',
      raw_output TEXT NOT NULL,
      parsed_output_json TEXT DEFAULT '{}',
      validation_status TEXT NOT NULL DEFAULT 'PENDING',
      validation_report_json TEXT DEFAULT '{}',
      item_count INTEGER DEFAULT 0,
      submitted_by INTEGER,
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      export_status TEXT DEFAULT 'NOT_EXPORTED',
      exported_at TEXT DEFAULT '');

    CREATE TABLE IF NOT EXISTS content_import_batches(
      id INTEGER PRIMARY KEY,
      batch_code TEXT UNIQUE NOT NULL,
      source_system TEXT DEFAULT 'MANUAL_FILE',
      source_prompt_pack_id TEXT DEFAULT '',
      source_prompt_pack_version TEXT DEFAULT '',
      filename TEXT NOT NULL,
      file_type TEXT NOT NULL,
      payload_checksum TEXT NOT NULL,
      row_count INTEGER DEFAULT 0,
      valid_count INTEGER DEFAULT 0,
      error_count INTEGER DEFAULT 0,
      warning_count INTEGER DEFAULT 0,
      status TEXT NOT NULL DEFAULT 'PREVIEWED',
      validation_report_json TEXT DEFAULT '{}',
      compatibility_before_json TEXT DEFAULT '{}',
      compatibility_after_json TEXT DEFAULT '{}',
      backup_record_id INTEGER,
      imported_by INTEGER,
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      confirmed_at TEXT DEFAULT '',
      rolled_back_at TEXT DEFAULT '',
      rolled_back_by INTEGER,
      rollback_note TEXT DEFAULT '');

    CREATE TABLE IF NOT EXISTS content_import_batch_rows(
      id INTEGER PRIMARY KEY,
      batch_id INTEGER NOT NULL,
      row_number INTEGER NOT NULL,
      sheet_name TEXT DEFAULT '',
      question_id TEXT DEFAULT '',
      row_json TEXT NOT NULL,
      errors_json TEXT DEFAULT '[]',
      warnings_json TEXT DEFAULT '[]',
      import_status TEXT DEFAULT 'PENDING',
      question_db_id INTEGER,
      UNIQUE(batch_id,row_number,sheet_name));

    CREATE TABLE IF NOT EXISTS pilot_backups(
      id INTEGER PRIMARY KEY,
      backup_code TEXT UNIQUE NOT NULL,
      reason TEXT NOT NULL,
      file_path TEXT NOT NULL,
      file_sha256 TEXT DEFAULT '',
      file_size_bytes INTEGER DEFAULT 0,
      integrity_status TEXT DEFAULT 'PENDING',
      created_by INTEGER,
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS pilot_feedback(
      id INTEGER PRIMARY KEY,
      feedback_code TEXT UNIQUE NOT NULL,
      reporter_user_id INTEGER NOT NULL,
      category TEXT NOT NULL,
      severity TEXT DEFAULT 'MEDIUM',
      description TEXT NOT NULL,
      question_id INTEGER,
      attempt_id INTEGER,
      written_attempt_id INTEGER,
      blueprint_id INTEGER,
      assembly_policy_id INTEGER,
      screenshot_path TEXT DEFAULT '',
      routing_target TEXT NOT NULL,
      status TEXT DEFAULT 'OPEN',
      assigned_to INTEGER,
      resolution TEXT DEFAULT '',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      resolved_at TEXT DEFAULT '');

    CREATE TABLE IF NOT EXISTS pilot_activity_events(
      id INTEGER PRIMARY KEY,
      event_type TEXT NOT NULL,
      actor_user_id INTEGER,
      subject_type TEXT DEFAULT '',
      subject_id TEXT DEFAULT '',
      metadata_json TEXT DEFAULT '{}',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS demo_cleanup_runs(
      id INTEGER PRIMARY KEY,
      run_code TEXT UNIQUE NOT NULL,
      backup_record_id INTEGER,
      demo_users_count INTEGER DEFAULT 0,
      demo_questions_count INTEGER DEFAULT 0,
      demo_attempts_count INTEGER DEFAULT 0,
      status TEXT DEFAULT 'PREVIEWED',
      report_json TEXT DEFAULT '{}',
      executed_by INTEGER,
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      executed_at TEXT DEFAULT '');

    CREATE TABLE IF NOT EXISTS knowledge_feature_controls(
      feature_code TEXT PRIMARY KEY,
      state TEXT DEFAULT 'HIDDEN',
      configuration_json TEXT DEFAULT '{}',
      updated_by INTEGER,
      updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS knowledge_articles(
      id INTEGER PRIMARY KEY,
      slug TEXT UNIQUE NOT NULL,
      title TEXT NOT NULL,
      summary TEXT DEFAULT '',
      body_text TEXT NOT NULL,
      author_name TEXT DEFAULT '',
      source_origin TEXT DEFAULT 'MANUAL',
      framework TEXT DEFAULT '',
      framework_version TEXT DEFAULT '',
      subject TEXT DEFAULT '',
      chapter TEXT DEFAULT '',
      status TEXT DEFAULT 'DRAFT',
      seo_title TEXT DEFAULT '',
      seo_description TEXT DEFAULT '',
      applicable_from TEXT DEFAULT '',
      applicable_to TEXT DEFAULT '',
      reviewed_at TEXT DEFAULT '',
      reviewed_by INTEGER,
      published_at TEXT DEFAULT '',
      created_by INTEGER,
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS knowledge_sources(
      id INTEGER PRIMARY KEY,
      article_id INTEGER NOT NULL,
      source_title TEXT NOT NULL,
      source_organisation TEXT DEFAULT '',
      source_url TEXT DEFAULT '',
      source_document_ref TEXT DEFAULT '',
      publication_date TEXT DEFAULT '',
      rights_status TEXT DEFAULT 'LINK_ONLY',
      notes TEXT DEFAULT '',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS growth_content_intake(
      id INTEGER PRIMARY KEY,
      external_draft_id TEXT DEFAULT '',
      source_system TEXT DEFAULT 'Growth Engine',
      payload_json TEXT NOT NULL,
      payload_checksum TEXT NOT NULL,
      status TEXT DEFAULT 'RECEIVED',
      converted_article_id INTEGER,
      imported_by INTEGER,
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE INDEX IF NOT EXISTS idx_prompt_packs_status ON powerhouse_prompt_packs(local_status,subject,chapter);
    CREATE INDEX IF NOT EXISTS idx_generation_batches_pack ON powerhouse_generation_batches(prompt_pack_db_id,created_at);
    CREATE INDEX IF NOT EXISTS idx_import_batches_status ON content_import_batches(status,created_at);
    CREATE INDEX IF NOT EXISTS idx_import_rows_batch ON content_import_batch_rows(batch_id,import_status,row_number);
    CREATE INDEX IF NOT EXISTS idx_pilot_feedback_status ON pilot_feedback(status,routing_target,severity,created_at);
    CREATE INDEX IF NOT EXISTS idx_pilot_events_type ON pilot_activity_events(event_type,created_at);
    CREATE INDEX IF NOT EXISTS idx_knowledge_articles_status ON knowledge_articles(status,published_at);
    """)
    ensure_column(c,'content_import_batches','source_file_path',"TEXT DEFAULT ''")
    controls=[
      ('content_intake','Power House Content Intake','PILOT',{'whole_batch_atomicity':True,'automatic_pre_import_backup':True}),
      ('pilot_issue_reporting','Pilot Issue Reporting','PILOT',{'routes_academic_issues_to_powerhouse':True}),
      ('pilot_analytics','Pilot Operations Analytics','PILOT',{'operational_not_growth_attribution':True}),
      ('knowledge_hub','Knowledge Hub','HIDDEN',{'human_approval_required':True,'copyright_controls':True})]
    for code,name,state,config in controls:
        c.execute("""INSERT OR IGNORE INTO pilot_feature_controls(feature_code,name,state,configuration_json)
          VALUES(?,?,?,?)""",(code,name,state,json.dumps(config,sort_keys=True)))
    c.execute("""INSERT OR IGNORE INTO knowledge_feature_controls(feature_code,state,configuration_json)
      VALUES('knowledge_hub','HIDDEN',?)""",(json.dumps({'human_review_required':True,'growth_engine_drafts_only':True}),))
    # Existing built-in content remains demonstrative unless explicitly imported and governed.
    c.execute("UPDATE questions SET content_environment='DEMO' WHERE COALESCE(is_demo,0)=1")
    c.execute("UPDATE questions SET content_environment='CANDIDATE' WHERE COALESCE(is_demo,0)=0 AND COALESCE(source_import_batch_id,0)=0 AND COALESCE(content_environment,'')='' ")


def migrate_v6_2_3(c):
    """Idempotent V6.2.3 student-command-centre, pathway, coach and connect schema."""
    ensure_column(c,'users','coach_enabled','INTEGER DEFAULT 1')
    ensure_column(c,'users','future_pathway_code',"TEXT DEFAULT ''")
    ensure_column(c,'pilot_feedback','context_json',"TEXT DEFAULT '{}'")
    ensure_column(c,'pilot_feedback','page_path',"TEXT DEFAULT ''")
    c.executescript("""
    CREATE TABLE IF NOT EXISTS student_pathway_preferences(
      student_id INTEGER PRIMARY KEY,
      pathway_code TEXT NOT NULL,
      status TEXT DEFAULT 'ACTIVE',
      selected_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS coach_nudge_events(
      id INTEGER PRIMARY KEY,
      student_id INTEGER NOT NULL,
      nudge_key TEXT NOT NULL,
      action TEXT NOT NULL,
      snoozed_until TEXT DEFAULT '',
      context_json TEXT DEFAULT '{}',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS platform_social_links(
      id INTEGER PRIMARY KEY,
      platform_code TEXT UNIQUE NOT NULL,
      display_name TEXT NOT NULL,
      url TEXT DEFAULT '',
      active INTEGER DEFAULT 0,
      sort_order INTEGER DEFAULT 0,
      updated_by INTEGER,
      updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE INDEX IF NOT EXISTS idx_coach_events_student_key
      ON coach_nudge_events(student_id,nudge_key,created_at);
    """)
    for code,name,order in [
      ('youtube','YouTube',10),('tiktok','TikTok',20),('facebook','Facebook',30),
      ('instagram','Instagram',40),('linkedin','LinkedIn',50),('x','X',60)]:
        c.execute("""INSERT OR IGNORE INTO platform_social_links(platform_code,display_name,sort_order)
          VALUES(?,?,?)""",(code,name,order))


def migrate_v6_2_5(c):
    """Idempotent V6.2.5 Sustainability, Public Trust and Daily Spark schema."""
    c.executescript("""
    CREATE TABLE IF NOT EXISTS sustainability_feature_controls(
      feature_code TEXT PRIMARY KEY,
      state TEXT NOT NULL DEFAULT 'LIVE',
      configuration_json TEXT DEFAULT '{}',
      updated_by INTEGER,
      updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS sustainability_content_blocks(
      id INTEGER PRIMARY KEY,
      content_code TEXT UNIQUE NOT NULL,
      heading TEXT NOT NULL,
      body_text TEXT NOT NULL,
      claim_stage TEXT NOT NULL DEFAULT 'CURRENT_PRACTICE',
      status TEXT NOT NULL DEFAULT 'DRAFT',
      version TEXT DEFAULT '1.0',
      owner TEXT DEFAULT 'ScoreMax',
      evidence_summary TEXT DEFAULT '',
      sort_order INTEGER DEFAULT 0,
      published_at TEXT DEFAULT '',
      updated_by INTEGER,
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS sustainability_policies(
      id INTEGER PRIMARY KEY,
      policy_code TEXT UNIQUE NOT NULL,
      title TEXT NOT NULL,
      summary TEXT DEFAULT '',
      body_text TEXT DEFAULT '',
      version TEXT DEFAULT '1.0',
      effective_date TEXT DEFAULT '',
      last_review_date TEXT DEFAULT '',
      next_review_date TEXT DEFAULT '',
      owner TEXT DEFAULT 'ScoreMax',
      status TEXT DEFAULT 'DRAFT',
      sort_order INTEGER DEFAULT 0,
      published_at TEXT DEFAULT '',
      updated_by INTEGER,
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS sustainability_commitments(
      id INTEGER PRIMARY KEY,
      commitment_code TEXT UNIQUE NOT NULL,
      title TEXT NOT NULL,
      category TEXT NOT NULL,
      claim_stage TEXT NOT NULL,
      description TEXT NOT NULL,
      baseline_text TEXT DEFAULT '',
      target_text TEXT DEFAULT '',
      target_date TEXT DEFAULT '',
      status TEXT DEFAULT 'PLANNED',
      owner TEXT DEFAULT 'ScoreMax',
      evidence_summary TEXT DEFAULT '',
      public_status TEXT DEFAULT 'PUBLISHED',
      sort_order INTEGER DEFAULT 0,
      updated_by INTEGER,
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS sustainability_progress_reports(
      id INTEGER PRIMARY KEY,
      reporting_period TEXT NOT NULL,
      title TEXT NOT NULL,
      summary TEXT DEFAULT '',
      body_text TEXT NOT NULL,
      status TEXT DEFAULT 'DRAFT',
      published_at TEXT DEFAULT '',
      created_by INTEGER,
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS sustainability_draft_intake(
      id INTEGER PRIMARY KEY,
      external_draft_id TEXT DEFAULT '',
      source_system TEXT DEFAULT 'Growth Engine',
      payload_json TEXT NOT NULL,
      payload_checksum TEXT NOT NULL,
      status TEXT DEFAULT 'DRAFT_REVIEW_REQUIRED',
      imported_by INTEGER,
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS daily_spark_feature_controls(
      feature_code TEXT PRIMARY KEY,
      name TEXT NOT NULL,
      state TEXT NOT NULL DEFAULT 'PILOT',
      configuration_json TEXT DEFAULT '{}',
      updated_by INTEGER,
      updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS daily_spark_words(
      id INTEGER PRIMARY KEY,
      word TEXT UNIQUE NOT NULL,
      pronunciation TEXT DEFAULT '',
      definition TEXT NOT NULL,
      example_sentence TEXT NOT NULL,
      synonym TEXT DEFAULT '',
      antonym TEXT DEFAULT '',
      exam_application TEXT DEFAULT '',
      difficulty_rank INTEGER DEFAULT 2,
      min_age INTEGER DEFAULT 10,
      max_age INTEGER DEFAULT 20,
      syllabus_tags_json TEXT DEFAULT '[]',
      source_name TEXT DEFAULT 'ScoreMax controlled vocabulary library',
      source_ref TEXT DEFAULT '',
      active INTEGER DEFAULT 1,
      content_version TEXT DEFAULT '1.0',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE TABLE IF NOT EXISTS daily_spark_assignments(
      id INTEGER PRIMARY KEY,
      student_id INTEGER NOT NULL,
      spark_date TEXT NOT NULL,
      stream TEXT NOT NULL,
      source_type TEXT NOT NULL,
      source_id INTEGER,
      content_version TEXT DEFAULT '1',
      payload_json TEXT NOT NULL,
      selection_reason TEXT DEFAULT '',
      status TEXT DEFAULT 'ASSIGNED',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      UNIQUE(student_id,spark_date,stream));

    CREATE TABLE IF NOT EXISTS daily_spark_events(
      id INTEGER PRIMARY KEY,
      student_id INTEGER NOT NULL,
      assignment_id INTEGER NOT NULL,
      event_type TEXT NOT NULL,
      metadata_json TEXT DEFAULT '{}',
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

    CREATE INDEX IF NOT EXISTS idx_daily_spark_assignments_student_date
      ON daily_spark_assignments(student_id,spark_date,stream);
    CREATE INDEX IF NOT EXISTS idx_daily_spark_events_assignment
      ON daily_spark_events(assignment_id,event_type,created_at);
    CREATE UNIQUE INDEX IF NOT EXISTS idx_daily_spark_event_once
      ON daily_spark_events(student_id,assignment_id,event_type);
    CREATE UNIQUE INDEX IF NOT EXISTS idx_daily_spark_one_answer
      ON daily_spark_events(assignment_id) WHERE event_type IN ('ANSWER_CORRECT','ANSWER_INCORRECT');
    CREATE INDEX IF NOT EXISTS idx_sustainability_commitments_public
      ON sustainability_commitments(public_status,claim_stage,sort_order);
    """)
    c.execute("""INSERT OR IGNORE INTO sustainability_feature_controls(feature_code,state,configuration_json)
      VALUES('sustainability_public','LIVE',?)""",(json.dumps({'separate_current_in_progress_future':True,'annual_progress_reporting':True},sort_keys=True),))
    ensure_column(c,'daily_spark_assignments','snoozed_until',"TEXT DEFAULT ''")
    for code,name,state,config in [
      ('academic_spark','Academic Spark','PILOT',{'formal_mastery':False,'source':'governed_scoremax_content','one_per_day':True}),
      ('word_of_the_day','Word of the Day','PILOT',{'live_ai_required':False,'one_per_day':True,'purpose':'vocabulary_enrichment'})]:
        c.execute("""INSERT OR IGNORE INTO daily_spark_feature_controls(feature_code,name,state,configuration_json)
          VALUES(?,?,?,?)""",(code,name,state,json.dumps(config,sort_keys=True)))

    blocks=[
      ('statement','Our Sustainability Statement',
       'For ScoreMax, sustainability means widening access to useful education, designing inclusively, using data and artificial intelligence responsibly, reducing avoidable waste, and improving the efficiency and accountability of the services we operate.',
       'CURRENT_PRACTICE','PUBLISHED',10,
       'This statement describes the governing scope of the ScoreMax sustainability programme; individual achievements are reported separately.'),
      ('policies_commitments','Our Current Policies and Commitments',
       'ScoreMax separates formal policies, current operating commitments and future targets. Policy records show their version, owner and review cycle so that public statements can be checked and updated.',
       'CURRENT_PRACTICE','PUBLISHED',20,
       'The product contains versioned governance, privacy, accessibility, content-review and public-commitment records.'),
      ('doing_now','What We Are Doing Now',
       'ScoreMax delivers supported assessment and feedback digitally, uses governed content-intake controls, keeps academic and public-content publishing under defined approval rules, and has introduced keyboard and accessibility foundations. These are current practices; their wider impact still needs measurement through real pilots.',
       'CURRENT_PRACTICE','PUBLISHED',30,
       'Supported by the current platform architecture and V6.2.4 accessibility foundations; environmental impact has not yet been quantified.'),
      ('future_plan','Our Future Sustainability Plan',
       'ScoreMax will establish measurable baselines, publish progress honestly, review accessibility with real users and assistive technology, and assess the efficiency and educational impact of its infrastructure and artificial-intelligence services.',
       'FUTURE_COMMITMENT','PUBLISHED',40,
       'Future commitments remain targets until evidence confirms completion.')]
    for row in blocks:
        c.execute("""INSERT OR IGNORE INTO sustainability_content_blocks(content_code,heading,body_text,claim_stage,status,sort_order,evidence_summary,published_at)
          VALUES(?,?,?,?,?,?,?,CURRENT_TIMESTAMP)""",row)

    policies=[
      ('ACCESSIBILITY','Accessibility and Inclusive Design','Design core journeys so they can be used by a broad range of learners and tested with real accessibility methods.','1.0','ScoreMax Product','PUBLISHED',10),
      ('RESPONSIBLE_AI','Responsible AI and Academic Governance','Use artificial intelligence within defined boundaries, preserve source and version evidence, and keep formal academic approval where it is required.','1.0','ScoreMax Academic Governance','PUBLISHED',20),
      ('DATA_USE','Responsible Data Use','Use learner data to support learning, safety and platform operation; avoid selling student data or using academic guidance as disguised advertising.','1.0','ScoreMax Data Governance','PUBLISHED',30),
      ('PARTNER_STANDARDS','Staff, Teacher and Partner Standards','Require professional boundaries, safeguarding controls and accountable conduct in ScoreMax-supported relationships.','1.0','ScoreMax Trust and Safety','PUBLISHED',40)]
    for code,title,summary,version,owner,status,order in policies:
        c.execute("""INSERT OR IGNORE INTO sustainability_policies(policy_code,title,summary,version,owner,status,sort_order,published_at)
          VALUES(?,?,?,?,?,?,?,CURRENT_TIMESTAMP)""",(code,title,summary,version,owner,status,order))

    commitments=[
      ('DIGITAL_ASSESSMENT','Provide digital assessment and feedback','Education access','CURRENT_PRACTICE','Provide supported test, feedback and progress journeys digitally so participating learners are not required to print routine ScoreMax activities.','Digital delivery is available; paper reduction has not yet been measured.','Establish a pilot paper-use baseline and report the measured result.','2027-12-31','ACTIVE','ScoreMax Product','Digital assessment workflows are implemented.',10),
      ('ACCESSIBILITY_ACCEPTANCE','Complete real accessibility acceptance','Accessibility','IN_PROGRESS','Test the core student journeys using keyboard-only navigation, 200% zoom, mobile touch and recognised screen-reader workflows.','Technical foundations implemented in V6.2.4.','Complete and document real-user and assistive-technology acceptance for all core student journeys.','2026-12-31','IN_PROGRESS','ScoreMax Product','Technical controls exist; human acceptance remains outstanding.',20),
      ('AI_GOVERNANCE_REVIEW','Publish an annual responsible-AI review','Responsible AI','FUTURE_COMMITMENT','Review model use, academic boundaries, material incidents and improvement actions at least annually.','No public annual report has yet been issued.','Publish the first annual responsible-AI and data-governance review.','2027-12-31','PLANNED','ScoreMax Academic Governance','Future target; not yet achieved.',30),
      ('INFRASTRUCTURE_BASELINE','Establish a digital-efficiency baseline','Operational efficiency','FUTURE_COMMITMENT','Measure material hosting and AI-service usage so future efficiency claims are evidence based.','No verified operational baseline yet.','Establish the first documented infrastructure and AI-usage baseline.','2027-12-31','PLANNED','ScoreMax Technology','Future target; not yet achieved.',40),
      ('ANNUAL_PROGRESS','Publish sustainability progress updates','Transparency','FUTURE_COMMITMENT','Report achievements, missed targets and revised actions without presenting ambition as completed work.','Public Sustainability section established in V6.2.5.','Publish the first annual Sustainability Progress Update.','2027-12-31','PLANNED','ScoreMax','Future target; not yet achieved.',50)]
    for row in commitments:
        c.execute("""INSERT OR IGNORE INTO sustainability_commitments(commitment_code,title,category,claim_stage,description,baseline_text,target_text,target_date,status,owner,evidence_summary,sort_order)
          VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",row)

    for item in WORD_LIBRARY:
        c.execute("""INSERT OR IGNORE INTO daily_spark_words(word,pronunciation,definition,example_sentence,synonym,antonym,exam_application,
          difficulty_rank,min_age,max_age,syllabus_tags_json,source_name,active,content_version)
          VALUES(?,?,?,?,?,?,?,?,?,?,?,?,1,'1.0')""",
          (item['word'],item['pronunciation'],item['definition'],item['example'],item['synonym'],item['antonym'],item['application'],
           item['difficulty'],item['min_age'],item['max_age'],json.dumps(item.get('tags',[]),sort_keys=True),'ScoreMax controlled vocabulary library'))


def daily_spark_feature_available(c, feature_code):
    row=c.execute("SELECT state FROM daily_spark_feature_controls WHERE feature_code=?",(feature_code,)).fetchone()
    return bool(row and row['state'] in ('PILOT','LIVE'))


def _daily_spark_existing(c, student_id, spark_date, stream):
    return c.execute("SELECT * FROM daily_spark_assignments WHERE student_id=? AND spark_date=? AND stream=?",
                     (student_id,spark_date,stream)).fetchone()


def _daily_spark_record_impression(c, student_id, assignment_id):
    c.execute("INSERT OR IGNORE INTO daily_spark_events(student_id,assignment_id,event_type) VALUES(?,?,'IMPRESSION')",
              (student_id,assignment_id))


def _daily_spark_assignment_view(c, assignment):
    if not assignment:
        return None
    data=dict(assignment); data['payload']=safe_json(data.get('payload_json'),{})
    events=c.execute("SELECT event_type,metadata_json,created_at FROM daily_spark_events WHERE assignment_id=? ORDER BY id",
                     (assignment['id'],)).fetchall()
    data['events']=[{'event_type':e['event_type'],'metadata':safe_json(e['metadata_json'],{}),'created_at':e['created_at']} for e in events]
    data['event_types']={e['event_type'] for e in events}
    data['completed']=bool({'ANSWER_CORRECT','ANSWER_INCORRECT','REVEAL'} & data['event_types'])
    answer_events=[e for e in data['events'] if e['event_type'] in {'ANSWER_CORRECT','ANSWER_INCORRECT'}]
    data['answer_event']=answer_events[-1] if answer_events else None
    data['revealed']='REVEAL' in data['event_types']
    data['saved']='SAVE' in data['event_types']
    return data


def _daily_spark_academic_question(c, student_id, spark_date):
    user=c.execute("SELECT academic_level,subjects,COALESCE(is_demo_account,0) is_demo_account FROM users WHERE id=?",(student_id,)).fetchone()
    demo_clause='' if user and int(user['is_demo_account'] or 0)==1 else " AND COALESCE(q.is_demo,0)=0 AND COALESCE(q.content_environment,'')<>'DEMO'"
    aliases=_programme_aliases(user['academic_level'] if user else '')
    scope_clause,scope_params=_programme_scope_sql(aliases,'q')
    spark_type_clause="lower(COALESCE(q.qtype,'')) IN ('mcq','single_choice','single choice','true/false','true_false','true false','fill_blank','fill blank','fill in the blank','numerical','numeric')"
    recent_ids=[r['source_id'] for r in c.execute("""SELECT source_id FROM daily_spark_assignments
      WHERE student_id=? AND stream='ACADEMIC' AND source_type='QUESTION' AND spark_date>=date(?,'-30 days')""",
      (student_id,spark_date)).fetchall() if r['source_id']]
    exclusion=''
    params=[student_id]+scope_params
    if recent_ids:
        exclusion=f" AND q.id NOT IN ({','.join('?' for _ in recent_ids)})"
        params+=recent_ids
    missed=c.execute(f"""SELECT q.*,a.created_at FROM attempt_answers aa JOIN attempts a ON a.id=aa.attempt_id
      JOIN questions q ON q.id=aa.question_db_id WHERE a.student_id=? AND aa.is_correct=0
      AND {live_question_clause('q')} AND {scope_clause} AND {spark_type_clause}{demo_clause}{exclusion}
      ORDER BY a.created_at DESC,q.id DESC LIMIT 1""",params).fetchone()
    if missed:
        return missed,'This revisits governed content from a question you previously answered incorrectly.'
    plan=get_active_study_plan(c,student_id)
    item=plan.get('next_item') if plan else None
    if item and (item['subject'] or item['chapter'] or item['topic']):
        where=[live_question_clause('q'),scope_clause,spark_type_clause]; p=list(scope_params)
        if demo_clause: where.append("COALESCE(q.is_demo,0)=0 AND COALESCE(q.content_environment,'')<>'DEMO'")
        if item['subject']: where.append('lower(q.subject)=lower(?)'); p.append(item['subject'])
        if item['chapter']: where.append('lower(q.chapter)=lower(?)'); p.append(item['chapter'])
        if item['topic']: where.append('lower(q.topic)=lower(?)'); p.append(item['topic'])
        if recent_ids: where.append(f"q.id NOT IN ({','.join('?' for _ in recent_ids)})"); p+=recent_ids
        rows=c.execute(f"SELECT q.* FROM questions q WHERE {' AND '.join(where)} ORDER BY q.id",p).fetchall()
        if rows:
            idx=int(hashlib.sha256(f"academic|{student_id}|{spark_date}|plan".encode()).hexdigest()[:12],16)%len(rows)
            return rows[idx],'This supports the next subject or topic in your active Study Plan.'
    rows=c.execute(f"SELECT q.* FROM questions q WHERE {live_question_clause('q')} AND {scope_clause} AND {spark_type_clause}{demo_clause}{exclusion} ORDER BY q.id",
                   scope_params+recent_ids).fetchall()
    if not rows and recent_ids:
        rows=c.execute(f"SELECT q.* FROM questions q WHERE {live_question_clause('q')} AND {scope_clause} AND {spark_type_clause}{demo_clause} ORDER BY q.id",scope_params).fetchall()
    if not rows:
        return None,''
    idx=int(hashlib.sha256(f"academic|{student_id}|{spark_date}".encode()).hexdigest()[:12],16)%len(rows)
    return rows[idx],'This is approved content from your current programme, selected to build regular recall.'


def ensure_daily_spark(c, student_id, stream, spark_date=None):
    spark_date=spark_date or iso_today(); stream=str(stream or '').upper()
    feature_code={'ACADEMIC':'academic_spark','WORD':'word_of_the_day'}.get(stream)
    if not feature_code: raise ValueError('Unknown Daily Spark stream.')
    if not daily_spark_feature_available(c,feature_code): return None
    existing=_daily_spark_existing(c,student_id,spark_date,stream)
    if existing:
        if existing['status']=='DISMISSED': return None
        if existing['status']=='SNOOZED' and existing['snoozed_until'] and existing['snoozed_until']>datetime.now().isoformat(timespec='seconds'): return None
        _daily_spark_record_impression(c,student_id,existing['id']); return _daily_spark_assignment_view(c,existing)
    if stream=='ACADEMIC':
        source,reason=_daily_spark_academic_question(c,student_id,spark_date)
        if not source: return None
        payload=academic_payload(source,reason=reason); source_type='QUESTION'; source_id=source['id']; version=payload['content_version']
    elif stream=='WORD':
        user=c.execute("SELECT dob,academic_level FROM users WHERE id=?",(student_id,)).fetchone()
        age=age_from_dob(user['dob'] if user else None)
        seen={int(r['source_id']) for r in c.execute("""SELECT source_id FROM daily_spark_assignments WHERE student_id=?
          AND stream='WORD' AND source_type='VOCABULARY' AND spark_date>=date(?,'-120 days')""",(student_id,spark_date)).fetchall() if r['source_id']}
        words=c.execute("SELECT * FROM daily_spark_words WHERE active=1 ORDER BY difficulty_rank,word").fetchall()
        source=choose_word(words,student_id=student_id,spark_date=spark_date,age=age,seen_ids=seen)
        if not source: return None
        payload=word_payload(source); source_type='VOCABULARY'; source_id=source['id']; version=payload['content_version']
        reason='Selected to broaden your English vocabulary at an appropriate level; syllabus relevance is helpful but not restrictive.'
    c.execute("""INSERT OR IGNORE INTO daily_spark_assignments(student_id,spark_date,stream,source_type,source_id,content_version,payload_json,selection_reason)
      VALUES(?,?,?,?,?,?,?,?)""",(student_id,spark_date,stream,source_type,source_id,version,json.dumps(payload,sort_keys=True),reason))
    row=c.execute("SELECT * FROM daily_spark_assignments WHERE student_id=? AND spark_date=? AND stream=?",(student_id,spark_date,stream)).fetchone()
    _daily_spark_record_impression(c,student_id,row['id'])
    return _daily_spark_assignment_view(c,row)


def daily_spark_snapshot(c, student_id, spark_date=None):
    spark_date=spark_date or iso_today()
    academic=ensure_daily_spark(c,student_id,'ACADEMIC',spark_date)
    word=ensure_daily_spark(c,student_id,'WORD',spark_date)
    c.commit()
    return {'date':spark_date,'academic':academic,'word':word,'available':bool(academic or word)}


def daily_spark_metrics(c):
    rows=c.execute("""SELECT a.stream,e.event_type,COUNT(*) n FROM daily_spark_events e
      JOIN daily_spark_assignments a ON a.id=e.assignment_id GROUP BY a.stream,e.event_type""").fetchall()
    by_stream={}
    for r in rows: by_stream.setdefault(r['stream'],{})[r['event_type']]=int(r['n'])
    streams={r['stream'] for r in c.execute("SELECT DISTINCT stream FROM daily_spark_assignments").fetchall()} | set(by_stream)
    for stream in streams:
        data=by_stream.setdefault(stream,{})
        impressions=c.execute("""SELECT COUNT(DISTINCT e.assignment_id) n FROM daily_spark_events e
          JOIN daily_spark_assignments a ON a.id=e.assignment_id WHERE a.stream=? AND e.event_type='IMPRESSION'""",(stream,)).fetchone()['n']
        engaged=c.execute("""SELECT COUNT(DISTINCT e.assignment_id) n FROM daily_spark_events e
          JOIN daily_spark_assignments a ON a.id=e.assignment_id WHERE a.stream=?
          AND e.event_type IN ('OPEN','REVEAL','ANSWER_CORRECT','ANSWER_INCORRECT','SAVE','REPORT')""",(stream,)).fetchone()['n']
        completed=c.execute("""SELECT COUNT(DISTINCT e.assignment_id) n FROM daily_spark_events e
          JOIN daily_spark_assignments a ON a.id=e.assignment_id WHERE a.stream=?
          AND e.event_type IN ('REVEAL','ANSWER_CORRECT','ANSWER_INCORRECT')""",(stream,)).fetchone()['n']
        data['IMPRESSION']=int(impressions or 0); data['engaged_assignments']=int(engaged or 0); data['completed_assignments']=int(completed or 0)
        data['open_rate']=round(100*engaged/impressions,1) if impressions else 0
        data['completion_rate']=round(100*completed/impressions,1) if impressions else 0
    return by_stream



def written_feature_control(c,feature_code):
    return c.execute("SELECT * FROM written_feature_controls WHERE feature_code=?",(feature_code,)).fetchone()


def written_feature_available(c,user_id,feature_code):
    user=c.execute("SELECT role,COALESCE(written_pilot_enabled,0) written_pilot_enabled FROM users WHERE id=?",(user_id,)).fetchone()
    if not user: return False
    if user['role']=='admin': return True
    control=written_feature_control(c,feature_code)
    if not control or control['state']=='HIDDEN': return False
    if control['state']=='PILOT': return bool(user['written_pilot_enabled'])
    today=datetime.now().date().isoformat()
    if control['available_from'] and control['available_from']>today: return False
    if control['available_to'] and control['available_to']<today: return False
    access=get_access_profile(c,user_id)
    required=ACCESS_CODES.get(control['required_access_code'],ACCESS_CODES['full_access'])['rank']
    return int(access.get('access_rank',0))>=int(required)


def written_package_question(c,question_id):
    return c.execute("""SELECT wq.*,wap.assessment_package_id,wap.assessment_package_version,wap.framework_id,
      wap.framework_version_id,wap.blueprint_snapshot_id,wap.subject_id,wap.chapter_id,wap.local_status package_status,
      wap.immutable_payload_json FROM written_questions wq JOIN written_assessment_packages wap ON wap.id=wq.package_id
      WHERE wq.id=? AND wq.active=1 AND wap.local_status='ACTIVE'""",(question_id,)).fetchone()


def written_question_payload(row):
    return safe_json(row['question_json'],{}) if row else {}


def written_evaluate_attempt(c,attempt_id,answer_version_id,creates_formal_evidence=True):
    attempt=c.execute("SELECT * FROM written_attempts WHERE id=?",(attempt_id,)).fetchone()
    version=c.execute("SELECT * FROM written_answer_versions WHERE id=? AND attempt_id=?",(answer_version_id,attempt_id)).fetchone()
    q=written_package_question(c,attempt['written_question_id']) if attempt else None
    if not attempt or not version or not q: raise ValueError('Written attempt evidence is incomplete.')
    payload=written_question_payload(q)
    answer=version['confirmed_transcript'] or version['answer_text'] or ''
    policy={'confirmed_confidence':0.72,'grader_a_version':'local-rubric-a-1','grader_b_version':'local-rubric-b-1',
            'reconciliation_policy_version':'local-conservative-1'}
    result=mark_written_response(payload,answer,policy)
    cur=c.execute("""INSERT INTO written_marking_runs(attempt_id,answer_version_id,proposed_mark,maximum_mark,percentage,
      confidence,result_state,command_verb_met,grader_a_json,grader_b_json,reconciliation_json,feedback_json,
      contradictions_json,misconceptions_json,validation_boundary) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
      (attempt_id,answer_version_id,result['proposed_mark'],result['maximum_mark'],result['percentage'],result['confidence'],
       result['status'],1 if result['command_verb_met'] else 0,json.dumps(result['grader_a']),json.dumps(result['grader_b']),
       json.dumps({'policy_version':result['reconciliation_policy_version']}),json.dumps(result['feedback']),
       json.dumps(result.get('contradictions',[])),json.dumps(result.get('misconceptions',[])),result['validation_boundary']))
    run_id=cur.lastrowid
    for pt in result['mark_points']:
        c.execute("""INSERT INTO written_mark_point_results(marking_run_id,point_id,description,available_marks,awarded_marks,
          status,evidence_json,improvement_instruction,grader_a_status,grader_b_status) VALUES(?,?,?,?,?,?,?,?,?,?)""",
          (run_id,pt['point_id'],pt['description'],pt['available_marks'],pt['awarded_marks'],pt['status'],json.dumps(pt['evidence']),
           pt['improvement_instruction'],pt['grader_a_status'],pt['grader_b_status']))
    c.execute("""UPDATE written_attempts SET status='MARKED',current_mark=?,maximum_mark=?,marking_confidence=?,result_state=?,
      grader_a_version=?,grader_b_version=?,reconciliation_policy_version=?,completed_at=? WHERE id=?""",
      (result['proposed_mark'],result['maximum_mark'],result['confidence'],result['status'],result['grader_a']['version'],
       result['grader_b']['version'],result['reconciliation_policy_version'],datetime.now().isoformat(timespec='seconds'),attempt_id))
    package=safe_json(q['immutable_payload_json'],{})
    if creates_formal_evidence and result['status']=='MARK_CONFIRMED' and attempt['support_level']=='independent':
        evidence_status='CONFIRMED' if attempt['novelty_status']=='unseen_reconfirmation' else 'AWAITING_UNSEEN_RECONFIRMATION'
        c.execute("""INSERT INTO written_mastery_evidence(student_id,attempt_id,package_id,framework_id,framework_version_id,
          subject_id,chapter_id,learning_outcome_ids_json,concept_ids_json,proposition_ids_json,command_verb,cognitive_demand,
          evidence_level,support_level,novelty_status,score_percentage,confidence,evidence_status,evidence_json)
          VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
          (attempt['student_id'],attempt_id,attempt['package_id'],q['framework_id'],q['framework_version_id'],q['subject_id'],q['chapter_id'],
           json.dumps(payload.get('learning_outcome_ids',[])),json.dumps(payload.get('concept_ids',[])),json.dumps(payload.get('proposition_ids',[])),
           payload.get('command_verb',''),payload.get('cognitive_demand',''),'independent_production',attempt['support_level'],
           attempt['novelty_status'],result['percentage'],result['confidence'],evidence_status,json.dumps(result)))
    missing=[x for x in result['mark_points'] if x['status']!='awarded']
    for pt in missing[:3]:
        activity=(payload.get('recovery_activity_map') or {}).get(pt['point_id'],'')
        c.execute("""INSERT INTO written_recovery_tasks(student_id,attempt_id,task_type,title,reason,approved_activity_id,scheduled_for)
          VALUES(?,?,?,?,?,?,date('now','+1 day'))""",(attempt['student_id'],attempt_id,'targeted_recovery',
          f"Repair: {pt['description'][:80]}",pt['improvement_instruction'],activity))
    if missing:
        plan=c.execute("SELECT * FROM study_plans WHERE student_id=? AND status='active' ORDER BY id DESC LIMIT 1",(attempt['student_id'],)).fetchone()
        if plan:
            title=f"Written-answer recovery: {missing[0]['description'][:70]}"
            exists=c.execute("SELECT 1 FROM study_plan_activities WHERE plan_id=? AND title=? AND status<>'completed'",(plan['id'],title)).fetchone()
            if not exists:
                c.execute("""INSERT INTO study_plan_activities(plan_id,student_id,activity_date,subject,chapter,topic,activity_type,title,
                  target_score,status,source_reason,priority,estimated_minutes,mandatory)
                  VALUES(?,?,date('now','+1 day'),?,?,?,'written_recovery',?,100,'planned',?,1,25,1)""",
                  (plan['id'],attempt['student_id'],q['subject_id'],q['chapter_id'],'',title,'Written response diagnostic evidence'))
    perfect=(abs(result['proposed_mark']-result['maximum_mark'])<0.001 and result['status']=='MARK_CONFIRMED' and
             attempt['support_level']=='independent' and attempt['evidence_type']=='independent_production')
    if perfect:
        c.execute("""INSERT OR IGNORE INTO written_exemplar_candidates(attempt_id,student_id,answer_version_id,eligibility_reason,
          exemplar_type) VALUES(?,?,?,'Perfect confirmed independent score','INDEPENDENT_FULL_MARK')""",
          (attempt_id,attempt['student_id'],answer_version_id))
    c.execute("INSERT INTO written_usage_ledger(student_id,attempt_id,operation,provider,units,metadata_json) VALUES(?,?,?,'local',1,?)",
              (attempt['student_id'],attempt_id,'written_marking',json.dumps({'marking_run_id':run_id})))
    return result,run_id


def materialize_written_exemplar_if_ready(c,candidate_id):
    candidate=c.execute("""SELECT wec.*,wa.written_question_id,wa.package_id,wa.package_version,wa.rubric_version,
      wa.reconciliation_policy_version,u.full_name FROM written_exemplar_candidates wec
      JOIN written_attempts wa ON wa.id=wec.attempt_id JOIN users u ON u.id=wec.student_id WHERE wec.id=?""",(candidate_id,)).fetchone()
    if not candidate or candidate['academic_status']!='APPROVED' or candidate['consent_status']!='OPTED_IN':
        return None
    consent=c.execute("SELECT * FROM written_exemplar_consents WHERE candidate_id=? AND consent_status='OPTED_IN'",(candidate_id,)).fetchone()
    if not consent:
        return None
    display='Anonymous student'
    if consent['attribution_preference']=='FIRST_NAME':
        display=(candidate['full_name'] or 'Student').split()[0]
    control=written_feature_control(c,'written_exemplar_library')
    status='PUBLISHED' if control and control['state']=='LIVE' else 'APPROVED_HIDDEN'
    now=datetime.now().isoformat(timespec='seconds')
    c.execute("""INSERT INTO written_exemplars(candidate_id,written_question_id,package_id,answer_version_id,exemplar_type,
      display_name,publication_status,approved_by,approved_at,published_at,package_version,rubric_version,marking_policy_version)
      VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)
      ON CONFLICT(candidate_id) DO UPDATE SET display_name=excluded.display_name,
      publication_status=excluded.publication_status,published_at=excluded.published_at""",
      (candidate_id,candidate['written_question_id'],candidate['package_id'],candidate['answer_version_id'],candidate['exemplar_type'],
       display,status,candidate['academic_reviewer_id'],candidate['academic_reviewed_at'] or now,
       now if status=='PUBLISHED' else '',candidate['package_version'],candidate['rubric_version'],candidate['reconciliation_policy_version']))
    return status


def written_attempt_view(c,attempt_id,student_id=None):
    where='wa.id=?'; params=[attempt_id]
    if student_id is not None: where+=' AND wa.student_id=?'; params.append(student_id)
    attempt=c.execute(f"""SELECT wa.*,wq.question_text,wq.command_verb,wq.maximum_marks,wq.question_json,
      wap.subject_id,wap.chapter_id,wap.assessment_package_id,wap.assessment_package_version
      FROM written_attempts wa JOIN written_questions wq ON wq.id=wa.written_question_id
      JOIN written_assessment_packages wap ON wap.id=wa.package_id WHERE {where}""",params).fetchone()
    if not attempt: return None
    versions=c.execute("SELECT * FROM written_answer_versions WHERE attempt_id=? ORDER BY version_no",(attempt_id,)).fetchall()
    runs=[]
    for row in c.execute("SELECT * FROM written_marking_runs WHERE attempt_id=? ORDER BY id",(attempt_id,)).fetchall():
        d=dict(row); d['feedback']=safe_json(row['feedback_json'],[]); d['contradictions']=safe_json(row['contradictions_json'],[])
        d['misconceptions']=safe_json(row['misconceptions_json'],[])
        d['points']=c.execute("SELECT * FROM written_mark_point_results WHERE marking_run_id=? ORDER BY id",(row['id'],)).fetchall()
        runs.append(d)
    recovery=c.execute("SELECT * FROM written_recovery_tasks WHERE attempt_id=? ORDER BY id",(attempt_id,)).fetchall()
    pages=c.execute("SELECT * FROM written_upload_pages WHERE attempt_id=? ORDER BY page_no",(attempt_id,)).fetchall()
    candidate=c.execute("SELECT * FROM written_exemplar_candidates WHERE attempt_id=?",(attempt_id,)).fetchone()
    return {'attempt':attempt,'versions':versions,'runs':runs,'recovery':recovery,'pages':pages,'candidate':candidate}

def canonical_question_type(q):
    raw=(q['qtype'] or '').strip().lower().replace('_',' ').replace('-',' ')
    mapping={
        'mcq':'single_choice','single choice':'single_choice','single answer mcq':'single_choice',
        'true/false':'true_false','true false':'true_false',
        'fill blank':'fill_blank','fill in the blank':'fill_blank',
        'multiple select':'multiple_select','multi select':'multiple_select',
        'numerical':'numerical','numeric':'numerical',
        'matching':'matching','ordering':'ordering','drag drop':'drag_drop',
        'short response':'short_response','extended response':'extended_response',
        'image hotspot':'image_hotspot','diagram label':'diagram_label'
    }
    return mapping.get(raw, raw.replace(' ','_') or 'single_choice')

def filter_live_question_ids(c, question_ids):
    ids=[int(x) for x in question_ids if str(x).isdigit() or isinstance(x,int)]
    if not ids:
        return []
    placeholders=','.join('?' for _ in ids)
    rows=c.execute(f"SELECT * FROM questions q WHERE q.id IN ({placeholders}) AND {live_question_clause('q')}",ids).fetchall()
    by_id={r['id']:r for r in rows}
    return [qid for qid in ids if qid in by_id and canonical_question_type(by_id[qid]) in LIVE_MARKABLE_TYPES]

def safe_json(raw, default):
    try:
        return json.loads(raw) if raw else default
    except (TypeError, ValueError, json.JSONDecodeError):
        return default

def mark_question_response(q, selected, blueprint_marking_rules=None):
    """Governed marking adapter with optional immutable blueprint-level score rules."""
    qtype=canonical_question_type(q)
    answer_cfg=safe_json(q['answer_config'], {}) if 'answer_config' in q.keys() else {}
    marking_cfg=safe_json(q['marking_config'], {}) if 'marking_config' in q.keys() else {}
    rules=blueprint_marking_rules if isinstance(blueprint_marking_rules,dict) else {}
    question_max=float(marking_cfg.get('marks') or (q['marks'] if 'marks' in q.keys() else 1) or 1)
    correct_marks=float(rules.get('correct_marks')) if rules.get('correct_marks') is not None else question_max
    incorrect_marks=float(rules.get('incorrect_marks')) if rules.get('incorrect_marks') is not None else float(marking_cfg.get('negative_marks') or 0)
    unanswered_marks=float(rules.get('unanswered_marks')) if rules.get('unanswered_marks') is not None else 0.0
    partial_credit=bool(rules.get('partial_credit_allowed')) if 'partial_credit_allowed' in rules else bool(marking_cfg.get('partial_credit'))
    response=(selected or '').strip()
    if not response:
        return False,unanswered_marks,''

    if qtype=='fill_blank':
        accepted=answer_cfg.get('accepted_answers') or [q['option_a'] or q['answer'] or '']
        trim=answer_cfg.get('trim_spaces',True)
        case_sensitive=answer_cfg.get('case_sensitive',False)
        candidate=response.strip() if trim else response
        expected=[str(x).strip() if trim else str(x) for x in accepted]
        if not case_sensitive:
            candidate=candidate.casefold(); expected=[x.casefold() for x in expected]
        ok=candidate in expected
        return ok,correct_marks if ok else incorrect_marks,''

    if qtype=='numerical':
        try:
            value=float(response); target=float(marking_cfg.get('correct_value')); tolerance=float(marking_cfg.get('tolerance',0)); ok=abs(value-target)<=tolerance
        except (TypeError,ValueError): ok=False
        return ok,correct_marks if ok else incorrect_marks,''

    if qtype=='multiple_select':
        selected_ids={x.strip() for x in response.split(',') if x.strip()}; correct_ids=set(marking_cfg.get('correct_option_ids') or [])
        ok=bool(correct_ids) and selected_ids==correct_ids
        if partial_credit and correct_ids and not ok:
            good=len(selected_ids & correct_ids); bad=len(selected_ids-correct_ids)
            fraction=(good-bad)/len(correct_ids)
            awarded=correct_marks*fraction
            # Blueprint incorrect marks is the governed floor for a submitted wrong response.
            awarded=max(incorrect_marks,min(correct_marks,awarded))
            return ok,round(awarded,4),''
        return ok,correct_marks if ok else incorrect_marks,''

    correct_ids=marking_cfg.get('correct_option_ids') or [q['answer'] or '']
    ok=response.upper() in {str(x).strip().upper() for x in correct_ids}
    misconception=''
    if not ok and response:
        for opt in answer_cfg.get('options',[]):
            if str(opt.get('id','')).strip().upper()==response.upper():
                misconception=str(opt.get('misconception') or opt.get('diagnostic_tag') or '').strip(); break
        if not misconception:
            tags=safe_json(q['misconception_tags'],[]) if 'misconception_tags' in q.keys() else []
            if isinstance(tags,list) and len(tags)==1: misconception=str(tags[0]).strip()
    return ok,correct_marks if ok else incorrect_marks,misconception

def evidence_strength(answered):
    if answered < 3: return 'Insufficient evidence'
    if answered < 5: return 'Emerging evidence'
    if answered < 8: return 'Moderate evidence'
    return 'Strong evidence'

def performance_status(accuracy, answered=0):
    if answered < 3: return 'More evidence needed'
    if accuracy >= 80: return 'Strong'
    if accuracy >= 70: return 'Secure'
    if accuracy >= 60: return 'Developing'
    return 'Needs attention'

def attempt_diagnostics(c, attempt_id, student_id=None):
    params=[attempt_id]; ownership=''
    if student_id is not None:
        ownership=' AND a.student_id=?'; params.append(student_id)
    attempt=c.execute('SELECT a.* FROM attempts a WHERE a.id=?'+ownership,params).fetchone()
    if not attempt: return None
    rows=c.execute("""SELECT aa.*,q.subject,q.chapter,q.topic,q.subtopic,q.learning_outcome,q.concept,
        q.difficulty,q.level,q.cognitive_skill,q.command_word,q.misconception_tags,q.question,q.explanation,q.answer
        FROM attempt_answers aa JOIN questions q ON q.id=aa.question_db_id WHERE aa.attempt_id=?""",(attempt_id,)).fetchall()
    def aggregate(field, fallback=None):
        groups={}
        for r in rows:
            key=(r[field] if field in r.keys() else '') or (r[fallback] if fallback and fallback in r.keys() else '') or 'Unmapped'
            g=groups.setdefault(key,{'name':key,'answered':0,'correct':0})
            g['answered']+=1; g['correct']+=int(r['is_correct'] or 0)
        out=[]
        for g in groups.values():
            g['accuracy']=round(100*g['correct']/g['answered'],1) if g['answered'] else 0
            g['evidence']=evidence_strength(g['answered']); g['status']=performance_status(g['accuracy'],g['answered']); out.append(g)
        return sorted(out,key=lambda x:(x['accuracy'],-x['answered']))
    by_topic=aggregate('subtopic','topic'); by_lo=aggregate('learning_outcome','subtopic')
    by_difficulty=aggregate('difficulty','level'); by_skill=aggregate('cognitive_skill'); by_command=aggregate('command_word')
    confidence={'confident_correct':0,'confident_wrong':0,'unsure_wrong':0,'guess_correct':0,'recorded':0}
    misconceptions={}
    for r in rows:
        conf=(r['confidence'] or '').strip().lower()
        if conf:
            confidence['recorded']+=1
            if conf=='confident' and r['is_correct']: confidence['confident_correct']+=1
            elif conf=='confident' and not r['is_correct']: confidence['confident_wrong']+=1
            elif conf in ('not sure','unsure') and not r['is_correct']: confidence['unsure_wrong']+=1
            elif conf in ('guessing','guess') and r['is_correct']: confidence['guess_correct']+=1
        trig=(r['misconception_triggered'] or '').strip()
        if trig: misconceptions[trig]=misconceptions.get(trig,0)+1
    misconception_list=[{'name':k,'count':v,'strength':'Likely' if v>=2 else 'Possible'} for k,v in misconceptions.items()]
    misconception_list.sort(key=lambda x:-x['count'])
    weak=next((x for x in by_lo if x['name']!='Unmapped' and x['answered']>=3 and x['accuracy']<70),None) or next((x for x in by_topic if x['name']!='Unmapped' and x['accuracy']<70),None)
    strong=max(by_topic,key=lambda x:x['accuracy'],default=None)
    level_perf={x['name']:x for x in by_difficulty}; mastery='Foundation'
    for level,threshold in [('Elite',80),('Distinction',75),('Exam Ready',70),('Foundation',70)]:
        r=level_perf.get(level)
        if r and r['answered']>=3 and r['accuracy']>=threshold: mastery=level; break
    cur=level_perf.get(mastery); mastery_state='Secure' if cur and cur['answered']>=5 and cur['accuracy']>=80 else 'Developing'
    if weak:
        next_action={'title':f"Recover {weak['name']}",'reason':f"Your current accuracy is {weak['accuracy']}% across {weak['answered']} question(s).",'kind':'recovery'}
    elif cur and cur['accuracy']>=80:
        next_action={'title':'Move to the next challenge level','reason':'Your current evidence is strong enough to increase the challenge.','kind':'progress'}
    else:
        next_action={'title':'Build more mastery evidence','reason':'Complete another targeted assessment so ScoreMax can make a stronger recommendation.','kind':'evidence'}
    recovery_comparison=None
    if 'assessment_kind' in attempt.keys() and attempt['assessment_kind']=='recovery' and attempt['recovery_parent_attempt_id']:
        focus_type=attempt['recovery_focus_type'] or 'topic'; focus_name=attempt['recovery_focus_name'] or ''
        before=focus_accuracy(c,attempt['recovery_parent_attempt_id'],focus_type,focus_name)
        after=focus_accuracy(c,attempt_id,focus_type,focus_name)
        if before is not None and after is not None:
            change=round(after-before,1)
            recovery_comparison={'focus':focus_name,'before':before,'after':after,'change':change,
                'recovered': after>=70 and change>0}
            if after>=70:
                next_action={'title':'Continue your mastery pathway','reason':f'Your recovery score for {focus_name} reached {after}%.','kind':'progress'}
            else:
                next_action={'title':f'Keep working on {focus_name}','reason':f'Your recovery score is {after}%. More targeted practice is recommended.','kind':'recovery'}
    return dict(attempt=attempt,rows=rows,by_topic=by_topic,by_lo=by_lo,by_difficulty=by_difficulty,
        by_skill=[x for x in by_skill if x['name']!='Unmapped'],by_command=[x for x in by_command if x['name']!='Unmapped'],
        confidence=confidence,misconceptions=misconception_list,weak=weak,strong=strong,mastery=mastery,mastery_state=mastery_state,next_action=next_action,
        recovery_comparison=recovery_comparison)

def diagnostic_focus(diagnostic):
    """Return the best recovery focus using LO first, then topic/subtopic."""
    weak=diagnostic.get('weak') if diagnostic else None
    if not weak:
        return None
    name=(weak.get('name') or '').strip()
    if not name or name=='Unmapped':
        return None
    # If the weak record is present in by_lo, treat it as an LO; otherwise topic.
    lo_names={x.get('name') for x in diagnostic.get('by_lo',[]) if x.get('name') and x.get('name')!='Unmapped'}
    return {'type':'learning_outcome' if name in lo_names else 'topic','name':name,
            'accuracy':weak.get('accuracy',0),'answered':weak.get('answered',0)}

def recovery_question_ids(c, student_id, diagnostic, count=8):
    """Select approved, active and preferably unseen questions for the diagnosed weakness."""
    focus=diagnostic_focus(diagnostic)
    if not focus:
        return focus, []
    attempt=diagnostic['attempt']
    clauses=[live_question_clause('q'), "q.subject=?"]
    params=[attempt['subject']]
    if focus['type']=='learning_outcome':
        clauses.append("q.learning_outcome=?"); params.append(focus['name'])
    else:
        clauses.append("(q.subtopic=? OR q.topic=?)"); params.extend([focus['name'],focus['name']])
    # Stay near the learner's current challenge level where possible, but do not make it mandatory.
    level=attempt['level'] or ''
    sql=f"""SELECT q.id,
      CASE WHEN EXISTS(SELECT 1 FROM attempt_answers aa JOIN attempts a ON a.id=aa.attempt_id
                       WHERE aa.question_db_id=q.id AND a.student_id=?) THEN 1 ELSE 0 END seen_q,
      CASE WHEN ?<>'' AND q.level=? THEN 0 ELSE 1 END level_penalty
      FROM questions q WHERE {' AND '.join(clauses)}
      ORDER BY seen_q, level_penalty, RANDOM() LIMIT ?"""
    rows=c.execute(sql,[student_id,level,level]+params+[max(6,min(40,count*2))]).fetchall()
    return focus,filter_live_question_ids(c,[r['id'] for r in rows])[:count]

def focus_accuracy(c, attempt_id, focus_type, focus_name):
    if not attempt_id or not focus_name:
        return None
    if focus_type=='learning_outcome':
        clause='q.learning_outcome=?'; params=[attempt_id,focus_name]
    else:
        clause='(q.subtopic=? OR q.topic=?)'; params=[attempt_id,focus_name,focus_name]
    row=c.execute(f"""SELECT COUNT(*) answered, COALESCE(SUM(aa.is_correct),0) correct
        FROM attempt_answers aa JOIN questions q ON q.id=aa.question_db_id
        WHERE aa.attempt_id=? AND {clause}""",params).fetchone()
    if not row or not row['answered']:
        return None
    return round(100.0*row['correct']/row['answered'],1)

def next_user_id(role, uid):
    prefix={'student':'STU','teacher':'TCH','parent':'PAR','admin':'ADM','reviewer':'REV'}.get(role,'USR')
    return f"{prefix}-{uid:06d}"

def code():
    chars='ABCDEFGHJKLMNPQRSTUVWXYZ23456789'; return ''.join(random.choice(chars) for _ in range(6))

def require(role=None):
    if 'user_id' not in session: return False
    return role is None or session.get('role')==role



def calculate_health_score(avg_score, tests_completed, recent_scores):
    if not tests_completed:
        return 60
    consistency=min(100,tests_completed*8)
    trend=50
    if len(recent_scores)>=2:
        change=recent_scores[0]-recent_scores[-1]
        trend=max(0,min(100,50+change*2))
    return round(avg_score*.60+consistency*.25+trend*.15)

def student_analytics(c,student_id):
    attempts=c.execute("SELECT * FROM attempts WHERE student_id=? ORDER BY created_at DESC",(student_id,)).fetchall()
    n=len(attempts)
    average=round(sum(float(a['score']) for a in attempts)/n,1) if n else 0
    recent=[float(a['score']) for a in attempts[:5]]
    health=calculate_health_score(average,n,recent)
    by_subtopic=c.execute("""SELECT COALESCE(q.subtopic,q.topic) area,q.subject,COUNT(*) answered,ROUND(100.0*SUM(aa.is_correct)/COUNT(*),1) accuracy FROM attempt_answers aa JOIN attempts a ON a.id=aa.attempt_id JOIN questions q ON q.id=aa.question_db_id WHERE a.student_id=? GROUP BY q.subject,COALESCE(q.subtopic,q.topic) ORDER BY accuracy ASC,answered DESC""",(student_id,)).fetchall()
    by_level=c.execute("""SELECT q.level,COUNT(*) answered,ROUND(100.0*SUM(aa.is_correct)/COUNT(*),1) accuracy FROM attempt_answers aa JOIN attempts a ON a.id=aa.attempt_id JOIN questions q ON q.id=aa.question_db_id WHERE a.student_id=? GROUP BY q.level ORDER BY CASE q.level WHEN 'Foundation' THEN 1 WHEN 'Exam Ready' THEN 2 WHEN 'Advanced' THEN 3 WHEN 'Distinction' THEN 4 WHEN 'Expert' THEN 5 WHEN 'Elite' THEN 6 ELSE 7 END""",(student_id,)).fetchall()
    weakest=by_subtopic[0] if by_subtopic else None
    strongest=max(by_subtopic,key=lambda r:float(r['accuracy'])) if by_subtopic else None
    rec=[]
    if weakest: rec.append(f"Prioritise {weakest['area']} in {weakest['subject']} — current accuracy is {weakest['accuracy']}%.")
    if n<3: rec.append('Complete at least three assessments so ScoreMax can build a more reliable progress profile.')
    elif average>=80: rec.append('Your overall performance is strong. Try the next mastery level on your strongest topics.')
    elif average<60: rec.append('Stay at your current mastery level and complete targeted recovery tests before progressing.')
    else: rec.append('Focus your next tests on weaker sub-topics rather than repeating strong areas.')
    return dict(attempts=attempts,tests_completed=n,avg_score=average,health=health,by_subtopic=by_subtopic,by_level=by_level,weakest=weakest,strongest=strongest,recommendations=rec)


SCOREMAX_LEVELS = ['Foundation','Exam Ready','Advanced','Distinction','Expert','Elite']

def scoremax_level_from_evidence(accuracy, answered, percentile=None, premium_path=True):
    """V5.1 provisional level model. Elite remains competitive and intentionally scarce."""
    a=float(accuracy or 0); n=int(answered or 0)
    if n < 4 or a < 55:
        return 'Foundation'
    if a < 68 or n < 8:
        return 'Exam Ready'
    # Upper levels require broader evidence. In live commercial mode these require the corresponding Access entitlement as well as earned evidence.
    if a < 78 or n < 12:
        return 'Advanced'
    if a < 86 or n < 18:
        return 'Distinction'
    if a < 92 or n < 25:
        return 'Expert'
    if percentile is not None and float(percentile) >= 95 and n >= 30:
        return 'Elite'
    return 'Expert'

def level_index(name):
    try: return SCOREMAX_LEVELS.index(name)
    except ValueError: return 0

def next_level_name(name):
    i=level_index(name)
    return SCOREMAX_LEVELS[min(i+1,len(SCOREMAX_LEVELS)-1)]

def next_level_progress(accuracy, current_level):
    """Simple visual progress only; final mastery policy can be recalibrated after real outcome data."""
    thresholds={'Foundation':55,'Exam Ready':68,'Advanced':78,'Distinction':86,'Expert':92}
    if current_level=='Elite': return 100
    lower={'Foundation':0,'Exam Ready':55,'Advanced':68,'Distinction':78,'Expert':86}.get(current_level,0)
    target=thresholds.get(current_level,100)
    if target<=lower: return 100
    return max(0,min(99,round(100*(float(accuracy or 0)-lower)/(target-lower))))


def student_declared_subjects(user):
    if not user:
        return []
    return parse_list(user['subjects'] if 'subjects' in user.keys() else '')


def student_catalogue_subjects(user):
    """Return visible study subjects independently from live question inventory."""
    declared=student_declared_subjects(user)
    subjects=list(declared)
    if is_matric_level(user['academic_level'] if user else ''):
        for name in MATRIC_COMMON_SUBJECTS:
            if name.casefold() not in {x.casefold() for x in subjects}:
                subjects.append(name)
    return subjects


def student_pathway_snapshot(c, student_id):
    user=c.execute("SELECT academic_level,future_pathway_code FROM users WHERE id=?",(student_id,)).fetchone()
    row=c.execute("SELECT * FROM student_pathway_preferences WHERE student_id=?",(student_id,)).fetchone()
    code=(row['pathway_code'] if row else (user['future_pathway_code'] if user else '')) or ''
    selected=pathway_definition(code)
    live_subjects={r['subject'].casefold() for r in c.execute(f"SELECT DISTINCT q.subject FROM questions q WHERE {live_question_clause('q')}").fetchall()}
    live_programmes={r['programme'].casefold() for r in c.execute(f"SELECT DISTINCT COALESCE(q.programme,'') programme FROM questions q WHERE {live_question_clause('q')}").fetchall() if r['programme']}
    active_frameworks={r['framework_name'].casefold() for r in c.execute("""SELECT DISTINCT af.name framework_name FROM assessment_blueprints ab JOIN assessment_frameworks af ON af.id=ab.framework_id WHERE ab.local_status='ACTIVE'""").fetchall() if r['framework_name']}
    options=[]
    for item in PATHWAY_CATALOGUE:
        subjects=[]
        for subject in item['subjects']:
            subjects.append({'name':subject,'availability':'LIVE' if subject.casefold() in live_subjects else 'COMING_SOON'})
        tests=[]
        for assessment in item['future_assessments']:
            key=assessment.casefold()
            live=key in active_frameworks or any(key in p or p in key for p in live_programmes if len(p)>3)
            tests.append({'name':assessment,'availability':'LIVE' if live else 'EXPLORE'})
        options.append({**item,'subjects_detail':subjects,'assessments_detail':tests,'selected':item['code']==code})
    return {
      'is_matric':is_matric_level(user['academic_level'] if user else ''),
      'selected_code':code,
      'selected':selected,
      'options':options,
    }


def active_social_links(c):
    return c.execute("SELECT * FROM platform_social_links WHERE active=1 AND COALESCE(url,'')<>'' ORDER BY sort_order,id").fetchall()


def _nudge_suppressed(c, student_id, nudge_key):
    row=c.execute("""SELECT action,snoozed_until,created_at FROM coach_nudge_events
      WHERE student_id=? AND nudge_key=? ORDER BY id DESC LIMIT 1""",(student_id,nudge_key)).fetchone()
    if not row:
        return False
    today=iso_today()
    if row['action']=='SNOOZE' and row['snoozed_until'] and row['snoozed_until']>=today:
        return True
    if row['action']=='DISMISS':
        created=_parse_date(str(row['created_at'] or '')[:10])
        return bool(created and (datetime.now().date()-created).days<30)
    return False


def scoremax_coach_candidates(c, student_id):
    """One evidence-aware next step. Support/commercial nudges are deliberately low priority."""
    user=c.execute("SELECT * FROM users WHERE id=?",(student_id,)).fetchone()
    if not user or not bool(user['coach_enabled']):
        return []
    candidates=[]
    active=c.execute("SELECT id FROM assessment_sessions WHERE student_id=? AND status='in_progress' ORDER BY id DESC LIMIT 1",(student_id,)).fetchone()
    if active:
        candidates.append({'key':'continue-assessment','priority':100,'tone':'urgent','title':'Continue your saved test','message':'Your answers are saved. Finish this assessment before starting something new.','endpoint':'take_test_v4','kwargs':{'assessment_id':active['id']},'action_label':'Continue test','reason':'An assessment is already in progress.'})
    plan=get_active_study_plan(c,student_id)
    if plan and plan.get('next_item'):
        x=plan['next_item']
        due_today=(x['activity_date'] or '')<=iso_today()
        candidates.append({'key':f"plan-{x['id']}",'priority':92 if due_today else 76,'tone':'academic','title':x['title'],'message':x['source_reason'] or 'This is the next evidence-building action in your Study Plan.','endpoint':'study_plan_page','kwargs':{},'action_label':'Open my plan','reason':'Your active Study Plan has a next action.'})
    pathway_data=student_pathway_snapshot(c,student_id)
    if pathway_data['is_matric'] and not pathway_data['selected_code']:
        candidates.append({'key':'choose-future-pathway','priority':88,'tone':'guidance','title':'What are you considering after Matric?','message':'Choose a possible route—or say you are still deciding—so ScoreMax can show relevant subjects and future tests.','endpoint':'student_pathways','kwargs':{},'action_label':'Explore my options','reason':'Your current level is Matric and no future pathway is saved.'})
    answered=c.execute("SELECT COUNT(*) n FROM attempt_answers aa JOIN attempts a ON a.id=aa.attempt_id WHERE a.student_id=?",(student_id,)).fetchone()['n']
    if not answered:
        candidates.append({'key':'first-diagnostic','priority':84,'tone':'academic','title':'Build your starting point','message':'Take a short diagnostic so ScoreMax can give you evidence-based guidance instead of generic advice.','endpoint':'test_setup','kwargs':{},'action_label':'Take a short test','reason':'There is not enough assessment evidence yet.'})
    recall=due_recall_items(c,student_id,limit=1)
    if recall:
        candidates.append({'key':f"recall-{recall[0]['id']}",'priority':82,'tone':'academic','title':'Recall is due','message':f"Reconfirm {recall[0]['area_name'] or recall[0]['concept_key']} before the evidence becomes stale.",'endpoint':'study_plan_page','kwargs':{},'action_label':'Open recall plan','reason':'A spaced-recall item is due.'})
    weak=student_weak_areas(c,student_id,limit=1)
    if weak:
        w=weak[0]
        candidates.append({'key':f"weak-{w['subject']}-{w['concept_key']}",'priority':78 if float(w['accuracy'] or 0)<60 else 70,'tone':'academic','title':f"Recover {w['area']}",'message':f"Your verified evidence in {w['subject']} is currently {w['accuracy']}%. Focus here before repeating strong areas.",'endpoint':'weak_areas_page','kwargs':{},'action_label':'Improve this area','reason':'This is the highest-priority verified weak area.'})
    if not (user['goal_name'] or user['goal_type']):
        candidates.append({'key':'set-goal','priority':45,'tone':'guidance','title':'Set a goal for clearer guidance','message':'A target exam or pathway helps ScoreMax prioritise the right work.','endpoint':'student_profile','kwargs':{},'action_label':'Set my goal','reason':'No student goal is saved.'})
    candidates=[x for x in candidates if not _nudge_suppressed(c,student_id,x['key'])]
    return sorted(candidates,key=lambda x:(-x['priority'],x['key']))


def scoremax_coach(c, student_id):
    rows=scoremax_coach_candidates(c,student_id)
    return rows[0] if rows else None

def subject_community_snapshot(c, programme, subject, student_id=None):
    """Aggregate distribution only. Absolute counts remain hidden in the student UI by default."""
    users=c.execute("""SELECT DISTINCT a.student_id FROM attempts a
      JOIN attempt_answers aa ON aa.attempt_id=a.id JOIN questions q ON q.id=aa.question_db_id
      WHERE COALESCE(NULLIF(q.programme,''),NULLIF(a.programme,''),'')=? AND q.subject=?""",(programme or '',subject)).fetchall()
    distribution={x:0 for x in SCOREMAX_LEVELS}
    total=0
    current=None
    for u in users:
        r=c.execute("""SELECT COUNT(*) answered,ROUND(100.0*SUM(aa.is_correct)/COUNT(*),1) accuracy
          FROM attempt_answers aa JOIN attempts a ON a.id=aa.attempt_id JOIN questions q ON q.id=aa.question_db_id
          WHERE a.student_id=? AND q.subject=? AND COALESCE(NULLIF(q.programme,''),NULLIF(a.programme,''),'')=?""",
          (u['student_id'],subject,programme or '')).fetchone()
        if not r or not r['answered']: continue
        level=scoremax_level_from_evidence(r['accuracy'],r['answered'])
        distribution[level]+=1; total+=1
        if student_id and int(u['student_id'])==int(student_id): current=level
    percentages={k:(round(100*v/total,1) if total else 0) for k,v in distribution.items()}
    return {'total':total,'distribution':distribution,'percentages':percentages,'current_level':current}

def health_status(accuracy=None, answered=0):
    answered=int(answered or 0)
    accuracy=float(accuracy or 0)
    if answered<=0:
        return 'Not Started'
    if answered<3:
        return 'Developing'
    if accuracy>=80:
        return 'Strong'
    if accuracy>=65:
        return 'Developing'
    if accuracy>=50:
        return 'Needs Attention'
    return 'Weak'

def _area_key_from_question(q):
    return str((q['concept_id'] if 'concept_id' in q.keys() else '') or
               (q['learning_outcome'] if 'learning_outcome' in q.keys() else '') or
               (q['concept'] if 'concept' in q.keys() else '') or
               (q['subtopic'] if 'subtopic' in q.keys() else '') or
               (q['topic'] if 'topic' in q.keys() else '') or
               (q['chapter'] if 'chapter' in q.keys() else '')).strip()

def _area_name_from_question(q):
    return str((q['learning_outcome'] if 'learning_outcome' in q.keys() else '') or
               (q['concept'] if 'concept' in q.keys() else '') or
               (q['subtopic'] if 'subtopic' in q.keys() else '') or
               (q['topic'] if 'topic' in q.keys() else '') or
               (q['chapter'] if 'chapter' in q.keys() else '')).strip()

def update_learning_intelligence_from_attempt(c, attempt_id, student_id):
    """Write assessment evidence into weak-area, misconception and recall state."""
    rows=c.execute("""SELECT aa.*,q.* FROM attempt_answers aa
      JOIN questions q ON q.id=aa.question_db_id WHERE aa.attempt_id=?""",(attempt_id,)).fetchall()
    if not rows:
        return
    now=datetime.now().isoformat(timespec='seconds')
    today=datetime.now().date()
    grouped={}
    for q in rows:
        key=_area_key_from_question(q)
        if not key:
            continue
        g=grouped.setdefault(key,{'subject':q['subject'] or '','chapter':q['chapter'] or '',
            'topic':q['topic'] or '','area_name':_area_name_from_question(q),'answered':0,'correct':0})
        g['answered']+=1; g['correct']+=int(q['is_correct'] or 0)
        misconception=(q['misconception_triggered'] or '').strip()
        if misconception and not q['is_correct']:
            prior=c.execute("SELECT * FROM student_misconceptions WHERE student_id=? AND misconception_key=?",
                            (student_id,misconception)).fetchone()
            ev=int(prior['evidence_count'] or 0)+1 if prior else 1
            confident=int(prior['confident_wrong_count'] or 0) if prior else 0
            if (q['confidence'] or '').strip().lower()=='confident':
                confident+=1
            status='Confirmed' if ev>=2 else 'Emerging Concern'
            c.execute("""INSERT INTO student_misconceptions(student_id,misconception_key,subject,area_name,evidence_count,
              confident_wrong_count,status,last_attempt_id,last_seen_at) VALUES(?,?,?,?,?,?,?,?,?)
              ON CONFLICT(student_id,misconception_key) DO UPDATE SET subject=excluded.subject,area_name=excluded.area_name,
              evidence_count=excluded.evidence_count,confident_wrong_count=excluded.confident_wrong_count,status=excluded.status,
              last_attempt_id=excluded.last_attempt_id,last_seen_at=excluded.last_seen_at,recovered_at=''""",
              (student_id,misconception,q['subject'] or '',_area_name_from_question(q),ev,confident,status,attempt_id,now))

    attempt=c.execute("SELECT * FROM attempts WHERE id=? AND student_id=?",(attempt_id,student_id)).fetchone()
    for key,g in grouped.items():
        prior=c.execute("SELECT * FROM student_learning_states WHERE student_id=? AND area_key=?",(student_id,key)).fetchone()
        ev=(int(prior['evidence_count'] or 0) if prior else 0)+g['answered']
        cor=(int(prior['correct_count'] or 0) if prior else 0)+g['correct']
        accuracy=round(100.0*cor/ev,1) if ev else 0
        status='Emerging Concern' if ev<3 else ('Weak Area' if accuracy<75 else 'Secure')

        # A clean, targeted recovery attempt can close an existing weakness; later recall still has to prove retention.
        attempt_kind=(attempt['assessment_kind'] or '') if attempt else ''
        focus_name=(attempt['recovery_focus_name'] or '') if attempt else ''
        this_pct=round(100.0*g['correct']/g['answered'],1) if g['answered'] else 0
        if prior and prior['status'] in ('Weak Area','Recovery') and attempt_kind in ('recovery','recall','reconfirmation') and g['answered']>=3 and this_pct>=80:
            status='Recovered'
        elif prior and prior['status']=='Recovered' and attempt_kind in ('recall','reconfirmation') and g['answered']>=3 and this_pct>=80:
            # Successful delayed recall/reconfirmation preserves a recovered area even when historic cumulative accuracy is still <75%.
            status='Recovered'
            # A misconception is only closed by evidence, never by a click. Match the recovered academic area conservatively.
            c.execute("""UPDATE student_misconceptions SET status='Recovered',recovered_at=?
              WHERE student_id=? AND status='Confirmed' AND subject=? AND area_name=?""",
              (now,student_id,g['subject'],g['area_name']))
        recovered_at=now if status=='Recovered' else ''
        c.execute("""INSERT INTO student_learning_states(student_id,area_key,subject,chapter,topic,area_name,
          evidence_count,correct_count,accuracy,status,last_attempt_id,last_evidence_at,recovered_at)
          VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)
          ON CONFLICT(student_id,area_key) DO UPDATE SET subject=excluded.subject,chapter=excluded.chapter,topic=excluded.topic,
          area_name=excluded.area_name,evidence_count=excluded.evidence_count,correct_count=excluded.correct_count,
          accuracy=excluded.accuracy,status=excluded.status,last_attempt_id=excluded.last_attempt_id,
          last_evidence_at=excluded.last_evidence_at,recovered_at=excluded.recovered_at""",
          (student_id,key,g['subject'],g['chapter'],g['topic'],g['area_name'],ev,cor,accuracy,status,attempt_id,now,recovered_at))

        # Recall scheduling is deliberately transparent/rule-based in V5.3.
        recall=c.execute("SELECT * FROM recall_items WHERE student_id=? AND concept_key=?",(student_id,key)).fetchone()
        successful=int(recall['successful_recalls'] or 0) if recall else 0
        if this_pct>=80:
            successful+=1
            interval=[7,21,45,90][min(successful-1,3)]
            next_due=today+timedelta(days=interval)
            rstatus='scheduled'
        else:
            interval=3
            next_due=today+timedelta(days=3)
            rstatus='repair_first'
        c.execute("""INSERT INTO recall_items(student_id,concept_key,subject,chapter,topic,area_name,interval_days,
          successful_recalls,last_score,last_attempt_at,next_due_date,status,updated_at)
          VALUES(?,?,?,?,?,?,?,?,?,?,?,?,CURRENT_TIMESTAMP)
          ON CONFLICT(student_id,concept_key) DO UPDATE SET subject=excluded.subject,chapter=excluded.chapter,topic=excluded.topic,
          area_name=excluded.area_name,interval_days=excluded.interval_days,successful_recalls=excluded.successful_recalls,
          last_score=excluded.last_score,last_attempt_at=excluded.last_attempt_at,next_due_date=excluded.next_due_date,
          status=excluded.status,updated_at=CURRENT_TIMESTAMP""",
          (student_id,key,g['subject'],g['chapter'],g['topic'],g['area_name'],interval,successful,this_pct,now,next_due.isoformat(),rstatus))

def due_recall_items(c, student_id, limit=6):
    return c.execute("""SELECT * FROM recall_items WHERE student_id=? AND status='scheduled' AND next_due_date<>''
      AND date(next_due_date)<=date('now') ORDER BY date(next_due_date),last_score LIMIT ?""",(student_id,limit)).fetchall()

def confirmed_misconceptions(c, student_id, limit=8):
    return c.execute("""SELECT * FROM student_misconceptions WHERE student_id=? AND status='Confirmed'
      ORDER BY confident_wrong_count DESC,evidence_count DESC,last_seen_at DESC LIMIT ?""",(student_id,limit)).fetchall()

def verified_plan_summary(c, student_id, week_start=None, week_end=None):
    params=[student_id]; where=['student_id=?']
    if week_start and week_end:
        where.append('activity_date BETWEEN ? AND ?'); params.extend([week_start,week_end])
    row=c.execute(f"""SELECT COUNT(*) total,
      SUM(CASE WHEN evidence_status='verified' THEN 1 ELSE 0 END) verified,
      SUM(CASE WHEN evidence_status='self_reported' THEN 1 ELSE 0 END) self_reported,
      SUM(CASE WHEN status='completed' THEN 1 ELSE 0 END) completed
      FROM study_plan_activities WHERE {' AND '.join(where)}""",params).fetchone()
    return {k:int(row[k] or 0) for k in ['total','verified','self_reported','completed']} if row else {'total':0,'verified':0,'self_reported':0,'completed':0}

def live_target_snapshot(c, student_id):
    target=c.execute("""SELECT * FROM student_admission_targets WHERE student_id=? AND active=1 ORDER BY id DESC LIMIT 1""",(student_id,)).fetchone()
    if not target:
        u=c.execute("SELECT goal_name,target_percentage FROM users WHERE id=?",(student_id,)).fetchone()
        if not u or not u['target_percentage']: return None
        target={'route_code':'PK_MEDICAL','institution_name':u['goal_name'] or 'Medical / Dental target','programme_name':'',
                'target_aggregate':u['target_percentage'],'admission_year':datetime.now().year}
    formula=c.execute("""SELECT * FROM admission_formulas WHERE route_code=? AND active=1
      ORDER BY CASE WHEN admission_year=? THEN 0 ELSE 1 END, admission_year DESC,id DESC LIMIT 1""",
      (target['route_code'],target['admission_year'] or datetime.now().year)).fetchone()
    if not formula: return None
    comps=safe_json(formula['components_json'],[]); by_key={x['key']:x for x in comps}
    def latest_like(pattern):
        return c.execute("""SELECT * FROM student_external_results WHERE student_id=? AND exam_type LIKE ? AND percentage IS NOT NULL
          ORDER BY exam_year DESC,id DESC LIMIT 1""",(student_id,pattern)).fetchone()
    matric=latest_like('Matric%')
    fsc_full=c.execute("""SELECT * FROM student_external_results WHERE student_id=? AND exam_type IN ('FSc','FSc Part II') AND percentage IS NOT NULL
      ORDER BY CASE exam_type WHEN 'FSc' THEN 0 ELSE 1 END,exam_year DESC,id DESC LIMIT 1""",(student_id,)).fetchone()
    fsc_part1=latest_like('FSc Part I%')
    fsc=fsc_full or fsc_part1
    mdcat=c.execute("""SELECT percentage FROM student_external_results WHERE student_id=? AND exam_type LIKE 'MDCAT%'
      AND percentage IS NOT NULL ORDER BY exam_year DESC,id DESC LIMIT 1""",(student_id,)).fetchone()
    if not mdcat:
        mdcat=c.execute("""SELECT score percentage FROM attempts WHERE student_id=? AND score IS NOT NULL AND
          (lower(COALESCE(programme,'')) LIKE '%mdcat%' OR lower(COALESCE(exam_title,'')) LIKE '%mdcat%')
          ORDER BY created_at DESC,id DESC LIMIT 1""",(student_id,)).fetchone()
    values={'matric':float(matric['percentage']) if matric else None,'fsc':float(fsc['percentage']) if fsc else None,
            'mdcat':float(mdcat['percentage']) if mdcat else None}
    target_agg=float(target['target_aggregate'] or 0) if target['target_aggregate'] is not None else None
    current=None; required=None; remaining_average=None; scenario=''
    if all(values.get(k) is not None for k in ('matric','fsc','mdcat')):
        current=round(sum(values[k]*float(by_key.get(k,{}).get('weight',0)) for k in ('matric','fsc','mdcat')),2)
    if target_agg is not None and values['matric'] is not None:
        mw=float(by_key.get('matric',{}).get('weight',.10)); fw=float(by_key.get('fsc',{}).get('weight',.40)); ew=float(by_key.get('mdcat',{}).get('weight',.50))
        if fsc_full and values['fsc'] is not None and ew:
            required=round((target_agg-values['matric']*mw-values['fsc']*fw)/ew,2)
            scenario='Required MDCAT score based on your completed Matric and FSc evidence.'
        elif fsc_part1 and values['fsc'] is not None and ew:
            required=round((target_agg-values['matric']*mw-values['fsc']*fw)/ew,2)
            scenario='Illustrative MDCAT requirement if your final FSc percentage remains at your current Part I percentage.'
        else:
            remaining_weight=max(.0001,1-mw)
            remaining_average=round((target_agg-values['matric']*mw)/remaining_weight,2)
            scenario='Average percentage needed across the remaining weighted components after your current Matric result.'
    return {'target':target,'formula':formula,'components':comps,'values':values,'current_projection':current,
            'required_entry_pct':required,'remaining_average_required':remaining_average,
            'gap':round(target_agg-current,2) if target_agg is not None and current is not None else None,
            'scenario_message':scenario,'fsc_is_partial':bool(fsc_part1 and not fsc_full)}


# ---------------------------------------------------------------------------
# V5.5 Assessment Blueprint Integration
# ---------------------------------------------------------------------------

def blueprint_sections(c, blueprint_id):
    return c.execute("""SELECT * FROM assessment_blueprint_sections WHERE blueprint_id=?
      ORDER BY section_order,subject""",(blueprint_id,)).fetchall()


def blueprint_joined(c, blueprint_id):
    return c.execute("""SELECT ab.*,af.name framework_name,af.powerhouse_framework_id,
      afv.version_name framework_version_name,afv.powerhouse_framework_version_id
      FROM assessment_blueprints ab
      JOIN assessment_frameworks af ON af.id=ab.framework_id
      JOIN assessment_framework_versions afv ON afv.id=ab.framework_version_id
      WHERE ab.id=?""",(blueprint_id,)).fetchone()


def active_assessment_blueprint(c, framework_name='', framework_version=''):
    clauses=["ab.local_status='ACTIVE'"]; params=[]
    if framework_name:
        clauses.append("lower(af.name)=lower(?)"); params.append(framework_name)
    if framework_version:
        clauses.append("lower(afv.version_name)=lower(?)"); params.append(framework_version)
    return c.execute(f"""SELECT ab.*,af.name framework_name,af.powerhouse_framework_id,
      afv.version_name framework_version_name,afv.powerhouse_framework_version_id
      FROM assessment_blueprints ab
      JOIN assessment_frameworks af ON af.id=ab.framework_id
      JOIN assessment_framework_versions afv ON afv.id=ab.framework_version_id
      WHERE {' AND '.join(clauses)} ORDER BY ab.activated_at DESC,ab.id DESC LIMIT 1""",params).fetchone()


def blueprint_payload_from_record(c, blueprint_id):
    bp=blueprint_joined(c,blueprint_id)
    if not bp:
        return None
    payload=safe_json(bp['immutable_payload_json'],{})
    if payload:
        return payload
    return {
      'schema_version':'1.0','blueprint_id':bp['powerhouse_blueprint_id'],
      'framework':{'id':bp['powerhouse_framework_id'],'name':bp['framework_name']},
      'framework_version':{'id':bp['powerhouse_framework_version_id'],'name':bp['framework_version_name']},
      'blueprint_version':bp['blueprint_version'],'status':bp['source_status'],'authority':bp['authority'],
      'total_questions':bp['total_questions'],'duration_minutes':bp['duration_minutes'],
      'sections':[{'subject':r['subject'],'question_count':r['question_count'],'weight_percent':r['weight_percent'],
                   'section_order':r['section_order']} for r in blueprint_sections(c,blueprint_id)]}


def active_assembly_policy(c, blueprint_id=None, framework_version_id=None, programme='', subject='', chapter='', assessment_type=''):
    """Resolve the most specific ACTIVE rigor/mastery policy for a future assessment.

    The official blueprint and the ScoreMax assembly policy remain separate.
    Policy changes affect only future forms and are pinned into each session/result.
    """
    candidates=[]
    if chapter:
        if programme or subject:
            candidates.append(('chapter','|'.join(x for x in (programme,subject,chapter) if x)))
        candidates.append(('chapter',chapter))
    if subject:
        if programme:
            candidates.append(('subject',f'{programme}|{subject}'))
        candidates.append(('subject',subject))
    if assessment_type:
        candidates.append(('assessment_type',assessment_type))
    if blueprint_id:
        candidates.append(('blueprint',str(blueprint_id)))
    if framework_version_id:
        candidates.append(('framework_version',str(framework_version_id)))
    if programme:
        candidates.append(('programme',programme))
    candidates.append(('global',''))
    for scope_type,scope_key in candidates:
        row=c.execute("""SELECT * FROM assessment_assembly_policies WHERE status='ACTIVE'
          AND scope_type=? AND lower(COALESCE(scope_key,''))=lower(?)
          ORDER BY approved_at DESC,id DESC LIMIT 1""",(scope_type,scope_key)).fetchone()
        if row:
            return row
    return None


def effective_mastery_requirements(base_policy, assembly_policy=None):
    """Create a transparent, pinned mastery standard without relabelling questions.

    Score 50 preserves the underlying mastery policy. Moving the standard adjusts
    future evidence volume, accuracy, target-band depth, novelty and verification
    interval within bounded academic guardrails.
    """
    if not base_policy:
        return {}
    base=dict(base_policy)
    standard=int(assembly_policy['mastery_standard_score'] or 50) if assembly_policy else 50
    delta=(max(0,min(100,standard))-50)/50.0
    min_accuracy=max(50.0,min(98.0,float(base.get('min_accuracy') or 0)+5.0*delta))
    min_questions=max(5,int(round(float(base.get('min_questions') or 10)*(1.0+0.25*delta))))
    min_forms=max(1,int(base.get('min_forms') or 1)+(1 if delta>=0.55 else (-1 if delta<=-0.75 else 0)))
    target_band=max(0.10,min(0.90,float(base.get('target_band_pct') or .25)+0.15*delta))
    unseen=max(0.20,min(0.98,float(base.get('unseen_family_pct') or .60)+0.20*delta))
    verification=max(14,int(round(float(base.get('verification_days') or 90)*(1.0-0.25*delta))))
    return {
      'level':base.get('level',''),'min_accuracy':round(min_accuracy,2),'min_questions':min_questions,
      'min_forms':min_forms,'target_band_pct':round(target_band,3),'unseen_family_pct':round(unseen,3),
      'verification_days':verification,'mastery_standard_score':standard,
      'base_mastery_policy_id':base.get('id'),'assembly_policy_id':assembly_policy['id'] if assembly_policy else None,
      'assembly_policy_version':assembly_policy['policy_version'] if assembly_policy else '1'
    }


def simulate_policy_impact(c, blueprint, mastery_standard_score=50):
    """Estimate form-level impact from real historical mastery evidence.

    This is deliberately transparent and conservative: it re-evaluates stored formal
    mastery-form outcomes against the proposed evidence thresholds. It does not claim
    to predict external exam percentiles and does not rewrite any historical record.
    """
    if not blueprint:
        return {'observed_forms':0,'levels':[],'confidence':'No blueprint selected','external_percentile_claim':False}
    fake={'id':None,'policy_version':'PREVIEW','mastery_standard_score':int(mastery_standard_score or 50)}
    rows=c.execute("""SELECT * FROM mastery_form_results
      WHERE demo_only=0 AND lower(COALESCE(programme,''))=lower(?) ORDER BY created_at""",
      (blueprint['framework_name'],)).fetchall()
    grouped={}
    for row in rows:
        grouped.setdefault(row['target_level'] or 'Foundation',[]).append(row)
    output=[]
    for level in MASTER_LEVELS:
        evidence=grouped.get(level,[])
        if not evidence: continue
        base=c.execute("SELECT * FROM mastery_policies WHERE level=?",(level,)).fetchone()
        effective=effective_mastery_requirements(base,fake) if base else {}
        proposed=0
        for row in evidence:
            meets=(float(row['score'] or 0)>=float(effective.get('min_accuracy') or 0)
                   and int(row['question_count'] or 0)>=int(effective.get('min_questions') or 0)
                   and int(row['breadth_ok'] or 0)==1
                   and float(row['unseen_family_ratio'] or 0)>=float(effective.get('unseen_family_pct') or 0))
            proposed+=1 if meets else 0
        output.append({'level':level,'observed_forms':len(evidence),
          'current_form_pass_rate':round(100.0*sum(int(r['passed'] or 0) for r in evidence)/len(evidence),1),
          'estimated_proposed_form_pass_rate':round(100.0*proposed/len(evidence),1),
          'proposed_min_accuracy':effective.get('min_accuracy'),
          'proposed_min_questions':effective.get('min_questions'),
          'proposed_unseen_family_pct':effective.get('unseen_family_pct')})
    n=len(rows)
    confidence='Low — fewer than 30 real forms' if n<30 else ('Moderate' if n<150 else 'Higher')
    return {'observed_forms':n,'levels':output,'confidence':confidence,
      'external_percentile_claim':False,
      'note':'Historical ScoreMax form re-evaluation only; external exam-outcome calibration remains a later evidence phase.'}

def record_blueprint_audit(c, blueprint_id, action, previous_status='', new_status='', reason='', impact=None, actor_user_id=None):
    c.execute("""INSERT INTO assessment_blueprint_audit(
      blueprint_id,action,actor_user_id,previous_status,new_status,reason,impact_snapshot_json)
      VALUES(?,?,?,?,?,?,?)""",(blueprint_id,action,actor_user_id,previous_status,new_status,reason,json.dumps(impact or {})))


def record_policy_audit(c, policy_id, action, previous_status='', new_status='', reason='', snapshot=None, actor_user_id=None):
    c.execute("""INSERT INTO assessment_policy_audit(
      policy_id,action,actor_user_id,previous_status,new_status,reason,snapshot_json)
      VALUES(?,?,?,?,?,?,?)""",(policy_id,action,actor_user_id,previous_status,new_status,reason,json.dumps(snapshot or {})))


def _powerhouse_blueprint_payload(blueprint):
    if not blueprint or 'immutable_payload_json' not in blueprint.keys(): return {}
    p=safe_json(blueprint['immutable_payload_json'],{})
    return p if isinstance(p,dict) and p.get('release_state')=='RELEASED' and p.get('blueprint_id') else {}


def _ph_blueprint_scope_match(q,rule):
    scope_type=str((rule or {}).get('scope_type') or '').upper(); scope_id=str((rule or {}).get('scope_id') or '')
    if not scope_type or not scope_id: return False
    if scope_type=='MARKET': return str(q['ph_market_id'] or '')==scope_id
    if scope_type=='PROGRAMME': return str(q['ph_programme_id'] or '')==scope_id
    if scope_type=='SUBJECT': return str(q['ph_subject_id'] or '')==scope_id
    if scope_type=='CHAPTER': return str(q['ph_chapter_id'] or '')==scope_id
    curr=safe_json(q['ph_curriculum_snapshot_json'],{}) if 'ph_curriculum_snapshot_json' in q.keys() else {}
    if scope_type=='SECTION': return str(curr.get('section_id') or '')==scope_id
    if scope_type=='TOPIC': return str(curr.get('topic_id') or '')==scope_id
    if scope_type=='SUBTOPIC': return str(curr.get('subtopic_id') or '')==scope_id
    if scope_type in {'LEARNING_OUTCOME','OUTCOME'}: return scope_id in {str(x) for x in (curr.get('learning_outcome_ids') or [])}
    if scope_type in {'TEACHING_LEARNING_OUTCOME','TLO'}: return scope_id in {str(x) for x in (curr.get('teaching_learning_outcome_ids') or [])}
    return False


def blueprint_question_pool(c, blueprint, subject, student_id=None):
    ph_payload=_powerhouse_blueprint_payload(blueprint)
    if ph_payload:
        params=[str(ph_payload.get('market_id') or ''),str(ph_payload.get('programme_id') or ''),str(ph_payload.get('subject_id') or '')]
        release_ids=[str(x) for x in (ph_payload.get('permitted_release_ids') or []) if str(x)]
        release_clause=''
        if release_ids:
            release_clause=' AND q.ph_release_id IN ('+','.join('?' for _ in release_ids)+')'; params.extend(release_ids)
        rows=c.execute(f"""SELECT q.* FROM questions q WHERE {live_question_clause('q')}
          AND q.ph_projection_owner='POWER_HOUSE' AND q.ph_market_id=? AND q.ph_programme_id=? AND q.ph_subject_id=? {release_clause}
          AND COALESCE(q.scoremax_ready,1)=1
          AND lower(COALESCE(q.rights_status,'')) IN
            ('owned','commissioned_ip_assigned','licensed_commercial','open_commercial','public_domain','scoremax original','licensed','permitted','public domain','approved')
          ORDER BY COALESCE(q.calibration_status,''),COALESCE(q.response_count,0) DESC,q.id""",params).fetchall()
        governed_sections=list(ph_payload.get('sections') or [])
        coverage=[r for sec in governed_sections for r in (sec.get('coverage_rules') or [])]
        type_rules=[r for sec in governed_sections for r in (sec.get('question_type_rules') or [])]
        allowed_types={integration_v1.SUPPORTED_BLUEPRINT_QUESTION_TYPES.get(str(r.get('question_type') or '').upper()) for r in type_rules}
        allowed_types.discard(None)
        rows=[r for r in rows if canonical_question_type(r) in LIVE_MARKABLE_TYPES]
        if allowed_types: rows=[r for r in rows if canonical_question_type(r) in allowed_types]
        if coverage: rows=[r for r in rows if any(_ph_blueprint_scope_match(r,rule) for rule in coverage)]
    else:
        framework=(blueprint['framework_name'] or '').strip() if blueprint else ''
        params=[subject]
        framework_clause=''
        if framework:
            framework_clause=" AND (lower(COALESCE(q.programme,''))=lower(?) OR lower(COALESCE(q.qualification,''))=lower(?) OR lower(COALESCE(q.programme,'')) LIKE lower(?))"
            params.extend([framework,framework,f'%{framework}%'])
        rows=c.execute(f"""SELECT q.* FROM questions q WHERE {live_question_clause('q')}
          AND q.subject=? {framework_clause}
          AND COALESCE(q.scoremax_ready,1)=1
          AND (COALESCE(q.assessment_purpose,'')='' OR lower(q.assessment_purpose) LIKE '%mock%')
          AND lower(COALESCE(q.rights_status,'ScoreMax Original')) IN
            ('scoremax original','licensed','permitted','public domain','approved')
          ORDER BY COALESCE(q.calibration_status,''),COALESCE(q.response_count,0) DESC,q.id""",params).fetchall()
        rows=[r for r in rows if canonical_question_type(r) in LIVE_MARKABLE_TYPES]
    if student_id:
        seen={r['question_db_id'] for r in c.execute("""SELECT DISTINCT aa.question_db_id FROM attempt_answers aa
          JOIN attempts a ON a.id=aa.attempt_id WHERE a.student_id=?""",(student_id,)).fetchall()}
        rows=sorted(rows,key=lambda r:(r['id'] in seen, -(int(r['response_count'] or 0)), r['id']))
    return rows


def _distribution_quotas(dist,total):
    values={str(k).upper():max(0.0,float(v or 0)) for k,v in (dist or {}).items()}
    if not values: return {}
    denom=sum(values.values())
    if denom<=0: return {}
    raw={k:total*v/denom for k,v in values.items()}; out={k:int(v) for k,v in raw.items()}
    while sum(out.values())<total:
        k=max(out,key=lambda x:(raw[x]-out[x],values[x],x)); out[k]+=1
    while sum(out.values())>total:
        choices=[k for k,v in out.items() if v>0]
        k=min(choices,key=lambda x:(raw[x]-out[x],-out[x],x)); out[k]-=1
    return out


def _select_powerhouse_blueprint_questions(rows,payload,seed=''):
    """Deterministically satisfy PH section coverage/type/mastery/cognitive rules or fail closed."""
    rng=random.Random(str(seed)); unused={r['id']:r for r in rows}; selected=[]; reports=[]; blockers=[]; position=1
    for sec in sorted(list(payload.get('sections') or []),key=lambda x:(int(x.get('order') or 0),str(x.get('section_id') or ''))):
        required=int(sec.get('question_count') or 0); coverage=list(sec.get('coverage_rules') or []); qrules=list(sec.get('question_type_rules') or [])
        mastery_quota=_distribution_quotas(sec.get('mastery_distribution') or {},required)
        cognitive_quota=_distribution_quotas(sec.get('cognitive_demand_distribution') or {},required)
        candidates=[]
        for r in unused.values():
            if coverage and not any(_ph_blueprint_scope_match(r,x) for x in coverage): continue
            qt=canonical_question_type(r)
            if qrules:
                allowed={integration_v1.SUPPORTED_BLUEPRINT_QUESTION_TYPES.get(str(x.get('question_type') or '').upper()) for x in qrules}
                if qt not in allowed: continue
            mastery=str(r['level'] or '').upper().replace(' ','_').replace('-','_')
            cog=str(r['ph_cognitive_demand'] or '').upper() if 'ph_cognitive_demand' in r.keys() else ''
            if mastery_quota and mastery not in mastery_quota: continue
            if cognitive_quota and cog not in cognitive_quota: continue
            candidates.append(r)
        rng.shuffle(candidates)
        candidates.sort(key=lambda r:(0 if (r['calibration_status'] or '').upper() in ('CALIBRATED','OPERATIONAL') else 1,
                                      -(int(r['response_count'] or 0)),r['id']))
        cov_counts=[0]*len(coverage); type_counts=[0]*len(qrules); mastery_counts={k:0 for k in mastery_quota}; cognitive_counts={k:0 for k in cognitive_quota}; picks=[]
        while len(picks)<required:
            best=None; best_score=None
            for r in candidates:
                if r['id'] not in unused: continue
                cov_hits=[i for i,x in enumerate(coverage) if _ph_blueprint_scope_match(r,x)]
                qt=canonical_question_type(r); type_hits=[i for i,x in enumerate(qrules) if integration_v1.SUPPORTED_BLUEPRINT_QUESTION_TYPES.get(str(x.get('question_type') or '').upper())==qt]
                if coverage and not cov_hits: continue
                if qrules and not type_hits: continue
                mastery=str(r['level'] or '').upper().replace(' ','_').replace('-','_')
                cog=str(r['ph_cognitive_demand'] or '').upper() if 'ph_cognitive_demand' in r.keys() else ''
                exceeds=False
                for i in cov_hits:
                    mx=int(coverage[i].get('maximum_questions') if coverage[i].get('maximum_questions') is not None else required)
                    if cov_counts[i]>=mx: exceeds=True
                for i in type_hits:
                    mx=int(qrules[i].get('maximum') if qrules[i].get('maximum') is not None else required)
                    if type_counts[i]>=mx: exceeds=True
                if mastery_quota and mastery_counts.get(mastery,0)>=mastery_quota.get(mastery,0): exceeds=True
                if cognitive_quota and cognitive_counts.get(cog,0)>=cognitive_quota.get(cog,0): exceeds=True
                if exceeds: continue
                deficit=0
                for i in cov_hits: deficit += max(0,int(coverage[i].get('minimum_questions') or 0)-cov_counts[i])*10
                for i in type_hits: deficit += max(0,int(qrules[i].get('minimum') or 0)-type_counts[i])*10
                if mastery_quota: deficit += max(0,mastery_quota.get(mastery,0)-mastery_counts.get(mastery,0))*8
                if cognitive_quota: deficit += max(0,cognitive_quota.get(cog,0)-cognitive_counts.get(cog,0))*8
                fam=str(r['family_key'] or r['family_id'] or r['id']); family_repeat=sum(1 for x in picks if str(x['family_key'] or x['family_id'] or x['id'])==fam)
                score=(deficit,-family_repeat,int(r['response_count'] or 0),-int(r['id']))
                if best is None or score>best_score: best=(r,cov_hits,type_hits,mastery,cog); best_score=score
            if best is None: break
            r,cov_hits,type_hits,mastery,cog=best; picks.append(r); unused.pop(r['id'],None)
            for i in cov_hits: cov_counts[i]+=1
            for i in type_hits: type_counts[i]+=1
            if mastery_quota: mastery_counts[mastery]=mastery_counts.get(mastery,0)+1
            if cognitive_quota: cognitive_counts[cog]=cognitive_counts.get(cog,0)+1
        failures=[]
        if len(picks)!=required: failures.append(f'selected {len(picks)} of {required}')
        for i,rule in enumerate(coverage):
            mn=int(rule.get('minimum_questions') or 0); mx=int(rule.get('maximum_questions') if rule.get('maximum_questions') is not None else required)
            if not (mn<=cov_counts[i]<=mx): failures.append(f"coverage {rule.get('scope_type')}::{rule.get('scope_id')}={cov_counts[i]} outside {mn}-{mx}")
        for i,rule in enumerate(qrules):
            mn=int(rule.get('minimum') or 0); mx=int(rule.get('maximum') if rule.get('maximum') is not None else required)
            if not (mn<=type_counts[i]<=mx): failures.append(f"type {rule.get('question_type')}={type_counts[i]} outside {mn}-{mx}")
        if mastery_quota and mastery_counts!=mastery_quota: failures.append(f'mastery distribution {mastery_counts} != {mastery_quota}')
        if cognitive_quota and cognitive_counts!=cognitive_quota: failures.append(f'cognitive distribution {cognitive_counts} != {cognitive_quota}')
        if failures:
            blockers.append(f"{sec.get('section_id') or sec.get('display_name')}: "+'; '.join(failures))
            for r in picks: unused[r['id']]=r
            continue
        correct_marks=float((payload.get('marking_rules') or {}).get('correct_marks') or 0)
        section_label=str(sec.get('display_name') or sec.get('section_id') or 'Section')
        for r in picks:
            selected.append({'question_id':r['id'],'position':position,'section_label':section_label,'display_number':str(position),
                             'marks':correct_marks if correct_marks>0 else float(r['marks'] or 1),'subject':str(payload.get('subject_id') or r['subject'])})
            position+=1
        reports.append({'section_id':sec.get('section_id'),'display_name':sec.get('display_name'),'required':required,'available':len(candidates),
                        'selected':len(picks),'coverage_counts':cov_counts,'question_type_counts':type_counts,
                        'mastery_counts':mastery_counts,'mastery_quotas':mastery_quota,'cognitive_counts':cognitive_counts,'cognitive_quotas':cognitive_quota,'blocking':False})
    required_total=sum(int(x.get('question_count') or 0) for x in (payload.get('sections') or []))
    if len(selected)!=required_total: blockers.append(f'Selected total is {len(selected)}; governed blueprint requires {required_total}.')
    return {'ready':not blockers,'blockers':blockers,'selected':selected,'sections':reports,'required_total':required_total,'selected_total':len(selected)}


def _select_subject_questions(rows, required, target_mix, seed=''):
    rng=random.Random(str(seed))
    by_diff={'Easy':[],'Moderate':[],'Difficult':[]}
    for r in rows:
        by_diff[normalize_difficulty(r['difficulty'] or r['level'])].append(r)
    for bucket in by_diff.values():
        rng.shuffle(bucket)
        bucket.sort(key=lambda r:(0 if (r['calibration_status'] or '').upper() in ('CALIBRATED','OPERATIONAL') else 1,
                                  -(int(r['response_count'] or 0))))
    quotas=allocate_counts(required,target_mix)
    selected=[]; used_families=set(); shortages={}
    # First pass: exact difficulty mix and one variant per family.
    for diff in ('Difficult','Moderate','Easy'):
        need=quotas.get(diff,0)
        for r in by_diff[diff]:
            family=(r['family_key'] or r['family_id'] or f"Q-{r['id']}")
            if family in used_families: continue
            selected.append(r); used_families.add(family)
            if sum(1 for x in selected if normalize_difficulty(x['difficulty'] or x['level'])==diff)>=need: break
        got=sum(1 for x in selected if normalize_difficulty(x['difficulty'] or x['level'])==diff)
        if got<need: shortages[diff]=need-got
    # Second pass: preserve family diversity while filling any mix shortage.
    for r in rows:
        if len(selected)>=required: break
        family=(r['family_key'] or r['family_id'] or f"Q-{r['id']}")
        if family in used_families: continue
        selected.append(r); used_families.add(family)
    # Last resort: a second variant can fill a subject count, but preflight records it.
    repeated=0
    chosen_ids={r['id'] for r in selected}
    for r in rows:
        if len(selected)>=required: break
        if r['id'] in chosen_ids: continue
        selected.append(r); chosen_ids.add(r['id']); repeated+=1
    return selected[:required],{'difficulty_quotas':quotas,'difficulty_shortages':shortages,
                                'repeated_family_variants':repeated,'distinct_families':len(used_families)}


def blueprint_bank_sufficiency(c, blueprint_id, target_parallel_mocks=3):
    bp=blueprint_joined(c,blueprint_id)
    if not bp: return None
    policy=active_assembly_policy(c,bp['id'],bp['framework_version_id'])
    rigor=int(policy['rigor_score'] or 50) if policy else 50
    official_mix=safe_json(bp['difficulty_distribution_json'],{})
    target_mix=rigor_mix(rigor,official_mix)
    subjects=[]; blockers=[]
    for section in blueprint_sections(c,blueprint_id):
        rows=blueprint_question_pool(c,bp,section['subject'])
        families={r['family_key'] or r['family_id'] or f"Q-{r['id']}" for r in rows}
        chapters={r['chapter'] for r in rows if (r['chapter'] or '').strip()}
        los={r['learning_outcome'] for r in rows if (r['learning_outcome'] or '').strip()}
        diff_counts={'Easy':0,'Moderate':0,'Difficult':0}
        for r in rows: diff_counts[normalize_difficulty(r['difficulty'] or r['level'])]+=1
        required=int(section['question_count'] or 0)
        safe_by_items=(len(rows)//required) if required else 0
        # 70% distinct-family expectation gives a conservative safe-form estimate.
        family_need=max(1,int(round(required*0.70))) if required else 1
        safe_by_families=(len(families)//family_need) if family_need else 0
        safe_forms=min(safe_by_items,safe_by_families)
        shortage=max(0,required*int(target_parallel_mocks)-len(rows))
        if len(rows)<required:
            blockers.append(f"{section['subject']} has {len(rows)} usable questions; {required} are required for one authentic mock.")
        subjects.append({'subject':section['subject'],'required_per_mock':required,'weight_percent':section['weight_percent'],
                         'usable_questions':len(rows),'family_count':len(families),'variant_count':len(rows),
                         'chapter_count':len(chapters),'learning_outcome_count':len(los),'difficulty_counts':diff_counts,
                         'safe_parallel_forms':safe_forms,'target_parallel_mocks':int(target_parallel_mocks),
                         'shortage_for_target':shortage,'status':'Blocked' if len(rows)<required else ('Thin' if safe_forms<target_parallel_mocks else 'Ready')})
    return {'blueprint':bp,'subjects':subjects,'blockers':blockers,'ready':not blockers,
            'target_parallel_mocks':int(target_parallel_mocks),'target_difficulty_mix':target_mix,
            'policy':policy}


def assemble_blueprint_mock(c, blueprint_id, student_id=None, policy_id=None, seed=''):
    bp=blueprint_joined(c,blueprint_id)
    if not bp: raise ValueError('Blueprint not found.')
    if bp['local_status']!='ACTIVE': raise ValueError('Only an ACTIVE blueprint can govern an authentic mock.')
    policy=c.execute("SELECT * FROM assessment_assembly_policies WHERE id=?",(policy_id,)).fetchone() if policy_id else active_assembly_policy(c,bp['id'],bp['framework_version_id'])
    rigor=int(policy['rigor_score'] or 50) if policy else 50
    official_mix=safe_json(bp['difficulty_distribution_json'],{})
    target_mix=rigor_mix(rigor,official_mix)
    ph_payload=_powerhouse_blueprint_payload(bp)
    if ph_payload:
        rows=blueprint_question_pool(c,bp,str(ph_payload.get('subject_id') or ''),student_id=student_id)
        governed=_select_powerhouse_blueprint_questions(rows,ph_payload,seed=f'{seed}|PH|{bp["id"]}')
        return {'ready':governed['ready'],'blockers':governed['blockers'],'warnings':[],
                'selected':governed['selected'],'sections':governed['sections'],
                'required_total':governed['required_total'],'selected_total':governed['selected_total'],
                'blueprint_id':bp['id'],'blueprint_source_id':bp['powerhouse_blueprint_id'],'blueprint_version':bp['blueprint_version'],
                'framework_version':bp['framework_version_name'],'policy_id':None,'policy_version':'PH_GOVERNED',
                'rigor_score':None,'target_difficulty_mix':{},'blueprint_snapshot':ph_payload}
    selected=[]; section_reports=[]; blockers=[]; warnings=[]; position=1
    for section in blueprint_sections(c,blueprint_id):
        required=int(section['question_count'] or 0)
        rows=blueprint_question_pool(c,bp,section['subject'],student_id=student_id)
        if len(rows)<required:
            blockers.append(f"{section['subject']}: requires {required}, only {len(rows)} governed ScoreMax-ready questions are available.")
            section_reports.append({'subject':section['subject'],'required':required,'available':len(rows),'selected':0,'blocking':True})
            continue
        section_mix=safe_json(section['difficulty_distribution_json'],{}) or official_mix
        mix=rigor_mix(rigor,section_mix)
        picks,detail=_select_subject_questions(rows,required,mix,seed=f"{seed}|{bp['id']}|{section['subject']}")
        if len(picks)<required:
            blockers.append(f"{section['subject']}: selection produced {len(picks)} of {required} required questions.")
            continue
        if detail['difficulty_shortages']:
            warnings.append(f"{section['subject']} could not exactly match the intended difficulty mix: {detail['difficulty_shortages']}.")
        if detail['repeated_family_variants']:
            warnings.append(f"{section['subject']} required {detail['repeated_family_variants']} additional variant(s) from already-used families.")
        for q in picks:
            selected.append({'question_id':q['id'],'position':position,'section_label':section['section_title'] or section['subject'],
                             'display_number':str(position),'marks':float(q['marks'] or 1),'subject':section['subject']})
            position+=1
        section_reports.append({'subject':section['subject'],'required':required,'available':len(rows),'selected':len(picks),
                                'family_count':detail['distinct_families'],'difficulty_mix':mix,
                                'difficulty_quotas':detail['difficulty_quotas'],'difficulty_shortages':detail['difficulty_shortages'],
                                'repeated_family_variants':detail['repeated_family_variants'],'blocking':False})
    if len(selected)!=int(bp['total_questions'] or 0):
        blockers.append(f"Selected total is {len(selected)}; blueprint requires {bp['total_questions']}.")
    return {'ready':not blockers,'blockers':blockers,'warnings':warnings,'selected':selected,
            'sections':section_reports,'required_total':int(bp['total_questions'] or 0),'selected_total':len(selected),
            'blueprint_id':bp['id'],'blueprint_source_id':bp['powerhouse_blueprint_id'],'blueprint_version':bp['blueprint_version'],
            'framework_version':bp['framework_version_name'],'policy_id':policy['id'] if policy else None,
            'policy_version':policy['policy_version'] if policy else '1','rigor_score':rigor,'target_difficulty_mix':target_mix}


def _allocate_blueprint_subject_counts(total_count, sections, shares=None):
    """Largest-remainder allocation that preserves the requested test length."""
    total=max(1,int(total_count or 1))
    rows=list(sections or [])
    if not rows:
        return {}
    weights={}
    for row in rows:
        subject=row['subject']
        weights[subject]=max(0.0,float((shares or {}).get(subject,row['weight_percent'] or 0)))
    weight_total=sum(weights.values()) or float(len(rows))
    raw={subject:total*weight/weight_total for subject,weight in weights.items()}
    allocated={subject:int(value) for subject,value in raw.items()}
    if total>=len(rows):
        for row in rows:
            if allocated[row['subject']]==0:
                allocated[row['subject']]=1
    while sum(allocated.values())>total:
        choices=[k for k,v in allocated.items() if v>1 or total<len(rows)]
        if not choices: break
        key=min(choices,key=lambda k:(raw[k]-allocated[k],-allocated[k]))
        allocated[key]-=1
    while sum(allocated.values())<total:
        key=max(allocated,key=lambda k:(raw[k]-allocated[k],weights[k]))
        allocated[key]+=1
    return allocated


def assemble_blueprint_practice(c, blueprint_id, student_id, total_count=60, practice_type='proportional_full', seed=''):
    """Assemble clearly-labelled blueprint-aware practice without impersonating an authentic mock.

    Authentic mocks never substitute between subjects. Non-authentic diagnostic/adaptive/proportional
    practice may transparently rebalance a shortage to other blueprint subjects, but the adjusted
    composition is stored and reported so it can never be mistaken for an official exam form.
    """
    bp=blueprint_joined(c,blueprint_id)
    if not bp or bp['local_status']!='ACTIVE':
        raise ValueError('An active assessment blueprint is required.')
    ph_payload=_powerhouse_blueprint_payload(bp)
    if ph_payload:
        access=get_access_profile(c,student_id)
        rows=[r for r in blueprint_question_pool(c,bp,str(ph_payload.get('subject_id') or ''),student_id=student_id)
              if mastery_rank(r['level'] or 'Foundation')<=mastery_rank(access['mastery_ceiling'])]
        governed=_select_powerhouse_blueprint_questions(rows,ph_payload,seed=f'{seed}|PH-PRACTICE|{student_id}')
        duration_minutes=max(1,(int(ph_payload.get('total_duration_seconds') or 0)+59)//60) if int(ph_payload.get('total_duration_seconds') or 0)>0 else None
        return {'ready':governed['ready'],'blockers':governed['blockers'],'warnings':[],
          'selected':governed['selected'],'sections':governed['sections'],'practice_type':'powerhouse_governed',
          'required_total':governed['required_total'],'selected_total':governed['selected_total'],
          'allocations':{str(ph_payload.get('subject_id') or ''):governed['required_total']},
          'official_allocations':{str(ph_payload.get('subject_id') or ''):governed['required_total']},
          'blueprint_id':bp['id'],'blueprint_source_id':bp['powerhouse_blueprint_id'],'blueprint_version':bp['blueprint_version'],
          'framework_version':bp['framework_version_name'],'blueprint_snapshot':ph_payload,
          'assembly_policy_id':None,'assembly_policy_version':'PH_GOVERNED','subject_policies':{},'priority_snapshot':{},
          'authenticity_status':'POWER_HOUSE_GOVERNED_BLUEPRINT_PRACTICE','duration_minutes':duration_minutes}
    practice_type=(practice_type or 'proportional_full').strip().lower()
    allowed={'proportional_full','proportional_half','diagnostic','adaptive'}
    if practice_type not in allowed:
        practice_type='proportional_full'
    total=max(10,min(int(bp['total_questions'] or 180),int(total_count or 60)))
    sections=blueprint_sections(c,blueprint_id)
    shares=None
    priority=None
    if practice_type in ('diagnostic','adaptive'):
        priority=blueprint_priority_snapshot(c,student_id,blueprint_id)
        priority_map={r['subject']:max(1.0,float(r['priority_score'] or 0)) for r in priority.get('subjects',[])}
        shares={}
        for sec in sections:
            official=float(sec['weight_percent'] or 0)
            need=priority_map.get(sec['subject'],1.0)
            shares[sec['subject']]=0.45*official+0.55*need
    allocations=_allocate_blueprint_subject_counts(total,sections,shares)
    access=get_access_profile(c,student_id)
    overall_policy=active_assembly_policy(c,bp['id'],bp['framework_version_id'],programme=bp['framework_name'],assessment_type=practice_type)
    selected=[]; reports=[]; blockers=[]; warnings=[]; position=1; subject_policies={}

    # Pre-build eligible pools so a non-authentic practice set can be transparently rebalanced
    # when one subject is currently under-stocked for this learner's Access ceiling.
    pools={}
    capacities={}
    for section in sections:
        subject=section['subject']
        rows=[r for r in blueprint_question_pool(c,bp,subject,student_id=student_id)
              if mastery_rank(r['level'] or 'Foundation')<=mastery_rank(access['mastery_ceiling'])]
        pools[subject]=rows
        capacities[subject]=len(rows)
    if sum(capacities.values())<total:
        blockers.append(f'Only {sum(capacities.values())} eligible governed questions are available across the blueprint for this learner; {total} are required.')

    adjusted=dict(allocations)
    shortfall=0
    for subject,required in list(adjusted.items()):
        cap=capacities.get(subject,0)
        if required>cap:
            shortfall+=required-cap
            adjusted[subject]=cap
            warnings.append(f'{subject} allocation reduced from {required} to {cap} because the current governed bank/Access capacity is limited.')

    if shortfall and not blockers:
        if practice_type in ('diagnostic','adaptive'):
            priority_order=[r['subject'] for r in (priority or {}).get('subjects',[])]
        else:
            priority_order=[r['subject'] for r in sorted(sections,key=lambda x:float(x['weight_percent'] or 0),reverse=True)]
        # Keep all blueprint subjects eligible, even if a sparse priority snapshot omitted one.
        priority_order+= [s['subject'] for s in sections if s['subject'] not in priority_order]
        while shortfall>0:
            progressed=False
            for subject in priority_order:
                if adjusted.get(subject,0)<capacities.get(subject,0):
                    adjusted[subject]=adjusted.get(subject,0)+1
                    shortfall-=1; progressed=True
                    if shortfall<=0: break
            if not progressed: break
        if shortfall:
            blockers.append(f'Practice assembly remains {shortfall} question(s) short after transparent redistribution.')
        else:
            warnings.append('Because this is a clearly labelled non-authentic practice set, subject allocation was transparently adjusted for current bank/Access capacity. The official blueprint itself was not changed.')
    allocations=adjusted

    for section in sections:
        subject=section['subject']; required=int(allocations.get(subject,0))
        if required<=0: continue
        rows=pools.get(subject,[])
        policy=active_assembly_policy(c,bp['id'],bp['framework_version_id'],programme=bp['framework_name'],subject=subject,assessment_type=practice_type) or overall_policy
        rigor=int(policy['rigor_score'] or 50) if policy else 50
        mix=rigor_mix(rigor,safe_json(section['difficulty_distribution_json'],{}) or safe_json(bp['difficulty_distribution_json'],{}))
        picks,detail=_select_subject_questions(rows,required,mix,seed=f'{seed}|{practice_type}|{subject}|{student_id}')
        if len(picks)<required:
            blockers.append(f'{subject}: requires {required} eligible questions for this practice set; only {len(picks)} could be selected at the current Access/policy level.')
            reports.append({'subject':subject,'required':required,'selected':len(picks),'blocking':True})
            continue
        subject_policies[subject]={'policy_id':policy['id'] if policy else None,'policy_version':policy['policy_version'] if policy else '1','rigor_score':rigor}
        for q in picks:
            selected.append({'question_id':q['id'],'position':position,'subject':subject})
            position+=1
        reports.append({'subject':subject,'required':required,'selected':len(picks),'difficulty_mix':mix,'detail':detail,'blocking':False})
        if detail.get('difficulty_shortages'):
            warnings.append(f'{subject} could not exactly meet the intended difficulty composition.')
    if len(selected)!=total:
        blockers.append(f'Practice assembly selected {len(selected)} of {total} required questions.')
    snapshot=blueprint_payload_from_record(c,blueprint_id) or {}
    return {'ready':not blockers,'blockers':blockers,'warnings':warnings,'selected':selected,'sections':reports,
      'practice_type':practice_type,'required_total':total,'selected_total':len(selected),'allocations':allocations,
      'official_allocations':_allocate_blueprint_subject_counts(total,sections,shares),
      'blueprint_id':bp['id'],'blueprint_source_id':bp['powerhouse_blueprint_id'],'blueprint_version':bp['blueprint_version'],
      'framework_version':bp['framework_version_name'],'blueprint_snapshot':snapshot,
      'assembly_policy_id':overall_policy['id'] if overall_policy else None,
      'assembly_policy_version':overall_policy['policy_version'] if overall_policy else '1',
      'subject_policies':subject_policies,'priority_snapshot':priority or {},
      'authenticity_status':'DIAGNOSTIC_BLUEPRINT_AWARE' if practice_type in ('diagnostic','adaptive') else 'PROPORTIONAL_BLUEPRINT_PRACTICE'}


def blueprint_priority_snapshot(c, student_id, blueprint_id, exam_date=''):
    bp=blueprint_joined(c,blueprint_id)
    if not bp: return {'subjects':[]}
    days=max(1,exam_days_remaining(exam_date) or 180)
    urgency=max(0.0,min(1.0,(90-days)/90.0))
    weak_by_subject={}
    for w in student_weak_areas(c,student_id,limit=100):
        weak_by_subject.setdefault(w['subject'] or '',[]).append(w)
    rows=[]
    for sec in blueprint_sections(c,blueprint_id):
        perf=c.execute("""SELECT COUNT(aa.id) answered,100.0*SUM(aa.is_correct)/NULLIF(COUNT(aa.id),0) accuracy,
          COUNT(DISTINCT q.chapter) chapters FROM attempts a JOIN attempt_answers aa ON aa.attempt_id=a.id
          JOIN questions q ON q.id=aa.question_db_id WHERE a.student_id=? AND q.subject=?""",(student_id,sec['subject'])).fetchone()
        answered=int(perf['answered'] or 0); accuracy=float(perf['accuracy'] or 0) if answered else 0.0
        bank=blueprint_question_pool(c,bp,sec['subject'])
        total_chapters=len({q['chapter'] for q in bank if (q['chapter'] or '').strip()})
        covered=int(perf['chapters'] or 0); coverage=100.0*covered/total_chapters if total_chapters else 0.0
        weakness=max(0.0,(75.0-accuracy)/75.0) if answered else 0.65
        weight=float(sec['weight_percent'] or 0)/100.0
        coverage_gap=max(0.0,1.0-coverage/100.0)
        priority=100*(0.45*weight+0.35*weakness+0.15*coverage_gap+0.05*urgency)
        reason=[]
        if weight>=0.25: reason.append(f"it contributes {float(sec['weight_percent']):g}% of the assessment")
        if weak_by_subject.get(sec['subject']): reason.append(f"you have {len(weak_by_subject[sec['subject']])} unrecovered weakness(es)")
        if answered and accuracy<60: reason.append(f"recent verified accuracy is critically low at {accuracy:.1f}%")
        if coverage<50: reason.append(f"verified syllabus coverage is only {coverage:.0f}%")
        if not reason: reason.append('maintaining balanced exam readiness')
        rows.append({'subject':sec['subject'],'weight_percent':float(sec['weight_percent'] or 0),'answered':answered,
                     'accuracy':round(accuracy,1),'coverage_percent':round(coverage,1),'priority_score':round(priority,1),
                     'reason':'Priority because '+', and '.join(reason)+'.'})
    rows.sort(key=lambda x:x['priority_score'],reverse=True)
    return {'blueprint_id':bp['id'],'blueprint_source_id':bp['powerhouse_blueprint_id'],'blueprint_version':bp['blueprint_version'],
            'framework':bp['framework_name'],'framework_version':bp['framework_version_name'],'subjects':rows}


def blueprint_projection_snapshot(c, student_id, blueprint_id):
    bp=blueprint_joined(c,blueprint_id)
    if not bp: return None
    section_rows=[]; overall_low=overall_mid=overall_high=0.0; evidence_scores=[]
    for sec in blueprint_sections(c,blueprint_id):
        perf=c.execute("""SELECT COUNT(aa.id) answered,100.0*SUM(aa.is_correct)/NULLIF(COUNT(aa.id),0) accuracy,
          COUNT(DISTINCT q.chapter) covered_chapters,
          AVG(CASE WHEN q.calibration_status IN ('CALIBRATED','OPERATIONAL') THEN 1.0 ELSE 0.0 END) calibrated_ratio
          FROM attempts a JOIN attempt_answers aa ON aa.attempt_id=a.id JOIN questions q ON q.id=aa.question_db_id
          WHERE a.student_id=? AND q.subject=?""",(student_id,sec['subject'])).fetchone()
        answered=int(perf['answered'] or 0); accuracy=float(perf['accuracy'] or 0) if answered else 0.0
        bank=blueprint_question_pool(c,bp,sec['subject'])
        total_chapters=len({q['chapter'] for q in bank if (q['chapter'] or '').strip()})
        coverage=100.0*int(perf['covered_chapters'] or 0)/total_chapters if total_chapters else 0.0
        label,evidence=confidence_label(answered,coverage,float(perf['calibrated_ratio'] or 0))
        qcount=int(sec['question_count'] or 0); midpoint=qcount*accuracy/100.0
        margin=qcount*({'Low':0.18,'Moderate':0.10,'High':0.055}[label])
        low=max(0.0,midpoint-margin); high=min(float(qcount),midpoint+margin)
        mastery=c.execute("""SELECT * FROM mastery_records WHERE student_id=? AND scope_type='subject'
          AND lower(subject)=lower(?) ORDER BY id DESC LIMIT 1""",(student_id,sec['subject'])).fetchone()
        mastery_status=effective_mastery_status(mastery) if mastery else 'Not yet verified'
        section_rows.append({'subject':sec['subject'],'question_count':qcount,'weight_percent':float(sec['weight_percent'] or 0),
                             'accuracy':round(accuracy,1),'answered':answered,'coverage_percent':round(coverage,1),
                             'projected_correct_low':round(low,1),'projected_correct_mid':round(midpoint,1),
                             'projected_correct_high':round(high,1),'confidence':label,'evidence_sufficiency':round(evidence,3),
                             'mastery_level':mastery['mastery_level'] if mastery else '', 'mastery_status':mastery_status})
        overall_low+=low; overall_mid+=midpoint; overall_high+=high; evidence_scores.append(evidence)
    overall_conf='Low'
    avg_evidence=sum(evidence_scores)/len(evidence_scores) if evidence_scores else 0
    if avg_evidence>=.72: overall_conf='High'
    elif avg_evidence>=.42: overall_conf='Moderate'
    previous=c.execute("""SELECT projection_json,created_at FROM student_blueprint_projections
      WHERE student_id=? AND blueprint_id=? ORDER BY id DESC LIMIT 1""",(student_id,blueprint_id)).fetchone()
    prior=safe_json(previous['projection_json'],{}) if previous else {}
    previous_mid=prior.get('projected_total_mid')
    change=None
    try:
        if previous_mid is not None: change=round(overall_mid-float(previous_mid),1)
    except (TypeError,ValueError):
        change=None
    return {'blueprint_id':bp['id'],'blueprint_source_id':bp['powerhouse_blueprint_id'],'blueprint_version':bp['blueprint_version'],
            'framework':bp['framework_name'],'framework_version':bp['framework_version_name'],'total_questions':bp['total_questions'],
            'subjects':section_rows,'projected_total_low':round(overall_low,1),'projected_total_mid':round(overall_mid,1),
            'projected_total_high':round(overall_high,1),'confidence':overall_conf,'evidence_sufficiency':round(avg_evidence,3),
            'change_since_previous_projection':change,'previous_projection_at':previous['created_at'] if previous else '',
            'generated_at':datetime.now().isoformat(timespec='seconds')}


def persist_blueprint_projection(c, student_id, projection):
    if not projection: return None
    cur=c.execute("""INSERT INTO student_blueprint_projections(student_id,blueprint_id,blueprint_version,framework_version,
      projection_json,confidence_label,evidence_sufficiency) VALUES(?,?,?,?,?,?,?)""",
      (student_id,projection['blueprint_id'],projection['blueprint_version'],projection['framework_version'],json.dumps(projection),
       projection['confidence'],projection['evidence_sufficiency']))
    return cur.lastrowid


def student_blueprint_snapshot(c, student_id):
    # Prefer a framework matching the learner goal/academic level, then fall back to the latest active blueprint.
    u=c.execute("SELECT academic_level,goal_name FROM users WHERE id=?",(student_id,)).fetchone()
    candidates=[]
    if u:
        candidates=[x for x in (u['academic_level'],u['goal_name']) if (x or '').strip()]
    bp=None
    for name in candidates:
        bp=active_assessment_blueprint(c,name)
        if bp: break
    # Do not show an unrelated active framework to a learner whose goal/level is known.
    # Only fall back when the learner has not yet selected an academic framework.
    if not bp and not candidates:
        bp=active_assessment_blueprint(c)
    if not bp: return None
    projection=blueprint_projection_snapshot(c,student_id,bp['id'])
    priority=blueprint_priority_snapshot(c,student_id,bp['id'])
    return {'blueprint':bp,'sections':blueprint_sections(c,bp['id']),'projection':projection,'priority':priority}

def student_weak_areas(c, student_id, limit=12):
    # V5.3: one wrong answer is not a confirmed weak area. Persistent state requires evidence.
    states=c.execute("""SELECT subject,'' programme,area_name area,area_key concept_key,evidence_count answered,
      accuracy,'' capsule_id,status FROM student_learning_states
      WHERE student_id=? AND status IN ('Weak Area','Recovery') AND evidence_count>=3
      ORDER BY accuracy ASC,evidence_count DESC LIMIT ?""",(student_id,limit)).fetchall()
    if states:
        return states
    # Backward-compatible fallback for databases created before V5.3 state exists.
    return c.execute("""SELECT q.subject,q.programme,
      COALESCE(NULLIF(q.learning_outcome,''),NULLIF(q.concept,''),NULLIF(q.subtopic,''),NULLIF(q.topic,''),q.chapter) area,
      COALESCE(NULLIF(q.concept_id,''),NULLIF(q.concept,''),NULLIF(q.learning_outcome,''),q.subtopic) concept_key,
      COUNT(*) answered,ROUND(100.0*SUM(aa.is_correct)/COUNT(*),1) accuracy,
      MAX(COALESCE(NULLIF(q.capsule_id,''),lc.capsule_id)) capsule_id,'Weak Area' status
      FROM attempt_answers aa JOIN attempts a ON a.id=aa.attempt_id JOIN questions q ON q.id=aa.question_db_id
      LEFT JOIN learning_capsules lc ON lc.active=1 AND lc.subject=q.subject
        AND (lc.concept_id<>'' AND lc.concept_id=q.concept_id OR lc.concept=q.concept)
      WHERE a.student_id=?
      GROUP BY q.subject,q.programme,area,concept_key
      HAVING COUNT(*)>=3 AND accuracy<75
      ORDER BY accuracy ASC,answered DESC LIMIT ?""",(student_id,limit)).fetchall()



STUDY_PLAN_PATHWAYS = {
  'Core': {
    'label':'Core','purpose':'A steady route to complete the syllabus, practise consistently and build secure exam readiness.',
    'target':'Secure performance','mastery':72,'days_per_week':5,'minutes_per_day':75,'mock_frequency_days':10,
    'coverage_pace':'Steady coverage','testing_rhythm':'Regular checkpoints','recovery_depth':'Essential weak-area recovery'
  },
  'Stretch': {
    'label':'Stretch','purpose':'A more demanding route with faster coverage, deeper practice and stronger evidence checks.',
    'target':'High performance','mastery':82,'days_per_week':6,'minutes_per_day':105,'mock_frequency_days':7,
    'coverage_pace':'Faster coverage','testing_rhythm':'Frequent checkpoints','recovery_depth':'Deeper recovery and reconfirmation'
  },
  'Peak': {
    'label':'Peak','purpose':'The highest-intensity preparation route for ambitious students, without confusing preparation with earned mastery.',
    'target':'Exceptional readiness','mastery':90,'days_per_week':6,'minutes_per_day':135,'mock_frequency_days':4,
    'coverage_pace':'Accelerated and comprehensive','testing_rhythm':'Frequent exam-standard testing','recovery_depth':'Maximum breadth, novelty and reconfirmation'
  },
  'Custom': {
    'label':'Custom','purpose':'A guided plan built around your target, current coverage, verified needs and exam date.',
    'target':'Personalised','mastery':80,'days_per_week':5,'minutes_per_day':72,'mock_frequency_days':7,
    'coverage_pace':'Availability-led','testing_rhythm':'Evidence-led','recovery_depth':'Personalised to verified need'
  }
}

def _parse_date(v):
    if not v:
        return None
    try:
        return datetime.strptime(v[:10],'%Y-%m-%d').date()
    except Exception:
        return None

def _student_accuracy(c, student_id):
    r=c.execute("""SELECT COUNT(*) answered,ROUND(100.0*SUM(aa.is_correct)/COUNT(*),1) accuracy
      FROM attempt_answers aa JOIN attempts a ON a.id=aa.attempt_id WHERE a.student_id=?""",(student_id,)).fetchone()
    return (int(r['answered'] or 0), float(r['accuracy'] or 0)) if r else (0,0.0)

def _programme_aliases(value):
    """Return conservative exact aliases for a learner programme/qualification.

    This intentionally avoids broad LIKE matching.  Current-study content must
    never fall through to another programme merely because the learner's own
    bank is empty.
    """
    text=(value or '').strip()
    if not text:
        return []
    aliases=[text]
    folded=text.casefold()
    if 'matric' in folded or 'ssc' in folded or 'class 9' in folded or 'class 10' in folded or 'grade 9' in folded or 'grade 10' in folded:
        aliases.extend(['Matric','SSC','Matric / Class 9–10','Class 9','Class 10'])
    elif folded in {'fsc part 1','f.sc part 1','fsc-i','hssc-i','hssc part 1'}:
        aliases.extend(['FSc Part 1','HSSC-I','HSSC Part 1'])
    elif folded in {'fsc part 2','f.sc part 2','fsc-ii','hssc-ii','hssc part 2'}:
        aliases.extend(['FSc Part 2','HSSC-II','HSSC Part 2'])
    # Stable case-insensitive de-duplication.
    out=[]; seen=set()
    for item in aliases:
        key=item.casefold()
        if key not in seen:
            seen.add(key); out.append(item)
    return out


def _programme_scope_sql(alias_values, table_alias='q'):
    aliases=[x for x in (alias_values or []) if (x or '').strip()]
    if not aliases:
        return '0=1', []
    placeholders=','.join('?' for _ in aliases)
    lowered=[x.casefold() for x in aliases]
    clause=f"(lower(COALESCE({table_alias}.programme,'')) IN ({placeholders}) OR lower(COALESCE({table_alias}.qualification,'')) IN ({placeholders}))"
    return clause, lowered+lowered


def _curriculum_chapters(c, student_id, target_exam=''):
    level=student_programme(c,student_id)
    requested=(target_exam or '').strip()
    scope_value=requested or level
    aliases=_programme_aliases(scope_value)
    scope_clause,scope_params=_programme_scope_sql(aliases,'q')
    rows=c.execute(f"""SELECT q.subject,q.chapter,MIN(q.id) first_id FROM questions q
      WHERE {live_question_clause('q')} AND COALESCE(q.chapter,'')<>'' AND {scope_clause}
      GROUP BY q.subject,q.chapter ORDER BY q.subject,first_id""",scope_params).fetchall()
    return [(r['subject'],r['chapter']) for r in rows]

def _tested_chapters(c, student_id):
    rows=c.execute("""SELECT DISTINCT q.subject,q.chapter FROM attempt_answers aa
      JOIN attempts a ON a.id=aa.attempt_id JOIN questions q ON q.id=aa.question_db_id
      WHERE a.student_id=? AND COALESCE(q.chapter,'')<>''""",(student_id,)).fetchall()
    return {(r['subject'],r['chapter']) for r in rows}

def _estimate_starting_coverage(c, student_id, target_exam=''):
    all_ch=_curriculum_chapters(c,student_id,target_exam)
    if not all_ch:
        return None
    tested=_tested_chapters(c,student_id)
    return round(100*len([x for x in all_ch if x in tested])/len(all_ch),1)

def recommend_plan_pathway(c, student_id):
    answered,accuracy=_student_accuracy(c,student_id)
    if answered<8:
        return 'Core'
    if accuracy>=86 and answered>=25:
        return 'Peak'
    if accuracy>=70 and answered>=12:
        return 'Stretch'
    return 'Core'

def _phase_for_date(plan, activity_date):
    d=_parse_date(activity_date)
    exam=_parse_date(plan['target_date'])
    ready=_parse_date(plan['readiness_date'])
    if not d:
        return 'Preparation'
    if exam and d>=exam:
        return 'Exam Complete'
    if ready and d>=ready:
        return '30-Day Exam Mode'
    return 'Preparation & Mastery'

def _sync_activity_phase_week(c, plan, activity_id, activity_date):
    d=_parse_date(activity_date)
    created=_parse_date(plan['created_at']) or datetime.now().date()
    week=max(1,((d-created).days//7)+1) if d else None
    c.execute("UPDATE study_plan_activities SET week_no=?,phase=? WHERE id=?",(week,_phase_for_date(plan,activity_date),activity_id))

def apply_attempt_to_plan_activity(c, attempt_id, student_id, meta):
    activity_id=meta.get('plan_activity_id')
    if not activity_id:
        return
    activity=c.execute("SELECT * FROM study_plan_activities WHERE id=? AND student_id=?",(activity_id,student_id)).fetchone()
    attempt=c.execute("SELECT * FROM attempts WHERE id=? AND student_id=?",(attempt_id,student_id)).fetchone()
    if not activity or not attempt:
        return
    threshold=float(activity['required_mastery'] or activity['target_score'] or 0)
    score=float(attempt['score'] or 0)
    outcome='passed' if not threshold or score>=threshold else 'needs_recovery'
    c.execute("""UPDATE study_plan_activities SET status='completed',evidence_status='verified',evidence_type='scoremax_attempt',
      linked_attempt_id=?,verified_score=?,outcome_status=?,completed_at=CURRENT_TIMESTAMP,last_evidence_at=CURRENT_TIMESTAMP
      WHERE id=?""",(attempt_id,score,outcome,activity_id))
    if outcome=='needs_recovery':
        # Create a protected recovery item once; do not silently pretend the target was achieved.
        exists=c.execute("""SELECT 1 FROM study_plan_activities WHERE plan_id=? AND status<>'completed' AND activity_type='recovery'
          AND subject=? AND COALESCE(chapter,'')=COALESCE(?, '') AND source_reason LIKE ? LIMIT 1""",
          (activity['plan_id'],activity['subject'] or '',activity['chapter'] or '',f'%attempt {attempt_id}%')).fetchone()
        if not exists:
            plan=c.execute("SELECT * FROM study_plans WHERE id=?",(activity['plan_id'],)).fetchone()
            exam=_parse_date(plan['target_date']) if plan else None
            next_date=datetime.now().date()+timedelta(days=2)
            if exam and next_date>=exam:
                next_date=max(datetime.now().date(),exam-timedelta(days=1))
            c.execute("""INSERT INTO study_plan_activities(plan_id,student_id,activity_date,subject,chapter,topic,activity_type,title,
              target_score,status,source_reason,week_no,phase,priority,estimated_minutes,required_mastery,mandatory,evidence_status,concept_key)
              VALUES(?,?,?,?,?,?, 'recovery',?,?, 'planned',?,?,?,?,?,?,?,?,?)""",
              (activity['plan_id'],student_id,next_date.isoformat(),activity['subject'] or '',activity['chapter'] or '',activity['topic'] or '',
               f"Recover and retest: {activity['title']}",threshold,
               f"Verified attempt {attempt_id} scored {score:.1f}% against a {threshold:.0f}% target.",
               1,'Weak Area Recovery',1,0,threshold,1,'scheduled',activity['concept_key'] or ''))

def add_due_recall_to_plan(c, student_id, plan):
    due=due_recall_items(c,student_id,limit=4)
    if not due:
        return 0
    exam=_parse_date(plan['target_date'])
    added=0
    for r in due:
        existing=c.execute("""SELECT 1 FROM study_plan_activities WHERE plan_id=? AND status<>'completed' AND activity_type='recall'
          AND concept_key=? LIMIT 1""",(plan['id'],r['concept_key'])).fetchone()
        if existing:
            continue
        d=max(datetime.now().date(),_parse_date(r['next_due_date']) or datetime.now().date())
        if exam and d>=exam:
            d=exam-timedelta(days=1)
        c.execute("""INSERT INTO study_plan_activities(plan_id,student_id,activity_date,subject,chapter,topic,activity_type,title,
          target_score,status,source_reason,week_no,phase,priority,estimated_minutes,required_mastery,mandatory,evidence_status,concept_key)
          VALUES(?,?,?,?,?,?,'recall',?,?,'planned',?,?,?,?,?,?,?,?,?)""",
          (plan['id'],student_id,d.isoformat(),r['subject'] or '',r['chapter'] or '',r['topic'] or '',
           f"Quick Recall: {r['area_name'] or r['concept_key']}",plan['readiness_target'] or 80,
           f"Recall is due after {r['interval_days']} days. ScoreMax is checking retention, not repeating busywork.",
           1,_phase_for_date(plan,d.isoformat()),2,0,plan['readiness_target'] or 80,1,'scheduled',r['concept_key']))
        added+=1
    # V6.3: feature-flagged universal maintenance queue feeds the same learner plan.
    if universal_mastery.feature_enabled(c,'universal_mastery_runtime',learner_key=f'USER:{student_id}'):
        for r in universal_mastery.universal_maintenance_actions(c,student_id,limit=4):
            concept_key=f"UM:{r['entity_type']}:{r['entity_id']}"
            existing=c.execute("""SELECT 1 FROM study_plan_activities WHERE plan_id=? AND status<>'completed'
              AND concept_key=? LIMIT 1""",(plan['id'],concept_key)).fetchone()
            if existing:
                continue
            d=_parse_date(r.get('due_at')) or datetime.now().date()
            if exam and d>=exam:
                d=max(datetime.now().date(),exam-timedelta(days=1))
            c.execute("""INSERT INTO study_plan_activities(plan_id,student_id,activity_date,subject,chapter,topic,activity_type,title,
              target_score,status,source_reason,week_no,phase,priority,estimated_minutes,required_mastery,mandatory,evidence_status,concept_key)
              VALUES(?,?,?,?,?,?,'reconfirmation',?,?,'planned',?,?,?,?,?,?,?,?,?)""",
              (plan['id'],student_id,d.isoformat(),r.get('subject',''),r.get('chapter',''),'',
               f"Keep it strong: {r.get('label') or r['entity_id']}",plan['readiness_target'] or 80,
               'Universal mastery maintenance is due; this checks freshness without declaring the learner forgotten.',
               1,_phase_for_date(plan,d.isoformat()),2,10,plan['readiness_target'] or 80,1,'scheduled',concept_key))
            added+=1
    return added

def add_priority_recovery_to_plan(c, student_id, plan):
    weak=student_weak_areas(c,student_id,limit=4)
    exam=_parse_date(plan['target_date'])
    added=0
    for w in weak:
        exists=c.execute("""SELECT 1 FROM study_plan_activities WHERE plan_id=? AND status<>'completed'
          AND activity_type IN ('weak_area','recovery') AND concept_key=? LIMIT 1""",(plan['id'],w['concept_key'] or '')).fetchone()
        if exists:
            continue
        d=datetime.now().date()+timedelta(days=1+added)
        if exam and d>=exam: d=exam-timedelta(days=1)
        c.execute("""INSERT INTO study_plan_activities(plan_id,student_id,activity_date,subject,chapter,topic,activity_type,title,
          target_score,status,source_reason,week_no,phase,priority,estimated_minutes,required_mastery,mandatory,evidence_status,concept_key)
          VALUES(?,?,?,?,?,?,'weak_area',?,?,'planned',?,?,?,?,?,?,?,?,?)""",
          (plan['id'],student_id,d.isoformat(),w['subject'] or '','','',f"Repair: {w['area']}",plan['readiness_target'] or 80,
           f"Evidence shows {w['accuracy']:.1f}% across {w['answered']} answers. Recovery stays protected until retested.",
           1,'Weak Area Recovery',1,0,plan['readiness_target'] or 80,1,'scheduled',w['concept_key'] or ''))
        added+=1
    # V6.3: add causal universal recovery actions only when the pilot flag is enabled.
    if universal_mastery.feature_enabled(c,'universal_mastery_runtime',learner_key=f'USER:{student_id}'):
        for r in universal_mastery.universal_recovery_actions(c,student_id,limit=4):
            concept_key=f"UM:{r['entity_type']}:{r['entity_id']}"
            exists=c.execute("""SELECT 1 FROM study_plan_activities WHERE plan_id=? AND status<>'completed'
              AND activity_type IN ('weak_area','recovery') AND concept_key=? LIMIT 1""",(plan['id'],concept_key)).fetchone()
            if exists:
                continue
            d=datetime.now().date()+timedelta(days=1+added)
            if exam and d>=exam:
                d=max(datetime.now().date(),exam-timedelta(days=1))
            reason=(r.get('cause_code') or 'EVIDENCE_GAP').replace('_',' ').title()
            c.execute("""INSERT INTO study_plan_activities(plan_id,student_id,activity_date,subject,chapter,topic,activity_type,title,
              target_score,status,source_reason,week_no,phase,priority,estimated_minutes,required_mastery,mandatory,evidence_status,concept_key)
              VALUES(?,?,?,?,?,?,'recovery',?,?,'planned',?,?,?,?,?,?,?,?,?)""",
              (plan['id'],student_id,d.isoformat(),r.get('subject',''),r.get('chapter',''),'',
               f"Repair: {r.get('label') or r['entity_id']}",plan['readiness_target'] or 80,
               f"Universal mastery diagnosis: {reason}. Recovery stays open until qualifying independent evidence is obtained.",
               1,'Weak Area Recovery',1,15,plan['readiness_target'] or 80,1,'scheduled',concept_key))
            added+=1
    return added

def get_active_study_plan(c, student_id):
    plan=c.execute("""SELECT * FROM study_plans WHERE student_id=? AND status='active'
      ORDER BY id DESC LIMIT 1""",(student_id,)).fetchone()
    if not plan:
        return None
    activities=c.execute("""SELECT * FROM study_plan_activities WHERE plan_id=?
      ORDER BY CASE WHEN activity_date='' THEN 1 ELSE 0 END,activity_date,priority,id""",(plan['id'],)).fetchall()
    completed=sum(1 for x in activities if x['status']=='completed')
    verified=sum(1 for x in activities if x['evidence_status']=='verified')
    self_reported=sum(1 for x in activities if x['evidence_status']=='self_reported')
    total=len(activities)
    pct=round(100*completed/total) if total else 0
    today=iso_today()
    today_items=[x for x in activities if x['activity_date']==today and x['status']!='completed']
    next_item=today_items[0] if today_items else next((x for x in activities if x['status']!='completed'),None)

    by_week={}
    by_phase={}
    for x in activities:
        wk=x['week_no'] or 0
        by_week.setdefault(wk,{'total':0,'done':0})
        by_week[wk]['total']+=1
        if x['status']=='completed':
            by_week[wk]['done']+=1
        ph=x['phase'] or 'Preparation'
        by_phase.setdefault(ph,{'total':0,'done':0})
        by_phase[ph]['total']+=1
        if x['status']=='completed':
            by_phase[ph]['done']+=1

    weekly=[{'week':k,'total':v['total'],'done':v['done'],'pct':round(100*v['done']/v['total']) if v['total'] else 0}
            for k,v in sorted(by_week.items()) if k]
    phases=[{'phase':k,'total':v['total'],'done':v['done'],'pct':round(100*v['done']/v['total']) if v['total'] else 0}
            for k,v in by_phase.items()]

    exam_date=_parse_date(plan['target_date'])
    readiness=_parse_date(plan['readiness_date'])
    days_remaining=(exam_date-datetime.now().date()).days if exam_date else None
    readiness_days=(readiness-datetime.now().date()).days if readiness else None
    priority_summary={
      'pending':sum(1 for x in activities if x['status']!='completed'),
      'high_priority':sum(1 for x in activities if x['status']!='completed' and int(x['priority'] or 0)>=80),
      'verified':verified,
      'recall_or_recovery':sum(1 for x in activities if x['status']!='completed' and (x['activity_type'] in ('recall','recovery','reconfirmation'))),
    }
    return {'plan':plan,'activities':activities,'completed':completed,'verified':verified,'self_reported':self_reported,'total':total,'completion_pct':pct,
            'next_item':next_item,'weekly':weekly,'phases':phases,'days_remaining':days_remaining,
            'readiness_days':readiness_days,'priority_summary':priority_summary}

def generate_scoremax_plan(c, student_id, pathway='Core', target_exam='', target_date='',
                           target_percentage=None, days_per_week=None, minutes_per_day=None,
                           starting_coverage=None, custom_settings=None):
    pathway=pathway if pathway in STUDY_PLAN_PATHWAYS else 'Core'
    cfg=STUDY_PLAN_PATHWAYS[pathway]
    today=datetime.now().date()
    exam_date=_parse_date(target_date)

    if not exam_date or exam_date<=today:
        raise ValueError('Choose a valid future exam date so ScoreMax can build an honest plan.')

    exam_mode_days=30
    readiness_date=exam_date-timedelta(days=exam_mode_days)
    days_to_readiness=max(1,(readiness_date-today).days)

    answered,accuracy=_student_accuracy(c,student_id)
    measured_coverage=_estimate_starting_coverage(c,student_id,target_exam)
    coverage=float(starting_coverage) if starting_coverage is not None else measured_coverage
    coverage=None if coverage is None else max(0,min(100,coverage))
    dpw=max(1,min(7,int(days_per_week or cfg['days_per_week'])))
    mpd=max(20,min(480,int(minutes_per_day or cfg['minutes_per_day'])))
    base_mastery=float(cfg['mastery'])
    target_num=float(target_percentage) if target_percentage is not None else base_mastery
    mastery=max(base_mastery,min(95.0,target_num-5 if target_num>=85 else target_num))

    c.execute("UPDATE study_plans SET status='inactive',updated_at=CURRENT_TIMESTAMP WHERE student_id=? AND status='active'",(student_id,))
    cur=c.execute("""INSERT INTO study_plans(student_id,pathway,source,title,target_exam,target_date,target_percentage,
      readiness_date,exam_mode_days,days_per_week,minutes_per_day,starting_coverage,starting_accuracy,readiness_target,
      pace_status,custom_settings_json)
      VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
      (student_id,pathway,'scoremax',f'ScoreMax {pathway} Plan',target_exam,target_date,target_percentage,
       readiness_date.isoformat(),exam_mode_days,dpw,mpd,coverage,accuracy if answered else None,mastery,'',
       json.dumps(custom_settings or {})))
    pid=cur.lastrowid
    access_at_build=get_access_profile(c,student_id)
    c.execute("UPDATE study_plans SET access_rank=?,mastery_ceiling=? WHERE id=?",(access_at_build.get('access_rank',0),access_at_build.get('mastery_ceiling','Foundation'),pid))
    c.execute("UPDATE users SET study_plan_pathway=?,study_plan_source='scoremax',study_plan_active=1 WHERE id=?",(pathway,student_id))

    all_chapters=_curriculum_chapters(c,student_id,target_exam)
    tested=_tested_chapters(c,student_id)
    remaining=[x for x in all_chapters if x not in tested]

    # V5.5: one authoritative blueprint informs priority without replacing verified learner need.
    active_bp=active_assessment_blueprint(c,target_exam) if target_exam else None
    priority_snapshot=None
    if active_bp:
        priority_snapshot=blueprint_priority_snapshot(c,student_id,active_bp['id'],target_date)
        priority_rank={x['subject']:x['priority_score'] for x in priority_snapshot.get('subjects',[])}
        remaining.sort(key=lambda pair:(-float(priority_rank.get(pair[0],0)),pair[0],pair[1]))
        c.execute("""UPDATE study_plans SET assessment_blueprint_id=?,blueprint_version=?,framework_version=?,
          subject_priority_snapshot_json=? WHERE id=?""",
          (active_bp['id'],active_bp['blueprint_version'],active_bp['framework_version_name'],json.dumps(priority_snapshot),pid))

    if all_chapters and coverage is not None and coverage>0 and len(tested)<round(len(all_chapters)*coverage/100):
        assumed_done=round(len(all_chapters)*coverage/100)
        remaining=all_chapters[assumed_done:]

    prep_weeks=max(1,(days_to_readiness+6)//7)
    available_slots=max(1,prep_weeks*dpw)
    base_sessions=len(remaining)
    recovery_sessions=max(2,round(len(all_chapters)*0.12)) if all_chapters else 2
    test_sessions=max(2,prep_weeks//2)
    needed_sessions=base_sessions+recovery_sessions+test_sessions
    load_ratio=needed_sessions/available_slots if available_slots else 99

    if load_ratio<=0.75:
        pace='Comfortable'
    elif load_ratio<=1.0:
        pace='Focused'
    elif load_ratio<=1.3:
        pace='Intensive'
    else:
        pace='Compressed'
    c.execute("UPDATE study_plans SET pace_status=? WHERE id=?",(pace,pid))

    work_dates=[]
    preferred=(custom_settings or {}).get('preferred_days') or []
    day_map={'Mon':0,'Tue':1,'Wed':2,'Thu':3,'Fri':4,'Sat':5,'Sun':6,'Monday':0,'Tuesday':1,'Wednesday':2,'Thursday':3,'Friday':4,'Saturday':5,'Sunday':6}
    preferred_idx=[day_map[x] for x in preferred if x in day_map]
    if not preferred_idx:
        preferred_idx=list(range(dpw))
    d=today
    while d<readiness_date:
        if d.weekday() in preferred_idx:
            work_dates.append(d)
        d+=timedelta(days=1)
    if not work_dates:
        work_dates=[today]

    items=[]
    def add_item(date,subject,chapter,kind,title,phase,priority=3,minutes=None,reason=''):
        week=((date-today).days//7)+1
        items.append((date.isoformat(),subject,chapter,'',kind,title,mastery,'planned',reason,week,phase,
                      priority,0,mastery,1,'scheduled',''))

    # Syllabus coverage before the protected readiness date.
    for idx,(subject,chapter) in enumerate(remaining):
        date=work_dates[min(idx,len(work_dates)-1)]
        phase='Syllabus Coverage' if idx < max(1,int(len(remaining)*0.75)) else 'Coverage + Mastery'
        bp_reason=''
        if priority_snapshot:
            row=next((x for x in priority_snapshot.get('subjects',[]) if x['subject']==subject),None)
            bp_reason=(' '+row['reason']) if row else ''
        add_item(date,subject,chapter,'chapter_test',f"{subject}: {chapter}",phase,2,
                 min(mpd,75),f"Scheduled to complete syllabus before {readiness_date.isoformat()}.{bp_reason}")

    # Weak-area recovery.
    weak=student_weak_areas(c,student_id,limit=8)
    occupied={x[0] for x in items}
    spare=[d for d in work_dates if d.isoformat() not in occupied]
    for i,w in enumerate(weak[:min(6,len(spare))]):
        date=spare[i]
        add_item(date,w['subject'],'','weak_area',f"Recover weak area: {w['area']}",'Weak Area Recovery',1,
                 min(mpd,60),f"Current evidence: {w['accuracy']}% accuracy.")

    # Regular progress checks.
    freq={'Core':14,'Stretch':10,'Peak':7,'Custom':10}[pathway]
    cursor=today+timedelta(days=freq)
    test_no=1
    while cursor<readiness_date:
        while cursor<readiness_date and cursor.weekday() not in preferred_idx:
            cursor+=timedelta(days=1)
        if cursor>=readiness_date: break
        add_item(cursor,'','','mixed_test',f"Progress Check {test_no}",'Testing & Revision',2,
                 min(mpd,90),"Checks whether syllabus coverage is becoming reliable exam performance.")
        test_no+=1
        cursor+=timedelta(days=freq)

    # Protected final 30 days: Exam Mode.
    d=readiness_date
    mock_every=cfg['mock_frequency_days']
    mock_no=1
    while d<exam_date:
        if (d-readiness_date).days % mock_every==0:
            bp_label=f" under {active_bp['powerhouse_blueprint_id']} v{active_bp['blueprint_version']}" if active_bp else ''
            add_item(d,'','','mock',f"Full Mock {mock_no}",'30-Day Exam Mode',1,
                     min(max(mpd,120),240),f"Exam Mode: authentic blueprint-driven practice{bp_label}; analyse, repair, retest.")
            mock_no+=1
        elif d.weekday() in preferred_idx:
            add_item(d,'','','recovery',"Mock Review & Targeted Repair",'30-Day Exam Mode',2,
                     min(mpd,75),"Use recent mock evidence to recover marks before the next timed attempt.")
        d+=timedelta(days=1)

    # Late joiners: honest priority mode instead of pretending the normal pathway is possible.
    if readiness_date<=today:
        items=[]
        remaining_days=max(1,(exam_date-today).days)
        for i in range(min(30,remaining_days)):
            day=today+timedelta(days=i)
            if i%3==0:
                add_item(day,'','','mock',f"Priority Full Mock {i//3+1}",'Late-Join Exam Mode',1,
                         min(max(mpd,120),240),"Late-join plan prioritises marks recovery and exam performance.")
            elif day.weekday() in preferred_idx:
                add_item(day,'','','recovery',"Priority Weak-Area Repair",'Late-Join Exam Mode',1,
                         min(mpd,75),"Focus on the highest-value weaknesses with the time remaining.")
        c.execute("UPDATE study_plans SET pace_status='Late-Join Priority' WHERE id=?",(pid,))

    for row in items:
        c.execute("""INSERT INTO study_plan_activities(plan_id,student_id,activity_date,subject,chapter,topic,
          activity_type,title,target_score,status,source_reason,week_no,phase,priority,estimated_minutes,required_mastery,mandatory,evidence_status,concept_key)
          VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",(pid,student_id)+row)

    universal_mastery.emit_growth_event(c,'STUDY_PLAN_CREATED',f'USER:{student_id}',{
      'plan_id':pid,'pathway':pathway,'target_exam':target_exam,'target_date':target_date})
    c.commit()
    return pid

def rebalance_study_plan(c, student_id):
    """V5.3 evidence-aware rebalance. Never schedules beyond the exam date and protects recovery/recall."""
    data=get_active_study_plan(c,student_id)
    if not data:
        return None
    plan=data['plan']
    today=datetime.now().date()
    exam=_parse_date(plan['target_date'])
    readiness=_parse_date(plan['readiness_date'])
    if not exam or exam<=today:
        return plan['id']

    add_priority_recovery_to_plan(c,student_id,plan)
    add_due_recall_to_plan(c,student_id,plan)
    data=get_active_study_plan(c,student_id)

    overdue=[x for x in data['activities'] if x['status']!='completed' and _parse_date(x['activity_date']) and _parse_date(x['activity_date'])<today]
    future_fixed={x['activity_date'] for x in data['activities'] if x['status']!='completed' and _parse_date(x['activity_date']) and _parse_date(x['activity_date'])>=today}
    preferred=safe_json(plan['custom_settings_json'],{}).get('preferred_days') or []
    day_map={'Mon':0,'Tue':1,'Wed':2,'Thu':3,'Fri':4,'Sat':5,'Sun':6,'Monday':0,'Tuesday':1,'Wednesday':2,'Thursday':3,'Friday':4,'Saturday':5,'Sunday':6}
    preferred_idx=[day_map[x] for x in preferred if x in day_map] or list(range(int(plan['days_per_week'] or 6)))
    d=today
    overflow=0
    for x in sorted(overdue,key=lambda r:(int(r['priority'] or 3),r['activity_date'],r['id'])):
        while d<exam and (d.isoformat() in future_fixed or d.weekday() not in preferred_idx):
            d+=timedelta(days=1)
        if d>=exam:
            # Keep the item visible but do not lie by scheduling after the exam.
            c.execute("UPDATE study_plan_activities SET status='deferred',outcome_status='insufficient_time',source_reason=source_reason||' Not rescheduled beyond the exam date.' WHERE id=?",(x['id'],))
            overflow+=1
            continue
        # Non-exam-mode syllabus work is not pushed into the protected final 30 days unless it is recovery/recall/testing.
        if readiness and d>=readiness and x['activity_type'] not in ('recovery','weak_area','recall','mixed_test','mock'):
            c.execute("UPDATE study_plan_activities SET status='deferred',outcome_status='exam_mode_protected',source_reason=source_reason||' Deferred to protect Exam Mode.' WHERE id=?",(x['id'],))
            overflow+=1
            continue
        c.execute("""UPDATE study_plan_activities SET activity_date=?,source_reason=source_reason||' Rebalanced from verified plan evidence.' WHERE id=?""",
                  (d.isoformat(),x['id']))
        _sync_activity_phase_week(c,plan,x['id'],d.isoformat())
        future_fixed.add(d.isoformat()); d+=timedelta(days=1)

    week=verified_plan_summary(c,student_id,(today-timedelta(days=today.weekday())).isoformat(),(today-timedelta(days=today.weekday())+timedelta(days=6)).isoformat())
    completion=(100.0*week['completed']/week['total']) if week['total'] else 0
    verified_rate=(100.0*week['verified']/week['total']) if week['total'] else 0
    if overflow:
        pace='Compressed'
    elif week['total'] and completion>=85 and verified_rate>=50:
        pace='On Track'
    elif week['total'] and completion<60:
        pace='Behind'
    else:
        pace='Focused'
    c.execute("UPDATE study_plans SET pace_status=?,last_rebalanced_at=?,updated_at=CURRENT_TIMESTAMP WHERE id=?",
              (pace,datetime.now().isoformat(timespec='seconds'),plan['id']))
    c.commit()
    return plan['id']

def maybe_weekly_rebalance(c, student_id):
    """Run evidence-based consolidation once per calendar week; high-signal attempts can still trigger immediate rebalance."""
    plan=c.execute("SELECT * FROM study_plans WHERE student_id=? AND status='active' ORDER BY id DESC LIMIT 1",(student_id,)).fetchone()
    if not plan:
        return None
    today=datetime.now().date(); monday=today-timedelta(days=today.weekday())
    last=_parse_date(plan['last_rebalanced_at'])
    if not last or last<monday:
        return rebalance_study_plan(c,student_id)
    return plan['id']

def weekly_progress_summary(c, student_id):
    today=datetime.now().date()
    monday=today-timedelta(days=today.weekday())
    sunday=monday+timedelta(days=6)
    plan_data=get_active_study_plan(c,student_id)
    planned=done=verified=self_reported=0
    if plan_data:
        rows=[x for x in plan_data['activities'] if monday.isoformat()<=x['activity_date']<=sunday.isoformat()]
        planned=len(rows)
        done=sum(1 for x in rows if x['status']=='completed')
        verified=sum(1 for x in rows if x['evidence_status']=='verified')
        self_reported=sum(1 for x in rows if x['evidence_status']=='self_reported')
    completion=round(100*done/planned,1) if planned else 0
    attempts=c.execute("""SELECT score FROM attempts WHERE student_id=? AND date(created_at) BETWEEN ? AND ?""",
                       (student_id,monday.isoformat(),sunday.isoformat())).fetchall()
    avg=round(sum(float(x['score'] or 0) for x in attempts)/len(attempts),1) if attempts else None
    if planned==0:
        status='Getting Started'
    elif completion>=85:
        status='On Track'
    elif completion>=60:
        status='Slightly Behind'
    else:
        status='Needs Attention'
    weak=student_weak_areas(c,student_id,limit=2)
    priorities=', '.join(f"{w['subject']}: {w['area']}" for w in weak) or 'Continue planned practice'
    summary=(f"You completed {done} of {planned} planned activities this week. "
             f"{'Keep the momentum going.' if status=='On Track' else 'Your plan can be rebalanced so you can recover without panic.'}")
    return {'week_start':monday.isoformat(),'week_end':sunday.isoformat(),'planned':planned,'done':done,'verified':verified,'self_reported':self_reported,
            'completion':completion,'tests':len(attempts),'avg':avg,'status':status,
            'priorities':priorities,'summary':summary}

def ensure_weekly_progress_report(c, student_id):
    w=weekly_progress_summary(c,student_id)
    plan=get_active_study_plan(c,student_id)
    u=c.execute("SELECT weekly_email_enabled,parent_weekly_email_enabled FROM users WHERE id=?",(student_id,)).fetchone()
    c.execute("""INSERT INTO weekly_progress_reports(student_id,week_start,week_end,plan_id,plan_completion,
      tests_completed,avg_score,status_label,summary_text,priorities_text,student_email_enabled,parent_email_enabled)
      VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
      ON CONFLICT(student_id,week_start) DO UPDATE SET
      plan_completion=excluded.plan_completion,tests_completed=excluded.tests_completed,avg_score=excluded.avg_score,
      status_label=excluded.status_label,summary_text=excluded.summary_text,priorities_text=excluded.priorities_text""",
      (student_id,w['week_start'],w['week_end'],plan['plan']['id'] if plan else None,w['completion'],w['tests'],w['avg'],
       w['status'],w['summary'],w['priorities'],int(u['weekly_email_enabled'] or 0) if u else 1,
       int(u['parent_weekly_email_enabled'] or 0) if u else 0))
    c.commit()
    return w


def motivation_message(c, student_id):
    d=student_dashboard_intelligence(c,student_id)
    weekly=weekly_progress_summary(c,student_id)
    if d.get('improvements'):
        x=d['improvements'][0]
        return {'headline':f"You've improved {x['area']}.",
                'message':f"{x['first']}% → {x['latest']}%. Your work is showing in the results.",
                'tone':'progress'}
    if weekly['status']=='On Track':
        return {'headline':"You're keeping your promise to yourself.",
                'message':"Your Study Plan is on track this week. Keep the momentum going.",
                'tone':'ontrack'}
    if weekly['status']=='Slightly Behind':
        return {'headline':"A small reset can change the week.",
                'message':"You're slightly behind your plan. Rebalance it and focus on the next priority — not the missed day.",
                'tone':'recover'}
    if weekly['status']=='Needs Attention':
        return {'headline':"You haven't lost the goal.",
                'message':"The plan needs attention, not panic. Start with one focused session and ScoreMax can rebuild the route from there.",
                'tone':'recover'}
    return {'headline':"Your journey starts with evidence.",
            'message':"Take your first tests so ScoreMax can understand where you stand and guide what comes next.",
            'tone':'start'}

def daily_focus(c, student_id):
    u=c.execute("SELECT * FROM users WHERE id=?",(student_id,)).fetchone()
    plan_data=get_active_study_plan(c,student_id)
    dates=c.execute("""SELECT * FROM student_exam_dates WHERE student_id=? AND active=1 AND exam_date>=?
      ORDER BY exam_date LIMIT 1""",(student_id,iso_today())).fetchone()
    days=exam_days_remaining(dates['exam_date']) if dates else None
    target=(u['goal_name'] or u['goal_type'] or '') if u else ''
    item=plan_data['next_item'] if plan_data else None
    misconceptions=confirmed_misconceptions(c,student_id,limit=1)
    if misconceptions:
        m=misconceptions[0]
        return {'headline':"Fix a misconception before it costs more marks",
                'message':f"ScoreMax has repeated evidence around {m['area_name'] or m['misconception_key']}.",
                'action':'Repair Misconception','kind':'weak','target':target,'days':days}
    recall=due_recall_items(c,student_id,limit=1)
    if recall and (not item or int(item['priority'] or 3)>1):
        r=recall[0]
        return {'headline':"Quick recall is due",'message':f"Refresh {r['area_name'] or r['concept_key']} before it fades.",
                'action':'Do Quick Recall','kind':'plan' if plan_data else 'recall','target':target,'days':days}
    if item:
        return {'headline':f"{days} days to {dates['exam_name'] or dates['subject']}" if days is not None else "Today's focus",
                'message':f"Your next priority is {item['title']}.",'action':'Continue My Plan','kind':'plan',
                'target':target,'days':days}
    weak=student_weak_areas(c,student_id,limit=1)
    if weak:
        return {'headline':f"{days} days to {dates['exam_name'] or dates['subject']}" if days is not None else "Keep moving",
                'message':f"Your biggest current opportunity is {weak[0]['area']} in {weak[0]['subject']}.",
                'action':'Improve Weak Areas','kind':'weak','target':target,'days':days}
    return {'headline':f"{days} days to {dates['exam_name'] or dates['subject']}" if days is not None else "Build your starting point",
            'message':"Take a short test so ScoreMax can guide what you should do next.",'action':'Test Me','kind':'test',
            'target':target,'days':days}

def student_dashboard_intelligence(c, student_id):
    """Build V5 dashboard cards from live assessment evidence without inventing mastery."""
    now=datetime.now()
    week_start=(now-timedelta(days=7)).isoformat(timespec='seconds')

    # Subject and chapter mastery summaries from response-level evidence.
    subject_rows=c.execute("""SELECT COALESCE(NULLIF(q.programme,''),NULLIF(a.programme,''),'') programme,q.subject,COUNT(*) answered,
      ROUND(100.0*SUM(aa.is_correct)/COUNT(*),1) accuracy,
      MAX(a.created_at) last_attempt
      FROM attempt_answers aa JOIN attempts a ON a.id=aa.attempt_id
      JOIN questions q ON q.id=aa.question_db_id
      WHERE a.student_id=? AND COALESCE(q.subject,'')<>''
      GROUP BY COALESCE(NULLIF(q.programme,''),NULLIF(a.programme,''),''),q.subject
      ORDER BY programme,q.subject""",(student_id,)).fetchall()
    chapter_rows=c.execute("""SELECT q.subject,q.chapter,COUNT(*) answered,
      ROUND(100.0*SUM(aa.is_correct)/COUNT(*),1) accuracy,
      MAX(a.created_at) last_attempt
      FROM attempt_answers aa JOIN attempts a ON a.id=aa.attempt_id
      JOIN questions q ON q.id=aa.question_db_id
      WHERE a.student_id=? AND COALESCE(q.chapter,'')<>''
      GROUP BY q.subject,q.chapter ORDER BY q.subject,q.chapter""",(student_id,)).fetchall()

    level_rows=c.execute("""SELECT q.subject,q.level,COUNT(*) answered,
      ROUND(100.0*SUM(aa.is_correct)/COUNT(*),1) accuracy
      FROM attempt_answers aa JOIN attempts a ON a.id=aa.attempt_id
      JOIN questions q ON q.id=aa.question_db_id
      WHERE a.student_id=? AND COALESCE(q.level,'')<>''
      GROUP BY q.subject,q.level""",(student_id,)).fetchall()
    levels_by_subject={}
    for r in level_rows:
        levels_by_subject.setdefault(r['subject'],{})[r['level']]={'answered':r['answered'],'accuracy':float(r['accuracy'] or 0)}

    order=SCOREMAX_LEVELS
    subjects=[]
    for r in subject_rows:
        evidence='Insufficient evidence' if r['answered']<4 else ('Emerging evidence' if r['answered']<8 else ('Moderate evidence' if r['answered']<15 else 'Strong evidence'))
        mastered=scoremax_level_from_evidence(r['accuracy'],r['answered'])
        next_level=next_level_name(mastered)
        level_progress=next_level_progress(r['accuracy'],mastered)
        community=subject_community_snapshot(c,r['programme'],r['subject'],student_id)
        subjects.append({'programme':r['programme'],'subject':r['subject'],'accuracy':float(r['accuracy'] or 0),'answered':r['answered'],'evidence':evidence,'mastery':mastered,'mastery_index':level_index(mastered),'next_level':next_level,'level_progress':level_progress,'community':community})

    chapters=[]
    for r in chapter_rows:
        chapters.append({'subject':r['subject'],'chapter':r['chapter'],'accuracy':float(r['accuracy'] or 0),'answered':r['answered'],
                         'status':'Strong' if float(r['accuracy'] or 0)>=80 else ('Developing' if float(r['accuracy'] or 0)>=60 else 'Needs attention')})

    # Weekly activity uses actual attempts and captured response-time seconds.
    weekly=c.execute("""SELECT COUNT(DISTINCT a.id) assessments,
      COALESCE(SUM(aa.response_time_seconds),0) seconds,
      COALESCE(SUM(CASE WHEN aa.is_correct=1 THEN 1 ELSE 0 END),0) correct,
      COUNT(aa.id) answered
      FROM attempts a LEFT JOIN attempt_answers aa ON aa.attempt_id=a.id
      WHERE a.student_id=? AND a.created_at>=?""",(student_id,week_start)).fetchone()
    weekly_data={'assessments':int(weekly['assessments'] or 0),'minutes':round(float(weekly['seconds'] or 0)/60),
                 'answered':int(weekly['answered'] or 0),'correct':int(weekly['correct'] or 0),'goal':3}

    # Find real improvement by comparing a student's earliest and latest result for the same tracked area.
    area_rows=c.execute("""SELECT a.id,a.created_at,q.subject,COALESCE(NULLIF(q.learning_outcome,''),NULLIF(q.subtopic,''),q.topic) area,
      COUNT(*) answered,100.0*SUM(aa.is_correct)/COUNT(*) accuracy
      FROM attempt_answers aa JOIN attempts a ON a.id=aa.attempt_id JOIN questions q ON q.id=aa.question_db_id
      WHERE a.student_id=? AND COALESCE(NULLIF(q.learning_outcome,''),NULLIF(q.subtopic,''),q.topic,'')<>''
      GROUP BY a.id,q.subject,COALESCE(NULLIF(q.learning_outcome,''),NULLIF(q.subtopic,''),q.topic)
      ORDER BY a.created_at""",(student_id,)).fetchall()
    histories={}
    for r in area_rows:
        histories.setdefault((r['subject'],r['area']),[]).append(float(r['accuracy'] or 0))
    improvements=[]
    for (subject,area), vals in histories.items():
        if len(vals)>=2:
            delta=round(vals[-1]-vals[0],1)
            if delta>0:
                improvements.append({'subject':subject,'area':area,'first':round(vals[0],1),'latest':round(vals[-1],1),'change':delta})
    improvements.sort(key=lambda x:x['change'],reverse=True)

    # Assignment architecture is already available even before teacher-side creation UI is activated.
    assignments=c.execute("""SELECT asg.*,u.full_name teacher_name,ast.status student_status
      FROM assignment_students ast JOIN assignments asg ON asg.id=ast.assignment_id
      JOIN users u ON u.id=asg.teacher_id
      WHERE ast.student_id=? AND asg.status='active' AND COALESCE(ast.status,'assigned')<>'completed'
      ORDER BY CASE WHEN COALESCE(asg.due_at,'')='' THEN 1 ELSE 0 END,asg.due_at LIMIT 5""",(student_id,)).fetchall()

    active=c.execute("""SELECT * FROM assessment_sessions WHERE student_id=? AND status='in_progress'
      ORDER BY started_at DESC LIMIT 1""",(student_id,)).fetchone()

    return {'subjects':subjects,'chapters':chapters,'weekly':weekly_data,'improvements':improvements[:3],
            'assignments':assignments,'active_assessment':active}


def institution_access(c,user_id):
    u=c.execute("SELECT role,primary_institution_id FROM users WHERE id=?",(user_id,)).fetchone()
    if not u: return None
    if u['role']=='admin' and u['primary_institution_id']:
        return {'institution_id':u['primary_institution_id'],'institution_role':'admin'}
    staff=c.execute("""SELECT institution_id,institution_role FROM institution_staff
      WHERE user_id=? AND active=1 ORDER BY CASE institution_role WHEN 'admin' THEN 1 WHEN 'manager' THEN 2 ELSE 3 END LIMIT 1""",(user_id,)).fetchone()
    return dict(staff) if staff else None


def institution_dashboard_data(c,institution_id):
    inst=c.execute("SELECT * FROM institutions WHERE id=?",(institution_id,)).fetchone()
    if not inst: return None
    student_count=c.execute("SELECT COUNT(*) n FROM users WHERE role='student' AND primary_institution_id=?",(institution_id,)).fetchone()['n']
    teacher_count=c.execute("SELECT COUNT(*) n FROM users WHERE role='teacher' AND primary_institution_id=?",(institution_id,)).fetchone()['n']
    class_count=c.execute("SELECT COUNT(*) n FROM classrooms WHERE institution_id=?",(institution_id,)).fetchone()['n']
    perf=c.execute("""SELECT COUNT(a.id) attempts,ROUND(AVG(a.score),1) avg_score,
      COUNT(DISTINCT CASE WHEN a.created_at>=datetime('now','-30 day') THEN a.student_id END) active_students
      FROM attempts a JOIN users u ON u.id=a.student_id WHERE u.primary_institution_id=?""",(institution_id,)).fetchone()
    classes=c.execute("""SELECT cl.id,cl.name,cl.level,cl.subject,u.full_name teacher_name,
      COUNT(DISTINCT cs.student_id) students,COUNT(DISTINCT a.id) attempts,ROUND(AVG(a.score),1) avg_score
      FROM classrooms cl LEFT JOIN users u ON u.id=cl.teacher_id
      LEFT JOIN classroom_students cs ON cs.classroom_id=cl.id
      LEFT JOIN attempts a ON a.student_id=cs.student_id
      WHERE cl.institution_id=? GROUP BY cl.id ORDER BY CASE WHEN AVG(a.score) IS NULL THEN 1 ELSE 0 END,AVG(a.score) ASC""",(institution_id,)).fetchall()
    weak=c.execute("""SELECT q.subject,COALESCE(NULLIF(q.learning_outcome,''),NULLIF(q.subtopic,''),q.topic) area,
      COUNT(*) answered,ROUND(100.0*SUM(aa.is_correct)/COUNT(*),1) accuracy,COUNT(DISTINCT a.student_id) students
      FROM users u JOIN attempts a ON a.student_id=u.id JOIN attempt_answers aa ON aa.attempt_id=a.id
      JOIN questions q ON q.id=aa.question_db_id WHERE u.primary_institution_id=?
      AND COALESCE(NULLIF(q.learning_outcome,''),NULLIF(q.subtopic,''),q.topic,'')<>''
      GROUP BY q.subject,COALESCE(NULLIF(q.learning_outcome,''),NULLIF(q.subtopic,''),q.topic)
      HAVING COUNT(*)>=3 ORDER BY accuracy ASC,answered DESC LIMIT 10""",(institution_id,)).fetchall()
    licence=c.execute("""SELECT il.*,p.name plan_name,
      (SELECT COUNT(*) FROM institution_license_users ilu WHERE ilu.institution_license_id=il.id) allocated
      FROM institution_licenses il JOIN plans p ON p.id=il.plan_id
      WHERE il.institution_id=? AND il.status='active' ORDER BY il.id DESC LIMIT 1""",(institution_id,)).fetchone()
    assignment=c.execute("""SELECT COUNT(*) total,
      SUM(CASE WHEN ast.status='completed' THEN 1 ELSE 0 END) completed
      FROM assignments ass JOIN classrooms cl ON cl.id=ass.classroom_id
      JOIN assignment_students ast ON ast.assignment_id=ass.id WHERE cl.institution_id=?""",(institution_id,)).fetchone()
    total=int(assignment['total'] or 0); done=int(assignment['completed'] or 0)
    return {'institution':inst,'student_count':student_count,'teacher_count':teacher_count,'class_count':class_count,
            'attempt_count':int(perf['attempts'] or 0),'avg_score':perf['avg_score'],'active_students':int(perf['active_students'] or 0),
            'classes':classes,'weak_areas':weak,'licence':licence,'assignment_completion':round(100.0*done/total,1) if total else 0}


def class_analytics(c,cid):
    students=c.execute("""SELECT u.id,u.system_user_id,u.full_name,u.mobile,cs.roll_no,COUNT(a.id) tests_completed,ROUND(AVG(a.score),1) avg_score,MAX(a.created_at) last_attempt FROM classroom_students cs JOIN users u ON u.id=cs.student_id LEFT JOIN attempts a ON a.student_id=u.id WHERE cs.classroom_id=? GROUP BY u.id,u.system_user_id,u.full_name,u.mobile,cs.roll_no ORDER BY CASE WHEN AVG(a.score) IS NULL THEN 1 ELSE 0 END,AVG(a.score) ASC""",(cid,)).fetchall()
    struggling=[s for s in students if s['avg_score'] is not None and float(s['avg_score'])<60]
    inactive=[s for s in students if int(s['tests_completed'] or 0)==0]
    weak=c.execute("""SELECT COALESCE(q.subtopic,q.topic) area,COUNT(*) answered,ROUND(100.0*SUM(aa.is_correct)/COUNT(*),1) accuracy FROM classroom_students cs JOIN attempts a ON a.student_id=cs.student_id JOIN attempt_answers aa ON aa.attempt_id=a.id JOIN questions q ON q.id=aa.question_db_id WHERE cs.classroom_id=? GROUP BY COALESCE(q.subtopic,q.topic) ORDER BY accuracy ASC LIMIT 5""",(cid,)).fetchall()
    return students,struggling,inactive,weak

def teacher_dashboard_intelligence(c, teacher_id):
    class_ids=[r['id'] for r in c.execute("SELECT id FROM classrooms WHERE teacher_id=?",(teacher_id,)).fetchall()]
    if not class_ids:
        return {'students':0,'active_assignments':0,'completion':0,'priority':None,'recent_assignments':[]}
    ph=','.join('?' for _ in class_ids)
    students=c.execute(f"SELECT COUNT(DISTINCT student_id) n FROM classroom_students WHERE classroom_id IN ({ph})",class_ids).fetchone()['n']
    active=c.execute("SELECT COUNT(*) n FROM assignments WHERE teacher_id=? AND status='active'",(teacher_id,)).fetchone()['n']
    completion_row=c.execute("""SELECT COUNT(*) total, SUM(CASE WHEN ast.status='completed' THEN 1 ELSE 0 END) done
        FROM assignment_students ast JOIN assignments a ON a.id=ast.assignment_id WHERE a.teacher_id=?""",(teacher_id,)).fetchone()
    total=int(completion_row['total'] or 0); done=int(completion_row['done'] or 0)
    completion=round(100.0*done/total,1) if total else 0
    priority=c.execute(f"""SELECT q.subject,COALESCE(NULLIF(q.learning_outcome,''),NULLIF(q.subtopic,''),q.topic) area,
        COUNT(*) answered,ROUND(100.0*SUM(aa.is_correct)/COUNT(*),1) accuracy,COUNT(DISTINCT a.student_id) students
        FROM classroom_students cs JOIN attempts a ON a.student_id=cs.student_id
        JOIN attempt_answers aa ON aa.attempt_id=a.id JOIN questions q ON q.id=aa.question_db_id
        WHERE cs.classroom_id IN ({ph}) AND COALESCE(NULLIF(q.learning_outcome,''),NULLIF(q.subtopic,''),q.topic,'')<>''
        GROUP BY q.subject,COALESCE(NULLIF(q.learning_outcome,''),NULLIF(q.subtopic,''),q.topic)
        HAVING COUNT(*)>=3 ORDER BY accuracy ASC,answered DESC LIMIT 1""",class_ids).fetchone()
    recent=c.execute("""SELECT a.*,cl.name class_name,COUNT(ast.student_id) assigned_count,
        SUM(CASE WHEN ast.status='completed' THEN 1 ELSE 0 END) completed_count,ROUND(AVG(at.score),1) recovery_avg
        FROM assignments a LEFT JOIN classrooms cl ON cl.id=a.classroom_id
        LEFT JOIN assignment_students ast ON ast.assignment_id=a.id
        LEFT JOIN attempts at ON at.id=ast.attempt_id
        WHERE a.teacher_id=? GROUP BY a.id ORDER BY a.created_at DESC LIMIT 5""",(teacher_id,)).fetchall()
    return {'students':students,'active_assignments':active,'completion':completion,'priority':priority,'recent_assignments':recent}

def class_intervention_analytics(c, cid):
    lo_rows=c.execute("""SELECT q.subject,COALESCE(NULLIF(q.learning_outcome,''),NULLIF(q.subtopic,''),q.topic) area,
      COUNT(*) answered,ROUND(100.0*SUM(aa.is_correct)/COUNT(*),1) accuracy,COUNT(DISTINCT a.student_id) students
      FROM classroom_students cs JOIN attempts a ON a.student_id=cs.student_id
      JOIN attempt_answers aa ON aa.attempt_id=a.id JOIN questions q ON q.id=aa.question_db_id
      WHERE cs.classroom_id=? AND COALESCE(NULLIF(q.learning_outcome,''),NULLIF(q.subtopic,''),q.topic,'')<>''
      GROUP BY q.subject,COALESCE(NULLIF(q.learning_outcome,''),NULLIF(q.subtopic,''),q.topic)
      HAVING COUNT(*)>=2 ORDER BY accuracy ASC,answered DESC LIMIT 12""",(cid,)).fetchall()
    priority=lo_rows[0] if lo_rows else None
    affected=[]
    if priority:
        affected=c.execute("""SELECT u.id,u.full_name,COUNT(*) answered,ROUND(100.0*SUM(aa.is_correct)/COUNT(*),1) accuracy
          FROM classroom_students cs JOIN users u ON u.id=cs.student_id
          JOIN attempts a ON a.student_id=u.id JOIN attempt_answers aa ON aa.attempt_id=a.id JOIN questions q ON q.id=aa.question_db_id
          WHERE cs.classroom_id=? AND COALESCE(NULLIF(q.learning_outcome,''),NULLIF(q.subtopic,''),q.topic)=?
          GROUP BY u.id,u.full_name HAVING COUNT(*)>=1 AND (100.0*SUM(aa.is_correct)/COUNT(*))<60 ORDER BY accuracy ASC""",(cid,priority['area'])).fetchall()
    misconceptions=c.execute("""SELECT aa.misconception_triggered misconception,COUNT(*) signals,COUNT(DISTINCT a.student_id) students
      FROM classroom_students cs JOIN attempts a ON a.student_id=cs.student_id JOIN attempt_answers aa ON aa.attempt_id=a.id
      WHERE cs.classroom_id=? AND COALESCE(aa.misconception_triggered,'')<>''
      GROUP BY aa.misconception_triggered ORDER BY signals DESC LIMIT 6""",(cid,)).fetchall()
    assignments=c.execute("""SELECT a.*,COUNT(ast.student_id) assigned_count,
      SUM(CASE WHEN ast.status='completed' THEN 1 ELSE 0 END) completed_count,ROUND(AVG(at.score),1) recovery_avg
      FROM assignments a LEFT JOIN assignment_students ast ON ast.assignment_id=a.id
      LEFT JOIN attempts at ON at.id=ast.attempt_id
      WHERE a.classroom_id=? GROUP BY a.id ORDER BY a.created_at DESC LIMIT 10""",(cid,)).fetchall()
    return {'lo_rows':lo_rows,'priority':priority,'affected':affected,'misconceptions':misconceptions,'assignments':assignments}

def ai_answer(c,student_id,prompt):
    a=student_analytics(c,student_id); p=(prompt or '').lower()
    if any(k in p for k in ['revise','study','next']):
        if a['weakest']:
            w=a['weakest']; return f"Your next priority should be {w['area']} in {w['subject']}. Your current accuracy there is {w['accuracy']}%. Take a targeted practice test, review every mistake, then retest."
        return 'Complete a few assessments first so I can identify your strongest and weakest areas.'
    if any(k in p for k in ['progress','doing','performance']): return f"You have completed {a['tests_completed']} tests with an average score of {a['avg_score']}% and an Academic Health Score of {a['health']}/100."
    if 'weak' in p:
        return f"Your weakest tracked area is {a['weakest']['area']} at {a['weakest']['accuracy']}% accuracy." if a['weakest'] else 'I do not yet have enough assessment history.'
    if 'strong' in p:
        return f"Your strongest tracked area is {a['strongest']['area']} at {a['strongest']['accuracy']}% accuracy." if a['strongest'] else 'I do not yet have enough assessment history.'
    return "Try asking: What should I revise next? How am I doing? What is my weakest area? What is my strongest area?"


def create_assessment_session(c,student_id,mode,duration_minutes,question_ids,meta):
    # One central safety gate: only question types the current renderer can mark reliably may enter live scored sessions.
    question_ids=list(dict.fromkeys(filter_live_question_ids(c,question_ids)))
    if not question_ids:
        raise ValueError('No approved auto-markable questions are available for this assessment yet.')
    expires_at=None
    governed_snapshot=(meta or {}).get('blueprint_snapshot') or {}
    governed_timing=bool(isinstance(governed_snapshot,dict) and governed_snapshot.get('release_state')=='RELEASED' and governed_snapshot.get('total_duration_seconds'))
    if (mode in ('exam','mock') or governed_timing) and duration_minutes:
        expires_at=(datetime.now()+timedelta(minutes=duration_minutes)).isoformat(timespec='seconds')
    ph_release_pins,ph_question_pins=integration_v1.build_session_content_pins(c,question_ids)
    cur=c.execute("""INSERT INTO assessment_sessions(
        student_id,mode,duration_minutes,expires_at,status,current_index,
        question_ids,flagged_ids,saved_answers,meta_json,assessment_blueprint_id,blueprint_source_id,
        blueprint_version,framework_version,blueprint_snapshot_json,assembly_policy_id,assembly_policy_version,
        ph_release_pins_json,ph_question_pins_json
    ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",(
        student_id,mode,duration_minutes,expires_at,'in_progress',0,
        ','.join(str(i) for i in question_ids),'','{}',json.dumps(meta),meta.get('assessment_blueprint_id'),
        meta.get('blueprint_source_id',''),meta.get('blueprint_version',''),meta.get('framework_version',''),
        json.dumps(meta.get('blueprint_snapshot') or {}),meta.get('assembly_policy_id'),meta.get('assembly_policy_version',''),
        integration_v1.canonical_json(ph_release_pins),integration_v1.canonical_json(ph_question_pins)
    ))
    c.commit()
    return cur.lastrowid

def get_assessment_session(c,assessment_id,student_id):
    return c.execute(
        'SELECT * FROM assessment_sessions WHERE id=? AND student_id=?',
        (assessment_id,student_id)
    ).fetchone()

def parse_ids(raw):
    return [int(x) for x in (raw or '').split(',') if x.strip().isdigit()]

def load_answers(row):
    try:
        return json.loads(row['saved_answers'] or '{}')
    except Exception:
        return {}

def load_json_map(row, field):
    try:
        return json.loads(row[field] or '{}') if field in row.keys() else {}
    except Exception:
        return {}

def seconds_left(row):
    if not row or row['mode'] not in ('exam','mock') or not row['expires_at']:
        return None
    try:
        return max(0,int((datetime.fromisoformat(row['expires_at'])-datetime.now()).total_seconds()))
    except Exception:
        return None

def iso_today():
    return datetime.now().date().isoformat()


def money_display(minor, currency='PKR'):
    if minor is None:
        return 'Price set at launch'
    amount=float(minor or 0)/100.0
    if currency=='PKR':
        return f'PKR {amount:,.0f}'
    if currency=='GBP':
        return f'£{amount:,.2f}'
    return f'{currency} {amount:,.2f}'


MASTER_LEVELS=['Foundation','Exam Ready','Advanced','Distinction','Expert','Elite']
ACCESS_CODES={
  'free_access': {'name':'Free Access','rank':0,'ceiling':'Foundation'},
  'level_1_access': {'name':'Level 1 Access','rank':1,'ceiling':'Exam Ready'},
  'level_2_access': {'name':'Level 2 Access','rank':2,'ceiling':'Distinction'},
  'full_access': {'name':'Full Access','rank':3,'ceiling':'Elite'}
}
LEGACY_FULL_ACCESS_CODES={'premium_monthly','premium_annual','institution_student'}

def mastery_rank(level):
    try: return MASTER_LEVELS.index(level)
    except ValueError: return 0

def _access_shape(code,name=None,source='free',entitlements=None,subscription=None,extra=None):
    canonical=code
    if code in LEGACY_FULL_ACCESS_CODES: canonical='full_access'
    cfg=ACCESS_CODES.get(canonical,ACCESS_CODES['free_access'])
    resolved_name=name or cfg['name']
    out={'plan_code':canonical,'legacy_plan_code':code,'plan_name':resolved_name,'name':resolved_name,'source':source,
         'entitlements':entitlements or {},'subscription':subscription,'access_rank':cfg['rank'],
         'mastery_ceiling':cfg['ceiling']}
    if extra: out.update(extra)
    return out

def get_access_profile(c, user_id):
    """Resolve the learner's cumulative V5.4 access independently from earned mastery."""
    user=c.execute('SELECT * FROM users WHERE id=?',(user_id,)).fetchone()
    if not user:
        return _access_shape('free_access','No plan','none',{})
    if user['role']!='student':
        fallback='free_teacher'
        plan=c.execute('SELECT * FROM plans WHERE code=?',(fallback,)).fetchone()
        return {'plan_code':fallback,'plan_name':plan['name'] if plan else 'Free Teacher','source':'free',
                'entitlements':safe_json(plan['entitlements_json'],{}) if plan else {},'subscription':None,
                'access_rank':0,'mastery_ceiling':'Foundation'}

    # Internal-live testing can explicitly open the complete learner journey without
    # changing earned mastery or weakening the default/free commercial semantics.
    # This is opt-in: SCOREMAX_INTERNAL_FULL_ACCESS=1 and paywall enforcement must be off.
    if INTERNAL_FULL_ACCESS_ENABLED and not COMMERCIAL_GATES_ENABLED:
        plan=c.execute("SELECT * FROM plans WHERE code='full_access'").fetchone()
        return _access_shape('full_access','Full Access','internal_live_full_access',
                             safe_json(plan['entitlements_json'],{}) if plan else {})

    override=(user['access_override_code'] or '').strip()
    if override in ACCESS_CODES:
        plan=c.execute('SELECT * FROM plans WHERE code=?',(override,)).fetchone()
        return _access_shape(override, plan['name'] if plan else ACCESS_CODES[override]['name'], 'admin_override',
                             safe_json(plan['entitlements_json'],{}) if plan else {})

    today=iso_today()
    inst=c.execute("""SELECT il.*,p.code plan_code,p.name plan_name,p.entitlements_json
        FROM institution_license_users ilu
        JOIN institution_licenses il ON il.id=ilu.institution_license_id
        JOIN plans p ON p.id=il.plan_id
        WHERE ilu.user_id=? AND il.status='active' AND il.starts_at<=?
          AND (il.ends_at IS NULL OR il.ends_at='' OR il.ends_at>=?)
        ORDER BY il.id DESC LIMIT 1""",(user_id,today,today)).fetchone()
    if inst:
        return _access_shape('full_access','Full Access','institution',safe_json(inst['entitlements_json'],{}),None,
                             {'institution_license':dict(inst)})

    sub=c.execute("""SELECT s.*,p.code plan_code,p.name plan_name,p.entitlements_json,p.currency,p.price_minor
        FROM subscriptions s JOIN plans p ON p.id=s.plan_id
        WHERE s.user_id=? AND s.status IN ('active','trial') AND s.starts_at<=?
          AND (s.ends_at IS NULL OR s.ends_at='' OR s.ends_at>=?)
        ORDER BY s.id DESC LIMIT 1""",(user_id,today,today)).fetchone()
    if sub:
        code=sub['plan_code']
        return _access_shape(code,sub['plan_name'],sub['source'] or 'subscription',safe_json(sub['entitlements_json'],{}),dict(sub))

    plan=c.execute("SELECT * FROM plans WHERE code='free_access'").fetchone()
    return _access_shape('free_access','Free Access','free',safe_json(plan['entitlements_json'],{}) if plan else {})

def has_entitlement(c, user_id, entitlement):
    return bool(get_access_profile(c,user_id)['entitlements'].get(entitlement,False))


def student_programme(c,user_id):
    row=c.execute("SELECT academic_level,COALESCE(active_programme,'') active_programme FROM users WHERE id=?",(user_id,)).fetchone()
    if not row: return ''
    return (row['active_programme'] or row['academic_level'] or '').strip()


def programme_short_label(value):
    folded=(value or '').strip().casefold()
    for item in STUDENT_PROGRAMME_CHOICES:
        if folded in {item['value'].casefold(),item['label'].casefold(),item['code'].casefold()}:
            return item['label']
    return (value or '').strip()


def student_programme_options(c,user_id):
    active=student_programme(c,user_id)
    out=[]
    for item in STUDENT_PROGRAMME_CHOICES:
        row=dict(item)
        row['active']=item['value'].casefold()==active.casefold()
        # Availability is informative only; programmes remain visible even when content is still coming.
        row['question_count']=int(c.execute(f"SELECT COUNT(*) n FROM questions q WHERE {live_question_clause('q')} AND (lower(COALESCE(q.programme,''))=lower(?) OR lower(COALESCE(q.qualification,''))=lower(?))",(item['value'],item['value'])).fetchone()['n'] or 0)
        out.append(row)
    return out


def student_coverage_profile(c,user_id):
    programme=student_programme(c,user_id)
    return commercial_access.effective_coverage(c,user_id,programme,commercial_gates_enabled=COMMERCIAL_GATES_ENABLED)


def student_subject_state(c,user_id,subject,programme=''):
    programme=programme or student_programme(c,user_id)
    aliases=_programme_aliases(programme)
    scope_clause,scope_params=_programme_scope_sql(aliases,'q')
    row=c.execute(f"SELECT COUNT(*) n FROM questions q WHERE {live_question_clause('q')} AND {scope_clause} AND lower(q.subject)=lower(?)",scope_params+[subject]).fetchone()
    return commercial_access.subject_state(c,user_id,programme,subject,available=bool(row and row['n']),commercial_gates_enabled=COMMERCIAL_GATES_ENABLED)


def subject_access_redirect(c,user_id,subject,programme=''):
    state=student_subject_state(c,user_id,subject,programme)
    if state=='LOCKED':
        return redirect(url_for('access_account',locked_subject=subject))
    return None



def access_allows_level(c,user_id,level):
    profile=get_access_profile(c,user_id)
    return mastery_rank(level)<=mastery_rank(profile['mastery_ceiling'])

def mastery_scope_key(scope_type,programme='',subject='',chapter=''):
    scope_type=(scope_type or '').lower()
    if scope_type=='chapter': return f"{programme}|{subject}|{chapter}"
    if scope_type=='subject': return f"{programme}|{subject}"
    return programme or 'overall'

def mastery_scope_ceiling(scope_type):
    return {'chapter':'Distinction','subject':'Expert','overall':'Elite'}.get((scope_type or '').lower(),'Distinction')

CHAPTER_MASTERY_LEVELS=['Foundation','Exam Ready','Advanced','Distinction']
CHAPTER_MASTERY_PROGRESS={'Foundation':25,'Exam Ready':50,'Advanced':75,'Distinction':100}

def chapter_mastery_opportunity(c,student_id,subject,chapter,programme=''):
    """Learner-facing chapter mastery snapshot.

    Existing mastery comes only from governed formal mastery_records. Potential mastery
    is the highest chapter level the *current reviewed production bank* can support
    under the active mastery/assembly requirements. It is intentionally not inferred
    from raw accuracy and demo questions never raise potential mastery.
    """
    programme=(programme or student_programme(c,student_id) or '').strip()
    scope_key=mastery_scope_key('chapter',programme,subject,chapter)
    rec=c.execute("""SELECT * FROM mastery_records WHERE student_id=? AND scope_type='chapter'
      AND scope_key=? ORDER BY id DESC LIMIT 1""",(student_id,scope_key)).fetchone()
    if not rec:
        # Preserve older records created before an exact programme key was available.
        rec=c.execute("""SELECT * FROM mastery_records WHERE student_id=? AND scope_type='chapter'
          AND lower(COALESCE(subject,''))=lower(?) AND COALESCE(chapter,'')=?
          ORDER BY updated_at DESC,id DESC LIMIT 1""",(student_id,subject,chapter)).fetchone()
    existing_level=(rec['mastery_level'] or '').strip() if rec else ''
    existing_status=effective_mastery_status(rec) if rec else 'Not verified'

    governing_blueprint=active_assessment_blueprint(c,programme) if programme else None
    assembly_policy=active_assembly_policy(c,
      governing_blueprint['id'] if governing_blueprint else None,
      governing_blueprint['framework_version_id'] if governing_blueprint else None,
      programme=programme,subject=subject,chapter=chapter,assessment_type='mastery')

    clauses=[live_question_clause('q'),"COALESCE(q.is_demo,0)=0",
             "COALESCE(q.calibration_status,'PROVISIONAL') NOT IN ('BLOCKED','REVIEW_NEGATIVE')",
             "lower(COALESCE(q.subject,''))=lower(?)","COALESCE(q.chapter,'')=?"]
    params=[subject,chapter]
    if programme:
        aliases=_programme_aliases(programme)
        scope_clause,scope_params=_programme_scope_sql(aliases,'q')
        clauses.append(scope_clause); params+=scope_params
    all_rows=c.execute(f"SELECT q.* FROM questions q WHERE {' AND '.join(clauses)}",params).fetchall()
    markable=[r for r in all_rows if canonical_question_type(r) in LIVE_MARKABLE_TYPES]

    potential_level=''
    for level in CHAPTER_MASTERY_LEVELS:
        policy=mastery_policy(c,level)
        if not policy: continue
        effective=effective_mastery_requirements(policy,assembly_policy)
        pool=[r for r in markable if mastery_rank(r['level'] or 'Foundation')<=mastery_rank(level)]
        min_q=int(effective.get('min_questions') or policy['min_questions'] or 10)
        if len(pool)<min_q: continue
        target=[r for r in pool if (r['level'] or 'Foundation')==level]
        target_needed=max(1,int(round(min_q*float(effective.get('target_band_pct') or policy['target_band_pct'] or .25))))
        if len(target)<target_needed: continue
        families={((r['family_id'] or '').strip() or f"Q-{r['id']}") for r in pool}
        target_families={((r['family_id'] or '').strip() or f"Q-{r['id']}") for r in target}
        if len(families)<min_q or len(target_families)<target_needed: continue
        potential_level=level

    # Earned mastery is never visually erased just because later inventory/access changed.
    if existing_level in CHAPTER_MASTERY_LEVELS and (not potential_level or mastery_rank(existing_level)>mastery_rank(potential_level)):
        potential_level=existing_level

    existing_pct=CHAPTER_MASTERY_PROGRESS.get(existing_level,0)
    potential_pct=CHAPTER_MASTERY_PROGRESS.get(potential_level,0)
    access=get_access_profile(c,student_id)
    access_ceiling=access.get('mastery_ceiling','Foundation')
    access_limited=bool(potential_level and mastery_rank(potential_level)>mastery_rank(access_ceiling))
    opportunity=max(0,potential_pct-existing_pct)
    return {
      'existing_level':existing_level or 'Not verified',
      'existing_status':existing_status,
      'existing_pct':existing_pct,
      'potential_level':potential_level or 'Bank building',
      'potential_pct':potential_pct,
      'opportunity_pct':opportunity,
      'access_ceiling':access_ceiling,
      'access_limited':access_limited,
      'production_questions':len(markable),
      'has_formal_mastery':bool(rec)
    }

def current_mastery_records(c,student_id):
    rows=c.execute("SELECT * FROM mastery_records WHERE student_id=? ORDER BY scope_type,scope_key",(student_id,)).fetchall()
    changed=False; out=[]
    for r in rows:
        d=dict(r); effective=effective_mastery_status(r)
        if effective=='Verification Due' and r['status']!='Verification Due':
            record_mastery_history(c,r,'verification_due',new_status='Verification Due',
                                   note='Verification date passed; fresh evidence is required.')
            c.execute("UPDATE mastery_records SET status='Verification Due',updated_at=CURRENT_TIMESTAMP WHERE id=?",(r['id'],))
            d['status']='Verification Due'; changed=True
        else:
            d['status']=effective
        out.append(d)
    if changed: c.commit()
    return out

def mastery_policy(c,level):
    return c.execute("SELECT * FROM mastery_policies WHERE mastery_level=? AND active=1",(level,)).fetchone()

def _mastery_previous_families(c,student_id,scope_type,scope_key,target_level):
    rows=c.execute("""SELECT family_ids_json FROM mastery_form_results
      WHERE student_id=? AND scope_type=? AND scope_key=? AND target_level=? ORDER BY id DESC LIMIT 5""",
      (student_id,scope_type,scope_key,target_level)).fetchall()
    seen=set()
    for r in rows:
        seen.update(str(x) for x in safe_json(r['family_ids_json'],[]) if x)
    return seen

def build_mastery_form(c,student_id,scope_type,target_level,programme='',subject='',chapter=''):
    scope_type=(scope_type or 'chapter').lower(); target_level=(target_level or 'Foundation').strip()
    if target_level not in MASTER_LEVELS: raise ValueError('Choose a valid mastery level.')
    if mastery_rank(target_level)>mastery_rank(mastery_scope_ceiling(scope_type)):
        raise ValueError(f"{target_level} is above the {scope_type.title()} mastery ceiling.")
    if not access_allows_level(c,student_id,target_level):
        raise PermissionError('Your current Access level does not unlock this mastery assessment yet.')
    policy=mastery_policy(c,target_level)
    if not policy: raise ValueError('Mastery policy is not configured.')
    if not programme:
        u=c.execute("SELECT academic_level FROM users WHERE id=?",(student_id,)).fetchone()
        programme=(u['academic_level'] or '').strip() if u else ''
    governing_blueprint=active_assessment_blueprint(c,programme) if programme else None
    assembly_policy=active_assembly_policy(c,
      governing_blueprint['id'] if governing_blueprint else None,
      governing_blueprint['framework_version_id'] if governing_blueprint else None,
      programme=programme,subject=subject,chapter=chapter,assessment_type='mastery')
    effective_policy=effective_mastery_requirements(policy,assembly_policy)
    clauses=[live_question_clause('q'),"COALESCE(q.calibration_status,'PROVISIONAL') NOT IN ('BLOCKED','REVIEW_NEGATIVE')"]
    params=[]
    if programme:
        clauses.append("(lower(COALESCE(q.programme,''))=lower(?) OR lower(COALESCE(q.qualification,''))=lower(?))")
        params.extend([programme,programme])
    if subject:
        clauses.append('q.subject=?'); params.append(subject)
    if scope_type=='chapter':
        if not chapter: raise ValueError('Choose a chapter for chapter mastery.')
        clauses.append('q.chapter=?'); params.append(chapter)
    all_rows=c.execute(f"SELECT q.* FROM questions q WHERE {' AND '.join(clauses)} ORDER BY RANDOM()",params).fetchall()
    rows=[r for r in all_rows if canonical_question_type(r) in LIVE_MARKABLE_TYPES and mastery_rank(r['level'] or 'Foundation')<=mastery_rank(target_level)]
    prod=[r for r in rows if not int(r['is_demo'] or 0)]
    demo=[r for r in rows if int(r['is_demo'] or 0)]
    min_q=int(effective_policy.get('min_questions') or policy['min_questions'] or 10)
    if prod:
        if len(prod)<min_q:
            raise ValueError('The reviewed production bank is not deep enough for a defensible mastery form yet. Add more reviewed question families in Power House.')
        pool=prod; demo_only=False
    else:
        pool=demo; demo_only=True
        if len(pool)<5: raise ValueError('Not enough demo questions are available to exercise this mastery journey.')
        min_q=min(min_q,len(pool))
    previous=_mastery_previous_families(c,student_id,scope_type,mastery_scope_key(scope_type,programme,subject,chapter),target_level)
    target_band=[r for r in pool if (r['level'] or 'Foundation')==target_level]
    easier=[r for r in pool if (r['level'] or 'Foundation')!=target_level]
    target_needed=max(1,int(round(min_q*float(effective_policy.get('target_band_pct') or policy['target_band_pct'] or .25))))
    if not demo_only and len(target_band)<target_needed:
        raise ValueError('This question bank does not yet contain enough target-level families for a defensible mastery form.')
    rigor=int(assembly_policy['rigor_score'] or 50) if assembly_policy else 50
    if rigor>=65:
        difficulty_order={'Difficult':0,'Moderate':1,'Easy':2}
    elif rigor<=35:
        difficulty_order={'Easy':0,'Moderate':1,'Difficult':2}
    else:
        difficulty_order={'Moderate':0,'Difficult':1,'Easy':2}
    def order_unseen(seq):
        return sorted(seq,key=lambda r:((r['family_id'] or '') in previous,
          difficulty_order.get(normalize_difficulty(r['difficulty'] or r['level']),1),random.random()))
    def add_question(r,chosen,used_families):
        fam=(r['family_id'] or f"Q-{r['id']}").strip()
        if fam in used_families: return False
        chosen.append(r); used_families.add(fam); return True

    # Expert and Elite are deliberately broader than chapter mastery. Expert must
    # sample a representative share of the available subject chapters; Elite must
    # represent every available subject in the programme. This is enforced while
    # constructing the form, rather than hoping a random draw happens to be broad.
    breadth_unit=''; breadth_required=0; breadth_available=[]
    if scope_type=='subject' and target_level=='Expert':
        breadth_unit='chapter'
        breadth_available=sorted({str(r['chapter'] or '').strip() for r in pool if str(r['chapter'] or '').strip()})
        if len(breadth_available)>=3:
            breadth_required=max(3,(len(breadth_available)+1)//2)
        else:
            breadth_required=3
    elif scope_type=='overall' and target_level=='Elite':
        breadth_unit='subject'
        breadth_available=sorted({str(r['subject'] or '').strip() for r in pool if str(r['subject'] or '').strip()})
        breadth_required=max(3,len(breadth_available))

    chosen=[]; used_families=set()
    if breadth_unit and not demo_only:
        if len(breadth_available)<breadth_required:
            label='chapters' if breadth_unit=='chapter' else 'subjects'
            raise ValueError(f'The reviewed bank is not broad enough for {target_level}: at least {breadth_required} distinct {label} are required.')
        units=list(breadth_available); random.shuffle(units)
        # Expert samples at least the representative minimum. Elite samples all available subjects.
        for unit in units[:breadth_required]:
            candidates=[r for r in target_band if str(r[breadth_unit] or '').strip()==unit]
            candidates+= [r for r in easier if str(r[breadth_unit] or '').strip()==unit]
            picked=False
            for r in order_unseen(candidates):
                if add_question(r,chosen,used_families): picked=True; break
            if not picked:
                raise ValueError(f'Not enough independent question families are available across the required {breadth_unit} coverage.')

    # Guarantee the target-level band after the breadth seed, then fill the rest
    # with independent families at or below the target level.
    for r in order_unseen(target_band):
        if len([x for x in chosen if (x['level'] or 'Foundation')==target_level])>=target_needed: break
        add_question(r,chosen,used_families)
    for r in order_unseen(target_band)+order_unseen(easier):
        if len(chosen)>=min_q: break
        add_question(r,chosen,used_families)
    if len(chosen)<min_q:
        # Demo forms may relax one-family-per-form purely to test the software journey.
        if demo_only:
            for r in pool:
                if r['id'] not in {x['id'] for x in chosen}: chosen.append(r)
                if len(chosen)>=min_q: break
        if len(chosen)<min_q: raise ValueError('Not enough independent question families for this mastery form.')
    if len([x for x in chosen if (x['level'] or 'Foundation')==target_level])<target_needed and not demo_only:
        raise ValueError('This mastery form cannot meet the required target-level question proportion yet.')
    unseen=sum(1 for r in chosen if (r['family_id'] or f"Q-{r['id']}") not in previous)
    unseen_ratio=round(unseen/len(chosen),3) if chosen else 0
    if previous and not demo_only and unseen_ratio<float(effective_policy.get('unseen_family_pct') or policy['unseen_family_pct'] or .6):
        raise ValueError('Not enough unseen question families are available for an independent reconfirmation form.')
    subjects={r['subject'] for r in chosen if r['subject']}; chapters={r['chapter'] for r in chosen if r['chapter']}
    if scope_type=='chapter':
        breadth_ok=len(chosen)>=min_q; breadth_covered=1 if breadth_ok else 0; breadth_required=max(1,breadth_required)
    elif scope_type=='subject':
        breadth_covered=len(chapters); breadth_required=breadth_required or 2; breadth_ok=breadth_covered>=breadth_required
    else:
        breadth_covered=len(subjects); breadth_required=breadth_required or 2; breadth_ok=breadth_covered>=breadth_required
    if not demo_only and not breadth_ok:
        raise ValueError(f'The available question bank is not broad enough for this mastery scope yet ({breadth_covered}/{breadth_required} required units represented).')
    meta={'scope':scope_type,'programme':programme,'subject':subject,'chapters':chapter,'level':target_level,
          'assessment_kind':'mastery','mastery_scope_type':scope_type,
          'mastery_scope_key':mastery_scope_key(scope_type,programme,subject,chapter),'mastery_target_level':target_level,
          'mastery_demo_only':demo_only,'mastery_breadth_ok':bool(breadth_ok),'mastery_unseen_family_ratio':unseen_ratio,
          'mastery_breadth_unit':breadth_unit or ('chapter' if scope_type=='subject' else 'subject' if scope_type=='overall' else 'question'),
          'mastery_breadth_required':breadth_required,'mastery_breadth_covered':breadth_covered,
          'mastery_family_ids':[r['family_id'] or f"Q-{r['id']}" for r in chosen],
          'mastery_effective_policy':effective_policy,
          'assessment_blueprint_id':governing_blueprint['id'] if governing_blueprint else None,
          'blueprint_source_id':governing_blueprint['powerhouse_blueprint_id'] if governing_blueprint else '',
          'blueprint_version':governing_blueprint['blueprint_version'] if governing_blueprint else '',
          'framework_version':governing_blueprint['framework_version_name'] if governing_blueprint else '',
          'blueprint_snapshot':blueprint_payload_from_record(c,governing_blueprint['id']) if governing_blueprint else {},
          'assembly_policy_id':assembly_policy['id'] if assembly_policy else None,
          'assembly_policy_version':assembly_policy['policy_version'] if assembly_policy else '1',
          'assessment_rigor_score':rigor}
    return chosen,meta

def _refresh_item_calibration(c,question_ids):
    now=datetime.now().isoformat(timespec='seconds')
    for qid in set(question_ids):
        q=c.execute("SELECT * FROM questions WHERE id=?",(qid,)).fetchone()
        if not q: continue
        if int(q['is_demo'] or 0):
            c.execute("UPDATE questions SET calibration_status='DEMO',calibrated_at=? WHERE id=?",(now,qid)); continue
        if (q['calibration_status'] or '')=='BLOCKED': continue
        stats=c.execute("SELECT COUNT(*) n,AVG(is_correct) facility FROM attempt_answers WHERE question_db_id=?",(qid,)).fetchone()
        n=int(stats['n'] or 0); facility=float(stats['facility'] or 0) if n else None
        disc=None
        if n>=8:
            split=c.execute("""SELECT AVG(CASE WHEN aa.is_correct=1 THEN a.score END) hi,
              AVG(CASE WHEN aa.is_correct=0 THEN a.score END) lo FROM attempt_answers aa JOIN attempts a ON a.id=aa.attempt_id
              WHERE aa.question_db_id=?""",(qid,)).fetchone()
            if split and split['hi'] is not None and split['lo'] is not None: disc=round((float(split['hi'])-float(split['lo']))/100.0,3)
        if n<5: status='PROVISIONAL'
        elif n<20: status='COLLECTING'
        elif disc is not None and disc<0: status='REVIEW_NEGATIVE'
        elif facility is not None and (facility<.15 or facility>.95): status='LOW_INFORMATION'
        else: status='EMPIRICAL'
        c.execute("UPDATE questions SET response_count=?,facility_value=?,discrimination_value=?,calibration_status=?,calibrated_at=? WHERE id=?",
                  (n,round(facility,3) if facility is not None else None,disc,status,now,qid))

def process_mastery_result(c,assessment,attempt_id,score,question_ids):
    meta=safe_json(assessment['meta_json'],{})
    if meta.get('assessment_kind')!='mastery': return
    demo_only=1 if meta.get('mastery_demo_only') else 0
    level=meta.get('mastery_target_level','Foundation')
    policy=mastery_policy(c,level)
    effective_policy=meta.get('mastery_effective_policy') or effective_mastery_requirements(policy,None)
    scope_type=meta.get('mastery_scope_type','chapter'); scope_key=meta.get('mastery_scope_key','')
    passed=bool(policy and score>=float(effective_policy.get('min_accuracy') or policy['min_accuracy'] or 0) and meta.get('mastery_breadth_ok'))
    effective_min_forms=int(effective_policy.get('min_forms') or policy['min_forms'] or 1) if policy else 1
    effective_verification_days=int(effective_policy.get('verification_days') or policy['verification_days'] or 90) if policy else 90
    policy_version=str(meta.get('assembly_policy_version') or effective_policy.get('assembly_policy_version') or '1')
    family_ids=meta.get('mastery_family_ids') or []
    c.execute("""INSERT INTO mastery_form_results(student_id,assessment_session_id,attempt_id,scope_type,scope_key,programme,subject,chapter,
      target_level,score,question_count,passed,demo_only,breadth_ok,unseen_family_ratio,family_ids_json,policy_snapshot_json,
      assembly_policy_id,assembly_policy_version,effective_policy_json)
      VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
      (assessment['student_id'],assessment['id'],attempt_id,scope_type,scope_key,meta.get('programme',''),meta.get('subject',''),meta.get('chapters',''),
       level,score,len(question_ids),1 if passed else 0,demo_only,1 if meta.get('mastery_breadth_ok') else 0,float(meta.get('mastery_unseen_family_ratio') or 0),
       json.dumps(family_ids),json.dumps(dict(policy) if policy else {}),meta.get('assembly_policy_id'),policy_version,json.dumps(effective_policy)))
    if demo_only or not policy: return

    existing=c.execute("SELECT * FROM mastery_records WHERE student_id=? AND scope_type=? AND scope_key=?",
                       (assessment['student_id'],scope_type,scope_key)).fetchone()
    if existing and effective_mastery_status(existing)=='Verification Due' and existing['status']!='Verification Due':
        record_mastery_history(c,existing,'verification_due',new_status='Verification Due',attempt_id=attempt_id,
                               assessment_session_id=assessment['id'],note='Verification date had passed when new mastery evidence arrived.')
        c.execute("UPDATE mastery_records SET status='Verification Due',updated_at=CURRENT_TIMESTAMP WHERE id=?",(existing['id'],))
        existing=c.execute("SELECT * FROM mastery_records WHERE id=?",(existing['id'],)).fetchone()

    held_level=(existing['mastery_level'] or 'Foundation') if existing else ''
    incoming_rank=mastery_rank(level); held_rank=mastery_rank(held_level) if existing else -1

    # Integrity rule: lower-level formal evidence is useful diagnostically, but it can
    # never renew, extend, downgrade or otherwise mutate a higher mastery record.
    if existing and incoming_rank<held_rank:
        record_mastery_history(c,existing,'lower_level_evidence_ignored',attempt_id=attempt_id,
                               assessment_session_id=assessment['id'],
                               note=f'{level} evidence cannot alter currently held {held_level} mastery.',
                               metadata={'submitted_level':level,'score':score,'passed':passed})
        return

    # Verification Due is only failed/reconfirmed by evidence at the level currently
    # held. A harder progression attempt is allowed, but failure at that harder level
    # must not count as a failed reconfirmation of the lower existing level.
    if existing and existing['status']=='Verification Due' and incoming_rank==held_rank:
        if not passed:
            fails=int(existing['failed_reconfirmations'] or 0)+1
            if fails>=2:
                new_level=MASTER_LEVELS[max(0,held_rank-1)]
                new_policy=mastery_policy(c,new_level)
                new_due=(datetime.now()+timedelta(days=int(new_policy['verification_days'] or 90) if new_policy else 90)).date().isoformat()
                record_mastery_history(c,existing,'reconfirmation_downgrade',new_level=new_level,new_status='Verified',attempt_id=attempt_id,
                                       assessment_session_id=assessment['id'],note='Two failed same-level reconfirmations lowered mastery by one level.')
                c.execute("UPDATE mastery_records SET mastery_level=?,status='Verified',failed_reconfirmations=0,verified_at=CURRENT_TIMESTAMP,verification_due_at=?,updated_at=CURRENT_TIMESTAMP WHERE id=?",
                          (new_level,new_due,existing['id']))
            else:
                record_mastery_history(c,existing,'reconfirmation_failed',new_status='Verification Due',attempt_id=attempt_id,
                                       assessment_session_id=assessment['id'],note='First failed same-level reconfirmation; mastery remains Verification Due.',
                                       metadata={'failed_reconfirmations':fails})
                c.execute("UPDATE mastery_records SET failed_reconfirmations=?,updated_at=CURRENT_TIMESTAMP WHERE id=?",(fails,existing['id']))
            return

        due_since=existing['verification_due_at'] or existing['updated_at'] or existing['verified_at'] or ''
        fresh_good=c.execute("""SELECT * FROM mastery_form_results
          WHERE student_id=? AND scope_type=? AND scope_key=? AND target_level=? AND demo_only=0 AND passed=1
            AND datetime(created_at)>=datetime(?) AND COALESCE(assembly_policy_version,'1')=?
          ORDER BY id DESC LIMIT ?""",
          (assessment['student_id'],scope_type,scope_key,held_level,due_since,policy_version,max(10,effective_min_forms+2))).fetchall()
        if len(fresh_good)<effective_min_forms:
            return
        now=datetime.now(); due=(now+timedelta(days=effective_verification_days)).date().isoformat()
        status='Verified'
        if held_level=='Elite' and os.environ.get('SCOREMAX_ELITE_CONFIRMATION','0')!='1': status='Elite Candidate'
        best=max(float(r['score'] or 0) for r in fresh_good)
        total=sum(int(r['question_count'] or 0) for r in fresh_good[:effective_min_forms])
        record_mastery_history(c,existing,'reconfirmed',new_level=held_level,new_status=status,attempt_id=attempt_id,
                               assessment_session_id=assessment['id'],note='Fresh same-level evidence satisfied the reconfirmation policy.',
                               metadata={'forms':len(fresh_good),'best_accuracy':best})
        c.execute("""UPDATE mastery_records SET programme=?,subject=?,chapter=?,mastery_level=?,status=?,verified_at=?,verification_due_at=?,
          best_accuracy=?,forms_passed=?,questions_total=?,failed_reconfirmations=0,updated_at=CURRENT_TIMESTAMP WHERE id=?""",
          (meta.get('programme',''),meta.get('subject',''),meta.get('chapters',''),held_level,status,now.isoformat(timespec='seconds'),due,
           best,len(fresh_good),total,existing['id']))
        return

    # Same-level evidence submitted before expiry can refresh mastery only if enough
    # fresh passing forms have been collected since the last verification. Old forms
    # cannot be recycled indefinitely, and a failed current form never refreshes it.
    if existing and incoming_rank==held_rank and existing['status'] in ('Verified','Elite Candidate'):
        if not passed:
            return
        fresh_since=existing['verified_at'] or existing['updated_at'] or ''
        fresh_good=c.execute("""SELECT * FROM mastery_form_results
          WHERE student_id=? AND scope_type=? AND scope_key=? AND target_level=? AND demo_only=0 AND passed=1
            AND datetime(created_at)>datetime(?) AND COALESCE(assembly_policy_version,'1')=?
          ORDER BY id DESC LIMIT ?""",
          (assessment['student_id'],scope_type,scope_key,held_level,fresh_since,policy_version,max(10,effective_min_forms+2))).fetchall()
        if len(fresh_good)<effective_min_forms:
            return
        now=datetime.now(); due=(now+timedelta(days=effective_verification_days)).date().isoformat()
        status='Verified' if held_level!='Elite' or os.environ.get('SCOREMAX_ELITE_CONFIRMATION','0')=='1' else 'Elite Candidate'
        best=max(float(r['score'] or 0) for r in fresh_good)
        total=sum(int(r['question_count'] or 0) for r in fresh_good[:effective_min_forms])
        record_mastery_history(c,existing,'verified_again',new_level=held_level,new_status=status,attempt_id=attempt_id,
                               assessment_session_id=assessment['id'],note='Fresh same-level evidence renewed mastery before its due date.')
        c.execute("""UPDATE mastery_records SET status=?,verified_at=?,verification_due_at=?,best_accuracy=?,forms_passed=?,questions_total=?,
          failed_reconfirmations=0,updated_at=CURRENT_TIMESTAMP WHERE id=?""",
          (status,now.isoformat(timespec='seconds'),due,best,len(fresh_good),total,existing['id']))
        return

    # New mastery or progression to a genuinely higher level.
    forms=c.execute("""SELECT * FROM mastery_form_results WHERE student_id=? AND scope_type=? AND scope_key=? AND target_level=? AND demo_only=0
      AND COALESCE(assembly_policy_version,'1')=? ORDER BY id DESC LIMIT ?""",
      (assessment['student_id'],scope_type,scope_key,level,policy_version,max(10,effective_min_forms+2))).fetchall()
    good=[r for r in forms if int(r['passed'] or 0)]
    if len(good)<effective_min_forms:
        return
    now=datetime.now(); due=(now+timedelta(days=effective_verification_days)).date().isoformat()
    status='Verified'
    if level=='Elite' and os.environ.get('SCOREMAX_ELITE_CONFIRMATION','0')!='1': status='Elite Candidate'
    best=max(float(r['score'] or 0) for r in good)
    total=sum(int(r['question_count'] or 0) for r in good[:effective_min_forms])
    if existing:
        record_mastery_history(c,existing,'upgraded',new_level=level,new_status=status,attempt_id=attempt_id,
                               assessment_session_id=assessment['id'],note=f'Formal evidence progressed mastery from {held_level} to {level}.',
                               metadata={'forms':len(good),'best_accuracy':best})
        c.execute("""UPDATE mastery_records SET programme=?,subject=?,chapter=?,mastery_level=?,status=?,verified_at=?,verification_due_at=?,
          best_accuracy=?,forms_passed=?,questions_total=?,failed_reconfirmations=0,updated_at=CURRENT_TIMESTAMP WHERE id=?""",
          (meta.get('programme',''),meta.get('subject',''),meta.get('chapters',''),level,status,now.isoformat(timespec='seconds'),due,
           best,len(good),total,existing['id']))
    else:
        cur=c.execute("""INSERT INTO mastery_records(student_id,scope_type,scope_key,programme,subject,chapter,mastery_level,status,verified_at,
          verification_due_at,best_accuracy,forms_passed,questions_total,failed_reconfirmations)
          VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,0)""",
          (assessment['student_id'],scope_type,scope_key,meta.get('programme',''),meta.get('subject',''),meta.get('chapters',''),level,status,
           now.isoformat(timespec='seconds'),due,best,len(good),total))
        rid=cur.lastrowid
        c.execute("""INSERT INTO mastery_history(mastery_record_id,student_id,scope_type,scope_key,event_type,previous_level,new_level,
          previous_status,new_status,attempt_id,assessment_session_id,note,metadata_json)
          VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",
          (rid,assessment['student_id'],scope_type,scope_key,'earned','',level,'',status,attempt_id,assessment['id'],
           'Initial mastery earned from formal evidence.',json.dumps({'forms':len(good),'best_accuracy':best})))

def challenge_readiness(c,student_id,scope_type,programme,subject='',chapter='',required_level=''):
    access=get_access_profile(c,student_id)
    level=required_level or ('Distinction' if scope_type.lower()=='chapter' else 'Expert' if scope_type.lower()=='subject' else 'Elite')
    if mastery_rank(level)>mastery_rank(access['mastery_ceiling']):
        return False,f"Requires access up to {level}."
    key=mastery_scope_key(scope_type,programme,subject,chapter)
    rec=c.execute("SELECT * FROM mastery_records WHERE student_id=? AND scope_type=? AND scope_key=?",
                  (student_id,scope_type.lower(),key)).fetchone()
    if not rec: return False,f"Earn verified {level} mastery first."
    current_status=effective_mastery_status(rec)
    if current_status=='Verification Due': return False,'Mastery verification is due.'
    if level=='Elite' and current_status!='Verified': return False,'Final Elite verification is required before the full-exam challenge unlocks.'
    if mastery_rank(rec['mastery_level'])<mastery_rank(level): return False,f"Earn verified {level} mastery first."
    return True,'Ready'

def create_subscription(c, user_id, plan_id, starts_at=None, ends_at=None, source='manual', provider='manual',
                        provider_ref='', auto_renew=0, status='active'):
    starts_at=starts_at or iso_today()
    cur=c.execute('''INSERT INTO subscriptions(user_id,plan_id,status,source,provider,provider_subscription_ref,starts_at,ends_at,renews_at,auto_renew)
        VALUES(?,?,?,?,?,?,?,?,?,?)''',(user_id,plan_id,status,source,provider,provider_ref,starts_at,ends_at,ends_at if auto_renew else None,auto_renew))
    return cur.lastrowid



def ensure_referral_code(c,user_id):
    u=c.execute("SELECT id,full_name,own_referral_code FROM users WHERE id=?",(user_id,)).fetchone()
    if not u: return ''
    if u['own_referral_code']: return u['own_referral_code']
    stem=''.join(ch for ch in (u['full_name'] or 'USER').upper() if ch.isalnum())[:5] or 'USER'
    code=f"{stem}{int(user_id):05d}"
    # Extremely unlikely collision; preserve deterministic readability while remaining unique.
    if c.execute("SELECT 1 FROM users WHERE own_referral_code=? AND id<>?",(code,user_id)).fetchone():
        code=f"SM{int(user_id):07d}"
    c.execute("UPDATE users SET own_referral_code=? WHERE id=?",(code,user_id))
    return code


def apply_referral_attribution(c,referred_user_id,entered_code='',source=''):
    """Lock the first valid referrer. Attribution cannot be silently reassigned later."""
    existing=c.execute("SELECT * FROM referral_attributions WHERE user_id=?",(referred_user_id,)).fetchone()
    if existing and existing['referrer_type']=='user' and existing['referrer_id']:
        return int(existing['referrer_id'])
    code=(entered_code or '').strip().upper(); referrer=None
    if code:
        referrer=c.execute("SELECT id,role,own_referral_code FROM users WHERE upper(own_referral_code)=?",(code,)).fetchone()
    if referrer and int(referrer['id'])!=int(referred_user_id):
        referred=c.execute("SELECT role FROM users WHERE id=?",(referred_user_id,)).fetchone()
        kind='TEACHER_RECRUITMENT' if referred and referred['role']=='teacher' and referrer['role']=='teacher' else 'DIRECT'
        c.execute("""INSERT INTO referral_attributions(user_id,referral_source,referral_code,referrer_type,referrer_id,attribution_kind,locked_at)
          VALUES(?,?,?,?,?,?,?) ON CONFLICT(user_id) DO NOTHING""",
          (referred_user_id,source or 'user_referral',referrer['own_referral_code'],'user',referrer['id'],kind,datetime.now().isoformat(timespec='seconds')))
        return int(referrer['id'])
    if (source or code) and not existing:
        c.execute("""INSERT INTO referral_attributions(user_id,referral_source,referral_code,referrer_type,referrer_id,attribution_kind,locked_at)
          VALUES(?,?,?,?,NULL,'UNRESOLVED','') ON CONFLICT(user_id) DO NOTHING""",(referred_user_id,source,code,'unresolved'))
    return None


def create_referral_reward(c,payment_transaction_id):
    """Create the direct reward plus at most one teacher-recruitment override."""
    tx=c.execute("SELECT * FROM payment_transactions WHERE id=?",(payment_transaction_id,)).fetchone()
    if not tx or tx['status']!='successful' or not tx['user_id'] or int(tx['net_amount_minor'] or 0)<=0: return None
    if c.execute("SELECT id FROM referral_rewards WHERE payment_transaction_id=?",(payment_transaction_id,)).fetchone(): return None
    attr=c.execute("SELECT * FROM referral_attributions WHERE user_id=?",(tx['user_id'],)).fetchone()
    if not attr or attr['referrer_type']!='user' or not attr['referrer_id']: return None
    ref=c.execute("SELECT id,role FROM users WHERE id=?",(attr['referrer_id'],)).fetchone()
    if not ref or int(ref['id'])==int(tx['user_id']): return None
    role_group='teacher_direct' if ref['role']=='teacher' else ('partner' if ref['role']=='admin' else 'student')
    programme=c.execute("SELECT * FROM referral_programs WHERE role_group=? AND active=1",(role_group,)).fetchone()
    if not programme: return None
    rate=float(programme['reward_rate'] or 0); amount=int(round(int(tx['net_amount_minor'] or 0)*rate))
    # Teacher-attributed paid conversions are preserved even before the founder sets a commission rate.
    # This prevents commercial attribution from being lost while rates remain deliberately unconfigured.
    if amount<=0 and role_group!='teacher_direct': return None
    reward_type=programme['reward_type']; hold_days=max(0,int(programme['hold_days'] or 0))
    available=(datetime.now()+timedelta(days=hold_days)).isoformat(timespec='seconds')

    override_referrer=None; override_rate=0.0; override_amount=0; override_available=''; override_version=0
    if ref['role']=='teacher':
        upstream=c.execute("""SELECT ra.referrer_id,u.role FROM referral_attributions ra JOIN users u ON u.id=ra.referrer_id
          WHERE ra.user_id=? AND ra.referrer_type='user' AND ra.attribution_kind='TEACHER_RECRUITMENT'""",(ref['id'],)).fetchone()
        if upstream and upstream['role']=='teacher' and int(upstream['referrer_id']) not in {int(ref['id']),int(tx['user_id'])}:
            rule=c.execute("SELECT * FROM referral_programs WHERE role_group='teacher_override' AND active=1").fetchone()
            if rule:
                override_referrer=int(upstream['referrer_id']); override_rate=float(rule['reward_rate'] or 0)
                override_amount=int(round(int(tx['net_amount_minor'] or 0)*override_rate))
                override_available=(datetime.now()+timedelta(days=max(0,int(rule['hold_days'] or 0)))).isoformat(timespec='seconds')
                override_version=int(rule['programme_version'] or 1)
    cur=c.execute("""INSERT INTO referral_rewards(payment_transaction_id,referrer_user_id,referred_user_id,reward_type,currency,
      qualifying_amount_minor,reward_rate,reward_amount_minor,status,available_at,notes,rule_version,
      override_referrer_user_id,override_reward_rate,override_reward_amount_minor,override_rule_version,override_status,override_available_at,override_notes)
      VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",(payment_transaction_id,ref['id'],tx['user_id'],reward_type,tx['currency'],
      tx['net_amount_minor'],rate,amount,'pending' if amount>0 else 'rate_not_configured',available,
      'Created from successful referred payment' if amount>0 else 'Paid teacher referral captured; commission rate not configured',int(programme['programme_version'] or 1),
      override_referrer,override_rate,override_amount,override_version,
      ('pending' if override_amount>0 else ('rate_not_configured' if override_referrer else '')),override_available,
      ('One-level teacher recruitment override' if override_amount>0 else ('Teacher-network attribution captured; override rate not configured' if override_referrer else ''))))
    return cur.lastrowid


def release_due_referral_rewards(c,user_id=None):
    """Release matured wallet rewards. Commission rewards remain payable/approvable in admin."""
    now=datetime.now().isoformat(timespec='seconds')
    params=[now]; where="status='pending' AND available_at<=?"
    if user_id:
        where+=" AND referrer_user_id=?"; params.append(user_id)
    rows=c.execute(f"SELECT * FROM referral_rewards WHERE {where}",params).fetchall()
    for r in rows:
        if r['reward_type']=='wallet_credit':
            c.execute("""INSERT INTO wallet_balances(user_id,currency,balance_minor,updated_at) VALUES(?,?,?,?)
              ON CONFLICT(user_id,currency) DO UPDATE SET balance_minor=balance_minor+excluded.balance_minor,updated_at=excluded.updated_at""",
              (r['referrer_user_id'],r['currency'],r['reward_amount_minor'],datetime.now().isoformat(timespec='seconds')))
            c.execute("""INSERT INTO wallet_transactions(user_id,referral_reward_id,currency,amount_minor,transaction_type,status,description)
              VALUES(?,?,?,?,?,?,?)""",(r['referrer_user_id'],r['id'],r['currency'],r['reward_amount_minor'],
              'referral_reward','posted','Referral reward released'))
            c.execute("UPDATE referral_rewards SET status='available' WHERE id=?",(r['id'],))
    c.commit()
    return len(rows)



def reverse_referral_reward(c,reward_id,reason='Payment/referral reversal',commit=True):
    """Reverse the existing authoritative reward row; optionally join a wider payment transaction."""
    r=c.execute("SELECT * FROM referral_rewards WHERE id=?",(reward_id,)).fetchone()
    if not r or r['status']=='reversed':
        return False
    if r['reward_type']=='wallet_credit' and r['status']=='available':
        bal=c.execute("SELECT balance_minor FROM wallet_balances WHERE user_id=? AND currency=?",
                      (r['referrer_user_id'],r['currency'])).fetchone()
        current=int(bal['balance_minor'] or 0) if bal else 0
        debit=min(current,int(r['reward_amount_minor'] or 0))
        if debit:
            c.execute("""UPDATE wallet_balances SET balance_minor=MAX(0,balance_minor-?),updated_at=?
              WHERE user_id=? AND currency=?""",(debit,datetime.now().isoformat(timespec='seconds'),
              r['referrer_user_id'],r['currency']))
            c.execute("""INSERT INTO wallet_transactions(user_id,referral_reward_id,currency,amount_minor,transaction_type,status,description)
              VALUES(?,?,?,?,?,?,?)""",(r['referrer_user_id'],r['id'],r['currency'],-debit,
              'referral_reversal','posted',reason))
    now=datetime.now().isoformat(timespec='seconds')
    c.execute("""UPDATE referral_rewards SET status='reversed',reversed_at=?,notes=TRIM(COALESCE(notes,'') || ' | ' || ?),
      override_status=CASE WHEN COALESCE(override_reward_amount_minor,0)>0 THEN 'reversed' ELSE override_status END,
      override_reversed_at=CASE WHEN COALESCE(override_reward_amount_minor,0)>0 THEN ? ELSE override_reversed_at END,
      override_notes=CASE WHEN COALESCE(override_reward_amount_minor,0)>0 THEN TRIM(COALESCE(override_notes,'') || ' | ' || ?) ELSE override_notes END WHERE id=?""",
              (now,reason,now,reason,reward_id))
    if commit:
        c.commit()
    return True


def _payment_lifecycle_state(tx):
    """Return ScoreMax's governed payment lifecycle state or INVALID.

    V6.5.10 deliberately does not invent transitions between terminal states. The tuple
    status/refund_status/refund_amount_minor/net_amount_minor must already be internally
    coherent before a governed terminal helper is allowed to mutate anything.
    """
    status=str(tx['status'] or '').strip().lower()
    refund_status=str(tx['refund_status'] or '').strip().lower()
    refund=int(tx['refund_amount_minor'] or 0)
    net=int(tx['net_amount_minor'] or 0)
    empty_refund_status=refund_status in {'','none','not_refunded'}
    if status in {'successful','cleared','paid'} and refund==0 and empty_refund_status:
        return 'CLEARED'
    if status in {'failed','declined'} and refund==0 and empty_refund_status:
        return 'FAILED'
    if status=='refunded' and net>0 and refund==net and refund_status=='refunded':
        return 'REFUNDED'
    if status in {'reversed','voided'} and refund==0 and refund_status in {'reversed','voided'}:
        return 'REVERSED'
    return 'INVALID'


def refund_payment_transaction(c,payment_transaction_id,refund_amount_minor=None,reason='Full payment refund'):
    """Governed full-refund transition with reward reversal in the same database commit.

    Allowed: CLEARED -> REFUNDED and exact repeat of an already-completed identical refund.
    Every other source state fails closed before any payment/reward/source-change mutation.
    """
    tx=c.execute("SELECT * FROM payment_transactions WHERE id=?",(payment_transaction_id,)).fetchone()
    if not tx:
        raise ValueError('payment transaction not found')
    net=int(tx['net_amount_minor'] or 0)
    amount=net if refund_amount_minor is None else int(refund_amount_minor)
    if net<=0 or amount<=0:
        raise ValueError('refund amount must be positive')
    if amount!=net:
        raise ValueError('partial refund reward policy is not governed; refusing transition')

    reward=c.execute("SELECT id,status,override_status FROM referral_rewards WHERE payment_transaction_id=?",(payment_transaction_id,)).fetchone()
    reward_done=(not reward) or (
        str(reward['status'] or '').lower()=='reversed'
        and str(reward['override_status'] or '').lower() in {'','reversed','rate_not_configured'}
    )
    lifecycle=_payment_lifecycle_state(tx)
    if lifecycle=='REFUNDED':
        if reward_done:
            return False
        raise ValueError('refunded payment has incomplete reward reversal; refusing transition')
    if lifecycle!='CLEARED':
        raise ValueError(f'unsupported payment transition {lifecycle}->REFUNDED; refusing transition')

    c.execute('SAVEPOINT scoremax_payment_refund')
    try:
        c.execute("""UPDATE payment_transactions
          SET status='refunded',refund_amount_minor=?,refund_status='refunded',notes=TRIM(COALESCE(notes,'') || ' | ' || ?)
          WHERE id=?""",(amount,reason,payment_transaction_id))
        if reward:
            reverse_referral_reward(c,reward['id'],reason,commit=False)
        c.execute('RELEASE SAVEPOINT scoremax_payment_refund')
        c.commit()
    except Exception:
        try:
            c.execute('ROLLBACK TO SAVEPOINT scoremax_payment_refund')
            c.execute('RELEASE SAVEPOINT scoremax_payment_refund')
        except Exception:
            pass
        c.rollback()
        raise
    return True


def reverse_payment_transaction(c,payment_transaction_id,reason='Payment reversal'):
    """Governed full reversal with a fail-closed terminal state machine.

    Allowed: CLEARED -> REVERSED and exact repeat of an already-completed identical reversal.
    REFUNDED/FAILED/INVALID states are rejected before any mutation.
    """
    tx=c.execute("SELECT * FROM payment_transactions WHERE id=?",(payment_transaction_id,)).fetchone()
    if not tx:
        raise ValueError('payment transaction not found')
    reward=c.execute("SELECT id,status,override_status FROM referral_rewards WHERE payment_transaction_id=?",(payment_transaction_id,)).fetchone()
    reward_done=(not reward) or (
        str(reward['status'] or '').lower()=='reversed'
        and str(reward['override_status'] or '').lower() in {'','reversed','rate_not_configured'}
    )
    lifecycle=_payment_lifecycle_state(tx)
    if lifecycle=='REVERSED':
        if reward_done:
            return False
        raise ValueError('reversed payment has incomplete reward reversal; refusing transition')
    if lifecycle!='CLEARED':
        raise ValueError(f'unsupported payment transition {lifecycle}->REVERSED; refusing transition')

    c.execute('SAVEPOINT scoremax_payment_reversal')
    try:
        c.execute("""UPDATE payment_transactions
          SET status='reversed',refund_amount_minor=0,refund_status='reversed',
              notes=TRIM(COALESCE(notes,'') || ' | ' || ?)
          WHERE id=?""",(reason,payment_transaction_id))
        if reward:
            reverse_referral_reward(c,reward['id'],reason,commit=False)
        c.execute('RELEASE SAVEPOINT scoremax_payment_reversal')
        c.commit()
    except Exception:
        try:
            c.execute('ROLLBACK TO SAVEPOINT scoremax_payment_reversal')
            c.execute('RELEASE SAVEPOINT scoremax_payment_reversal')
        except Exception:
            pass
        c.rollback()
        raise
    return True


def feature_available(c,user_id,entitlement):
    if not COMMERCIAL_GATES_ENABLED:
        return True
    return has_entitlement(c,user_id,entitlement)


def record_payment(c, user_id, plan_id, gross_minor, currency='PKR', status='successful', provider='manual',
                   provider_ref='', payment_method='manual', discount_minor=0, tax_minor=0, promo_code='', notes=''):
    gross_minor=max(0,int(gross_minor or 0)); discount_minor=max(0,int(discount_minor or 0)); tax_minor=max(0,int(tax_minor or 0))
    net=max(0,gross_minor-discount_minor+tax_minor)
    referral=''
    if user_id:
        u=c.execute('SELECT referral_source FROM users WHERE id=?',(user_id,)).fetchone(); referral=(u['referral_source'] if u else '') or ''
    cur=c.execute('''INSERT INTO payment_transactions(user_id,plan_id,provider,provider_transaction_ref,currency,gross_amount_minor,discount_minor,tax_minor,net_amount_minor,status,payment_method,paid_at,referral_source,promo_code,notes)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',(user_id,plan_id,provider,provider_ref,currency,gross_minor,discount_minor,tax_minor,net,status,payment_method,datetime.now().isoformat(timespec='seconds') if status=='successful' else None,referral,promo_code,notes))
    txid=cur.lastrowid
    reward_id=None
    if status=='successful':
        reward_id=create_referral_reward(c,txid)
    # Asynchronous commercial boundary: Growth Engine can consume this later without
    # receiving authority over payments, academic content or mastery state.
    universal_mastery.emit_growth_event(c,'PAYMENT_RECORDED',f'USER:{user_id}' if user_id else '',{
      'payment_transaction_id':txid,'plan_id':plan_id,'status':status,'currency':currency,
      'gross_amount_minor':gross_minor,'net_amount_minor':net,'referral_reward_id':reward_id})
    if reward_id:
        rr=c.execute("SELECT * FROM referral_rewards WHERE id=?",(reward_id,)).fetchone()
        if rr:
            universal_mastery.emit_growth_event(c,'TEACHER_REFERRAL_CONVERSION',f'USER:{user_id}',{
              'payment_transaction_id':txid,'direct_referrer_user_id':rr['referrer_user_id'],
              'override_referrer_user_id':rr['override_referrer_user_id'],
              'qualifying_amount_minor':rr['qualifying_amount_minor'],'currency':rr['currency'],
              'direct_reward_amount_minor':rr['reward_amount_minor'],
              'override_reward_amount_minor':rr['override_reward_amount_minor']})
    return txid



def ranking_band(score, percentile=None):
    """V5.1 challenge category. Elite requires exceptional score plus top-5% standing."""
    s=float(score or 0)
    if s>=92 and percentile is not None and float(percentile)>=95: return 'Elite'
    if s>=90: return 'Expert'
    if s>=82: return 'Distinction'
    if s>=72: return 'Advanced'
    if s>=60: return 'Exam Ready'
    return 'Foundation'


def refresh_challenge_rankings(c, challenge_id):
    """Rank only students who explicitly consented. Exact rank is reserved for qualifying high performers."""
    challenge=c.execute("SELECT * FROM challenges WHERE id=?",(challenge_id,)).fetchone()
    if not challenge or not int(challenge['ranking_enabled'] or 0):
        return []
    rows=c.execute("""SELECT ce.*,u.full_name,u.ranking_display_name
      FROM challenge_entries ce JOIN users u ON u.id=ce.student_id
      WHERE ce.challenge_id=? AND ce.status='completed' AND ce.ranking_consent=1
      ORDER BY ce.score DESC,ce.elapsed_seconds ASC,ce.completed_at ASC""",(challenge_id,)).fetchall()
    n=len(rows)
    snapshot=[]
    for idx,r in enumerate(rows,1):
        # Percentile is intentionally broad; students below the top qualifying bands never need to see a demotivating exact place.
        percentile=100.0 if n<=1 else round(100.0*(n-idx)/(n-1),1)
        band=ranking_band(r['score'],percentile)
        exact=idx if (float(r['score'] or 0)>=float(challenge['exact_rank_min_score'] or 80) and band in ('Distinction','Expert','Elite')) else None
        c.execute("""UPDATE challenge_entries SET ranking_band=?,exact_rank=?,percentile=? WHERE id=?""",
                  (band,exact,percentile,r['id']))
        snapshot.append({'student_id':r['student_id'],'display_name':r['ranking_display_name'] or r['full_name'],
                         'score':float(r['score'] or 0),'band':band,'rank':exact,'percentile':percentile})
    c.execute("INSERT INTO ranking_snapshots(challenge_id,participant_count,snapshot_json) VALUES(?,?,?)",
              (challenge_id,n,json.dumps(snapshot)))
    c.commit()
    return snapshot


def challenge_catalogue(c, student_id):
    now=datetime.now().isoformat(timespec='seconds')
    rows=c.execute("""SELECT ch.*,
      (SELECT COUNT(*) FROM challenge_questions cq WHERE cq.challenge_id=ch.id) mapped_questions,
      ce.status entry_status,ce.score entry_score,ce.ranking_band,ce.exact_rank,ce.percentile,ce.ranking_consent
      FROM challenges ch LEFT JOIN challenge_entries ce ON ce.challenge_id=ch.id AND ce.student_id=?
      WHERE ch.status IN ('published','closed') ORDER BY ch.challenge_month DESC,ch.subject""",(student_id,)).fetchall()
    current=[]; past=[]
    for r in rows:
        if r['opens_at']<=now<=r['closes_at'] and r['status']=='published': current.append(r)
        else: past.append(r)
    return current,past


def exam_centre_data(c, student_id):
    """Return Exam Centre catalogue without confusing official papers with ScoreMax originals."""
    papers=c.execute("""SELECT ep.*,eb.title legacy_blueprint_title,eb.country,eb.qualification,eb.exam_board,
        COALESCE(af.name,eb.programme,'') programme,COALESCE(eb.subject,'') subject,eb.paper_name,
        COALESCE(ep.duration_minutes,ab.duration_minutes,eb.duration_minutes) effective_duration,
        COALESCE(ep.total_marks,ab.total_questions,eb.total_marks) effective_marks,
        ab.powerhouse_blueprint_id authoritative_blueprint_id,ab.blueprint_version authoritative_blueprint_version,
        af.name authoritative_framework,afv.version_name authoritative_framework_version,
        (SELECT COUNT(*) FROM exam_paper_questions pq WHERE pq.paper_id=ep.id) question_count
        FROM exam_papers ep LEFT JOIN exam_blueprints eb ON eb.id=ep.blueprint_id
        LEFT JOIN assessment_blueprints ab ON ab.id=ep.assessment_blueprint_id
        LEFT JOIN assessment_frameworks af ON af.id=ab.framework_id
        LEFT JOIN assessment_framework_versions afv ON afv.id=ab.framework_version_id
        WHERE ep.active=1 ORDER BY COALESCE(af.name,eb.subject),ep.paper_kind,ep.official_year DESC,ep.title""").fetchall()
    grouped={'official_past_paper':[],'scoremax_mock':[]}
    for r in papers:
        grouped.setdefault(r['paper_kind'],[]).append(r)

    history=c.execute("""SELECT a.*,COUNT(aa.id) answered,
        COALESCE(SUM(aa.marks_awarded),0) marks_awarded
        FROM attempts a LEFT JOIN attempt_answers aa ON aa.attempt_id=a.id
        WHERE a.student_id=? AND a.assessment_kind IN ('past_paper','mock')
        GROUP BY a.id ORDER BY a.created_at DESC LIMIT 20""",(student_id,)).fetchall()
    dates=c.execute("""SELECT * FROM student_exam_dates WHERE student_id=? AND active=1 AND exam_date>=?
        ORDER BY exam_date""",(student_id,iso_today())).fetchall()
    return {'papers':grouped,'exam_history':history,'exam_dates':dates}


def paper_question_ids(c, paper_id):
    return [r['question_id'] for r in c.execute(
        "SELECT question_id FROM exam_paper_questions WHERE paper_id=? ORDER BY position",(paper_id,)).fetchall()]


def exam_paper_access_requirement(c, paper_id):
    """Return the highest mastery level present in the governed live paper.

    Exam papers are kept intact: ScoreMax blocks a paper that exceeds the student's
    Access ceiling instead of silently dropping harder questions and changing the paper.
    """
    ids=filter_live_question_ids(c,paper_question_ids(c,paper_id))
    if not ids:
        return '',[]
    marks=','.join('?' for _ in ids)
    rows=c.execute(f"SELECT id,COALESCE(level,'Foundation') level FROM questions WHERE id IN ({marks})",ids).fetchall()
    highest=max((r['level'] or 'Foundation' for r in rows),key=mastery_rank,default='Foundation')
    return highest,ids

def start_exam_paper_session(c, student_id, paper, guided=False):
    required_level,ids=exam_paper_access_requirement(c,paper['id'])
    if not ids:
        return None
    access=get_access_profile(c,student_id)
    if mastery_rank(required_level)>mastery_rank(access['mastery_ceiling']):
        raise PermissionError(f"This paper contains {required_level}-level questions and requires Access up to {required_level}.")
    legacy_blueprint=c.execute("SELECT * FROM exam_blueprints WHERE id=?",(paper['blueprint_id'],)).fetchone() if paper['blueprint_id'] else None
    assessment_bp=blueprint_joined(c,paper['assessment_blueprint_id']) if paper['assessment_blueprint_id'] else None
    if (paper['authenticity_status'] or '')=='AUTHENTIC_BLUEPRINT' and not assessment_bp:
        raise ValueError('This authentic mock has lost its governing blueprint reference and cannot be started.')
    kind='past_paper' if paper['paper_kind']=='official_past_paper' else 'mock'
    mode='practice' if guided else 'mock'
    duration=None if guided else int(paper['duration_minutes'] or (assessment_bp['duration_minutes'] if assessment_bp else 0) or
                                      (legacy_blueprint['duration_minutes'] if legacy_blueprint else 0) or 0) or None
    snapshot=safe_json(paper['blueprint_snapshot_json'],{})
    if assessment_bp and not snapshot:
        snapshot=blueprint_payload_from_record(c,assessment_bp['id']) or {}
    programme=assessment_bp['framework_name'] if assessment_bp else (legacy_blueprint['programme'] if legacy_blueprint else '')
    subject=legacy_blueprint['subject'] if legacy_blueprint else ''
    meta={
      'scope':'paper','programme':programme,'subject':subject,
      'chapters':'','topic':'','subtopic':'','level':'',
      'assessment_kind':kind,'exam_paper_id':paper['id'],'exam_paper_code':paper['code'],
      'exam_title':paper['title'],'guided_mode':bool(guided),
      'official_year':paper['official_year'] or '',
      'assessment_blueprint_id':paper['assessment_blueprint_id'],
      'blueprint_source_id':paper['blueprint_source_id'] or (assessment_bp['powerhouse_blueprint_id'] if assessment_bp else ''),
      'blueprint_version':paper['blueprint_version'] or (assessment_bp['blueprint_version'] if assessment_bp else ''),
      'framework_version':paper['framework_version'] or (assessment_bp['framework_version_name'] if assessment_bp else ''),
      'blueprint_snapshot':snapshot,
      'assembly_policy_id':paper['assembly_policy_id'],
      'assembly_policy_version':paper['assembly_policy_version'] or '',
      'authenticity_status':paper['authenticity_status'] or 'LEGACY_UNPINNED'
    }
    return create_assessment_session(c,student_id,mode,duration,ids,meta)


def exam_days_remaining(exam_date):
    try:
        return max(0,(datetime.fromisoformat(exam_date).date()-datetime.now().date()).days)
    except Exception:
        return None


def save_assessment(c,assessment_id,index=None,answers=None,flagged=None,status=None,confidence=None,response_times=None):
    sets=[]
    params=[]
    if index is not None:
        sets.append('current_index=?'); params.append(index)
    if answers is not None:
        sets.append('saved_answers=?'); params.append(json.dumps(answers))
    if flagged is not None:
        sets.append('flagged_ids=?'); params.append(','.join(str(x) for x in sorted(flagged)))
    if status is not None:
        sets.append('status=?'); params.append(status)
    if confidence is not None:
        sets.append('confidence_json=?'); params.append(json.dumps(confidence))
    if response_times is not None:
        sets.append('response_times_json=?'); params.append(json.dumps(response_times))
    if sets:
        params.append(assessment_id)
        c.execute(f"UPDATE assessment_sessions SET {', '.join(sets)} WHERE id=?",params)
        c.commit()


@app.context_processor
def inject_ui_preferences():
    enabled=True; subject_nav=[]; unread_notifications=0; show_subject_nav=False; active_subject=''
    learner_ui_global=session.get('role') in ('student','qa_student')
    social_links=[]; coach_global=None; show_coach=False; student_nav_section=''; student_secondary_nav=[]
    programme_nav=[]; active_programme=''; active_programme_label=''
    try:
        if session.get('user_id'):
            c=db(); u=c.execute("SELECT help_tips_enabled FROM users WHERE id=?",(session['user_id'],)).fetchone()
            if u is not None: enabled=bool(u['help_tips_enabled'])
            social_links=active_social_links(c)
            if session.get('role')=='student':
                endpoint=(getattr(request,'endpoint','') or '').strip()
                section_endpoints={
                  'dashboard':{'student_dashboard'},
                  'learn':{'subject_browser','subject_detail','chapter_page','weak_areas_page','mastery_page','written_practice_home','written_question_page'},
                  'plan':{'study_plan_page','create_study_plan','start_study_plan_activity','complete_study_plan_activity'},
                  'tests':{'test_setup','test_start','scoremax_test_me','take_test_v4','assessment_review_v4'},
                  'exams':{'exam_centre','exam_structure_page','save_exam_date'},
                  'progress':{'student_analytics_page','student_profile'},
                  'knowledge':{'knowledge_home','knowledge_article'},
                  'more':{'student_pathways','teacher_directory','academic_messages_inbox','sustainability_page','faq_page','pilot_report_issue','access_account','account_settings'}
                }
                student_nav_section=next((name for name,endpoints in section_endpoints.items() if endpoint in endpoints),'dashboard')
                secondary={
                  'dashboard':[('Overview','student_dashboard',''),('Subjects','subject_browser',''),('Daily Spark','student_dashboard','#daily-spark'),('My position','student_analytics_page','')],
                  'learn':[('Subjects','subject_browser',''),('Quick Practice','test_setup',''),('Assess Mastery','mastery_page',''),('Weak Areas','weak_areas_page',''),('Written Practice','written_practice_home','')],
                  'plan':[('Today','study_plan_page','#today'),('Priority Queue','study_plan_page',''),('Full Route','study_plan_page','#route'),('Plan Settings','study_plan_page','#settings')],
                  'tests':[('Test Me','test_setup','#quick'),('Choose Scope','test_setup','#scope'),('Design Test','test_setup','#custom'),('Challenges','challenges_page','')],
                  'exams':[('Exam Centre','exam_centre',''),('Mocks','exam_centre','#mocks'),('Past Papers','exam_centre','#past-papers'),('Exam Structure','exam_structure_page','')],
                  'progress':[('Overview','student_analytics_page',''),('Subjects','student_analytics_page','#subjects'),('Mastery','student_analytics_page','#mastery'),('Weak Areas','weak_areas_page',''),('Activity','student_analytics_page','#activity')],
                  'knowledge':[('Knowledge Hub','knowledge_home',''),('Daily Spark','student_dashboard','#daily-spark'),('Saved Content','knowledge_home','#saved'),('Vocabulary','student_dashboard','#daily-spark')],
                  'more':[('Pathways','student_pathways',''),('Find a Teacher','teacher_directory',''),('Messages','academic_messages_inbox',''),('Access','access_account',''),('Settings','account_settings',''),('Help','faq_page','')]
                }
                student_secondary_nav=secondary.get(student_nav_section,[])
                subject_nav_sections=('dashboard','learn','plan','tests','exams','progress')
                show_subject_nav=endpoint in set().union(*(section_endpoints[name] for name in subject_nav_sections))
                view_args=getattr(request,'view_args',None) or {}
                if endpoint=='subject_detail': active_subject=str(view_args.get('subject') or '').strip()
                elif endpoint=='chapter_page': active_subject=(getattr(request,'args',{}) or {}).get('subject','').strip()
                elif endpoint in {'test_setup','weak_areas_page'}: active_subject=(getattr(request,'args',{}) or {}).get('subject','').strip()
                subject_nav=_subject_map(c,session['user_id']) if '_subject_map' in globals() else []
                active_programme=student_programme(c,session['user_id']) if 'student_programme' in globals() else ''
                active_programme_label=programme_short_label(active_programme) if 'programme_short_label' in globals() else active_programme
                programme_nav=student_programme_options(c,session['user_id']) if 'student_programme_options' in globals() else []
                unread_notifications=c.execute("SELECT COUNT(*) n FROM user_notifications WHERE user_id=? AND COALESCE(read_at,'')=''",(session['user_id'],)).fetchone()['n']
                show_coach=endpoint not in {'student_dashboard','take_test_v4','assessment_review_v4','submit_assessment_v4','login','register'}
                coach_global=scoremax_coach(c,session['user_id']) if 'scoremax_coach' in globals() else None
            c.close()
    except Exception:
        enabled=True; subject_nav=[]; unread_notifications=0; show_subject_nav=False; active_subject=''; social_links=[]; coach_global=None; show_coach=False
    return {
      'help_tips_enabled_global':enabled,'subject_nav_global':subject_nav,'show_subject_nav_global':show_subject_nav,
      'active_subject_global':active_subject,'unread_notifications_global':unread_notifications,
      'social_links_global':social_links,'scoremax_coach_global':coach_global,'show_scoremax_coach_global':show_coach,
      'student_nav_section_global':student_nav_section,'student_secondary_nav_global':student_secondary_nav,
      'programme_nav_global':programme_nav,'active_programme_global':active_programme,
      'active_programme_label_global':active_programme_label,'scoremax_release_version':SCOREMAX_RELEASE_VERSION,
      'learner_ui_global':learner_ui_global
    }

@app.route('/')
def index():
    # Public Daily Spark is deliberately generic: no account, assignment, mastery
    # or learner-history write occurs before login.
    today=iso_today()
    public_word=WORD_LIBRARY[int(hashlib.sha256(('public-spark|'+today).encode()).hexdigest()[:12],16)%len(WORD_LIBRARY)]
    public_spark={
      'date':today,'word':public_word['word'],'pronunciation':public_word.get('pronunciation',''),
      'definition':public_word['definition'],'example':public_word['example'],
      'synonym':public_word.get('synonym',''),'antonym':public_word.get('antonym',''),
    }
    c=db()
    programme_order=['FSc Part 1','Grade 9','Grade 10','FSc Part 2','MDCAT']
    cards=[]
    for programme in programme_order:
        packages=commercial_access.package_rows(c,programme,include_coming=True)
        if not packages and programme=='Grade 9':
            packages=[x for x in commercial_access.package_rows(c,'Matric / Class 9–10',include_coming=True) if x['code']=='grade9_full']
        status='AVAILABLE' if any(x['status']=='ACTIVE' for x in packages) else 'COMING_SOON'
        subjects=[]
        for pkg in packages:
            for subject in pkg.get('subjects',[]):
                if subject not in subjects: subjects.append(subject)
        if programme=='FSc Part 1' and not subjects:
            subjects=['Biology','Chemistry','Physics']
        cards.append({'name':programme,'status':status,'subjects':subjects,
          'package_count':sum(1 for x in packages if x['status']=='ACTIVE'),
          'action':'Start learning' if status=='AVAILABLE' else 'Coming soon'})
    c.close()
    cards.sort(key=lambda x:(x['status']!='AVAILABLE',programme_order.index(x['name'])))
    requested=(request.args.get('programme') or 'FSc Part 1').strip()
    if requested not in {x['value'] for x in STUDENT_PROGRAMME_CHOICES}: requested='FSc Part 1'
    return render_template('index.html',public_spark=public_spark,public_programmes=cards,
      public_programme_choices=STUDENT_PROGRAMME_CHOICES,public_active_programme=requested)


@app.route('/healthz')
def healthz():
    try:
        c=db(); c.execute('SELECT 1').fetchone()
        integration_v1.activate_due_releases(c)
        pending=c.execute("SELECT COUNT(*) n FROM integration_outbox WHERE status IN ('PENDING','RETRY')").fetchone()['n'] if c.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name='integration_outbox'").fetchone() else 0
        c.commit(); c.close()
        return jsonify({'status':'ok','version':'6.2.8.1','release_version':SCOREMAX_RELEASE_VERSION,'build_name':SCOREMAX_BUILD_NAME,
          'universal_mastery_architecture':'0.8','integration_contract_version':'1','integration_schema_version':'1.1.0',
          'integration_outbox_pending':int(pending)}),200
    except Exception:
        app.logger.exception('Health check failed')
        return jsonify({'status':'error'}),503


def _integration_record_transport_diagnostic(contract_name,expected_source,code):
    if not contract_name:
        return
    c=None
    try:
        c=db(); integration_v1.init_schema(c)
        integration_v1.record_transport_diagnostic(c,contract_name,f"{expected_source} -> SCOREMAX",code)
        c.commit()
    except Exception:
        if c:
            try: c.rollback()
            except Exception: pass
        app.logger.exception('Unable to persist redacted integration transport diagnostic')
    finally:
        if c:
            try: c.close()
            except Exception: pass


def _integration_parse_verified(expected_source='POWER_HOUSE',contract_name=''):
    ok,status,code=integration_v1.verify_inbound_http(request,expected_source=expected_source)
    if not ok:
        _integration_record_transport_diagnostic(contract_name,expected_source,code)
        return None,(jsonify({'error':'integration_auth_failed','code':code}),status)
    try:
        raw=request.get_data(cache=True,as_text=True)
        envelope=integration_v1.strict_json_loads(raw) if raw else request.get_json(force=True,silent=False)
        if not isinstance(envelope,dict): raise ValueError('Top-level integration message must be an object.')
    except Exception as exc:
        _integration_record_transport_diagnostic(contract_name,expected_source,'INVALID_JSON')
        return None,(jsonify({'error':'invalid_json','detail':str(exc)}),400)
    if str(request.headers.get('X-Message-Id') or '')!=str(envelope.get('message_id') or ''):
        _integration_record_transport_diagnostic(contract_name,expected_source,'MESSAGE_ID_HEADER_MISMATCH')
        return None,(jsonify({'error':'message_id_header_mismatch'}),400)
    if str(request.headers.get('X-Content-SHA256') or '')!=str(envelope.get('payload_checksum_sha256') or ''):
        _integration_record_transport_diagnostic(contract_name,expected_source,'CONTENT_CHECKSUM_HEADER_MISMATCH')
        return None,(jsonify({'error':'content_checksum_header_mismatch'}),400)
    return envelope,None

@app.route('/api/integration/v1/power-house/content-releases',methods=['POST'])
def integration_power_house_content_release():
    envelope,error=_integration_parse_verified('POWER_HOUSE','PH_SM_APPROVED_CONTENT_V1')
    if error: return error
    c=db()
    try:
        receipt,status=integration_v1.admit_content_envelope(c,envelope,request.headers.get('X-Content-SHA256',''))
        return jsonify(receipt),status
    finally:
        c.close()


@app.route('/api/integration/v1/power-house/assessment-blueprints',methods=['POST'])
def integration_power_house_assessment_blueprint():
    envelope,error=_integration_parse_verified('POWER_HOUSE','PH_SM_ASSESSMENT_BLUEPRINT_V1')
    if error: return error
    c=db()
    try:
        receipt,status=integration_v1.admit_blueprint_envelope(c,envelope,request.headers.get('X-Content-SHA256',''))
        return jsonify(receipt),status
    finally:
        c.close()


@app.route('/api/integration/v1/health')
def integration_health_api():
    # Either configured peer may inspect the redacted operational health surface.
    ok,status,code=integration_v1.verify_inbound_http(request,expected_source='POWER_HOUSE')
    if not ok:
        ok2,status2,code2=integration_v1.verify_inbound_http(request,expected_source='GROWTH_ENGINE')
        if not ok2:
            return jsonify({'error':'integration_auth_failed','code':code2 or code}),max(status,status2)
    c=db()
    try:
        return jsonify(integration_v1.integration_health(c)),200
    finally:
        c.close()

@app.route('/api/internal/universal-mastery/status')
def universal_mastery_status():
    if not require('admin'):
        return jsonify({'error':'admin_required'}),403
    c=db()
    try:
        return jsonify(universal_mastery.runtime_status(c))
    finally:
        c.close()

@app.route('/register/<role>',methods=['GET','POST'])
def register(role):
    if role not in ('student','teacher','parent'):
        return redirect(url_for('index'))
    c=db(); institutions=c.execute('SELECT * FROM institutions WHERE active=1 ORDER BY name').fetchall()
    if request.method=='POST':
        register_key=f"register:{role}:{request.remote_addr or 'unknown'}"
        if not rate_limit(register_key,limit=5,window_seconds=3600):
            c.close(); flash('Too many account-creation attempts from this connection. Please wait and try again.','error'); return redirect(request.url)
        full_name=request.form.get('full_name','').strip()
        email=request.form.get('email','').strip().lower()
        password=request.form.get('password','')
        if not full_name or not email:
            c.close(); flash('Name and email are required.','error'); return redirect(request.url)
        if len(password)<8:
            c.close(); flash('Use at least 8 characters for your password.','error'); return redirect(request.url)
        if c.execute("SELECT 1 FROM users WHERE lower(COALESCE(email,''))=?",(email,)).fetchone():
            c.close(); flash('If this email can be used for a new ScoreMax account, the account has been created. You can now sign in; otherwise use Forgot password.','success'); return redirect(url_for('login'))
        temp_username='pending-'+secrets.token_urlsafe(10).lower()
        cur=c.execute("""INSERT INTO users(system_user_id,role,full_name,father_name,dob,mobile,email,username,password_hash,province,division,district,board,academic_level,subjects,login_provider)
          VALUES('',?,?,?,?,?,?,?,?,?,?,?,?,?,?, 'password')""",(
          role,full_name,request.form.get('father_name','').strip(),request.form.get('dob','').strip(),request.form.get('mobile','').strip(),
          email,temp_username,generate_password_hash(password),request.form.get('province','').strip(),request.form.get('division','').strip(),
          request.form.get('district','').strip(),request.form.get('board','').strip(),request.form.get('academic_level','').strip(),
          ','.join(request.form.getlist('subjects'))))
        uid=cur.lastrowid
        prefix={'student':'STU','teacher':'TCH','parent':'PAR'}[role]
        assigned_id=f'{prefix}-{uid:06d}'
        c.execute("UPDATE users SET system_user_id=?,username=? WHERE id=?",(assigned_id,f'sm-{uid:08d}',uid))
        if role=='student':
            selected_programme=request.form.get('academic_level','').strip()
            c.execute("UPDATE users SET parent_link_code=?,active_programme=? WHERE id=?",('P-'+secrets.token_urlsafe(12).upper(),selected_programme,uid))
        institution_name=request.form.get('institution_name','').strip()
        if institution_name:
            inst=c.execute("SELECT id FROM institutions WHERE lower(name)=lower(?)",(institution_name,)).fetchone()
            if inst: c.execute("UPDATE users SET primary_institution_id=? WHERE id=?",(inst['id'],uid))
        if role=='teacher': ensure_referral_code(c,uid)
        apply_referral_attribution(c,uid,request.form.get('referral_code',''),request.form.get('referral_source',''))
        universal_mastery.emit_growth_event(c,'REGISTERED',f'USER:{uid}',{
          'role':role,'academic_level':request.form.get('academic_level','').strip(),
          'referral_source':request.form.get('referral_source','').strip()})
        c.commit(); c.close()
        flash(f'Account created. Log in with {email} or your ScoreMax ID {assigned_id}.','success')
        return redirect(url_for('login'))
    c.close()
    return render_template('register.html',role=role,institutions=institutions,referral_source=request.args.get('source',''),referral_code=request.args.get('ref',''),selected_programme=request.args.get('programme',''))

def normalize_login_identity(value):
    """Normalize an email, assigned username, or formal ScoreMax User ID for exact lookup."""
    return str(value or '').strip().lower()


def resolve_login_user(c, identity):
    """Return exactly one account matching email, username, or system_user_id.

    Cross-field collisions are rejected rather than selecting whichever row SQLite returns
    first. The caller uses the same neutral failure message for missing and ambiguous values.
    """
    normalized=normalize_login_identity(identity)
    if not normalized or len(normalized)>320:
        return None
    matches=c.execute("""SELECT * FROM users
      WHERE lower(COALESCE(email,''))=?
         OR lower(COALESCE(username,''))=?
         OR lower(COALESCE(system_user_id,''))=?
      ORDER BY id LIMIT 2""",(normalized,normalized,normalized)).fetchall()
    if len(matches)!=1:
        if len(matches)>1:
            app.logger.warning('Ambiguous login identifier rejected because it matched more than one account.')
        return None
    return matches[0]


@app.route('/login',methods=['GET','POST'])
def login():
    if request.method=='POST':
        # Keep the old `email` form key as a compatibility fallback for older local clients.
        identity=request.form.get('identity')
        if identity is None:
            identity=request.form.get('email','')
        key='login:'+str(request.remote_addr or 'unknown')
        if not rate_limit(key,limit=8,window_seconds=300):
            flash('Too many login attempts. Please wait a few minutes and try again.','error')
            return render_template('login.html')
        c=db(); u=resolve_login_user(c,identity)
        if u and (not u['account_status'] or u['account_status']=='active') and check_password_hash(u['password_hash'],request.form.get('password','')):
            _RATE_BUCKETS.pop(key,None)
            c.execute("UPDATE users SET last_login_at=? WHERE id=?",(datetime.now().isoformat(timespec='seconds'),u['id']))
            if u['role']!='qa_student':
                universal_mastery.emit_growth_event(c,'LOGIN',f"USER:{u['id']}",{'role':u['role']})
            c.commit(); c.close()
            session.clear(); session.update(user_id=u['id'],role=u['role'],full_name=u['full_name'],session_version=int(u['session_version'] or 0)); _csrf_token(); return redirect(url_for('reviewer_home') if u['role']=='reviewer' else url_for('dashboard'))
        c.close(); flash('Invalid login details.','error')
    return render_template('login.html')

def send_transactional_email(recipient, subject, body):
    """Minimal provider-neutral SMTP delivery for pilot password resets.

    Configure SCOREMAX_SMTP_HOST/PORT/USER/PASSWORD/FROM. Local development keeps
    the terminal fallback; production logs a delivery error rather than exposing a token.
    """
    host=os.environ.get('SCOREMAX_SMTP_HOST','').strip()
    sender=os.environ.get('SCOREMAX_SMTP_FROM','').strip()
    if not host or not sender or not recipient:
        return False
    port=int(os.environ.get('SCOREMAX_SMTP_PORT','587') or 587)
    username=os.environ.get('SCOREMAX_SMTP_USER','').strip()
    password=os.environ.get('SCOREMAX_SMTP_PASSWORD','')
    use_ssl=os.environ.get('SCOREMAX_SMTP_SSL','0')=='1' or port==465
    msg=EmailMessage(); msg['From']=sender; msg['To']=recipient; msg['Subject']=subject; msg.set_content(body)
    try:
        if use_ssl:
            smtp=smtplib.SMTP_SSL(host,port,timeout=15,context=ssl.create_default_context())
        else:
            smtp=smtplib.SMTP(host,port,timeout=15); smtp.ehlo(); smtp.starttls(context=ssl.create_default_context()); smtp.ehlo()
        with smtp:
            if username: smtp.login(username,password)
            smtp.send_message(msg)
        return True
    except Exception:
        app.logger.exception('Transactional email delivery failed')
        return False

@app.route('/forgot-password',methods=['GET','POST'])
def forgot_password():
    if request.method=='POST':
        email=request.form.get('email','').strip().lower(); key='reset:'+str(request.remote_addr or 'unknown')
        if rate_limit(key,limit=5,window_seconds=900):
            c=db(); u=c.execute("SELECT id,email FROM users WHERE lower(COALESCE(email,''))=? AND account_status='active'",(email,)).fetchone()
            if u:
                raw=secrets.token_urlsafe(32); token_hash=hashlib.sha256(raw.encode()).hexdigest()
                expires=(datetime.now()+timedelta(minutes=30)).isoformat(timespec='seconds')
                c.execute("UPDATE password_reset_tokens SET used_at=? WHERE user_id=? AND COALESCE(used_at,'')=''",(datetime.now().isoformat(timespec='seconds'),u['id']))
                c.execute("INSERT INTO password_reset_tokens(user_id,token_hash,expires_at) VALUES(?,?,?)",(u['id'],token_hash,expires)); c.commit()
                reset_url=url_for('reset_password',token=raw,_external=True)
                sent=send_transactional_email(u['email'],'Reset your ScoreMax password',
                    f"A password reset was requested for your ScoreMax account.\n\nUse this single-use link within 30 minutes:\n{reset_url}\n\nIf you did not request this, you can ignore this email.")
                if not sent and SCOREMAX_ENV!='production':
                    print('[ScoreMax V6.5.10] Local password reset URL:',reset_url)
            c.close()
        flash("If an account exists for this email, we've sent password-reset instructions.",'success')
        return redirect(url_for('forgot_password'))
    return render_template('forgot_password.html')

@app.route('/reset-password/<token>',methods=['GET','POST'])
def reset_password(token):
    token_hash=hashlib.sha256((token or '').encode()).hexdigest(); c=db()
    row=c.execute("SELECT * FROM password_reset_tokens WHERE token_hash=? AND COALESCE(used_at,'')=''",(token_hash,)).fetchone()
    valid=False
    if row:
        try: valid=datetime.fromisoformat(row['expires_at'])>=datetime.now()
        except Exception: valid=False
    if request.method=='POST':
        password=request.form.get('password','')
        if not valid:
            c.close(); flash('This password-reset link is invalid or has expired.','error'); return redirect(url_for('forgot_password'))
        if len(password)<8:
            c.close(); flash('Use at least 8 characters for your password.','error'); return render_template('reset_password.html',valid=True)
        now_reset=datetime.now().isoformat(timespec='seconds')
        c.execute("UPDATE users SET password_hash=?,session_version=COALESCE(session_version,0)+1 WHERE id=?",(generate_password_hash(password),row['user_id']))
        c.execute("UPDATE password_reset_tokens SET used_at=? WHERE user_id=? AND COALESCE(used_at,'')=''",(now_reset,row['user_id']))
        c.commit(); c.close(); session.clear(); flash('Password changed. For your security, all existing ScoreMax sessions have been signed out.','success'); return redirect(url_for('login'))
    c.close(); return render_template('reset_password.html',valid=valid)

@app.route('/logout')
def logout(): session.clear(); return redirect(url_for('index'))

def progress_chart_data(c, student_id, limit=12):
    """Return lightweight chart-ready data without external JS libraries."""
    rows=c.execute("""SELECT id,score,created_at,subject FROM attempts
      WHERE student_id=? AND score IS NOT NULL ORDER BY created_at DESC LIMIT ?""",(student_id,limit)).fetchall()
    rows=list(reversed(rows))
    trend=[]
    for r in rows:
        dt=str(r['created_at'] or '')[:10]
        try:
            label=datetime.strptime(dt,'%Y-%m-%d').strftime('%d %b')
        except Exception:
            label=dt or 'Test'
        trend.append({'label':label,'score':round(float(r['score'] or 0),1),'subject':r['subject'] or ''})

    subjects=c.execute("""SELECT q.subject,COUNT(*) answered,
      ROUND(100.0*SUM(aa.is_correct)/COUNT(*),1) accuracy
      FROM attempt_answers aa JOIN attempts a ON a.id=aa.attempt_id
      JOIN questions q ON q.id=aa.question_db_id
      WHERE a.student_id=? AND COALESCE(q.subject,'')<>''
      GROUP BY q.subject ORDER BY accuracy DESC""",(student_id,)).fetchall()
    subject_data=[{'subject':r['subject'],'accuracy':round(float(r['accuracy'] or 0),1),'answered':int(r['answered'] or 0)} for r in subjects]

    plan=get_active_study_plan(c,student_id)
    weekly=[]
    if plan:
        for item in list(plan.get('weekly',[]))[:10]:
            wk=item.get('week',0)
            total=item.get('total',0)
            done=item.get('done',0)
            pct=item.get('pct',round(100*done/total) if total else 0)
            weekly.append({'week':f'W{wk}','completion':pct,'done':done,'total':total})
    return {'trend':trend,'subjects':subject_data,'plan_weeks':weekly}

@app.route('/dashboard')
def dashboard():
    if not require(): return redirect(url_for('login'))
    return redirect(url_for(session['role']+'_dashboard'))

def student_mastery_hero(c,student_id,programme):
    """Simple learner-facing mastery identity without inventing a cross-subject potential score."""
    records=[dict(r) for r in current_mastery_records(c,student_id) if (r['programme'] or '').casefold()==(programme or '').casefold()]
    overall=[r for r in records if r['scope_type']=='overall']
    subjects=[r for r in records if r['scope_type']=='subject']
    verified=[r for r in overall+subjects if r.get('mastery_level')]
    if not verified:
        return {'established':False,'headline':'Find your starting point','current_level':'Not established',
          'status':'Starting point','next_level':'Foundation','programme':programme_short_label(programme),
          'message':'Complete a short diagnostic so ScoreMax can start building your mastery picture.','progress':0}
    # Prefer an explicit programme/overall record. Otherwise show the strongest *earned* subject level and label it honestly.
    chosen=max(overall or subjects,key=lambda r:mastery_rank(r.get('mastery_level') or 'Foundation'))
    level=chosen.get('mastery_level') or 'Foundation'; status=effective_mastery_status(chosen)
    chapter_ladder=['Foundation','Exam Ready','Advanced','Distinction']
    next_level=chapter_ladder[min(chapter_ladder.index(level)+1,len(chapter_ladder)-1)] if level in chapter_ladder else level
    pct={'Foundation':25,'Exam Ready':50,'Advanced':75,'Distinction':100}.get(level,100)
    label='Programme mastery' if chosen.get('scope_type')=='overall' else f"Strongest verified subject: {chosen.get('subject') or 'your programme'}"
    return {'established':True,'headline':label,'current_level':level,'status':status,'next_level':next_level,
      'programme':programme_short_label(programme),'message':'Keep building with fresh evidence. ScoreMax will tell you when verification is due.','progress':pct}


@app.route('/student')
def student_dashboard():
    if not require('student'): return redirect(url_for('login'))
    c=db()
    user=c.execute('''SELECT u.*,i.name institution_name FROM users u LEFT JOIN institutions i ON i.id=u.primary_institution_id WHERE u.id=?''',(session['user_id'],)).fetchone()
    classes=c.execute('''SELECT classrooms.* FROM classroom_students JOIN classrooms ON classrooms.id=classroom_students.classroom_id WHERE student_id=?''',(session['user_id'],)).fetchall()
    a=student_analytics(c,session['user_id'])
    d=student_dashboard_intelligence(c,session['user_id'])
    if d.get('subjects'):
        total_answers=sum(int(x.get('answered') or 0) for x in d['subjects'])
        weighted=sum(float(x.get('accuracy') or 0)*int(x.get('answered') or 0) for x in d['subjects'])
        overall_accuracy=round(weighted/total_answers,1) if total_answers else 0
        overall_level=scoremax_level_from_evidence(overall_accuracy,total_answers)
        overall_next=next_level_name(overall_level)
        overall_progress=next_level_progress(overall_accuracy,overall_level)
        d['overall_journey']={'accuracy':overall_accuracy,'answered':total_answers,'level':overall_level,
                              'level_index':level_index(overall_level),'next_level':overall_next,'progress':overall_progress}
    else:
        d['overall_journey']=None
    access=get_access_profile(c,session['user_id'])
    mastery_records=current_mastery_records(c,session['user_id'])
    active_programme=student_programme(c,session['user_id'])
    mastery_hero=student_mastery_hero(c,session['user_id'],active_programme)
    demo_question_count=c.execute(f"SELECT COUNT(*) n FROM questions q WHERE q.is_demo=1 AND {live_question_clause('q')}").fetchone()['n']
    formal_mastery_count=c.execute("SELECT COUNT(*) n FROM mastery_records WHERE student_id=?",(session['user_id'],)).fetchone()['n']
    ensure_referral_code(c,session['user_id'])
    release_due_referral_rewards(c,session['user_id'])
    referral_user=c.execute("SELECT own_referral_code FROM users WHERE id=?",(session['user_id'],)).fetchone()
    exam_data=exam_centre_data(c,session['user_id'])
    maybe_weekly_rebalance(c,session['user_id'])
    study_plan=get_active_study_plan(c,session['user_id'])
    focus=daily_focus(c,session['user_id'])
    weekly_report=ensure_weekly_progress_report(c,session['user_id'])
    motivation=motivation_message(c,session['user_id'])
    charts=progress_chart_data(c,session['user_id'])
    target_snapshot=live_target_snapshot(c,session['user_id'])
    recall_due=due_recall_items(c,session['user_id'],limit=4)
    misconceptions=confirmed_misconceptions(c,session['user_id'],limit=4)
    subject_map=_subject_map(c,session['user_id'])
    pathway_data=student_pathway_snapshot(c,session['user_id'])
    coach=scoremax_coach(c,session['user_id'])
    daily_spark=daily_spark_snapshot(c,session['user_id'])
    knowledge_control=c.execute("SELECT state FROM knowledge_feature_controls WHERE feature_code='knowledge_hub'").fetchone()
    latest_knowledge=c.execute("SELECT slug,title,summary,subject,published_at FROM knowledge_articles WHERE status='PUBLISHED' ORDER BY published_at DESC,id DESC LIMIT 2").fetchall() if knowledge_control and knowledge_control['state']=='LIVE' else []
    social_links=active_social_links(c)
    blueprint_snapshot=student_blueprint_snapshot(c,session['user_id'])
    if blueprint_snapshot and blueprint_snapshot.get('projection'):
        latest=c.execute("""SELECT id FROM student_blueprint_projections WHERE student_id=? AND blueprint_id=?
          AND date(created_at)=date('now') ORDER BY id DESC LIMIT 1""",
          (session['user_id'],blueprint_snapshot['projection']['blueprint_id'])).fetchone()
        if not latest:
            persist_blueprint_projection(c,session['user_id'],blueprint_snapshot['projection'])
            c.commit()
    countdowns=[]
    for x in exam_data['exam_dates']:
        countdowns.append({'subject':x['subject'],'exam_name':x['exam_name'],'exam_date':x['exam_date'],'days':exam_days_remaining(x['exam_date'])})
    c.close()
    return render_template('student.html',user=user,classes=classes,attempts=a['attempts'][:6],average=a['avg_score'],n=a['tests_completed'],health=a['health'],by_subtopic=a['by_subtopic'],by_level=a['by_level'],weakest=a['weakest'],strongest=a['strongest'],recommendations=a['recommendations'],access=access,exam_countdowns=countdowns,referral_user=referral_user,study_plan=study_plan,daily_focus=focus,weekly_report=weekly_report,motivation=motivation,charts=charts,target_snapshot=target_snapshot,recall_due=recall_due,misconceptions=misconceptions,subject_map=subject_map,pathway_data=pathway_data,scoremax_coach=coach,daily_spark=daily_spark,latest_knowledge=latest_knowledge,social_links=social_links,mastery_records=mastery_records,demo_question_count=demo_question_count,formal_mastery_count=formal_mastery_count,blueprint_snapshot=blueprint_snapshot,mastery_hero=mastery_hero,active_programme=active_programme,**d)

@app.route('/student/join',methods=['POST'])
def join():
    if not require('student'): return redirect(url_for('login'))
    c=db(); cl=c.execute('SELECT * FROM classrooms WHERE join_code=?',(request.form['join_code'].upper(),)).fetchone()
    if not cl: flash('Classroom code not found.','error')
    else: c.execute('INSERT OR IGNORE INTO classroom_students VALUES(?,?,?)',(cl['id'],session['user_id'],request.form['roll_no'])); c.commit(); flash('Classroom joined.','success')
    c.close(); return redirect(url_for('student_dashboard'))



@app.route('/parent')
def parent_dashboard():
    if not require('parent'): return redirect(url_for('login'))
    c=db()
    parent=c.execute("SELECT * FROM users WHERE id=?",(session['user_id'],)).fetchone()
    links=c.execute("""SELECT psl.*,u.full_name,u.system_user_id,u.academic_level,u.goal_name,u.goal_type,u.target_percentage
      FROM parent_student_links psl JOIN users u ON u.id=psl.student_user_id
      WHERE psl.parent_user_id=? AND psl.status='active' ORDER BY u.full_name""",(session['user_id'],)).fetchall()
    students=[]
    for link in links:
        sid=link['student_user_id']
        analytics=student_analytics(c,sid)
        progress=student_dashboard_intelligence(c,sid)
        plan=get_active_study_plan(c,sid)
        weekly=ensure_weekly_progress_report(c,sid)
        exam_data=exam_centre_data(c,sid)
        countdown=None
        if exam_data['exam_dates']:
            x=exam_data['exam_dates'][0]
            countdown={'name':x['exam_name'] or x['subject'],'days':exam_days_remaining(x['exam_date'])}
        charts=progress_chart_data(c,sid)
        verified=verified_plan_summary(c,sid,weekly['week_start'],weekly['week_end'])
        target_snapshot=live_target_snapshot(c,sid)
        blueprint_snapshot=student_blueprint_snapshot(c,sid)
        students.append({'link':link,'analytics':analytics,'progress':progress,'plan':plan,'weekly':weekly,'verified':verified,
                         'countdown':countdown,'subjects':progress.get('subjects',[])[:6],'charts':charts,'target_snapshot':target_snapshot,
                         'blueprint_snapshot':blueprint_snapshot})
    c.close()
    return render_template('parent_dashboard.html',parent=parent,students=students)

@app.route('/parent/link-student',methods=['POST'])
def parent_link_student():
    if not require('parent'): return redirect(url_for('login'))
    if not rate_limit('parent-link:'+str(session['user_id']),limit=5,window_seconds=600):
        flash('Too many linking attempts. Please wait before trying again.','error'); return redirect(url_for('parent_dashboard'))
    student_email=request.form.get('student_email','').strip().lower()
    link_code=request.form.get('link_code','').strip().upper()
    c=db()
    student=c.execute("SELECT * FROM users WHERE role='student' AND lower(COALESCE(email,''))=?",(student_email,)).fetchone()
    if not student or not secrets.compare_digest((student['parent_link_code'] or '').upper(),link_code):
        c.close(); flash('Student email or parent access code is incorrect.','error'); return redirect(url_for('parent_dashboard'))
    c.execute("""INSERT INTO parent_student_links(parent_user_id,student_user_id,relationship,status)
      VALUES(?,?,?,'active') ON CONFLICT(parent_user_id,student_user_id) DO UPDATE SET status='active',relationship=excluded.relationship""",
      (session['user_id'],student['id'],request.form.get('relationship','Parent')))
    parent=c.execute("SELECT full_name FROM users WHERE id=?",(session['user_id'],)).fetchone()
    c.execute("INSERT INTO user_notifications(user_id,notification_type,title,message) VALUES(?,?,?,?)",
      (student['id'],'privacy','Parent/guardian access linked',f"{parent['full_name'] if parent else 'A parent/guardian'} linked to your ScoreMax progress. You can revoke access from My Target & Results."))
    c.commit(); c.close()
    flash('Student linked. The student has been notified and can revoke access at any time.','success')
    return redirect(url_for('parent_dashboard'))

@app.route('/student/parent-link/<int:link_id>/revoke',methods=['POST'])
def student_revoke_parent_link(link_id):
    if not require('student'): return redirect(url_for('login'))
    c=db(); link=c.execute("SELECT * FROM parent_student_links WHERE id=? AND student_user_id=?",(link_id,session['user_id'])).fetchone()
    if link:
        c.execute("UPDATE parent_student_links SET status='revoked' WHERE id=?",(link_id,))
        c.commit(); flash('Parent/guardian access revoked.','success')
    c.close(); return redirect(url_for('student_profile'))

@app.route('/student/study-plan/rebalance',methods=['POST'])
def rebalance_my_study_plan():
    if not require('student'): return redirect(url_for('login'))
    c=db()
    pid=rebalance_study_plan(c,session['user_id'])
    c.close()
    flash('Your Study Plan has been rebalanced around the time you have left.','success' if pid else 'error')
    return redirect(url_for('study_plan_page'))

@app.route('/student/preferences',methods=['POST'])
def student_preferences():
    if not require('student'): return redirect(url_for('login'))
    c=db()
    c.execute("""UPDATE users SET help_tips_enabled=?,weekly_email_enabled=?,parent_weekly_email_enabled=? WHERE id=?""",
      (1 if request.form.get('help_tips_enabled')=='1' else 0,
       1 if request.form.get('weekly_email_enabled')=='1' else 0,
       1 if request.form.get('parent_weekly_email_enabled')=='1' else 0,
       session['user_id']))
    c.commit(); c.close()
    flash('Preferences updated.','success')
    return redirect(request.referrer or url_for('student_profile'))

@app.route('/student/coach/action',methods=['POST'])
def student_coach_action():
    if not require('student'): return redirect(url_for('login'))
    action=request.form.get('action','OPEN').strip().upper()
    key=request.form.get('nudge_key','').strip()
    if action not in {'OPEN','SNOOZE','DISMISS'}:
        action='OPEN'
    c=db(); candidates=scoremax_coach_candidates(c,session['user_id']); nudge=next((x for x in candidates if x['key']==key),None)
    if not nudge:
        c.close(); return redirect(request.referrer or url_for('student_dashboard'))
    snoozed=''
    if action=='SNOOZE': snoozed=(datetime.now().date()+timedelta(days=3)).isoformat()
    if action=='DISMISS': snoozed=(datetime.now().date()+timedelta(days=30)).isoformat()
    c.execute("INSERT INTO coach_nudge_events(student_id,nudge_key,action,snoozed_until,context_json) VALUES(?,?,?,?,?)",
      (session['user_id'],key,action,snoozed,json.dumps({'reason':nudge['reason'],'endpoint':nudge['endpoint']})))
    c.commit(); c.close()
    if action=='OPEN':
        return redirect(url_for(nudge['endpoint'],**nudge.get('kwargs',{})))
    flash('ScoreMax Coach will pause this suggestion for a while.','success')
    return redirect(request.referrer or url_for('student_dashboard'))


@app.route('/student/daily-spark/action',methods=['POST'])
def student_daily_spark_action():
    if not require('student'): return redirect(url_for('login'))
    try: assignment_id=int(request.form.get('assignment_id') or 0)
    except (TypeError,ValueError): assignment_id=0
    action=(request.form.get('action') or '').strip().upper()
    c=db(); assignment=c.execute("SELECT * FROM daily_spark_assignments WHERE id=? AND student_id=?",(assignment_id,session['user_id'])).fetchone()
    if not assignment:
        c.close(); flash('That Daily Spark is not available for this account.','error'); return redirect(url_for('student_dashboard')+'#daily-spark')
    prior={r['event_type'] for r in c.execute("SELECT event_type FROM daily_spark_events WHERE assignment_id=? AND student_id=?",(assignment_id,session['user_id'])).fetchall()}
    metadata={}
    if action=='ANSWER':
        if assignment['stream']!='ACADEMIC' or assignment['source_type']!='QUESTION':
            c.close(); abort(400,description='Only Academic Sparks accept question answers.')
        if {'ANSWER_CORRECT','ANSWER_INCORRECT'} & prior:
            c.close(); flash('You have already completed today’s Academic Spark.','info'); return redirect(url_for('student_dashboard')+'#daily-spark')
        q=c.execute(f"SELECT * FROM questions q WHERE q.id=? AND {live_question_clause('q')}",(assignment['source_id'],)).fetchone()
        if not q:
            c.close(); flash('This Spark is no longer available because its source content changed.','error'); return redirect(url_for('student_dashboard')+'#daily-spark')
        selected=(request.form.get('selected_answer') or '').strip()
        result=mark_question_response(q,selected)
        if isinstance(result,tuple):
            correct=bool(result[0]); awarded=float(result[1] or 0); misconception=str(result[2] or '')
        elif isinstance(result,dict):
            correct=bool(result.get('is_correct')); awarded=float(result.get('awarded_marks') or 0); misconception=str(result.get('misconception') or '')
        else:
            correct=bool(result); awarded=1.0 if correct else 0.0; misconception=''
        metadata={'selected_answer':selected,'is_correct':correct,'awarded_marks':awarded,'misconception':misconception,'question_id':q['question_id']}
        event='ANSWER_CORRECT' if correct else 'ANSWER_INCORRECT'
        cur=c.execute("INSERT OR IGNORE INTO daily_spark_events(student_id,assignment_id,event_type,metadata_json) VALUES(?,?,?,?)",
                  (session['user_id'],assignment_id,event,json.dumps(metadata,sort_keys=True)))
        if not cur.rowcount:
            c.close(); flash('You have already completed today’s Academic Spark.','info'); return redirect(url_for('student_dashboard')+'#daily-spark')
        c.execute("UPDATE daily_spark_assignments SET status='COMPLETED' WHERE id=?",(assignment_id,))
        flash('Correct — a useful piece of recall.' if correct else 'Not quite. Review the explanation and keep going.','success' if correct else 'info')
    elif action in {'OPEN','REVEAL','SAVE','SNOOZE','DISMISS'}:
        if action not in prior:
            c.execute("INSERT OR IGNORE INTO daily_spark_events(student_id,assignment_id,event_type,metadata_json) VALUES(?,?,?,?)",
                      (session['user_id'],assignment_id,action,'{}'))
        if action=='REVEAL': c.execute("UPDATE daily_spark_assignments SET status='COMPLETED' WHERE id=?",(assignment_id,))
        if action=='SNOOZE': c.execute("UPDATE daily_spark_assignments SET status='SNOOZED',snoozed_until=? WHERE id=?",((datetime.now()+timedelta(hours=2)).isoformat(timespec='seconds'),assignment_id))
        if action=='DISMISS': c.execute("UPDATE daily_spark_assignments SET status='DISMISSED' WHERE id=?",(assignment_id,))
    elif action=='REPORT':
        reason=(request.form.get('report_reason') or 'The student reported this Daily Spark.').strip()[:1000]
        code='FB-'+datetime.now().strftime('%Y%m%d')+'-'+secrets.token_hex(3).upper()
        target='Power House' if assignment['stream']=='ACADEMIC' else 'ScoreMax'
        context={'daily_spark_assignment_id':assignment_id,'stream':assignment['stream'],'source_type':assignment['source_type'],'source_id':assignment['source_id']}
        c.execute("""INSERT INTO pilot_feedback(feedback_code,reporter_user_id,category,severity,description,question_id,routing_target,context_json,page_path)
          VALUES(?,?,?,?,?,?,?,?,?)""",(code,session['user_id'],'ACADEMIC_CONTENT' if assignment['stream']=='ACADEMIC' else 'GENERAL','MEDIUM',reason,
          assignment['source_id'] if assignment['source_type']=='QUESTION' else None,target,json.dumps(context,sort_keys=True),'/student#daily-spark'))
        c.execute("INSERT OR IGNORE INTO daily_spark_events(student_id,assignment_id,event_type,metadata_json) VALUES(?,?,?,?)",
                  (session['user_id'],assignment_id,'REPORT',json.dumps({'feedback_code':code},sort_keys=True)))
        flash('Thank you. The Spark has been reported with its technical context.','success')
    else:
        c.close(); abort(400,description='Unknown Daily Spark action.')
    c.commit(); c.close(); return redirect(url_for('student_dashboard')+'#daily-spark')


@app.route('/institution')
def institution_dashboard():
    if not session.get('user_id'): return redirect(url_for('login'))
    c=db(); access=institution_access(c,session['user_id'])
    if not access:
        c.close(); flash('Institution dashboard access is not assigned to this account.','error'); return redirect(url_for('dashboard'))
    data=institution_dashboard_data(c,access['institution_id'])
    active_structures=[exam_structure_snapshot(c,r['id']) for r in c.execute("SELECT id FROM assessment_blueprints WHERE local_status='ACTIVE' ORDER BY activated_at DESC").fetchall()]
    c.close()
    if not data: return redirect(url_for('dashboard'))
    return render_template('institution_dashboard.html',institution_role=access['institution_role'],active_structures=active_structures,**data)


@app.route('/institution/class/<int:cid>')
def institution_class_view(cid):
    if not session.get('user_id'): return redirect(url_for('login'))
    c=db(); access=institution_access(c,session['user_id'])
    if not access:
        c.close(); return redirect(url_for('dashboard'))
    cl=c.execute("""SELECT cl.*,u.full_name teacher_name FROM classrooms cl LEFT JOIN users u ON u.id=cl.teacher_id
      WHERE cl.id=? AND cl.institution_id=?""",(cid,access['institution_id'])).fetchone()
    if not cl:
        c.close(); flash('Class is outside your institution.','error'); return redirect(url_for('institution_dashboard'))
    students,struggling,inactive,weak_areas=class_analytics(c,cid)
    intervention=class_intervention_analytics(c,cid)
    c.close()
    return render_template('institution_class.html',cl=cl,students=students,struggling=struggling,inactive=inactive,weak_areas=weak_areas,**intervention)


@app.route('/teacher')
def teacher_dashboard():
    if not require('teacher'): return redirect(url_for('login'))
    c=db(); ensure_referral_code(c,session['user_id']); c.commit(); user=c.execute('''SELECT u.*,i.name institution_name FROM users u LEFT JOIN institutions i ON i.id=u.primary_institution_id WHERE u.id=?''',(session['user_id'],)).fetchone()
    classes=c.execute('''SELECT classrooms.*,(SELECT COUNT(*) FROM classroom_students cs WHERE cs.classroom_id=classrooms.id) student_count FROM classrooms WHERE teacher_id=?''',(session['user_id'],)).fetchall()
    intel=teacher_dashboard_intelligence(c,session['user_id'])
    active_structures=[exam_structure_snapshot(c,r['id']) for r in c.execute("SELECT id FROM assessment_blueprints WHERE local_status='ACTIVE' ORDER BY activated_at DESC").fetchall()]
    c.close()
    return render_template('teacher.html',user=user,classes=classes,active_structures=active_structures,**intel)

@app.route('/teacher/create',methods=['POST'])
def create_class():
    if not require('teacher'): return redirect(url_for('login'))
    c=db(); t=c.execute('SELECT primary_institution_id FROM users WHERE id=?',(session['user_id'],)).fetchone(); join_code=code()
    c.execute('INSERT INTO classrooms(teacher_id,institution_id,name,level,subject,join_code) VALUES(?,?,?,?,?,?)',(session['user_id'],t['primary_institution_id'],request.form['name'],request.form['level'],request.form['subject'],join_code)); c.commit(); c.close(); flash(f'Classroom created. Code: {join_code}','success'); return redirect(url_for('teacher_dashboard'))

@app.route('/teacher/class/<int:cid>')
def classroom(cid):
    if not require('teacher'): return redirect(url_for('login'))
    c=db(); cl=c.execute('SELECT * FROM classrooms WHERE id=? AND teacher_id=?',(cid,session['user_id'])).fetchone()
    if not cl: c.close(); flash('Classroom not found.','error'); return redirect(url_for('teacher_dashboard'))
    students,struggling,inactive,weak_areas=class_analytics(c,cid); intervention=class_intervention_analytics(c,cid)
    class_bp=active_assessment_blueprint(c,cl['level'] or '')
    blueprint_summary=None
    if class_bp:
        projections=[]
        for st in students[:100]:
            proj=blueprint_projection_snapshot(c,st['id'],class_bp['id'])
            if proj: projections.append({'student_id':st['id'],'full_name':st['full_name'],'projection':proj})
        blueprint_summary={'blueprint':class_bp,'sections':blueprint_sections(c,class_bp['id']),
                           'bank':blueprint_bank_sufficiency(c,class_bp['id']),'student_projections':projections}
    c.close()
    return render_template('classroom.html',cl=cl,students=students,struggling=struggling,inactive=inactive,weak_areas=weak_areas,blueprint_summary=blueprint_summary,**intervention)

def assignment_question_pool(c, subject, focus_type='', focus_name='', level=''):
    """Return the full governed/markable pool for teacher assignment validation and start."""
    clauses=[live_question_clause('q'),'q.subject=?']; params=[subject]
    if focus_name:
        if focus_type=='learning_outcome':
            clauses.append('q.learning_outcome=?'); params.append(focus_name)
        elif focus_type=='diagnostic':
            clauses.append("COALESCE(NULLIF(q.learning_outcome,''),NULLIF(q.subtopic,''),q.topic)=?"); params.append(focus_name)
        else:
            clauses.append('(q.subtopic=? OR q.topic=?)'); params.extend([focus_name,focus_name])
    if level:
        clauses.append('q.level=?'); params.append(level)
    rows=c.execute(f"SELECT q.* FROM questions q WHERE {' AND '.join(clauses)}",params).fetchall()
    return [q for q in rows if canonical_question_type(q) in LIVE_MARKABLE_TYPES]

@app.route('/teacher/class/<int:cid>/assign',methods=['POST'])
def teacher_assign(cid):
    if not require('teacher'): return redirect(url_for('login'))
    c=db(); cl=c.execute('SELECT * FROM classrooms WHERE id=? AND teacher_id=?',(cid,session['user_id'])).fetchone()
    if not cl: c.close(); flash('Classroom not found.','error'); return redirect(url_for('teacher_dashboard'))
    title=request.form.get('title','').strip() or 'ScoreMax assessment'
    subject=request.form.get('subject','').strip() or cl['subject']
    focus_type=request.form.get('focus_type','diagnostic').strip()
    focus_name=request.form.get('focus_name','').strip()
    level=request.form.get('level','').strip()
    mode=request.form.get('assessment_mode','exam').strip() or 'exam'
    due_at=request.form.get('due_at','').strip()
    count=max(3,min(30,int(request.form.get('question_count','8') or 8)))
    duration_raw=request.form.get('duration_minutes','').strip(); duration=int(duration_raw) if duration_raw.isdigit() else None
    baseline_raw=request.form.get('baseline_accuracy','').strip(); baseline=float(baseline_raw) if baseline_raw else None
    members={r['student_id'] for r in c.execute('SELECT student_id FROM classroom_students WHERE classroom_id=?',(cid,)).fetchall()}
    selected=[int(x) for x in request.form.getlist('student_ids') if str(x).isdigit() and int(x) in members]
    if not selected: selected=list(members)
    if not selected:
        c.close(); flash('Add students to the class before creating an assignment.','error'); return redirect(url_for('classroom',cid=cid))

    pool=assignment_question_pool(c,subject,focus_type,focus_name,level)
    if len(pool)<count:
        c.close(); flash(f'Only {len(pool)} approved, governed questions are available for this assignment. Reduce the count or add/review more content.','error')
        return redirect(url_for('classroom',cid=cid))
    insufficient=[]
    for sid in selected:
        access=get_access_profile(c,sid)
        eligible=sum(1 for q in pool if mastery_rank(q['level'] or 'Foundation')<=mastery_rank(access['mastery_ceiling']))
        if eligible<count:
            u=c.execute('SELECT full_name FROM users WHERE id=?',(sid,)).fetchone()
            insufficient.append((u['full_name'] if u else f'Student {sid}',eligible,access['mastery_ceiling']))
    if insufficient:
        name,eligible,ceiling=insufficient[0]
        c.close(); flash(f'Assignment not sent: {name} can access only {eligible} eligible questions at the current {ceiling} ceiling (requested {count}).','error')
        return redirect(url_for('classroom',cid=cid))
    cur=c.execute("""INSERT INTO assignments(teacher_id,classroom_id,title,subject,level,assessment_mode,due_at,status,assignment_kind,focus_type,focus_name,question_count,duration_minutes,created_from_diagnostic,baseline_accuracy)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",(session['user_id'],cid,title,subject,level,mode,due_at,'active','recovery' if focus_name else 'standard',focus_type,focus_name,count,duration,1 if focus_name else 0,baseline))
    aid=cur.lastrowid
    c.executemany('INSERT OR IGNORE INTO assignment_students(assignment_id,student_id,status) VALUES(?,?,?)',[(aid,sid,'assigned') for sid in selected])
    c.commit(); c.close(); flash(f'Assignment sent to {len(selected)} student(s).','success'); return redirect(url_for('classroom',cid=cid))


@app.route('/teacher/class/<int:cid>/assign-study-plan',methods=['POST'])
def teacher_assign_study_plan(cid):
    if not require('teacher'): return redirect(url_for('login'))
    pathway=request.form.get('pathway','Core')
    if pathway not in STUDY_PLAN_PATHWAYS: pathway='Core'
    target_exam=request.form.get('target_exam','').strip()
    target_date=request.form.get('target_date','').strip()
    raw=request.form.get('target_percentage','').strip()
    target=float(raw) if raw else None
    selected=[int(x) for x in request.form.getlist('student_ids') if str(x).isdigit()]
    c=db()
    cl=c.execute("SELECT * FROM classrooms WHERE id=? AND teacher_id=?",(cid,session['user_id'])).fetchone()
    if not cl:
        c.close(); flash('Classroom not found.','error'); return redirect(url_for('teacher_dashboard'))
    allowed={r['student_id'] for r in c.execute("SELECT student_id FROM classroom_students WHERE classroom_id=?",(cid,)).fetchall()}
    if not selected: selected=list(allowed)
    selected=[sid for sid in selected if sid in allowed]
    try:
        for sid in selected:
            pid=generate_scoremax_plan(c,sid,pathway,target_exam,target_date,target)
            c.execute("UPDATE study_plans SET source='institution',assigned_by_user_id=? WHERE id=?",(session['user_id'],pid))
            c.execute("UPDATE users SET study_plan_source='institution' WHERE id=?",(sid,))
    except ValueError as exc:
        c.rollback(); c.close(); flash(str(exc),'error'); return redirect(url_for('classroom',cid=cid))
    c.commit(); c.close()
    flash(f'{pathway} Study Plan assigned to {len(selected)} student(s).','success')
    return redirect(url_for('classroom',cid=cid))

@app.route('/student/assignment/<int:assignment_id>/start',methods=['POST'])
def start_assignment(assignment_id):
    if not require('student'): return redirect(url_for('login'))
    c=db(); row=c.execute("""SELECT a.*,ast.status student_status FROM assignments a JOIN assignment_students ast ON ast.assignment_id=a.id
      WHERE a.id=? AND ast.student_id=? AND a.status='active'""",(assignment_id,session['user_id'])).fetchone()
    if not row: c.close(); flash('Assignment not available.','error'); return redirect(url_for('student_dashboard'))
    qrows=assignment_question_pool(c,row['subject'],row['focus_type'],row['focus_name'],row['level'])
    random.shuffle(qrows)
    access=get_access_profile(c,session['user_id'])
    qids=[q['id'] for q in qrows if mastery_rank(q['level'] or 'Foundation')<=mastery_rank(access['mastery_ceiling'])][:int(row['question_count'] or 8)]
    if len(qids)<3:
        c.close(); flash('Not enough approved questions are available for this assignment yet.','error'); return redirect(url_for('student_dashboard'))
    meta={'scope':'teacher_assignment','subject':row['subject'],'level':row['level'] or '', 'assignment_id':assignment_id,'title':row['title'],'focus_type':row['focus_type'],'focus_name':row['focus_name']}
    sid=create_assessment_session(c,session['user_id'],row['assessment_mode'] or 'exam',row['duration_minutes'],qids,meta)
    c.execute("UPDATE assignment_students SET status='in_progress' WHERE assignment_id=? AND student_id=?",(assignment_id,session['user_id'])); c.commit(); c.close()
    return redirect(url_for('take_test_v4',assessment_id=sid))

@app.route('/api/curriculum')
def curriculum_api():
    if not require():
        return jsonify([]), 401
    field=request.args.get('field'); allowed={'programme','subject','chapter','topic','subtopic'}
    if field not in allowed: return jsonify([])
    clauses=['active=1']; params=[]
    for k in ['programme','subject','chapter','topic']:
        v=request.args.get(k,'')
        if v: clauses.append(f'{k}=?'); params.append(v)
    c=db(); rows=c.execute(f"SELECT DISTINCT {field} v FROM curriculum WHERE {' AND '.join(clauses)} AND COALESCE({field},'')<>'' ORDER BY {field}",params).fetchall(); c.close(); return jsonify([r['v'] for r in rows])



@app.route('/challenges')
def challenges_page():
    if not require('student'): return redirect(url_for('login'))
    c=db(); current_raw,past_raw=challenge_catalogue(c,session['user_id'])
    user=c.execute("SELECT ranking_opt_in,ranking_display_name,academic_level FROM users WHERE id=?",(session['user_id'],)).fetchone()
    access=get_access_profile(c,session['user_id']); records=current_mastery_records(c,session['user_id'])
    def enrich(rows):
        out=[]
        for row in rows:
            d=dict(row); scope=(d.get('scope_type') or d.get('challenge_type') or 'Subject').lower()
            if scope not in ('chapter','subject','overall'): scope='chapter' if d.get('chapter') else 'subject'
            programme=d.get('programme') or d.get('qualification') or (user['academic_level'] if user else '') or ''
            required=d.get('required_mastery_level') or ('Distinction' if scope=='chapter' else 'Expert' if scope=='subject' else 'Elite')
            ready,reason=challenge_readiness(c,session['user_id'],scope,programme,d.get('subject',''),d.get('chapter',''),required)
            d.update(ready=ready,readiness_reason=reason,required_mastery=required,scope_resolved=scope); out.append(d)
        return out
    current=enrich(current_raw); past=enrich(past_raw); c.close()
    return render_template('challenges.html',current=current,past=past,user=user,access=access,mastery_records=records)

@app.route('/challenges/create',methods=['POST'])
def create_student_challenge():
    if not require('student'): return redirect(url_for('login'))
    subject=request.form.get('subject','').strip(); chapter=request.form.get('chapter','').strip(); topic=request.form.get('topic','').strip()
    challenge_type=request.form.get('challenge_type','chapter').lower(); programme=request.form.get('programme','').strip()
    if challenge_type not in ('chapter','subject','covered','half','full'): challenge_type='chapter'
    count=max(10,min(40,int(request.form.get('question_count') or 20))); c=db()
    if not programme:
        u=c.execute("SELECT academic_level FROM users WHERE id=?",(session['user_id'],)).fetchone(); programme=(u['academic_level'] or '').strip() if u else ''
    if challenge_type in ('chapter','covered'):
        scope_type='chapter'; required='Distinction'
        if not subject or not chapter: c.close(); flash('Choose a subject and chapter first.','error'); return redirect(url_for('challenges_page'))
    elif challenge_type in ('subject','half'):
        scope_type='subject'; required='Expert'
        if not subject: c.close(); flash('Choose a subject first.','error'); return redirect(url_for('challenges_page'))
    else:
        scope_type='overall'; required='Elite'
    ready,reason=challenge_readiness(c,session['user_id'],scope_type,programme,subject,chapter,required)
    if not ready:
        c.close(); flash(reason,'error'); return redirect(url_for('challenges_page'))
    clauses=[live_question_clause('q')]; params=[]
    if programme: clauses.append('programme=?'); params.append(programme)
    if subject: clauses.append('subject=?'); params.append(subject)
    if challenge_type in ('chapter','covered') and chapter: clauses.append('chapter=?'); params.append(chapter)
    raw=c.execute(f"SELECT q.* FROM questions q WHERE {' AND '.join(clauses)} ORDER BY RANDOM()",params).fetchall()
    access=get_access_profile(c,session['user_id']); pool=[q for q in raw if canonical_question_type(q) in LIVE_MARKABLE_TYPES and mastery_rank(q['level'] or 'Foundation')<=mastery_rank(access['mastery_ceiling'])]
    # Prefer independent families for challenge integrity.
    qs=[]; fams=set()
    for q in pool:
        fam=q['family_id'] or f"Q-{q['id']}"
        if fam in fams: continue
        fams.add(fam); qs.append(q)
        if len(qs)>=count: break
    if len(qs)<5:
        c.close(); flash('Not enough independent approved question families exist for that challenge yet.','error'); return redirect(url_for('challenges_page'))
    now=datetime.now(); closes=now+timedelta(days=7); code=f"USR-{session['user_id']}-{int(now.timestamp())}"
    # Closed-pilot safety: public student-created challenge titles are generated by ScoreMax,
    # avoiding unmoderated free-text publication to other learners.
    title=f"{chapter or subject or programme or 'Full Exam'} {challenge_type.title()} Challenge"; duration=max(5,round(len(qs)*0.75))
    cur=c.execute("""INSERT INTO challenges(code,title,country,subject,qualification,exam_board,challenge_month,description,
      duration_minutes,question_count,opens_at,closes_at,premium_required,ranking_enabled,exact_rank_min_score,max_attempts,status,
      challenge_type,chapter,topic,created_by_user_id,official,scope_type,programme,required_mastery_level,scope_label)
      VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,0,?,?,?,?)""",
      (code,title,'Pakistan',subject,programme,'','Community Challenge','Created by a ScoreMax student after verified mastery.',duration,len(qs),
       now.isoformat(timespec='seconds'),closes.isoformat(timespec='seconds'),0,1,80,1,'published',challenge_type,chapter,topic,session['user_id'],
       scope_type,programme,required,challenge_type.title()))
    cid=cur.lastrowid
    for pos,q in enumerate(qs,1): c.execute("INSERT INTO challenge_questions(challenge_id,question_id,position) VALUES(?,?,?)",(cid,q['id'],pos))
    c.commit(); c.close(); flash('Your mastery-gated challenge is live for 7 days.','success'); return redirect(url_for('challenge_detail',challenge_id=cid))

@app.route('/ranking/preferences',methods=['POST'])
def ranking_preferences():
    if not require('student'): return redirect(url_for('login'))
    opt=1 if request.form.get('ranking_opt_in')=='1' else 0
    display=request.form.get('ranking_display_name','').strip()[:40]
    c=db(); c.execute("UPDATE users SET ranking_opt_in=?,ranking_display_name=? WHERE id=?",(opt,display,session['user_id']))
    c.commit(); c.close()
    flash('Ranking preference updated. Participation is voluntary.','success')
    return redirect(request.referrer or url_for('challenges_page'))


@app.route('/challenges/<int:challenge_id>')
def challenge_detail(challenge_id):
    if not require('student'): return redirect(url_for('login'))
    c=db()
    ch=c.execute("""SELECT ch.*,(SELECT COUNT(*) FROM challenge_questions cq WHERE cq.challenge_id=ch.id) mapped_questions
      FROM challenges ch WHERE ch.id=? AND ch.status IN ('published','closed')""",(challenge_id,)).fetchone()
    entry=c.execute("SELECT * FROM challenge_entries WHERE challenge_id=? AND student_id=?",(challenge_id,session['user_id'])).fetchone()
    user=c.execute("SELECT ranking_opt_in,ranking_display_name FROM users WHERE id=?",(session['user_id'],)).fetchone()
    access=get_access_profile(c,session['user_id'])
    leaderboard=[]
    if ch and int(ch['ranking_enabled'] or 0):
        leaderboard=c.execute("""SELECT ce.score,ce.ranking_band,ce.exact_rank,ce.percentile,
          COALESCE(NULLIF(u.ranking_display_name,''),u.full_name) display_name
          FROM challenge_entries ce JOIN users u ON u.id=ce.student_id
          WHERE ce.challenge_id=? AND ce.status='completed' AND ce.ranking_consent=1 AND ce.exact_rank IS NOT NULL
          ORDER BY ce.exact_rank LIMIT 20""",(challenge_id,)).fetchall()
    c.close()
    if not ch: return redirect(url_for('challenges_page'))
    return render_template('challenge_detail.html',challenge=ch,entry=entry,user=user,access=access,leaderboard=leaderboard)


@app.route('/challenges/<int:challenge_id>/start',methods=['POST'])
def start_challenge(challenge_id):
    if not require('student'): return redirect(url_for('login'))
    c=db()
    ch=c.execute("SELECT * FROM challenges WHERE id=? AND status='published'",(challenge_id,)).fetchone()
    if not ch:
        c.close(); flash('Challenge is not available.','error'); return redirect(url_for('challenges_page'))
    now=datetime.now()
    try:
        opens=datetime.fromisoformat(ch['opens_at']); closes=datetime.fromisoformat(ch['closes_at'])
    except Exception:
        c.close(); flash('Challenge schedule is invalid.','error'); return redirect(url_for('challenge_detail',challenge_id=challenge_id))
    if not (opens<=now<=closes):
        c.close(); flash('This challenge is not currently open.','error'); return redirect(url_for('challenge_detail',challenge_id=challenge_id))
    access=get_access_profile(c,session['user_id'])
    scope=(ch['scope_type'] or ('chapter' if ch['chapter'] else 'subject')).lower()
    if scope not in ('chapter','subject','overall'): scope='subject'
    programme=ch['programme'] or ch['qualification'] or ''
    required=ch['required_mastery_level'] or ('Distinction' if scope=='chapter' else 'Expert' if scope=='subject' else 'Elite')
    ready,reason=challenge_readiness(c,session['user_id'],scope,programme,ch['subject'] or '',ch['chapter'] or '',required)
    if not ready:
        c.close(); flash(reason,'error'); return redirect(url_for('challenges_page'))
    prior=c.execute("SELECT * FROM challenge_entries WHERE challenge_id=? AND student_id=?",(challenge_id,session['user_id'])).fetchone()
    if prior and prior['status'] in ('started','completed'):
        c.close(); flash('This challenge allows one verified attempt.','error'); return redirect(url_for('challenge_detail',challenge_id=challenge_id))
    ids=[r['question_id'] for r in c.execute("SELECT question_id FROM challenge_questions WHERE challenge_id=? ORDER BY position",(challenge_id,)).fetchall()]
    if len(ids)<1:
        c.close(); flash('Challenge questions are not ready yet.','error'); return redirect(url_for('challenge_detail',challenge_id=challenge_id))
    user=c.execute("SELECT ranking_opt_in FROM users WHERE id=?",(session['user_id'],)).fetchone()
    consent=1 if request.form.get('ranking_consent')=='1' and int(user['ranking_opt_in'] or 0) else 0
    meta={'scope':'challenge','subject':ch['subject'],'programme':ch['qualification'],'chapters':'','topic':'','subtopic':'','level':'',
          'assessment_kind':'challenge','challenge_id':challenge_id,'challenge_title':ch['title'],'ranking_consent':consent}
    aid=create_assessment_session(c,session['user_id'],'mock',int(ch['duration_minutes'] or 30),ids,meta)
    c.execute("""INSERT OR REPLACE INTO challenge_entries(challenge_id,student_id,assessment_session_id,status,started_at,ranking_consent)
      VALUES(?,?,?,?,?,?)""",(challenge_id,session['user_id'],aid,'started',datetime.now().isoformat(timespec='seconds'),consent))
    c.commit(); c.close()
    return redirect(url_for('take_test_v4',assessment_id=aid))



# ---------------------------------------------------------------------------
# V5.5 Final — Stakeholder-facing exam structure and blueprint compliance
# ---------------------------------------------------------------------------

def assessment_type_label(authenticity_status='', paper_kind='', scope=''):
    """Return an honest learner-facing relationship to the official blueprint."""
    status=(authenticity_status or '').upper()
    kind=(paper_kind or '').lower()
    scope=(scope or '').lower()
    if status=='AUTHENTIC_BLUEPRINT' or kind=='authentic_mock':
        return {'code':'AUTHENTIC_FULL_MOCK','label':'Authentic Full Mock','tone':'good',
                'description':'Follows the pinned official assessment blueprint exactly.'}
    if status=='PROPORTIONAL_BLUEPRINT_PRACTICE':
        return {'code':'BLUEPRINT_PROPORTIONAL','label':'Blueprint-Proportional Practice','tone':'mid',
                'description':'Uses a shorter proportional form and is not a full authentic mock.'}
    if status=='DIAGNOSTIC_BLUEPRINT_AWARE' or scope in ('diagnostic','recovery'):
        return {'code':'DIAGNOSTIC','label':'Diagnostic Assessment','tone':'mid',
                'description':'Intentionally prioritises learner needs and may differ from official weighting.'}
    if scope in ('subject','chapter','topic'):
        labels={'subject':'Subject Mock / Test','chapter':'Chapter Test','topic':'Topic Test'}
        return {'code':scope.upper(),'label':labels[scope],'tone':'neutral',
                'description':'A local-scope assessment; it does not represent the complete official examination.'}
    if kind=='official_past_paper':
        return {'code':'OFFICIAL_PAST_PAPER','label':'Official Past Paper','tone':'good',
                'description':'A historical paper governed by its original examination structure and rights status.'}
    return {'code':'LOCAL_PRACTICE','label':'Practice Assessment','tone':'neutral',
            'description':'A ScoreMax practice form; no claim of exact full-blueprint compliance is made.'}


def exam_structure_snapshot(c, blueprint_id):
    bp=blueprint_joined(c,blueprint_id)
    if not bp: return None
    sections=[]
    for row in blueprint_sections(c,blueprint_id):
        sections.append({'subject':row['subject'],'question_count':int(row['question_count'] or 0),
                         'weight_percent':float(row['weight_percent'] or 0),
                         'duration_minutes':row['duration_minutes'],
                         'rules':safe_json(row['rules_json'],{}),
                         'difficulty_distribution':safe_json(row['difficulty_distribution_json'],{})})
    return {'blueprint':bp,'sections':sections,
            'difficulty_distribution':safe_json(bp['difficulty_distribution_json'],{}),
            'rules':safe_json((blueprint_payload_from_record(c,blueprint_id) or {}).get('rules'),{}),
            'payload':blueprint_payload_from_record(c,blueprint_id) or {}}


def paper_blueprint_compliance(c, paper_id):
    paper=c.execute("""SELECT ep.*,ab.total_questions,ab.duration_minutes blueprint_duration,
      ab.powerhouse_blueprint_id,ab.blueprint_version,af.name framework_name,afv.version_name framework_version_name
      FROM exam_papers ep LEFT JOIN assessment_blueprints ab ON ab.id=ep.assessment_blueprint_id
      LEFT JOIN assessment_frameworks af ON af.id=ab.framework_id
      LEFT JOIN assessment_framework_versions afv ON afv.id=ab.framework_version_id WHERE ep.id=?""",(paper_id,)).fetchone()
    if not paper: return None
    relation=assessment_type_label(paper['authenticity_status'],paper['paper_kind'])
    actual_rows=c.execute("""SELECT q.subject,COUNT(*) question_count,
      SUM(CASE WHEN q.status='Approved' AND q.review_status='Approved' AND q.active=1 THEN 1 ELSE 0 END) approved_count,
      COUNT(DISTINCT q.family_key) family_count,
      SUM(CASE WHEN COALESCE(q.rights_status,'')<>'' AND q.scoremax_ready=1 THEN 1 ELSE 0 END) ready_rights_count
      FROM exam_paper_questions pq JOIN questions q ON q.id=pq.question_id
      WHERE pq.paper_id=? GROUP BY q.subject ORDER BY q.subject""",(paper_id,)).fetchall()
    actual={r['subject']:{'count':int(r['question_count'] or 0),'approved':int(r['approved_count'] or 0),
                          'families':int(r['family_count'] or 0),'ready_rights':int(r['ready_rights_count'] or 0)} for r in actual_rows}
    criteria=[]; blocking=[]; warnings=[]
    total_actual=sum(x['count'] for x in actual.values())
    if paper['assessment_blueprint_id']:
        for sec in blueprint_sections(c,paper['assessment_blueprint_id']):
            required=int(sec['question_count'] or 0); got=actual.get(sec['subject'],{}).get('count',0)
            ok=(got==required)
            criteria.append({'criterion':sec['subject'],'required':required,'actual':got,'status':'PASS' if ok else 'FAIL'})
            if not ok: blocking.append(f"{sec['subject']} requires {required} questions; this paper contains {got}.")
        total_required=int(paper['total_questions'] or 0)
        ok_total=(total_actual==total_required)
        criteria.insert(0,{'criterion':'Total questions','required':total_required,'actual':total_actual,'status':'PASS' if ok_total else 'FAIL'})
        if not ok_total: blocking.append(f'Total requires {total_required}; this paper contains {total_actual}.')
    else:
        criteria.append({'criterion':'Blueprint version pinned','required':'Required for authentic claim','actual':'Not pinned','status':'WARNING'})
        warnings.append('This is a legacy or local paper and cannot claim current authentic-blueprint compliance.')
    approved=sum(x['approved'] for x in actual.values())
    ready=sum(x['ready_rights'] for x in actual.values())
    criteria.append({'criterion':'Approved active questions','required':total_actual,'actual':approved,'status':'PASS' if approved==total_actual else 'FAIL'})
    criteria.append({'criterion':'Rights and ScoreMax readiness','required':total_actual,'actual':ready,'status':'PASS' if ready==total_actual else 'FAIL'})
    if approved!=total_actual: blocking.append('One or more questions are not approved and active.')
    if ready!=total_actual: blocking.append('One or more questions do not satisfy rights/readiness requirements.')
    policy_ok=bool(paper['assembly_policy_version'])
    criteria.append({'criterion':'Assembly policy pinned','required':'Version required','actual':paper['assembly_policy_version'] or 'Not pinned','status':'PASS' if policy_ok else 'WARNING'})
    if not policy_ok: warnings.append('No assembly-policy version is pinned to this paper.')
    structural_ok=(not blocking and bool(paper['assessment_blueprint_id'])) if relation['code']=='AUTHENTIC_FULL_MOCK' else not blocking
    return {'paper':paper,'relation':relation,'criteria':criteria,'blocking':blocking,'warnings':warnings,
            'structural_status':'COMPLIANT' if structural_ok else ('NOT_APPLICABLE' if not paper['assessment_blueprint_id'] else 'NON_COMPLIANT'),
            'content_status':'COMPLIANT' if approved==total_actual and ready==total_actual else 'NON_COMPLIANT',
            'rigor_status':'PINNED' if policy_ok else 'UNPINNED','actual_by_subject':actual,'total_actual':total_actual}


def attempt_blueprint_result(c, attempt_id):
    attempt=c.execute("SELECT * FROM attempts WHERE id=?",(attempt_id,)).fetchone()
    if not attempt: return None
    rows=c.execute("""SELECT q.subject,COUNT(*) total,SUM(aa.is_correct) correct,
      ROUND(100.0*SUM(aa.is_correct)/NULLIF(COUNT(*),0),1) accuracy
      FROM attempt_answers aa JOIN questions q ON q.id=aa.question_db_id
      WHERE aa.attempt_id=? GROUP BY q.subject ORDER BY q.subject""",(attempt_id,)).fetchall()
    actual={r['subject']:r for r in rows}
    snapshot=safe_json(attempt['blueprint_snapshot_json'],{})
    sections=snapshot.get('sections') or []
    breakdown=[]
    if sections:
        for sec in sections:
            subject=sec.get('subject') or sec.get('section_title') or ''
            r=actual.get(subject)
            breakdown.append({'subject':subject,'official_questions':int(sec.get('question_count') or 0),
                              'weight_percent':float(sec.get('weight_percent') or 0),
                              'attempted_questions':int(r['total'] or 0) if r else 0,
                              'correct':int(r['correct'] or 0) if r else 0,
                              'accuracy':float(r['accuracy'] or 0) if r and r['accuracy'] is not None else 0.0})
    else:
        breakdown=[{'subject':r['subject'],'official_questions':None,'weight_percent':None,
                    'attempted_questions':int(r['total'] or 0),'correct':int(r['correct'] or 0),
                    'accuracy':float(r['accuracy'] or 0)} for r in rows]
    authenticity=''
    paper_kind=''
    if attempt['exam_paper_id']:
        pinned_paper=c.execute("SELECT authenticity_status,paper_kind FROM exam_papers WHERE id=?",(attempt['exam_paper_id'],)).fetchone()
        if pinned_paper:
            authenticity=pinned_paper['authenticity_status'] or ''
            paper_kind=pinned_paper['paper_kind'] or ''
    elif attempt['assessment_blueprint_id'] and (attempt['assessment_kind'] or '').lower() in ('mock','authentic_mock','exam_paper'):
        authenticity='AUTHENTIC_BLUEPRINT'
    relation=assessment_type_label(authenticity,paper_kind,attempt['scope'])
    return {'attempt':attempt,'snapshot':snapshot,'breakdown':breakdown,'relation':relation,
            'blueprint_pinned':bool(attempt['assessment_blueprint_id']),'blueprint_version':attempt['blueprint_version'] or '',
            'assembly_policy_version':attempt['assembly_policy_version'] or ''}


@app.route('/exam-structure')
def exam_structure_page():
    if not session.get('user_id'): return redirect(url_for('login'))
    c=db()
    active=c.execute("""SELECT ab.id,ab.powerhouse_blueprint_id,ab.blueprint_version,ab.total_questions,ab.duration_minutes,
      ab.authority,ab.activation_date,af.name framework_name,afv.version_name framework_version_name
      FROM assessment_blueprints ab JOIN assessment_frameworks af ON af.id=ab.framework_id
      JOIN assessment_framework_versions afv ON afv.id=ab.framework_version_id
      WHERE ab.local_status='ACTIVE' ORDER BY af.name,afv.version_name""").fetchall()
    structures=[exam_structure_snapshot(c,r['id']) for r in active]
    c.close()
    return render_template('exam_structure.html',structures=structures)


@app.route('/exam-structure/<int:blueprint_id>')
def exam_structure_detail(blueprint_id):
    if not session.get('user_id'): return redirect(url_for('login'))
    c=db(); structure=exam_structure_snapshot(c,blueprint_id)
    if not structure:
        c.close(); abort(404)
    bank=blueprint_bank_sufficiency(c,blueprint_id)
    c.close()
    return render_template('exam_structure_detail.html',structure=structure,bank=bank)


@app.route('/exam-centre/paper/<int:paper_id>/compliance')
def exam_paper_compliance_report(paper_id):
    if not session.get('user_id'): return redirect(url_for('login'))
    c=db(); report=paper_blueprint_compliance(c,paper_id)
    if not report:
        c.close(); abort(404)
    c.close()
    return render_template('exam_paper_compliance.html',report=report)


@app.route('/exam-centre')
def exam_centre():
    if not require('student'): return redirect(url_for('login'))
    c=db()
    data=exam_centre_data(c,session['user_id'])
    access=get_access_profile(c,session['user_id'])
    active_structures=c.execute("""SELECT ab.id,ab.powerhouse_blueprint_id,ab.blueprint_version,ab.total_questions,ab.duration_minutes,
      af.name framework_name,afv.version_name framework_version_name FROM assessment_blueprints ab
      JOIN assessment_frameworks af ON af.id=ab.framework_id JOIN assessment_framework_versions afv ON afv.id=ab.framework_version_id
      WHERE ab.local_status='ACTIVE' ORDER BY af.name""").fetchall()
    c.close()
    return render_template('exam_centre.html',access=access,days_remaining=exam_days_remaining,active_structures=active_structures,**data)


@app.route('/exam-centre/paper/<int:paper_id>')
def exam_paper_detail(paper_id):
    if not require('student'): return redirect(url_for('login'))
    c=db()
    paper=c.execute("""SELECT ep.*,eb.title legacy_blueprint_title,eb.country,eb.qualification,eb.exam_board,
        COALESCE(af.name,eb.programme,'') programme,COALESCE(eb.subject,'') subject,eb.paper_name,
        COALESCE(ab.duration_minutes,eb.duration_minutes) blueprint_duration,
        COALESCE(ab.total_questions,eb.total_marks) blueprint_marks,
        ab.powerhouse_blueprint_id authoritative_blueprint_id,af.name authoritative_framework,
        afv.version_name authoritative_framework_version
        FROM exam_papers ep LEFT JOIN exam_blueprints eb ON eb.id=ep.blueprint_id
        LEFT JOIN assessment_blueprints ab ON ab.id=ep.assessment_blueprint_id
        LEFT JOIN assessment_frameworks af ON af.id=ab.framework_id
        LEFT JOIN assessment_framework_versions afv ON afv.id=ab.framework_version_id
        WHERE ep.id=? AND ep.active=1""",(paper_id,)).fetchone()
    if not paper:
        c.close(); flash('Exam paper not found.','error'); return redirect(url_for('exam_centre'))
    questions=c.execute(f"""SELECT pq.*,q.question_id,q.qtype,q.level,q.chapter,q.topic
        FROM exam_paper_questions pq JOIN questions q ON q.id=pq.question_id
        WHERE pq.paper_id=? AND {live_question_clause('q')} ORDER BY pq.position""",(paper_id,)).fetchall()
    access=get_access_profile(c,session['user_id'])
    required_level=max(((q['level'] or 'Foundation') for q in questions),key=mastery_rank,default='Foundation') if questions else ''
    access_ok=not required_level or mastery_rank(required_level)<=mastery_rank(access['mastery_ceiling'])
    compliance=paper_blueprint_compliance(c,paper_id)
    structure=exam_structure_snapshot(c,paper['assessment_blueprint_id']) if paper['assessment_blueprint_id'] else None
    c.close()
    return render_template('exam_paper_detail.html',paper=paper,questions=questions,access=access,required_level=required_level,access_ok=access_ok,compliance=compliance,structure=structure)


@app.route('/exam-centre/paper/<int:paper_id>/start',methods=['POST'])
def start_exam_paper(paper_id):
    if not require('student'): return redirect(url_for('login'))
    guided=request.form.get('mode')=='guided'
    c=db()
    paper=c.execute("SELECT * FROM exam_papers WHERE id=? AND active=1",(paper_id,)).fetchone()
    if not paper:
        c.close(); flash('Exam paper not found.','error'); return redirect(url_for('exam_centre'))
    if paper['paper_kind']=='official_past_paper' and paper['reproduction_status'] not in ('permitted','licensed','public_domain'):
        c.close(); flash('This historical paper is not available for in-platform simulation until reproduction permission is confirmed.','error')
        return redirect(url_for('exam_paper_detail',paper_id=paper_id))
    access=get_access_profile(c,session['user_id'])
    if int(paper['premium_required'] or 0) and COMMERCIAL_GATES_ENABLED and int(access.get('access_rank',0))<1:
        c.close(); flash('This exam requires at least Level 1 Access.','error'); return redirect(url_for('access_account'))
    try:
        assessment_id=start_exam_paper_session(c,session['user_id'],paper,guided=guided)
    except PermissionError as exc:
        c.close(); flash(str(exc),'error'); return redirect(url_for('access_account'))
    c.close()
    if not assessment_id:
        flash('This paper has not been populated with approved questions yet.','error')
        return redirect(url_for('exam_paper_detail',paper_id=paper_id))
    return redirect(url_for('take_test_v4',assessment_id=assessment_id))


@app.route('/exam-centre/exam-date',methods=['POST'])
def save_exam_date():
    if not require('student'): return redirect(url_for('login'))
    subject=request.form.get('subject','').strip()
    exam_name=request.form.get('exam_name','').strip()
    exam_date=request.form.get('exam_date','').strip()
    if not subject or not exam_date:
        flash('Subject and exam date are required.','error'); return redirect(url_for('exam_centre'))
    try:
        if datetime.fromisoformat(exam_date).date()<datetime.now().date():
            raise ValueError()
    except Exception:
        flash('Please choose a valid future exam date.','error'); return redirect(url_for('exam_centre'))
    c=db()
    c.execute("INSERT INTO student_exam_dates(student_id,subject,exam_name,exam_date) VALUES(?,?,?,?)",
              (session['user_id'],subject,exam_name,exam_date))
    c.commit(); c.close()
    flash('Exam countdown added.','success')
    return redirect(url_for('exam_centre'))



def _subject_map(c, student_id):
    user=c.execute("SELECT academic_level,COALESCE(active_programme,'') active_programme,subjects FROM users WHERE id=?",(student_id,)).fetchone()
    programme=student_programme(c,student_id)
    # Declared subjects belong to the student's home programme only; switching programmes must not bleed context.
    visible=student_catalogue_subjects(user) if user and programme.casefold()==(user['academic_level'] or '').strip().casefold() else []
    matric_student=is_matric_level(programme)
    aliases=_programme_aliases(programme)
    scope_clause,scope_params=_programme_scope_sql(aliases,'q')
    live_rows=c.execute(f"""SELECT DISTINCT q.subject FROM questions q
      WHERE {live_question_clause('q')} AND COALESCE(q.subject,'')<>'' AND {scope_clause}
      ORDER BY q.subject""",scope_params).fetchall()
    visible_keys={x.casefold() for x in visible}
    for row in live_rows:
        # Current-study pages show only the learner's own programme.  Future
        # routes and other markets belong in Pathway Explorer.
        if matric_student and row['subject'].casefold() not in visible_keys:
            continue
        if row['subject'].casefold() not in visible_keys:
            visible.append(row['subject']); visible_keys.add(row['subject'].casefold())
    data=[]
    for subject in visible:
        subject_scope=f"{scope_clause} AND lower(q.subject)=lower(?)"
        params=scope_params+[subject]
        qcount=c.execute(f"SELECT COUNT(*) n FROM questions q WHERE {live_question_clause('q')} AND {subject_scope}",params).fetchone()['n']
        canonical=c.execute(f"SELECT q.subject FROM questions q WHERE {live_question_clause('q')} AND {subject_scope} LIMIT 1",params).fetchone()
        display=canonical['subject'] if canonical else subject
        r=c.execute(f"""SELECT COUNT(aa.id) answered,ROUND(100.0*SUM(aa.is_correct)/COUNT(aa.id),1) accuracy
          FROM attempt_answers aa JOIN attempts a ON a.id=aa.attempt_id JOIN questions q ON q.id=aa.question_db_id
          WHERE a.student_id=? AND {scope_clause} AND lower(q.subject)=lower(?)""",
          [student_id]+scope_params+[display]).fetchone()
        answered=int(r['answered'] or 0) if r else 0
        accuracy=float(r['accuracy'] or 0) if r and r['accuracy'] is not None else 0
        chapters=[]
        chapter_rows=c.execute(f"""SELECT q.chapter,MIN(q.id) first_id FROM questions q WHERE {live_question_clause('q')}
          AND {scope_clause} AND lower(q.subject)=lower(?) AND COALESCE(q.chapter,'')<>''
          GROUP BY q.chapter ORDER BY first_id""",scope_params+[display]).fetchall()
        for ch in chapter_rows:
            cr=c.execute(f"""SELECT COUNT(aa.id) answered,ROUND(100.0*SUM(aa.is_correct)/COUNT(aa.id),1) accuracy
              FROM attempt_answers aa JOIN attempts a ON a.id=aa.attempt_id JOIN questions q ON q.id=aa.question_db_id
              WHERE a.student_id=? AND {scope_clause} AND lower(q.subject)=lower(?) AND q.chapter=?""",
              [student_id]+scope_params+[display,ch['chapter']]).fetchone()
            ca=int(cr['answered'] or 0) if cr else 0
            cp=float(cr['accuracy'] or 0) if cr and cr['accuracy'] is not None else 0
            identity=chapter_identity(c,programme,display,ch['chapter'])
            chapters.append({**identity,'answered':ca,'accuracy':cp,'status':health_status(cp,ca)})
        access_state=commercial_access.subject_state(c,student_id,programme,display,
          available=bool(qcount),commercial_gates_enabled=COMMERCIAL_GATES_ENABLED)
        availability='LIVE' if access_state=='INCLUDED' else access_state
        status=health_status(accuracy,answered) if access_state=='INCLUDED' else ('Locked' if access_state=='LOCKED' else 'Coming Soon')
        data.append({'subject':display,'answered':answered,'accuracy':accuracy,'status':status,'chapters':chapters,
                     'availability':availability,'access_state':access_state,'available_questions':int(qcount or 0)})
    return data

@app.route('/student/programme',methods=['POST'])
def student_programme_switch():
    if not require('student'): return redirect(url_for('login'))
    code=(request.form.get('programme_code') or '').strip().lower()
    item=STUDENT_PROGRAMME_BY_CODE.get(code)
    if not item:
        flash('Choose FSc 1, FSc 2 or MDCAT.','error'); return redirect(request.referrer or url_for('student_dashboard'))
    c=db(); c.execute("UPDATE users SET active_programme=? WHERE id=?",(item['value'],session['user_id']))
    universal_mastery.emit_growth_event(c,'PROGRAMME_SWITCHED',f"USER:{session['user_id']}",{'programme':item['value']})
    c.commit(); c.close()
    target=(request.form.get('return_to') or '').strip()
    if target.startswith('/') and not target.startswith('//'):
        return redirect(target)
    return redirect(url_for('student_dashboard'))


@app.route('/student/subjects')
def subject_browser():
    if not require('student'): return redirect(url_for('login'))
    c=db(); subjects=_subject_map(c,session['user_id']); pathway_data=student_pathway_snapshot(c,session['user_id']); c.close()
    return render_template('subject_browser.html',subjects=subjects,pathway_data=pathway_data)


@app.route('/student/pathways',methods=['GET','POST'])
def student_pathways():
    if not require('student'): return redirect(url_for('login'))
    c=db()
    if request.method=='POST':
        code=request.form.get('pathway_code','').strip().upper()
        item=pathway_definition(code)
        if not item:
            c.close(); flash('Choose one of the available pathways.','error'); return redirect(url_for('student_pathways'))
        c.execute("""INSERT INTO student_pathway_preferences(student_id,pathway_code,status,selected_at,updated_at)
          VALUES(?,?,'ACTIVE',CURRENT_TIMESTAMP,CURRENT_TIMESTAMP)
          ON CONFLICT(student_id) DO UPDATE SET pathway_code=excluded.pathway_code,status='ACTIVE',updated_at=CURRENT_TIMESTAMP""",
          (session['user_id'],code))
        c.execute("UPDATE users SET future_pathway_code=? WHERE id=?",(code,session['user_id']))
        c.execute("INSERT INTO coach_nudge_events(student_id,nudge_key,action,context_json) VALUES(?,?,'COMPLETED',?)",
          (session['user_id'],'choose-future-pathway',json.dumps({'pathway_code':code})))
        c.commit(); c.close(); flash(f"Saved: {item['title']}. ScoreMax will use this as guidance, not a permanent lock.",'success')
        return redirect(url_for('student_pathways'))
    data=student_pathway_snapshot(c,session['user_id']); subjects=_subject_map(c,session['user_id']); c.close()
    return render_template('student_pathways.html',pathway_data=data,subjects=subjects)

@app.route('/student/subject/<path:subject>')
def subject_detail(subject):
    if not require('student'): return redirect(url_for('login'))
    requested=(subject or '').strip()
    c=db(); all_subjects=_subject_map(c,session['user_id'])
    selected=next((x for x in all_subjects if x['subject'].casefold()==requested.casefold()),None)
    if not selected:
        c.close(); flash('Subject not found.','error'); return redirect(url_for('subject_browser'))
    if selected.get('access_state')=='LOCKED':
        c.close(); flash(f'{selected["subject"]} is not included in your current subject package.','info')
        return redirect(url_for('access_account',locked_subject=selected['subject']))
    selected=dict(selected)
    programme=student_programme(c,session['user_id'])
    selected['chapters']=[dict(ch,mastery=chapter_mastery_opportunity(c,session['user_id'],selected['subject'],ch['chapter'],programme))
                          for ch in selected.get('chapters',[])]
    c.close()
    # One subject route renders one subject only. It must never fall back to another subject.
    return render_template('subject_detail.html',selected=selected)

@app.route('/student/chapter')
def chapter_page():
    if not require('student'): return redirect(url_for('login'))
    subject=request.args.get('subject','').strip(); chapter=request.args.get('chapter','').strip()
    if not subject or not chapter: return redirect(url_for('subject_browser'))
    c=db()
    blocked=subject_access_redirect(c,session['user_id'],subject)
    if blocked:
        c.close(); flash(f'{subject} is not included in your current subject package.','info'); return blocked
    topics=[]
    topic_names=[r['topic'] for r in c.execute(f"SELECT DISTINCT q.topic FROM questions q WHERE {live_question_clause('q')} AND q.subject=? AND q.chapter=? AND COALESCE(q.topic,'')<>'' ORDER BY q.topic",(subject,chapter)).fetchall()]
    for topic in topic_names:
        r=c.execute("""SELECT COUNT(aa.id) answered,ROUND(100.0*SUM(aa.is_correct)/COUNT(aa.id),1) accuracy
          FROM attempt_answers aa JOIN attempts a ON a.id=aa.attempt_id JOIN questions q ON q.id=aa.question_db_id
          WHERE a.student_id=? AND q.subject=? AND q.chapter=? AND q.topic=?""",(session['user_id'],subject,chapter,topic)).fetchone()
        answered=int(r['answered'] or 0) if r else 0; accuracy=float(r['accuracy'] or 0) if r and r['accuracy'] is not None else 0
        topics.append({'topic':topic,'answered':answered,'accuracy':accuracy,'status':health_status(accuracy,answered)})
    cr=c.execute("""SELECT COUNT(aa.id) answered,ROUND(100.0*SUM(aa.is_correct)/COUNT(aa.id),1) accuracy
      FROM attempt_answers aa JOIN attempts a ON a.id=aa.attempt_id JOIN questions q ON q.id=aa.question_db_id
      WHERE a.student_id=? AND q.subject=? AND q.chapter=?""",(session['user_id'],subject,chapter)).fetchone()
    answered=int(cr['answered'] or 0) if cr else 0; accuracy=float(cr['accuracy'] or 0) if cr and cr['accuracy'] is not None else 0
    weak=[w for w in student_weak_areas(c,session['user_id'],limit=20) if w['subject']==subject]
    recall=[r for r in c.execute("SELECT * FROM recall_items WHERE student_id=? AND subject=? AND chapter=? ORDER BY next_due_date",(session['user_id'],subject,chapter)).fetchall()]
    misconceptions=[m for m in confirmed_misconceptions(c,session['user_id'],limit=20) if m['subject']==subject]
    capsules=c.execute("SELECT * FROM learning_capsules WHERE active=1 AND review_status='Approved' AND subject=? AND chapter=? ORDER BY topic,concept",(subject,chapter)).fetchall()
    plan=c.execute("""SELECT spa.* FROM study_plan_activities spa JOIN study_plans sp ON sp.id=spa.plan_id
      WHERE spa.student_id=? AND sp.status='active' AND spa.subject=? AND spa.chapter=? AND spa.status<>'completed'
      ORDER BY spa.activity_date,spa.priority LIMIT 5""",(session['user_id'],subject,chapter)).fetchall()
    programme=student_programme(c,session['user_id'])
    mastery_opportunity=chapter_mastery_opportunity(c,session['user_id'],subject,chapter,programme)
    chapter_display=chapter_identity(c,programme,subject,chapter)
    subject_nav=_subject_map(c,session['user_id']); c.close()
    chapter_data={'subject':subject,**chapter_display,'answered':answered,'accuracy':accuracy,'status':health_status(accuracy,answered),
                  'mastery':mastery_opportunity}
    return render_template('chapter_page.html',chapter=chapter_data,topics=topics,weak=weak,recall=recall,
                           misconceptions=misconceptions,capsules=capsules,plan_items=plan,subject_nav=subject_nav)

@app.route('/student/weak-areas')
def weak_areas_page():
    if not require('student'): return redirect(url_for('login'))
    c=db()
    weak=student_weak_areas(c,session['user_id'])
    c.close()
    return render_template('weak_areas.html',weak=weak)

@app.route('/student/weak-areas/start',methods=['POST'])
def weak_areas_start():
    if not require('student'): return redirect(url_for('login'))
    selected=request.form.getlist('areas')
    count=max(5,min(30,int(request.form.get('count') or 15)))
    if not selected:
        flash('Choose at least one weak area to practise.','error')
        return redirect(url_for('weak_areas_page'))
    c=db()
    weak=student_weak_areas(c,session['user_id'],limit=30)
    wanted=[w for w in weak if f"{w['subject']}||{w['concept_key']}" in selected]
    ids=[]
    # Weight weaker areas more heavily while avoiding the exact questions the student has already seen when possible.
    for w in wanted:
        per_area=max(2,round(count/max(1,len(wanted))))
        rows=c.execute(f"""SELECT q.id,
          CASE WHEN EXISTS(SELECT 1 FROM attempt_answers aa JOIN attempts a ON a.id=aa.attempt_id
            WHERE aa.question_db_id=q.id AND a.student_id=?) THEN 1 ELSE 0 END seen
          FROM questions q WHERE q.subject=? AND {live_question_clause('q')}
          AND (q.concept_id=? OR q.concept=? OR q.learning_outcome=? OR q.subtopic=?)
          ORDER BY seen ASC,RANDOM() LIMIT ?""",
          (session['user_id'],w['subject'],w['concept_key'],w['concept_key'],w['area'],w['area'],per_area+2)).fetchall()
        ids.extend([r['id'] for r in rows])
    ids=filter_live_question_ids(c,list(dict.fromkeys(ids)))[:count]
    if len(ids)<3:
        c.close(); flash('Not enough approved questions are available for the selected weak areas yet.','error')
        return redirect(url_for('weak_areas_page'))
    subject=wanted[0]['subject'] if wanted else ''
    programme=wanted[0]['programme'] if wanted else ''
    meta={'scope':'weak_areas','programme':programme,'subject':subject,'chapters':'','topic':'','subtopic':'','level':'',
          'assessment_kind':'weak_area_practice','selected_weak_areas':[w['area'] for w in wanted]}
    assessment_id=create_assessment_session(c,session['user_id'],'practice',None,ids,meta)
    c.close()
    return redirect(url_for('take_test_v4',assessment_id=assessment_id))

@app.route('/learn/<capsule_id>')
def learning_capsule(capsule_id):
    if not require('student'): return redirect(url_for('login'))
    c=db()
    cap=c.execute("""SELECT * FROM learning_capsules WHERE capsule_id=? AND active=1
      AND review_status='Approved'""",(capsule_id,)).fetchone()
    c.close()
    if not cap:
        flash('Quick help for this concept is not available yet.','error')
        return redirect(request.referrer or url_for('student_dashboard'))
    return render_template('learning_capsule.html',cap=cap)

@app.route('/student/profile',methods=['GET','POST'])
def student_profile():
    if not require('student'): return redirect(url_for('login'))
    c=db()
    if request.method=='POST':
        action=request.form.get('action','goal')
        if action=='goal':
            goal_type=request.form.get('goal_type','').strip()[:60]
            goal_name=request.form.get('goal_name','').strip()[:120]
            raw=request.form.get('target_percentage','').strip()
            try: target=float(raw) if raw else None
            except ValueError: target=None
            if target is not None and not 0 <= target <= 100:
                flash('Target percentage must be between 0 and 100.','error'); c.close(); return redirect(url_for('student_profile'))
            c.execute("UPDATE users SET goal_type=?,goal_name=?,target_percentage=? WHERE id=?",
                      (goal_type,goal_name,target,session['user_id']))
            if target is not None and goal_type in ('MBBS','BDS'):
                c.execute("UPDATE student_admission_targets SET active=0 WHERE student_id=? AND route_code='PK_MEDICAL'",(session['user_id'],))
                c.execute("""INSERT INTO student_admission_targets(student_id,route_code,institution_name,programme_name,target_aggregate,admission_year,active)
                  VALUES(?,?,?,?,?,?,1)""",(session['user_id'],'PK_MEDICAL',goal_name,goal_type,target,datetime.now().year))
            c.commit(); flash('Your goal has been saved.','success')
        elif action=='result':
            exam_type=request.form.get('exam_type','').strip()
            subject=request.form.get('subject','').strip()
            board=request.form.get('board_authority','').strip()
            year=int(request.form.get('exam_year') or datetime.now().year)
            obtained=float(request.form.get('marks_obtained') or 0)
            total=float(request.form.get('total_marks') or 0)
            pct=round(100*obtained/total,2) if total>0 else None
            # Snapshot current ScoreMax evidence in that subject where available.
            row=c.execute("""SELECT COUNT(*) answered,ROUND(100.0*SUM(aa.is_correct)/COUNT(*),1) accuracy
              FROM attempt_answers aa JOIN attempts a ON a.id=aa.attempt_id
              WHERE a.student_id=? AND (?='' OR a.subject=?)""",(session['user_id'],subject,subject)).fetchone()
            lvl=scoremax_level_from_evidence(row['accuracy'],row['answered']) if row and row['answered'] else ''
            cur=c.execute("""INSERT INTO student_external_results(student_id,exam_type,subject,board_authority,exam_year,
              marks_obtained,total_marks,percentage,grade,verification_status,scoremax_level_snapshot,scoremax_score_snapshot)
              VALUES(?,?,?,?,?,?,?,?,?,'self_reported',?,?)""",(session['user_id'],exam_type,subject,board,year,
              obtained,total,pct,request.form.get('grade','').strip(),lvl,row['accuracy'] if row else None))
            result_id=cur.lastrowid
            subjects=request.form.getlist('subject_name')
            obtained_list=request.form.getlist('subject_obtained')
            total_list=request.form.getlist('subject_total')
            grade_list=request.form.getlist('subject_grade')
            for i,name in enumerate(subjects):
                name=(name or '').strip()
                if not name: continue
                try:
                    so=float(obtained_list[i] or 0); st=float(total_list[i] or 0)
                except Exception: continue
                sp=round(100*so/st,2) if st>0 else None
                sg=(grade_list[i] if i<len(grade_list) else '').strip()
                c.execute("""INSERT INTO external_result_subjects(external_result_id,student_id,subject,marks_obtained,total_marks,percentage,grade)
                  VALUES(?,?,?,?,?,?,?)""",(result_id,session['user_id'],name,so,st,sp,sg))
            c.commit(); flash('Achievement saved. Add another result whenever you are ready.','success')
    user=c.execute("SELECT * FROM users WHERE id=?",(session['user_id'],)).fetchone()
    results=c.execute("SELECT * FROM student_external_results WHERE student_id=? ORDER BY exam_year DESC,id DESC",
                      (session['user_id'],)).fetchall()
    subject_results=c.execute("SELECT * FROM external_result_subjects WHERE student_id=? ORDER BY external_result_id,id",
                              (session['user_id'],)).fetchall()
    progress=student_dashboard_intelligence(c,session['user_id'])
    parent_links=c.execute("""SELECT psl.*,u.full_name parent_name FROM parent_student_links psl JOIN users u ON u.id=psl.parent_user_id
      WHERE psl.student_user_id=? AND psl.status='active' ORDER BY psl.id DESC""",(session['user_id'],)).fetchall()
    notifications=c.execute("SELECT * FROM user_notifications WHERE user_id=? ORDER BY id DESC LIMIT 8",(session['user_id'],)).fetchall()
    target_snapshot=live_target_snapshot(c,session['user_id'])
    formula=c.execute("SELECT * FROM admission_formulas WHERE route_code='PK_MEDICAL' AND active=1 ORDER BY admission_year DESC,id DESC LIMIT 1").fetchone()
    medical_components=safe_json(formula['components_json'],[]) if formula else []
    c.close()
    return render_template('student_profile.html',user=user,results=results,subject_results=subject_results,subjects=progress.get('subjects',[]),
                           parent_links=parent_links,notifications=notifications,target_snapshot=target_snapshot,medical_formula=formula,medical_components=medical_components)

@app.route('/account/profile',methods=['GET','POST'])
def account_settings():
    if not require(): return redirect(url_for('login'))
    c=db(); u=c.execute("SELECT * FROM users WHERE id=?",(session['user_id'],)).fetchone()
    if request.method=='POST':
        full_name=request.form.get('full_name','').strip()[:120]
        if not full_name: c.close(); flash('Name cannot be empty.','error'); return redirect(url_for('account_settings'))
        c.execute("""UPDATE users SET full_name=?,mobile=?,academic_level=?,board=?,province=?,division=?,district=?,profile_completed=1 WHERE id=?""",
          (full_name,request.form.get('mobile','').strip(),request.form.get('academic_level','').strip(),request.form.get('board','').strip(),
           request.form.get('province','').strip(),request.form.get('division','').strip(),request.form.get('district','').strip(),session['user_id']))
        c.commit(); session['full_name']=full_name; flash('Profile & Settings updated.','success'); u=c.execute("SELECT * FROM users WHERE id=?",(session['user_id'],)).fetchone()
    c.close(); return render_template('account_settings.html',user=u)

@app.route('/account/access',methods=['GET','POST'])
def access_account():
    if not require('student'): return redirect(url_for('dashboard'))
    c=db(); programme=student_programme(c,session['user_id'])
    if request.method=='POST':
        try:
            package_id=int(request.form.get('coverage_package_id') or 0)
            access_code=request.form.get('access_plan_code','').strip()
            request_id=commercial_access.request_checkout(c,student_id=session['user_id'],coverage_package_id=package_id,access_plan_code=access_code)
            c.commit()
            flash(f'Access request {request_id} saved. The live payment gateway is not connected in this controlled-pilot build; an Admin can activate it after payment verification.','success')
        except ValueError as exc:
            c.rollback(); flash(str(exc),'error')
        c.close(); return redirect(url_for('access_account'))
    access=get_access_profile(c,session['user_id'])
    plans=[]
    for code in ['free_access','level_1_access','level_2_access','full_access']:
        r=c.execute("SELECT * FROM plans WHERE code=?",(code,)).fetchone(); d=dict(r) if r else {'code':code,'name':ACCESS_CODES[code]['name'],'price_minor':None,'currency':'PKR'}
        d['ceiling']=ACCESS_CODES[code]['ceiling']; d['rank']=ACCESS_CODES[code]['rank']; d['price_display']=money_display(d.get('price_minor'),d.get('currency','PKR')); plans.append(d)
    packages=commercial_access.package_rows(c,programme,include_coming=True)
    # Keep future programme options visible without forcing an extra catalogue click.
    if not packages: packages=commercial_access.package_rows(c,'',include_coming=True)
    coverage=commercial_access.effective_coverage(c,session['user_id'],programme,commercial_gates_enabled=COMMERCIAL_GATES_ENABLED)
    c.close(); return render_template('access.html',access=access,plans=plans,levels=MASTER_LEVELS,packages=packages,
      coverage=coverage,programme=programme,locked_subject=request.args.get('locked_subject',''),commercial_gates=COMMERCIAL_GATES_ENABLED)

@app.route('/student/mastery',methods=['GET','POST'])
def mastery_page():
    if not require('student'): return redirect(url_for('login'))
    c=db(); access=get_access_profile(c,session['user_id']); records=current_mastery_records(c,session['user_id'])
    programmes=[r['programme'] for r in c.execute(f"SELECT DISTINCT q.programme FROM questions q WHERE {live_question_clause('q')} AND COALESCE(q.programme,'')<>'' ORDER BY q.programme").fetchall()]
    subjects=[r['subject'] for r in c.execute(f"SELECT DISTINCT q.subject FROM questions q WHERE {live_question_clause('q')} AND COALESCE(q.subject,'')<>'' ORDER BY q.subject").fetchall()]
    chapter_rows=[dict(r) for r in c.execute(f"SELECT COALESCE(q.programme,'') programme,q.subject,q.chapter,MIN(q.id) first_id FROM questions q WHERE {live_question_clause('q')} AND COALESCE(q.chapter,'')<>'' GROUP BY programme,q.subject,q.chapter ORDER BY q.subject,first_id").fetchall()]
    chapters=[dict(row,**{k:v for k,v in chapter_identity(c,row['programme'],row['subject'],row['chapter']).items() if k!='chapter'}) for row in chapter_rows]
    policies=[dict(r) for r in c.execute("SELECT * FROM mastery_policies WHERE active=1 ORDER BY level_rank").fetchall()]
    if request.method=='POST':
        scope_type=request.form.get('scope_type','chapter').lower(); target=request.form.get('target_level','Foundation')
        programme=request.form.get('programme','').strip(); subject=request.form.get('subject','').strip(); chapter=request.form.get('chapter','').strip()
        try:
            if subject:
                state=student_subject_state(c,session['user_id'],subject,programme)
                if state=='LOCKED': raise PermissionError(f'{subject} is not included in your current subject package.')
                if state=='COMING_SOON': raise ValueError(f'{subject} is Coming Soon for your programme.')
            rows,meta=build_mastery_form(c,session['user_id'],scope_type,target,programme,subject,chapter)
            duration=max(10,min(120,int(round(len(rows)*1.5))))
            aid=create_assessment_session(c,session['user_id'],'exam',duration,[r['id'] for r in rows],meta); c.close()
            return redirect(url_for('take_test_v4',assessment_id=aid))
        except PermissionError as exc:
            c.close(); flash(str(exc),'error'); return redirect(url_for('access_account'))
        except ValueError as exc:
            c.close(); flash(str(exc),'error'); return redirect(url_for('mastery_page'))
    c.close(); return render_template('mastery.html',access=access,records=records,programmes=programmes,subjects=subjects,chapters=chapters,policies=policies,levels=MASTER_LEVELS)

@app.route('/student/demo-progress',methods=['POST'])
def load_demo_progress():
    if not require('student'): return redirect(url_for('login'))
    if SCOREMAX_ENV=='production':
        flash('Demo progress is disabled in the live pilot environment.','error'); return redirect(url_for('student_dashboard'))
    c=db(); existing=c.execute("SELECT COUNT(*) n FROM attempts WHERE student_id=? AND assessment_kind='demo_progress'",(session['user_id'],)).fetchone()['n']
    if existing:
        c.close(); flash('Demo progress is already loaded for this account.','success'); return redirect(url_for('student_dashboard'))
    specs=[('Biology',90),('Biology',82),('Chemistry',70),('Chemistry',62),('Physics',40),('Physics',35)]
    for subject,pct in specs:
        qs=c.execute("SELECT * FROM questions WHERE is_demo=1 AND subject=? ORDER BY id LIMIT 10",(subject,)).fetchall()
        if not qs: continue
        correct_n=round(len(qs)*pct/100)
        cur=c.execute("""INSERT INTO attempts(student_id,scope,programme,subject,chapters,level,score,correct_count,total_count,assessment_kind)
          VALUES(?, 'demo','FSc Part 1',?,'','',?,?,?, 'demo_progress')""",
          (session['user_id'],subject,round(100*correct_n/len(qs),1),correct_n,len(qs)))
        attempt_id=cur.lastrowid
        answer_rows=[]
        for i,q in enumerate(qs):
            ok=1 if i<correct_n else 0
            misconception='Demo repeated Physics misconception' if subject=='Physics' and not ok else ''
            answer_rows.append((attempt_id,q['id'],'A' if ok else 'B',ok,1 if ok else 0,int(q['question_version'] or 1),misconception,'confident' if not ok else 'unsure',20))
        c.executemany("""INSERT INTO attempt_answers(attempt_id,question_db_id,selected_answer,is_correct,marks_awarded,question_version,misconception_triggered,confidence,response_time_seconds)
          VALUES(?,?,?,?,?,?,?,?,?)""",answer_rows)
        update_learning_intelligence_from_attempt(c,attempt_id,session['user_id']); _refresh_item_calibration(c,[q['id'] for q in qs])
    c.commit(); c.close(); flash('Demo progress loaded. It changes analytics only and creates no formal mastery.','success'); return redirect(url_for('student_dashboard'))

@app.route('/about')
def about_page():
    return render_template('about.html')

@app.route('/how-it-works')
def how_it_works():
    return render_template('how_it_works.html')

@app.route('/faq')
def faq_page():
    return render_template('faq.html')

@app.route('/updates')
def updates_page():
    c=db()
    updates=c.execute("""SELECT * FROM content_updates WHERE status='published'
      ORDER BY COALESCE(important_date,published_date,created_at) DESC,id DESC LIMIT 50""").fetchall()
    c.close()
    return render_template('updates.html',updates=updates)

@app.route('/contact',methods=['GET','POST'])
def contact_page():
    if request.method=='POST':
        c=db()
        c.execute("""INSERT INTO contact_messages(name,email,mobile,enquiry_type,message)
          VALUES(?,?,?,?,?)""",(request.form.get('name','').strip(),request.form.get('email','').strip(),
          request.form.get('mobile','').strip(),request.form.get('enquiry_type','general'),
          request.form.get('message','').strip()))
        c.commit(); c.close(); flash('Thanks. Your message has been received.','success')
        return redirect(url_for('contact_page'))
    return render_template('contact.html')

@app.route('/admin/learning-capsules')
def admin_learning_capsules():
    if not require('admin'): return redirect(url_for('login'))
    c=db(); capsules=c.execute("SELECT * FROM learning_capsules ORDER BY subject,chapter,concept").fetchall(); c.close()
    return render_template('admin_learning_capsules.html',capsules=capsules)


@app.route('/student/study-plan')
def study_plan_page():
    if not require('student'): return redirect(url_for('login'))
    c=db()
    maybe_weekly_rebalance(c,session['user_id'])
    data=get_active_study_plan(c,session['user_id'])
    if data:
        data=dict(data)
        enriched=[]
        for item in data.get('activities',[]):
            d=dict(item)
            ident=chapter_identity(c,data['plan']['target_exam'] or student_programme(c,session['user_id']),d.get('subject',''),d.get('chapter','')) if d.get('chapter') else None
            d['chapter_display']=ident['display_label'] if ident else ''
            enriched.append(d)
        data['activities']=enriched
        if data.get('next_item'):
            next_id=data['next_item']['id']
            data['next_item']=next((x for x in enriched if x.get('id')==next_id),dict(data['next_item']))
    recommended=recommend_plan_pathway(c,session['user_id'])
    recall_due=due_recall_items(c,session['user_id'],limit=6)
    misconceptions=confirmed_misconceptions(c,session['user_id'],limit=6)
    access=get_access_profile(c,session['user_id'])
    mastery=current_mastery_records(c,session['user_id'])
    pathway_previews={}
    c.close()
    return render_template('study_plan.html',plan_data=data,recommended=recommended,pathways=STUDY_PLAN_PATHWAYS,
      pathway_previews=pathway_previews,recall_due=recall_due,misconceptions=misconceptions,access=access,mastery_records=mastery)

@app.route('/student/study-plan/create',methods=['POST'])
def create_study_plan():
    if not require('student'): return redirect(url_for('login'))
    pathway=request.form.get('pathway','Core')
    source=request.form.get('source','scoremax')
    if source=='self':
        pathway='Custom'
    target_exam=request.form.get('target_exam','').strip()
    target_date=request.form.get('target_date','').strip()
    raw=request.form.get('target_percentage','').strip()
    target=float(raw) if raw else None
    cfg=STUDY_PLAN_PATHWAYS.get(pathway,STUDY_PLAN_PATHWAYS['Core'])
    # V6.2.8: scheduling is priority/evidence-led. Internal cadence values remain
    # implementation details and are never presented as required study time.
    days_per_week=int(cfg['days_per_week'])
    minutes_per_day=int(cfg['minutes_per_day'])
    coverage_raw=request.form.get('starting_coverage','').strip()
    starting_coverage=float(coverage_raw) if coverage_raw else None
    c=db()
    access=get_access_profile(c,session['user_id'])
    if int(access.get('access_rank',0))<1:
        c.close(); flash('Level 1 Access or above is needed to build a live Study Plan. You can still preview the pathways.','error'); return redirect(url_for('access_account'))
    try:
        generate_scoremax_plan(c,session['user_id'],pathway,target_exam,target_date,target,
                           days_per_week=days_per_week,minutes_per_day=minutes_per_day,
                           starting_coverage=starting_coverage,
                           custom_settings={'planning_model':'PRIORITY_EVIDENCE_SEQUENCE'})
    except ValueError as exc:
        c.close(); flash(str(exc),'error'); return redirect(url_for('study_plan_page'))
    c.execute("UPDATE study_plans SET source=? WHERE student_id=? AND status='active'",
              ('self' if source=='self' else 'scoremax',session['user_id']))
    c.commit(); c.close()
    flash(f'{pathway} Study Plan is ready.','success')
    return redirect(url_for('study_plan_page'))

@app.route('/student/study-plan/activity/<int:activity_id>/complete',methods=['POST'])
def complete_study_plan_activity(activity_id):
    if not require('student'): return redirect(url_for('login'))
    c=db()
    c.execute("""UPDATE study_plan_activities SET status='completed',completed_at=CURRENT_TIMESTAMP,
      evidence_status='self_reported',evidence_type='student_confirmation',outcome_status='completed_outside_scoremax',last_evidence_at=CURRENT_TIMESTAMP
      WHERE id=? AND student_id=?""",(activity_id,session['user_id']))
    c.commit(); c.close()
    flash('Saved as self-reported completion. ScoreMax marks test-based work as Verified only when it sees the attempt.','success')
    return redirect(request.referrer or url_for('study_plan_page'))

@app.route('/student/study-plan/activity/<int:activity_id>/start',methods=['POST'])
def start_study_plan_activity(activity_id):
    if not require('student'): return redirect(url_for('login'))
    c=db(); x=c.execute("SELECT * FROM study_plan_activities WHERE id=? AND student_id=?",(activity_id,session['user_id'])).fetchone()
    if not x or x['status']=='completed':
        c.close(); flash('That plan activity is no longer available.','error'); return redirect(url_for('study_plan_page'))
    kind=x['activity_type'] or 'practice'
    if x['subject']:
        state=student_subject_state(c,session['user_id'],x['subject'])
        if state!='INCLUDED':
            c.close(); flash(f"{x['subject']} is not included in your current package." if state=='LOCKED' else f"{x['subject']} is Coming Soon.",'info')
            return redirect(url_for('access_account',locked_subject=x['subject']) if state=='LOCKED' else url_for('subject_browser'))
    if kind=='mock':
        plan=c.execute("SELECT * FROM study_plans WHERE id=? AND student_id=?",(x['plan_id'],session['user_id'])).fetchone()
        if plan and plan['assessment_blueprint_id']:
            try:
                preflight=assemble_blueprint_mock(c,plan['assessment_blueprint_id'],student_id=session['user_id'],seed=f"plan-{activity_id}")
            except ValueError as exc:
                c.close(); flash(str(exc),'error'); return redirect(url_for('study_plan_page'))
            if not preflight['ready']:
                message=' Authentic full mock blocked: '+' '.join(preflight['blockers'][:3])
                c.close(); flash(message,'error'); return redirect(url_for('study_plan_page'))
            bp=blueprint_joined(c,plan['assessment_blueprint_id'])
            snapshot=blueprint_payload_from_record(c,bp['id']) or {}
            meta={'scope':'study_plan','programme':bp['framework_name'],'subject':'','chapters':'','topic':'','subtopic':'','level':'',
                  'assessment_kind':'mock','plan_activity_id':activity_id,'assessment_blueprint_id':bp['id'],
                  'blueprint_source_id':bp['powerhouse_blueprint_id'],'blueprint_version':bp['blueprint_version'],
                  'framework_version':bp['framework_version_name'],'blueprint_snapshot':snapshot,
                  'assembly_policy_id':preflight['policy_id'],'assembly_policy_version':preflight['policy_version'],
                  'authenticity_status':'AUTHENTIC_BLUEPRINT'}
            aid=create_assessment_session(c,session['user_id'],'mock',int(bp['duration_minutes'] or 0) or None,
                                          [q['question_id'] for q in preflight['selected']],meta)
            c.close(); return redirect(url_for('take_test_v4',assessment_id=aid))
    clauses=[live_question_clause('q')]
    params=[]
    # Only auto-markable types can write verified evidence into the plan.
    qtype_sql="("+",".join("?" for _ in LIVE_MARKABLE_TYPES)+")"
    clauses.append(f"lower(replace(replace(q.qtype,'_',' '),'-',' ')) IS NOT NULL")
    if x['subject']:
        clauses.append('q.subject=?'); params.append(x['subject'])
    if x['chapter']:
        clauses.append('q.chapter=?'); params.append(x['chapter'])
    concept=(x['concept_key'] or '').strip()
    if concept:
        clauses.append("(q.concept_id=? OR q.concept=? OR q.learning_outcome=? OR q.subtopic=?)"); params.extend([concept]*4)
    candidates=c.execute(f"SELECT q.* FROM questions q WHERE {' AND '.join(clauses)} ORDER BY RANDOM()",params).fetchall()
    access=get_access_profile(c,session['user_id'])
    candidates=[q for q in candidates if canonical_question_type(q) in LIVE_MARKABLE_TYPES and mastery_rank(q['level'] or 'Foundation')<=mastery_rank(access['mastery_ceiling'])]
    seen={r['question_db_id'] for r in c.execute("SELECT aa.question_db_id FROM attempt_answers aa JOIN attempts a ON a.id=aa.attempt_id WHERE a.student_id=?",(session['user_id'],)).fetchall()}
    unseen=[q for q in candidates if q['id'] not in seen]
    pool=unseen+candidates
    unique=[]; used=set()
    for q in pool:
        if q['id'] not in used: unique.append(q); used.add(q['id'])
    count=5 if kind=='recall' else (20 if kind in ('mixed_test','mock') else 10)
    rows=unique[:count]
    if len(rows)<3:
        c.close(); flash('There are not enough approved auto-markable questions for this activity yet.','error'); return redirect(url_for('study_plan_page'))
    mode='mock' if kind=='mock' else 'practice'
    meta={'scope':'study_plan','programme':'','subject':x['subject'] or '','chapters':x['chapter'] or '',
          'topic':x['topic'] or '','subtopic':'','level':'','assessment_kind':'recall' if kind=='recall' else ('recovery' if kind in ('recovery','weak_area') else kind),
          'recovery_focus_type':'concept' if concept else '','recovery_focus_name':concept,'plan_activity_id':activity_id}
    aid=create_assessment_session(c,session['user_id'],mode,None,[q['id'] for q in rows],meta)
    c.close(); return redirect(url_for('take_test_v4',assessment_id=aid))

@app.route('/student/test-me',methods=['POST'])
def scoremax_test_me():
    if not require('student'): return redirect(url_for('login'))
    subject=request.form.get('subject','').strip()
    count=max(10,min(40,int(request.form.get('count') or 20)))
    if not subject:
        flash('Choose a subject first.','error'); return redirect(url_for('test_setup'))
    c=db()
    programme=student_programme(c,session['user_id'])
    state=student_subject_state(c,session['user_id'],subject,programme)
    if state!='INCLUDED':
        c.close()
        flash(f'{subject} is not included in your current subject package.' if state=='LOCKED' else f'{subject} is Coming Soon.','info')
        return redirect(url_for('access_account',locked_subject=subject) if state=='LOCKED' else url_for('subject_browser'))
    # Prioritise unseen questions and weaker/untested coverage while keeping a mixed challenge level.
    weak=student_weak_areas(c,session['user_id'],limit=8)
    weak_keys=[w['concept_key'] for w in weak if w['subject']==subject]
    weak_slots=",".join("?" for _ in weak_keys) or "''"
    rows=c.execute(f"""SELECT q.*,
      CASE WHEN EXISTS(SELECT 1 FROM attempt_answers aa JOIN attempts a ON a.id=aa.attempt_id
        WHERE aa.question_db_id=q.id AND a.student_id=?) THEN 1 ELSE 0 END seen,
      CASE WHEN q.concept_id IN ({weak_slots}) OR q.concept IN ({weak_slots}) THEN 0 ELSE 1 END weak_priority
      FROM questions q WHERE {live_question_clause('q')} AND q.subject=?
        AND (?='' OR lower(COALESCE(q.programme,''))=lower(?) OR lower(COALESCE(q.qualification,''))=lower(?))
      ORDER BY seen ASC,weak_priority ASC,RANDOM() LIMIT ?""",
        [session['user_id']]+weak_keys+weak_keys+[subject,programme,programme,programme,count*3]).fetchall()
    access=get_access_profile(c,session['user_id'])
    rows=[r for r in rows if canonical_question_type(r) in LIVE_MARKABLE_TYPES and mastery_rank(r['level'] or 'Foundation')<=mastery_rank(access['mastery_ceiling'])][:count]
    if not rows:
        c.close(); flash('No approved questions are available for this subject yet.','error'); return redirect(url_for('test_setup'))
    meta={'scope':'scoremax_test','programme':programme,'subject':subject,'chapters':'','topic':'','subtopic':'','level':'',
          'assessment_kind':'scoremax_test'}
    aid=create_assessment_session(c,session['user_id'],'practice',None,[r['id'] for r in rows],meta)
    c.close()
    return redirect(url_for('take_test_v4',assessment_id=aid))

@app.route('/student/blueprint-practice/start',methods=['POST'])
def start_blueprint_practice():
    if not require('student'): return redirect(url_for('login'))
    practice_type=request.form.get('practice_type','proportional_full')
    try:
        count=max(10,min(180,int(request.form.get('count') or (90 if practice_type=='proportional_half' else 60))))
    except ValueError:
        count=60
    c=db(); u=c.execute("SELECT academic_level,goal_name FROM users WHERE id=?",(session['user_id'],)).fetchone()
    bp=None
    for name in [x for x in ((u['academic_level'] if u else ''),(u['goal_name'] if u else '')) if (x or '').strip()]:
        bp=active_assessment_blueprint(c,name)
        if bp: break
    if not bp:
        c.close(); flash('No active assessment blueprint matches your current exam goal yet.','error'); return redirect(url_for('student_dashboard'))
    try:
        assembled=assemble_blueprint_practice(c,bp['id'],session['user_id'],count,practice_type,seed=datetime.now().isoformat())
    except (ValueError,PermissionError) as exc:
        c.close(); flash(str(exc),'error'); return redirect(url_for('student_dashboard'))
    if not assembled['ready']:
        c.close(); flash('Blueprint-aware practice could not be assembled: '+' '.join(assembled['blockers'][:2]),'error'); return redirect(url_for('student_dashboard'))
    label={'proportional_full':'blueprint_proportional_practice','proportional_half':'blueprint_half_practice',
      'diagnostic':'blueprint_diagnostic','adaptive':'blueprint_adaptive','powerhouse_governed':'powerhouse_governed_blueprint_practice'}[assembled['practice_type']]
    meta={'scope':assembled['practice_type'],'programme':bp['framework_name'],'subject':'','chapters':'','topic':'','subtopic':'','level':'',
      'assessment_kind':label,'authenticity_status':assembled['authenticity_status'],
      'assessment_blueprint_id':bp['id'],'blueprint_source_id':bp['powerhouse_blueprint_id'],
      'blueprint_version':bp['blueprint_version'],'framework_version':bp['framework_version_name'],
      'blueprint_snapshot':assembled['blueprint_snapshot'],'assembly_policy_id':assembled['assembly_policy_id'],
      'assembly_policy_version':assembled['assembly_policy_version'],'blueprint_subject_allocations':assembled['allocations'],
      'subject_policy_versions':assembled['subject_policies'],'diagnostic_deviation':assembled['practice_type'] in ('diagnostic','adaptive')}
    assessment_id=create_assessment_session(c,session['user_id'],'practice',assembled.get('duration_minutes'),[x['question_id'] for x in assembled['selected']],meta)
    c.close(); return redirect(url_for('take_test_v4',assessment_id=assessment_id))


@app.route('/test/setup')
def test_setup():
    if not require('student'): return redirect(url_for('login'))
    c=db()
    programmes=[r['programme'] for r in c.execute(f"""SELECT DISTINCT programme FROM (
        SELECT programme FROM curriculum WHERE COALESCE(programme,'')<>''
        UNION SELECT q.programme FROM questions q WHERE {live_question_clause('q')} AND COALESCE(q.programme,'')<>''
      ) ORDER BY programme""").fetchall()]
    subject_programmes=[dict(r) for r in c.execute(f"SELECT DISTINCT COALESCE(q.programme,'') programme,q.subject FROM questions q WHERE {live_question_clause('q')} AND COALESCE(q.subject,'')<>'' ORDER BY programme,q.subject").fetchall()]
    subjects=sorted({r['subject'] for r in subject_programmes})
    chapter_rows=[dict(r) for r in c.execute(f"SELECT COALESCE(q.programme,'') programme,q.subject,q.chapter,MIN(q.id) first_id FROM questions q WHERE {live_question_clause('q')} AND COALESCE(q.chapter,'')<>'' GROUP BY programme,q.subject,q.chapter ORDER BY programme,q.subject,first_id").fetchall()]
    chapters=[dict(row,**{k:v for k,v in chapter_identity(c,row['programme'],row['subject'],row['chapter']).items() if k!='chapter'}) for row in chapter_rows]
    topics=[dict(r) for r in c.execute(f"SELECT DISTINCT COALESCE(q.programme,'') programme,q.subject,q.topic FROM questions q WHERE {live_question_clause('q')} AND COALESCE(q.topic,'')<>'' ORDER BY programme,q.subject,q.topic").fetchall()]
    access=get_access_profile(c,session['user_id'])
    subject_cards=_subject_map(c,session['user_id'])
    included={x['subject'].casefold() for x in subject_cards if x.get('access_state')=='INCLUDED'}
    subjects=[x for x in subjects if x.casefold() in included]
    subject_programmes=[x for x in subject_programmes if x['subject'].casefold() in included]
    c.close()
    prefill={k:request.args.get(k,'') for k in ['programme','subject','chapter','topic','subtopic','level','scope','mode']}
    return render_template('test_setup.html',programmes=programmes,subjects=subjects,subject_programmes=subject_programmes,chapters=chapters,topics=topics,prefill=prefill,access=access,levels=MASTER_LEVELS,subject_cards=subject_cards)

@app.route('/test/start',methods=['POST'])
def test_start():
    if not require('student'):
        return redirect(url_for('login'))

    programme=request.form.get('programme_override','').strip() or request.form.get('programme','').strip()
    subject=request.form['subject']
    scope=request.form['scope']
    count=max(1,min(100,int(request.form['count'])))
    level=request.form.get('level_override','').strip() or request.form.get('level','').strip()
    mode=request.form.get('mode','practice')
    access_check=db(); access=get_access_profile(access_check,session['user_id'])
    state=student_subject_state(access_check,session['user_id'],subject,programme)
    access_check.close()
    if state=='LOCKED':
        flash(f'{subject} is not included in your current subject package.','info')
        return redirect(url_for('access_account',locked_subject=subject))
    if state=='COMING_SOON':
        flash(f'{subject} is Coming Soon for your programme.','info')
        return redirect(url_for('subject_browser'))
    if level and mastery_rank(level)>mastery_rank(access['mastery_ceiling']):
        flash(f"{level} questions require higher Access. Your current ceiling is {access['mastery_ceiling']}.",'error')
        return redirect(url_for('access_account'))
    duration_raw=request.form.get('duration_minutes','').strip()
    duration=int(duration_raw) if duration_raw.isdigit() else None

    clauses=[live_question_clause('q'),'q.subject=?']
    params=[subject]
    if programme:
        clauses+=['q.programme=?']; params+=[programme]
    if level:
        clauses+=['q.level=?']; params+=[level]
    if scope=='topic' and request.form.get('topic'):
        clauses+=['q.topic=?']; params+=[request.form['topic']]
    if scope=='topic' and request.form.get('subtopic'):
        clauses+=['q.subtopic=?']; params+=[request.form['subtopic']]
    if scope=='chapter' and request.form.get('chapter'):
        clauses+=['q.chapter=?']; params+=[request.form['chapter']]

    chapters=request.form.getlist('chapters')
    if scope=='half':
        ctmp=db()
        if programme:
            all_chapters=[r['chapter'] for r in ctmp.execute(f"""SELECT q.chapter,MIN(q.id) first_id FROM questions q
              WHERE {live_question_clause('q')} AND q.subject=? AND q.programme=? AND COALESCE(q.chapter,'')<>'' GROUP BY q.chapter ORDER BY first_id""",(subject,programme)).fetchall()]
        else:
            all_chapters=[r['chapter'] for r in ctmp.execute(f"""SELECT q.chapter,MIN(q.id) first_id FROM questions q
              WHERE {live_question_clause('q')} AND q.subject=? AND COALESCE(q.chapter,'')<>'' GROUP BY q.chapter ORDER BY first_id""",(subject,)).fetchall()]
        ctmp.close()
        if all_chapters:
            split=(len(all_chapters)+1)//2
            chapters=all_chapters[:split] if request.form.get('half_choice','first')=='first' else all_chapters[split:]
            if chapters:
                clauses+=[f"q.chapter IN ({','.join('?' for _ in chapters)})"]
                params+=chapters

    if not chapters:
        chapters=[x.strip() for x in request.form.get('chapters_text','').split(',') if x.strip()]
    if scope=='multi' and chapters:
        clauses+=[f"q.chapter IN ({','.join('?' for _ in chapters)})"]
        params+=chapters

    sql=f"""SELECT q.*,
    CASE WHEN EXISTS(
      SELECT 1 FROM attempt_answers aa
      JOIN attempts a ON a.id=aa.attempt_id
      WHERE aa.question_db_id=q.id AND a.student_id=?
    ) THEN 1 ELSE 0 END seen_q,
    CASE WHEN q.family_id<>'' AND EXISTS(
      SELECT 1 FROM attempt_answers aa2
      JOIN attempts a2 ON a2.id=aa2.attempt_id
      JOIN questions q2 ON q2.id=aa2.question_db_id
      WHERE a2.student_id=? AND q2.family_id=q.family_id
    ) THEN 1 ELSE 0 END seen_family
    FROM questions q
    WHERE {' AND '.join(clauses)}
    ORDER BY seen_q,seen_family,RANDOM()
    LIMIT ?"""

    c=db()
    rows=c.execute(sql,[session['user_id'],session['user_id']]+params+[max(count*12,count)]).fetchall()
    access=get_access_profile(c,session['user_id'])
    rows=[r for r in rows if canonical_question_type(r) in LIVE_MARKABLE_TYPES and mastery_rank(r['level'] or 'Foundation')<=mastery_rank(access['mastery_ceiling'])]
    governing_blueprint=active_assessment_blueprint(c,programme) if programme else None
    assessment_type='subject_mock' if mode=='mock' else ('half_syllabus_practice' if scope=='half' else 'practice')
    assembly_policy=active_assembly_policy(c,
      governing_blueprint['id'] if governing_blueprint else None,
      governing_blueprint['framework_version_id'] if governing_blueprint else None,
      programme=programme,subject=subject,chapter=request.form.get('chapter',''),assessment_type=assessment_type)
    rigor=int(assembly_policy['rigor_score'] or 50) if assembly_policy else 50
    official_mix=safe_json(governing_blueprint['difficulty_distribution_json'],{}) if governing_blueprint else {}
    selected,selection_detail=_select_subject_questions(rows,min(count,len(rows)),rigor_mix(rigor,official_mix),
      seed=f"student-test|{session['user_id']}|{datetime.now().isoformat()}") if rows else ([],{})
    rows=selected
    if not rows:
        c.close()
        flash('No matching questions yet.','error')
        return redirect(url_for('test_setup'))

    meta={
      'scope':scope,'programme':programme,'subject':subject,
      'chapters':','.join(chapters),'topic':request.form.get('topic',''),
      'subtopic':request.form.get('subtopic',''),'level':level,
      'assessment_kind':assessment_type,'authenticity_status':'BLUEPRINT_AWARE_PRACTICE' if governing_blueprint else 'STANDARD_PRACTICE',
      'assessment_blueprint_id':governing_blueprint['id'] if governing_blueprint else None,
      'blueprint_source_id':governing_blueprint['powerhouse_blueprint_id'] if governing_blueprint else '',
      'blueprint_version':governing_blueprint['blueprint_version'] if governing_blueprint else '',
      'framework_version':governing_blueprint['framework_version_name'] if governing_blueprint else '',
      'blueprint_snapshot':blueprint_payload_from_record(c,governing_blueprint['id']) if governing_blueprint else {},
      'assembly_policy_id':assembly_policy['id'] if assembly_policy else None,
      'assembly_policy_version':assembly_policy['policy_version'] if assembly_policy else '1',
      'assessment_rigor_score':rigor,'selection_detail':selection_detail
    }
    assessment_id=create_assessment_session(
        c,session['user_id'],mode,duration,[r['id'] for r in rows],meta
    )
    c.close()
    return redirect(url_for('take_test_v4',assessment_id=assessment_id))



@app.route('/test/session/<int:assessment_id>',methods=['GET','POST'])
def take_test_v4(assessment_id):
    if not require('student'):
        return redirect(url_for('login'))

    c=db()
    a=get_assessment_session(c,assessment_id,session['user_id'])
    if not a:
        c.close()
        flash('Assessment session not found.','error')
        return redirect(url_for('test_setup'))

    if a['status']!='in_progress':
        c.close()
        return redirect(url_for('student_dashboard'))

    remain=seconds_left(a)
    if remain==0 and a['mode'] in ('exam','mock'):
        c.close()
        return redirect(url_for('assessment_review_v4',assessment_id=assessment_id))

    ids=parse_ids(a['question_ids'])
    idx=max(0,min(int(a['current_index'] or 0),len(ids)-1))
    answers=load_answers(a)
    confidence=load_json_map(a,'confidence_json')
    response_times=load_json_map(a,'response_times_json')
    flagged=set(parse_ids(a['flagged_ids']))

    if request.method=='POST':
        qid=ids[idx]
        answer_values=request.form.getlist('answer')
        answer=','.join(answer_values) if len(answer_values)>1 else (answer_values[0] if answer_values else '')
        action=request.form.get('action','next')
        conf=request.form.get('confidence','').strip()
        try: elapsed=max(0,int(request.form.get('response_time_seconds','0') or 0))
        except ValueError: elapsed=0

        if answer!='':
            answers[str(qid)]=answer
        if conf: confidence[str(qid)]=conf
        if elapsed: response_times[str(qid)]=max(int(response_times.get(str(qid),0) or 0),elapsed)

        if request.form.get('flagged')=='1':
            flagged.add(qid)
        else:
            flagged.discard(qid)

        if action=='previous':
            idx=max(0,idx-1)
        elif action=='next':
            idx=min(len(ids)-1,idx+1)
        elif action.startswith('jump:'):
            try:
                idx=max(0,min(len(ids)-1,int(action.split(':',1)[1])))
            except Exception:
                pass
        elif action=='review':
            save_assessment(c,assessment_id,idx,answers,flagged,confidence=confidence,response_times=response_times)
            c.close()
            return redirect(url_for('assessment_review_v4',assessment_id=assessment_id))

        save_assessment(c,assessment_id,idx,answers,flagged,confidence=confidence,response_times=response_times)
        a=get_assessment_session(c,assessment_id,session['user_id'])
        remain=seconds_left(a)

    q=c.execute('SELECT * FROM questions WHERE id=?',(ids[idx],)).fetchone()
    q=integration_v1.pinned_question(a,ids[idx],q)
    exam_meta=safe_json(a['meta_json'],{})
    qtype=canonical_question_type(q)
    answer_cfg=safe_json(q['answer_config'], {})
    marking_cfg=safe_json(q['marking_config'], {})
    options=answer_cfg.get('options') or [
        {'id':code,'text':q[key]} for code,key in [('A','option_a'),('B','option_b'),('C','option_c'),('D','option_d')] if q[key]
    ]
    c.close()
    answered_count=sum(1 for qid in ids if answers.get(str(qid),'')!='')

    return render_template(
        'take_test_v4.html',
        assessment=a,assessment_id=assessment_id,q=q,idx=idx,total=len(ids),
        ids=ids,answers=answers,flagged=flagged,
        remaining_seconds=remain,answered_count=answered_count,
        qtype=qtype,options=options,answer_cfg=answer_cfg,marking_cfg=marking_cfg,confidence=confidence,response_times=response_times,
        exam_meta=exam_meta
    )

@app.route('/test/session/<int:assessment_id>/review')
def assessment_review_v4(assessment_id):
    if not require('student'):
        return redirect(url_for('login'))

    c=db()
    a=get_assessment_session(c,assessment_id,session['user_id'])
    if not a:
        c.close()
        return redirect(url_for('student_dashboard'))

    ids=parse_ids(a['question_ids'])
    answers=load_answers(a)
    flagged=set(parse_ids(a['flagged_ids']))
    unanswered=[qid for qid in ids if answers.get(str(qid),'')=='']
    c.close()

    return render_template(
        'assessment_review_v4.html',
        assessment=a,assessment_id=assessment_id,total=len(ids),
        answered=len(ids)-len(unanswered),unanswered=unanswered,flagged=flagged
    )

@app.route('/test/session/<int:assessment_id>/submit',methods=['POST'])
def submit_assessment_v4(assessment_id):
    if not require('student'):
        return redirect(url_for('login'))

    c=db()
    # Atomic idempotency claim. BEGIN IMMEDIATE serialises competing submit writers on
    # SQLite; the unique attempts.assessment_session_id index is the second safety layer.
    c.execute('BEGIN IMMEDIATE')
    a=get_assessment_session(c,assessment_id,session['user_id'])
    if not a:
        c.rollback(); c.close()
        return redirect(url_for('student_dashboard'))
    if (a['status'] or '')!='in_progress':
        existing_attempt=(a['submitted_attempt_id'] if 'submitted_attempt_id' in a.keys() else None)
        if not existing_attempt:
            row=c.execute("SELECT id FROM attempts WHERE student_id=? AND assessment_session_id=? ORDER BY id LIMIT 1",
                          (session['user_id'],assessment_id)).fetchone()
            existing_attempt=row['id'] if row else None
        c.rollback(); c.close()
        if existing_attempt:
            return redirect(url_for('result',aid=int(existing_attempt)))
        flash('This assessment has already been submitted.','info')
        return redirect(url_for('student_dashboard'))
    claimed=c.execute("UPDATE assessment_sessions SET status='submitting' WHERE id=? AND student_id=? AND status='in_progress'",
                      (assessment_id,session['user_id']))
    if claimed.rowcount!=1:
        row=c.execute("SELECT id FROM attempts WHERE student_id=? AND assessment_session_id=? ORDER BY id LIMIT 1",
                      (session['user_id'],assessment_id)).fetchone()
        c.rollback(); c.close()
        if row:
            return redirect(url_for('result',aid=int(row['id'])))
        flash('This assessment is already being submitted. Please open your results in a moment.','info')
        return redirect(url_for('student_dashboard'))

    ids=parse_ids(a['question_ids'])
    answers=load_answers(a)
    confidence=load_json_map(a,'confidence_json')
    response_times=load_json_map(a,'response_times_json')
    qs=','.join('?' for _ in ids)
    rows=c.execute(f'SELECT * FROM questions WHERE id IN ({qs})',ids).fetchall()
    byid={r['id']:r for r in rows}

    correct=0
    blueprint_snapshot=safe_json(a['blueprint_snapshot_json'],{}) if 'blueprint_snapshot_json' in a.keys() else {}
    blueprint_marking_rules=(blueprint_snapshot.get('marking_rules') or {}) if isinstance(blueprint_snapshot,dict) else {}
    total_awarded_marks=0.0
    answer_rows=[]
    for qid in ids:
        current_q=byid[qid]
        q=integration_v1.pinned_question(a,qid,current_q)
        pin=integration_v1.answer_pin(a,qid)
        selected=answers.get(str(qid),'').strip()
        ok,marks_awarded,misconception=mark_question_response(q,selected,blueprint_marking_rules)
        correct+=int(ok); total_awarded_marks+=float(marks_awarded or 0)
        answer_rows.append((qid,selected,int(ok),marks_awarded,int(q['question_version'] or 1),misconception,
          confidence.get(str(qid),''),int(response_times.get(str(qid),0) or 0),
          pin.get('question_id',''),pin.get('question_version_id',''),pin.get('question_checksum_sha256',''),
          pin.get('release_id',''),pin.get('release_version',''),pin.get('release_checksum_sha256',''),
          integration_v1.canonical_json(pin.get('projection') or {})))

    meta=json.loads(a['meta_json'] or '{}')
    if blueprint_marking_rules and blueprint_marking_rules.get('correct_marks') is not None:
        possible=float(blueprint_marking_rules.get('correct_marks') or 0)*len(ids)
        score=round(max(0.0,100.0*total_awarded_marks/possible),1) if possible>0 else 0.0
    else:
        score=round(correct/len(ids)*100,1)

    cur=c.execute(
        """INSERT INTO attempts(
            student_id,scope,programme,subject,chapters,topic,subtopic,level,
            score,correct_count,total_count,assessment_kind,recovery_parent_attempt_id,recovery_focus_type,recovery_focus_name,
            exam_paper_id,exam_paper_code,exam_title,guided_mode,assessment_blueprint_id,blueprint_source_id,
            blueprint_version,framework_version,blueprint_snapshot_json,assembly_policy_id,assembly_policy_version,assessment_session_id,
            ph_release_pins_json,ph_question_pins_json
        ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (
            session['user_id'],meta.get('scope',''),meta.get('programme',''),
            meta.get('subject',''),meta.get('chapters',''),meta.get('topic',''),
            meta.get('subtopic',''),meta.get('level',''),
            score,correct,len(ids),meta.get('assessment_kind','standard'),meta.get('recovery_parent_attempt_id'),
            meta.get('recovery_focus_type',''),meta.get('recovery_focus_name',''),
            meta.get('exam_paper_id'),meta.get('exam_paper_code',''),meta.get('exam_title',''),1 if meta.get('guided_mode') else 0,
            meta.get('assessment_blueprint_id'),meta.get('blueprint_source_id',''),meta.get('blueprint_version',''),
            meta.get('framework_version',''),json.dumps(meta.get('blueprint_snapshot') or {}),meta.get('assembly_policy_id'),
            meta.get('assembly_policy_version',''),assessment_id,a['ph_release_pins_json'] if 'ph_release_pins_json' in a.keys() else '{}',
            a['ph_question_pins_json'] if 'ph_question_pins_json' in a.keys() else '{}'
        )
    )
    attempt_id=cur.lastrowid

    c.executemany(
        '''INSERT INTO attempt_answers(
            attempt_id,question_db_id,selected_answer,is_correct,marks_awarded,question_version,misconception_triggered,confidence,response_time_seconds,
            ph_question_id,ph_question_version_id,ph_question_checksum_sha256,ph_release_id,ph_release_version,ph_release_checksum_sha256,ph_question_snapshot_json
        ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
        [(attempt_id,*r) for r in answer_rows]
    )
    # Commit the one-and-only attempt identity together with the submitted session state.
    # save_assessment performs the commit, making later retries deterministic redirects.
    c.execute("UPDATE assessment_sessions SET submitted_attempt_id=? WHERE id=? AND student_id=?",
              (attempt_id,assessment_id,session['user_id']))
    save_assessment(c,assessment_id,status='submitted')
    update_learning_intelligence_from_attempt(c,attempt_id,session['user_id'])
    _refresh_item_calibration(c,ids)
    process_mastery_result(c,a,attempt_id,score,ids)
    # Shadow/PILOT universal evidence capture. Unmapped legacy questions are skipped rather
    # than assigned invented academic identities; legacy mastery remains authoritative.
    universal_mastery.capture_scoremax_attempt(c,attempt_id=attempt_id,assessment_session_id=assessment_id,
      student_id=session['user_id'],meta=meta)
    apply_attempt_to_plan_activity(c,attempt_id,session['user_id'],meta)
    # Important evidence can immediately influence the plan; weekly rebalance remains the main consolidation point.
    if meta.get('plan_activity_id') or meta.get('assessment_kind') in ('mock','recovery','recall'):
        rebalance_study_plan(c,session['user_id'])
    assignment_id=meta.get('assignment_id')
    if assignment_id:
        c.execute("""UPDATE assignment_students SET status='completed',attempt_id=?,completed_at=CURRENT_TIMESTAMP
          WHERE assignment_id=? AND student_id=?""",(attempt_id,assignment_id,session['user_id']))
    challenge_id=meta.get('challenge_id')
    if challenge_id:
        elapsed=sum(int(v or 0) for v in response_times.values())
        c.execute("""UPDATE challenge_entries SET status='completed',attempt_id=?,completed_at=?,score=?,correct_count=?,total_count=?,elapsed_seconds=?
          WHERE challenge_id=? AND student_id=?""",(attempt_id,datetime.now().isoformat(timespec='seconds'),score,correct,len(ids),elapsed,challenge_id,session['user_id']))
        refresh_challenge_rankings(c,int(challenge_id))
    c.commit()
    c.close()
    return redirect(url_for('result',aid=attempt_id))

@app.route('/recovery/start/<int:aid>',methods=['POST'])
def start_recovery(aid):
    if not require('student'):
        return redirect(url_for('login'))
    c=db()
    if not feature_available(c,session['user_id'],'recovery_engine'):
        c.close(); flash('Personalised weak-area practice requires a higher ScoreMax Access level.','error'); return redirect(url_for('access_account'))
    diagnostic=attempt_diagnostics(c,aid,session['user_id'])
    if not diagnostic:
        c.close(); flash('Your earlier test could not be found.','error'); return redirect(url_for('student_dashboard'))
    focus,ids=recovery_question_ids(c,session['user_id'],diagnostic,count=8)
    if not focus or len(ids)<3:
        c.close(); flash('There are not enough approved practice questions for this weak area yet.','error'); return redirect(url_for('result',aid=aid))
    original=diagnostic['attempt']
    meta={
      'scope':'recovery','programme':original['programme'] or '', 'subject':original['subject'] or '',
      'chapters':original['chapters'] or '', 'topic':focus['name'] if focus['type']=='topic' else '',
      'subtopic':'', 'level':original['level'] or '', 'assessment_kind':'recovery',
      'recovery_parent_attempt_id':aid,'recovery_focus_type':focus['type'],'recovery_focus_name':focus['name']
    }
    assessment_id=create_assessment_session(c,session['user_id'],'practice',None,ids,meta)
    c.close()
    return redirect(url_for('take_test_v4',assessment_id=assessment_id))

@app.route('/test/take',methods=['GET','POST'])
def take_test():
    # V5.3 retires the legacy scoring path so all live tests use the validated renderer/marker.
    if not require('student'): return redirect(url_for('login'))
    flash('This test route has moved to the current ScoreMax assessment engine.','success')
    return redirect(url_for('test_setup'))


@app.route('/result/<int:aid>')
def result(aid):
    if not require('student'): return redirect(url_for('login'))
    c=db()
    d=attempt_diagnostics(c,aid,session['user_id'])
    if not d:
        c.close(); return redirect(url_for('student_dashboard'))
    details=c.execute('''SELECT aa.*,q.question,q.answer,q.explanation,q.level,q.subtopic,q.family_id,q.variant,
      q.learning_outcome,q.concept,q.concept_id,q.capsule_id,q.difficulty,q.cognitive_skill,q.command_word,q.subject,q.chapter
      FROM attempt_answers aa JOIN questions q ON q.id=aa.question_db_id WHERE aa.attempt_id=?''',(aid,)).fetchall()
    details=[integration_v1.overlay_attempt_question(r) for r in details]
    weak_capsules=c.execute("""SELECT DISTINCT lc.* FROM attempt_answers aa JOIN questions q ON q.id=aa.question_db_id
      JOIN learning_capsules lc ON lc.active=1 AND lc.review_status='Approved'
      AND (lc.capsule_id=q.capsule_id OR (lc.concept_id<>'' AND lc.concept_id=q.concept_id) OR lc.concept=q.concept)
      WHERE aa.attempt_id=? AND aa.is_correct=0 LIMIT 5""",(aid,)).fetchall()
    access=get_access_profile(c,session['user_id'])
    advanced_allowed=(not COMMERCIAL_GATES_ENABLED) or bool(access['entitlements'].get('advanced_diagnostics',False))
    blueprint_result=attempt_blueprint_result(c,aid)
    c.close()
    return render_template('result.html',a=d['attempt'],details=details,diagnostic=d,advanced_allowed=advanced_allowed,access=access,weak_capsules=weak_capsules,blueprint_result=blueprint_result)


@app.route('/student/analytics')
def student_analytics_page():
    if not require('student'): return redirect(url_for('login'))
    c=db(); a=student_analytics(c,session['user_id']); c.close(); return render_template('student_analytics.html',**a)

@app.route('/student/ask-ai',methods=['GET','POST'])
def ask_ai():
    if not require('student'): return redirect(url_for('login'))
    answer=None; question=''
    if request.method=='POST':
        question=request.form.get('question','').strip(); c=db(); answer=ai_answer(c,session['user_id'],question); c.close()
    return render_template('ask_ai.html',answer=answer,question=question)

@app.route('/student/progress-email-preview')
def progress_email_preview():
    if not require('student'): return redirect(url_for('login'))
    c=db()
    user=c.execute('SELECT * FROM users WHERE id=?',(session['user_id'],)).fetchone()
    weekly=ensure_weekly_progress_report(c,session['user_id'])
    plan=get_active_study_plan(c,session['user_id'])
    charts=progress_chart_data(c,session['user_id'])
    c.close()
    return render_template('progress_email_preview.html',user=user,weekly=weekly,plan=plan,charts=charts)


def question_snapshot(row):
    keys=['question_id','family_id','variant','programme','subject','chapter','topic','subtopic','qtype','level','question',
          'option_a','option_b','option_c','option_d','answer','explanation','status','country','qualification','exam_board',
          'curriculum_version','learning_outcome','concept','concept_id','capsule_id','misconception_id','difficulty','cognitive_skill','command_word','marks',
          'estimated_time_seconds','stimulus_type','stimulus_data','answer_config','marking_config','feedback_config',
          'misconception_tags','prerequisite_tags','question_version','review_status','reviewer','reviewed_at','source_type',
          'secure_bank','language','translation_group','active']
    return {k: row[k] for k in keys if k in row.keys()}

def question_health(c, question_id):
    row=c.execute('''SELECT q.id,q.question_id,q.question,q.subject,q.chapter,q.topic,q.subtopic,q.level,q.review_status,q.active,
        COUNT(aa.id) attempts,
        ROUND(100.0*AVG(CASE WHEN aa.is_correct=1 THEN 1.0 ELSE 0.0 END),1) correct_pct,
        ROUND(AVG(CASE WHEN aa.response_time_seconds>0 THEN aa.response_time_seconds END),1) avg_seconds
        FROM questions q LEFT JOIN attempt_answers aa ON aa.question_db_id=q.id
        WHERE q.id=? GROUP BY q.id''',(question_id,)).fetchone()
    if not row: return None
    reports=c.execute("SELECT COUNT(*) n FROM question_review_events WHERE question_id=? AND action='Student Report'",(question_id,)).fetchone()['n']
    attempts=row['attempts'] or 0; correct=row['correct_pct']
    flags=[]
    if reports>=3: flags.append(f'{reports} reports')
    if attempts>=20 and correct is not None and correct<20: flags.append('Unusually low success rate')
    if attempts>=20 and correct is not None and correct>95: flags.append('Possibly too easy')
    if row['avg_seconds'] and row['avg_seconds']>180: flags.append('Long response time')
    status='Review recommended' if flags else ('Collecting data' if attempts<20 else 'Healthy')
    return {'row':row,'reports':reports,'flags':flags,'status':status}

def validate_import_row(r, row_number):
    errors=[]; warnings=[]
    qid=str(r.get('Question ID','') or '').strip()
    family_id=str(r.get('Family ID','') or '').strip()
    qtype=str(r.get('Type','') or '').strip()
    question=str(r.get('Question','') or '').strip()
    answer=str(r.get('Answer','') or '').strip()
    subject=str(r.get('Subject','') or '').strip()
    chapter=str(r.get('Chapter','') or '').strip()
    level=str(r.get('Level','') or '').strip()
    difficulty=str(r.get('Difficulty','') or '').strip()
    allowed={'MCQ','Single Choice','True/False','Fill Blank','Multiple Select','Numerical','Matching','Ordering','Drag Drop','Short Response','Extended Response','Image Hotspot','Diagram Label'}
    if not qid: errors.append('Missing Question ID')
    if not family_id: errors.append('Missing Family ID (required for governed question-family/variant integrity)')
    if not question: errors.append('Missing question text')
    if not subject: errors.append('Missing subject')
    if not chapter: warnings.append('Missing chapter')
    if not qtype: errors.append('Missing question type')
    elif qtype not in allowed: warnings.append(f'Unrecognised type: {qtype}')
    if qtype in {'MCQ','Single Choice','True/False','Fill Blank','Multiple Select','Numerical'} and not answer:
        errors.append('Missing answer')
    if not str(r.get('Explanation','') or '').strip(): warnings.append('Missing explanation')
    if not str(r.get('Learning Outcome','') or '').strip(): warnings.append('Missing learning outcome')
    if not str(r.get('Concept','') or '').strip(): warnings.append('Missing concept')
    if not str(r.get('Curriculum Version','') or '').strip(): warnings.append('Missing curriculum version')
    if str(r.get('Status','') or '').strip().lower()=='approved' or str(r.get('Review Status','') or '').strip().lower()=='approved':
        warnings.append('Spreadsheet approval fields are ignored; V5.5 always imports as Draft + inactive')
    if not level: warnings.append('Missing mastery level')
    if not difficulty: errors.append('Missing Difficulty (required separately from Mastery Level)')
    rights=str(r.get('Rights Status','') or '').strip()
    ready=str(r.get('ScoreMax Ready','') or '').strip().lower()
    if not rights: errors.append('Missing Rights Status (e.g. ScoreMax Original, Licensed, Permitted, Public Domain)')
    elif rights.lower() not in {'scoremax original','licensed','permitted','public domain','approved','permission pending','unknown'}:
        warnings.append(f'Unrecognised Rights Status: {rights}')
    if ready not in {'yes','true','1','no','false','0'}: errors.append('ScoreMax Ready must be Yes or No')
    if not str(r.get('Assessment Purpose','') or '').strip(): warnings.append('Missing Assessment Purpose; defaults to practice|test|mock|mastery')
    return {'row_number':row_number,'question_id':qid or '(blank)','errors':errors,'warnings':warnings,'row':r}


@app.route('/referrals')
def referral_account():
    if not require(): return redirect(url_for('login'))
    c=db()
    release_due_referral_rewards(c,session['user_id'])
    code=ensure_referral_code(c,session['user_id'])
    user=c.execute("SELECT id,role,full_name,own_referral_code FROM users WHERE id=?",(session['user_id'],)).fetchone()
    balances=c.execute("SELECT currency,balance_minor FROM wallet_balances WHERE user_id=? ORDER BY currency",(session['user_id'],)).fetchall()
    stats=dict(c.execute("""SELECT COUNT(DISTINCT ra.user_id) signups,
      COUNT(DISTINCT CASE WHEN rr.id IS NOT NULL THEN ra.user_id END) paid_referrals
      FROM referral_attributions ra LEFT JOIN referral_rewards rr ON rr.referred_user_id=ra.user_id
      WHERE ra.referrer_type='user' AND ra.referrer_id=?""",(session['user_id'],)).fetchone())
    stats['recruited_teachers']=c.execute("""SELECT COUNT(*) n FROM referral_attributions ra JOIN users u ON u.id=ra.user_id
      WHERE ra.referrer_type='user' AND ra.referrer_id=? AND ra.attribution_kind='TEACHER_RECRUITMENT' AND u.role='teacher'""",(session['user_id'],)).fetchone()['n']
    stats['downstream_paying_students']=c.execute("SELECT COUNT(DISTINCT referred_user_id) n FROM referral_rewards WHERE override_referrer_user_id=?",(session['user_id'],)).fetchone()['n']
    money=c.execute("""SELECT currency,
      COALESCE(SUM(CASE WHEN referrer_user_id=? AND status<>'reversed' THEN reward_amount_minor ELSE 0 END),0) direct_minor,
      COALESCE(SUM(CASE WHEN override_referrer_user_id=? AND override_status<>'reversed' THEN override_reward_amount_minor ELSE 0 END),0) override_minor
      FROM referral_rewards WHERE referrer_user_id=? OR override_referrer_user_id=? GROUP BY currency ORDER BY currency""",
      (session['user_id'],session['user_id'],session['user_id'],session['user_id'])).fetchall()
    rewards=c.execute("""SELECT rr.*,u.full_name referred_name,ref.full_name direct_referrer_name
      FROM referral_rewards rr JOIN users u ON u.id=rr.referred_user_id JOIN users ref ON ref.id=rr.referrer_user_id
      WHERE rr.referrer_user_id=? OR rr.override_referrer_user_id=? ORDER BY rr.id DESC LIMIT 100""",
      (session['user_id'],session['user_id'])).fetchall()
    wallet=c.execute("SELECT * FROM wallet_transactions WHERE user_id=? ORDER BY id DESC LIMIT 30",(session['user_id'],)).fetchall()
    c.close()
    base=request.host_url.rstrip('/')
    referral_link=base+url_for('register',role='student')+f'?ref={code}&source=teacher_referral'
    teacher_referral_link=base+url_for('register',role='teacher')+f'?ref={code}&source=teacher_referral'
    return render_template('referrals.html',user=user,stats=stats,rewards=rewards,wallet=wallet,balances=balances,
      referral_link=referral_link,teacher_referral_link=teacher_referral_link,reward_money=money)


@app.route('/admin/referrals',methods=['GET','POST'])
def admin_referrals():
    if not require('admin'): return redirect(url_for('login'))
    c=db()
    release_due_referral_rewards(c)
    if request.method=='POST':
        action=request.form.get('action','')
        if action=='programme':
            role_group=request.form.get('role_group','student')
            if role_group not in ('student','partner','teacher_direct','teacher_override'): role_group='student'
            rate=max(0,min(1,float(request.form.get('reward_rate') or 0)))
            hold=max(0,min(90,int(request.form.get('hold_days') or 14)))
            reward_type='wallet_credit' if role_group=='student' else 'commission'
            existing_rule=c.execute("SELECT programme_version FROM referral_programs WHERE role_group=?",(role_group,)).fetchone()
            version=int(existing_rule['programme_version'] or 1)+1 if existing_rule else 1
            c.execute("""INSERT INTO referral_programs(role_group,reward_type,reward_rate,hold_days,active,updated_at,programme_version)
              VALUES(?,?,?,?,1,?,?) ON CONFLICT(role_group) DO UPDATE SET reward_type=excluded.reward_type,
              reward_rate=excluded.reward_rate,hold_days=excluded.hold_days,active=1,updated_at=excluded.updated_at,programme_version=excluded.programme_version""",
              (role_group,reward_type,rate,hold,datetime.now().isoformat(timespec='seconds'),version))
            c.commit(); c.close(); flash('Referral programme settings updated.','success')
            return redirect(url_for('admin_referrals'))
        reward_id=int(request.form.get('reward_id') or 0)
        r=c.execute("SELECT * FROM referral_rewards WHERE id=?",(reward_id,)).fetchone()
        if r:
            now=datetime.now().isoformat(timespec='seconds')
            if action=='reverse' and r['status'] not in ('paid','reversed'):
                reverse_referral_reward(c,reward_id,'Admin reversal')
                flash('Referral reward reversed and any released wallet credit adjusted.','success')
            elif r['reward_type']=='commission' and action=='approve':
                if r['status']=='pending' and r['available_at']<=now:
                    c.execute("UPDATE referral_rewards SET status='approved' WHERE id=?",(reward_id,))
                    c.commit(); flash('Commission approved.','success')
                else:
                    flash('Commission is still within its hold period or is not pending.','error')
            elif r['reward_type']=='commission' and action=='paid' and r['status'] in ('approved','available'):
                c.execute("UPDATE referral_rewards SET status='paid',paid_at=? WHERE id=?",(now,reward_id))
                c.commit(); flash('Commission marked paid.','success')
            elif action=='approve_override' and int(r['override_reward_amount_minor'] or 0)>0 and r['override_status']=='pending' and (r['override_available_at'] or '')<=now:
                c.execute("UPDATE referral_rewards SET override_status='approved' WHERE id=?",(reward_id,)); c.commit(); flash('Teacher-recruitment override approved.','success')
            elif action=='paid_override' and int(r['override_reward_amount_minor'] or 0)>0 and r['override_status'] in ('approved','available'):
                c.execute("UPDATE referral_rewards SET override_status='paid',override_paid_at=? WHERE id=?",(now,reward_id)); c.commit(); flash('Teacher-recruitment override marked paid.','success')
            elif action=='reverse_override' and int(r['override_reward_amount_minor'] or 0)>0 and r['override_status'] not in ('paid','reversed'):
                c.execute("UPDATE referral_rewards SET override_status='reversed',override_reversed_at=?,override_notes=TRIM(COALESCE(override_notes,'') || ' | Admin reversal') WHERE id=?",(now,reward_id)); c.commit(); flash('Teacher-recruitment override reversed.','success')
        c.close(); return redirect(url_for('admin_referrals'))
    summary=c.execute("""SELECT COUNT(DISTINCT ra.user_id) attributed_signups,
      COUNT(DISTINCT rr.id) qualifying_rewards
      FROM referral_attributions ra LEFT JOIN referral_rewards rr ON rr.referred_user_id=ra.user_id
      WHERE ra.referrer_type='user'""").fetchone()
    reward_totals=c.execute("""SELECT currency,
      COALESCE(SUM(CASE WHEN status IN ('available','approved','paid') THEN reward_amount_minor ELSE 0 END),0) released_minor
      FROM referral_rewards GROUP BY currency ORDER BY currency""").fetchall()
    rewards=c.execute("""SELECT rr.*,ref.full_name referrer_name,ref.role referrer_role,new.full_name referred_name,
      over.full_name override_referrer_name,pt.provider_transaction_ref,pt.gross_amount_minor,pt.net_amount_minor,pt.paid_at,
      COALESCE(cp.name,p.name,'') package_name,COALESCE(cp.programme,new.academic_level,'') package_programme
      FROM referral_rewards rr
      JOIN users ref ON ref.id=rr.referrer_user_id JOIN users new ON new.id=rr.referred_user_id
      JOIN payment_transactions pt ON pt.id=rr.payment_transaction_id
      LEFT JOIN users over ON over.id=rr.override_referrer_user_id
      LEFT JOIN plans p ON p.id=pt.plan_id LEFT JOIN subscriptions sub ON sub.id=pt.subscription_id
      LEFT JOIN coverage_packages cp ON cp.id=sub.coverage_package_id
      ORDER BY rr.id DESC LIMIT 100""").fetchall()
    programmes=c.execute("SELECT * FROM referral_programs ORDER BY role_group").fetchall()
    leaders=c.execute("""SELECT u.id,u.full_name,u.role,u.own_referral_code,
      COUNT(DISTINCT ra.user_id) signups,COUNT(DISTINCT rr.id) paid_conversions,
      COALESCE(SUM(rr.reward_amount_minor),0) rewards_minor
      FROM users u LEFT JOIN referral_attributions ra ON ra.referrer_type='user' AND ra.referrer_id=u.id
      LEFT JOIN referral_rewards rr ON rr.referrer_user_id=u.id
      WHERE COALESCE(u.own_referral_code,'')<>'' GROUP BY u.id
      HAVING signups>0 ORDER BY paid_conversions DESC,signups DESC LIMIT 50""").fetchall()
    c.close()
    return render_template('admin_referrals.html',summary=summary,reward_totals=reward_totals,rewards=rewards,leaders=leaders,programmes=programmes)


@app.route('/admin/referrals/export.xlsx')
def admin_referrals_export():
    if not require('admin'): return redirect(url_for('login'))
    month=(request.args.get('month') or datetime.now().strftime('%Y-%m')).strip()
    if not re.fullmatch(r'\d{4}-\d{2}',month): abort(400,description='month must be YYYY-MM')
    c=db()
    rows=c.execute("""SELECT rr.*,ref.full_name referrer_name,ref.role referrer_role,ref.own_referral_code referrer_code,
      stu.full_name student_name,stu.system_user_id student_system_id,stu.academic_level student_programme,
      over.full_name override_referrer_name,over.own_referral_code override_referrer_code,
      pt.gross_amount_minor,pt.net_amount_minor,pt.refund_amount_minor,pt.paid_at,pt.currency,
      COALESCE(cp.name,p.name,'') package_name,COALESCE(cp.programme,stu.academic_level,'') package_programme
      FROM referral_rewards rr JOIN users ref ON ref.id=rr.referrer_user_id JOIN users stu ON stu.id=rr.referred_user_id
      JOIN payment_transactions pt ON pt.id=rr.payment_transaction_id
      LEFT JOIN users over ON over.id=rr.override_referrer_user_id
      LEFT JOIN plans p ON p.id=pt.plan_id LEFT JOIN subscriptions sub ON sub.id=pt.subscription_id
      LEFT JOIN coverage_packages cp ON cp.id=sub.coverage_package_id
      WHERE substr(COALESCE(pt.paid_at,pt.created_at),1,7)=?
        AND (ref.role='teacher' OR rr.override_referrer_user_id IS NOT NULL)
      ORDER BY ref.full_name,pt.paid_at,rr.id""",(month,)).fetchall()
    def _excel_safe_text(value):
        s='' if value is None else str(value)
        return "'"+s if s[:1] in {'=','+','-','@'} else s
    wb=Workbook(); ws=wb.active; ws.title='Teacher Monthly Summary'
    ws.append(['Teacher','Referral Code','Direct Paying Students','Downstream Paying Students','Total Paying Students',
      'Direct Gross Sales','Direct Eligible Sales','Network Eligible Sales','Direct Commission','Teacher Referral Override','Total Due','Paid','Balance','Currency'])
    summary={}
    def bucket(user_id,currency,name,code):
        return summary.setdefault((user_id,currency),{'name':name,'code':code,'direct_students':set(),'network_students':set(),
          'gross':0,'eligible':0,'network_eligible':0,'direct':0,'override':0,'paid':0})
    for r in rows:
        d=bucket(r['referrer_user_id'],r['currency'],r['referrer_name'],r['referrer_code'])
        d['direct_students'].add(r['referred_user_id']); d['gross']+=int(r['gross_amount_minor'] or 0); d['eligible']+=int(r['qualifying_amount_minor'] or 0); d['direct']+=int(r['reward_amount_minor'] or 0)
        if r['status']=='paid': d['paid']+=int(r['reward_amount_minor'] or 0)
        if r['override_referrer_user_id']:
            o=bucket(r['override_referrer_user_id'],r['currency'],r['override_referrer_name'],r['override_referrer_code'])
            o['network_students'].add(r['referred_user_id']); o['network_eligible']+=int(r['qualifying_amount_minor'] or 0); o['override']+=int(r['override_reward_amount_minor'] or 0)
            if r['override_status']=='paid': o['paid']+=int(r['override_reward_amount_minor'] or 0)
    for (_,currency),d in sorted(summary.items(),key=lambda kv:(kv[1]['name'] or '')):
        all_students=d['direct_students']|d['network_students']; total=d['direct']+d['override']; balance=total-d['paid']
        ws.append([_excel_safe_text(d['name']),_excel_safe_text(d['code']),len(d['direct_students']),len(d['network_students']),len(all_students),
          round(d['gross']/100,2),round(d['eligible']/100,2),round(d['network_eligible']/100,2),round(d['direct']/100,2),round(d['override']/100,2),round(total/100,2),round(d['paid']/100,2),round(balance/100,2),_excel_safe_text(currency)])

    detail=wb.create_sheet('Student Referral Detail')
    detail.append(['Teacher','Referral Code','Student ID','Student','Package','Programme','Payment Date','Gross Amount','Eligible Amount','Direct Rate','Direct Reward','Direct Status',
      'Original Teacher Override','Override Referral Code','Override Rate','Override Reward','Override Status','Refund','Currency'])
    for r in rows:
        detail.append([_excel_safe_text(r['referrer_name']),_excel_safe_text(r['referrer_code']),_excel_safe_text(r['student_system_id']),_excel_safe_text(r['student_name']),
          _excel_safe_text(r['package_name']),_excel_safe_text(r['package_programme']),_excel_safe_text(r['paid_at']),round(int(r['gross_amount_minor'] or 0)/100,2),
          round(int(r['qualifying_amount_minor'] or 0)/100,2),float(r['reward_rate'] or 0),round(int(r['reward_amount_minor'] or 0)/100,2),_excel_safe_text(r['status']),
          _excel_safe_text(r['override_referrer_name']),_excel_safe_text(r['override_referrer_code']),float(r['override_reward_rate'] or 0),
          round(int(r['override_reward_amount_minor'] or 0)/100,2),_excel_safe_text(r['override_status']),round(int(r['refund_amount_minor'] or 0)/100,2),_excel_safe_text(r['currency'])])

    over=wb.create_sheet('Teacher-to-Teacher Rewards')
    over.append(['Original Teacher','Referral Code','Recruited Teacher','Paying Students Generated','Eligible Sales','Override Rate','Override Reward','Status','Currency'])
    groups={}
    for r in rows:
        if not r['override_referrer_user_id']: continue
        key=(r['override_referrer_user_id'],r['referrer_user_id'],r['currency'],float(r['override_reward_rate'] or 0),r['override_status'])
        g=groups.setdefault(key,{'a':r['override_referrer_name'],'code':r['override_referrer_code'],'b':r['referrer_name'],'students':set(),'eligible':0,'reward':0})
        g['students'].add(r['referred_user_id']); g['eligible']+=int(r['qualifying_amount_minor'] or 0); g['reward']+=int(r['override_reward_amount_minor'] or 0)
    for (_,_,currency,rate,status),g in groups.items():
        over.append([_excel_safe_text(g['a']),_excel_safe_text(g['code']),_excel_safe_text(g['b']),len(g['students']),round(g['eligible']/100,2),rate,round(g['reward']/100,2),_excel_safe_text(status),_excel_safe_text(currency)])
    for sheet in wb.worksheets:
        sheet.freeze_panes='A2'; sheet.auto_filter.ref=sheet.dimensions
        for col in sheet.columns:
            letter=col[0].column_letter; sheet.column_dimensions[letter].width=min(38,max(12,max(len(str(x.value or '')) for x in col)+2))
    out=io.BytesIO(); wb.save(out); out.seek(0); c.close()
    return send_file(out,as_attachment=True,download_name=f'ScoreMax_Teacher_Referral_Rewards_{month}.xlsx',mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/account/subscription')
def subscription_account():
    if not require(): return redirect(url_for('login'))
    if session.get('role')=='student': return redirect(url_for('access_account'))
    c=db(); access=get_access_profile(c,session['user_id'])
    audience='teacher' if session.get('role')=='teacher' else 'student'
    plans=c.execute('SELECT * FROM plans WHERE active=1 AND audience=? ORDER BY sort_order',(audience,)).fetchall()
    tx=c.execute('''SELECT pt.*,p.name plan_name FROM payment_transactions pt LEFT JOIN plans p ON p.id=pt.plan_id
                    WHERE pt.user_id=? ORDER BY pt.id DESC LIMIT 20''',(session['user_id'],)).fetchall()
    plan_cards=[]
    for p in plans:
        d=dict(p); d['price_display']=money_display(p['price_minor'],p['currency']); d['entitlements']=safe_json(p['entitlements_json'],{}); plan_cards.append(d)
    c.close(); return render_template('subscription.html',access=access,plans=plan_cards,transactions=tx)

@app.route('/account/subscription/cancel-renewal',methods=['POST'])
def cancel_subscription_renewal():
    if not require(): return redirect(url_for('login'))
    c=db(); sub=c.execute("SELECT * FROM subscriptions WHERE user_id=? AND status IN ('active','trial') ORDER BY id DESC LIMIT 1",(session['user_id'],)).fetchone()
    if sub:
        c.execute("UPDATE subscriptions SET auto_renew=0,cancelled_at=?,cancellation_reason=?,updated_at=? WHERE id=?",
                  (datetime.now().isoformat(timespec='seconds'),'User requested cancellation',datetime.now().isoformat(timespec='seconds'),sub['id']))
        c.commit(); flash('Automatic renewal has been turned off. Access remains until the current plan ends.','success')
    else: flash('No renewable subscription found.','error')
    c.close(); return redirect(url_for('subscription_account'))

@app.route('/admin/payments',methods=['GET','POST'])
def admin_payments():
    if not require('admin'): return redirect(url_for('login'))
    c=db()
    if request.method=='POST':
        action=request.form.get('action','record_payment')
        if action=='coverage_package':
            code=request.form.get('code','').strip().lower().replace(' ','_'); name=request.form.get('name','').strip(); programme=request.form.get('programme','').strip()
            if not code or not name or not programme: raise ValueError('Package code, name and programme are required.')
            coverage_type=request.form.get('coverage_type','SUBJECTS').strip().upper(); subjects=[x.strip() for x in request.form.get('subjects','').split(',') if x.strip()]
            raw=request.form.get('price','').strip(); price_minor=None if raw=='' else int(round(float(raw)*100))
            c.execute('''INSERT INTO coverage_packages(code,name,programme,description,coverage_type,subjects_json,price_minor,currency,billing_period,status,sort_order)
              VALUES(?,?,?,?,?,?,?,?,?,?,?) ON CONFLICT(code) DO UPDATE SET name=excluded.name,programme=excluded.programme,
              description=excluded.description,coverage_type=excluded.coverage_type,subjects_json=excluded.subjects_json,
              price_minor=excluded.price_minor,currency=excluded.currency,billing_period=excluded.billing_period,status=excluded.status,updated_at=CURRENT_TIMESTAMP''',
              (code,name,programme,request.form.get('description','').strip(),coverage_type,json.dumps(subjects),price_minor,
               request.form.get('currency','PKR'),request.form.get('billing_period','monthly'),request.form.get('package_status','ACTIVE'),int(request.form.get('sort_order') or 0)))
            c.commit(); flash('Subject package saved.','success')
        elif action=='assign_package':
            student_id=int(request.form.get('user_id') or 0); package_id=int(request.form.get('coverage_package_id') or 0); access_code=request.form.get('access_plan_code','').strip()
            start=request.form.get('starts_at') or iso_today(); end=request.form.get('ends_at') or ''
            commercial_access.assign_entitlement(c,student_id=student_id,coverage_package_id=package_id,access_plan_code=access_code,
              starts_at=start,ends_at=end,source=request.form.get('source','manual'),notes=request.form.get('notes',''),actor_user_id=session['user_id'])
            c.execute('UPDATE users SET access_override_code=? WHERE id=?',(access_code,student_id))
            c.commit(); flash('Student subject package and Access level activated.','success')
        elif action=='update_plan':
            plan_id=int(request.form.get('plan_id','0') or 0); raw=request.form.get('price','').strip(); currency=request.form.get('currency','PKR').strip() or 'PKR'
            price_minor=None if raw=='' else int(round(float(raw)*100))
            c.execute('UPDATE plans SET price_minor=?,currency=? WHERE id=?',(price_minor,currency,plan_id)); c.commit(); flash('Plan price updated.','success')
        elif action=='promo':
            code=request.form.get('code','').strip().upper(); discount_type=request.form.get('discount_type','percent'); value=int(request.form.get('discount_value','0') or 0)
            try:
                c.execute('''INSERT INTO promo_codes(code,discount_type,discount_value,starts_at,expires_at,usage_limit,active,campaign) VALUES(?,?,?,?,?,?,1,?)''',
                          (code,discount_type,value,request.form.get('starts_at',''),request.form.get('expires_at',''),int(request.form.get('usage_limit') or 0) or None,request.form.get('campaign','')))
                c.commit(); flash('Promo code created.','success')
            except sqlite3.IntegrityError: flash('Promo code already exists.','error')
        elif action=='institution_license':
            institution_id=int(request.form.get('institution_id','0') or 0); seats=max(1,int(request.form.get('seat_count','1') or 1)); start=request.form.get('starts_at') or iso_today(); end=request.form.get('ends_at') or ''
            plan=c.execute("SELECT * FROM plans WHERE code='institution_student'").fetchone()
            cur=c.execute('''INSERT INTO institution_licenses(institution_id,plan_id,seat_count,starts_at,ends_at,status,invoice_ref,payment_status,notes) VALUES(?,?,?,?,?,'active',?,?,?)''',
                          (institution_id,plan['id'],seats,start,end,request.form.get('invoice_ref',''),request.form.get('payment_status','pending'),request.form.get('notes','')))
            lid=cur.lastrowid
            students=c.execute("SELECT id FROM users WHERE role='student' AND primary_institution_id=? ORDER BY id LIMIT ?",(institution_id,seats)).fetchall()
            c.executemany('INSERT OR IGNORE INTO institution_license_users(institution_license_id,user_id) VALUES(?,?)',[(lid,r['id']) for r in students])
            c.commit(); flash(f'Institution licence created; {len(students)} existing student seat(s) allocated.','success')
        else:
            user_id=int(request.form.get('user_id','0') or 0); plan_id=int(request.form.get('plan_id','0') or 0); status=request.form.get('status','successful')
            coverage_package_id=int(request.form.get('coverage_package_id') or 0)
            plan=c.execute('SELECT * FROM plans WHERE id=?',(plan_id,)).fetchone(); raw=request.form.get('amount','').strip(); gross=int(round(float(raw)*100)) if raw else int(plan['price_minor'] or 0)
            txid=record_payment(c,user_id,plan_id,gross,currency=request.form.get('currency') or plan['currency'] or 'PKR',status=status,provider=request.form.get('provider','manual'),provider_ref=request.form.get('provider_ref',''),payment_method=request.form.get('payment_method','manual'),promo_code=request.form.get('promo_code','').strip().upper(),notes=request.form.get('notes',''))
            if status=='successful' and user_id and plan and plan['code'] not in ('free_student','free_teacher','institution_plan'):
                start=iso_today(); period=plan['billing_period']; days=365 if period=='annual' else (30 if period=='monthly' else 365); end=(datetime.now().date()+timedelta(days=days)).isoformat()
                sid=create_subscription(c,user_id,plan_id,start,end,source='manual',provider=request.form.get('provider','manual'),provider_ref=request.form.get('provider_ref',''))
                if coverage_package_id:
                    c.execute('UPDATE subscriptions SET coverage_package_id=? WHERE id=?',(coverage_package_id,sid))
                    commercial_access.assign_entitlement(c,student_id=user_id,coverage_package_id=coverage_package_id,access_plan_code=plan['code'],
                      starts_at=start,ends_at=end,source='payment',notes=request.form.get('notes',''),actor_user_id=session['user_id'])
                c.execute('UPDATE payment_transactions SET subscription_id=? WHERE id=?',(sid,txid))
            c.commit(); flash('Payment record saved.' + (' Subscription activated.' if status=='successful' else ''),'success')
        c.close(); return redirect(url_for('admin_payments'))

    metrics=c.execute('''SELECT COUNT(*) transactions,
        SUM(CASE WHEN status='successful' THEN 1 ELSE 0 END) successful,
        SUM(CASE WHEN status='failed' THEN 1 ELSE 0 END) failed
        FROM payment_transactions''').fetchone()
    revenue_rows=c.execute('''SELECT currency,COALESCE(SUM(net_amount_minor),0) revenue_minor FROM payment_transactions
        WHERE status='successful' GROUP BY currency ORDER BY currency''').fetchall()
    active=c.execute("SELECT COUNT(*) n FROM subscriptions WHERE status IN ('active','trial') AND (ends_at IS NULL OR ends_at='' OR ends_at>=?)",(iso_today(),)).fetchone()['n']
    transactions=c.execute('''SELECT pt.*,u.full_name,u.system_user_id,p.name plan_name FROM payment_transactions pt
        LEFT JOIN users u ON u.id=pt.user_id LEFT JOIN plans p ON p.id=pt.plan_id ORDER BY pt.id DESC LIMIT 200''').fetchall()
    subscriptions=c.execute('''SELECT s.*,u.full_name,u.system_user_id,p.name plan_name FROM subscriptions s
        LEFT JOIN users u ON u.id=s.user_id JOIN plans p ON p.id=s.plan_id ORDER BY s.id DESC LIMIT 200''').fetchall()
    plans=[dict(r) for r in c.execute('SELECT * FROM plans WHERE active=1 ORDER BY sort_order').fetchall()]
    for p in plans: p['price_display']=money_display(p['price_minor'],p['currency'])
    students=c.execute("SELECT id,system_user_id,full_name FROM users WHERE role='student' ORDER BY full_name").fetchall()
    institutions=c.execute("SELECT id,name FROM institutions WHERE active=1 ORDER BY name").fetchall()
    promos=c.execute('SELECT * FROM promo_codes ORDER BY id DESC LIMIT 100').fetchall()
    licences=c.execute('''SELECT il.*,i.name institution_name,p.name plan_name,
        (SELECT COUNT(*) FROM institution_license_users ilu WHERE ilu.institution_license_id=il.id) allocated
        FROM institution_licenses il JOIN institutions i ON i.id=il.institution_id JOIN plans p ON p.id=il.plan_id ORDER BY il.id DESC''').fetchall()
    coverage_packages=commercial_access.package_rows(c,'',include_coming=True)
    package_entitlements=c.execute('''SELECT spe.*,u.full_name,u.system_user_id,cp.name package_name,cp.programme
      FROM student_package_entitlements spe JOIN users u ON u.id=spe.student_id JOIN coverage_packages cp ON cp.id=spe.coverage_package_id
      ORDER BY spe.id DESC LIMIT 200''').fetchall()
    checkout_requests=c.execute('''SELECT cr.*,u.full_name,u.system_user_id,cp.name package_name FROM checkout_requests cr
      JOIN users u ON u.id=cr.student_id JOIN coverage_packages cp ON cp.id=cr.coverage_package_id ORDER BY cr.id DESC LIMIT 100''').fetchall()
    revenue_display=' · '.join(money_display(r['revenue_minor'],r['currency']) for r in revenue_rows) or money_display(0,'PKR')
    dashboard={'transactions':metrics['transactions'] or 0,'successful':metrics['successful'] or 0,'failed':metrics['failed'] or 0,'active_subscriptions':active,'revenue_display':revenue_display}
    c.close(); return render_template('admin_payments.html',m=dashboard,transactions=transactions,subscriptions=subscriptions,plans=plans,students=students,institutions=institutions,promos=promos,licences=licences,
      coverage_packages=coverage_packages,package_entitlements=package_entitlements,checkout_requests=checkout_requests)


@app.route('/admin/analytics')
def admin_analytics():
    if not require('admin'): return redirect(url_for('login'))
    c=db(); users=c.execute("SELECT COUNT(CASE WHEN role='student' THEN 1 END) students,COUNT(CASE WHEN role='teacher' THEN 1 END) teachers FROM users").fetchone(); attempts=c.execute('SELECT COUNT(*) attempts,ROUND(AVG(score),1) avg_score FROM attempts').fetchone(); subjects=c.execute('SELECT subject,COUNT(*) attempts,ROUND(AVG(score),1) avg_score FROM attempts GROUP BY subject ORDER BY attempts DESC').fetchall(); c.close(); return render_template('admin_analytics.html',users=users,attempts=attempts,subjects=subjects)

@app.route('/admin')
def admin_dashboard():
    if not require('admin'): return redirect(url_for('login'))
    c=db()
    queries={'students':"SELECT COUNT(*) n FROM users WHERE role='student'",'teachers':"SELECT COUNT(*) n FROM users WHERE role='teacher'",'institutions':'SELECT COUNT(*) n FROM institutions WHERE active=1','classrooms':'SELECT COUNT(*) n FROM classrooms','questions':'SELECT COUNT(*) n FROM questions','attempts':'SELECT COUNT(*) n FROM attempts'}
    metrics={k:c.execute(q).fetchone()['n'] for k,q in queries.items()}
    metrics['draft']=c.execute("SELECT COUNT(*) n FROM questions WHERE COALESCE(review_status,status,'Draft')='Draft'").fetchone()['n']
    metrics['review']=c.execute("SELECT COUNT(*) n FROM questions WHERE COALESCE(review_status,status,'') IN ('Ready for Review','Under Review','Changes Required')").fetchone()['n']
    metrics['live']=c.execute(f"SELECT COUNT(*) n FROM questions q WHERE {live_question_clause('q')}").fetchone()['n']
    metrics['secure']=c.execute("SELECT COUNT(*) n FROM questions WHERE COALESCE(secure_bank,0)=1 AND COALESCE(active,1)=1").fetchone()['n']
    metrics['active_subscriptions']=c.execute("SELECT COUNT(*) n FROM subscriptions WHERE status IN ('active','trial') AND (ends_at IS NULL OR ends_at='' OR ends_at>=?)",(iso_today(),)).fetchone()['n']
    integration_health=integration_v1.integration_health(c)
    c.close(); return render_template('admin.html',m=metrics,commercial_gates_enabled=COMMERCIAL_GATES_ENABLED,integration_health=integration_health)

@app.route('/admin/integration-health')
def admin_integration_health():
    if not require('admin'): return redirect(url_for('login'))
    c=db()
    try:
        health=integration_v1.integration_health(c)
        quarantined=c.execute("SELECT * FROM integration_quarantine WHERE status='OPEN' ORDER BY id DESC LIMIT 100").fetchall()
        recent=c.execute("SELECT * FROM integration_dispatch_attempts ORDER BY id DESC LIMIT 100").fetchall()
        staged=c.execute("""SELECT r.*,a.authorized_at,a.authorized_by,a.reason activation_reason,a.activation_status
          FROM integration_ph_content_releases r LEFT JOIN integration_ph_product_activation_authorizations a
            ON a.release_id=r.release_id AND a.release_version=r.release_version AND a.package_checksum_sha256=r.package_checksum_sha256
          WHERE r.local_status='STAGED' AND r.release_operation='PUBLISH_SNAPSHOT'
          ORDER BY r.admitted_at DESC LIMIT 100""").fetchall()
        return render_template('admin_integration_health.html',health=health,quarantined=quarantined,recent=recent,staged=staged,preflight=integration_v1.production_preflight(strict=False))
    finally:
        c.close()


@app.route('/admin/integration-health/power-house/activate',methods=['POST'])
def admin_activate_power_house_release():
    if not require('admin'): return redirect(url_for('login'))
    rid=request.form.get('release_id','').strip(); ver=request.form.get('release_version','').strip()
    chk=request.form.get('package_checksum_sha256','').strip(); reason=request.form.get('reason','').strip()
    c=db()
    try:
        u=c.execute('SELECT system_user_id FROM users WHERE id=?',(session.get('user_id'),)).fetchone()
        actor=str(u['system_user_id'] if u and u['system_user_id'] else f"ADMIN_USER::{session.get('user_id')}")
        result=integration_v1.authorize_product_activation(c,rid,ver,chk,actor,reason)
        if result.get('status')=='ACTIVE':
            c.commit(); flash('Power House release activated for learners.' if result.get('activated_count') else 'Power House release was already activated; no duplicate state was created.','success')
        else:
            c.rollback(); flash('Activation blocked: '+str(result.get('code') or 'UNKNOWN'),'error')
    except Exception:
        c.rollback(); app.logger.exception('Power House product activation failed'); flash('Activation failed safely; no learner release was changed.','error')
    finally:
        c.close()
    return redirect(url_for('admin_integration_health'))


@app.route('/admin/challenges')
def admin_challenges():
    if not require('admin'): return redirect(url_for('login'))
    c=db()
    challenges=c.execute("""SELECT ch.*,(SELECT COUNT(*) FROM challenge_questions cq WHERE cq.challenge_id=ch.id) mapped_questions,
      (SELECT COUNT(*) FROM challenge_entries ce WHERE ce.challenge_id=ch.id AND ce.status='completed') completions
      FROM challenges ch ORDER BY ch.challenge_month DESC,ch.subject""").fetchall()
    c.close(); return render_template('admin_challenges.html',challenges=challenges)


@app.route('/admin/challenges/create',methods=['POST'])
def admin_create_challenge():
    if not require('admin'): return redirect(url_for('login'))
    code=request.form.get('code','').strip(); title=request.form.get('title','').strip(); subject=request.form.get('subject','').strip()
    if not code or not title or not subject:
        flash('Code, title and subject are required.','error'); return redirect(url_for('admin_challenges'))
    c=db()
    try:
        c.execute("""INSERT INTO challenges(code,title,country,subject,qualification,exam_board,challenge_month,description,
          duration_minutes,question_count,opens_at,closes_at,premium_required,ranking_enabled,exact_rank_min_score,max_attempts,status,
          challenge_type,chapter,topic,official)
          VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,1)""",(code,title,request.form.get('country','Pakistan').strip() or 'Pakistan',subject,
          request.form.get('qualification','').strip(),request.form.get('exam_board','').strip(),request.form.get('challenge_month','').strip(),
          request.form.get('description','').strip(),int(request.form.get('duration_minutes') or 30),int(request.form.get('question_count') or 20),
          request.form.get('opens_at','').strip(),request.form.get('closes_at','').strip(),1 if request.form.get('premium_required')=='1' else 0,
          1 if request.form.get('ranking_enabled')=='1' else 0,float(request.form.get('exact_rank_min_score') or 80),1,
          request.form.get('status','draft'),request.form.get('challenge_type','subject'),
          request.form.get('chapter','').strip(),request.form.get('topic','').strip()))
        c.commit(); flash('Challenge created.','success')
    except sqlite3.IntegrityError:
        flash('Challenge code already exists.','error')
    c.close(); return redirect(url_for('admin_challenges'))


@app.route('/admin/challenges/<int:challenge_id>',methods=['GET','POST'])
def admin_challenge_builder(challenge_id):
    if not require('admin'): return redirect(url_for('login'))
    c=db(); ch=c.execute("SELECT * FROM challenges WHERE id=?",(challenge_id,)).fetchone()
    if not ch:
        c.close(); return redirect(url_for('admin_challenges'))
    if request.method=='POST':
        if (paper['authenticity_status'] or '')=='AUTHENTIC_BLUEPRINT':
            c.close(); flash('Authentic blueprint-pinned mocks are immutable. Generate a new mock/policy version instead of editing this paper.','error')
            return redirect(url_for('admin_exam_paper_questions',paper_id=paper_id))
        qcode=request.form.get('question_id','').strip()
        q=c.execute(f"""SELECT q.* FROM questions q WHERE q.question_id=? AND {live_question_clause('q')}
          AND COALESCE(q.secure_bank,0)=1""",(qcode,)).fetchone()
        if not q:
            flash('Challenge questions must be approved, active and marked Secure Bank.','error')
        else:
            pos=int(request.form.get('position') or 0)
            if pos<=0: pos=c.execute("SELECT COALESCE(MAX(position),0)+1 n FROM challenge_questions WHERE challenge_id=?",(challenge_id,)).fetchone()['n']
            try:
                c.execute("INSERT INTO challenge_questions(challenge_id,question_id,position) VALUES(?,?,?)",(challenge_id,q['id'],pos))
                c.commit(); flash('Secure question added.','success')
            except sqlite3.IntegrityError:
                flash('Question or position already exists in this challenge.','error')
    questions=c.execute("""SELECT cq.position,q.question_id,q.question,q.subject,q.level FROM challenge_questions cq
      JOIN questions q ON q.id=cq.question_id WHERE cq.challenge_id=? ORDER BY cq.position""",(challenge_id,)).fetchall()
    entries=c.execute("""SELECT ce.*,u.full_name,u.ranking_display_name FROM challenge_entries ce JOIN users u ON u.id=ce.student_id
      WHERE ce.challenge_id=? ORDER BY ce.score DESC,ce.elapsed_seconds ASC""",(challenge_id,)).fetchall()
    c.close(); return render_template('admin_challenge_builder.html',challenge=ch,questions=questions,entries=entries)



@app.route('/admin/assessment-blueprints')
def admin_assessment_blueprints():
    if not require('admin'): return redirect(url_for('login'))
    c=db()
    rows=c.execute("""SELECT ab.*,af.name framework_name,afv.version_name framework_version_name,
      (SELECT COUNT(*) FROM assessment_blueprint_sections s WHERE s.blueprint_id=ab.id) section_count,
      (SELECT COUNT(*) FROM exam_papers ep WHERE ep.assessment_blueprint_id=ab.id) mock_count,
      (SELECT COUNT(*) FROM attempts a WHERE a.assessment_blueprint_id=ab.id) attempt_count
      FROM assessment_blueprints ab JOIN assessment_frameworks af ON af.id=ab.framework_id
      JOIN assessment_framework_versions afv ON afv.id=ab.framework_version_id
      ORDER BY ab.imported_at DESC,ab.id DESC""").fetchall()
    policy_rows=c.execute("""SELECT ap.*,ab.powerhouse_blueprint_id,afv.version_name framework_version_name
      FROM assessment_assembly_policies ap LEFT JOIN assessment_blueprints ab ON ab.id=ap.blueprint_id
      LEFT JOIN assessment_framework_versions afv ON afv.id=ap.framework_version_id
      ORDER BY CASE ap.status WHEN 'ACTIVE' THEN 0 WHEN 'DRAFT' THEN 1 ELSE 2 END,ap.created_at DESC""").fetchall()
    policies=[]
    for row in policy_rows:
        item=dict(row); item['preview']=safe_json(row['preview_json'],{}); policies.append(item)
    requests=c.execute("SELECT * FROM content_requirement_requests ORDER BY created_at DESC LIMIT 100").fetchall()
    c.close()
    return render_template('admin_assessment_blueprints.html',blueprints=rows,policies=policies,content_requests=requests)


@app.route('/admin/assessment-blueprints/import',methods=['POST'])
def admin_import_assessment_blueprint():
    if not require('admin'): return redirect(url_for('login'))
    upload=request.files.get('blueprint_file')
    if not upload or not upload.filename:
        flash('Choose a Power House blueprint JSON file.','error'); return redirect(url_for('admin_assessment_blueprints'))
    if not upload.filename.lower().endswith('.json'):
        flash('Blueprint transport must be a JSON file.','error'); return redirect(url_for('admin_assessment_blueprints'))
    try:
        raw=upload.read().decode('utf-8-sig')
        payload=json.loads(raw)
        if not isinstance(payload,dict): raise ValueError('Top-level JSON must be an object.')
    except Exception as exc:
        flash(f'Blueprint JSON could not be read: {exc}','error'); return redirect(url_for('admin_assessment_blueprints'))
    secret=os.environ.get('SCOREMAX_POWERHOUSE_SHARED_SECRET','').strip()
    require_sig=os.environ.get('SCOREMAX_REQUIRE_POWERHOUSE_SIGNATURE','0')=='1' or SCOREMAX_ENV=='production'
    report=validate_blueprint_payload(payload,shared_secret=secret,require_signature=require_sig)
    n=report['normalized']; checksum=report['calculated_checksum']
    c=db()
    existing=c.execute("""SELECT * FROM assessment_blueprints WHERE powerhouse_blueprint_id=? AND blueprint_version=?""",
                       (n['blueprint_id'],n['blueprint_version'])).fetchone()
    if existing:
        if secrets.compare_digest(existing['payload_checksum'] or '',checksum):
            c.execute("""INSERT INTO assessment_blueprint_sync_events(blueprint_id,action,sync_status,checksum,message,actor_user_id)
              VALUES(?,'DUPLICATE_IMPORT','UNCHANGED',?,'Identical immutable snapshot already exists.',?)""",
              (existing['id'],checksum,session['user_id']))
            c.commit(); c.close(); flash('That exact blueprint version is already imported; no duplicate was created.','info')
            return redirect(url_for('admin_assessment_blueprint_detail',blueprint_id=existing['id']))
        c.execute("""INSERT INTO assessment_blueprint_sync_events(blueprint_id,action,sync_status,checksum,message,actor_user_id)
          VALUES(?,'CHECKSUM_MISMATCH','SYNC_ERROR',?,'Same blueprint ID/version arrived with different immutable content.',?)""",
          (existing['id'],checksum,session['user_id']))
        c.commit(); c.close(); flash('Rejected: this Blueprint ID/version already exists with a different checksum. Power House must issue a new version.','error')
        return redirect(url_for('admin_assessment_blueprint_detail',blueprint_id=existing['id']))
    fw=c.execute("SELECT * FROM assessment_frameworks WHERE powerhouse_framework_id=?",(n['framework_id'],)).fetchone()
    if not fw:
        fid=c.execute("""INSERT INTO assessment_frameworks(powerhouse_framework_id,name,country,authority)
          VALUES(?,?,?,?)""",(n['framework_id'],n['framework_name'],str(payload.get('country') or ''),n['authority'])).lastrowid
    else:
        fid=fw['id']
        if (fw['name'] or '').casefold()!=n['framework_name'].casefold():
            report['warnings'].append('Framework name differs from the existing Power House framework ID; existing identity was preserved.')
    fver=c.execute("""SELECT * FROM assessment_framework_versions WHERE framework_id=? AND powerhouse_framework_version_id=?""",
                   (fid,n['framework_version_id'])).fetchone()
    if not fver:
        fvid=c.execute("""INSERT INTO assessment_framework_versions(framework_id,powerhouse_framework_version_id,version_name,effective_from,effective_to,status)
          VALUES(?,?,?,?,?,'ACTIVE')""",(fid,n['framework_version_id'],n['framework_version_name'],n['activation_date'],n['superseded_date'])).lastrowid
    else: fvid=fver['id']
    local_status='VALIDATED' if report['valid'] else 'REJECTED'
    signature_status='VERIFIED' if secret and report['valid'] else ('NOT_CONFIGURED' if not secret else 'FAILED')
    cur=c.execute("""INSERT INTO assessment_blueprints(
      powerhouse_blueprint_id,framework_id,framework_version_id,blueprint_version,source_status,local_status,
      authority,source_reference,governance_note,total_questions,duration_minutes,difficulty_distribution_json,
      activation_date,superseded_date,source_created_at,source_approved_at,source_approved_by,source_policy_version,
      imported_by,payload_checksum,signature_status,sync_status,validation_report_json,immutable_payload_json)
      VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
      (n['blueprint_id'],fid,fvid,n['blueprint_version'],n['source_status'],local_status,n['authority'],n['source_reference'],
       n['governance_note'],int(n['total_questions'] or 0),n['duration_minutes'],json.dumps(n['difficulty_distribution']),
       n['activation_date'],n['superseded_date'],n['source_created_at'],n['source_approved_at'],n['source_approved_by'],
       n['source_policy_version'],session['user_id'],checksum,signature_status,'VALIDATED' if report['valid'] else 'REJECTED',
       json.dumps(report),canonical_json(payload)))
    bid=cur.lastrowid
    for row in n['sections']:
        if not row['subject'] or not row['question_count'] or not row['weight_percent']: continue
        c.execute("""INSERT INTO assessment_blueprint_sections(blueprint_id,section_order,section_code,section_title,subject,
          question_count,weight_percent,duration_minutes,difficulty_distribution_json,rules_json)
          VALUES(?,?,?,?,?,?,?,?,?,?)""",(bid,row['section_order'],row['section_code'],row['section_title'],row['subject'],
          row['question_count'],row['weight_percent'],row['duration_minutes'],json.dumps(row['difficulty_distribution']),json.dumps(row['rules'])))
    c.execute("""INSERT INTO assessment_blueprint_sync_events(blueprint_id,action,sync_status,checksum,message,actor_user_id,metadata_json)
      VALUES(?,'IMPORT',?,?,?,?,?)""",(bid,'VALIDATED' if report['valid'] else 'REJECTED',checksum,
      'Immutable Power House snapshot imported.',session['user_id'],json.dumps({'filename':upload.filename})))
    record_blueprint_audit(c,bid,'IMPORTED','',local_status,'Power House JSON transport imported.',report,session['user_id'])
    c.commit(); c.close()
    flash('Blueprint imported and validated.' if report['valid'] else 'Blueprint imported as REJECTED. Review blocking validation errors.','success' if report['valid'] else 'error')
    return redirect(url_for('admin_assessment_blueprint_detail',blueprint_id=bid))


@app.route('/admin/assessment-blueprints/<int:blueprint_id>')
def admin_assessment_blueprint_detail(blueprint_id):
    if not require('admin'): return redirect(url_for('login'))
    c=db(); bp=blueprint_joined(c,blueprint_id)
    if not bp: c.close(); return redirect(url_for('admin_assessment_blueprints'))
    sections=blueprint_sections(c,blueprint_id)
    report=safe_json(bp['validation_report_json'],{})
    active=active_assessment_blueprint(c,bp['framework_name'],bp['framework_version_name'])
    impact=compare_blueprints(blueprint_payload_from_record(c,active['id']) if active and active['id']!=bp['id'] else None,
                              blueprint_payload_from_record(c,bp['id']) or {})
    bank=blueprint_bank_sufficiency(c,blueprint_id)
    impact['candidate_bank_ready']=bool(bank and bank['ready'])
    impact['candidate_bank_blockers']=bank['blockers'] if bank else []
    impact['historical_results_immutable']=True
    impact['existing_mocks_using_current_version']=c.execute("SELECT COUNT(*) n FROM exam_papers WHERE assessment_blueprint_id=?",
      (active['id'],)).fetchone()['n'] if active and active['id']!=bp['id'] else 0
    impact['study_plan_implication']='Future plans use the new subject weights after activation; existing plan/history snapshots are retained.'
    audits=c.execute("SELECT * FROM assessment_blueprint_audit WHERE blueprint_id=? ORDER BY created_at DESC",(blueprint_id,)).fetchall()
    syncs=c.execute("SELECT * FROM assessment_blueprint_sync_events WHERE blueprint_id=? ORDER BY created_at DESC",(blueprint_id,)).fetchall()
    papers=c.execute("""SELECT ep.*,(SELECT COUNT(*) FROM exam_paper_questions pq WHERE pq.paper_id=ep.id) question_count
      FROM exam_papers ep WHERE ep.assessment_blueprint_id=? ORDER BY ep.created_at DESC""",(blueprint_id,)).fetchall()
    policy_rows=c.execute("""SELECT * FROM assessment_assembly_policies WHERE blueprint_id=? OR
      (scope_type='framework_version' AND scope_key=?) OR scope_type='global'
      ORDER BY CASE status WHEN 'ACTIVE' THEN 0 ELSE 1 END,created_at DESC""",(blueprint_id,str(bp['framework_version_id']))).fetchall()
    policies=[]
    for row in policy_rows:
        item=dict(row); item['preview']=safe_json(row['preview_json'],{}); policies.append(item)
    c.close()
    return render_template('admin_assessment_blueprint_detail.html',bp=bp,sections=sections,validation=report,
                           impact=impact,bank=bank,audits=audits,syncs=syncs,papers=papers,policies=policies)


@app.route('/admin/assessment-blueprints/<int:blueprint_id>/activate',methods=['POST'])
def admin_activate_assessment_blueprint(blueprint_id):
    if not require('admin'): return redirect(url_for('login'))
    reason=request.form.get('reason','').strip()
    c=db(); bp=blueprint_joined(c,blueprint_id)
    if not bp: c.close(); return redirect(url_for('admin_assessment_blueprints'))
    payload=blueprint_payload_from_record(c,blueprint_id) or {}
    secret=os.environ.get('SCOREMAX_POWERHOUSE_SHARED_SECRET','').strip()
    report=validate_blueprint_payload(payload,shared_secret=secret,
      require_signature=os.environ.get('SCOREMAX_REQUIRE_POWERHOUSE_SIGNATURE','0')=='1' or SCOREMAX_ENV=='production')
    if not report['valid'] or bp['local_status'] not in ('VALIDATED','IMPORTED'):
        c.close(); flash('Blueprint cannot be activated until all blocking validation errors are resolved through a new Power House version.','error')
        return redirect(url_for('admin_assessment_blueprint_detail',blueprint_id=blueprint_id))
    current=active_assessment_blueprint(c,bp['framework_name'],bp['framework_version_name'])
    if current and current['id']==blueprint_id:
        c.close(); flash('This blueprint is already active.','info'); return redirect(url_for('admin_assessment_blueprint_detail',blueprint_id=blueprint_id))
    if current:
        # Prevent accidental downgrade by numeric version where possible.
        try:
            older=float(bp['blueprint_version'])<float(current['blueprint_version'])
        except Exception: older=False
        if older and request.form.get('allow_older')!='1':
            c.close(); flash('Activation blocked: this is older than the current active blueprint. Use an explicitly governed emergency override.','error')
            return redirect(url_for('admin_assessment_blueprint_detail',blueprint_id=blueprint_id))
    impact=compare_blueprints(blueprint_payload_from_record(c,current['id']) if current else None,payload)
    if current:
        c.execute("""UPDATE assessment_blueprints SET local_status='SUPERSEDED',superseded_date=?,superseded_by_blueprint_id=?
          WHERE id=?""",(iso_today(),blueprint_id,current['id']))
        record_blueprint_audit(c,current['id'],'SUPERSEDED','ACTIVE','SUPERSEDED',reason or 'New approved blueprint activated.',impact,session['user_id'])
    c.execute("""UPDATE assessment_blueprints SET local_status='ACTIVE',activated_at=?,activated_by=?,sync_status='ACTIVE'
      WHERE id=?""",(datetime.now().isoformat(timespec='seconds'),session['user_id'],blueprint_id))
    record_blueprint_audit(c,blueprint_id,'ACTIVATED',bp['local_status'],'ACTIVE',reason or 'Authorised admin activation.',impact,session['user_id'])
    c.commit(); c.close(); flash('Blueprint activated. Future authentic mocks, projections and plans use this version; history remains pinned.','success')
    return redirect(url_for('admin_assessment_blueprint_detail',blueprint_id=blueprint_id))


@app.route('/admin/assessment-blueprints/<int:blueprint_id>/archive',methods=['POST'])
def admin_archive_assessment_blueprint(blueprint_id):
    if not require('admin'): return redirect(url_for('login'))
    c=db(); bp=blueprint_joined(c,blueprint_id)
    if not bp: c.close(); return redirect(url_for('admin_assessment_blueprints'))
    if bp['local_status']=='ACTIVE':
        c.close(); flash('Activate a replacement before archiving an active blueprint.','error'); return redirect(url_for('admin_assessment_blueprint_detail',blueprint_id=blueprint_id))
    c.execute("UPDATE assessment_blueprints SET local_status='ARCHIVED' WHERE id=?",(blueprint_id,))
    record_blueprint_audit(c,blueprint_id,'ARCHIVED',bp['local_status'],'ARCHIVED',request.form.get('reason','').strip(),{},session['user_id'])
    c.commit(); c.close(); flash('Blueprint archived; immutable history remains available.','success')
    return redirect(url_for('admin_assessment_blueprints'))


@app.route('/admin/assessment-blueprints/<int:blueprint_id>/export')
def admin_export_assessment_blueprint(blueprint_id):
    if not require('admin'): return redirect(url_for('login'))
    c=db(); bp=blueprint_joined(c,blueprint_id); payload=blueprint_payload_from_record(c,blueprint_id) if bp else None; c.close()
    if not payload: abort(404)
    data=canonical_json(payload).encode('utf-8')
    return send_file(io.BytesIO(data),mimetype='application/json',as_attachment=True,
                     download_name=f"{bp['powerhouse_blueprint_id']}-v{bp['blueprint_version']}-snapshot.json")


@app.route('/admin/assessment-blueprints/<int:blueprint_id>/generate-mock',methods=['POST'])
def admin_generate_blueprint_mock(blueprint_id):
    if not require('admin'): return redirect(url_for('login'))
    c=db(); bp=blueprint_joined(c,blueprint_id)
    if not bp: c.close(); return redirect(url_for('admin_assessment_blueprints'))
    preflight=assemble_blueprint_mock(c,blueprint_id,seed=request.form.get('seed','') or datetime.now().isoformat())
    if not preflight['ready']:
        record_blueprint_audit(c,blueprint_id,'MOCK_PREFLIGHT_BLOCKED',bp['local_status'],bp['local_status'],
          'Authentic mock not generated because inventory did not meet the blueprint.',preflight,session['user_id'])
        c.commit(); c.close(); flash('Authentic mock blocked: '+' '.join(preflight['blockers'][:3]),'error')
        return redirect(url_for('admin_assessment_blueprint_detail',blueprint_id=blueprint_id))
    snapshot=blueprint_payload_from_record(c,blueprint_id) or {}
    code=(request.form.get('code','').strip() or f"{bp['framework_name']}-{bp['framework_version_name']}-BP{bp['blueprint_version']}-{datetime.now().strftime('%Y%m%d%H%M%S')}")
    title=request.form.get('title','').strip() or f"{bp['framework_name']} {bp['framework_version_name']} Authentic Mock"
    cur=c.execute("""INSERT INTO exam_papers(code,title,paper_kind,official_year,source_label,reproduction_status,duration_minutes,total_marks,
      instructions,premium_required,active,assessment_blueprint_id,blueprint_source_id,blueprint_version,framework_version,
      blueprint_snapshot_json,assembly_policy_id,assembly_policy_version,authenticity_status,preflight_json,generated_at)
      VALUES(?,?,'scoremax_mock',?,'ScoreMax Blueprint Assembly','scoremax_original',?,?,?,0,1,?,?,?,?,?,?,?,?,?,?)""",
      (code,title,bp['framework_version_name'],bp['duration_minutes'],float(bp['total_questions'] or 0),
       f"Authentic mock governed by {bp['powerhouse_blueprint_id']} v{bp['blueprint_version']}.",blueprint_id,
       bp['powerhouse_blueprint_id'],bp['blueprint_version'],bp['framework_version_name'],canonical_json(snapshot),
       preflight['policy_id'],preflight['policy_version'],'AUTHENTIC_BLUEPRINT',json.dumps(preflight),datetime.now().isoformat(timespec='seconds')))
    paper_id=cur.lastrowid
    c.executemany("""INSERT INTO exam_paper_questions(paper_id,question_id,position,section_label,display_number,marks)
      VALUES(?,?,?,?,?,?)""",[(paper_id,x['question_id'],x['position'],x['section_label'],x['display_number'],x['marks']) for x in preflight['selected']])
    record_blueprint_audit(c,blueprint_id,'AUTHENTIC_MOCK_GENERATED',bp['local_status'],bp['local_status'],title,preflight,session['user_id'])
    c.commit(); c.close(); flash('Authentic blueprint-pinned mock generated successfully.','success')
    return redirect(url_for('admin_exam_paper_questions',paper_id=paper_id))


@app.route('/admin/assessment-blueprints/<int:blueprint_id>/request-content',methods=['POST'])
def admin_request_blueprint_content(blueprint_id):
    if not require('admin'): return redirect(url_for('login'))
    c=db(); bank=blueprint_bank_sufficiency(c,blueprint_id,int(request.form.get('target_parallel_mocks') or 3))
    if not bank: c.close(); return redirect(url_for('admin_assessment_blueprints'))
    created=0
    for row in bank['subjects']:
        if row['shortage_for_target']<=0 and row['status']=='Ready': continue
        code=f"CRR-{blueprint_id}-{re.sub(r'[^A-Z0-9]','',row['subject'].upper())[:8]}-{int(time.time())}-{created}"
        reason=(f"Blueprint requires {row['required_per_mock']} {row['subject']} questions per authentic mock. "
                f"Current governed bank has {row['usable_questions']} items across {row['family_count']} families; "
                f"safe parallel depth is {row['safe_parallel_forms']} form(s).")
        c.execute("""INSERT INTO content_requirement_requests(request_code,framework_version_id,blueprint_id,subject,
          assets_required,families_required,intended_use,priority,reason,existing_bank_evidence_json,created_by)
          VALUES(?,?,?,?,?,?,?,?,?,?,?)""",(code,bank['blueprint']['framework_version_id'],blueprint_id,row['subject'],
          max(row['shortage_for_target'],row['required_per_mock']-row['usable_questions'],0),
          max(0,int(row['required_per_mock']*.70)-row['family_count']),'authentic_mock','Critical' if row['status']=='Blocked' else 'High',
          reason,json.dumps(row),session['user_id']))
        created+=1
    c.commit(); c.close(); flash(f'{created} structured Power House content requirement request(s) created.','success' if created else 'info')
    return redirect(url_for('admin_assessment_blueprint_detail',blueprint_id=blueprint_id))


@app.route('/admin/content-requirements/<int:request_id>/export')
def admin_export_content_requirement(request_id):
    if not require('admin'): return redirect(url_for('login'))
    c=db()
    row=c.execute("""SELECT cr.*,ab.powerhouse_blueprint_id,ab.blueprint_version,af.powerhouse_framework_id,af.name framework_name,
      afv.powerhouse_framework_version_id,afv.version_name framework_version_name
      FROM content_requirement_requests cr
      LEFT JOIN assessment_blueprints ab ON ab.id=cr.blueprint_id
      LEFT JOIN assessment_frameworks af ON af.id=ab.framework_id
      LEFT JOIN assessment_framework_versions afv ON afv.id=cr.framework_version_id
      WHERE cr.id=?""",(request_id,)).fetchone()
    if not row:
        c.close(); abort(404)
    payload={
      'schema_version':'1.0','request_id':row['request_code'],'source_system':'ScoreMax','status':row['status'],
      'framework':{'id':row['powerhouse_framework_id'] or '','name':row['framework_name'] or ''},
      'framework_version':{'id':row['powerhouse_framework_version_id'] or '','name':row['framework_version_name'] or ''},
      'blueprint':{'id':row['powerhouse_blueprint_id'] or '','version':row['blueprint_version'] or ''},
      'requirement':{'subject':row['subject'],'chapter':row['chapter'],'learning_outcome':row['learning_outcome'],
        'mastery_level':row['mastery_level'],'difficulty':row['difficulty'],'assets_required':row['assets_required'],
        'families_required':row['families_required'],'intended_use':row['intended_use'],'deadline':row['deadline'],
        'priority':row['priority'],'reason':row['reason']},
      'existing_bank_evidence':safe_json(row['existing_bank_evidence_json'],{}),
      'created_at':row['created_at']
    }
    payload['checksum']=calculate_checksum(payload)
    c.execute("UPDATE content_requirement_requests SET status='EXPORTED',updated_at=CURRENT_TIMESTAMP WHERE id=?",(request_id,))
    c.commit(); c.close()
    data=canonical_json(payload).encode('utf-8')
    return send_file(io.BytesIO(data),mimetype='application/json',as_attachment=True,download_name=f"{row['request_code']}.json")


@app.route('/admin/assessment-policies',methods=['POST'])
def admin_create_assessment_policy():
    if not require('admin'): return redirect(url_for('login'))
    scope_type=request.form.get('scope_type','global').strip()
    if scope_type not in ('global','framework_version','blueprint','programme','subject','chapter','assessment_type'): scope_type='global'
    scope_key=request.form.get('scope_key','').strip()
    version=request.form.get('policy_version','').strip() or datetime.now().strftime('%Y.%m.%d.%H%M')
    rigor=max(0,min(100,int(request.form.get('rigor_score') or 50)))
    standard=max(0,min(100,int(request.form.get('mastery_standard_score') or 50)))
    blueprint_id=int(request.form.get('blueprint_id') or 0) or None
    c=db(); bp=blueprint_joined(c,blueprint_id) if blueprint_id else None
    if scope_type=='blueprint' and blueprint_id:
        scope_key=str(blueprint_id)
    elif scope_type=='framework_version' and bp:
        scope_key=str(bp['framework_version_id'])
    preview={}
    if bp:
        current=blueprint_bank_sufficiency(c,bp['id'])
        preview={'blueprint':bp['powerhouse_blueprint_id'],'current_bank_ready':bool(current and current['ready']),
                 'current_mix':current['target_difficulty_mix'] if current else {},'proposed_mix':rigor_mix(rigor,safe_json(bp['difficulty_distribution_json'],{})),
                 'historical_simulation':simulate_policy_impact(c,bp,standard),
                 'historical_mastery_is_not_rewritten':True,'material_tightening_action':'Verification Due'}
    code=request.form.get('policy_code','').strip() or f"SMX-RIGOR-{re.sub(r'[^A-Z0-9]','-',scope_type.upper())}-{int(time.time())}"
    selection={'unseen_family_ratio':round(.50+.004*rigor,2),'duplicate_family_limit':1,
               'target_difficulty_mix':preview.get('proposed_mix') or rigor_mix(rigor),
               'cognitive_demand_bias':'higher' if rigor>=65 else ('accessible' if rigor<=35 else 'balanced')}
    evidence={'mastery_standard_score':standard,'historical_results_immutable':True,
              'policy_tightening_action':'Verification Due','academic_approval_required':True}
    try:
        cur=c.execute("""INSERT INTO assessment_assembly_policies(policy_code,policy_version,scope_type,scope_key,framework_version_id,
          blueprint_id,name,rigor_score,mastery_standard_score,selection_config_json,evidence_config_json,status,created_by,reason,preview_json)
          VALUES(?,?,?,?,?,?,?,?,?,?,?,'DRAFT',?,?,?)""",(code,version,scope_type,scope_key,
          bp['framework_version_id'] if bp else None,blueprint_id,request.form.get('name','').strip() or f'Rigor Policy {version}',
          rigor,standard,json.dumps(selection),json.dumps(evidence),session['user_id'],request.form.get('reason','').strip(),json.dumps(preview)))
        record_policy_audit(c,cur.lastrowid,'CREATED','', 'DRAFT',request.form.get('reason','').strip(),
          {'scope_type':scope_type,'scope_key':scope_key,'rigor_score':rigor,'mastery_standard_score':standard,'preview':preview},session['user_id'])
        c.commit(); flash('Rigor/mastery policy saved as DRAFT. Preview it before academic activation.','success')
    except sqlite3.IntegrityError:
        flash('That policy code or scope/version already exists.','error')
    c.close(); return redirect(request.referrer or url_for('admin_assessment_blueprints'))


@app.route('/admin/assessment-policies/<int:policy_id>/activate',methods=['POST'])
def admin_activate_assessment_policy(policy_id):
    if not require('admin'): return redirect(url_for('login'))
    c=db(); policy=c.execute("SELECT * FROM assessment_assembly_policies WHERE id=?",(policy_id,)).fetchone()
    if not policy: c.close(); return redirect(url_for('admin_assessment_blueprints'))
    if policy['status']!='DRAFT':
        c.close(); flash('Only a reviewed DRAFT policy can be activated.','error'); return redirect(request.referrer or url_for('admin_assessment_blueprints'))
    reason=request.form.get('reason','').strip() or policy['reason'] or 'Academic approval'
    prior=c.execute("""SELECT * FROM assessment_assembly_policies WHERE status='ACTIVE' AND scope_type=? AND scope_key=? AND id<>?
      ORDER BY approved_at DESC,id DESC LIMIT 1""",(policy['scope_type'],policy['scope_key'],policy_id)).fetchone()
    c.execute("""UPDATE assessment_assembly_policies SET status='SUPERSEDED',superseded_at=?
      WHERE status='ACTIVE' AND scope_type=? AND scope_key=? AND id<>?""",
      (datetime.now().isoformat(timespec='seconds'),policy['scope_type'],policy['scope_key'],policy_id))
    c.execute("""UPDATE assessment_assembly_policies SET status='ACTIVE',approved_by=?,approved_at=?,reason=? WHERE id=?""",
      (session['user_id'],datetime.now().isoformat(timespec='seconds'),reason,policy_id))
    # Material tightening never hard-downgrades history. Only records in the governed
    # scope become Verification Due; lower/equal standards do not disturb mastery.
    old_standard=int(prior['mastery_standard_score'] or 50) if prior else 50
    new_standard=int(policy['mastery_standard_score'] or 50)
    if new_standard>old_standard:
        clauses=["status IN ('Verified','Elite Candidate')"]; params=[]
        if policy['scope_type']=='blueprint' and policy['blueprint_id']:
            bp=blueprint_joined(c,policy['blueprint_id'])
            if bp:
                clauses.append("lower(programme)=lower(?)"); params.append(bp['framework_name'])
        elif policy['scope_type']=='framework_version' and policy['framework_version_id']:
            fv=c.execute("""SELECT af.name FROM assessment_framework_versions afv JOIN assessment_frameworks af ON af.id=afv.framework_id
              WHERE afv.id=?""",(policy['framework_version_id'],)).fetchone()
            if fv: clauses.append("lower(programme)=lower(?)"); params.append(fv['name'])
        elif policy['scope_type']=='subject' and policy['scope_key']:
            clauses.append("lower(subject)=lower(?)"); params.append(policy['scope_key'])
        elif policy['scope_type']=='chapter' and policy['scope_key']:
            clauses.append("lower(chapter)=lower(?)"); params.append(policy['scope_key'])
        c.execute(f"UPDATE mastery_records SET status='Verification Due',updated_at=CURRENT_TIMESTAMP WHERE {' AND '.join(clauses)}",params)
    record_policy_audit(c,policy_id,'ACTIVATED',policy['status'],'ACTIVE',reason,
      {'prior_policy_id':prior['id'] if prior else None,'previous_standard':old_standard,'new_standard':new_standard,
       'rigor_score':policy['rigor_score'],'scope_type':policy['scope_type'],'scope_key':policy['scope_key'],
       'historical_results_immutable':True},session['user_id'])
    c.commit(); c.close(); flash('Versioned assessment policy activated. Historical attempts remain unchanged.','success')
    return redirect(request.referrer or url_for('admin_assessment_blueprints'))

@app.route('/admin/exams')
def admin_exams():
    if not require('admin'): return redirect(url_for('login'))
    c=db()
    blueprints=c.execute("""SELECT eb.*,(SELECT COUNT(*) FROM exam_papers ep WHERE ep.blueprint_id=eb.id) paper_count
        FROM exam_blueprints eb ORDER BY eb.subject,eb.title""").fetchall()
    papers=c.execute("""SELECT ep.*,eb.subject,eb.exam_board,eb.qualification,
        (SELECT COUNT(*) FROM exam_paper_questions pq WHERE pq.paper_id=ep.id) question_count
        FROM exam_papers ep LEFT JOIN exam_blueprints eb ON eb.id=ep.blueprint_id
        ORDER BY ep.created_at DESC""").fetchall()
    c.close()
    return render_template('admin_exams.html',blueprints=blueprints,papers=papers)


@app.route('/admin/exams/blueprint',methods=['POST'])
def admin_create_exam_blueprint():
    if not require('admin'): return redirect(url_for('login'))
    vals={k:request.form.get(k,'').strip() for k in ['code','title','country','qualification','exam_board','programme','subject','curriculum_version','paper_name']}
    if not vals['code'] or not vals['title'] or not vals['subject']:
        flash('Blueprint code, title and subject are required.','error'); return redirect(url_for('admin_exams'))
    duration=int(request.form.get('duration_minutes') or 0) or None
    marks=float(request.form.get('total_marks') or 0) or None
    c=db()
    try:
        c.execute("""INSERT INTO exam_blueprints(code,title,country,qualification,exam_board,programme,subject,curriculum_version,paper_name,duration_minutes,total_marks)
          VALUES(?,?,?,?,?,?,?,?,?,?,?)""",(vals['code'],vals['title'],vals['country'] or 'Pakistan',vals['qualification'],vals['exam_board'],vals['programme'],vals['subject'],vals['curriculum_version'],vals['paper_name'],duration,marks))
        c.commit(); flash('Exam blueprint created.','success')
    except sqlite3.IntegrityError:
        flash('That blueprint code already exists.','error')
    c.close(); return redirect(url_for('admin_exams'))


@app.route('/admin/exams/paper',methods=['POST'])
def admin_create_exam_paper():
    if not require('admin'): return redirect(url_for('login'))
    blueprint_id=int(request.form.get('blueprint_id') or 0) or None
    code=request.form.get('code','').strip(); title=request.form.get('title','').strip()
    kind=request.form.get('paper_kind','scoremax_mock')
    if kind not in ('scoremax_mock','official_past_paper'): kind='scoremax_mock'
    if not code or not title:
        flash('Paper code and title are required.','error'); return redirect(url_for('admin_exams'))
    reproduction=request.form.get('reproduction_status','scoremax_original')
    # Official/historical papers are not published as reproducible by default.
    if kind=='official_past_paper' and reproduction not in ('permitted','licensed','public_domain'):
        reproduction='permission_pending'
    c=db()
    try:
        c.execute("""INSERT INTO exam_papers(blueprint_id,code,title,paper_kind,official_year,source_label,source_url,reproduction_status,
          duration_minutes,total_marks,instructions,premium_required,active)
          VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",(blueprint_id,code,title,kind,request.form.get('official_year','').strip(),
          request.form.get('source_label','').strip(),request.form.get('source_url','').strip(),reproduction,
          int(request.form.get('duration_minutes') or 0) or None,float(request.form.get('total_marks') or 0) or None,
          request.form.get('instructions','').strip(),1 if request.form.get('premium_required')=='1' else 0,
          1 if request.form.get('active')=='1' else 0))
        c.commit(); flash('Exam paper created.','success')
    except sqlite3.IntegrityError:
        flash('That paper code already exists.','error')
    c.close(); return redirect(url_for('admin_exams'))


@app.route('/admin/exams/paper/<int:paper_id>',methods=['GET','POST'])
def admin_exam_paper_questions(paper_id):
    if not require('admin'): return redirect(url_for('login'))
    c=db()
    paper=c.execute("""SELECT ep.*,COALESCE(eb.subject,'') subject,eb.exam_board,
      ab.powerhouse_blueprint_id authoritative_blueprint_id,af.name authoritative_framework,
      afv.version_name authoritative_framework_version
      FROM exam_papers ep LEFT JOIN exam_blueprints eb ON eb.id=ep.blueprint_id
      LEFT JOIN assessment_blueprints ab ON ab.id=ep.assessment_blueprint_id
      LEFT JOIN assessment_frameworks af ON af.id=ab.framework_id
      LEFT JOIN assessment_framework_versions afv ON afv.id=ab.framework_version_id WHERE ep.id=?""",(paper_id,)).fetchone()
    if not paper:
        c.close(); return redirect(url_for('admin_exams'))
    if request.method=='POST':
        qcode=request.form.get('question_id','').strip()
        q=c.execute(f"""SELECT q.* FROM questions q WHERE q.question_id=? AND {live_question_clause('q')}""",(qcode,)).fetchone()
        if not q:
            flash('Approved active question not found.','error')
        else:
            position=int(request.form.get('position') or 0)
            if position<=0:
                position=c.execute("SELECT COALESCE(MAX(position),0)+1 n FROM exam_paper_questions WHERE paper_id=?",(paper_id,)).fetchone()['n']
            try:
                c.execute("""INSERT INTO exam_paper_questions(paper_id,question_id,position,section_label,display_number,marks)
                  VALUES(?,?,?,?,?,?)""",(paper_id,q['id'],position,request.form.get('section_label','').strip(),
                  request.form.get('display_number','').strip(),float(request.form.get('marks') or q['marks'] or 1)))
                c.commit(); flash('Question added to paper.','success')
            except sqlite3.IntegrityError:
                flash('That question or position is already used in this paper.','error')
    questions=c.execute("""SELECT pq.*,q.question_id,q.question,q.qtype,q.level FROM exam_paper_questions pq
      JOIN questions q ON q.id=pq.question_id WHERE pq.paper_id=? ORDER BY pq.position""",(paper_id,)).fetchall()
    c.close()
    return render_template('admin_exam_paper.html',paper=paper,questions=questions)


@app.route('/admin/questions')
def admin_questions():
    if not require('admin'): return redirect(url_for('login'))
    c=db()
    status=request.args.get('status','').strip(); subject=request.args.get('subject','').strip(); q=request.args.get('q','').strip()
    clauses=['1=1']; params=[]
    if status:
        clauses.append("COALESCE(review_status,status,'Draft')=?"); params.append(status)
    if subject:
        clauses.append('subject=?'); params.append(subject)
    if q:
        clauses.append('(question_id LIKE ? OR question LIKE ? OR topic LIKE ? OR learning_outcome LIKE ?)'); params.extend([f'%{q}%']*4)
    rows=c.execute(f'''SELECT id,question_id,subject,chapter,topic,subtopic,qtype,level,question,review_status,status,active,secure_bank,question_version,reviewer,reviewed_at
        FROM questions WHERE {' AND '.join(clauses)} ORDER BY id DESC LIMIT 500''',params).fetchall()
    subjects=c.execute("SELECT DISTINCT subject FROM questions WHERE COALESCE(subject,'')<>'' ORDER BY subject").fetchall()
    counts={r['review_status'] or 'Unspecified':r['n'] for r in c.execute("SELECT COALESCE(review_status,status,'Unspecified') review_status,COUNT(*) n FROM questions GROUP BY COALESCE(review_status,status,'Unspecified')").fetchall()}
    c.close(); return render_template('admin_questions.html',questions=rows,subjects=subjects,counts=counts,filters={'status':status,'subject':subject,'q':q})

@app.route('/admin/questions/<int:qid>')
def admin_question_detail(qid):
    if not require('admin'): return redirect(url_for('login'))
    c=db(); q=c.execute('SELECT * FROM questions WHERE id=?',(qid,)).fetchone()
    if not q: c.close(); return redirect(url_for('admin_questions'))
    events=c.execute('''SELECT e.*,u.full_name reviewer_name FROM question_review_events e LEFT JOIN users u ON u.id=e.reviewer_id WHERE e.question_id=? ORDER BY e.id DESC LIMIT 100''',(qid,)).fetchall()
    versions=c.execute('SELECT * FROM question_versions WHERE question_id=? ORDER BY version DESC',(qid,)).fetchall()
    family=c.execute('SELECT * FROM question_families WHERE family_key=?',(q['family_key'] or '',)).fetchone()
    health=question_health(c,qid); c.close()
    return render_template('admin_question_detail.html',q=q,family=family,events=events,versions=versions,health=health,answer_config=safe_json(q['answer_config'],{}),marking_config=safe_json(q['marking_config'],{}))

@app.route('/admin/question-families')
def admin_question_families():
    if not require('admin'): return redirect(url_for('login'))
    c=db(); rows=c.execute("""SELECT qf.*,
      COUNT(q.id) question_count,
      SUM(CASE WHEN q.review_status='Approved' AND q.status='Approved' AND q.active=1 THEN 1 ELSE 0 END) approved_questions
      FROM question_families qf LEFT JOIN questions q ON q.family_key=qf.family_key
      GROUP BY qf.family_key ORDER BY qf.subject,qf.family_id""").fetchall(); c.close()
    return render_template('admin_question_families.html',families=rows)

@app.route('/admin/question-families/<family_key>/review',methods=['POST'])
def admin_question_family_review(family_key):
    if not require('admin'): return redirect(url_for('login'))
    action=request.form.get('action','').strip()
    status_map={'approve':'Approved','changes':'Changes Required','retire':'Retired','restore':'Approved'}
    if action not in status_map:
        flash('Unknown family review action.','error'); return redirect(url_for('admin_question_families'))
    c=db(); fam=c.execute('SELECT * FROM question_families WHERE family_key=?',(family_key,)).fetchone()
    if not fam:
        c.close(); flash('Question family not found.','error'); return redirect(url_for('admin_question_families'))
    new_status=status_map[action]; active=1 if new_status=='Approved' else 0
    reviewed_at=datetime.utcnow().isoformat(timespec='seconds') if new_status in {'Approved','Retired'} else ''
    c.execute("UPDATE question_families SET review_status=?,active=?,reviewer_id=?,reviewed_at=?,updated_at=CURRENT_TIMESTAMP WHERE family_key=?",
              (new_status,active,session.get('user_id'),reviewed_at,family_key))
    c.commit(); c.close()
    flash(f'Question family moved to {new_status}. Individual questions still require their own approval.','success')
    return redirect(url_for('admin_question_families'))

@app.route('/admin/questions/<int:qid>/review',methods=['POST'])
def admin_question_review(qid):
    if not require('admin'): return redirect(url_for('login'))
    action=request.form.get('action','').strip(); reason=request.form.get('reason_code','').strip(); note=request.form.get('note','').strip()
    status_map={'ready':'Ready for Review','start':'Under Review','approve':'Approved','changes':'Changes Required','reject':'Rejected','retire':'Retired','restore':'Approved'}
    if action not in status_map: flash('Unknown review action.','error'); return redirect(url_for('admin_question_detail',qid=qid))
    c=db(); row=c.execute('SELECT * FROM questions WHERE id=?',(qid,)).fetchone()
    if not row: c.close(); return redirect(url_for('admin_questions'))
    version=int(row['question_version'] or 1)
    c.execute('INSERT OR IGNORE INTO question_versions(question_id,version,snapshot_json,change_note) VALUES(?,?,?,?)',(qid,version,json.dumps(question_snapshot(row)),f'Before {status_map[action]}'))
    active=0 if action in {'reject','retire'} else 1
    reviewed_at=datetime.utcnow().isoformat(timespec='seconds') if action in {'approve','reject','retire','restore'} else (row['reviewed_at'] or '')
    reviewer=session.get('full_name','Platform Admin')
    c.execute('UPDATE questions SET review_status=?,status=?,active=?,reviewer=?,reviewed_at=? WHERE id=?',(status_map[action],status_map[action],active,reviewer,reviewed_at,qid))
    c.execute('INSERT INTO question_review_events(question_id,action,reviewer_id,reason_code,note) VALUES(?,?,?,?,?)',(qid,status_map[action],session.get('user_id'),reason,note))
    family=c.execute('SELECT review_status,active FROM question_families WHERE family_key=?',(row['family_key'] or '',)).fetchone()
    family_live=bool(family and family['review_status']=='Approved' and int(family['active'] or 0)==1)
    c.commit(); c.close()
    if status_map[action]=='Approved' and not family_live:
        flash('Question approved, but it is NOT live until its question family is also Approved.','success')
    else:
        flash(f'Question moved to {status_map[action]}.','success')
    return redirect(url_for('admin_question_detail',qid=qid))

@app.route('/admin/questions/<int:qid>/edit',methods=['GET','POST'])
def admin_question_edit(qid):
    if not require('admin'): return redirect(url_for('login'))
    c=db(); q=c.execute('SELECT * FROM questions WHERE id=?',(qid,)).fetchone()
    if not q: c.close(); return redirect(url_for('admin_questions'))
    if request.method=='POST':
        old_version=int(q['question_version'] or 1)
        c.execute('INSERT OR IGNORE INTO question_versions(question_id,version,snapshot_json,change_note) VALUES(?,?,?,?)',(qid,old_version,json.dumps(question_snapshot(q)),request.form.get('change_note','Edited question')))
        new_version=old_version+1
        fields=['question','option_a','option_b','option_c','option_d','answer','explanation','chapter','topic','subtopic','learning_outcome','concept','difficulty','cognitive_skill','command_word','misconception_tags','prerequisite_tags','source_type']
        vals=[]
        for f in fields:
            v=request.form.get(f,'').strip()
            if f in {'misconception_tags','prerequisite_tags'}:
                v=json.dumps([x.strip() for x in v.split('|') if x.strip()])
            vals.append(v)
        new_learning_outcome=request.form.get('learning_outcome','').strip()
        new_concept=request.form.get('concept','').strip()
        c.execute(f"UPDATE questions SET {','.join(f+'=?' for f in fields)},question_version=?,review_status='Changes Required',status='Changes Required',reviewer='',reviewed_at='' WHERE id=?",vals+[new_version,qid])
        if (new_learning_outcome!=(q['learning_outcome'] or '').strip() or new_concept!=(q['concept'] or '').strip()) and (q['family_key'] or '').strip():
            c.execute("UPDATE question_families SET review_status='Changes Required',active=0,updated_at=CURRENT_TIMESTAMP WHERE family_key=?",(q['family_key'],))
        c.execute("UPDATE questions SET answer_config='{}',marking_config='{}' WHERE id=?",(qid,))
        c.execute('INSERT INTO question_review_events(question_id,action,reviewer_id,note) VALUES(?,?,?,?)',(qid,'Edited',session.get('user_id'),request.form.get('change_note','')))
        c.commit(); c.close(); flash(f'Question saved as version {new_version} and returned for review.','success'); return redirect(url_for('admin_question_detail',qid=qid))
    c.close(); return render_template('admin_question_edit.html',q=q)

@app.route('/admin/question-health')
def admin_question_health():
    if not require('admin'): return redirect(url_for('login'))
    c=db(); ids=[r['id'] for r in c.execute('SELECT id FROM questions WHERE COALESCE(active,1)=1 ORDER BY id DESC LIMIT 500').fetchall()]
    items=[question_health(c,i) for i in ids]; items=[x for x in items if x]
    priority={'Review recommended':0,'Collecting data':1,'Healthy':2}; items.sort(key=lambda x:(priority.get(x['status'],9),-(x['reports'] or 0),-(x['row']['attempts'] or 0)))
    c.close(); return render_template('admin_question_health.html',items=items)

@app.route('/admin/mastery-rules',methods=['GET','POST'])
def admin_mastery_rules():
    if not require('admin'): return redirect(url_for('login'))
    c=db()
    if request.method=='POST':
        level=request.form.get('mastery_level','')
        if level in MASTER_LEVELS:
            try:
                c.execute("""UPDATE mastery_policies SET min_forms=?,min_questions=?,min_accuracy=?,verification_days=?,external_percentile_target=?,
                  target_band_pct=?,unseen_family_pct=?,updated_at=CURRENT_TIMESTAMP WHERE mastery_level=?""",
                  (max(1,int(request.form.get('min_forms') or 1)),max(5,int(request.form.get('min_questions') or 10)),
                   max(0,min(100,float(request.form.get('min_accuracy') or 70))),max(1,int(request.form.get('verification_days') or 90)),
                   float(request.form.get('external_percentile_target')) if request.form.get('external_percentile_target','').strip() else None,
                   max(0,min(1,float(request.form.get('target_band_pct') or .25))),max(0,min(1,float(request.form.get('unseen_family_pct') or .6))),level))
                c.commit(); flash(f'{level} mastery rule updated.','success')
            except ValueError: flash('Check the mastery-rule numbers and try again.','error')
        c.close(); return redirect(url_for('admin_mastery_rules'))
    policies=c.execute("SELECT * FROM mastery_policies ORDER BY level_rank").fetchall(); c.close()
    return render_template('admin_mastery_rules.html',policies=policies)

@app.route('/admin/item-calibration',methods=['GET','POST'])
def admin_item_calibration():
    if not require('admin'): return redirect(url_for('login'))
    c=db()
    if request.method=='POST':
        qid=int(request.form.get('question_id') or 0); status=request.form.get('calibration_status','PROVISIONAL')
        allowed={'PROVISIONAL','COLLECTING','EMPIRICAL','LOW_INFORMATION','REVIEW_NEGATIVE','BLOCKED','DEMO'}
        q=c.execute("SELECT * FROM questions WHERE id=?",(qid,)).fetchone()
        if q and status in allowed:
            if int(q['is_demo'] or 0): status='DEMO'
            c.execute("UPDATE questions SET calibration_status=?,calibrated_at=? WHERE id=?",(status,datetime.now().isoformat(timespec='seconds'),qid)); c.commit()
            flash('Item calibration status updated.','success')
        c.close(); return redirect(url_for('admin_item_calibration'))
    items=c.execute("""SELECT id,question_id,subject,chapter,level,is_demo,response_count,facility_value,discrimination_value,calibration_status
      FROM questions WHERE COALESCE(active,1)=1 ORDER BY is_demo DESC,response_count DESC,id DESC LIMIT 500""").fetchall(); c.close()
    return render_template('admin_item_calibration.html',items=items)

@app.route('/admin/users/<int:user_id>/access',methods=['POST'])
def admin_user_access(user_id):
    if not require('admin'): return redirect(url_for('login'))
    code=request.form.get('access_code','free_access')
    if code not in ACCESS_CODES: code='free_access'
    c=db(); user=c.execute("SELECT id,role,access_override_code FROM users WHERE id=?",(user_id,)).fetchone()
    if not user or user['role']!='student':
        c.close(); flash('Student account not found.','error'); return redirect(url_for('admin_users'))
    previous=(user['access_override_code'] or '').strip()
    c.execute("UPDATE users SET access_override_code=? WHERE id=?",(code,user_id))
    c.execute("""INSERT INTO access_change_history(student_id,actor_user_id,source,previous_access_code,new_access_code,note)
      VALUES(?,?,?,?,?,?)""",(user_id,session.get('user_id'),'admin_override',previous,code,'Manual pilot/admin Access override'))
    c.commit(); c.close()
    flash(f"Student access set to {ACCESS_CODES[code]['name']} for local/pilot testing.",'success'); return redirect(url_for('admin_users'))

@app.route('/admin/users')
def admin_users():
    if not require('admin'): return redirect(url_for('login'))
    c=db(); users=c.execute('''SELECT u.*,i.name institution_name FROM users u LEFT JOIN institutions i ON i.id=u.primary_institution_id ORDER BY role,full_name''').fetchall(); c.close(); return render_template('admin_users.html',users=users)

@app.route('/admin/governance-audit')
def admin_governance_audit():
    if not require('admin'): return redirect(url_for('login'))
    c=db()
    mastery=c.execute("""SELECT mh.*,u.full_name FROM mastery_history mh
      LEFT JOIN users u ON u.id=mh.student_id ORDER BY mh.id DESC LIMIT 250""").fetchall()
    access=c.execute("""SELECT ah.*,u.full_name,actor.full_name actor_name FROM access_change_history ah
      LEFT JOIN users u ON u.id=ah.student_id LEFT JOIN users actor ON actor.id=ah.actor_user_id
      ORDER BY ah.id DESC LIMIT 250""").fetchall()
    c.close(); return render_template('admin_governance_audit.html',mastery_events=mastery,access_events=access)



@app.route('/admin/users/<int:user_id>/reset-password',methods=['POST'])
def admin_reset_user_password(user_id):
    if not require('admin'): return redirect(url_for('login'))
    password=request.form.get('new_password','').strip()
    if len(password)<8:
        flash('Use a password of at least 8 characters.','error'); return redirect(url_for('admin_users'))
    c=db(); c.execute("UPDATE users SET password_hash=? WHERE id=?",(generate_password_hash(password),user_id)); c.commit(); c.close()
    flash('Password reset successfully.','success')
    return redirect(url_for('admin_users'))

@app.route('/admin/users/<int:user_id>/status',methods=['POST'])
def admin_user_status(user_id):
    if not require('admin'): return redirect(url_for('login'))
    status=request.form.get('status','active')
    if status not in ('active','disabled'): status='active'
    c=db(); c.execute("UPDATE users SET account_status=? WHERE id=?",(status,user_id)); c.commit(); c.close()
    flash('Account status updated.','success')
    return redirect(url_for('admin_users'))

@app.route('/admin/users/create-demo',methods=['POST'])
def admin_create_demo_user():
    if not require('admin'): return redirect(url_for('login'))
    role=request.form.get('role','student')
    if role not in ('student','teacher'): role='student'
    username=request.form.get('username','').strip().lower()
    password=request.form.get('password','').strip()
    full_name=request.form.get('full_name','Demo Tester').strip()
    dob=request.form.get('dob','').strip()
    try:
        parsed_dob=datetime.fromisoformat(dob).date() if dob else None
        if not parsed_dob or parsed_dob>=datetime.now().date(): raise ValueError('invalid DOB')
    except Exception:
        flash('A valid past date of birth is required for safe pilot messaging tests.','error'); return redirect(url_for('admin_users'))
    if not username or len(password)<8:
        flash('Demo login/email and an 8+ character password are required.','error'); return redirect(url_for('admin_users'))
    c=db()
    try:
        email=username if '@' in username else f"{re.sub(r'[^a-z0-9]+','.',username).strip('.') or 'demo'}@scoremax.local"
        cur=c.execute("""INSERT INTO users(role,full_name,dob,email,username,password_hash,account_status,academic_level,login_provider,
          teacher_marketplace_pilot_enabled,academic_messages_pilot_enabled,is_demo_account)
          VALUES(?,?,?,?,?,?, 'active',?, 'password',?,?,1)""",(role,full_name,dob,email,username,generate_password_hash(password),
          request.form.get('academic_level','Matric Part 2' if role=='student' else ''),
          1 if request.form.get('teacher_marketplace_pilot_enabled')=='1' else 0,
          1 if request.form.get('academic_messages_pilot_enabled')=='1' else 0))
        uid=cur.lastrowid; c.execute("UPDATE users SET system_user_id=? WHERE id=?",(next_user_id(role,uid),uid)); c.commit()
        flash('Demo/test account created.','success')
    except sqlite3.IntegrityError:
        flash('That username already exists.','error')
    c.close()
    return redirect(url_for('admin_users'))

@app.route('/admin/institutions/<int:institution_id>')
def admin_institution_detail(institution_id):
    if not require('admin'): return redirect(url_for('login'))
    c=db(); data=institution_dashboard_data(c,institution_id)
    staff=c.execute("""SELECT ist.*,u.full_name,u.email,u.username FROM institution_staff ist JOIN users u ON u.id=ist.user_id
      WHERE ist.institution_id=? AND ist.active=1 ORDER BY u.full_name""",(institution_id,)).fetchall()
    candidates=c.execute("""SELECT id,full_name,email,username,role FROM users
      WHERE role IN ('teacher','admin') AND (primary_institution_id=? OR primary_institution_id IS NULL)
      ORDER BY full_name""",(institution_id,)).fetchall()
    c.close()
    if not data: return redirect(url_for('institutions'))
    return render_template('admin_institution_detail.html',staff=staff,candidates=candidates,**data)


@app.route('/admin/institutions/<int:institution_id>/staff',methods=['POST'])
def admin_institution_staff(institution_id):
    if not require('admin'): return redirect(url_for('login'))
    user_id=int(request.form.get('user_id') or 0); role=request.form.get('institution_role','viewer')
    if role not in ('viewer','manager','admin'): role='viewer'
    c=db()
    if not c.execute("SELECT 1 FROM institutions WHERE id=?",(institution_id,)).fetchone():
        c.close(); return redirect(url_for('institutions'))
    if not c.execute("SELECT 1 FROM users WHERE id=?",(user_id,)).fetchone():
        c.close(); flash('User not found.','error'); return redirect(url_for('admin_institution_detail',institution_id=institution_id))
    c.execute("""INSERT OR REPLACE INTO institution_staff(institution_id,user_id,institution_role,active)
      VALUES(?,?,?,1)""",(institution_id,user_id,role))
    c.commit(); c.close(); flash('Institution dashboard access updated.','success')
    return redirect(url_for('admin_institution_detail',institution_id=institution_id))


@app.route('/admin/institutions',methods=['GET','POST'])
def institutions():
    if not require('admin'): return redirect(url_for('login'))
    c=db()
    if request.method=='POST':
        c.execute('INSERT INTO institutions(institution_code,name,province,division,district,board,institution_type) VALUES(?,?,?,?,?,?,?)',(f'INST-{random.randint(100000,999999)}',request.form['name'],request.form.get('province',''),request.form.get('division',''),request.form.get('district',''),request.form.get('board',''),request.form.get('institution_type','College'))); c.commit(); flash('Institution added.','success')
    rows=c.execute('SELECT * FROM institutions WHERE active=1 ORDER BY province,division,district,name').fetchall(); c.close(); return render_template('institutions.html',institutions=rows)

# ---------------------------------------------------------------------------
# V6.2 Pilot Readiness, Content Intake, Prompt Bridge and Knowledge Hub
# ---------------------------------------------------------------------------

def pilot_record_event(c,event_type,subject_type='',subject_id='',metadata=None,actor_user_id=None):
    c.execute("""INSERT INTO pilot_activity_events(event_type,actor_user_id,subject_type,subject_id,metadata_json)
      VALUES(?,?,?,?,?)""",(event_type,actor_user_id,subject_type,str(subject_id or ''),json.dumps(metadata or {},sort_keys=True)))


def _pilot_backup(c,reason,actor_user_id=None):
    c.commit()
    code='BKP-'+datetime.now().strftime('%Y%m%d-%H%M%S')+'-'+secrets.token_hex(3).upper()
    folder=Path(os.environ.get('SCOREMAX_BACKUP_DIR',str(BASE/'pilot_backups'))).resolve(); folder.mkdir(parents=True,exist_ok=True)
    path=folder/f'{code}.db'
    ok,message=sqlite_backup(DB,path)
    digest=''; size=0
    if ok and path.exists():
        digest=hashlib.sha256(path.read_bytes()).hexdigest(); size=path.stat().st_size
    cur=c.execute("""INSERT INTO pilot_backups(backup_code,reason,file_path,file_sha256,file_size_bytes,integrity_status,created_by)
      VALUES(?,?,?,?,?,?,?)""",(code,reason,str(path),digest,size,'OK' if ok else 'FAILED',actor_user_id))
    c.commit()
    return {'id':cur.lastrowid,'code':code,'path':str(path),'sha256':digest,'size':size,'ok':ok,'message':message}


def _active_blueprint_readiness(c):
    bp=c.execute("""SELECT id,powerhouse_blueprint_id,blueprint_version FROM assessment_blueprints
      WHERE local_status='ACTIVE' ORDER BY activated_at DESC,id DESC LIMIT 1""").fetchone()
    if not bp: return {'available':False,'message':'No active assessment blueprint.'}
    report=blueprint_bank_sufficiency(c,bp['id'])
    return {'available':True,'blueprint_id':bp['id'],'source_id':bp['powerhouse_blueprint_id'],
            'version':bp['blueprint_version'],'ready':report['ready'],'subjects':report['subjects'],
            'blockers':report['blockers']}


def _norm_import_header(value):
    value=str(value or '').strip().casefold().replace('_',' ').replace('-',' ')
    value=re.sub(r'\s*/\s*',' / ',value)
    return re.sub(r'\s+',' ',value).strip()


_IMPORT_CANONICAL_ALIASES={
  'Question ID':('question id','question_id','canonical content id','external question id'),
  'Family ID':('family id','question family id','family_id'),
  'Variant':('variant','variant id','question variant'),
  'Programme':('programme','program','academic level'),
  'Country':('country','market'),
  'Qualification':('qualification','qualification level'),
  'Exam Board':('exam board','board'),
  'Curriculum Version':('curriculum version','syllabus version','framework version'),
  'Subject':('subject',),
  'Chapter':('chapter','chapter / unit','unit'),
  'Chapter Number':('chapter number','chapter no','chapter #'),
  'Chapter Name':('chapter name','chapter title'),
  'Topic':('topic',),
  'Sub-topic':('sub topic','subtopic','sub-topic'),
  'Type':('type','question type','question_type','item type'),
  'Question':('question','question / task','question/task','question text','stem','prompt'),
  'Answer':('answer','key answer','correct answer','answer key','key'),
  'Explanation':('explanation','explanation / marking rubric','explanation/marking rubric','marking rubric','rubric','rationale'),
  'Level':('level','mastery','mastery level'),
  'Difficulty':('difficulty','difficulty band'),
  'Learning Outcome':('learning outcome','chapter outcome'),
  'Concept':('concept','mastery claim'),
  'Concept ID':('concept id','knowledge node id','mastery node id'),
  'Cognitive Skill':('cognitive skill','cognitive demand'),
  'Command Word':('command word',),
  'Marks':('marks','mark'),
  'Estimated Time Seconds':('estimated time seconds','time seconds','estimated seconds'),
  'Rights Status':('rights status','rights'),
  'ScoreMax Ready':('scoremax ready','scoremax-ready','release ready','release readiness'),
  'Assessment Purpose':('assessment purpose','question purpose'),
  'Difficulty Source':('difficulty source',),
  'Source Type':('source type','source'),
  'Secure Bank':('secure bank',),
  'Misconception Tags':('misconception tags','misconceptions'),
  'Prerequisite Tags':('prerequisite tags','prerequisites'),
  'Status':('status',),
  'Review Status':('review status',),
  'R2 Status':('r2 status','reviewer 2 status'),
  'Readiness':('readiness',),
  'Release Status':('release status',),
  'Stimulus / Context':('stimulus / context','stimulus/context','stimulus','context'),
  'Statements / Options':('statements / options','statements/options','options','options / statements'),
  'Parent Seed ID':('parent seed id','seed id','reasoning seed id'),
  'Dependency Type':('dependency type','dependency','evidence role'),
  'Independent Mastery Weight':('independent mastery weight','independent weight','mastery weight'),
  'Architecture Question ID':('architecture question id',),
  'Architecture Layer':('architecture layer','learner layer'),
  'Transfer Level':('transfer level',),
  'Delivery Context':('delivery context',),
}


def _parse_import_options(text):
    text=str(text or '').strip()
    if not text: return {}
    found={}
    # Supports common A. / A) / A: / Option A forms, one per line.
    for line in text.replace('\r\n','\n').replace('\r','\n').split('\n'):
        m=re.match(r'^\s*(?:option\s+)?([A-D])\s*[\).:\-]\s*(.+?)\s*$',line,re.I)
        if m: found[m.group(1).upper()]=m.group(2).strip()
    return found if len(found)>=2 else {}


def _normalize_import_row(row):
    """Add canonical intake keys without deleting unfamiliar source columns.

    This is deliberately conservative: it recognizes equivalent headers but never
    invents IDs, mastery relationships, difficulty or approval state.
    """
    out=dict(row)
    by_norm={_norm_import_header(k):k for k in row.keys() if not str(k).startswith('_') and not str(k).startswith('__')}
    for canonical,aliases in _IMPORT_CANONICAL_ALIASES.items():
        if str(out.get(canonical,'') or '').strip():
            continue
        for alias in aliases:
            src=by_norm.get(_norm_import_header(alias))
            if src is not None and row.get(src) not in (None,''):
                out[canonical]=row.get(src); break
    # Preserve direct A-D columns and parse combined options only when explicit columns are absent.
    for letter in 'ABCD':
        if out.get(letter) in (None,''):
            for alias in (letter, f'Option {letter}', f'option_{letter.lower()}'):
                src=by_norm.get(_norm_import_header(alias))
                if src is not None and row.get(src) not in (None,''):
                    out[letter]=row.get(src); break
    if not any(str(out.get(x,'') or '').strip() for x in 'ABCD'):
        out.update(_parse_import_options(out.get('Statements / Options','')))
    return out


def _read_import_upload(upload):
    rows=[]; raw_bytes=upload.read(); file_type='CSV'
    filename=upload.filename or 'content-import'
    if filename.lower().endswith('.xlsx'):
        file_type='XLSX'
        # Reuse the already-proven worksheet detector: instructional/reference sheets
        # are ignored when question-bearing sheets can be identified.
        detected=reviewer_workspace.parse_upload(filename,raw_bytes)
        for idx,row in enumerate(detected,2):
            raw=dict(row)
            raw['_sheet']=raw.get('__source_sheet') or raw.get('_sheet') or 'Imported questions'
            raw['_row']=raw.get('__source_row') or raw.get('_row') or idx
            rows.append(_normalize_import_row(raw))
    else:
        decoded=raw_bytes.decode('utf-8-sig')
        parsed=list(csv.DictReader(io.StringIO(decoded)))
        for idx,row in enumerate(parsed,start=2):
            row['_sheet']='CSV'; row['_row']=idx; rows.append(_normalize_import_row(row))
    return filename,file_type,raw_bytes,rows


def _insert_import_question(c,row,batch_id):
    def val(key,default=''): return str(row.get(key,default) or default).strip()
    proposed_family_key=canonical_family_key(val('Family ID'),val('Country','Pakistan'),val('Qualification',val('Programme')),val('Exam Board'),val('Curriculum Version'),val('Programme'),val('Subject'))
    family_preexisting=bool(c.execute("SELECT 1 FROM question_families WHERE family_key=?",(proposed_family_key,)).fetchone())
    family_key=upsert_question_family(c,row,review_status='Draft',active=0)
    if not family_key: raise ValueError('Family ID is required.')
    if not family_preexisting:
        c.execute("UPDATE question_families SET source_import_batch_id=? WHERE family_key=?",(batch_id,family_key))
    qtype=val('Type','MCQ'); level=val('Level','Foundation'); difficulty=val('Difficulty')
    canonical_type=canonical_question_type({'qtype':qtype})
    options=[{'id':code,'text':val(code)} for code in ('A','B','C','D') if val(code)]
    if canonical_type=='true_false' and not options:
        options=[{'id':'A','text':'True'},{'id':'B','text':'False'}]
    answer_value=val('Answer')
    answer_config={'options':options} if canonical_type in {'single_choice','true_false','multiple_select'} else ({'accepted_answers':[answer_value],'case_sensitive':False,'trim_spaces':True} if canonical_type=='fill_blank' else {})
    marking_config={'marks':float(row.get('Marks',1) or 1)}
    if canonical_type in {'single_choice','true_false'}: marking_config['correct_option_ids']=[answer_value]
    elif canonical_type=='multiple_select': marking_config['correct_option_ids']=[x.strip() for x in answer_value.split('|') if x.strip()]
    elif canonical_type=='numerical': marking_config.update({'correct_value':answer_value,'tolerance':str(row.get('Numerical Tolerance','0') or '0')})
    stimulus_text=val('Stimulus / Context')
    stimulus_data=json.dumps({'text':stimulus_text}) if stimulus_text else '{}'
    columns=[
      'question_id','family_id','variant','programme','subject','chapter','topic','subtopic','qtype','level','question',
      'option_a','option_b','option_c','option_d','answer','explanation','status','country','qualification','exam_board',
      'curriculum_version','learning_outcome','concept','concept_id','capsule_id','misconception_id','difficulty','cognitive_skill',
      'command_word','marks','estimated_time_seconds','stimulus_type','stimulus_data','misconception_tags','prerequisite_tags','source_type','review_status',
      'secure_bank','family_key','active','rights_status','scoremax_ready','assessment_purpose','difficulty_source',
      'source_import_batch_id','content_environment','answer_config','marking_config','feedback_config']
    values=[
      val('Question ID'),val('Family ID'),val('Variant'),val('Programme'),val('Subject'),val('Chapter'),val('Topic'),val('Sub-topic'),
      qtype,level,val('Question'),val('A'),val('B'),val('C'),val('D'),val('Answer'),val('Explanation'),'Draft',
      val('Country','Pakistan'),val('Qualification',val('Programme')),val('Exam Board'),val('Curriculum Version'),val('Learning Outcome'),
      val('Concept'),val('Concept ID'),val('Capsule ID'),val('Misconception ID'),difficulty,val('Cognitive Skill'),val('Command Word'),
      float(row.get('Marks',1) or 1),int(float(row.get('Estimated Time Seconds',60) or 60)),
      'text' if stimulus_text else '',stimulus_data,
      json.dumps([x.strip() for x in val('Misconception Tags').split('|') if x.strip()]),
      json.dumps([x.strip() for x in val('Prerequisite Tags').split('|') if x.strip()]),
      val('Source Type','Power House Import'),'Draft',1 if val('Secure Bank').lower() in {'1','yes','true','secure'} else 0,
      family_key,0,val('Rights Status'),1 if val('ScoreMax Ready','No').lower() in {'1','yes','true'} else 0,
      val('Assessment Purpose','practice|test|mock|mastery'),val('Difficulty Source','authoring'),batch_id,'CANDIDATE',json.dumps(answer_config),json.dumps(marking_config),
      json.dumps({'import_batch_id':batch_id,'source_sheet':row.get('Source Worksheet') or row.get('_sheet') or row.get('__source_sheet'),'source_row':row.get('Source Row') or row.get('_row') or row.get('__source_row')})]
    placeholders=','.join('?' for _ in columns)
    cur=c.execute(f"INSERT INTO questions({','.join(columns)}) VALUES({placeholders})",values)
    qid=cur.lastrowid
    ensure_curriculum_mapping_for_question(c,qid,row)
    c.execute("""INSERT INTO question_review_events(question_id,action,reviewer_id,note)
      VALUES(?,?,?,?)""",(qid,'Imported',session.get('user_id'),f'V6.4 governed batch {batch_id}: Draft + inactive; spreadsheet approval ignored.'))
    c.execute("""INSERT INTO curriculum(programme,subject,chapter,topic,subtopic) SELECT ?,?,?,?,? WHERE NOT EXISTS(
      SELECT 1 FROM curriculum WHERE programme=? AND subject=? AND chapter=? AND topic=? AND COALESCE(subtopic,'')=?)""",
      (val('Programme'),val('Subject'),val('Chapter'),val('Topic'),val('Sub-topic'),
       val('Programme'),val('Subject'),val('Chapter'),val('Topic'),val('Sub-topic')))
    upsert_chapter_catalogue(c,val('Programme'),val('Subject'),val('Chapter'),val('Chapter Number'),val('Chapter Name'),
      metadata_source='GOVERNED_IMPORT' if (val('Chapter Number') or val('Chapter Name')) else 'SOURCE_LABEL',review_status='Candidate')
    # Preserve explicit universal-mastery relationship metadata if and only if it was supplied.
    # No Node/Family/Seed identity is fabricated for legacy rows.
    lineage_fields=('Architecture Question ID','Parent Seed ID','Dependency Type','Independent Mastery Weight','Architecture Layer','Transfer Level','Delivery Context')
    if any(val(k) for k in lineage_fields):
        raw_purpose=val('Assessment Purpose','SUBJECT_MASTERY').split('|')[0].upper().replace(' ','_')
        purpose=raw_purpose if raw_purpose in universal_mastery.PURPOSES else 'SUBJECT_MASTERY'
        payload={
          'architecture_question_id':val('Architecture Question ID') or '',
          'question_db_id':qid,'external_question_id':val('Question ID'),
          'purpose':purpose,
          'architecture_layer':val('Architecture Layer','L2_UNDERSTAND'),
          'pedagogical_type':qtype,'independent_mastery_weight':float(val('Independent Mastery Weight','0') or 0),
          'parent_seed_id':val('Parent Seed ID'),'dependency_type':val('Dependency Type','DEPENDENT'),
          'transfer_level':val('Transfer Level','NEAR_COPY'),'delivery_context':val('Delivery Context','BLOCKED'),
          'environment':'QA_SANDBOX_ONLY','version':'1','status':'CANDIDATE'
        }
        # This is intentionally candidate/QA until a governed live mapping package exists.
        universal_mastery.upsert_question_architecture(c,payload)
    return qid


@app.route('/admin/pilot-readiness')
def admin_pilot_readiness():
    if not require('admin'): return redirect(url_for('login'))
    c=db(); readiness=_active_blueprint_readiness(c)
    controls=c.execute("SELECT * FROM pilot_feature_controls ORDER BY name").fetchall()
    imports=c.execute("SELECT * FROM content_import_batches ORDER BY id DESC LIMIT 12").fetchall()
    prompts=c.execute("SELECT * FROM powerhouse_prompt_packs ORDER BY id DESC LIMIT 8").fetchall()
    feedback=c.execute("SELECT routing_target,status,COUNT(*) n FROM pilot_feedback GROUP BY routing_target,status").fetchall()
    backups=c.execute("SELECT * FROM pilot_backups ORDER BY id DESC LIMIT 8").fetchall()
    failed_jobs=c.execute("SELECT wpj.*,wa.student_id FROM written_processing_jobs wpj LEFT JOIN written_attempts wa ON wa.id=wpj.attempt_id WHERE wpj.state LIKE 'FAILED%' ORDER BY wpj.id DESC LIMIT 20").fetchall()
    demos={'users':c.execute("SELECT COUNT(*) n FROM users WHERE COALESCE(is_demo_account,0)=1 AND account_status<>'archived_demo'").fetchone()['n'],
           'questions':c.execute("SELECT COUNT(*) n FROM questions WHERE COALESCE(is_demo,0)=1 AND COALESCE(active,0)=1").fetchone()['n'],
           'attempts':c.execute("SELECT COUNT(*) n FROM attempts WHERE assessment_kind='demo_progress'").fetchone()['n']}
    c.close(); return render_template('admin_pilot_readiness.html',readiness=readiness,controls=controls,imports=imports,prompts=prompts,feedback=feedback,backups=backups,demos=demos,failed_jobs=failed_jobs)


@app.route('/admin/pilot-readiness/feature/<feature_code>',methods=['POST'])
def admin_pilot_feature_update(feature_code):
    if not require('admin'): return redirect(url_for('login'))
    state=request.form.get('state','PILOT').upper()
    if state not in {'HIDDEN','PILOT','LIVE','PAUSED'}: state='PILOT'
    c=db(); c.execute("UPDATE pilot_feature_controls SET state=?,updated_by=?,updated_at=CURRENT_TIMESTAMP WHERE feature_code=?",(state,session['user_id'],feature_code))
    pilot_record_event(c,'PILOT_FEATURE_UPDATED','pilot_feature',feature_code,{'state':state},session['user_id']); c.commit(); c.close()
    flash('Pilot feature state updated.','success'); return redirect(url_for('admin_pilot_readiness'))


@app.route('/admin/pilot-readiness/backup',methods=['POST'])
def admin_pilot_backup_create():
    if not require('admin'): return redirect(url_for('login'))
    c=db(); result=_pilot_backup(c,request.form.get('reason','Manual pre-pilot backup'),session['user_id']); c.close()
    flash(f"Backup {result['code']} created and integrity checked." if result['ok'] else f"Backup failed: {result['message']}",'success' if result['ok'] else 'error')
    return redirect(url_for('admin_pilot_readiness'))


@app.route('/admin/pilot-readiness/backup/<int:backup_id>/download')
def admin_pilot_backup_download(backup_id):
    if not require('admin'): return redirect(url_for('login'))
    c=db(); row=c.execute("SELECT * FROM pilot_backups WHERE id=? AND integrity_status='OK'",(backup_id,)).fetchone(); c.close()
    if not row: abort(404)
    path=Path(row['file_path']).resolve(); root=Path(os.environ.get('SCOREMAX_BACKUP_DIR',str(BASE/'pilot_backups'))).resolve()
    if root not in path.parents or not path.exists(): abort(404)
    return send_file(path,as_attachment=True,download_name=f"{row['backup_code']}.db")


@app.route('/admin/pilot-readiness/demo-cleanup',methods=['POST'])
def admin_demo_cleanup():
    if not require('admin'): return redirect(url_for('login'))
    if request.form.get('confirmation','').strip()!='ARCHIVE DEMO DATA':
        flash('Type ARCHIVE DEMO DATA exactly. This safeguard prevents accidental pilot-data cleanup.','error'); return redirect(url_for('admin_pilot_readiness'))
    c=db(); backup=_pilot_backup(c,'Automatic backup before V6.2 demo-data quarantine',session['user_id'])
    if not backup['ok']: c.close(); flash('Demo cleanup blocked because the safety backup failed.','error'); return redirect(url_for('admin_pilot_readiness'))
    demo_users=[r['id'] for r in c.execute("SELECT id FROM users WHERE COALESCE(is_demo_account,0)=1 AND role<>'admin'").fetchall()]
    demo_questions=[r['id'] for r in c.execute("SELECT id FROM questions WHERE COALESCE(is_demo,0)=1").fetchall()]
    demo_attempts=[r['id'] for r in c.execute("SELECT id FROM attempts WHERE assessment_kind='demo_progress' OR student_id IN (%s)"%(','.join('?'*len(demo_users)) if demo_users else 'NULL'),demo_users).fetchall()]
    if demo_attempts:
        q=','.join('?'*len(demo_attempts)); c.execute(f"DELETE FROM attempt_answers WHERE attempt_id IN ({q})",demo_attempts); c.execute(f"DELETE FROM attempts WHERE id IN ({q})",demo_attempts)
    if demo_users:
        q=','.join('?'*len(demo_users)); c.execute(f"DELETE FROM assessment_sessions WHERE student_id IN ({q})",demo_users)
        for table in ('mastery_form_results','mastery_records','mastery_history','study_plan_activities','study_plans','student_learning_states','student_misconceptions','recall_items','student_blueprint_projections','weekly_progress_reports'):
            if 'student_id' in table_columns(c,table): c.execute(f"DELETE FROM {table} WHERE student_id IN ({q})",demo_users)
        c.execute(f"UPDATE users SET account_status='archived_demo',session_version=COALESCE(session_version,0)+1 WHERE id IN ({q})",demo_users)
    if demo_questions:
        q=','.join('?'*len(demo_questions)); c.execute(f"UPDATE questions SET active=0,status='Archived',review_status='Archived',scoremax_ready=0,content_environment='DEMO' WHERE id IN ({q})",demo_questions)
    report={'demo_users_archived':len(demo_users),'demo_questions_archived':len(demo_questions),'demo_attempts_removed':len(demo_attempts)}
    code='DCL-'+datetime.now().strftime('%Y%m%d%H%M%S')
    c.execute("""INSERT INTO demo_cleanup_runs(run_code,backup_record_id,demo_users_count,demo_questions_count,demo_attempts_count,status,report_json,executed_by,executed_at)
      VALUES(?,?,?,?,?,'COMPLETED',?,?,?)""",(code,backup['id'],len(demo_users),len(demo_questions),len(demo_attempts),json.dumps(report),session['user_id'],datetime.now().isoformat(timespec='seconds')))
    pilot_record_event(c,'DEMO_DATA_QUARANTINED','demo_cleanup',code,report,session['user_id']); c.commit(); c.close()
    flash(f"Demo data quarantined safely: {len(demo_users)} account(s), {len(demo_questions)} question(s), {len(demo_attempts)} attempt(s).",'success')
    return redirect(url_for('admin_pilot_readiness'))



@app.route('/admin/pilot-readiness/job/<int:job_id>/retry',methods=['POST'])
def admin_pilot_job_retry(job_id):
    if not require('admin'): return redirect(url_for('login'))
    c=db(); job=c.execute("SELECT * FROM written_processing_jobs WHERE id=?",(job_id,)).fetchone()
    if not job or not str(job['state'] or '').startswith('FAILED'):
        c.close(); flash('Only a failed retryable processing job can be re-queued.','error'); return redirect(url_for('admin_pilot_readiness'))
    c.execute("""UPDATE written_processing_jobs SET state='QUEUED_RETRY',retry_count=COALESCE(retry_count,0)+1,
      error_message='',updated_at=CURRENT_TIMESTAMP WHERE id=?""",(job_id,))
    pilot_record_event(c,'FAILED_JOB_REQUEUED','written_processing_job',job_id,{'previous_state':job['state'],'retry_count':int(job['retry_count'] or 0)+1},session['user_id']); c.commit(); c.close()
    flash('Failed processing job re-queued with its original idempotency key.','success'); return redirect(url_for('admin_pilot_readiness'))

@app.route('/admin/powerhouse-prompt-bridge')
def admin_powerhouse_prompt_bridge():
    if not require('admin'): return redirect(url_for('login'))
    selected_id=int(request.args.get('pack_id') or 0)
    c=db(); packs=c.execute("SELECT * FROM powerhouse_prompt_packs ORDER BY id DESC").fetchall()
    selected=c.execute("SELECT * FROM powerhouse_prompt_packs WHERE id=?",(selected_id,)).fetchone() if selected_id else (packs[0] if packs else None)
    batches=c.execute("""SELECT pgb.*,pp.prompt_pack_id,pp.prompt_pack_version FROM powerhouse_generation_batches pgb
      JOIN powerhouse_prompt_packs pp ON pp.id=pgb.prompt_pack_db_id ORDER BY pgb.id DESC LIMIT 30""").fetchall(); c.close()
    return render_template('admin_prompt_bridge.html',packs=packs,selected=selected,batches=batches)


@app.route('/admin/powerhouse-prompt-bridge/import',methods=['POST'])
def admin_prompt_pack_import():
    if not require('admin'): return redirect(url_for('login'))
    upload=request.files.get('prompt_pack_file')
    if not upload or not upload.filename: flash('Choose a Power House prompt-pack JSON file.','error'); return redirect(url_for('admin_powerhouse_prompt_bridge'))
    try: payload=json.loads(upload.read().decode('utf-8-sig'))
    except Exception as exc: flash(f'Prompt pack could not be read: {exc}','error'); return redirect(url_for('admin_powerhouse_prompt_bridge'))
    prompt_secret=os.environ.get('SCOREMAX_POWERHOUSE_PROMPT_SECRET','').strip(); require_signature=SCOREMAX_ENV=='production' or os.environ.get('SCOREMAX_REQUIRE_SIGNED_PROMPT_PACKS','0')=='1'
    report=validate_prompt_pack(payload,shared_secret=prompt_secret,require_signature=require_signature)
    if not report['valid']: flash('Prompt pack rejected: '+' '.join(report['errors'][:4]),'error'); return redirect(url_for('admin_powerhouse_prompt_bridge'))
    c=db(); existing=c.execute("SELECT * FROM powerhouse_prompt_packs WHERE prompt_pack_id=? AND prompt_pack_version=?",(report['prompt_pack_id'],report['prompt_pack_version'])).fetchone()
    if existing:
        c.close(); flash('That exact prompt pack is already present.' if existing['payload_checksum']==report['checksum'] else 'Rejected: same prompt-pack ID/version has different content. Power House must issue a new version.','info' if existing['payload_checksum']==report['checksum'] else 'error')
        return redirect(url_for('admin_powerhouse_prompt_bridge',pack_id=existing['id']))
    cur=c.execute("""INSERT INTO powerhouse_prompt_packs(prompt_pack_id,prompt_pack_version,source_status,local_status,framework,framework_version,subject,chapter,
      learning_outcome_ids_json,source_evidence_ids_json,prompt_text,expected_output_schema_json,immutable_payload_json,payload_checksum,imported_by)
      VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",(report['prompt_pack_id'],report['prompt_pack_version'],str(payload.get('status','')), 'READY_TO_COPY',
      str(payload.get('framework','')),str(payload.get('framework_version','')),str(payload.get('subject','')),str(payload.get('chapter','')),
      json.dumps(payload.get('learning_outcome_ids',[])),json.dumps(payload.get('source_evidence_ids',[])),str(payload.get('prompt_text','')),
      json.dumps(payload.get('expected_output_schema',{})),canonical_json(payload),report['checksum'],session['user_id']))
    pilot_record_event(c,'PROMPT_PACK_IMPORTED','prompt_pack',cur.lastrowid,{'prompt_pack_id':report['prompt_pack_id'],'warnings':report['warnings']},session['user_id']); c.commit(); c.close()
    flash('Approved provider-neutral prompt pack imported. Power House remains the academic authority.','success'); return redirect(url_for('admin_powerhouse_prompt_bridge',pack_id=cur.lastrowid))


@app.route('/admin/powerhouse-prompt-bridge/<int:pack_id>/copy',methods=['POST'])
def admin_prompt_pack_copy(pack_id):
    if not require('admin'): return redirect(url_for('login'))
    c=db(); c.execute("UPDATE powerhouse_prompt_packs SET copied_count=copied_count+1,last_copied_at=? WHERE id=?",(datetime.now().isoformat(timespec='seconds'),pack_id))
    pilot_record_event(c,'PROMPT_PACK_COPIED','prompt_pack',pack_id,{},session['user_id']); c.commit(); c.close()
    flash('Copy event recorded. Use the Copy Prompt button to place the complete text on your clipboard.','success'); return redirect(url_for('admin_powerhouse_prompt_bridge',pack_id=pack_id))


@app.route('/admin/powerhouse-prompt-bridge/<int:pack_id>/submit',methods=['POST'])
def admin_generation_output_submit(pack_id):
    if not require('admin'): return redirect(url_for('login'))
    provider=request.form.get('provider','').strip(); model=request.form.get('model','').strip(); raw=request.form.get('generated_output','').strip()
    upload=request.files.get('generated_output_file')
    if upload and upload.filename: raw=upload.read().decode('utf-8-sig')
    if not provider or not raw: flash('Provider and generated JSON output are required.','error'); return redirect(url_for('admin_powerhouse_prompt_bridge',pack_id=pack_id))
    c=db(); pack=c.execute("SELECT * FROM powerhouse_prompt_packs WHERE id=?",(pack_id,)).fetchone()
    if not pack: c.close(); abort(404)
    prompt_payload=safe_json(pack['immutable_payload_json'],{})
    report=parse_generation_output(raw,dict(pack))
    cur=c.execute("""INSERT INTO powerhouse_generation_batches(prompt_pack_db_id,provider,model,provider_run_id,raw_output,parsed_output_json,
      validation_status,validation_report_json,item_count,submitted_by) VALUES(?,?,?,?,?,?,?,?,?,?)""",(pack_id,provider,model,request.form.get('provider_run_id','').strip(),raw,
      json.dumps(report['parsed']) if report['parsed'] is not None else '{}','VALIDATED_CANDIDATE' if report['valid'] else 'REJECTED_LOCAL_CHECK',json.dumps({'errors':report['errors'],'warnings':report['warnings']}),report['item_count'],session['user_id']))
    pilot_record_event(c,'MANUAL_AI_OUTPUT_RECEIVED','generation_batch',cur.lastrowid,{'provider':provider,'valid':report['valid'],'item_count':report['item_count']},session['user_id']); c.commit(); c.close()
    flash(f"Candidate output stored: {report['item_count']} item(s). It still requires Power House validation and academic approval." if report['valid'] else 'Output stored but failed local structural checks: '+' '.join(report['errors'][:3]),'success' if report['valid'] else 'error')
    return redirect(url_for('admin_powerhouse_prompt_bridge',pack_id=pack_id))


@app.route('/admin/powerhouse-prompt-bridge/batch/<int:batch_id>/export')
def admin_generation_batch_export(batch_id):
    if not require('admin'): return redirect(url_for('login'))
    c=db(); row=c.execute("""SELECT pgb.*,pp.prompt_pack_id,pp.prompt_pack_version,pp.immutable_payload_json FROM powerhouse_generation_batches pgb
      JOIN powerhouse_prompt_packs pp ON pp.id=pgb.prompt_pack_db_id WHERE pgb.id=?""",(batch_id,)).fetchone()
    if not row: c.close(); abort(404)
    parsed=safe_json(row['parsed_output_json'],{})
    transport=generation_transport({'prompt_pack_id':row['prompt_pack_id'],'prompt_pack_version':row['prompt_pack_version']},dict(row),parsed)
    c.execute("UPDATE powerhouse_generation_batches SET export_status='EXPORTED_TO_POWER_HOUSE',exported_at=? WHERE id=?",(datetime.now().isoformat(timespec='seconds'),batch_id)); c.commit(); c.close()
    buf=io.BytesIO(json.dumps(transport,indent=2,ensure_ascii=False).encode('utf-8'))
    return send_file(buf,as_attachment=True,download_name=f"{safe_filename(row['prompt_pack_id'])}-manual-ai-return-{batch_id}.json",mimetype='application/json')


@app.route('/admin/import',methods=['GET','POST'])
def admin_import():
    if not require('admin'): return redirect(url_for('login'))
    c=db()
    if request.method=='POST':
        upload=request.files.get('file')
        if not upload or not upload.filename: c.close(); flash('Choose an Excel or CSV file.','error'); return redirect(url_for('admin_import'))
        extension=Path(upload.filename).suffix.lower()
        if extension not in {'.xlsx','.csv'}:
            c.close(); flash('Only .xlsx and .csv question-bank files are accepted.','error'); return redirect(url_for('admin_import'))
        try: filename,file_type,raw_bytes,rows=_read_import_upload(upload)
        except Exception as exc: c.close(); flash(f'Import file could not be read: {exc}','error'); return redirect(url_for('admin_import'))
        if not rows:
            c.close(); flash('The import contains no question rows.','error'); return redirect(url_for('admin_import'))
        intake_mode=(request.form.get('intake_mode') or request.args.get('mode') or 'STANDARD').strip().upper()
        if intake_mode not in {'STANDARD','EMERGENCY_DIRECT'}: intake_mode='STANDARD'
        max_rows=3000 if intake_mode=='EMERGENCY_DIRECT' else int(os.environ.get('SCOREMAX_MAX_IMPORT_ROWS','10000') or 10000)
        if len(rows)>max_rows:
            c.close(); flash(f'Import contains {len(rows)} rows; the governed limit for this intake mode is {max_rows}.','error'); return redirect(url_for('admin_import',mode='emergency' if intake_mode=='EMERGENCY_DIRECT' else None))
        seen=set(); checks=[]
        for i,row in enumerate(rows):
            check=validate_import_row(row,row.get('_row',i+2)); qid=check['question_id']
            if qid!='(blank)' and qid in seen: check['errors'].append(f'Duplicate Question ID inside this file: {qid}')
            if qid!='(blank)' and c.execute("SELECT 1 FROM questions WHERE question_id=?",(qid,)).fetchone(): check['errors'].append(f'Question ID already exists in ScoreMax: {qid}')
            seen.add(qid); checks.append(check)
        errors=sum(bool(x['errors']) for x in checks); warnings=sum(bool(x['warnings']) for x in checks); valid=len(checks)-errors
        code='IMP-'+datetime.now().strftime('%Y%m%d%H%M%S')+'-'+secrets.token_hex(2).upper()
        configured=os.environ.get('SCOREMAX_CONTENT_INTAKE_DIR','').strip()
        if SCOREMAX_ENV=='production' and not configured:
            c.close(); flash('Production content intake requires SCOREMAX_CONTENT_INTAKE_DIR to point to protected storage.','error'); return redirect(url_for('admin_import'))
        intake_dir=Path(configured) if configured else BASE/'content_intake_uploads'; intake_dir=intake_dir.resolve(); intake_dir.mkdir(parents=True,exist_ok=True)
        digest=hashlib.sha256(raw_bytes).hexdigest(); source_path=intake_dir/f"{code}-{digest[:12]}{extension}"; source_path.write_bytes(raw_bytes)
        report={'errors':errors,'warnings':warnings,'rows':len(checks),'atomic_import_required':True,'source_file_preserved':True}
        cur=c.execute("""INSERT INTO content_import_batches(batch_code,source_system,source_prompt_pack_id,source_prompt_pack_version,filename,file_type,payload_checksum,
          row_count,valid_count,error_count,warning_count,status,validation_report_json,compatibility_before_json,imported_by,source_file_path,intake_mode)
          VALUES(?,?,?,?,?,?,?,?,?,?,?,'PREVIEWED',?,?,?,?,?)""",(code,request.form.get('source_system','MANUAL_FILE'),request.form.get('source_prompt_pack_id','').strip(),
          request.form.get('source_prompt_pack_version','').strip(),filename,file_type,digest,len(checks),valid,errors,warnings,json.dumps(report),json.dumps(_active_blueprint_readiness(c)),session['user_id'],str(source_path),intake_mode))
        batch_id=cur.lastrowid
        for check in checks:
            clean={k:v for k,v in check['row'].items() if not k.startswith('_')}
            clean['Source Worksheet']=str(check['row'].get('_sheet','') or '')
            clean['Source Row']=int(check['row_number'])
            c.execute("""INSERT INTO content_import_batch_rows(batch_id,row_number,sheet_name,question_id,row_json,errors_json,warnings_json)
              VALUES(?,?,?,?,?,?,?)""",(batch_id,int(check['row_number']),str(check['row'].get('_sheet','')),check['question_id'],json.dumps(clean,default=str),json.dumps(check['errors']),json.dumps(check['warnings'])))
        pilot_record_event(c,'CONTENT_IMPORT_PREVIEWED','content_import_batch',batch_id,report,session['user_id']); c.commit()
        batch=c.execute("SELECT * FROM content_import_batches WHERE id=?",(batch_id,)).fetchone(); batch_rows=c.execute("SELECT * FROM content_import_batch_rows WHERE batch_id=? ORDER BY id",(batch_id,)).fetchall(); c.close()
        return render_template('import_validation.html',batch=batch,checks=batch_rows,total=len(checks),valid=valid,errors=errors,warnings=warnings,compat_before=safe_json(batch['compatibility_before_json'],{}),compat_after={},candidate_summary=[])
    batches=c.execute("SELECT * FROM content_import_batches ORDER BY id DESC LIMIT 30").fetchall(); c.close()
    emergency=request.args.get('mode','').strip().lower()=='emergency'
    return render_template('import.html',batches=batches,emergency=emergency)



@app.route('/admin/import/batch/<int:batch_id>/source')
def admin_import_batch_source(batch_id):
    if not require('admin'): return redirect(url_for('login'))
    c=db(); row=c.execute("SELECT * FROM content_import_batches WHERE id=?",(batch_id,)).fetchone(); c.close()
    if not row or not row['source_file_path']: abort(404)
    path=Path(row['source_file_path']).resolve(); configured=os.environ.get('SCOREMAX_CONTENT_INTAKE_DIR','').strip()
    root=(Path(configured) if configured else BASE/'content_intake_uploads').resolve()
    if root not in path.parents or not path.exists() or hashlib.sha256(path.read_bytes()).hexdigest()!=row['payload_checksum']: abort(404)
    return send_file(path,as_attachment=True,download_name=row['filename'])

@app.route('/admin/import/batch/<int:batch_id>')
def admin_import_batch_detail(batch_id):
    if not require('admin'): return redirect(url_for('login'))
    c=db(); batch=c.execute("SELECT * FROM content_import_batches WHERE id=?",(batch_id,)).fetchone()
    if not batch: c.close(); abort(404)
    rows=c.execute("SELECT * FROM content_import_batch_rows WHERE batch_id=? ORDER BY id",(batch_id,)).fetchall()
    candidate_summary=c.execute("""SELECT subject,COUNT(*) questions,COUNT(DISTINCT family_key) families,
      SUM(CASE WHEN scoremax_ready=1 THEN 1 ELSE 0 END) marked_ready,COUNT(DISTINCT difficulty) difficulty_bands
      FROM questions WHERE source_import_batch_id=? GROUP BY subject ORDER BY subject""",(batch_id,)).fetchall()
    c.close()
    return render_template('import_validation.html',batch=batch,checks=rows,total=batch['row_count'],valid=batch['valid_count'],errors=batch['error_count'],warnings=batch['warning_count'],compat_before=safe_json(batch['compatibility_before_json'],{}),compat_after=safe_json(batch['compatibility_after_json'],{}),candidate_summary=candidate_summary)


@app.route('/admin/import/confirm',methods=['POST'])
def admin_import_confirm():
    if not require('admin'): return redirect(url_for('login'))
    batch_id=int(request.form.get('batch_id') or 0); c=db()
    # Backward-compatible migration path for V5.5/V6.0 local sessions. New browser
    # previews never store large row payloads in the session; they persist them in tables.
    if not batch_id:
        legacy_rows=session.pop('pending_import_rows',[])
        if legacy_rows:
            checks=[validate_import_row(row,index+2) for index,row in enumerate(legacy_rows)]
            errors=sum(bool(x['errors']) for x in checks); warnings=sum(bool(x['warnings']) for x in checks)
            code='IMP-LEGACY-'+datetime.now().strftime('%Y%m%d%H%M%S')+'-'+secrets.token_hex(2).upper()
            cur=c.execute("""INSERT INTO content_import_batches(batch_code,source_system,filename,file_type,payload_checksum,row_count,valid_count,error_count,warning_count,status,validation_report_json,compatibility_before_json,imported_by)
              VALUES(?,'LEGACY_SESSION_COMPAT','legacy-session-import','SESSION',?,?,?,?,?,'PREVIEWED',?,?,?)""",
              (code,hashlib.sha256(canonical_json(legacy_rows).encode('utf-8')).hexdigest(),len(checks),len(checks)-errors,errors,warnings,json.dumps({'legacy_compatibility':True}),json.dumps(_active_blueprint_readiness(c)),session['user_id']))
            batch_id=cur.lastrowid
            for check in checks:
                c.execute("""INSERT INTO content_import_batch_rows(batch_id,row_number,sheet_name,question_id,row_json,errors_json,warnings_json)
                  VALUES(?,?,?,?,?,?,?)""",(batch_id,check['row_number'],'LEGACY',check['question_id'],json.dumps(check['row'],default=str),json.dumps(check['errors']),json.dumps(check['warnings'])))
            c.commit()
    batch=c.execute("SELECT * FROM content_import_batches WHERE id=?",(batch_id,)).fetchone()
    if not batch: c.close(); flash('Import batch not found.','error'); return redirect(url_for('admin_import'))
    if batch['status']!='PREVIEWED': c.close(); flash('Only a PREVIEWED batch can be confirmed.','error'); return redirect(url_for('admin_import_batch_detail',batch_id=batch_id))
    if int(batch['error_count'] or 0)>0: c.close(); flash('Whole-batch import blocked until every row-level error is corrected. No partial import was performed.','error'); return redirect(url_for('admin_import_batch_detail',batch_id=batch_id))
    backup=_pilot_backup(c,f"Automatic pre-import backup for {batch['batch_code']}",session['user_id'])
    if not backup['ok']: c.close(); flash('Import blocked because the automatic safety backup failed.','error'); return redirect(url_for('admin_import_batch_detail',batch_id=batch_id))
    rows=c.execute("SELECT * FROM content_import_batch_rows WHERE batch_id=? ORDER BY id",(batch_id,)).fetchall(); inserted=[]
    try:
        c.execute('BEGIN IMMEDIATE')
        for stored in rows:
            row=safe_json(stored['row_json'],{})
            qid=_insert_import_question(c,row,batch_id); inserted.append(qid)
            c.execute("UPDATE content_import_batch_rows SET import_status='IMPORTED',question_db_id=? WHERE id=?",(qid,stored['id']))
        c.execute("""UPDATE content_import_batches SET status='IMPORTED',backup_record_id=?,confirmed_at=?,compatibility_after_json=? WHERE id=?""",
          (backup['id'],datetime.now().isoformat(timespec='seconds'),json.dumps(_active_blueprint_readiness(c)),batch_id))
        pilot_record_event(c,'CONTENT_IMPORT_CONFIRMED','content_import_batch',batch_id,{'inserted':len(inserted),'backup':backup['code']},session['user_id']); c.commit()
    except Exception as exc:
        c.rollback(); c.execute("UPDATE content_import_batches SET status='FAILED_ATOMIC_IMPORT',validation_report_json=? WHERE id=?",(json.dumps({'runtime_error':str(exc)}),batch_id)); c.commit(); c.close()
        flash(f'Atomic import failed; no questions were committed. Safety backup retained. Error: {exc}','error'); return redirect(url_for('admin_import_batch_detail',batch_id=batch_id))
    c.close(); flash(f'{len(inserted)} questions imported atomically as Draft + inactive. Batch rollback remains available until any item is used.','success')
    return redirect(url_for('admin_import_batch_detail',batch_id=batch_id))


@app.route('/admin/import/batch/<int:batch_id>/release-eligible',methods=['POST'])
def admin_import_release_eligible(batch_id):
    if not require('admin'): return redirect(url_for('login'))
    c=db(); batch=c.execute("SELECT * FROM content_import_batches WHERE id=?",(batch_id,)).fetchone()
    if not batch or batch['status']!='IMPORTED' or batch['intake_mode']!='EMERGENCY_DIRECT':
        c.close(); flash('Only an imported Emergency Direct Intake batch can use this release action.','error'); return redirect(url_for('admin_import_batch_detail',batch_id=batch_id))
    attestation=(request.form.get('attestation') or '').strip()
    if attestation!='I CONFIRM THIS IS A FROZEN ACADEMICALLY APPROVED RELEASE':
        c.close(); flash('Release blocked. Enter the exact academic-release attestation.','error'); return redirect(url_for('admin_import_batch_detail',batch_id=batch_id))
    # Eligibility is deterministic and conservative: ScoreMax Ready + acceptable rights + no hold/R2 language in transported review/status fields.
    rows=c.execute("SELECT id,row_json,question_db_id FROM content_import_batch_rows WHERE batch_id=? AND question_db_id IS NOT NULL",(batch_id,)).fetchall()
    eligible=[]; excluded=[]
    for r in rows:
        payload=safe_json(r['row_json'],{})
        ready=str(payload.get('ScoreMax Ready','') or '').strip().lower() in {'1','yes','true'}
        rights=str(payload.get('Rights Status','') or '').strip().lower() in {'scoremax original','licensed','permitted','public domain','approved'}
        governance=' '.join(str(payload.get(k,'') or '') for k in (
          'Status','Review Status','R2 Status','Readiness','Release Status','Review Requirement','Dual Review Status')).casefold()
        r2_required=any(str(payload.get(k,'') or '').strip().casefold() in {'1','yes','true','required','y','dual_review_required','r2_required'}
          for k in ('Reviewer 2 Required','Reviewer2 Required','Dual Review Required'))
        if 'dual_review_required' in str(payload.get('Review Requirement','') or '').strip().casefold():
            r2_required=True
        blocked=r2_required or any(term in governance for term in (
          'hold','unresolved','r2 required','r2_required','pending r2','awaiting r2','not ready','reject','needs changing'))
        (eligible if ready and rights and not blocked else excluded).append(int(r['question_db_id']))
    backup=_pilot_backup(c,f"Automatic backup before emergency release {batch['batch_code']}",session['user_id'])
    if not backup['ok']: c.close(); flash('Release blocked because the safety backup failed.','error'); return redirect(url_for('admin_import_batch_detail',batch_id=batch_id))
    try:
        c.execute('BEGIN IMMEDIATE')
        if eligible:
            # Keep release bounded below common SQLite host-parameter limits. A 3,000-row
            # emergency chapter must not fail simply because one UPDATE contains thousands of IDs.
            family_keys=set()
            for start in range(0,len(eligible),500):
                ids=eligible[start:start+500]; marks=','.join('?' for _ in ids)
                family_keys.update(r['family_key'] for r in c.execute(
                  f"SELECT DISTINCT family_key FROM questions WHERE id IN ({marks}) AND COALESCE(family_key,'')<>''",ids).fetchall())
                c.execute(f"UPDATE questions SET status='Approved',review_status='Approved',active=1,content_environment='PRODUCTION' WHERE id IN ({marks})",ids)
            family_keys=sorted(family_keys)
            for start in range(0,len(family_keys),500):
                keys=family_keys[start:start+500]; family_marks=','.join('?' for _ in keys)
                c.execute(f"UPDATE question_families SET review_status='Approved',active=1,updated_at=CURRENT_TIMESTAMP WHERE family_key IN ({family_marks})",keys)
        now=datetime.now().isoformat(timespec='seconds')
        c.execute("UPDATE content_import_batches SET release_status='RELEASED_ELIGIBLE',release_attested_at=?,release_attested_by=?,release_note=?,released_count=?,released_at=? WHERE id=?",
          (now,session['user_id'],request.form.get('release_note','').strip(),len(eligible),now,batch_id))
        pilot_record_event(c,'EMERGENCY_DIRECT_RELEASE','content_import_batch',batch_id,{'released':len(eligible),'excluded':len(excluded),'backup':backup['code']},session['user_id'])
        c.commit()
    except Exception as exc:
        c.rollback(); c.close(); flash(f'Emergency release failed atomically: {exc}','error'); return redirect(url_for('admin_import_batch_detail',batch_id=batch_id))
    c.close(); flash(f'{len(eligible)} release-eligible questions activated; {len(excluded)} remained inactive.','success')
    return redirect(url_for('admin_import_batch_detail',batch_id=batch_id))


@app.route('/admin/import/batch/<int:batch_id>/rollback',methods=['POST'])
def admin_import_batch_rollback(batch_id):
    if not require('admin'): return redirect(url_for('login'))
    c=db(); batch=c.execute("SELECT * FROM content_import_batches WHERE id=?",(batch_id,)).fetchone()
    if not batch or batch['status']!='IMPORTED': c.close(); flash('Only an imported, unused batch can be rolled back.','error'); return redirect(url_for('admin_import'))
    question_ids=[r['id'] for r in c.execute("SELECT id FROM questions WHERE source_import_batch_id=?",(batch_id,)).fetchall()]
    used=0; governed_changes=0
    if question_ids:
        q=','.join('?'*len(question_ids))
        governed_changes=c.execute(f"""SELECT COUNT(*) n FROM questions WHERE id IN ({q}) AND
          (COALESCE(status,'Draft')<>'Draft' OR COALESCE(review_status,'Draft')<>'Draft' OR COALESCE(active,0)=1)""",question_ids).fetchone()['n']
        used+=c.execute(f"SELECT COUNT(*) n FROM attempt_answers WHERE question_db_id IN ({q})",question_ids).fetchone()['n']
        used+=c.execute(f"SELECT COUNT(*) n FROM exam_paper_questions WHERE question_id IN ({q})",question_ids).fetchone()['n']
        used+=c.execute(f"SELECT COUNT(*) n FROM challenge_questions WHERE question_id IN ({q})",question_ids).fetchone()['n']
    if used or governed_changes:
        c.close(); flash('Rollback blocked because one or more imported questions have entered evidence, a governed paper, or academic review. Deactivate and supersede them instead.','error'); return redirect(url_for('admin_import_batch_detail',batch_id=batch_id))
    backup=_pilot_backup(c,f"Automatic backup before rollback of {batch['batch_code']}",session['user_id'])
    if not backup['ok']: c.close(); flash('Rollback blocked because the safety backup failed.','error'); return redirect(url_for('admin_import_batch_detail',batch_id=batch_id))
    if question_ids:
        q=','.join('?'*len(question_ids))
        for table in ('question_review_events','question_versions','question_curriculum_map'):
            c.execute(f"DELETE FROM {table} WHERE question_id IN ({q})",question_ids)
        c.execute(f"DELETE FROM questions WHERE id IN ({q})",question_ids)
    c.execute("DELETE FROM question_families WHERE source_import_batch_id=? AND NOT EXISTS(SELECT 1 FROM questions q WHERE q.family_key=question_families.family_key)",(batch_id,))
    c.execute("UPDATE content_import_batch_rows SET import_status='ROLLED_BACK',question_db_id=NULL WHERE batch_id=?",(batch_id,))
    c.execute("UPDATE content_import_batches SET status='ROLLED_BACK',rolled_back_at=?,rolled_back_by=?,rollback_note=? WHERE id=?",(datetime.now().isoformat(timespec='seconds'),session['user_id'],request.form.get('note','Whole-batch rollback before use.'),batch_id))
    pilot_record_event(c,'CONTENT_IMPORT_ROLLED_BACK','content_import_batch',batch_id,{'removed_questions':len(question_ids),'backup':backup['code']},session['user_id']); c.commit(); c.close()
    flash(f'Batch rolled back safely; {len(question_ids)} unused candidate questions removed.','success'); return redirect(url_for('admin_import_batch_detail',batch_id=batch_id))


@app.route('/report-issue',methods=['GET','POST'])
def pilot_report_issue():
    if not require(): return redirect(url_for('login'))
    def as_int(value):
        try: return int(value or 0) or None
        except (TypeError,ValueError): return None
    c=db()
    values=request.form if request.method=='POST' else request.args
    question_id=as_int(values.get('question_id')); attempt_id=as_int(values.get('attempt_id')); written_attempt_id=as_int(values.get('written_attempt_id'))
    context={'source':(values.get('source') or '').strip()[:80],'page':(values.get('page') or getattr(request,'referrer','') or '').strip()[:500]}
    question=None; attempt=None; written=None
    if question_id:
        question=c.execute("SELECT id,question_id,subject,chapter,question FROM questions WHERE id=?",(question_id,)).fetchone()
        if not question: question_id=None
    if attempt_id:
        attempt=c.execute("SELECT id,subject,scope,assessment_blueprint_id,assembly_policy_id FROM attempts WHERE id=? AND student_id=?",(attempt_id,session['user_id'])).fetchone()
        if not attempt:
            attempt_id=None
        elif question_id and not c.execute("SELECT 1 FROM attempt_answers WHERE attempt_id=? AND question_db_id=? LIMIT 1",(attempt_id,question_id)).fetchone():
            question_id=None; question=None
    if written_attempt_id:
        written=c.execute("SELECT id,question_id,result_status FROM written_attempts WHERE id=? AND student_id=?",(written_attempt_id,session['user_id'])).fetchone()
        if not written: written_attempt_id=None
    if request.method=='POST':
        category=request.form.get('category','Other').strip(); description=request.form.get('description','').strip(); severity=request.form.get('severity','MEDIUM').upper()
        if severity not in {'LOW','MEDIUM','HIGH','CRITICAL'}: severity='MEDIUM'
        if len(description)<12: c.close(); flash('Please describe the issue clearly.','error'); return redirect(request.referrer or url_for('pilot_report_issue'))
        screenshot_path=''; upload=request.files.get('screenshot')
        if upload and upload.filename:
            ext=Path(upload.filename).suffix.lower()
            if ext not in {'.png','.jpg','.jpeg','.webp'}: c.close(); flash('Screenshots must be PNG, JPG or WebP.','error'); return redirect(request.referrer or url_for('pilot_report_issue'))
            configured=os.environ.get('SCOREMAX_PILOT_UPLOAD_DIR','').strip()
            if SCOREMAX_ENV=='production' and not configured:
                c.close(); flash('Screenshot upload is unavailable until protected storage is configured.','error'); return redirect(request.referrer or url_for('pilot_report_issue'))
            folder=(Path(configured) if configured else BASE/'pilot_uploads')/'feedback'; folder=folder.resolve(); folder.mkdir(parents=True,exist_ok=True)
            name=f"FB-{secrets.token_hex(8)}{ext}"; path=folder/name; upload.save(path); screenshot_path=str(path)
        code='FDB-'+datetime.now().strftime('%Y%m%d%H%M%S')+'-'+secrets.token_hex(2).upper(); target=feedback_route(category)
        blueprint_id=attempt['assessment_blueprint_id'] if attempt else None; policy_id=attempt['assembly_policy_id'] if attempt else None
        c.execute("""INSERT INTO pilot_feedback(feedback_code,reporter_user_id,category,severity,description,question_id,attempt_id,written_attempt_id,blueprint_id,assembly_policy_id,screenshot_path,routing_target,context_json,page_path)
          VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",(code,session['user_id'],category,severity,description,question_id,attempt_id,written_attempt_id,blueprint_id,policy_id,screenshot_path,target,json.dumps(context),context['page']))
        pilot_record_event(c,'PILOT_FEEDBACK_SUBMITTED','pilot_feedback',code,{'category':category,'routing_target':target,'severity':severity,'context':context},session['user_id']); c.commit(); c.close()
        flash(f'Issue {code} submitted. ScoreMax captured the technical context automatically.','success'); return redirect(url_for('student_dashboard') if session.get('role')=='student' else url_for('dashboard'))
    context_label='General platform issue'
    if question: context_label=f"Question in {question['subject'] or 'ScoreMax'}"+(f" · {question['chapter']}" if question['chapter'] else '')
    elif attempt: context_label=f"Test attempt · {attempt['subject'] or attempt['scope'] or 'assessment'}"
    elif written: context_label='Written-answer attempt'
    c.close()
    return render_template('report_issue.html',question=question,attempt=attempt,written_attempt=written,context_label=context_label,
      question_id=question_id,attempt_id=attempt_id,written_attempt_id=written_attempt_id,source=context['source'],page_path=context['page'])


@app.route('/admin/pilot-feedback')
def admin_pilot_feedback():
    if not require('admin'): return redirect(url_for('login'))
    c=db(); rows=c.execute("""SELECT pf.*,u.full_name reporter_name,q.question_id source_question_id FROM pilot_feedback pf
      JOIN users u ON u.id=pf.reporter_user_id LEFT JOIN questions q ON q.id=pf.question_id
      ORDER BY CASE pf.severity WHEN 'CRITICAL' THEN 1 WHEN 'HIGH' THEN 2 WHEN 'MEDIUM' THEN 3 ELSE 4 END,pf.id DESC""").fetchall(); c.close()
    return render_template('admin_pilot_feedback.html',rows=rows)


@app.route('/admin/pilot-feedback/<int:feedback_id>',methods=['POST'])
def admin_pilot_feedback_update(feedback_id):
    if not require('admin'): return redirect(url_for('login'))
    status=request.form.get('status','OPEN').upper()
    if status not in {'OPEN','TRIAGED','SENT_TO_POWER_HOUSE','IN_PROGRESS','RESOLVED','CLOSED'}: status='OPEN'
    c=db(); c.execute("""UPDATE pilot_feedback SET status=?,assigned_to=?,resolution=?,resolved_at=? WHERE id=?""",(status,session['user_id'],request.form.get('resolution','').strip(),datetime.now().isoformat(timespec='seconds') if status in {'RESOLVED','CLOSED'} else '',feedback_id))
    pilot_record_event(c,'PILOT_FEEDBACK_UPDATED','pilot_feedback',feedback_id,{'status':status},session['user_id']); c.commit(); c.close()
    flash('Pilot issue updated.','success'); return redirect(url_for('admin_pilot_feedback'))


@app.route('/admin/pilot-analytics')
def admin_pilot_analytics():
    if not require('admin'): return redirect(url_for('login'))
    c=db(); since=(datetime.now()-timedelta(days=30)).isoformat(timespec='seconds')
    metrics={
      'students':c.execute("SELECT COUNT(*) n FROM users WHERE role='student' AND account_status='active' AND COALESCE(is_demo_account,0)=0").fetchone()['n'],
      'active_students_30d':c.execute("SELECT COUNT(DISTINCT student_id) n FROM attempts WHERE created_at>=?",(since,)).fetchone()['n'],
      'attempts_30d':c.execute("SELECT COUNT(*) n FROM attempts WHERE created_at>=? AND assessment_kind<>'demo_progress'",(since,)).fetchone()['n'],
      'completed_sessions_30d':c.execute("SELECT COUNT(*) n FROM assessment_sessions WHERE status='completed' AND started_at>=?",(since,)).fetchone()['n'],
      'live_questions':c.execute(f"SELECT COUNT(*) n FROM questions q WHERE {live_question_clause('q')}").fetchone()['n'],
      'candidate_questions':c.execute("SELECT COUNT(*) n FROM questions WHERE content_environment='CANDIDATE'").fetchone()['n'],
      'open_feedback':c.execute("SELECT COUNT(*) n FROM pilot_feedback WHERE status NOT IN ('RESOLVED','CLOSED')").fetchone()['n'],
      'failed_jobs':c.execute("SELECT COUNT(*) n FROM written_processing_jobs WHERE state LIKE 'FAILED%'").fetchone()['n'],
      'written_attempts_30d':c.execute("SELECT COUNT(*) n FROM written_attempts WHERE created_at>=?",(since,)).fetchone()['n'],
      'teacher_messages_30d':c.execute("SELECT COUNT(*) n FROM academic_messages WHERE created_at>=?",(since,)).fetchone()['n']}
    by_subject=c.execute("""SELECT subject,COUNT(*) attempts,ROUND(AVG(score),1) average_score FROM attempts
      WHERE created_at>=? AND assessment_kind<>'demo_progress' GROUP BY subject ORDER BY attempts DESC""",(since,)).fetchall()
    imports=c.execute("SELECT status,COUNT(*) n,SUM(row_count) rows FROM content_import_batches GROUP BY status ORDER BY status").fetchall()
    feedback=c.execute("SELECT category,routing_target,status,COUNT(*) n FROM pilot_feedback GROUP BY category,routing_target,status ORDER BY n DESC").fetchall()
    readiness=_active_blueprint_readiness(c); c.close()
    return render_template('admin_pilot_analytics_v62.html',metrics=metrics,by_subject=by_subject,imports=imports,feedback=feedback,readiness=readiness)


@app.route('/sustainability')
def sustainability_page():
    c=db(); control=c.execute("SELECT state FROM sustainability_feature_controls WHERE feature_code='sustainability_public'").fetchone()
    allowed=bool(control and control['state']=='LIVE') or session.get('role')=='admin'
    blocks=c.execute("SELECT * FROM sustainability_content_blocks WHERE status='PUBLISHED' ORDER BY sort_order,id").fetchall() if allowed else []
    policies=c.execute("SELECT * FROM sustainability_policies WHERE status='PUBLISHED' ORDER BY sort_order,id").fetchall() if allowed else []
    commitments=c.execute("SELECT * FROM sustainability_commitments WHERE public_status='PUBLISHED' ORDER BY sort_order,id").fetchall() if allowed else []
    reports=c.execute("SELECT * FROM sustainability_progress_reports WHERE status='PUBLISHED' ORDER BY reporting_period DESC,id DESC").fetchall() if allowed else []
    c.close(); return render_template('sustainability.html',available=allowed,blocks=blocks,policies=policies,commitments=commitments,reports=reports)


@app.route('/admin/sustainability',methods=['GET','POST'])
def admin_sustainability():
    if not require('admin'): return redirect(url_for('login'))
    c=db()
    if request.method=='POST':
        action=(request.form.get('action') or '').strip()
        if action=='feature':
            state=(request.form.get('state') or 'HIDDEN').upper()
            if state not in {'HIDDEN','PILOT','LIVE'}: c.close(); abort(400)
            c.execute("UPDATE sustainability_feature_controls SET state=?,updated_by=?,updated_at=CURRENT_TIMESTAMP WHERE feature_code='sustainability_public'",(state,session['user_id']))
        elif action=='block':
            code=(request.form.get('content_code') or '').strip(); heading=(request.form.get('heading') or '').strip(); body=(request.form.get('body_text') or '').strip()
            stage=(request.form.get('claim_stage') or 'CURRENT_PRACTICE').upper(); status=(request.form.get('status') or 'DRAFT').upper()
            if not code or not heading or not body or stage not in {'CURRENT_PRACTICE','IN_PROGRESS','FUTURE_COMMITMENT'} or status not in {'DRAFT','REVIEW','PUBLISHED'}:
                c.close(); flash('Complete the content block using valid governance labels.','error'); return redirect(url_for('admin_sustainability'))
            c.execute("""UPDATE sustainability_content_blocks SET heading=?,body_text=?,claim_stage=?,status=?,version=?,owner=?,evidence_summary=?,
              published_at=CASE WHEN ?='PUBLISHED' THEN COALESCE(NULLIF(published_at,''),CURRENT_TIMESTAMP) ELSE published_at END,
              updated_by=?,updated_at=CURRENT_TIMESTAMP WHERE content_code=?""",
              (heading,body,stage,status,(request.form.get('version') or '1.0').strip(),(request.form.get('owner') or 'ScoreMax').strip(),
               (request.form.get('evidence_summary') or '').strip(),status,session['user_id'],code))
        elif action=='policy':
            code=(request.form.get('policy_code') or '').strip().upper(); title=(request.form.get('title') or '').strip()
            if not code or not title: c.close(); flash('Policy code and title are required.','error'); return redirect(url_for('admin_sustainability'))
            status=(request.form.get('status') or 'DRAFT').upper()
            c.execute("""INSERT INTO sustainability_policies(policy_code,title,summary,body_text,version,effective_date,last_review_date,next_review_date,owner,status,sort_order,published_at,updated_by)
              VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?) ON CONFLICT(policy_code) DO UPDATE SET title=excluded.title,summary=excluded.summary,body_text=excluded.body_text,
              version=excluded.version,effective_date=excluded.effective_date,last_review_date=excluded.last_review_date,next_review_date=excluded.next_review_date,
              owner=excluded.owner,status=excluded.status,sort_order=excluded.sort_order,published_at=excluded.published_at,updated_by=excluded.updated_by,updated_at=CURRENT_TIMESTAMP""",
              (code,title,(request.form.get('summary') or '').strip(),(request.form.get('body_text') or '').strip(),(request.form.get('version') or '1.0').strip(),
               request.form.get('effective_date',''),request.form.get('last_review_date',''),request.form.get('next_review_date',''),(request.form.get('owner') or 'ScoreMax').strip(),status,
               int(request.form.get('sort_order') or 100),datetime.now().isoformat(timespec='seconds') if status=='PUBLISHED' else '',session['user_id']))
        elif action=='commitment':
            code=(request.form.get('commitment_code') or '').strip().upper(); title=(request.form.get('title') or '').strip(); description=(request.form.get('description') or '').strip()
            stage=(request.form.get('claim_stage') or 'FUTURE_COMMITMENT').upper()
            if not code or not title or not description or stage not in {'CURRENT_PRACTICE','IN_PROGRESS','FUTURE_COMMITMENT'}:
                c.close(); flash('Commitment code, title, description and a valid claim stage are required.','error'); return redirect(url_for('admin_sustainability'))
            c.execute("""INSERT INTO sustainability_commitments(commitment_code,title,category,claim_stage,description,baseline_text,target_text,target_date,status,owner,evidence_summary,public_status,sort_order,updated_by)
              VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?) ON CONFLICT(commitment_code) DO UPDATE SET title=excluded.title,category=excluded.category,claim_stage=excluded.claim_stage,
              description=excluded.description,baseline_text=excluded.baseline_text,target_text=excluded.target_text,target_date=excluded.target_date,status=excluded.status,
              owner=excluded.owner,evidence_summary=excluded.evidence_summary,public_status=excluded.public_status,sort_order=excluded.sort_order,updated_by=excluded.updated_by,updated_at=CURRENT_TIMESTAMP""",
              (code,title,(request.form.get('category') or 'Other').strip(),stage,description,(request.form.get('baseline_text') or '').strip(),
               (request.form.get('target_text') or '').strip(),request.form.get('target_date',''),(request.form.get('commitment_status') or 'PLANNED').upper(),
               (request.form.get('owner') or 'ScoreMax').strip(),(request.form.get('evidence_summary') or '').strip(),(request.form.get('public_status') or 'DRAFT').upper(),
               int(request.form.get('sort_order') or 100),session['user_id']))
        elif action=='progress':
            title=(request.form.get('title') or '').strip(); body=(request.form.get('body_text') or '').strip(); period=(request.form.get('reporting_period') or '').strip()
            if not title or not body or not period: c.close(); flash('Reporting period, title and report body are required.','error'); return redirect(url_for('admin_sustainability'))
            status=(request.form.get('status') or 'DRAFT').upper()
            c.execute("""INSERT INTO sustainability_progress_reports(reporting_period,title,summary,body_text,status,published_at,created_by)
              VALUES(?,?,?,?,?,?,?)""",(period,title,(request.form.get('summary') or '').strip(),body,status,datetime.now().isoformat(timespec='seconds') if status=='PUBLISHED' else '',session['user_id']))
        elif action=='growth_import':
            raw=(request.form.get('draft_json') or '').strip()
            try: payload=json.loads(raw); canonical=json.dumps(payload,sort_keys=True,separators=(',',':'))
            except Exception as exc: c.close(); flash(f'Draft JSON could not be read: {exc}','error'); return redirect(url_for('admin_sustainability'))
            checksum=hashlib.sha256(canonical.encode()).hexdigest()
            c.execute("""INSERT INTO sustainability_draft_intake(external_draft_id,source_system,payload_json,payload_checksum,imported_by)
              VALUES(?,?,?,?,?)""",(str(payload.get('draft_id') or ''),str(payload.get('source_system') or 'Growth Engine'),canonical,checksum,session['user_id']))
            flash('Sustainability draft imported for human review; nothing was published.','success')
        else: c.close(); abort(400)
        pilot_record_event(c,'SUSTAINABILITY_ADMIN_UPDATED','sustainability',action,{},session['user_id']); c.commit()
    control=c.execute("SELECT * FROM sustainability_feature_controls WHERE feature_code='sustainability_public'").fetchone()
    blocks=c.execute("SELECT * FROM sustainability_content_blocks ORDER BY sort_order,id").fetchall(); policies=c.execute("SELECT * FROM sustainability_policies ORDER BY sort_order,id").fetchall()
    commitments=c.execute("SELECT * FROM sustainability_commitments ORDER BY sort_order,id").fetchall(); reports=c.execute("SELECT * FROM sustainability_progress_reports ORDER BY id DESC").fetchall()
    drafts=c.execute("SELECT * FROM sustainability_draft_intake ORDER BY id DESC LIMIT 20").fetchall(); c.close()
    return render_template('admin_sustainability.html',control=control,blocks=blocks,policies=policies,commitments=commitments,reports=reports,drafts=drafts)


@app.route('/admin/daily-spark',methods=['GET','POST'])
def admin_daily_spark():
    if not require('admin'): return redirect(url_for('login'))
    c=db()
    if request.method=='POST':
        action=(request.form.get('action') or '').strip()
        if action=='feature':
            code=(request.form.get('feature_code') or '').strip(); state=(request.form.get('state') or 'HIDDEN').upper()
            if code not in {'academic_spark','word_of_the_day'} or state not in {'HIDDEN','PILOT','LIVE'}: c.close(); abort(400)
            c.execute("UPDATE daily_spark_feature_controls SET state=?,updated_by=?,updated_at=CURRENT_TIMESTAMP WHERE feature_code=?",(state,session['user_id'],code))
        elif action=='word':
            word=(request.form.get('word') or '').strip(); definition=(request.form.get('definition') or '').strip(); example=(request.form.get('example_sentence') or '').strip()
            if not word or not definition or not example: c.close(); flash('Word, definition and example sentence are required.','error'); return redirect(url_for('admin_daily_spark'))
            c.execute("""INSERT INTO daily_spark_words(word,pronunciation,definition,example_sentence,synonym,antonym,exam_application,difficulty_rank,min_age,max_age,syllabus_tags_json,source_name,source_ref,active,content_version)
              VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) ON CONFLICT(word) DO UPDATE SET pronunciation=excluded.pronunciation,definition=excluded.definition,
              example_sentence=excluded.example_sentence,synonym=excluded.synonym,antonym=excluded.antonym,exam_application=excluded.exam_application,
              difficulty_rank=excluded.difficulty_rank,min_age=excluded.min_age,max_age=excluded.max_age,syllabus_tags_json=excluded.syllabus_tags_json,
              source_name=excluded.source_name,source_ref=excluded.source_ref,active=excluded.active,content_version=excluded.content_version,updated_at=CURRENT_TIMESTAMP""",
              (word,(request.form.get('pronunciation') or '').strip(),definition,example,(request.form.get('synonym') or '').strip(),(request.form.get('antonym') or '').strip(),
               (request.form.get('exam_application') or '').strip(),max(1,min(5,int(request.form.get('difficulty_rank') or 2))),max(5,min(25,int(request.form.get('min_age') or 10))),
               max(5,min(30,int(request.form.get('max_age') or 20))),json.dumps(parse_list(request.form.get('syllabus_tags','')),sort_keys=True),
               (request.form.get('source_name') or 'ScoreMax controlled vocabulary library').strip(),(request.form.get('source_ref') or '').strip(),
               1 if request.form.get('active')=='1' else 0,(request.form.get('content_version') or '1.0').strip()))
        elif action=='toggle_word':
            wid=int(request.form.get('word_id') or 0); active=1 if request.form.get('active')=='1' else 0
            c.execute("UPDATE daily_spark_words SET active=?,updated_at=CURRENT_TIMESTAMP WHERE id=?",(active,wid))
        else: c.close(); abort(400)
        pilot_record_event(c,'DAILY_SPARK_ADMIN_UPDATED','daily_spark',action,{},session['user_id']); c.commit()
    controls=c.execute("SELECT * FROM daily_spark_feature_controls ORDER BY feature_code").fetchall(); words=c.execute("SELECT * FROM daily_spark_words ORDER BY active DESC,difficulty_rank,word").fetchall()
    metrics=daily_spark_metrics(c); recent=c.execute("""SELECT a.spark_date,a.stream,u.full_name,e.event_type,e.created_at FROM daily_spark_events e
      JOIN daily_spark_assignments a ON a.id=e.assignment_id JOIN users u ON u.id=e.student_id ORDER BY e.id DESC LIMIT 40""").fetchall(); c.close()
    return render_template('admin_daily_spark.html',controls=controls,words=words,metrics=metrics,recent=recent)


@app.route('/connect')
def connect_page():
    c=db(); links=active_social_links(c); c.close()
    return render_template('connect.html',social_links=links)


@app.route('/admin/social-links',methods=['GET','POST'])
def admin_social_links():
    if not require('admin'): return redirect(url_for('login'))
    c=db()
    if request.method=='POST':
        for row in c.execute("SELECT * FROM platform_social_links ORDER BY sort_order,id").fetchall():
            code=row['platform_code']; url=request.form.get(f'url_{code}','').strip(); active=1 if request.form.get(f'active_{code}')=='1' else 0
            if url and not re.match(r'^https://[^\s]+$',url,re.I):
                c.close(); flash(f"{row['display_name']} must use a complete https:// URL.",'error'); return redirect(url_for('admin_social_links'))
            if active and not url:
                c.close(); flash(f"Add the verified {row['display_name']} URL before enabling it.",'error'); return redirect(url_for('admin_social_links'))
            c.execute("UPDATE platform_social_links SET url=?,active=?,updated_by=?,updated_at=CURRENT_TIMESTAMP WHERE id=?",
              (url,active,session['user_id'],row['id']))
        pilot_record_event(c,'SOCIAL_LINKS_UPDATED','platform_social_links','all',{},session['user_id']); c.commit(); flash('Official ScoreMax channel links updated.','success')
    links=c.execute("SELECT * FROM platform_social_links ORDER BY sort_order,id").fetchall(); c.close()
    return render_template('admin_social_links.html',links=links)


@app.route('/knowledge')
def knowledge_home():
    c=db(); control=c.execute("SELECT state FROM knowledge_feature_controls WHERE feature_code='knowledge_hub'").fetchone(); allowed=bool(control and control['state']=='LIVE') or session.get('role')=='admin'
    articles=c.execute("SELECT * FROM knowledge_articles WHERE status='PUBLISHED' ORDER BY published_at DESC,id DESC").fetchall() if allowed else []
    c.close(); return render_template('knowledge_home.html',articles=articles,available=allowed)


@app.route('/knowledge/<slug>')
def knowledge_article(slug):
    c=db(); control=c.execute("SELECT state FROM knowledge_feature_controls WHERE feature_code='knowledge_hub'").fetchone(); allowed=bool(control and control['state']=='LIVE') or session.get('role')=='admin'
    row=c.execute("SELECT * FROM knowledge_articles WHERE slug=? AND (status='PUBLISHED' OR ?=1)",(slug,1 if session.get('role')=='admin' else 0)).fetchone() if allowed else None
    sources=c.execute("SELECT * FROM knowledge_sources WHERE article_id=? ORDER BY id",(row['id'],)).fetchall() if row else []
    c.close()
    if not row: abort(404)
    return render_template('knowledge_article.html',article=row,sources=sources)


@app.route('/admin/knowledge',methods=['GET','POST'])
def admin_knowledge():
    if not require('admin'): return redirect(url_for('login'))
    c=db()
    if request.method=='POST':
        action=request.form.get('action','article')
        if action=='feature':
            state=request.form.get('state','HIDDEN').upper()
            if state not in {'HIDDEN','PILOT','LIVE'}: state='HIDDEN'
            c.execute("UPDATE knowledge_feature_controls SET state=?,updated_by=?,updated_at=CURRENT_TIMESTAMP WHERE feature_code='knowledge_hub'",(state,session['user_id']))
            c.execute("UPDATE pilot_feature_controls SET state=?,updated_by=?,updated_at=CURRENT_TIMESTAMP WHERE feature_code='knowledge_hub'",(state,session['user_id']))
        elif action=='growth_import':
            upload=request.files.get('growth_draft_file')
            try: payload=json.loads(upload.read().decode('utf-8-sig')) if upload and upload.filename else json.loads(request.form.get('growth_draft_json','{}'))
            except Exception as exc: c.close(); flash(f'Growth draft could not be read: {exc}','error'); return redirect(url_for('admin_knowledge'))
            digest=pilot_payload_checksum(payload)
            cur=c.execute("INSERT INTO growth_content_intake(external_draft_id,payload_json,payload_checksum,imported_by) VALUES(?,?,?,?)",(str(payload.get('draft_id','')),canonical_json(payload),digest,session['user_id']))
            title=str(payload.get('title','Untitled Growth draft')).strip(); slug=safe_filename(str(payload.get('slug','') or title).lower()).replace('_','-')
            if c.execute("SELECT 1 FROM knowledge_articles WHERE slug=?",(slug,)).fetchone(): slug=f"{slug}-{secrets.token_hex(2)}"
            article=c.execute("""INSERT INTO knowledge_articles(slug,title,summary,body_text,author_name,source_origin,framework,framework_version,subject,chapter,status,created_by)
              VALUES(?,?,?,?,?,'GROWTH_ENGINE',?,?,?,?, 'DRAFT',?)""",(slug,title,str(payload.get('summary','')),str(payload.get('body_text','')),str(payload.get('author_name','ScoreMax Growth Engine')),
              str(payload.get('framework','')),str(payload.get('framework_version','')),str(payload.get('subject','')),str(payload.get('chapter','')),session['user_id'])).lastrowid
            c.execute("UPDATE growth_content_intake SET status='CONVERTED_TO_DRAFT',converted_article_id=? WHERE id=?",(article,cur.lastrowid))
        else:
            title=request.form.get('title','').strip(); body=request.form.get('body_text','').strip()
            if not title or not body: c.close(); flash('Title and article body are required.','error'); return redirect(url_for('admin_knowledge'))
            slug=safe_filename(request.form.get('slug','').strip().lower() or title.lower()).replace('_','-')
            if c.execute("SELECT 1 FROM knowledge_articles WHERE slug=?",(slug,)).fetchone(): slug=f"{slug}-{secrets.token_hex(2)}"
            status=request.form.get('status','DRAFT').upper()
            if status not in {'DRAFT','REVIEW','PUBLISHED','ARCHIVED'}: status='DRAFT'
            published=datetime.now().isoformat(timespec='seconds') if status=='PUBLISHED' else ''
            cur=c.execute("""INSERT INTO knowledge_articles(slug,title,summary,body_text,author_name,source_origin,framework,framework_version,subject,chapter,status,seo_title,seo_description,reviewed_at,reviewed_by,published_at,created_by)
              VALUES(?,?,?,?,?,'MANUAL',?,?,?,?,?,?,?,?,?,?,?)""",(slug,title,request.form.get('summary',''),body,request.form.get('author_name',session.get('full_name','ScoreMax')),
              request.form.get('framework',''),request.form.get('framework_version',''),request.form.get('subject',''),request.form.get('chapter',''),status,
              request.form.get('seo_title',''),request.form.get('seo_description',''),datetime.now().isoformat(timespec='seconds') if status in {'REVIEW','PUBLISHED'} else '',session['user_id'] if status in {'REVIEW','PUBLISHED'} else None,published,session['user_id']))
            if request.form.get('source_title','').strip():
                c.execute("""INSERT INTO knowledge_sources(article_id,source_title,source_organisation,source_url,source_document_ref,publication_date,rights_status,notes)
                  VALUES(?,?,?,?,?,?,?,?)""",(cur.lastrowid,request.form.get('source_title',''),request.form.get('source_organisation',''),request.form.get('source_url',''),request.form.get('source_document_ref',''),request.form.get('source_publication_date',''),request.form.get('rights_status','LINK_ONLY'),request.form.get('source_notes','')))
        c.commit(); flash('Knowledge Hub control updated.','success')
    control=c.execute("SELECT * FROM knowledge_feature_controls WHERE feature_code='knowledge_hub'").fetchone(); articles=c.execute("SELECT * FROM knowledge_articles ORDER BY id DESC").fetchall(); intake=c.execute("SELECT * FROM growth_content_intake ORDER BY id DESC LIMIT 20").fetchall(); c.close()
    return render_template('admin_knowledge.html',control=control,articles=articles,intake=intake)

# ---------------------------------------------------------------------------
# V6.0 Written Response Intelligence routes
# ---------------------------------------------------------------------------

@app.route('/written-practice')
def written_practice_home():
    if not require('student'): return redirect(url_for('login'))
    c=db()
    available=written_feature_available(c,session['user_id'],'written_response_engine')
    questions=[]
    if available:
        questions=c.execute("""SELECT wq.*,wap.subject_id,wap.chapter_id,wap.assessment_package_id,wap.assessment_package_version
          FROM written_questions wq JOIN written_assessment_packages wap ON wap.id=wq.package_id
          WHERE wq.active=1 AND wap.local_status='ACTIVE' ORDER BY wap.subject_id,wap.chapter_id,wq.id""").fetchall()
    recent=c.execute("""SELECT wa.*,wq.question_text,wap.subject_id,wap.chapter_id FROM written_attempts wa
      JOIN written_questions wq ON wq.id=wa.written_question_id JOIN written_assessment_packages wap ON wap.id=wa.package_id
      WHERE wa.student_id=? ORDER BY wa.created_at DESC LIMIT 12""",(session['user_id'],)).fetchall()
    c.close(); return render_template('written_practice.html',available=available,questions=questions,recent=recent)


@app.route('/written-practice/question/<int:question_id>')
def written_question_page(question_id):
    if not require('student'): return redirect(url_for('login'))
    c=db()
    if not written_feature_available(c,session['user_id'],'written_response_engine'):
        c.close(); flash('Written Response Intelligence is currently in controlled pilot testing.','info'); return redirect(url_for('written_practice_home'))
    q=written_package_question(c,question_id)
    if not q: c.close(); abort(404)
    payload=written_question_payload(q)
    handwriting=written_feature_available(c,session['user_id'],'written_handwriting')
    build_answer=written_feature_available(c,session['user_id'],'written_build_answer') and bool(payload.get('scaffold_activities'))
    parent_attempt_id=int(request.args.get('parent_attempt_id') or 0) if str(request.args.get('parent_attempt_id') or '').isdigit() else 0
    unseen=request.args.get('unseen')=='1'
    c.close(); return render_template('written_question.html',q=q,payload=payload,handwriting=handwriting,build_answer=build_answer,parent_attempt_id=parent_attempt_id,unseen=unseen)


@app.route('/written-practice/question/<int:question_id>/submit',methods=['POST'])
def written_submit_typed(question_id):
    if not require('student'): return redirect(url_for('login'))
    if not rate_limit(f"written-submit:{session.get('user_id')}",limit=20,window_seconds=600):
        flash('Too many written submissions in a short period. Review your feedback before trying again.','error'); return redirect(url_for('written_question_page',question_id=question_id))
    answer=request.form.get('answer_text','').strip(); mode=request.form.get('attempt_mode','practice')
    if mode not in ('practice','mock'): mode='practice'
    parent_attempt_id=int(request.form.get('parent_attempt_id') or 0) if str(request.form.get('parent_attempt_id') or '').isdigit() else None
    unseen=request.form.get('unseen')=='1'
    if not answer: flash('Write an answer before submitting.','error'); return redirect(url_for('written_question_page',question_id=question_id))
    c=db()
    if not written_feature_available(c,session['user_id'],'written_response_engine'):
        c.close(); abort(403)
    q=written_package_question(c,question_id)
    if not q: c.close(); abort(404)
    payload=written_question_payload(q)
    cur=c.execute("""INSERT INTO written_attempts(student_id,written_question_id,package_id,parent_attempt_id,attempt_mode,entry_method,evidence_type,
      support_level,novelty_status,status,maximum_mark,package_version,rubric_version,mastery_policy_version,rigor_policy_version,
      original_submitted_at) VALUES(?,?,?,?,?,'typed','independent_production','independent',?,'SUBMITTED',?,?,?,?,?,?)""",
      (session['user_id'],question_id,q['package_id'],parent_attempt_id,mode,'unseen_reconfirmation' if unseen else 'seen_family',float(q['maximum_marks']),q['assessment_package_version'],
       str(payload.get('rubric_version','1')),str(payload.get('mastery_policy_version','')),str(payload.get('rigor_policy_version','')),
       datetime.now().isoformat(timespec='seconds')))
    attempt_id=cur.lastrowid
    vcur=c.execute("""INSERT INTO written_answer_versions(attempt_id,version_no,version_type,answer_text,word_count,is_frozen)
      VALUES(?,1,'ORIGINAL_TYPED',?,?,?)""",(attempt_id,answer,len(answer.split()),1 if mode=='mock' else 0))
    written_evaluate_attempt(c,attempt_id,vcur.lastrowid,creates_formal_evidence=True)
    c.commit(); c.close(); return redirect(url_for('written_attempt_result',attempt_id=attempt_id))


@app.route('/written-practice/attempt/<int:attempt_id>')
def written_attempt_result(attempt_id):
    if not require('student'): return redirect(url_for('login'))
    c=db(); data=written_attempt_view(c,attempt_id,session['user_id'])
    if not data: c.close(); abort(404)
    exemplar_live=written_feature_available(c,session['user_id'],'written_exemplar_library')
    c.close(); return render_template('written_result.html',data=data,exemplar_live=exemplar_live)


@app.route('/written-practice/attempt/<int:attempt_id>/improve',methods=['POST'])
def written_improve_attempt(attempt_id):
    if not require('student'): return redirect(url_for('login'))
    answer=request.form.get('answer_text','').strip()
    if not answer: flash('Your improved answer cannot be blank.','error'); return redirect(url_for('written_attempt_result',attempt_id=attempt_id))
    c=db(); attempt=c.execute("SELECT * FROM written_attempts WHERE id=? AND student_id=?",(attempt_id,session['user_id'])).fetchone()
    if not attempt: c.close(); abort(404)
    if attempt['attempt_mode']=='mock': c.close(); flash('Mock evidence is frozen and cannot be substantively amended.','error'); return redirect(url_for('written_attempt_result',attempt_id=attempt_id))
    version_no=int(c.execute("SELECT COALESCE(MAX(version_no),0)+1 n FROM written_answer_versions WHERE attempt_id=?",(attempt_id,)).fetchone()['n'])
    cur=c.execute("""INSERT INTO written_answer_versions(attempt_id,version_no,version_type,answer_text,word_count,is_frozen)
      VALUES(?,?,'FEEDBACK_LED_IMPROVEMENT',?,?,0)""",(attempt_id,version_no,answer,len(answer.split())))
    written_evaluate_attempt(c,attempt_id,cur.lastrowid,creates_formal_evidence=False)
    c.commit(); c.close(); flash('Improved version marked. Your original evidence remains unchanged.','success')
    return redirect(url_for('written_attempt_result',attempt_id=attempt_id))


@app.route('/written-practice/attempt/<int:attempt_id>/unseen',methods=['POST'])
def written_start_unseen(attempt_id):
    if not require('student'): return redirect(url_for('login'))
    c=db(); parent=c.execute("SELECT * FROM written_attempts WHERE id=? AND student_id=?",(attempt_id,session['user_id'])).fetchone()
    if not parent: c.close(); abort(404)
    q=written_package_question(c,parent['written_question_id']); payload=written_question_payload(q)
    ids=[str(x) for x in payload.get('unseen_reconfirmation_ids',[]) if str(x)]
    unseen=None
    if ids:
        placeholders=','.join('?' for _ in ids)
        unseen=c.execute(f"""SELECT wq.* FROM written_questions wq JOIN written_assessment_packages wap ON wap.id=wq.package_id
          WHERE wq.question_source_id IN ({placeholders}) AND wq.id<>? AND wap.local_status='ACTIVE' LIMIT 1""",ids+[parent['written_question_id']]).fetchone()
    if not unseen:
        unseen=c.execute("""SELECT * FROM written_questions WHERE package_id=? AND question_family_id=? AND id<>? AND active=1 LIMIT 1""",
                         (parent['package_id'],q['question_family_id'],parent['written_question_id'])).fetchone()
    if not unseen:
        c.close(); flash('No approved unseen reconfirmation variant is available from Power House yet.','info'); return redirect(url_for('written_attempt_result',attempt_id=attempt_id))
    c.close(); return redirect(url_for('written_question_page',question_id=unseen['id'],parent_attempt_id=attempt_id,unseen='1'))


@app.route('/written-practice/question/<int:question_id>/build-answer')
def written_build_answer_page(question_id):
    if not require('student'): return redirect(url_for('login'))
    c=db(); q=written_package_question(c,question_id)
    if not q or not written_feature_available(c,session['user_id'],'written_build_answer'):
        c.close(); abort(404)
    payload=written_question_payload(q); scaffolds=payload.get('scaffold_activities') or []
    c.close(); return render_template('written_build_answer.html',q=q,payload=payload,scaffolds=scaffolds)


@app.route('/written-practice/question/<int:question_id>/handwriting',methods=['GET','POST'])
def written_handwriting_upload(question_id):
    if not require('student'): return redirect(url_for('login'))
    c=db(); q=written_package_question(c,question_id)
    if not q or not written_feature_available(c,session['user_id'],'written_handwriting'):
        c.close(); flash('Handwriting processing is in controlled pilot testing.','info'); return redirect(url_for('written_question_page',question_id=question_id))
    if request.method=='GET':
        c.close(); return render_template('written_handwriting.html',q=q)
    if not rate_limit(f"written-upload:{session.get('user_id')}",limit=10,window_seconds=3600):
        c.close(); flash('Handwriting upload limit reached for this hour.','error'); return redirect(url_for('written_question_page',question_id=question_id))
    if SCOREMAX_ENV=='production' and os.environ.get('SCOREMAX_WRITTEN_STORAGE_PROVIDER','').strip()!='secure_object_store':
        c.close(); flash('Secure private object storage must be configured before production handwriting uploads are enabled.','error'); return redirect(url_for('written_question_page',question_id=question_id))
    files=request.files.getlist('pages') if hasattr(request.files,'getlist') else []
    files=[x for x in files if getattr(x,'filename','')]
    if not files or len(files)>6:
        c.close(); flash('Upload between one and six notebook pages.','error'); return redirect(request.url)
    payload=written_question_payload(q)
    cur=c.execute("""INSERT INTO written_attempts(student_id,written_question_id,package_id,attempt_mode,entry_method,evidence_type,
      support_level,novelty_status,status,maximum_mark,package_version,rubric_version,original_submitted_at)
      VALUES(?,?,?,'practice','handwritten','independent_production','independent','seen_family','IMAGE_CHECK_RUNNING',?,?,?,?)""",
      (session['user_id'],question_id,q['package_id'],float(q['maximum_marks']),q['assessment_package_version'],str(payload.get('rubric_version','1')),
       datetime.now().isoformat(timespec='seconds')))
    attempt_id=cur.lastrowid
    folder=BASE/'private_uploads'/'written'/f"attempt_{attempt_id}_{secrets.token_hex(6)}"; folder.mkdir(parents=True,exist_ok=True)
    quality_block=False
    for page_no,upload in enumerate(files,1):
        try:
            img=Image.open(upload.stream); img.load(); img=img.convert('RGB')
            width,height=img.size
            gray=img.convert('L'); stat=ImageStat.Stat(gray); brightness=float(stat.mean[0]); contrast=float(stat.stddev[0])
            edge=ImageStat.Stat(gray.filter(ImageFilter.FIND_EDGES)).mean[0]
            issues=[]
            if width<900 or height<1100: issues.append('low_resolution')
            if contrast<22: issues.append('low_contrast_or_shadow')
            if brightness<45: issues.append('too_dark')
            if brightness>238: issues.append('glare_or_overexposure')
            if edge<6: issues.append('possible_blur')
            score=max(0,100-18*len(issues)); status='RETAKE_REQUIRED' if score<65 else 'PASSED'
            quality_block=quality_block or status!='PASSED'
            target=folder/f"page_{page_no}.jpg"; img.save(target,'JPEG',quality=92,optimize=True)
            digest=hashlib.sha256(target.read_bytes()).hexdigest()
            c.execute("""INSERT INTO written_upload_pages(attempt_id,page_no,storage_path,original_filename,mime_type,width,height,
              quality_score,quality_status,quality_json,original_file_hash) VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
              (attempt_id,page_no,str(target),upload.filename,'image/jpeg',width,height,score,status,json.dumps({'brightness':round(brightness,2),'contrast':round(contrast,2),'edge_score':round(edge,2),'issues':issues}),digest))
        except Exception:
            quality_block=True
    state='RETAKE_REQUIRED' if quality_block else 'OCR_PROVIDER_REQUIRED'
    c.execute("UPDATE written_attempts SET status=?,result_state=? WHERE id=?",(state,state,attempt_id))
    c.execute("""INSERT INTO written_processing_jobs(attempt_id,job_type,state,provider,provider_version,idempotency_key,input_json)
      VALUES(?,'OCR',?,'NOT_CONFIGURED','',?,?)""",(attempt_id,state,f'ocr-{attempt_id}-1',json.dumps({'page_count':len(files)})))
    c.commit(); c.close(); return redirect(url_for('written_attempt_result',attempt_id=attempt_id))


@app.route('/written-practice/page/<int:page_id>')
def written_private_page(page_id):
    if not session.get('user_id'): return redirect(url_for('login'))
    c=db(); page=c.execute("""SELECT wp.*,wa.student_id FROM written_upload_pages wp JOIN written_attempts wa ON wa.id=wp.attempt_id WHERE wp.id=?""",(page_id,)).fetchone()
    if not page or (session.get('role')!='admin' and page['student_id']!=session.get('user_id')):
        c.close(); abort(404)
    path=page['storage_path']; c.close(); return send_file(path,mimetype='image/jpeg')


@app.route('/admin/written-attempts/<int:attempt_id>/simulate-ocr',methods=['POST'])
def admin_written_simulate_ocr(attempt_id):
    if not require('admin'): return redirect(url_for('login'))
    if SCOREMAX_ENV=='production': abort(404)
    transcript=request.form.get('transcript','').strip()
    c=db(); attempt=c.execute("SELECT * FROM written_attempts WHERE id=? AND entry_method='handwritten'",(attempt_id,)).fetchone()
    if not attempt or not transcript: c.close(); abort(400)
    c.execute("""INSERT INTO written_answer_versions(attempt_id,version_no,version_type,original_ocr_text,confirmed_transcript,word_count,is_frozen)
      VALUES(?,1,'OCR_TRANSCRIPT',?,?,?,0)""",(attempt_id,transcript,transcript,len(transcript.split())))
    c.execute("UPDATE written_attempts SET status='OCR_CONFIRMATION_REQUIRED',result_state='OCR_CONFIRMATION_REQUIRED' WHERE id=?",(attempt_id,))
    c.execute("UPDATE written_processing_jobs SET state='OCR_COMPLETE_SIMULATED',provider='LOCAL_ADMIN_SIMULATION',output_json=?,updated_at=CURRENT_TIMESTAMP WHERE attempt_id=? AND job_type='OCR'",
              (json.dumps({'transcript':transcript}),attempt_id))
    c.commit(); c.close(); flash('Pilot OCR transcript injected. Student confirmation is still required.','success')
    return redirect(url_for('admin_written_response'))


@app.route('/written-practice/attempt/<int:attempt_id>/confirm-transcript',methods=['POST'])
def written_confirm_transcript(attempt_id):
    if not require('student'): return redirect(url_for('login'))
    confirmed=request.form.get('confirmed_transcript','').strip()
    c=db(); attempt=c.execute("SELECT * FROM written_attempts WHERE id=? AND student_id=? AND entry_method='handwritten'",(attempt_id,session['user_id'])).fetchone()
    version=c.execute("SELECT * FROM written_answer_versions WHERE attempt_id=? ORDER BY version_no LIMIT 1",(attempt_id,)).fetchone() if attempt else None
    if not attempt or not version or not confirmed: c.close(); abort(400)
    original=version['original_ocr_text'] or ''
    corrections={'original_ocr':original,'confirmed':confirmed,'changed':original!=confirmed,'confirmed_at':datetime.now().isoformat(timespec='seconds')}
    c.execute("UPDATE written_answer_versions SET confirmed_transcript=?,correction_log_json=?,is_frozen=1 WHERE id=?",
              (confirmed,json.dumps([corrections]),version['id']))
    c.execute("UPDATE written_attempts SET status='SUBMITTED',result_state='' WHERE id=?",(attempt_id,))
    written_evaluate_attempt(c,attempt_id,version['id'],creates_formal_evidence=True)
    c.commit(); c.close(); return redirect(url_for('written_attempt_result',attempt_id=attempt_id))


@app.route('/written-practice/exemplars')
def written_exemplar_library():
    if not require('student'): return redirect(url_for('login'))
    c=db()
    if not written_feature_available(c,session['user_id'],'written_exemplar_library'):
        c.close(); flash('The approved exemplar library is not yet released.','info'); return redirect(url_for('written_practice_home'))
    rows=c.execute("""SELECT we.*,wq.question_text,wav.answer_text,wav.confirmed_transcript,wap.subject_id,wap.chapter_id
      FROM written_exemplars we JOIN written_questions wq ON wq.id=we.written_question_id
      JOIN written_answer_versions wav ON wav.id=we.answer_version_id JOIN written_assessment_packages wap ON wap.id=we.package_id
      WHERE we.publication_status='PUBLISHED' ORDER BY we.published_at DESC""").fetchall()
    c.close(); return render_template('written_exemplars.html',exemplars=rows)


@app.route('/written-practice/exemplar-candidate/<int:candidate_id>/consent',methods=['POST'])
def written_exemplar_consent(candidate_id):
    if not require('student'): return redirect(url_for('login'))
    decision=request.form.get('consent_status','DECLINED')
    attribution=request.form.get('attribution_preference','ANONYMOUS')
    if decision not in ('OPTED_IN','DECLINED'): decision='DECLINED'
    if attribution not in ('ANONYMOUS','FIRST_NAME'): attribution='ANONYMOUS'
    c=db(); candidate=c.execute("SELECT * FROM written_exemplar_candidates WHERE id=? AND student_id=?",(candidate_id,session['user_id'])).fetchone()
    if not candidate: c.close(); abort(404)
    c.execute("""INSERT INTO written_exemplar_consents(candidate_id,student_id,consent_status,attribution_preference,
      guardian_confirmation,consent_text_version,consented_at) VALUES(?,?,?,?,?,'V6-CONSENT-1',?)
      ON CONFLICT(candidate_id,student_id) DO UPDATE SET consent_status=excluded.consent_status,
      attribution_preference=excluded.attribution_preference,guardian_confirmation=excluded.guardian_confirmation,
      consent_text_version=excluded.consent_text_version,consented_at=excluded.consented_at""",
      (candidate_id,session['user_id'],decision,attribution,1 if request.form.get('guardian_confirmation')=='1' else 0,datetime.now().isoformat(timespec='seconds')))
    c.execute("UPDATE written_exemplar_candidates SET consent_status=? WHERE id=?",(decision,candidate_id))
    if decision=='OPTED_IN': materialize_written_exemplar_if_ready(c,candidate_id)
    c.commit(); c.close(); flash('Your exemplar-sharing choice has been recorded separately from the platform Terms.','success')
    return redirect(url_for('written_attempt_result',attempt_id=candidate['attempt_id']))


@app.route('/admin/written-response')
def admin_written_response():
    if not require('admin'): return redirect(url_for('login'))
    c=db(); controls=c.execute("SELECT * FROM written_feature_controls ORDER BY feature_code").fetchall()
    packages=c.execute("""SELECT wap.*,(SELECT COUNT(*) FROM written_questions wq WHERE wq.package_id=wap.id) question_count
      FROM written_assessment_packages wap ORDER BY imported_at DESC""").fetchall()
    candidates=c.execute("""SELECT wec.*,u.full_name,wq.question_text,wa.current_mark,wa.maximum_mark
      FROM written_exemplar_candidates wec JOIN users u ON u.id=wec.student_id JOIN written_attempts wa ON wa.id=wec.attempt_id
      JOIN written_questions wq ON wq.id=wa.written_question_id ORDER BY wec.created_at DESC""").fetchall()
    pilots=c.execute("SELECT id,full_name,email,username,role,written_pilot_enabled FROM users WHERE role='student' ORDER BY full_name LIMIT 300").fetchall()
    pending_ocr=c.execute("""SELECT wa.*,u.full_name,wq.question_text FROM written_attempts wa JOIN users u ON u.id=wa.student_id
      JOIN written_questions wq ON wq.id=wa.written_question_id WHERE wa.entry_method='handwritten' AND wa.status IN ('OCR_PROVIDER_REQUIRED','OCR_CONFIRMATION_REQUIRED','RETAKE_REQUIRED') ORDER BY wa.id DESC""").fetchall()
    c.close(); return render_template('admin_written_response.html',controls=controls,packages=packages,candidates=candidates,pilots=pilots,pending_ocr=pending_ocr)


@app.route('/admin/written-response/import',methods=['POST'])
def admin_import_written_package():
    if not require('admin'): return redirect(url_for('login'))
    upload=request.files.get('package_file')
    if not upload or not upload.filename.lower().endswith('.json'):
        flash('Choose an approved Power House written-assessment JSON package.','error'); return redirect(url_for('admin_written_response'))
    try: payload=json.loads(upload.read().decode('utf-8'))
    except Exception as exc: flash(f'Package could not be read: {exc}','error'); return redirect(url_for('admin_written_response'))
    written_secret=os.environ.get('SCOREMAX_POWERHOUSE_WRITTEN_SECRET','').strip()
    report=validate_assessment_package(payload,shared_secret=written_secret,require_signature=(SCOREMAX_ENV=='production'))
    if not report['valid']:
        flash('Package rejected: '+' '.join(report['errors'][:4]),'error'); return redirect(url_for('admin_written_response'))
    checksum=report['checksum']; pid=str(payload['assessment_package_id']); version=str(payload['assessment_package_version'])
    c=db(); existing=c.execute("SELECT * FROM written_assessment_packages WHERE assessment_package_id=? AND assessment_package_version=?",(pid,version)).fetchone()
    if existing:
        c.close()
        flash('That exact immutable package is already imported.' if existing['export_checksum']==checksum else 'Rejected: same package ID/version has different content. Power House must issue a new version.','info' if existing['export_checksum']==checksum else 'error')
        return redirect(url_for('admin_written_response'))
    cur=c.execute("""INSERT INTO written_assessment_packages(assessment_package_id,assessment_package_version,framework_id,
      framework_version_id,blueprint_snapshot_id,subject_id,chapter_id,academic_approval_status,local_status,approved_at,
      source_approved_by,export_checksum,immutable_payload_json,imported_by) VALUES(?,?,?,?,?,?,?,?, 'IMPORTED',?,?,?,?,?)""",
      (pid,version,str(payload['framework_id']),str(payload['framework_version_id']),str(payload.get('blueprint_snapshot_id','')),
       str(payload['subject_id']),str(payload['chapter_id']),str(payload['academic_approval_status']),str(payload.get('approved_at','')),
       str(payload.get('approved_by','')),checksum,json.dumps(payload,sort_keys=True),session['user_id']))
    package_db_id=cur.lastrowid
    for q in payload.get('questions',[]):
        c.execute("""INSERT INTO written_questions(package_id,question_source_id,question_family_id,variant_id,question_type,
          question_text,command_verb,maximum_marks,estimated_time,difficulty,cognitive_demand,mastery_level,purpose,question_json,active)
          VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,1)""",(package_db_id,str(q['question_id']),str(q['question_family_id']),str(q.get('variant_id','')),
          str(q['question_type']),str(q['question_text']),str(q['command_verb']),float(q['maximum_marks']),int(q.get('estimated_time',0) or 0),
          str(q.get('difficulty','')),str(q.get('cognitive_demand','')),str(q.get('mastery_level','')),str(q.get('purpose','practice')),json.dumps(q,sort_keys=True)))
    c.commit(); c.close(); flash(f"Approved immutable package imported with {report['question_count']} written question(s). Activate it after review.",'success')
    return redirect(url_for('admin_written_response'))


@app.route('/admin/written-response/packages/<int:package_id>/activate',methods=['POST'])
def admin_activate_written_package(package_id):
    if not require('admin'): return redirect(url_for('login'))
    c=db(); package=c.execute("SELECT * FROM written_assessment_packages WHERE id=?",(package_id,)).fetchone()
    if not package or package['local_status'] not in ('IMPORTED','SUPERSEDED'):
        c.close(); flash('Only a reviewed imported package can be activated.','error'); return redirect(url_for('admin_written_response'))
    c.execute("UPDATE written_assessment_packages SET local_status='SUPERSEDED' WHERE framework_id=? AND framework_version_id=? AND subject_id=? AND chapter_id=? AND local_status='ACTIVE'",
              (package['framework_id'],package['framework_version_id'],package['subject_id'],package['chapter_id']))
    c.execute("UPDATE written_assessment_packages SET local_status='ACTIVE',activated_by=?,activated_at=? WHERE id=?",
              (session['user_id'],datetime.now().isoformat(timespec='seconds'),package_id))
    c.commit(); c.close(); flash('Written-assessment package activated. Historical attempts remain pinned to their original version.','success')
    return redirect(url_for('admin_written_response'))


@app.route('/admin/written-response/features/<feature_code>',methods=['POST'])
def admin_update_written_feature(feature_code):
    if not require('admin'): return redirect(url_for('login'))
    state=request.form.get('state','HIDDEN')
    if state not in ('HIDDEN','PILOT','LIVE'): state='HIDDEN'
    access=request.form.get('required_access_code','full_access')
    if access not in ACCESS_CODES: access='full_access'
    c=db(); c.execute("""UPDATE written_feature_controls SET state=?,required_access_code=?,available_from=?,available_to=?,
      updated_by=?,updated_at=CURRENT_TIMESTAMP WHERE feature_code=?""",(state,access,request.form.get('available_from',''),request.form.get('available_to',''),session['user_id'],feature_code))
    if feature_code=='written_exemplar_library':
        if state=='LIVE': c.execute("UPDATE written_exemplars SET publication_status='PUBLISHED',published_at=COALESCE(NULLIF(published_at,''),CURRENT_TIMESTAMP) WHERE publication_status='APPROVED_HIDDEN'")
        else: c.execute("UPDATE written_exemplars SET publication_status='APPROVED_HIDDEN' WHERE publication_status='PUBLISHED'")
    c.commit(); c.close(); flash('Written-response feature control updated.','success'); return redirect(url_for('admin_written_response'))


@app.route('/admin/written-response/pilot/<int:user_id>',methods=['POST'])
def admin_toggle_written_pilot(user_id):
    if not require('admin'): return redirect(url_for('login'))
    c=db(); c.execute("UPDATE users SET written_pilot_enabled=? WHERE id=? AND role='student'",(1 if request.form.get('enabled')=='1' else 0,user_id))
    c.commit(); c.close(); flash('Pilot access updated.','success'); return redirect(url_for('admin_written_response'))


@app.route('/admin/written-response/exemplars/<int:candidate_id>/review',methods=['POST'])
def admin_review_written_exemplar(candidate_id):
    if not require('admin'): return redirect(url_for('login'))
    decision=request.form.get('decision','REJECTED')
    if decision not in ('APPROVED','REJECTED'): decision='REJECTED'
    c=db(); candidate=c.execute("""SELECT wec.*,wa.written_question_id,wa.package_id,wa.package_version,wa.rubric_version,
      u.full_name,wav.version_type FROM written_exemplar_candidates wec JOIN written_attempts wa ON wa.id=wec.attempt_id
      JOIN users u ON u.id=wec.student_id JOIN written_answer_versions wav ON wav.id=wec.answer_version_id WHERE wec.id=?""",(candidate_id,)).fetchone()
    if not candidate: c.close(); abort(404)
    c.execute("UPDATE written_exemplar_candidates SET academic_status=?,academic_reviewer_id=?,academic_reviewed_at=?,academic_note=? WHERE id=?",
              (decision,session['user_id'],datetime.now().isoformat(timespec='seconds'),request.form.get('note',''),candidate_id))
    if decision=='APPROVED': materialize_written_exemplar_if_ready(c,candidate_id)
    c.commit(); c.close(); flash('Academic exemplar review recorded. Publication still requires explicit student consent and release control.','success')
    return redirect(url_for('admin_written_response'))


# ---------------------------------------------------------------------------
# V6.1 Teacher Discovery & Academic Messages routes
# ---------------------------------------------------------------------------

def community_feature_control(c,feature_code):
    return c.execute("SELECT * FROM community_feature_controls WHERE feature_code=?",(feature_code,)).fetchone()


def community_feature_available(c,user_id,feature_code):
    user=c.execute("""SELECT role,COALESCE(teacher_marketplace_pilot_enabled,0) teacher_marketplace_pilot_enabled,
      COALESCE(academic_messages_pilot_enabled,0) academic_messages_pilot_enabled FROM users WHERE id=?""",(user_id,)).fetchone()
    if not user: return False
    if user['role']=='admin': return True
    control=community_feature_control(c,feature_code)
    if not control or control['state']=='HIDDEN': return False
    if control['state']=='LIVE': return True
    if feature_code=='teacher_discovery': return bool(user['teacher_marketplace_pilot_enabled'])
    return bool(user['academic_messages_pilot_enabled'])




COMMUNITY_AGREEMENTS={
  'ACADEMIC_MESSAGES_USER': {'version':'V6.1-MESSAGES-1','roles':('student','teacher'),'name':'Academic Messages Community Rules'},
  'TEACHER_MARKETPLACE_CONDUCT': {'version':'V6.1-TEACHER-1','roles':('teacher',),'name':'Teacher Marketplace Professional Conduct'}
}


def community_agreement_ready(c,user_id,agreement_code):
    cfg=COMMUNITY_AGREEMENTS.get(agreement_code)
    if not cfg: return False
    user=c.execute("SELECT role FROM users WHERE id=?",(user_id,)).fetchone()
    if not user or user['role'] not in cfg['roles']: return False
    return bool(c.execute("""SELECT 1 FROM community_user_agreements WHERE user_id=? AND agreement_code=? AND agreement_version=?
      AND status='ACCEPTED' LIMIT 1""",(user_id,agreement_code,cfg['version'])).fetchone())


def required_community_agreements(role):
    return [dict(code=code,**cfg) for code,cfg in COMMUNITY_AGREEMENTS.items() if role in cfg['roles']]


def user_age_status(c,user_id):
    row=c.execute("SELECT dob FROM users WHERE id=?",(user_id,)).fetchone()
    raw=(row['dob'] or '').strip() if row else ''
    if not raw: return {'known':False,'age':None,'minor':None}
    try:
        born=datetime.fromisoformat(raw).date()
        today=datetime.now().date()
        age=today.year-born.year-((today.month,today.day)<(born.month,born.day))
        return {'known':True,'age':age,'minor':age<18}
    except Exception:
        return {'known':False,'age':None,'minor':None}


def guardian_messaging_consent(c,student_id):
    return c.execute("""SELECT agc.*,u.full_name parent_name FROM academic_guardian_consents agc
      JOIN parent_student_links psl ON psl.student_user_id=agc.student_id AND psl.parent_user_id=agc.parent_user_id AND psl.status='active'
      JOIN users u ON u.id=agc.parent_user_id
      WHERE agc.student_id=? AND agc.consent_scope='TEACHER_DISCOVERY_AND_MESSAGES' AND agc.status='APPROVED'
      ORDER BY agc.id DESC LIMIT 1""",(student_id,)).fetchone()


def student_messaging_safety_ready(c,student_id):
    age=user_age_status(c,student_id)
    if not age['known']:
        return {'allowed':False,'reason':'Complete or correct the student date of birth before using teacher messaging.','age':age,'consent':None}
    if not age['minor']:
        return {'allowed':True,'reason':'','age':age,'consent':None}
    consent=guardian_messaging_consent(c,student_id)
    return {'allowed':bool(consent),'reason':'' if consent else 'An active linked parent or guardian must approve Teacher Discovery and Academic Messages first.','age':age,'consent':consent}


@app.route('/community/agreements')
def community_agreements():
    if session.get('role') not in ('student','teacher'): return redirect(url_for('dashboard'))
    c=db(); items=[]
    for cfg in required_community_agreements(session.get('role')):
        cfg['accepted']=community_agreement_ready(c,session['user_id'],cfg['code']); items.append(cfg)
    c.close(); return render_template('community_agreements.html',agreements=items)


@app.route('/community/agreements/<agreement_code>',methods=['POST'])
def community_agreement_accept(agreement_code):
    cfg=COMMUNITY_AGREEMENTS.get(agreement_code)
    if not cfg or session.get('role') not in cfg['roles']: abort(403)
    decision=request.form.get('decision','ACCEPTED')
    if decision not in ('ACCEPTED','REVOKED'): decision='REVOKED'
    now=datetime.now().isoformat(timespec='seconds')
    c=db(); c.execute("""INSERT INTO community_user_agreements(user_id,agreement_code,agreement_version,status,accepted_at,revoked_at,ip_evidence)
      VALUES(?,?,?,?,?,?,?) ON CONFLICT(user_id,agreement_code,agreement_version) DO UPDATE SET status=excluded.status,
      accepted_at=excluded.accepted_at,revoked_at=excluded.revoked_at,ip_evidence=excluded.ip_evidence""",
      (session['user_id'],agreement_code,cfg['version'],decision,now if decision=='ACCEPTED' else '',now if decision=='REVOKED' else '',clean_text(request.headers.get('X-Forwarded-For','local'),120)))
    if decision=='REVOKED' and agreement_code=='ACADEMIC_MESSAGES_USER':
        c.execute("UPDATE academic_conversation_members SET status='AGREEMENT_REVOKED' WHERE user_id=?",(session['user_id'],))
        if session.get('role')=='student': c.execute("UPDATE academic_group_members SET status='AGREEMENT_REVOKED' WHERE user_id=?",(session['user_id'],))
    c.commit(); c.close(); flash('Community agreement accepted.' if decision=='ACCEPTED' else 'Agreement revoked. Messaging access has been suspended.','success')
    return redirect(url_for('community_agreements'))

@app.route('/parent/academic-messaging')
def parent_academic_messaging_controls():
    if not require('parent'): return redirect(url_for('dashboard'))
    c=db(); links=c.execute("""SELECT psl.*,u.full_name,u.dob,u.academic_level FROM parent_student_links psl
      JOIN users u ON u.id=psl.student_user_id WHERE psl.parent_user_id=? AND psl.status='active' ORDER BY u.full_name""",(session['user_id'],)).fetchall()
    students=[]
    for link in links:
        consent=c.execute("""SELECT * FROM academic_guardian_consents WHERE student_id=? AND parent_user_id=?
          AND consent_scope='TEACHER_DISCOVERY_AND_MESSAGES' ORDER BY id DESC LIMIT 1""",(link['student_user_id'],session['user_id'])).fetchone()
        students.append({'link':link,'age':user_age_status(c,link['student_user_id']),'consent':consent})
    c.close(); return render_template('parent_messaging_controls.html',students=students)


@app.route('/parent/academic-messaging/<int:student_id>',methods=['POST'])
def parent_academic_messaging_decision(student_id):
    if not require('parent'): return redirect(url_for('dashboard'))
    decision=request.form.get('decision','REVOKED')
    if decision not in ('APPROVED','REVOKED'): decision='REVOKED'
    c=db(); link=c.execute("""SELECT * FROM parent_student_links WHERE parent_user_id=? AND student_user_id=? AND status='active'""",
      (session['user_id'],student_id)).fetchone()
    if not link: c.close(); abort(403)
    now=datetime.now().isoformat(timespec='seconds')
    c.execute("""INSERT INTO academic_guardian_consents(student_id,parent_user_id,consent_scope,status,consent_text_version,granted_at,revoked_at)
      VALUES(?,?,'TEACHER_DISCOVERY_AND_MESSAGES',?,'V6.1-GUARDIAN-1',?,?)
      ON CONFLICT(student_id,parent_user_id,consent_scope) DO UPDATE SET status=excluded.status,consent_text_version=excluded.consent_text_version,
      granted_at=excluded.granted_at,revoked_at=excluded.revoked_at""",(student_id,session['user_id'],decision,now if decision=='APPROVED' else '',now if decision=='REVOKED' else ''))
    if decision=='REVOKED':
        conversation_ids=[r['conversation_id'] for r in c.execute("SELECT conversation_id FROM academic_conversation_members WHERE user_id=?",(student_id,)).fetchall()]
        c.execute("UPDATE academic_conversation_members SET status='GUARDIAN_REVOKED' WHERE user_id=?",(student_id,))
        c.execute("UPDATE academic_group_members SET status='GUARDIAN_REVOKED' WHERE user_id=?",(student_id,))
        if conversation_ids:
            placeholders=','.join('?' for _ in conversation_ids)
            c.execute(f"UPDATE academic_conversations SET status='LOCKED' WHERE conversation_type='ONE_TO_ONE' AND id IN ({placeholders})",conversation_ids)
    c.commit(); c.close()
    flash('Academic messaging consent approved.' if decision=='APPROVED' else 'Consent revoked. Existing direct conversations were locked and group membership was suspended.','success')
    return redirect(url_for('parent_academic_messaging_controls'))

def community_agreements_ready(c,user_id,role):
    required=['ACADEMIC_MESSAGES_USER']
    if role=='teacher': required.append('TEACHER_MARKETPLACE_CONDUCT')
    return all(community_agreement_ready(c,user_id,code) for code in required)



def teacher_rating_summary(c,teacher_id):
    row=c.execute("""SELECT COUNT(*) review_count,COALESCE(ROUND(AVG(rating),1),0) average_rating
      FROM teacher_reviews WHERE teacher_id=? AND moderation_status='PUBLISHED'""",(teacher_id,)).fetchone()
    return {'review_count':int(row['review_count'] or 0),'average_rating':float(row['average_rating'] or 0)}


def academic_conversation_access(c,conversation_id,user_id):
    return c.execute("""SELECT ac.*,acm.member_role,acm.status member_status FROM academic_conversations ac
      JOIN academic_conversation_members acm ON acm.conversation_id=ac.id
      WHERE ac.id=? AND acm.user_id=? AND acm.status='ACTIVE'""",(conversation_id,user_id)).fetchone()


def academic_users_blocked(c,user_a,user_b):
    return bool(c.execute("""SELECT 1 FROM academic_user_blocks WHERE active=1 AND
      ((blocker_id=? AND blocked_id=?) OR (blocker_id=? AND blocked_id=?)) LIMIT 1""",(user_a,user_b,user_b,user_a)).fetchone())


def get_or_create_group_conversation(c,group):
    conv=c.execute("SELECT * FROM academic_conversations WHERE group_id=? AND conversation_type='GROUP'",(group['id'],)).fetchone()
    if conv: return conv
    code_value='GRP-'+secrets.token_hex(5).upper()
    cur=c.execute("""INSERT INTO academic_conversations(conversation_code,conversation_type,teacher_id,group_id,purpose,academic_context_json)
      VALUES(?,'GROUP',?,?,?,?)""",(code_value,group['teacher_id'],group['id'],group['name'],json.dumps({'subject':group['subject'],'framework':group['framework']})))
    conv_id=cur.lastrowid
    c.execute("""INSERT OR IGNORE INTO academic_conversation_members(conversation_id,user_id,member_role,status)
      VALUES(?,?, 'teacher','ACTIVE')""",(conv_id,group['teacher_id']))
    return c.execute("SELECT * FROM academic_conversations WHERE id=?",(conv_id,)).fetchone()


@app.route('/teachers')
def teacher_directory():
    c=db()
    if not community_feature_available(c,session['user_id'],'teacher_discovery'):
        c.close(); flash('Teacher Discovery is currently available only to approved pilot users.','info'); return redirect(url_for('dashboard'))
    filters={k:clean_text(request.args.get(k,''),80) for k in ('q','subject','framework','service_type','delivery_mode')}
    profiles=c.execute("""SELECT tp.*,u.full_name,u.district,u.province FROM teacher_profiles tp JOIN users u ON u.id=tp.teacher_id
      WHERE tp.profile_status='PUBLISHED' AND u.account_status='active' ORDER BY tp.updated_at DESC""").fetchall()
    cards=[]
    for row in profiles:
        profile=dict(row); rating=teacher_rating_summary(c,row['teacher_id']); profile.update(rating)
        profile['average_rating_x10']=int(rating['average_rating']*10)
        listings=c.execute("SELECT * FROM teacher_service_listings WHERE teacher_id=? AND listing_status='PUBLISHED' ORDER BY service_type,title",(row['teacher_id'],)).fetchall()
        matched=[]
        for listing_row in listings:
            listing=dict(listing_row)
            hay=' '.join([profile.get('headline',''),profile.get('bio',''),listing.get('title',''),listing.get('description',''),listing.get('subject',''),listing.get('framework','')]).casefold()
            if filters['q'] and filters['q'].casefold() not in hay: continue
            if filters['subject'] and filters['subject'].casefold()!=listing.get('subject','').casefold() and filters['subject'].casefold() not in [x.casefold() for x in parse_list(profile.get('subjects_json'))]: continue
            if filters['framework'] and filters['framework'].casefold()!=listing.get('framework','').casefold() and filters['framework'].casefold() not in [x.casefold() for x in parse_list(profile.get('frameworks_json'))]: continue
            if filters['service_type'] and filters['service_type'].upper()!=listing.get('service_type','').upper(): continue
            if filters['delivery_mode'] and filters['delivery_mode'].upper()!=listing.get('delivery_mode','').upper(): continue
            listing['match_score']=teacher_match_score(profile,listing,filters); matched.append(listing)
        if matched:
            matched.sort(key=lambda x:x['match_score'],reverse=True); cards.append({'profile':profile,'listings':matched})
    cards.sort(key=lambda x:(max([l['match_score'] for l in x['listings']] or [0]),x['profile']['average_rating']),reverse=True)
    subjects=[r['subject'] for r in c.execute("SELECT DISTINCT subject FROM teacher_service_listings WHERE listing_status='PUBLISHED' AND subject<>'' ORDER BY subject").fetchall()]
    frameworks=[r['framework'] for r in c.execute("SELECT DISTINCT framework FROM teacher_service_listings WHERE listing_status='PUBLISHED' AND framework<>'' ORDER BY framework").fetchall()]
    c.close(); return render_template('teacher_directory.html',cards=cards,filters=filters,subjects=subjects,frameworks=frameworks)


@app.route('/teachers/<int:teacher_id>')
def teacher_public_profile(teacher_id):
    c=db()
    if not community_feature_available(c,session['user_id'],'teacher_discovery'):
        c.close(); return redirect(url_for('dashboard'))
    profile=c.execute("""SELECT tp.*,u.full_name,u.district,u.province FROM teacher_profiles tp JOIN users u ON u.id=tp.teacher_id
      WHERE tp.teacher_id=?""",(teacher_id,)).fetchone()
    if not profile or (profile['profile_status']!='PUBLISHED' and session.get('user_id')!=teacher_id and session.get('role')!='admin'):
        c.close(); abort(404)
    listings=c.execute("SELECT * FROM teacher_service_listings WHERE teacher_id=? AND listing_status='PUBLISHED' ORDER BY service_type,title",(teacher_id,)).fetchall()
    groups=c.execute("SELECT * FROM academic_groups WHERE teacher_id=? AND status='PUBLISHED' ORDER BY name",(teacher_id,)).fetchall()
    reviews=c.execute("""SELECT tr.*,u.full_name student_name FROM teacher_reviews tr JOIN users u ON u.id=tr.student_id
      WHERE tr.teacher_id=? AND tr.moderation_status='PUBLISHED' ORDER BY tr.id DESC LIMIT 50""",(teacher_id,)).fetchall()
    rating=teacher_rating_summary(c,teacher_id); c.close()
    return render_template('teacher_profile_public.html',profile=profile,listings=listings,groups=groups,reviews=reviews,rating=rating)


@app.route('/teacher/marketplace')
def teacher_marketplace_dashboard():
    if not require('teacher'): return redirect(url_for('dashboard'))
    c=db()
    if not community_feature_available(c,session['user_id'],'teacher_discovery'):
        c.close(); flash('Teacher Discovery pilot access has not been enabled for this account.','info'); return redirect(url_for('teacher_dashboard'))
    profile=c.execute("SELECT * FROM teacher_profiles WHERE teacher_id=?",(session['user_id'],)).fetchone()
    listings=c.execute("SELECT * FROM teacher_service_listings WHERE teacher_id=? ORDER BY id DESC",(session['user_id'],)).fetchall()
    enquiries=c.execute("""SELECT te.*,u.full_name student_name,tsl.title listing_title FROM teacher_enquiries te
      JOIN users u ON u.id=te.student_id LEFT JOIN teacher_service_listings tsl ON tsl.id=te.listing_id
      WHERE te.teacher_id=? ORDER BY te.id DESC""",(session['user_id'],)).fetchall()
    groups=c.execute("SELECT * FROM academic_groups WHERE teacher_id=? ORDER BY id DESC",(session['user_id'],)).fetchall()
    pending_members=c.execute("""SELECT agm.*,ag.name,u.full_name FROM academic_group_members agm JOIN academic_groups ag ON ag.id=agm.group_id
      JOIN users u ON u.id=agm.user_id WHERE ag.teacher_id=? AND agm.status='PENDING' ORDER BY agm.group_id,u.full_name""",(session['user_id'],)).fetchall()
    completeness=profile_completeness(dict(profile)) if profile else 0
    c.close(); return render_template('teacher_marketplace.html',profile=profile,listings=listings,enquiries=enquiries,groups=groups,pending_members=pending_members,completeness=completeness)


@app.route('/teacher/marketplace/profile',methods=['POST'])
def teacher_marketplace_profile_save():
    if not require('teacher'): return redirect(url_for('dashboard'))
    c=db()
    if not community_agreements_ready(c,session['user_id'],'teacher'):
        c.close(); flash('Accept the current Academic Messages and Teacher Marketplace conduct rules first.','info'); return redirect(url_for('community_agreements'))
    c.close()
    fields={
      'headline':clean_text(request.form.get('headline'),140),'bio':clean_text(request.form.get('bio'),2500),
      'subjects_json':json.dumps(parse_list(request.form.get('subjects'))),'frameworks_json':json.dumps(parse_list(request.form.get('frameworks'))),
      'qualifications_text':clean_text(request.form.get('qualifications_text'),1500),'experience_years':max(0,min(60,int(request.form.get('experience_years') or 0))),
      'languages_json':json.dumps(parse_list(request.form.get('languages'))),'delivery_modes_json':json.dumps(parse_list(request.form.get('delivery_modes'))),
      'platforms_json':json.dumps(parse_list(request.form.get('platforms'))),'location_text':clean_text(request.form.get('location_text'),160),
      'price_from_minor':max(0,int(float(request.form.get('price_from') or 0)*100)),'currency':'PKR',
      'availability_text':clean_text(request.form.get('availability_text'),500),'response_expectation_hours':max(1,min(168,int(request.form.get('response_expectation_hours') or 24))),
      'office_hours':clean_text(request.form.get('office_hours'),300),'intro_video_url':clean_text(request.form.get('intro_video_url'),500),
      'allow_one_to_one':1 if request.form.get('allow_one_to_one') else 0,'allow_groups':1 if request.form.get('allow_groups') else 0}
    public_policy=detect_message_policy(' '.join([fields['headline'],fields['bio'],fields['qualifications_text'],fields['location_text'],fields['availability_text'],fields['office_hours'],' '.join(parse_list(fields['platforms_json']))]),'teacher','TEXT')
    if public_policy['moderation_status']=='HELD':
        flash('Remove phone numbers, email addresses, WhatsApp details, payment requests or external links from the public teacher profile.','error'); return redirect(url_for('teacher_marketplace_dashboard'))
    requested='PENDING_REVIEW' if request.form.get('submit_for_review')=='1' else 'DRAFT'
    c=db(); existing=c.execute("SELECT * FROM teacher_profiles WHERE teacher_id=?",(session['user_id'],)).fetchone()
    if existing:
        c.execute("""UPDATE teacher_profiles SET headline=?,bio=?,subjects_json=?,frameworks_json=?,qualifications_text=?,experience_years=?,
          languages_json=?,delivery_modes_json=?,platforms_json=?,location_text=?,price_from_minor=?,currency=?,availability_text=?,
          response_expectation_hours=?,office_hours=?,intro_video_url=?,allow_one_to_one=?,allow_groups=?,profile_status=?,updated_at=CURRENT_TIMESTAMP
          WHERE teacher_id=?""",tuple(fields.values())+(requested,session['user_id']))
    else:
        c.execute("""INSERT INTO teacher_profiles(teacher_id,headline,bio,subjects_json,frameworks_json,qualifications_text,experience_years,
          languages_json,delivery_modes_json,platforms_json,location_text,price_from_minor,currency,availability_text,response_expectation_hours,
          office_hours,intro_video_url,allow_one_to_one,allow_groups,profile_status) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
          (session['user_id'],)+tuple(fields.values())+(requested,))
    c.commit(); c.close(); flash('Teacher profile saved'+(' and submitted for moderation.' if requested=='PENDING_REVIEW' else ' as a draft.'),'success')
    return redirect(url_for('teacher_marketplace_dashboard'))


@app.route('/teacher/marketplace/listings/create',methods=['POST'])
def teacher_listing_create():
    if not require('teacher'): return redirect(url_for('dashboard'))
    c=db()
    if not community_agreements_ready(c,session['user_id'],'teacher'):
        c.close(); flash('Accept the current community and professional-conduct rules first.','info'); return redirect(url_for('community_agreements'))
    profile=c.execute("SELECT * FROM teacher_profiles WHERE teacher_id=?",(session['user_id'],)).fetchone()
    if not profile:
        c.close(); flash('Create your teacher profile before adding a service.','error'); return redirect(url_for('teacher_marketplace_dashboard'))
    payload=dict(request.form); payload['price_minor']=max(0,int(float(request.form.get('price') or 0)*100)); payload['platforms']=request.form.get('platforms','')
    public_policy=detect_message_policy(' '.join([clean_text(request.form.get('title'),120),clean_text(request.form.get('description'),1800),clean_text(request.form.get('availability_text'),400),clean_text(request.form.get('platforms'),300)]),'teacher','TEXT')
    if public_policy['moderation_status']=='HELD':
        c.close(); flash('Remove phone numbers, email addresses, WhatsApp details, payment requests or external links from the public listing.','error'); return redirect(url_for('teacher_marketplace_dashboard'))
    report=validate_teacher_listing(payload)
    if not report['valid']:
        c.close(); flash(' '.join(report['errors']),'error'); return redirect(url_for('teacher_marketplace_dashboard'))
    status='PENDING_REVIEW' if request.form.get('submit_for_review')=='1' else 'DRAFT'
    c.execute("""INSERT INTO teacher_service_listings(teacher_id,profile_id,service_type,title,description,subject,framework,chapter_scope,
      delivery_mode,platform_options_json,price_minor,currency,pricing_unit,capacity,availability_text,listing_status)
      VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",(session['user_id'],profile['id'],report['service_type'],clean_text(request.form.get('title'),120),
      clean_text(request.form.get('description'),1800),clean_text(request.form.get('subject'),80),clean_text(request.form.get('framework'),80),
      clean_text(request.form.get('chapter_scope'),250),clean_text(request.form.get('delivery_mode','ONLINE'),20).upper(),json.dumps(report['platforms']),
      report['price_minor'],'PKR',clean_text(request.form.get('pricing_unit','PER_SESSION'),30).upper(),max(1,int(request.form.get('capacity') or 1)),
      clean_text(request.form.get('availability_text'),400),status))
    c.commit(); c.close(); flash('Service listing saved'+(' for moderation.' if status=='PENDING_REVIEW' else ' as a draft.'),'success'); return redirect(url_for('teacher_marketplace_dashboard'))


@app.route('/teacher/marketplace/listings/<int:listing_id>/status',methods=['POST'])
def teacher_listing_status(listing_id):
    if not require('teacher'): return redirect(url_for('dashboard'))
    requested=request.form.get('status','DRAFT')
    if requested not in ('DRAFT','PENDING_REVIEW','PAUSED'): requested='DRAFT'
    c=db(); c.execute("UPDATE teacher_service_listings SET listing_status=?,updated_at=CURRENT_TIMESTAMP WHERE id=? AND teacher_id=?",(requested,listing_id,session['user_id']))
    c.commit(); c.close(); flash('Listing status updated.','success'); return redirect(url_for('teacher_marketplace_dashboard'))


@app.route('/teachers/<int:teacher_id>/enquire',methods=['POST'])
def teacher_enquiry_create(teacher_id):
    if not require('student'): return redirect(url_for('dashboard'))
    c=db()
    if not community_agreements_ready(c,session['user_id'],'student'):
        c.close(); flash('Accept the current Academic Messages community rules first.','info'); return redirect(url_for('community_agreements'))
    if not community_feature_available(c,session['user_id'],'academic_messages'):
        c.close(); flash('Academic Messages pilot access is not enabled for your account.','info'); return redirect(url_for('teacher_directory'))
    safety=student_messaging_safety_ready(c,session['user_id'])
    if not safety['allowed']:
        c.close(); flash(safety['reason'],'error'); return redirect(url_for('teacher_directory'))
    listing_id=int(request.form.get('listing_id') or 0)
    listing=c.execute("SELECT * FROM teacher_service_listings WHERE id=? AND teacher_id=? AND listing_status='PUBLISHED'",(listing_id,teacher_id)).fetchone()
    if not listing:
        c.close(); flash('That teacher service is not currently available.','error'); return redirect(url_for('teacher_public_profile',teacher_id=teacher_id))
    support_need=clean_text(request.form.get('support_need'),700); initial=clean_text(request.form.get('initial_message'),1200)
    if not support_need:
        c.close(); flash('Tell the teacher what academic support you need.','error'); return redirect(url_for('teacher_public_profile',teacher_id=teacher_id))
    policy=detect_message_policy(initial,'student','TEXT')
    if policy['moderation_status']=='HELD':
        c.close(); flash('Please remove personal contact details, payment requests or unsafe wording. ScoreMax protects your private details.','error'); return redirect(url_for('teacher_public_profile',teacher_id=teacher_id))
    recent_enquiries=c.execute("SELECT COUNT(*) n FROM teacher_enquiries WHERE student_id=? AND created_at>=datetime('now','-1 day')",(session['user_id'],)).fetchone()['n']
    if recent_enquiries>=5 or not rate_limit(f"teacher-enquiry:{session['user_id']}",limit=5,window_seconds=86400):
        c.close(); flash('You have reached today’s enquiry limit.','error'); return redirect(url_for('teacher_public_profile',teacher_id=teacher_id))
    duplicate=c.execute("SELECT 1 FROM teacher_enquiries WHERE student_id=? AND teacher_id=? AND listing_id=? AND status='PENDING'",(session['user_id'],teacher_id,listing_id)).fetchone()
    if duplicate:
        c.close(); flash('You already have a pending enquiry for this service.','info'); return redirect(url_for('academic_messages_inbox'))
    context={'linked_attempt_id':int(request.form.get('linked_attempt_id') or 0),'linked_study_plan_activity_id':int(request.form.get('linked_study_plan_activity_id') or 0)}
    guardian_status='APPROVED' if safety['consent'] else 'NOT_REQUIRED'
    guardian_parent_id=safety['consent']['parent_user_id'] if safety['consent'] else None
    c.execute("""INSERT INTO teacher_enquiries(student_id,teacher_id,listing_id,subject,framework,chapter,support_need,preferred_mode,initial_message,academic_context_json,guardian_consent_status,guardian_parent_user_id)
      VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",(session['user_id'],teacher_id,listing_id,listing['subject'],listing['framework'],clean_text(request.form.get('chapter'),100),
      support_need,clean_text(request.form.get('preferred_mode'),40),initial,json.dumps(context),guardian_status,guardian_parent_id))
    c.commit(); c.close(); flash('Your enquiry was sent without sharing your phone number. The teacher must accept before messaging begins.','success'); return redirect(url_for('academic_messages_inbox'))


@app.route('/teacher/enquiries/<int:enquiry_id>/respond',methods=['POST'])
def teacher_enquiry_respond(enquiry_id):
    if not require('teacher'): return redirect(url_for('dashboard'))
    decision=request.form.get('decision','DECLINED')
    if decision not in ('ACCEPTED','DECLINED'): decision='DECLINED'
    c=db()
    if not community_agreements_ready(c,session['user_id'],'teacher'):
        c.close(); flash('Accept the current community and professional-conduct rules first.','info'); return redirect(url_for('community_agreements'))
    enquiry=c.execute("SELECT * FROM teacher_enquiries WHERE id=? AND teacher_id=?",(enquiry_id,session['user_id'])).fetchone()
    if not enquiry or enquiry['status']!='PENDING':
        c.close(); flash('That enquiry is no longer pending.','error'); return redirect(url_for('teacher_marketplace_dashboard'))
    if decision=='ACCEPTED':
        safety=student_messaging_safety_ready(c,enquiry['student_id'])
        if not safety['allowed']:
            c.execute("UPDATE teacher_enquiries SET status='CANCELLED_SAFETY',responded_at=? WHERE id=?",(datetime.now().isoformat(timespec='seconds'),enquiry_id)); c.commit(); c.close()
            flash('This enquiry cannot be accepted because the student account no longer meets age/guardian messaging requirements.','error'); return redirect(url_for('teacher_marketplace_dashboard'))
    now=datetime.now().isoformat(timespec='seconds')
    c.execute("UPDATE teacher_enquiries SET status=?,responded_at=?,accepted_at=? WHERE id=?",(decision,now,now if decision=='ACCEPTED' else '',enquiry_id))
    if decision=='ACCEPTED':
        code_value='MSG-'+secrets.token_hex(5).upper()
        cur=c.execute("""INSERT INTO academic_conversations(conversation_code,conversation_type,teacher_id,student_id,enquiry_id,purpose,academic_context_json)
          VALUES(?,'ONE_TO_ONE',?,?,?,?,?)""",(code_value,session['user_id'],enquiry['student_id'],enquiry_id,enquiry['support_need'],enquiry['academic_context_json']))
        conv_id=cur.lastrowid
        c.executemany("INSERT INTO academic_conversation_members(conversation_id,user_id,member_role,status) VALUES(?,?,?,'ACTIVE')",
          [(conv_id,session['user_id'],'teacher'),(conv_id,enquiry['student_id'],'student')])
        c.execute("""INSERT INTO teacher_engagements(enquiry_id,conversation_id,student_id,teacher_id,service_type,status)
          VALUES(?,?,?,?,COALESCE((SELECT service_type FROM teacher_service_listings WHERE id=?),'ONE_TO_ONE'),'ACTIVE')""",
          (enquiry_id,conv_id,enquiry['student_id'],session['user_id'],enquiry['listing_id']))
        c.execute("INSERT INTO academic_messages(conversation_id,sender_id,message_type,body,moderation_status) VALUES(?,?,'SYSTEM',?,'VISIBLE')",
          (conv_id,session['user_id'],'The teacher accepted this academic-support enquiry. Personal phone numbers remain private.'))
    c.commit(); c.close(); flash('Enquiry accepted. Academic Messages is now available.' if decision=='ACCEPTED' else 'Enquiry declined.','success')
    return redirect(url_for('teacher_marketplace_dashboard'))


@app.route('/teacher/groups/create',methods=['POST'])
def teacher_group_create():
    if not require('teacher'): return redirect(url_for('dashboard'))
    c=db()
    if not community_agreements_ready(c,session['user_id'],'teacher'):
        c.close(); flash('Accept the current community and professional-conduct rules first.','info'); return redirect(url_for('community_agreements'))
    if not community_feature_available(c,session['user_id'],'teacher_group_channels'):
        c.close(); flash('Teacher-led group channels are not enabled for this account.','info'); return redirect(url_for('teacher_marketplace_dashboard'))
    listing_id=int(request.form.get('listing_id') or 0)
    listing=c.execute("SELECT * FROM teacher_service_listings WHERE id=? AND teacher_id=? AND service_type='GROUP'",(listing_id,session['user_id'])).fetchone()
    if not listing:
        c.close(); flash('Create a group service listing first.','error'); return redirect(url_for('teacher_marketplace_dashboard'))
    policy=request.form.get('posting_policy','TEACHER_ONLY')
    if policy not in ('TEACHER_ONLY','ALL_MEMBERS'): policy='TEACHER_ONLY'
    cur=c.execute("""INSERT INTO academic_groups(teacher_id,listing_id,name,subject,framework,description,group_type,posting_policy,join_policy,max_members,status)
      VALUES(?,?,?,?,?,?, 'TUITION',?,'REQUEST',?,'PENDING_REVIEW')""",(session['user_id'],listing_id,clean_text(request.form.get('name'),120),listing['subject'],listing['framework'],
      clean_text(request.form.get('description'),1200),policy,max(2,min(200,int(request.form.get('max_members') or listing['capacity'] or 30)))))
    group=c.execute("SELECT * FROM academic_groups WHERE id=?",(cur.lastrowid,)).fetchone(); get_or_create_group_conversation(c,group)
    c.commit(); c.close(); flash('Teacher-led group submitted for moderation.','success'); return redirect(url_for('teacher_marketplace_dashboard'))


@app.route('/teacher/groups/<int:group_id>/join',methods=['POST'])
def teacher_group_join(group_id):
    if not require('student'): return redirect(url_for('dashboard'))
    c=db()
    if not community_agreements_ready(c,session['user_id'],'student'):
        c.close(); flash('Accept the current Academic Messages community rules first.','info'); return redirect(url_for('community_agreements'))
    group=c.execute("SELECT * FROM academic_groups WHERE id=? AND status='PUBLISHED'",(group_id,)).fetchone()
    safety=student_messaging_safety_ready(c,session['user_id'])
    if not safety['allowed']:
        c.close(); flash(safety['reason'],'error'); return redirect(url_for('teacher_directory'))
    if not group or not community_feature_available(c,session['user_id'],'teacher_group_channels'):
        c.close(); flash('That group is not available.','error'); return redirect(url_for('teacher_directory'))
    active_count=c.execute("SELECT COUNT(*) n FROM academic_group_members WHERE group_id=? AND status='ACTIVE'",(group_id,)).fetchone()['n']
    if active_count>=group['max_members']:
        c.close(); flash('This group is currently full.','error'); return redirect(url_for('teacher_public_profile',teacher_id=group['teacher_id']))
    c.execute("""INSERT INTO academic_group_members(group_id,user_id,member_role,status) VALUES(?,?,'student','PENDING')
      ON CONFLICT(group_id,user_id) DO UPDATE SET status='PENDING'""",(group_id,session['user_id']))
    c.commit(); c.close(); flash('Your request to join the teacher-led group has been sent.','success'); return redirect(url_for('academic_messages_inbox'))


@app.route('/teacher/groups/<int:group_id>/members/<int:user_id>',methods=['POST'])
def teacher_group_member_decision(group_id,user_id):
    if not require('teacher'): return redirect(url_for('dashboard'))
    decision=request.form.get('decision','REJECTED')
    if decision not in ('APPROVED','REJECTED'): decision='REJECTED'
    c=db()
    if not community_agreements_ready(c,session['user_id'],'teacher'):
        c.close(); flash('Accept the current community and professional-conduct rules first.','info'); return redirect(url_for('community_agreements'))
    group=c.execute("SELECT * FROM academic_groups WHERE id=? AND teacher_id=?",(group_id,session['user_id'])).fetchone()
    member=c.execute("SELECT * FROM academic_group_members WHERE group_id=? AND user_id=?",(group_id,user_id)).fetchone()
    if not group or not member or member['status']!='PENDING':
        c.close(); flash('That membership request is unavailable.','error'); return redirect(url_for('teacher_marketplace_dashboard'))
    if decision=='APPROVED':
        safety=student_messaging_safety_ready(c,user_id)
        if not safety['allowed']:
            c.execute("UPDATE academic_group_members SET status='REJECTED',approved_by=? WHERE group_id=? AND user_id=?",(session['user_id'],group_id,user_id)); c.commit(); c.close()
            flash('Membership could not be approved because student age/guardian messaging requirements are not currently met.','error'); return redirect(url_for('teacher_marketplace_dashboard'))
        active_count=c.execute("SELECT COUNT(*) n FROM academic_group_members WHERE group_id=? AND status='ACTIVE'",(group_id,)).fetchone()['n']
        if active_count>=group['max_members']:
            c.close(); flash('Group capacity has been reached.','error'); return redirect(url_for('teacher_marketplace_dashboard'))
        c.execute("UPDATE academic_group_members SET status='ACTIVE',joined_at=?,approved_by=? WHERE group_id=? AND user_id=?",(datetime.now().isoformat(timespec='seconds'),session['user_id'],group_id,user_id))
        conv=get_or_create_group_conversation(c,group)
        c.execute("INSERT OR IGNORE INTO academic_conversation_members(conversation_id,user_id,member_role,status) VALUES(?,?,'student','ACTIVE')",(conv['id'],user_id))
        c.execute("INSERT INTO academic_messages(conversation_id,sender_id,message_type,body,moderation_status) VALUES(?,?,'SYSTEM',?,'VISIBLE')",(conv['id'],session['user_id'],'A student joined this teacher-led academic group.'))
    else:
        c.execute("UPDATE academic_group_members SET status='REJECTED',approved_by=? WHERE group_id=? AND user_id=?",(session['user_id'],group_id,user_id))
    c.commit(); c.close(); flash('Group membership decision recorded.','success'); return redirect(url_for('teacher_marketplace_dashboard'))


@app.route('/messages')
def academic_messages_inbox():
    c=db()
    if session.get('role') in ('student','teacher') and not community_agreements_ready(c,session['user_id'],session.get('role')):
        c.close(); flash('Accept the current community rules before opening Academic Messages.','info'); return redirect(url_for('community_agreements'))
    if not community_feature_available(c,session['user_id'],'academic_messages'):
        c.close(); flash('Academic Messages is currently pilot-only.','info'); return redirect(url_for('dashboard'))
    conversations=c.execute("""SELECT ac.*,acm.member_role,
      CASE WHEN ac.conversation_type='GROUP' THEN ag.name ELSE COALESCE(other.full_name,'Academic conversation') END display_name,
      (SELECT body FROM academic_messages am WHERE am.conversation_id=ac.id AND am.moderation_status='VISIBLE' ORDER BY am.id DESC LIMIT 1) last_message,
      (SELECT created_at FROM academic_messages am WHERE am.conversation_id=ac.id AND am.moderation_status='VISIBLE' ORDER BY am.id DESC LIMIT 1) last_message_at
      FROM academic_conversation_members acm JOIN academic_conversations ac ON ac.id=acm.conversation_id
      LEFT JOIN academic_groups ag ON ag.id=ac.group_id
      LEFT JOIN users other ON other.id=CASE WHEN ac.teacher_id=? THEN ac.student_id ELSE ac.teacher_id END
      WHERE acm.user_id=? AND acm.status='ACTIVE' ORDER BY COALESCE(last_message_at,ac.created_at) DESC""",(session['user_id'],session['user_id'])).fetchall()
    enquiries=[]
    if session.get('role')=='student':
        enquiries=c.execute("""SELECT te.*,u.full_name teacher_name,tsl.title listing_title FROM teacher_enquiries te JOIN users u ON u.id=te.teacher_id
          LEFT JOIN teacher_service_listings tsl ON tsl.id=te.listing_id WHERE te.student_id=? ORDER BY te.id DESC""",(session['user_id'],)).fetchall()
    c.close(); return render_template('academic_messages_inbox.html',conversations=conversations,enquiries=enquiries)


@app.route('/messages/<int:conversation_id>')
def academic_conversation(conversation_id):
    c=db(); conv=academic_conversation_access(c,conversation_id,session['user_id'])
    if not conv: c.close(); abort(403)
    messages=c.execute("""SELECT am.*,u.full_name,u.role sender_role FROM academic_messages am JOIN users u ON u.id=am.sender_id
      WHERE am.conversation_id=? AND (am.moderation_status='VISIBLE' OR am.sender_id=?) ORDER BY am.id""",(conversation_id,session['user_id'])).fetchall()
    members=c.execute("""SELECT acm.*,u.full_name,u.role FROM academic_conversation_members acm JOIN users u ON u.id=acm.user_id
      WHERE acm.conversation_id=? AND acm.status='ACTIVE' ORDER BY CASE acm.member_role WHEN 'teacher' THEN 0 ELSE 1 END,u.full_name""",(conversation_id,)).fetchall()
    group=c.execute("SELECT * FROM academic_groups WHERE id=?",(conv['group_id'],)).fetchone() if conv['group_id'] else None
    engagement=c.execute("SELECT * FROM teacher_engagements WHERE conversation_id=? AND student_id=?",(conversation_id,session['user_id'])).fetchone() if session.get('role')=='student' else c.execute("SELECT * FROM teacher_engagements WHERE conversation_id=? LIMIT 1",(conversation_id,)).fetchone()
    if messages:
        c.execute("UPDATE academic_conversation_members SET last_read_message_id=? WHERE conversation_id=? AND user_id=?",(messages[-1]['id'],conversation_id,session['user_id'])); c.commit()
    c.close(); return render_template('academic_conversation.html',conversation=conv,messages=messages,members=members,group=group,engagement=engagement)


@app.route('/messages/<int:conversation_id>/send',methods=['POST'])
def academic_message_send(conversation_id):
    c=db()
    if session.get('role') in ('student','teacher') and not community_agreements_ready(c,session['user_id'],session.get('role')):
        c.close(); flash('Accept the current community rules before sending messages.','info'); return redirect(url_for('community_agreements'))
    conv=academic_conversation_access(c,conversation_id,session['user_id'])
    if not conv or conv['status']!='ACTIVE': c.close(); abort(403)
    if not community_feature_available(c,session['user_id'],'academic_messages'):
        c.close(); abort(403)
    if session.get('role')=='student':
        safety=student_messaging_safety_ready(c,session['user_id'])
        if not safety['allowed']:
            c.close(); flash(safety['reason'],'error'); return redirect(url_for('academic_messages_inbox'))
    if conv['conversation_type']=='ONE_TO_ONE':
        other_id=conv['teacher_id'] if session['user_id']!=conv['teacher_id'] else conv['student_id']
        if academic_users_blocked(c,session['user_id'],other_id):
            c.close(); flash('Messaging is unavailable because one participant has blocked the other.','error'); return redirect(url_for('academic_conversation',conversation_id=conversation_id))
    if conv['conversation_type']=='GROUP':
        group=c.execute("SELECT * FROM academic_groups WHERE id=?",(conv['group_id'],)).fetchone()
        if not group or group['status']!='PUBLISHED': c.close(); abort(403)
        if group['posting_policy']=='TEACHER_ONLY' and session.get('role')!='teacher':
            c.close(); flash('Only the teacher can post in this announcement-style group.','error'); return redirect(url_for('academic_conversation',conversation_id=conversation_id))
    recent_messages=c.execute("SELECT COUNT(*) n FROM academic_messages WHERE sender_id=? AND created_at>=datetime('now','-1 hour') AND message_type<>'SYSTEM'",(session['user_id'],)).fetchone()['n']
    if recent_messages>=40 or not rate_limit(f"academic-message:{session['user_id']}",limit=40,window_seconds=3600):
        c.close(); flash('Messaging rate limit reached. Please try again later.','error'); return redirect(url_for('academic_conversation',conversation_id=conversation_id))
    message_type=request.form.get('message_type','TEXT').upper()
    if message_type not in ('TEXT','MEETING_LINK','ASSESSMENT_LINK','STUDY_PLAN_LINK'): message_type='TEXT'
    body=clean_text(request.form.get('body'),6000)
    if not body: c.close(); flash('Write a message before sending.','error'); return redirect(url_for('academic_conversation',conversation_id=conversation_id))
    policy=detect_message_policy(body,session.get('role','student'),message_type)
    cur=c.execute("""INSERT INTO academic_messages(conversation_id,sender_id,message_type,body,moderation_status,policy_flags_json)
      VALUES(?,?,?,?,?,?)""",(conversation_id,session['user_id'],message_type,policy['clean_body'],policy['moderation_status'],json.dumps(policy['flags'])))
    if policy['moderation_status']=='HELD':
        c.execute("""INSERT INTO academic_message_reports(reporter_id,conversation_id,message_id,reported_user_id,category,detail,status)
          VALUES(?,?,?,?,?,'Automatically held by ScoreMax safety policy.','OPEN')""",(session['user_id'],conversation_id,cur.lastrowid,session['user_id'],'AUTOMATED_POLICY_HOLD'))
    c.commit(); c.close()
    flash('Message sent.' if policy['moderation_status']=='VISIBLE' else 'Message held for safety review. Remove personal contact details or unsafe content before trying again.','success' if policy['moderation_status']=='VISIBLE' else 'error')
    return redirect(url_for('academic_conversation',conversation_id=conversation_id))


@app.route('/messages/<int:conversation_id>/report',methods=['POST'])
def academic_message_report(conversation_id):
    c=db(); conv=academic_conversation_access(c,conversation_id,session['user_id'])
    if not conv: c.close(); abort(403)
    message_id=int(request.form.get('message_id') or 0); reported_user_id=int(request.form.get('reported_user_id') or 0)
    category=request.form.get('category','OTHER').upper()
    if category not in ('HARASSMENT','SPAM','SCAM','INAPPROPRIATE','SAFEGUARDING','CONTACT_SHARING','OTHER'): category='OTHER'
    c.execute("""INSERT INTO academic_message_reports(reporter_id,conversation_id,message_id,reported_user_id,category,detail)
      VALUES(?,?,?,?,?,?)""",(session['user_id'],conversation_id,message_id or None,reported_user_id or None,category,clean_text(request.form.get('detail'),1000)))
    c.commit(); c.close(); flash('Report submitted to ScoreMax moderation.','success'); return redirect(url_for('academic_conversation',conversation_id=conversation_id))


@app.route('/messages/<int:conversation_id>/block',methods=['POST'])
def academic_message_block(conversation_id):
    c=db(); conv=academic_conversation_access(c,conversation_id,session['user_id'])
    if not conv: c.close(); abort(403)
    blocked_id=int(request.form.get('blocked_user_id') or 0)
    if not c.execute("SELECT 1 FROM academic_conversation_members WHERE conversation_id=? AND user_id=?",(conversation_id,blocked_id)).fetchone() or blocked_id==session['user_id']:
        c.close(); abort(400)
    c.execute("""INSERT INTO academic_user_blocks(blocker_id,blocked_id,active,reason) VALUES(?,?,1,?)
      ON CONFLICT(blocker_id,blocked_id) DO UPDATE SET active=1,reason=excluded.reason,updated_at=CURRENT_TIMESTAMP""",
      (session['user_id'],blocked_id,clean_text(request.form.get('reason'),300)))
    if conv['conversation_type']=='ONE_TO_ONE': c.execute("UPDATE academic_conversations SET status='LOCKED' WHERE id=?",(conversation_id,))
    c.commit(); c.close(); flash('User blocked. Direct messaging in this conversation is now locked.','success'); return redirect(url_for('academic_messages_inbox'))


@app.route('/messages/<int:conversation_id>/confirm-session',methods=['POST'])
def academic_confirm_engagement(conversation_id):
    if session.get('role') not in ('student','teacher'): return redirect(url_for('dashboard'))
    c=db(); conv=academic_conversation_access(c,conversation_id,session['user_id'])
    if not conv or conv['conversation_type']!='ONE_TO_ONE': c.close(); abort(403)
    engagement=c.execute("SELECT * FROM teacher_engagements WHERE conversation_id=?",(conversation_id,)).fetchone()
    if not engagement: c.close(); abort(404)
    field='teacher_confirmed' if session.get('role')=='teacher' else 'student_confirmed'
    c.execute(f"UPDATE teacher_engagements SET {field}=1,session_date=COALESCE(NULLIF(session_date,''),?) WHERE id=?",(clean_text(request.form.get('session_date'),20),engagement['id']))
    updated=c.execute("SELECT * FROM teacher_engagements WHERE id=?",(engagement['id'],)).fetchone()
    if updated['teacher_confirmed'] and updated['student_confirmed']:
        c.execute("UPDATE teacher_engagements SET status='VERIFIED_COMPLETED',completed_at=? WHERE id=?",(datetime.now().isoformat(timespec='seconds'),engagement['id']))
    c.commit(); c.close(); flash('Session confirmation recorded. A rating becomes available only after both sides confirm.','success'); return redirect(url_for('academic_conversation',conversation_id=conversation_id))


@app.route('/teachers/<int:teacher_id>/reviews',methods=['POST'])
def teacher_review_create(teacher_id):
    if not require('student'): return redirect(url_for('dashboard'))
    engagement_id=int(request.form.get('engagement_id') or 0); rating=max(1,min(5,int(request.form.get('rating') or 0)))
    c=db()
    if not community_agreements_ready(c,session['user_id'],'student'):
        c.close(); flash('Accept the current Academic Messages community rules first.','info'); return redirect(url_for('community_agreements'))
    engagement=c.execute("""SELECT * FROM teacher_engagements WHERE id=? AND student_id=? AND teacher_id=? AND status='VERIFIED_COMPLETED'""",
      (engagement_id,session['user_id'],teacher_id)).fetchone()
    if not engagement:
        c.close(); flash('Only a verified completed teacher interaction can be reviewed.','error'); return redirect(url_for('teacher_public_profile',teacher_id=teacher_id))
    review_text=clean_text(request.form.get('review_text'),1200); policy=detect_message_policy(review_text,'student','TEXT')
    moderation='HELD' if policy['moderation_status']=='HELD' else 'PENDING'
    try:
        c.execute("""INSERT INTO teacher_reviews(engagement_id,student_id,teacher_id,rating,review_text,moderation_status,policy_flags_json)
          VALUES(?,?,?,?,?,?,?)""",(engagement_id,session['user_id'],teacher_id,rating,review_text,moderation,json.dumps(policy['flags'])))
        c.commit(); flash('Your verified review was submitted for moderation.','success')
    except sqlite3.IntegrityError:
        flash('You have already reviewed this verified interaction.','info')
    c.close(); return redirect(url_for('teacher_public_profile',teacher_id=teacher_id))


@app.route('/admin/community')
def admin_community():
    if not require('admin'): return redirect(url_for('login'))
    c=db(); controls=c.execute("SELECT * FROM community_feature_controls ORDER BY feature_code").fetchall()
    profiles=c.execute("""SELECT tp.*,u.full_name,u.email FROM teacher_profiles tp JOIN users u ON u.id=tp.teacher_id
      WHERE tp.profile_status IN ('PENDING_REVIEW','PUBLISHED','SUSPENDED') ORDER BY CASE tp.profile_status WHEN 'PENDING_REVIEW' THEN 0 ELSE 1 END,tp.id DESC""").fetchall()
    listings=c.execute("""SELECT tsl.*,u.full_name FROM teacher_service_listings tsl JOIN users u ON u.id=tsl.teacher_id
      WHERE tsl.listing_status IN ('PENDING_REVIEW','PUBLISHED','REJECTED') ORDER BY CASE tsl.listing_status WHEN 'PENDING_REVIEW' THEN 0 ELSE 1 END,tsl.id DESC""").fetchall()
    groups=c.execute("""SELECT ag.*,u.full_name FROM academic_groups ag JOIN users u ON u.id=ag.teacher_id
      WHERE ag.status IN ('PENDING_REVIEW','PUBLISHED','REJECTED','CLOSED') ORDER BY CASE ag.status WHEN 'PENDING_REVIEW' THEN 0 ELSE 1 END,ag.id DESC""").fetchall()
    reports=c.execute("""SELECT amr.*,reporter.full_name reporter_name,reported.full_name reported_name,am.body message_body FROM academic_message_reports amr
      JOIN users reporter ON reporter.id=amr.reporter_id LEFT JOIN users reported ON reported.id=amr.reported_user_id LEFT JOIN academic_messages am ON am.id=amr.message_id
      WHERE amr.status='OPEN' ORDER BY amr.id DESC""").fetchall()
    reviews=c.execute("""SELECT tr.*,student.full_name student_name,teacher.full_name teacher_name FROM teacher_reviews tr JOIN users student ON student.id=tr.student_id
      JOIN users teacher ON teacher.id=tr.teacher_id WHERE tr.moderation_status IN ('PENDING','HELD') ORDER BY tr.id DESC""").fetchall()
    users=c.execute("SELECT id,full_name,role,teacher_marketplace_pilot_enabled,academic_messages_pilot_enabled FROM users WHERE role IN ('student','teacher') ORDER BY role,full_name").fetchall()
    c.close(); return render_template('admin_community.html',controls=controls,profiles=profiles,listings=listings,groups=groups,reports=reports,reviews=reviews,users=users)


@app.route('/admin/community/features/<feature_code>',methods=['POST'])
def admin_community_feature_update(feature_code):
    if not require('admin'): return redirect(url_for('login'))
    state=request.form.get('state','HIDDEN')
    if state not in ('HIDDEN','PILOT','LIVE'): state='HIDDEN'
    if feature_code=='student_direct_messages' and state!='HIDDEN':
        flash('Unrestricted student-to-student direct messages remain blocked until safeguarding operations are independently approved.','error'); return redirect(url_for('admin_community'))
    if state=='LIVE' and SCOREMAX_ENV=='production':
        if os.environ.get('SCOREMAX_COMMUNITY_LIVE','0')!='1' or not os.environ.get('SCOREMAX_SAFETY_CONTACT','').strip():
            flash('Production community launch requires SCOREMAX_COMMUNITY_LIVE=1 and SCOREMAX_SAFETY_CONTACT.','error'); return redirect(url_for('admin_community'))
    c=db(); c.execute("UPDATE community_feature_controls SET state=?,updated_by=?,updated_at=CURRENT_TIMESTAMP WHERE feature_code=?",(state,session['user_id'],feature_code))
    c.commit(); c.close(); flash('Community feature state updated.','success'); return redirect(url_for('admin_community'))


@app.route('/admin/community/pilot/<int:user_id>',methods=['POST'])
def admin_community_pilot(user_id):
    if not require('admin'): return redirect(url_for('login'))
    feature=request.form.get('feature','teacher_discovery'); enabled=1 if request.form.get('enabled')=='1' else 0
    field='teacher_marketplace_pilot_enabled' if feature=='teacher_discovery' else 'academic_messages_pilot_enabled'
    c=db(); c.execute(f"UPDATE users SET {field}=? WHERE id=? AND role IN ('student','teacher')",(enabled,user_id)); c.commit(); c.close()
    flash('Pilot access updated.','success'); return redirect(url_for('admin_community'))


@app.route('/admin/community/teachers/<int:teacher_id>/profile',methods=['POST'])
def admin_teacher_profile_review(teacher_id):
    if not require('admin'): return redirect(url_for('login'))
    decision=request.form.get('decision','REJECTED')
    if decision not in ('PUBLISHED','REJECTED','SUSPENDED'): decision='REJECTED'
    c=db(); profile=c.execute("SELECT * FROM teacher_profiles WHERE teacher_id=?",(teacher_id,)).fetchone()
    if not profile: c.close(); abort(404)
    if decision=='PUBLISHED':
        if profile['identity_verification_status']!='VERIFIED':
            c.close(); flash('Identity verification is required before a teacher profile can be published.','error'); return redirect(url_for('admin_community'))
        if profile_completeness(dict(profile))<60:
            c.close(); flash('Teacher profile must be at least 60% complete before publication.','error'); return redirect(url_for('admin_community'))
    c.execute("""UPDATE teacher_profiles SET profile_status=?,moderation_note=?,moderated_by=?,moderated_at=?,updated_at=CURRENT_TIMESTAMP WHERE teacher_id=?""",
      (decision,clean_text(request.form.get('note'),600),session['user_id'],datetime.now().isoformat(timespec='seconds'),teacher_id))
    c.commit(); c.close(); flash('Teacher profile moderation recorded.','success'); return redirect(url_for('admin_community'))


@app.route('/admin/community/teachers/<int:teacher_id>/verify',methods=['POST'])
def admin_teacher_verification(teacher_id):
    if not require('admin'): return redirect(url_for('login'))
    vtype=request.form.get('verification_type','identity')
    field={'identity':'identity_verification_status','qualification':'qualification_verification_status','experience':'experience_verification_status'}.get(vtype)
    status=request.form.get('status','UNVERIFIED')
    if not field or status not in ('UNVERIFIED','PENDING','VERIFIED','REJECTED'): return redirect(url_for('admin_community'))
    c=db(); profile=c.execute("SELECT * FROM teacher_profiles WHERE teacher_id=?",(teacher_id,)).fetchone()
    if not profile: c.close(); abort(404)
    previous=profile[field]
    c.execute(f"UPDATE teacher_profiles SET {field}=?,updated_at=CURRENT_TIMESTAMP WHERE teacher_id=?",(status,teacher_id))
    c.execute("""INSERT INTO teacher_verification_events(teacher_id,verification_type,previous_status,new_status,reviewer_id,evidence_note)
      VALUES(?,?,?,?,?,?)""",(teacher_id,vtype.upper(),previous,status,session['user_id'],clean_text(request.form.get('evidence_note'),700)))
    c.commit(); c.close(); flash('Verification status updated. Verification is not the same as ScoreMax academic endorsement.','success'); return redirect(url_for('admin_community'))


@app.route('/admin/community/listings/<int:listing_id>',methods=['POST'])
def admin_teacher_listing_review(listing_id):
    if not require('admin'): return redirect(url_for('login'))
    decision=request.form.get('decision','REJECTED')
    if decision not in ('PUBLISHED','REJECTED','PAUSED'): decision='REJECTED'
    c=db(); listing=c.execute("SELECT * FROM teacher_service_listings WHERE id=?",(listing_id,)).fetchone()
    if not listing: c.close(); abort(404)
    profile=c.execute("SELECT * FROM teacher_profiles WHERE teacher_id=?",(listing['teacher_id'],)).fetchone()
    if decision=='PUBLISHED' and (not profile or profile['profile_status']!='PUBLISHED'):
        c.close(); flash('Publish the moderated teacher profile before publishing its service listing.','error'); return redirect(url_for('admin_community'))
    c.execute("""UPDATE teacher_service_listings SET listing_status=?,moderation_note=?,moderated_by=?,moderated_at=?,updated_at=CURRENT_TIMESTAMP WHERE id=?""",
      (decision,clean_text(request.form.get('note'),600),session['user_id'],datetime.now().isoformat(timespec='seconds'),listing_id))
    c.commit(); c.close(); flash('Service listing moderation recorded.','success'); return redirect(url_for('admin_community'))


@app.route('/admin/community/groups/<int:group_id>',methods=['POST'])
def admin_teacher_group_review(group_id):
    if not require('admin'): return redirect(url_for('login'))
    decision=request.form.get('decision','REJECTED')
    if decision not in ('PUBLISHED','REJECTED','CLOSED'): decision='REJECTED'
    c=db(); group=c.execute("SELECT * FROM academic_groups WHERE id=?",(group_id,)).fetchone()
    if not group: c.close(); abort(404)
    listing=c.execute("SELECT * FROM teacher_service_listings WHERE id=?",(group['listing_id'],)).fetchone()
    if decision=='PUBLISHED' and (not listing or listing['listing_status']!='PUBLISHED'):
        c.close(); flash('The linked group service must be published first.','error'); return redirect(url_for('admin_community'))
    c.execute("UPDATE academic_groups SET status=?,updated_at=CURRENT_TIMESTAMP WHERE id=?",(decision,group_id)); c.commit(); c.close()
    flash('Teacher-led group moderation recorded.','success'); return redirect(url_for('admin_community'))


@app.route('/admin/community/reports/<int:report_id>',methods=['POST'])
def admin_academic_report_resolve(report_id):
    if not require('admin'): return redirect(url_for('login'))
    action=request.form.get('action','RESOLVED')
    if action not in ('RESOLVED','DISMISSED','REMOVE_MESSAGE','SUSPEND_USER'): action='RESOLVED'
    c=db(); report=c.execute("SELECT * FROM academic_message_reports WHERE id=?",(report_id,)).fetchone()
    if not report: c.close(); abort(404)
    if action=='REMOVE_MESSAGE' and report['message_id']:
        c.execute("UPDATE academic_messages SET moderation_status='REMOVED',removed_at=?,removed_by=?,removal_reason=? WHERE id=?",
          (datetime.now().isoformat(timespec='seconds'),session['user_id'],clean_text(request.form.get('resolution'),500),report['message_id']))
    if action=='SUSPEND_USER' and report['reported_user_id']:
        c.execute("UPDATE users SET account_status='disabled',session_version=COALESCE(session_version,0)+1 WHERE id=?",(report['reported_user_id'],))
    c.execute("UPDATE academic_message_reports SET status=?,resolution=?,resolved_by=?,resolved_at=? WHERE id=?",
      ('DISMISSED' if action=='DISMISSED' else 'RESOLVED',clean_text(request.form.get('resolution'),800),session['user_id'],datetime.now().isoformat(timespec='seconds'),report_id))
    c.commit(); c.close(); flash('Moderation action recorded.','success'); return redirect(url_for('admin_community'))


@app.route('/admin/community/reviews/<int:review_id>',methods=['POST'])
def admin_teacher_review_moderate(review_id):
    if not require('admin'): return redirect(url_for('login'))
    decision=request.form.get('decision','REMOVED')
    if decision not in ('PUBLISHED','REMOVED'): decision='REMOVED'
    c=db(); c.execute("UPDATE teacher_reviews SET moderation_status=?,moderated_by=?,moderated_at=? WHERE id=?",
      (decision,session['user_id'],datetime.now().isoformat(timespec='seconds'),review_id)); c.commit(); c.close()
    flash('Teacher review moderation recorded.','success'); return redirect(url_for('admin_community'))



def _qa_identity_or_none(c):
    return qa_synthetic.identity_for_user(c,session.get('user_id')) if session.get('role')=='qa_student' else None


@app.route('/qa/synthetic')
def qa_synthetic_home():
    c=db(); identity=_qa_identity_or_none(c)
    if not identity:
        c.close(); return redirect(url_for('login'))
    batches=c.execute("""SELECT b.*,
      (SELECT COUNT(*) FROM mastery_lab_questions q WHERE q.batch_id=b.id AND q.active=1) question_count
      FROM mastery_lab_batches b ORDER BY b.id DESC LIMIT 30""").fetchall()
    external_question_id=str(request.args.get('external_question_id','') or '').strip()
    external_version=str(request.args.get('external_version','') or '').strip()
    if external_question_id and external_version:
        questions=c.execute("""SELECT q.id,q.batch_id,q.external_question_id,q.external_version,q.family_type,q.response_mode,
          q.programme,q.subject,q.chapter,q.topic,q.mastery_level,q.question_text,
          b.batch_code FROM mastery_lab_questions q JOIN mastery_lab_batches b ON b.id=q.batch_id
          WHERE q.active=1 AND q.external_question_id=? AND q.external_version=? ORDER BY q.id""",(external_question_id,external_version)).fetchall()
    elif external_question_id:
        questions=c.execute("""SELECT q.id,q.batch_id,q.external_question_id,q.external_version,q.family_type,q.response_mode,
          q.programme,q.subject,q.chapter,q.topic,q.mastery_level,q.question_text,
          b.batch_code FROM mastery_lab_questions q JOIN mastery_lab_batches b ON b.id=q.batch_id
          WHERE q.active=1 AND q.external_question_id=? ORDER BY q.id""",(external_question_id,)).fetchall()
    else:
        questions=c.execute("""SELECT q.id,q.batch_id,q.external_question_id,q.external_version,q.family_type,q.response_mode,
          q.programme,q.subject,q.chapter,q.topic,q.mastery_level,q.question_text,
          b.batch_code FROM mastery_lab_questions q JOIN mastery_lab_batches b ON b.id=q.batch_id
          WHERE q.active=1 ORDER BY q.id LIMIT 250""").fetchall()
    recent=c.execute("""SELECT s.*,q.external_question_id,q.question_text FROM mastery_lab_e2e_sessions s
      JOIN mastery_lab_learner_identities i ON i.id=s.identity_id
      JOIN mastery_lab_questions q ON q.id=s.question_id
      WHERE i.user_id=? ORDER BY s.id DESC LIMIT 20""",(session['user_id'],)).fetchall()
    c.close()
    return render_template('qa_synthetic_home.html',identity=identity,batches=batches,questions=questions,recent=recent)


@app.route('/qa/synthetic/start',methods=['POST'])
def qa_synthetic_start():
    c=db(); identity=_qa_identity_or_none(c)
    if not identity:
        c.close(); return redirect(url_for('login'))
    try:
        question_id=int(request.form.get('question_id','0') or 0)
        expected_mode=request.form.get('expected_mode','OBSERVE_ONLY').strip().upper()
        if identity['learner_kind']=='VISUAL_SEMANTIC':
            expected_mode='OBSERVE_ONLY'
        created=qa_synthetic.create_e2e_session(
            c,user_id=session['user_id'],question_id=question_id,expected_mode=expected_mode)
        sid=int(created['id'])
    except Exception as exc:
        c.close(); flash(f'QA session could not be created: {exc}','error'); return redirect(url_for('qa_synthetic_home'))
    c.close(); return redirect(url_for('qa_synthetic_session',session_id=sid))


@app.route('/qa/synthetic/session/<int:session_id>',methods=['GET','POST'])
def qa_synthetic_session(session_id):
    c=db(); identity=_qa_identity_or_none(c)
    if not identity:
        c.close(); return redirect(url_for('login'))
    summary=qa_synthetic.session_summary(c,session_id,user_id=session['user_id'])
    if not summary:
        c.close(); abort(404)
    q=summary['question']
    if request.method=='POST':
        answer_values=request.form.getlist('answer')
        selected=','.join(answer_values) if len(answer_values)>1 else (answer_values[0] if answer_values else '')
        # Score the same response through both the live ScoreMax marking adapter and the
        # existing Mastery Laboratory scorer. No live attempt/mastery function is invoked.
        live_ok,live_marks,misconception=mark_question_response(q,selected,{})
        live_result={'is_correct':bool(live_ok),'awarded_marks':float(live_marks or 0),'misconception':misconception}
        lab_question=c.execute("SELECT * FROM mastery_lab_questions WHERE id=?",(summary['question_id'],)).fetchone()
        lab_result=mastery_lab.score_lab_response(dict(lab_question),selected)
        outcome=qa_synthetic.record_attempt(
            c,session_id=session_id,user_id=session['user_id'],response=selected,
            live_result=live_result,lab_result=lab_result)
        c.close()
        return redirect(url_for('qa_synthetic_result',session_id=session_id,attempt_seq=outcome['attempt_seq']))
    answer_cfg=safe_json(q['answer_config'],{})
    marking_cfg=safe_json(q['marking_config'],{})
    options=answer_cfg.get('options') or [
        {'id':code,'text':q[key]} for code,key in [('A','option_a'),('B','option_b'),('C','option_c'),('D','option_d')] if q.get(key)
    ]
    qtype=canonical_question_type(q)
    # One-question shell reuses the production learner template and question card.
    assessment={'mode':'practice'}
    render_material={'question':q,'qtype':qtype,'options':options,'answer_cfg':answer_cfg,'marking_cfg':marking_cfg}
    render_checksum=hashlib.sha256(json.dumps(render_material,ensure_ascii=False,sort_keys=True,default=str,separators=(',',':')).encode('utf-8')).hexdigest()
    c.execute("UPDATE mastery_lab_e2e_sessions SET render_checksum_sha256=? WHERE id=?",(render_checksum,session_id)); c.commit(); c.close()
    return render_template(
        'take_test_v4.html',assessment=assessment,assessment_id=0,q=q,idx=0,total=1,
        ids=[q['id']],answers={},flagged=set(),remaining_seconds=None,answered_count=0,
        qtype=qtype,options=options,answer_cfg=answer_cfg,marking_cfg=marking_cfg,
        confidence={},response_times={},exam_meta={},qa_sandbox=True,qa_session_id=session_id,
        qa_identity=identity,qa_render_checksum=render_checksum)


@app.route('/qa/synthetic/session/<int:session_id>/result/<int:attempt_seq>')
def qa_synthetic_result(session_id,attempt_seq):
    c=db(); identity=_qa_identity_or_none(c)
    if not identity:
        c.close(); return redirect(url_for('login'))
    summary=qa_synthetic.session_summary(c,session_id,user_id=session['user_id'])
    attempt=c.execute("""SELECT * FROM mastery_lab_e2e_attempts WHERE e2e_session_id=? AND attempt_seq=?""",
                      (session_id,attempt_seq)).fetchone()
    c.close()
    if not summary or not attempt: abort(404)
    return render_template('qa_synthetic_result.html',summary=summary,attempt=attempt,identity=identity)


@app.route('/qa/synthetic/session/<int:session_id>/visual-capture',methods=['POST'])
def qa_synthetic_visual_capture(session_id):
    c=db(); identity=_qa_identity_or_none(c)
    if not identity or identity['learner_kind']!='VISUAL_SEMANTIC':
        c.close(); abort(403)
    summary=qa_synthetic.session_summary(c,session_id,user_id=session['user_id'])
    if not summary:
        c.close(); abort(404)
    upload=request.files.get('screenshot')
    if not upload or not upload.filename:
        c.close(); return jsonify({'ok':False,'error':'screenshot_required'}),400
    ext=Path(upload.filename).suffix.lower()
    if ext not in {'.png','.jpg','.jpeg','.webp'}:
        c.close(); return jsonify({'ok':False,'error':'unsupported_image_type'}),400
    root=Path(os.environ.get('SCOREMAX_QA_EVIDENCE_DIR',BASE/'qa_evidence'))
    root.mkdir(parents=True,exist_ok=True)
    path=root/f"SM-QA-{session_id}-{secrets.token_hex(6)}{ext}"
    upload.save(path)
    try:
        with Image.open(path) as im:
            im.verify()
        viewport=safe_json(request.form.get('viewport_json','{}'),{})
        metadata=safe_json(request.form.get('render_metadata_json','{}'),{})
        result=qa_synthetic.record_visual_capture(
            c,session_id=session_id,user_id=session['user_id'],screenshot_path=str(path),
            viewport=viewport,render_metadata=metadata)
    except Exception as exc:
        path.unlink(missing_ok=True); c.close()
        return jsonify({'ok':False,'error':str(exc)}),400
    c.close(); return jsonify({'ok':True,**result})


@app.route('/qa/synthetic/session/<int:session_id>/visual-judgement',methods=['POST'])
def qa_synthetic_visual_judgement(session_id):
    c=db(); identity=_qa_identity_or_none(c)
    if not identity or identity['learner_kind']!='VISUAL_SEMANTIC':
        c.close(); abort(403)
    summary=qa_synthetic.session_summary(c,session_id,user_id=session['user_id'])
    if not summary:
        c.close(); abort(404)
    try:
        visual_review_id=int(request.form.get('visual_review_id','0') or 0)
        findings=safe_json(request.form.get('findings_json','[]'),[])
        if not isinstance(findings,list): raise ValueError('findings_json must be a list')
        result=qa_synthetic.record_visual_judgement(
            c,visual_review_id=visual_review_id,session_id=session_id,user_id=session['user_id'],
            judgement=request.form.get('judgement',''),findings=findings,
            judge_type=request.form.get('judge_type',''),judge_version=request.form.get('judge_version',''))
    except Exception as exc:
        c.close(); return jsonify({'ok':False,'error':str(exc)}),400
    c.close(); return jsonify({'ok':True,**result})



@app.route('/admin/mastery-lab')
def admin_mastery_lab():
    if not require('admin'): return redirect(url_for('login'))
    c=db()
    control=c.execute("SELECT * FROM mastery_lab_feature_controls WHERE feature_code='mastery_laboratory'").fetchone()
    batches=c.execute("""SELECT b.*,
      (SELECT COUNT(*) FROM mastery_lab_runs r WHERE r.batch_id=b.id) run_count,
      (SELECT COUNT(*) FROM mastery_lab_blockers x WHERE x.batch_id=b.id AND x.status='OPEN') open_blockers
      FROM mastery_lab_batches b ORDER BY b.id DESC LIMIT 30""").fetchall()
    profiles=c.execute("SELECT * FROM mastery_lab_synthetic_profiles WHERE active=1 ORDER BY profile_code").fetchall()
    family_support=sorted(mastery_lab.SUPPORTED_FAMILIES)
    state_support=mastery_lab.MASTERY_STATES
    blocker_counts=c.execute("SELECT severity,status,COUNT(*) n FROM mastery_lab_blockers GROUP BY severity,status").fetchall()
    c.close()
    return render_template('admin_mastery_lab.html',control=control,batches=batches,profiles=profiles,
                           family_support=family_support,state_support=state_support,blocker_counts=blocker_counts,
                           release_flags=mastery_lab.LAB_RELEASE_FLAGS)


@app.route('/admin/mastery-lab/sample')
def admin_mastery_lab_sample():
    if not require('admin'): return redirect(url_for('login'))
    payload=mastery_lab.sample_candidate_corpus()
    raw=json.dumps(payload,ensure_ascii=False,indent=2).encode('utf-8')
    return send_file(io.BytesIO(raw),as_attachment=True,download_name='scoremax_mastery_lab_technical_sample_v6_2_6.json',mimetype='application/json')


@app.route('/admin/mastery-lab/import',methods=['POST'])
def admin_mastery_lab_import():
    if not require('admin'): return redirect(url_for('login'))
    action=request.form.get('action','upload')
    try:
        if action=='sample':
            payload=mastery_lab.sample_candidate_corpus()
            rows=payload['questions']; file_type='json'; filename='technical_sample_v6_2_6.json'
        else:
            uploaded=request.files.get('candidate_file')
            if not uploaded or not uploaded.filename:
                raise ValueError('Choose a JSON, CSV or XLSX candidate corpus.')
            raw=uploaded.read()
            if not raw:
                raise ValueError('The selected candidate file is empty.')
            rows,file_type=mastery_lab.parse_candidate_payload(raw,uploaded.filename)
            filename=uploaded.filename
        if len(rows)>10000:
            raise ValueError('A Mastery Laboratory batch is limited to 10,000 candidate rows.')
        c=db()
        result=mastery_lab.import_candidate_batch(c,rows,filename=filename,file_type=file_type,
            imported_by=session.get('user_id'),source_reference=request.form.get('source_reference','').strip())
        c.close()
        if not result.get('ok'):
            for message in result.get('errors',[])[:8]: flash(message,'error')
            if len(result.get('errors',[]))>8: flash(f"{len(result['errors'])-8} additional validation error(s) were retained in the report.",'error')
            return redirect(url_for('admin_mastery_lab'))
        flash(f"{result['batch_code']} imported {result['imported_count']} QA-only candidate(s); {result['warning_count']} warning(s) remain visible.",'success')
        return redirect(url_for('admin_mastery_lab_batch',batch_id=result['batch_id']))
    except (ValueError,json.JSONDecodeError) as exc:
        flash(str(exc),'error')
        return redirect(url_for('admin_mastery_lab'))


@app.route('/admin/mastery-lab/batch/<int:batch_id>')
def admin_mastery_lab_batch(batch_id):
    if not require('admin'): return redirect(url_for('login'))
    c=db(); summary=mastery_lab.batch_summary(c,batch_id)
    if not summary:
        c.close(); abort(404)
    profiles=c.execute("SELECT * FROM mastery_lab_synthetic_profiles WHERE active=1 ORDER BY profile_code").fetchall()
    validation=mastery_lab.safe_json(summary['batch']['validation_report_json'],{})
    release_flags=mastery_lab.safe_json(summary['batch']['release_flags_json'],{})
    c.close()
    return render_template('admin_mastery_lab_batch.html',summary=summary,profiles=profiles,
                           validation=validation,release_flags=release_flags)


@app.route('/admin/mastery-lab/batch/<int:batch_id>/simulate',methods=['POST'])
def admin_mastery_lab_simulate(batch_id):
    if not require('admin'): return redirect(url_for('login'))
    profile=request.form.get('profile_code','').strip().upper()
    c=db()
    try:
        if profile=='ALL':
            results=[]
            for code in mastery_lab.SYNTHETIC_PROFILES:
                results.append(mastery_lab.simulate_profile(c,batch_id,code,created_by=session.get('user_id')))
            latest=results[-1]['run_id'] if results else None
            mastery_lab.evaluate_all_gates(c,batch_id,latest)
            flash(f"Replayed all {len(results)} synthetic learner histories.",'success')
        else:
            result=mastery_lab.simulate_profile(c,batch_id,profile,created_by=session.get('user_id'))
            mastery_lab.evaluate_all_gates(c,batch_id,result['run_id'])
            flash(f"{profile.replace('_',' ').title()} completed in {result['final_state']}.",'success')
        c.close()
    except ValueError as exc:
        c.close(); flash(str(exc),'error')
    return redirect(url_for('admin_mastery_lab_batch',batch_id=batch_id))


@app.route('/admin/mastery-lab/batch/<int:batch_id>/evaluate',methods=['POST'])
def admin_mastery_lab_evaluate(batch_id):
    if not require('admin'): return redirect(url_for('login'))
    c=db(); latest=c.execute("SELECT id FROM mastery_lab_runs WHERE batch_id=? AND status='COMPLETED' ORDER BY id DESC LIMIT 1",(batch_id,)).fetchone()
    mastery_lab.evaluate_all_gates(c,batch_id,latest['id'] if latest else None); c.close()
    flash('All available Mastery Laboratory assurance gates were re-evaluated.','success')
    return redirect(url_for('admin_mastery_lab_batch',batch_id=batch_id))


@app.route('/admin/mastery-lab/run/<int:run_id>')
def admin_mastery_lab_run(run_id):
    if not require('admin'): return redirect(url_for('login'))
    c=db(); summary=mastery_lab.run_summary(c,run_id)
    if not summary:
        c.close(); abort(404)
    metrics=mastery_lab.safe_json(summary['run']['metrics_json'],{})
    rationale=mastery_lab.safe_json(summary['run']['rationale_json'],{})
    c.close()
    return render_template('admin_mastery_lab_run.html',summary=summary,metrics=metrics,rationale=rationale)


@app.route('/admin/mastery-lab/run/<int:run_id>/export')
def admin_mastery_lab_run_export(run_id):
    if not require('admin'): return redirect(url_for('login'))
    c=db(); summary=mastery_lab.run_summary(c,run_id)
    if not summary:
        c.close(); abort(404)
    def serialise(row):
        return dict(row) if hasattr(row,'keys') else row
    payload={
      'contract':'ScoreMax Mastery Laboratory V6.2.6',
      'release_flags':mastery_lab.LAB_RELEASE_FLAGS,
      'run':serialise(summary['run']),
      'state_history':[serialise(x) for x in summary['states']],
      'evidence':[serialise(x) for x in summary['evidence']],
      'recovery_needs':[serialise(x) for x in summary['recovery']],
      'responses':[serialise(x) for x in summary['responses']]}
    c.close(); raw=json.dumps(payload,ensure_ascii=False,indent=2).encode('utf-8')
    return send_file(io.BytesIO(raw),as_attachment=True,download_name=f"mastery_lab_run_{run_id}.json",mimetype='application/json')


@app.route('/admin/mastery-lab/blocker/<int:blocker_id>',methods=['POST'])
def admin_mastery_lab_blocker(blocker_id):
    if not require('admin'): return redirect(url_for('login'))
    decision=request.form.get('decision','RESOLVED').upper()
    if decision not in {'RESOLVED','OPEN','ACCEPTED_RISK'}: decision='OPEN'
    resolution=request.form.get('resolution','').strip()
    c=db(); blocker=c.execute("SELECT * FROM mastery_lab_blockers WHERE id=?",(blocker_id,)).fetchone()
    if not blocker:
        c.close(); abort(404)
    c.execute("UPDATE mastery_lab_blockers SET status=?,resolution=?,resolved_at=? WHERE id=?",
      (decision,resolution,datetime.now().isoformat(timespec='seconds') if decision!='OPEN' else '',blocker_id))
    c.execute("INSERT INTO mastery_lab_audit_events(actor_user_id,event_type,subject_type,subject_id,metadata_json) VALUES(?,?,?,?,?)",
      (session.get('user_id'),'BLOCKER_STATUS_CHANGED','mastery_lab_blocker',str(blocker_id),json.dumps({'status':decision,'resolution':resolution})))
    c.commit(); c.close(); flash('Mastery Laboratory blocker status updated.','success')
    return redirect(request.referrer or url_for('admin_mastery_lab'))



@app.route('/teacher-of-the-year')
def teacher_of_year_page():
    return render_template('teacher_of_year.html')


@app.route('/admin/reviewer-workspace',methods=['GET','POST'])
def admin_reviewer_workspace():
    if not require('admin'): return redirect(url_for('login'))
    generated_invites=[]; pending_import=None
    c=db()
    if request.method=='POST':
        action=request.form.get('action','').strip()
        try:
            if action=='create_reviewer':
                name=request.form.get('full_name','').strip(); email=request.form.get('email','').strip().lower()
                if not name or not email: raise ValueError('Reviewer name and email are required.')
                existing=c.execute("SELECT * FROM users WHERE lower(COALESCE(email,''))=?",(email,)).fetchone()
                if existing and existing['role']!='reviewer': raise ValueError('That email belongs to another ScoreMax role.')
                if existing: reviewer_id=existing['id']
                else:
                    cur=c.execute("""INSERT INTO users(system_user_id,role,full_name,email,username,password_hash,account_status)
                      VALUES('','reviewer',?,?,?,?,'active')""",(name,email,'pending-reviewer-'+secrets.token_urlsafe(8).lower(),generate_password_hash(secrets.token_urlsafe(32))))
                    reviewer_id=cur.lastrowid
                    c.execute("UPDATE users SET system_user_id=?,username=? WHERE id=?",(f'REV-{reviewer_id:06d}',f'reviewer-{reviewer_id:06d}',reviewer_id))
                c.commit(); flash('Reviewer account is ready. Assign confidential batches next.','success')
            elif action=='preview_import':
                upload=request.files.get('review_file')
                if not upload or not upload.filename: raise ValueError('Choose a JSON, CSV or XLSX question-bank file.')
                rows=reviewer_workspace.parse_upload(upload.filename,upload.read())
                result=reviewer_import.preview_import(c,rows,title=request.form.get('title','').strip(),filename=upload.filename,
                    created_by=session['user_id'],chapter=request.form.get('chapter','').strip(),topic=request.form.get('topic','').strip())
                return redirect(url_for('admin_reviewer_workspace',preview=result['id']))
            elif action=='confirm_import':
                import_id=int(request.form.get('import_id') or 0)
                mapping={field:request.form.get('map_'+field,'').strip() for field in reviewer_import.CANONICAL_FIELDS}
                result=reviewer_import.confirm_import(c,import_id,mapping,actor_user_id=session['user_id'])
                flash(f"Imported {result['valid_rows']:,} questions into {result['batch_count']} review batch(es)."
                      +(f" {result['invalid_rows']} incomplete row(s) were excluded and recorded." if result['invalid_rows'] else ''),'success')
                return redirect(url_for('admin_reviewer_workspace'))
            elif action=='demo_import':
                rows=reviewer_import.demo_rows(24)
                staged=reviewer_import.preview_import(c,rows,title='Reviewer Experience Demo',filename='scoremax-reviewer-demo.csv',
                    chapter='Reviewer Demo',topic='Academic checking',created_by=session['user_id'])
                mapping=reviewer_import.suggest_mapping(reviewer_import.headers(rows))
                result=reviewer_import.confirm_import(c,staged['id'],mapping,actor_user_id=session['user_id'])
                flash(f"Reviewer demo created with {result['valid_rows']} safe sample questions. Assign it to a test reviewer to inspect the experience.",'success')
            elif action in {'assign','assign_many'}:
                batch_ids=[]
                if action=='assign':
                    batch_ids=[int(request.form.get('batch_id') or 0)]
                else:
                    batch_ids=[int(x) for x in request.form.getlist('batch_ids') if str(x).isdigit()]
                batch_ids=list(dict.fromkeys(x for x in batch_ids if x))
                if not batch_ids: raise ValueError('Choose at least one review batch.')
                reviewer_id=int(request.form.get('reviewer_user_id') or 0); due=request.form.get('due_at','').strip()
                existing=c.execute(f"SELECT batch_id FROM reviewer_assignments WHERE round_no=1 AND batch_id IN ({','.join('?' for _ in batch_ids)})",batch_ids).fetchall()
                if existing: raise ValueError('One or more selected batches already have a first-review assignment.')
                group_code='RVG-'+secrets.token_hex(5).upper(); created=[]
                try:
                    for index,batch_id in enumerate(batch_ids):
                        result=reviewer_workspace.create_assignment(c,batch_id=batch_id,reviewer_user_id=reviewer_id,
                          created_by=session['user_id'],due_at=due,assignment_group_code=group_code,issue_invitation=index==0)
                        created.append(result)
                except Exception:
                    # Compensating cleanup keeps a multi-batch assignment all-or-nothing for the Admin.
                    ids=[x['assignment_id'] for x in created]
                    if ids:
                        marks=','.join('?' for _ in ids)
                        cc=db(); cc.execute('BEGIN IMMEDIATE')
                        qids=[r['question_id'] for r in cc.execute(f'SELECT question_id FROM reviewer_assignment_items WHERE assignment_id IN ({marks})',ids).fetchall()]
                        cc.execute(f'DELETE FROM reviewer_assignment_items WHERE assignment_id IN ({marks})',ids)
                        cc.execute(f'DELETE FROM reviewer_assignments WHERE id IN ({marks})',ids)
                        for qid in qids:
                            if not cc.execute('SELECT 1 FROM reviewer_assignment_items WHERE question_id=?',(qid,)).fetchone():
                                cc.execute('DELETE FROM reviewer_question_outcomes WHERE question_id=?',(qid,))
                        cc.commit(); cc.close()
                    raise
                first=created[0]
                generated_invites=[{'link':url_for('reviewer_invite',token=first['raw_token'],_external=True),
                  'verification_code':first['verification_code'],'assignment_count':len(created),'group_code':group_code}]
                flash(f"Assigned {len(created)} batch(es). One secure activation link unlocks all of them for this reviewer.",'success')
            elif action=='reissue_invite':
                result=reviewer_workspace.reissue_invitation(c,int(request.form.get('assignment_id') or 0),session['user_id'])
                generated_invites=[{'link':url_for('reviewer_invite',token=result['raw_token'],_external=True),
                  'verification_code':result['verification_code'],'assignment_count':1,'group_code':''}]
                flash('Reviewer invitation reissued. Send the new link and code through separate channels.','success')
            elif action=='second_review':
                batch_id=int(request.form.get('batch_id') or 0); reviewer_id=int(request.form.get('reviewer_user_id') or 0)
                question_ids=[r['question_id'] for r in c.execute("""SELECT rqo.question_id FROM reviewer_question_outcomes rqo
                  JOIN reviewer_questions rq ON rq.id=rqo.question_id WHERE rq.batch_id=? AND rqo.status='SECOND_REVIEW_REQUIRED'""",(batch_id,)).fetchall()]
                if not question_ids: raise ValueError('This batch has no questions awaiting second review.')
                first=c.execute("SELECT id,reviewer_user_id FROM reviewer_assignments WHERE batch_id=? AND round_no=1 ORDER BY id LIMIT 1",(batch_id,)).fetchone()
                result=reviewer_workspace.create_assignment(c,batch_id=batch_id,reviewer_user_id=reviewer_id,created_by=session['user_id'],
                    round_no=2,parent_assignment_id=first['id'] if first else None,question_ids=question_ids,due_at=request.form.get('due_at','').strip())
                generated_invites=[{'link':url_for('reviewer_invite',token=result['raw_token'],_external=True),
                  'verification_code':result['verification_code'],'assignment_count':1,'group_code':result.get('assignment_group_code','')}]
                flash(f"Independent second-review assignment created for {len(question_ids)} questions. Send the link and code separately.",'success')
        except (ValueError,PermissionError,KeyError,json.JSONDecodeError) as exc:
            if c.in_transaction: c.rollback()
            flash(str(exc),'error')
    preview_id=int(request.args.get('preview') or 0)
    if preview_id:
        row=reviewer_import.get_preview(c,preview_id,session['user_id'])
        if row and row['status']=='PREVIEW':
            raw_preview_rows=json.loads(row['rows_json'] or '[]')
            pending_import=dict(row); pending_import['columns']=reviewer_import.headers(raw_preview_rows)
            pending_import['mapping']=json.loads(row['mapping_json'] or '{}')
            pending_import['profile']=reviewer_import.detect_profile(pending_import['columns'],raw_preview_rows)
            pending_import['sheet_counts']=reviewer_import.sheet_counts(raw_preview_rows)
            normalized,errors=reviewer_import.validate_preview(raw_preview_rows,pending_import['mapping'],chapter=row['chapter'] or '',topic=row['topic'] or '')
            pending_import['preview_rows']=normalized[:5]; pending_import['preview_valid']=len(normalized); pending_import['preview_errors']=errors[:10]
    reviewers=c.execute("SELECT id,full_name,email,system_user_id FROM users WHERE role='reviewer' AND COALESCE(account_status,'active')='active' ORDER BY full_name").fetchall()
    batches=c.execute("""SELECT rb.*,ri.import_code,ri.title import_title,
      (SELECT COUNT(*) FROM reviewer_assignments ra WHERE ra.batch_id=rb.id) assignment_count,
      (SELECT COUNT(*) FROM reviewer_question_outcomes rqo JOIN reviewer_questions rq ON rq.id=rqo.question_id WHERE rq.batch_id=rb.id AND rqo.status='SECOND_REVIEW_REQUIRED') second_review_required,
      (SELECT COUNT(*) FROM reviewer_question_outcomes rqo JOIN reviewer_questions rq ON rq.id=rqo.question_id WHERE rq.batch_id=rb.id AND rqo.status='ADJUDICATION_REQUIRED') adjudication_required
      FROM reviewer_batches rb LEFT JOIN reviewer_imports ri ON ri.id=rb.import_id ORDER BY rb.id DESC""").fetchall()
    imports=c.execute("SELECT * FROM reviewer_imports WHERE status='CONFIRMED' ORDER BY id DESC LIMIT 100").fetchall()
    assignments=c.execute("""SELECT ra.*,rb.title,rb.batch_code,u.full_name reviewer_name,u.email reviewer_email,
      (SELECT COUNT(*) FROM reviewer_assignment_items x WHERE x.assignment_id=ra.id) total_items,
      (SELECT COUNT(*) FROM reviewer_assignment_items x WHERE x.assignment_id=ra.id AND x.status='COMPLETED') completed_items,
      (SELECT COALESCE(SUM(active_seconds),0) FROM reviewer_assignment_items x WHERE x.assignment_id=ra.id) active_seconds
      FROM reviewer_assignments ra JOIN reviewer_batches rb ON rb.id=ra.batch_id JOIN users u ON u.id=ra.reviewer_user_id ORDER BY ra.id DESC""").fetchall()
    c.close()
    return render_template('admin_reviewer_workspace.html',reviewers=reviewers,batches=batches,imports=imports,assignments=assignments,
      generated_invites=generated_invites,pending_import=pending_import,reviewer_fields=reviewer_import.CANONICAL_FIELDS,
      reviewer_field_labels=reviewer_import.FIELD_LABELS)


@app.route('/admin/reviewer-workspace/template')
def reviewer_import_template():
    if not require('admin'): return redirect(url_for('login'))
    wb=Workbook(); ws=wb.active; ws.title='Questions'
    ws.append(['Question ID','Chapter','Topic','Question Type','Question / Task','Stimulus / Context','Statements / Options','Key Answer','Explanation / Marking Rubric','Mastery','Review Requirement','Reviewer 2 Required'])
    ws.append(['BIO-001','Cell Biology','Cell Transport','STANDARD_MCQ','Osmosis is the movement of:','',
      'Options:\nA. Water through a selectively permeable membrane\nB. Solute using ATP\nC. Proteins through ribosomes\nD. Gases against a gradient',
      'A','Osmosis is passive movement of water.','Foundation','STANDARD_FIRST_REVIEW','NO'])
    out=io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out,as_attachment=True,download_name='ScoreMax_reviewer_import_template.xlsx',mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/admin/reviewer-workspace/import/<int:import_id>/errors')
def reviewer_import_errors(import_id):
    if not require('admin'): return redirect(url_for('login'))
    c=db(); row=reviewer_import.get_preview(c,import_id); c.close()
    if not row: abort(404)
    errors=json.loads(row['error_rows_json'] or '[]')
    out=io.StringIO(); writer=csv.DictWriter(out,fieldnames=['sheet','row','reason'],extrasaction='ignore'); writer.writeheader(); writer.writerows(errors)
    return send_file(io.BytesIO(out.getvalue().encode('utf-8-sig')),as_attachment=True,download_name=f"{row['import_code']}_errors.csv",mimetype='text/csv')


@app.route('/admin/reviewer-workspace/assignment/<int:assignment_id>')
def admin_reviewer_assignment(assignment_id):
    if not require('admin'): return redirect(url_for('login'))
    c=db(); assignment=c.execute("""SELECT ra.*,rb.title,rb.batch_code,u.full_name reviewer_name,u.email reviewer_email
      FROM reviewer_assignments ra JOIN reviewer_batches rb ON rb.id=ra.batch_id JOIN users u ON u.id=ra.reviewer_user_id WHERE ra.id=?""",(assignment_id,)).fetchone()
    if not assignment: c.close(); abort(404)
    items=c.execute("""SELECT rai.*,rq.external_question_id,rq.chapter,rq.topic,rq.question_text,rq.mastery_level,rq.calibration_expected_decision,
      rqo.status outcome_status FROM reviewer_assignment_items rai JOIN reviewer_questions rq ON rq.id=rai.question_id
      LEFT JOIN reviewer_question_outcomes rqo ON rqo.question_id=rq.id WHERE rai.assignment_id=? ORDER BY rai.display_order""",(assignment_id,)).fetchall()
    progress=reviewer_workspace.assignment_progress(c,assignment_id); quality=reviewer_workspace.assignment_quality(c,assignment_id)
    events=c.execute("SELECT * FROM reviewer_audit_events WHERE assignment_id=? ORDER BY id DESC LIMIT 100",(assignment_id,)).fetchall(); c.close()
    return render_template('admin_reviewer_assignment.html',assignment=assignment,items=items,progress=progress,quality=quality,events=events)


@app.route('/review/invite/<token>',methods=['GET','POST'])
def reviewer_invite(token):
    hashed=reviewer_workspace.token_hash(token); c=db()
    assignment=c.execute("""SELECT ra.*,u.full_name,u.email,rb.title,rb.question_count FROM reviewer_assignments ra
      JOIN users u ON u.id=ra.reviewer_user_id JOIN reviewer_batches rb ON rb.id=ra.batch_id
      WHERE ra.invitation_token_hash=? AND ra.status='INVITED'""",(hashed,)).fetchone()
    valid=bool(assignment); account_conflict=False
    if valid:
        try: valid=datetime.fromisoformat(assignment['invitation_expires_at'])>=datetime.now() and not assignment['invitation_locked_at']
        except Exception: valid=False
    if valid and session.get('user_id') and int(session.get('user_id'))!=int(assignment['reviewer_user_id']):
        account_conflict=True; valid=False
    if request.method=='POST' and valid:
        if request.form.get('confidentiality_accept')!='yes':
            c.close(); flash('You must accept the confidentiality terms to continue.','error'); return redirect(url_for('reviewer_invite',token=token))
        password=request.form.get('password','')
        if len(password)<10:
            c.close(); flash('Use a password of at least 10 characters.','error'); return redirect(url_for('reviewer_invite',token=token))
        try:
            reviewer_workspace.accept_invitation(c,token,assignment['reviewer_user_id'],request.form.get('verification_code',''),generate_password_hash(password))
        except ValueError as exc:
            c.close(); flash(str(exc),'error'); return redirect(url_for('reviewer_invite',token=token))
        u=c.execute('SELECT * FROM users WHERE id=?',(assignment['reviewer_user_id'],)).fetchone(); c.close()
        session.clear(); session.update(user_id=u['id'],role='reviewer',full_name=u['full_name'],session_version=int(u['session_version'] or 0)); _csrf_token()
        return redirect(url_for('reviewer_assignment',assignment_id=assignment['id']))
    masked=mask_email(assignment['email']) if assignment else ''
    c.close(); return render_template('reviewer_invite.html',assignment=assignment,valid=valid,token=token,masked_email=masked,account_conflict=account_conflict)


@app.route('/review')
def reviewer_home():
    if not require('reviewer'): return redirect(url_for('login'))
    c=db(); assignments=c.execute("""SELECT ra.*,rb.title,rb.chapter,rb.topic,rb.question_count,
      (SELECT COUNT(*) FROM reviewer_assignment_items x WHERE x.assignment_id=ra.id AND x.status='COMPLETED') completed_items,
      (SELECT COALESCE(SUM(active_seconds),0) FROM reviewer_assignment_items x WHERE x.assignment_id=ra.id) active_seconds
      FROM reviewer_assignments ra JOIN reviewer_batches rb ON rb.id=ra.batch_id WHERE ra.reviewer_user_id=? ORDER BY ra.id DESC""",(session['user_id'],)).fetchall(); c.close()
    return render_template('reviewer_home.html',assignments=assignments)


@app.route('/review/continue')
def reviewer_continue():
    if not require('reviewer'): return redirect(url_for('login'))
    c=db(); row=c.execute('''SELECT ra.id FROM reviewer_assignments ra
      WHERE ra.reviewer_user_id=? AND ra.status='IN_PROGRESS'
      AND EXISTS(SELECT 1 FROM reviewer_assignment_items x WHERE x.assignment_id=ra.id AND x.status<>'COMPLETED')
      ORDER BY ra.assigned_at,ra.id LIMIT 1''',(session['user_id'],)).fetchone(); c.close()
    return redirect(url_for('reviewer_assignment',assignment_id=row['id'])) if row else redirect(url_for('reviewer_home'))


@app.route('/review/assignment/<int:assignment_id>')
def reviewer_assignment(assignment_id):
    if not require('reviewer'): return redirect(url_for('login'))
    c=db(); assignment=c.execute("SELECT * FROM reviewer_assignments WHERE id=? AND reviewer_user_id=?",(assignment_id,session['user_id'])).fetchone()
    if not assignment: c.close(); abort(404)
    next_item=reviewer_workspace.next_unfinished_item(c,assignment_id)
    if next_item:
        target=next_item['id']; c.close(); return redirect(url_for('reviewer_item',assignment_id=assignment_id,item_id=target))
    progress=reviewer_workspace.assignment_progress(c,assignment_id)
    next_assignment=c.execute('''SELECT ra.id,rb.title FROM reviewer_assignments ra JOIN reviewer_batches rb ON rb.id=ra.batch_id
      WHERE ra.reviewer_user_id=? AND ra.status='IN_PROGRESS' AND ra.id<>?
      AND EXISTS(SELECT 1 FROM reviewer_assignment_items x WHERE x.assignment_id=ra.id AND x.status<>'COMPLETED')
      ORDER BY ra.assigned_at,ra.id LIMIT 1''',(session['user_id'],assignment_id)).fetchone(); c.close()
    return render_template('reviewer_complete.html',assignment=assignment,progress=progress,next_assignment=next_assignment)


@app.route('/review/assignment/<int:assignment_id>/item/<int:item_id>')
def reviewer_item(assignment_id,item_id):
    if not require('reviewer'): return redirect(url_for('login'))
    c=db()
    try: item=reviewer_workspace.open_item(c,assignment_id,item_id,session['user_id'])
    except PermissionError: c.close(); abort(403)
    progress=reviewer_workspace.assignment_progress(c,assignment_id)
    navigator=c.execute("SELECT id,display_order,status FROM reviewer_assignment_items WHERE assignment_id=? ORDER BY display_order",(assignment_id,)).fetchall()
    c.close(); return render_template('reviewer_item.html',item=item,options=json.loads(item['options_json'] or '[]'),progress=progress,navigator=navigator,
      decisions=sorted(reviewer_workspace.DECISIONS),suitability=sorted(reviewer_workspace.MASTERY_SUITABILITY))


@app.route('/review/item/<int:item_id>/time',methods=['POST'])
def reviewer_timer(item_id):
    if not require('reviewer'): return jsonify({'error':'forbidden'}),403
    if not rate_limit(f"reviewer-time:{session['user_id']}:{item_id}",limit=20,window_seconds=60):
        return jsonify({'error':'rate_limited','accepted_seconds':0}),429
    c=db()
    try: accepted=reviewer_workspace.record_active_time(c,item_id,session['user_id'],int(request.form.get('seconds') or 0),
          {'visible':request.form.get('visible')=='1','activity':request.form.get('activity','')})
    except PermissionError: c.close(); return jsonify({'error':'forbidden'}),403
    c.close(); return jsonify({'ok':True,'accepted_seconds':accepted})


@app.route('/review/item/<int:item_id>/reveal',methods=['POST'])
def reviewer_reveal(item_id):
    if not require('reviewer'): return jsonify({'error':'forbidden'}),403
    c=db()
    try: reviewer_workspace.reveal_answer(c,item_id,session['user_id'])
    except PermissionError: c.close(); return jsonify({'error':'forbidden'}),403
    row=c.execute("""SELECT rq.correct_answer,rq.explanation FROM reviewer_assignment_items rai JOIN reviewer_questions rq ON rq.id=rai.question_id
      JOIN reviewer_assignments ra ON ra.id=rai.assignment_id WHERE rai.id=? AND ra.reviewer_user_id=?""",(item_id,session['user_id'])).fetchone(); c.close()
    return jsonify({'ok':True,'correct_answer':row['correct_answer'],'explanation':row['explanation']})


@app.route('/review/assignment/<int:assignment_id>/item/<int:item_id>/submit',methods=['POST'])
def reviewer_submit(assignment_id,item_id):
    if not require('reviewer'): return redirect(url_for('login'))
    c=db()
    try:
        result=reviewer_workspace.submit_decision(c,item_id=item_id,reviewer_user_id=session['user_id'],decision=request.form.get('decision',''),
          mastery_suitability=request.form.get('mastery_suitability',''),comments=request.form.get('comments',''),independent_answer=request.form.get('independent_answer',''))
    except (ValueError,PermissionError) as exc:
        c.close(); flash(str(exc),'error'); return redirect(url_for('reviewer_item',assignment_id=assignment_id,item_id=item_id))
    next_item=reviewer_workspace.next_unfinished_item(c,assignment_id); c.close()
    if next_item: return redirect(url_for('reviewer_item',assignment_id=assignment_id,item_id=next_item['id']))
    return redirect(url_for('reviewer_assignment',assignment_id=assignment_id))

@app.route('/admin/template')
def template():
    if not require('admin'): return redirect(url_for('login'))
    wb=Workbook(); ws=wb.active; ws.title='Chapter 1'
    headers=['Question ID','Family ID','Variant','Programme','Subject','Chapter','Topic','Sub-topic','Type','Level','Question','A','B','C','D','Answer','Explanation','Status',
             'Country','Qualification','Exam Board','Curriculum Version','Learning Outcome','Concept','Concept ID','Capsule ID','Misconception ID','Difficulty','Cognitive Skill','Command Word','Marks','Estimated Time Seconds','Misconception Tags','Prerequisite Tags','Source Type','Review Status','Secure Bank','Rights Status','ScoreMax Ready','Assessment Purpose','Difficulty Source','Family Construct','Family Invariants']
    ws.append(headers)
    ws.append(['PHY101','MOTION010','A','FSc Part 1','Physics','Measurements','Motion','Speed','MCQ','Foundation','Which quantity is speed?','Distance / time','Time / distance','Mass / time','Distance × time','A','Speed is distance travelled per unit time.','Draft',
               'Pakistan','FSc Part 1','','2026','LO-1.1','Speed','PHY-SPEED','LC-PHY-SPEED','MIS-SPEED-01','Moderate','Recall','Identify',1,60,'confuses speed with velocity|uses inverse formula','distance|time','ScoreMax Original','Draft','No','ScoreMax Original','Yes','practice|test|mock|mastery','authoring','Calculate and interpret speed from distance and time','same speed construct|same unit relationship|surface values may vary'])
    lc=wb.create_sheet('Learning Capsules')
    lc.append(['Capsule ID','Programme','Subject','Chapter','Topic','Sub-topic','Learning Outcome','Concept ID','Concept',
               'Simple Explanation','Remember This','Formula / Rule','Worked Example','Common Mistake','Memory Tip',
               'Visual Brief','Micro-video URL','Quick Check Question IDs','Recovery Question IDs','Review Status','Reviewer','Version'])
    lc.append(['LC-PHY-SPEED','FSc Part 1','Physics','Measurements','Motion','Speed','LO-1.1','PHY-SPEED','Speed',
               'Speed tells you how much distance is travelled in a given time.','Speed = distance divided by time.',
               'speed = distance / time','A car travels 120 km in 2 hours, so its average speed is 60 km/h.',
               'Do not divide time by distance.','Distance over time.','Simple distance-time visual','','','','Draft','',1])
    mc=wb.create_sheet('Misconceptions')
    mc.append(['Misconception ID','Concept ID','Subject','Title','Student Explanation','Corrective Hint','Active'])
    mc.append(['MIS-SPEED-01','PHY-SPEED','Physics','Uses inverse speed formula',
               'The student divides time by distance instead of distance by time.',
               'Put distance on top and time underneath: speed = distance ÷ time.','Yes'])
    p=BASE/'question_import_template_v5_5.xlsx'; wb.save(p); return send_file(p,as_attachment=True)


if __name__=='__main__':
    init()
    app.run(host=os.environ.get('SCOREMAX_HOST','127.0.0.1'),port=int(os.environ.get('SCOREMAX_PORT','5000')),debug=False)
