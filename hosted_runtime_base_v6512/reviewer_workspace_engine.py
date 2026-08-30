"""ScoreMax V6.2.8.1 confidential Academic Reviewer Workspace.

The reviewer workspace stores minimal question snapshots only. It deliberately exposes no
live-bank, mastery-engine, curriculum-graph, student, analytics or product-architecture data.
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
import re
import secrets
import sqlite3
from datetime import datetime, timedelta
from typing import Any, Mapping, Sequence

MAX_ASSIGNMENT_ITEMS = 100
DECISIONS = {
    'ACCEPT_UNCHANGED',
    'CORRECTION_REQUIRED',
    'MASTERY_LEVEL_UNSUITABLE',
    'REJECT',
    'UNSURE',
}
MASTERY_SUITABILITY = {'SUITABLE', 'UNSUITABLE', 'UNSURE'}
REVIEWER_ROLE = 'reviewer'
INVITATION_CODE_ATTEMPT_LIMIT = 8
MIN_TIMER_INTERVAL_SECONDS = 4
MAX_TIMER_TICK_SECONDS = 30


def verification_code() -> str:
    alphabet='ABCDEFGHJKLMNPQRSTUVWXYZ23456789'
    return '-'.join(''.join(secrets.choice(alphabet) for _ in range(4)) for _ in range(2))


def _columns(c, table: str) -> set[str]:
    return {str(r['name'] if hasattr(r,'keys') else r[1]) for r in c.execute(f'PRAGMA table_info({table})').fetchall()}


def _meaningful_comment(value: str) -> bool:
    text=_text(value)
    words=re.findall(r"[A-Za-z0-9]+", text)
    letters=re.findall(r"[A-Za-z]", text)
    return len(text)>=12 and len(words)>=2 and len(letters)>=6 and len(set(text.casefold()))>=4


def canonical_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(',', ':'))


def checksum(value: Any) -> str:
    return hashlib.sha256(canonical_json(value).encode('utf-8')).hexdigest()


def token_hash(raw: str) -> str:
    return hashlib.sha256((raw or '').encode('utf-8')).hexdigest()


def _text(value: Any) -> str:
    return '' if value is None else str(value).strip()


def _list(value: Any) -> list:
    if value is None or value == '':
        return []
    if isinstance(value, list):
        return value
    try:
        parsed=json.loads(str(value))
        if isinstance(parsed,list): return parsed
    except Exception:
        pass
    return [x.strip() for x in re.split(r'[|;]',str(value)) if x.strip()]


def init_reviewer_schema(c) -> None:
    c.executescript('''
    CREATE TABLE IF NOT EXISTS reviewer_feature_controls(
      feature_code TEXT PRIMARY KEY,state TEXT NOT NULL DEFAULT 'QA_ONLY',configuration_json TEXT DEFAULT '{}',
      updated_by INTEGER,updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS reviewer_batches(
      id INTEGER PRIMARY KEY,batch_code TEXT UNIQUE,title TEXT NOT NULL,chapter TEXT,topic TEXT,
      source_filename TEXT,source_checksum TEXT NOT NULL,status TEXT NOT NULL DEFAULT 'READY',
      question_count INTEGER NOT NULL DEFAULT 0,created_by INTEGER,created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS reviewer_questions(
      id INTEGER PRIMARY KEY,batch_id INTEGER NOT NULL,external_question_id TEXT NOT NULL,display_order INTEGER NOT NULL,
      chapter TEXT,topic TEXT,question_text TEXT NOT NULL,options_json TEXT DEFAULT '[]',correct_answer TEXT NOT NULL,
      explanation TEXT NOT NULL,mastery_level TEXT NOT NULL,calibration_expected_decision TEXT DEFAULT '',
      snapshot_checksum TEXT NOT NULL,created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
      UNIQUE(batch_id,external_question_id),UNIQUE(batch_id,display_order));
    CREATE TABLE IF NOT EXISTS reviewer_assignments(
      id INTEGER PRIMARY KEY,assignment_code TEXT UNIQUE,batch_id INTEGER NOT NULL,reviewer_user_id INTEGER NOT NULL,
      round_no INTEGER NOT NULL DEFAULT 1,parent_assignment_id INTEGER,status TEXT NOT NULL DEFAULT 'INVITED',
      assigned_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,due_at TEXT,started_at TEXT,completed_at TEXT,last_question_id INTEGER,
      invitation_token_hash TEXT UNIQUE,invitation_verification_hash TEXT,invitation_verification_attempts INTEGER NOT NULL DEFAULT 0,
      invitation_locked_at TEXT,invitation_expires_at TEXT,confidentiality_accepted_at TEXT,created_by INTEGER,
      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS reviewer_assignment_items(
      id INTEGER PRIMARY KEY,assignment_id INTEGER NOT NULL,question_id INTEGER NOT NULL,display_order INTEGER NOT NULL,
      round_no INTEGER NOT NULL DEFAULT 1,status TEXT NOT NULL DEFAULT 'NOT_STARTED',independent_answer TEXT DEFAULT '',answer_revealed_at TEXT,
      decision TEXT DEFAULT '',mastery_suitability TEXT DEFAULT '',comments TEXT DEFAULT '',
      first_opened_at TEXT,last_opened_at TEXT,submitted_at TEXT,active_seconds INTEGER NOT NULL DEFAULT 0,
      open_count INTEGER NOT NULL DEFAULT 0,return_count INTEGER NOT NULL DEFAULT 0,edit_count INTEGER NOT NULL DEFAULT 0,
      last_ping_at TEXT,previous_decision TEXT DEFAULT '',risk_flags_json TEXT DEFAULT '[]',
      UNIQUE(assignment_id,question_id),UNIQUE(assignment_id,display_order));
    CREATE TABLE IF NOT EXISTS reviewer_time_events(
      id INTEGER PRIMARY KEY,assignment_item_id INTEGER NOT NULL,event_type TEXT NOT NULL,seconds INTEGER NOT NULL DEFAULT 0,
      metadata_json TEXT DEFAULT '{}',created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS reviewer_question_outcomes(
      id INTEGER PRIMARY KEY,question_id INTEGER UNIQUE,first_assignment_item_id INTEGER,second_assignment_item_id INTEGER,
      status TEXT NOT NULL DEFAULT 'FIRST_REVIEW_PENDING',first_decision TEXT DEFAULT '',second_decision TEXT DEFAULT '',
      adjudication_status TEXT DEFAULT '',updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS reviewer_audit_events(
      id INTEGER PRIMARY KEY,actor_user_id INTEGER,event_type TEXT NOT NULL,assignment_id INTEGER,assignment_item_id INTEGER,
      metadata_json TEXT DEFAULT '{}',created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    CREATE INDEX IF NOT EXISTS idx_reviewer_items_assignment ON reviewer_assignment_items(assignment_id,display_order);
    CREATE INDEX IF NOT EXISTS idx_reviewer_assignments_user ON reviewer_assignments(reviewer_user_id,status);
    CREATE INDEX IF NOT EXISTS idx_reviewer_time_item ON reviewer_time_events(assignment_item_id,created_at);
    ''')
    assignment_columns=_columns(c,'reviewer_assignments')
    additions={
      'invitation_verification_hash':'TEXT',
      'invitation_verification_attempts':'INTEGER NOT NULL DEFAULT 0',
      'invitation_locked_at':'TEXT',
    }
    for name,definition in additions.items():
        if name not in assignment_columns:
            c.execute(f'ALTER TABLE reviewer_assignments ADD COLUMN {name} {definition}')
    question_columns=_columns(c,'reviewer_questions')
    for name,definition in {
        'stimulus_context':"TEXT DEFAULT ''",
        'review_content':"TEXT DEFAULT ''",
        'question_type':"TEXT DEFAULT ''",
        'review_priority':"TEXT DEFAULT ''",
        'review_requirement':"TEXT DEFAULT ''",
        'reviewer2_required':'INTEGER NOT NULL DEFAULT 0',
        'source_sheet':"TEXT DEFAULT ''",
        'source_row':'INTEGER'
    }.items():
        if name not in question_columns:
            c.execute(f'ALTER TABLE reviewer_questions ADD COLUMN {name} {definition}')
    if 'round_no' not in _columns(c,'reviewer_assignment_items'):
        c.execute('ALTER TABLE reviewer_assignment_items ADD COLUMN round_no INTEGER NOT NULL DEFAULT 1')
    c.execute('''UPDATE reviewer_assignment_items SET round_no=COALESCE((
      SELECT ra.round_no FROM reviewer_assignments ra WHERE ra.id=reviewer_assignment_items.assignment_id),1)''')
    # Pre-patch unused invitations have no separately deliverable verification code. Invalidate them safely.
    legacy_invites=c.execute("""SELECT id FROM reviewer_assignments WHERE status='INVITED'
      AND COALESCE(invitation_verification_hash,'')=''""").fetchall()
    for row in legacy_invites:
        c.execute("UPDATE reviewer_assignments SET status='INVITATION_REISSUE_REQUIRED',invitation_token_hash=NULL WHERE id=?",(row['id'],))
        c.execute("INSERT INTO reviewer_audit_events(event_type,assignment_id,metadata_json) VALUES(?,?,?)",
                  ('REVIEW_INVITATION_REISSUE_REQUIRED',row['id'],json.dumps({'reason':'V6.2.7.1 two-part verification migration'})))
    # Hard database backstops. Refuse to conceal pre-existing invariant violations.
    duplicate=c.execute('SELECT source_checksum,COUNT(*) n FROM reviewer_batches GROUP BY source_checksum HAVING COUNT(*)>1 LIMIT 1').fetchone()
    if duplicate:
        raise RuntimeError('Reviewer migration blocked: duplicate review-batch checksums must be resolved before V6.2.7.1.')
    duplicate=c.execute('SELECT batch_id,COUNT(*) n FROM reviewer_assignments WHERE round_no=1 GROUP BY batch_id HAVING COUNT(*)>1 LIMIT 1').fetchone()
    if duplicate:
        raise RuntimeError('Reviewer migration blocked: more than one first-review assignment exists for a batch.')
    duplicate=c.execute('SELECT question_id,COUNT(*) n FROM reviewer_assignment_items WHERE round_no=2 GROUP BY question_id HAVING COUNT(*)>1 LIMIT 1').fetchone()
    if duplicate:
        raise RuntimeError('Reviewer migration blocked: a question has overlapping second-review assignments.')
    c.executescript('''
      CREATE UNIQUE INDEX IF NOT EXISTS uq_reviewer_batch_checksum ON reviewer_batches(source_checksum);
      CREATE UNIQUE INDEX IF NOT EXISTS uq_reviewer_first_assignment ON reviewer_assignments(batch_id) WHERE round_no=1;
      CREATE UNIQUE INDEX IF NOT EXISTS uq_reviewer_second_question ON reviewer_assignment_items(question_id) WHERE round_no=2;
    ''')
    c.execute('''INSERT OR IGNORE INTO reviewer_feature_controls(feature_code,state,configuration_json)
      VALUES('academic_reviewer_workspace','QA_ONLY',?)''',(json.dumps({
        'max_assignment_items':MAX_ASSIGNMENT_ITEMS,'idle_timeout_seconds':120,'fast_item_seconds':10,
        'timer_server_reconciliation':True,'invitation_two_part_verification':True,
        'confidentiality_boundary':'MINIMAL_REVIEW_SNAPSHOT_ONLY'}),))
    c.execute('''UPDATE reviewer_feature_controls SET configuration_json=?,updated_at=CURRENT_TIMESTAMP
      WHERE feature_code='academic_reviewer_workspace' ''',(json.dumps({
        'max_assignment_items':MAX_ASSIGNMENT_ITEMS,'idle_timeout_seconds':120,'fast_item_seconds':10,
        'timer_server_reconciliation':True,'invitation_two_part_verification':True,
        'confidentiality_boundary':'MINIMAL_REVIEW_SNAPSHOT_ONLY'}),))


def _norm_spreadsheet_header(value: Any) -> str:
    text=_text(value).casefold().replace('_',' ').replace('-',' ')
    text=re.sub(r'\s*/\s*',' / ',text)
    return re.sub(r'\s+',' ',text).strip()


_QUESTION_SHEET_ALIASES={
    'question','question text','stem','prompt','item','content','question / task','question task'
}
_ANSWER_SHEET_ALIASES={
    'answer','correct answer','answer key','key','correct option','key answer','explanation / marking rubric','marking rubric','rubric'
}


def _find_question_header_row(ws, scan_rows: int=15):
    best=None
    for row_number,row in enumerate(ws.iter_rows(min_row=1,max_row=min(int(ws.max_row or 1),scan_rows),values_only=True),1):
        headers=[_text(x) for x in row]
        normalized={_norm_spreadsheet_header(x) for x in headers if _text(x)}
        q=normalized & _QUESTION_SHEET_ALIASES
        a=normalized & _ANSWER_SHEET_ALIASES
        if q and a:
            score=len(q)*10+len(a)*10+len(normalized)
            candidate=(score,row_number,headers)
            if best is None or candidate[0]>best[0]:
                best=candidate
    return best[1:] if best else None


def _worksheet_rows(ws, header_row: int, headers: Sequence[str], sheet_order: int) -> list[dict]:
    clean=[]
    used={}
    for index,value in enumerate(headers):
        name=_text(value) or f'Unnamed column {index+1}'
        count=used.get(name,0)+1
        used[name]=count
        clean.append(name if count==1 else f'{name} ({count})')
    rows=[]
    for source_row,row in enumerate(ws.iter_rows(min_row=header_row+1,values_only=True),header_row+1):
        if not any(x not in (None,'') for x in row):
            continue
        record={clean[i]:row[i] for i in range(min(len(clean),len(row)))}
        record['__source_sheet']=ws.title
        record['__source_row']=source_row
        record['__source_sheet_order']=sheet_order
        rows.append(record)
    return rows


def parse_upload(filename: str, raw: bytes) -> list[dict]:
    ext=(filename.rsplit('.',1)[-1] if '.' in filename else '').lower()
    if ext=='json':
        payload=json.loads(raw.decode('utf-8-sig'))
        rows=payload.get('questions',[]) if isinstance(payload,dict) else payload
        if not isinstance(rows,list): raise ValueError('JSON must contain a question list.')
        return [dict(x) for x in rows]
    if ext=='csv':
        return [dict(x) for x in csv.DictReader(io.StringIO(raw.decode('utf-8-sig')))]
    if ext in {'xlsx','xlsm'}:
        from openpyxl import load_workbook
        wb=load_workbook(io.BytesIO(raw),read_only=True,data_only=True)
        detected=[]
        try:
            for sheet_order,ws in enumerate(wb.worksheets,1):
                header=_find_question_header_row(ws)
                if not header:
                    continue
                header_row,headers=header
                rows=_worksheet_rows(ws,header_row,headers,sheet_order)
                if rows:
                    detected.extend(rows)
            if detected:
                return detected
            # Backward-compatible manual-mapping fallback for an unfamiliar one-sheet workbook.
            ws=wb.active
            values=list(ws.iter_rows(values_only=True))
            if not values:
                return []
            headers=[_text(x) for x in values[0]]
            return [
                {**{headers[i]:row[i] for i in range(min(len(headers),len(row)))},
                 '__source_sheet':ws.title,'__source_row':index,'__source_sheet_order':1}
                for index,row in enumerate(values[1:],2) if any(x not in (None,'') for x in row)
            ]
        finally:
            wb.close()
    raise ValueError('Use JSON, CSV or XLSX.')


def normalize_question(row: Mapping[str,Any], index: int) -> dict:
    def pick(*names,default=''):
        for name in names:
            if name in row and row.get(name) not in (None,''): return row.get(name)
        return default
    options=pick('options','Options',default=[])
    if not options:
        options=[]
        for key in ('A','B','C','D','E','F'):
            val=pick(f'option_{key.lower()}',f'Option {key}',key,default='')
            if _text(val): options.append({'id':key,'text':_text(val)})
    else:
        options=_list(options)
        normalized=[]
        for i,opt in enumerate(options):
            if isinstance(opt,dict): normalized.append({'id':_text(opt.get('id') or chr(65+i)),'text':_text(opt.get('text') or opt.get('label'))})
            else: normalized.append({'id':chr(65+i),'text':_text(opt)})
        options=normalized
    q={
      'external_question_id':_text(pick('question_id','Question ID','id',default=f'REVIEW-{index:04d}')),
      'chapter':_text(pick('chapter','Chapter')),
      'topic':_text(pick('topic','Topic','subtopic','Subtopic')),
      'question_text':_text(pick('question','Question','question_text','Question Text')),
      'options':options,
      'correct_answer':_text(pick('answer','Answer','correct_answer','Correct Answer')),
      'explanation':_text(pick('explanation','Explanation')),
      'mastery_level':_text(pick('mastery_level','Mastery Level','level','Level')),
      'calibration_expected_decision':_text(pick('calibration_expected_decision','Calibration Expected Decision')).upper(),
      'display_order':index,
    }
    missing=[k for k in ('external_question_id','question_text','correct_answer','explanation','mastery_level') if not q[k]]
    if missing: raise ValueError(f"Row {index}: missing {', '.join(missing)}")
    if q['calibration_expected_decision'] and q['calibration_expected_decision'] not in DECISIONS:
        raise ValueError(f'Row {index}: invalid calibration expected decision.')
    return q


def import_batch(c, rows: Sequence[Mapping[str,Any]], *, title: str, filename: str, created_by: int,
                 chapter: str='', topic: str='') -> dict:
    if not rows: raise ValueError('The review batch is empty.')
    if len(rows)>MAX_ASSIGNMENT_ITEMS: raise ValueError(f'A reviewer batch may contain at most {MAX_ASSIGNMENT_ITEMS} questions.')
    normalized=[normalize_question(r,i+1) for i,r in enumerate(rows)]
    ids=[x['external_question_id'].casefold() for x in normalized]
    if len(ids)!=len(set(ids)): raise ValueError('Duplicate question IDs exist inside the review batch.')
    payload={'title':title,'chapter':chapter,'topic':topic,'questions':normalized}
    digest=checksum(payload); batch_code='RVB-'+secrets.token_hex(5).upper()
    try:
        c.execute('BEGIN IMMEDIATE')
        if c.execute('SELECT id FROM reviewer_batches WHERE source_checksum=?',(digest,)).fetchone():
            raise ValueError('This exact review batch has already been imported.')
        cur=c.execute('''INSERT INTO reviewer_batches(batch_code,title,chapter,topic,source_filename,source_checksum,status,question_count,created_by)
          VALUES(?,?,?,?,?,?,'READY',?,?)''',(batch_code,_text(title) or 'Academic Review Batch',_text(chapter),_text(topic),_text(filename),digest,len(normalized),created_by))
        batch_id=cur.lastrowid
        for q in normalized:
            snapshot=checksum({k:v for k,v in q.items() if k!='display_order'})
            c.execute('''INSERT INTO reviewer_questions(batch_id,external_question_id,display_order,chapter,topic,question_text,options_json,
              correct_answer,explanation,mastery_level,calibration_expected_decision,snapshot_checksum)
              VALUES(?,?,?,?,?,?,?,?,?,?,?,?)''',(batch_id,q['external_question_id'],q['display_order'],q['chapter'] or chapter,q['topic'] or topic,
              q['question_text'],json.dumps(q['options'],ensure_ascii=False),q['correct_answer'],q['explanation'],q['mastery_level'],q['calibration_expected_decision'],snapshot))
        c.execute('INSERT INTO reviewer_audit_events(actor_user_id,event_type,metadata_json) VALUES(?,?,?)',
                  (created_by,'REVIEW_BATCH_IMPORTED',json.dumps({'batch_id':batch_id,'count':len(normalized),'checksum':digest})))
        c.commit(); return {'batch_id':batch_id,'batch_code':batch_code,'count':len(normalized),'checksum':digest}
    except sqlite3.IntegrityError as exc:
        c.rollback()
        if 'source_checksum' in str(exc) or 'uq_reviewer_batch_checksum' in str(exc):
            raise ValueError('This exact review batch has already been imported.') from exc
        raise
    except Exception:
        c.rollback(); raise


def create_assignment(c, *, batch_id: int, reviewer_user_id: int, created_by: int, due_at: str='',
                      round_no: int=1, parent_assignment_id: int|None=None, question_ids: Sequence[int]|None=None,
                      assignment_group_code: str='', issue_invitation: bool=True) -> dict:
    round_no=int(round_no or 0)
    if round_no not in (1,2): raise ValueError('Review round must be 1 or 2.')
    raw=secrets.token_urlsafe(32) if issue_invitation else ''; hashed=token_hash(raw) if raw else None
    raw_code=verification_code() if issue_invitation else ''; code_hash=token_hash(raw_code.replace('-','').upper()) if raw_code else None
    assignment_code='RVA-'+secrets.token_hex(5).upper(); expiry=(datetime.now()+timedelta(days=7)).isoformat(timespec='seconds') if issue_invitation else None
    group_code=_text(assignment_group_code) or ('RVG-'+secrets.token_hex(5).upper())
    try:
        c.execute('BEGIN IMMEDIATE')
        batch=c.execute('SELECT * FROM reviewer_batches WHERE id=?',(batch_id,)).fetchone()
        if not batch: raise ValueError('Review batch not found.')
        reviewer=c.execute("SELECT * FROM users WHERE id=? AND role='reviewer' AND COALESCE(account_status,'active')='active'",(reviewer_user_id,)).fetchone()
        if not reviewer: raise ValueError('Active reviewer account not found.')
        parent=None
        if round_no==1:
            if parent_assignment_id: raise ValueError('A first review cannot have a parent assignment.')
            if c.execute('SELECT id FROM reviewer_assignments WHERE batch_id=? AND round_no=1',(batch_id,)).fetchone():
                raise ValueError('The first review has already been assigned for this batch.')
        else:
            if not parent_assignment_id: raise ValueError('A second review requires the original first-review assignment.')
            parent=c.execute('SELECT id,batch_id,reviewer_user_id,round_no FROM reviewer_assignments WHERE id=?',(parent_assignment_id,)).fetchone()
            if not parent or int(parent['round_no'])!=1 or int(parent['batch_id'])!=int(batch_id):
                raise ValueError('The second review must reference this batch’s valid first-review assignment.')
            if int(parent['reviewer_user_id'])==int(reviewer_user_id):
                raise ValueError('The second reviewer must be independent from the first reviewer.')
            if not question_ids: raise ValueError('A second review must contain questions requiring independent review.')
        if question_ids:
            clean=list(dict.fromkeys(int(x) for x in question_ids))
            marks=','.join('?' for _ in clean)
            qrows=c.execute(f'SELECT id,display_order FROM reviewer_questions WHERE batch_id=? AND id IN ({marks}) ORDER BY display_order',[batch_id]+clean).fetchall()
            if len(qrows)!=len(clean): raise ValueError('One or more selected review questions do not belong to this batch.')
        else:
            qrows=c.execute('SELECT id,display_order FROM reviewer_questions WHERE batch_id=? ORDER BY display_order',(batch_id,)).fetchall()
        if not qrows or len(qrows)>MAX_ASSIGNMENT_ITEMS: raise ValueError('Assignment must contain 1 to 100 questions.')
        if round_no==2:
            qids=[int(r['id']) for r in qrows]; marks=','.join('?' for _ in qids)
            eligible=c.execute(f"SELECT question_id FROM reviewer_question_outcomes WHERE status='SECOND_REVIEW_REQUIRED' AND question_id IN ({marks})",qids).fetchall()
            if {int(r['question_id']) for r in eligible}!=set(qids):
                raise ValueError('Every second-review question must currently require independent review.')
            overlap=c.execute(f'''SELECT 1 FROM reviewer_assignment_items WHERE round_no=2 AND question_id IN ({marks}) LIMIT 1''',qids).fetchone()
            if overlap: raise ValueError('One or more questions already have an independent second-review assignment.')
        assignment_status='INVITED' if issue_invitation else 'PENDING_ACTIVATION'
        cur=c.execute('''INSERT INTO reviewer_assignments(assignment_code,batch_id,reviewer_user_id,round_no,parent_assignment_id,status,due_at,
          invitation_token_hash,invitation_verification_hash,invitation_expires_at,created_by,assignment_group_code)
          VALUES(?,?,?,?,?,?,?,?,?,?,?,?)''',
          (assignment_code,batch_id,reviewer_user_id,round_no,parent_assignment_id,assignment_status,_text(due_at),hashed,code_hash,expiry,created_by,group_code))
        assignment_id=cur.lastrowid
        for row in qrows:
            c.execute('INSERT INTO reviewer_assignment_items(assignment_id,question_id,display_order,round_no) VALUES(?,?,?,?)',
                      (assignment_id,row['id'],row['display_order'],round_no))
            c.execute("INSERT OR IGNORE INTO reviewer_question_outcomes(question_id,status) VALUES(?,'FIRST_REVIEW_PENDING')",(row['id'],))
        c.execute('INSERT INTO reviewer_audit_events(actor_user_id,event_type,assignment_id,metadata_json) VALUES(?,?,?,?)',
                  (created_by,'REVIEW_ASSIGNMENT_CREATED',assignment_id,json.dumps({'round_no':round_no,'question_count':len(qrows),'two_part_invitation':bool(issue_invitation),'assignment_group_code':group_code})))
        c.commit(); return {'assignment_id':assignment_id,'assignment_code':assignment_code,'raw_token':raw,
                           'verification_code':raw_code,'expires_at':expiry,'assignment_group_code':group_code,'invitation_issued':bool(issue_invitation)}
    except sqlite3.IntegrityError as exc:
        c.rollback(); message=str(exc)
        if 'uq_reviewer_first_assignment' in message or 'reviewer_assignments.batch_id' in message:
            raise ValueError('The first review has already been assigned for this batch.') from exc
        if 'uq_reviewer_second_question' in message or 'reviewer_assignment_items.question_id' in message:
            raise ValueError('One or more questions already have an independent second-review assignment.') from exc
        raise
    except Exception:
        c.rollback(); raise


def reissue_invitation(c, assignment_id: int, actor_user_id: int) -> dict:
    raw=secrets.token_urlsafe(32); raw_code=verification_code(); expiry=(datetime.now()+timedelta(days=7)).isoformat(timespec='seconds')
    try:
        c.execute('BEGIN IMMEDIATE')
        row=c.execute('SELECT id,status FROM reviewer_assignments WHERE id=?',(assignment_id,)).fetchone()
        if not row: raise ValueError('Reviewer assignment not found.')
        if row['status'] not in ('INVITED','INVITATION_REISSUE_REQUIRED'):
            raise ValueError('Only an unused reviewer invitation can be reissued.')
        c.execute("""UPDATE reviewer_assignments SET status='INVITED',invitation_token_hash=?,invitation_verification_hash=?,
          invitation_verification_attempts=0,invitation_locked_at=NULL,invitation_expires_at=? WHERE id=?""",
          (token_hash(raw),token_hash(raw_code.replace('-','').upper()),expiry,assignment_id))
        c.execute('INSERT INTO reviewer_audit_events(actor_user_id,event_type,assignment_id,metadata_json) VALUES(?,?,?,?)',
                  (actor_user_id,'REVIEW_INVITATION_REISSUED',assignment_id,json.dumps({'expires_at':expiry,'two_part_invitation':True})))
        c.commit(); return {'assignment_id':assignment_id,'raw_token':raw,'verification_code':raw_code,'expires_at':expiry}
    except Exception:
        c.rollback(); raise


def accept_invitation(c, raw_token: str, reviewer_user_id: int, verification: str, password_hash: str='') -> dict:
    if not _text(password_hash): raise ValueError('A verified reviewer password is required to activate the invitation.')
    try:
        c.execute('BEGIN IMMEDIATE')
        row=c.execute('''SELECT * FROM reviewer_assignments WHERE invitation_token_hash=? AND reviewer_user_id=? AND status='INVITED' ''',
                      (token_hash(raw_token),reviewer_user_id)).fetchone()
        if not row: raise ValueError('Invitation is invalid or no longer available.')
        if row['invitation_locked_at']: raise ValueError('Invitation verification is locked. Ask the administrator for a new invitation.')
        if datetime.fromisoformat(row['invitation_expires_at'])<datetime.now(): raise ValueError('Invitation has expired.')
        supplied=token_hash(_text(verification).replace('-','').upper())
        expected=_text(row['invitation_verification_hash'])
        if not expected or not secrets.compare_digest(supplied,expected):
            attempts=int(row['invitation_verification_attempts'] or 0)+1
            locked=datetime.now().isoformat(timespec='seconds') if attempts>=INVITATION_CODE_ATTEMPT_LIMIT else None
            c.execute('UPDATE reviewer_assignments SET invitation_verification_attempts=?,invitation_locked_at=? WHERE id=?',(attempts,locked,row['id']))
            c.execute('INSERT INTO reviewer_audit_events(actor_user_id,event_type,assignment_id,metadata_json) VALUES(?,?,?,?)',
                      (reviewer_user_id,'REVIEW_INVITATION_CODE_REJECTED',row['id'],json.dumps({'attempt':attempts,'locked':bool(locked)})))
            c.commit()
            raise ValueError('The separate reviewer verification code is incorrect.')
        now=datetime.now().isoformat(timespec='seconds')
        if password_hash:
            c.execute('UPDATE users SET password_hash=? WHERE id=?',(password_hash,reviewer_user_id))
        c.execute('''UPDATE reviewer_assignments SET status='IN_PROGRESS',started_at=COALESCE(started_at,?),
          confidentiality_accepted_at=?,invitation_token_hash=NULL,invitation_verification_hash=NULL,
          invitation_verification_attempts=0 WHERE id=?''',(now,now,row['id']))
        group_code=_text(row['assignment_group_code']) if 'assignment_group_code' in row.keys() else ''
        activated=0
        if group_code:
            cur=c.execute('''UPDATE reviewer_assignments SET status='IN_PROGRESS',started_at=COALESCE(started_at,?),
              confidentiality_accepted_at=COALESCE(confidentiality_accepted_at,?)
              WHERE reviewer_user_id=? AND assignment_group_code=? AND status='PENDING_ACTIVATION' ''',
              (now,now,reviewer_user_id,group_code))
            activated=int(cur.rowcount or 0)
        c.execute('INSERT INTO reviewer_audit_events(actor_user_id,event_type,assignment_id,metadata_json) VALUES(?,?,?,?)',
          (reviewer_user_id,'REVIEW_INVITATION_ACCEPTED',row['id'],json.dumps({'additional_assignments_activated':activated,'assignment_group_code':group_code})))
        c.commit(); return dict(row)
    except ValueError:
        if c.in_transaction: c.rollback()
        raise
    except Exception:
        c.rollback(); raise


def next_unfinished_item(c, assignment_id: int):
    return c.execute("SELECT * FROM reviewer_assignment_items WHERE assignment_id=? AND status<>'COMPLETED' ORDER BY display_order LIMIT 1",(assignment_id,)).fetchone()


def open_item(c, assignment_id: int, item_id: int, reviewer_user_id: int):
    item=c.execute('''SELECT rai.*,rq.*,ra.reviewer_user_id,ra.status assignment_status,ra.assignment_code,rb.title batch_title
      FROM reviewer_assignment_items rai JOIN reviewer_assignments ra ON ra.id=rai.assignment_id
      JOIN reviewer_questions rq ON rq.id=rai.question_id JOIN reviewer_batches rb ON rb.id=rq.batch_id
      WHERE rai.id=? AND rai.assignment_id=? AND ra.reviewer_user_id=?''',(item_id,assignment_id,reviewer_user_id)).fetchone()
    if not item: raise PermissionError('Review item unavailable.')
    if item['assignment_status'] not in ('IN_PROGRESS','COMPLETED'): raise PermissionError('Assignment is not active.')
    now=datetime.now().isoformat(timespec='seconds')
    prior=int(item['open_count'] or 0)
    c.execute('''UPDATE reviewer_assignment_items SET first_opened_at=COALESCE(first_opened_at,?),last_opened_at=?,
      open_count=open_count+1,return_count=return_count+? WHERE id=?''',(now,now,1 if prior else 0,item_id))
    c.execute('UPDATE reviewer_assignments SET last_question_id=? WHERE id=?',(item['question_id'],assignment_id))
    c.commit()
    return c.execute('''SELECT rai.*,rq.*,ra.assignment_code,rb.title batch_title FROM reviewer_assignment_items rai
      JOIN reviewer_assignments ra ON ra.id=rai.assignment_id JOIN reviewer_questions rq ON rq.id=rai.question_id
      JOIN reviewer_batches rb ON rb.id=rq.batch_id WHERE rai.id=?''',(item_id,)).fetchone()


def record_active_time(c, item_id: int, reviewer_user_id: int, seconds: int, metadata: Mapping[str,Any]|None=None) -> int:
    try:
        c.execute('BEGIN IMMEDIATE')
        row=c.execute('''SELECT rai.id,rai.question_id,rai.status,rai.last_opened_at,rai.last_ping_at,
          ra.status assignment_status,ra.last_question_id FROM reviewer_assignment_items rai
          JOIN reviewer_assignments ra ON ra.id=rai.assignment_id
          WHERE rai.id=? AND ra.reviewer_user_id=?''',(item_id,reviewer_user_id)).fetchone()
        if not row or row['assignment_status']!='IN_PROGRESS' or row['status']=='COMPLETED':
            raise PermissionError('Review timer unavailable.')
        if int(row['last_question_id'] or 0)!=int(row['question_id']) or not row['last_opened_at']:
            raise PermissionError('Only the currently open review item can receive active time.')
        requested=max(0,min(MAX_TIMER_TICK_SECONDS,int(seconds or 0)))
        if not requested:
            c.rollback(); return 0
        now_dt=datetime.now(); baseline_raw=row['last_ping_at'] or row['last_opened_at']
        try: elapsed=max(0.0,(now_dt-datetime.fromisoformat(baseline_raw)).total_seconds())
        except Exception: elapsed=0.0
        # Server elapsed time is authoritative. Rapid/replayed/multi-tab ticks cannot multiply credit.
        accepted=min(requested,int(elapsed)) if elapsed>=MIN_TIMER_INTERVAL_SECONDS else 0
        now=now_dt.isoformat(timespec='seconds')
        event='ACTIVE_TICK' if accepted else 'ACTIVE_TICK_DISCARDED'
        if accepted:
            c.execute('UPDATE reviewer_assignment_items SET active_seconds=active_seconds+?,last_ping_at=? WHERE id=?',(accepted,now,item_id))
        c.execute('INSERT INTO reviewer_time_events(assignment_item_id,event_type,seconds,metadata_json) VALUES(?,?,?,?)',
                  (item_id,event,accepted,json.dumps({**dict(metadata or {}),'requested_seconds':requested,'server_elapsed_seconds':round(elapsed,3)})))
        c.commit(); return accepted
    except PermissionError:
        c.rollback(); raise
    except Exception:
        c.rollback(); raise


def reveal_answer(c, item_id: int, reviewer_user_id: int) -> None:
    row=c.execute('''SELECT rai.id FROM reviewer_assignment_items rai JOIN reviewer_assignments ra ON ra.id=rai.assignment_id
      WHERE rai.id=? AND ra.reviewer_user_id=?''',(item_id,reviewer_user_id)).fetchone()
    if not row: raise PermissionError('Review item unavailable.')
    now=datetime.now().isoformat(timespec='seconds')
    c.execute('UPDATE reviewer_assignment_items SET answer_revealed_at=COALESCE(answer_revealed_at,?) WHERE id=?',(now,item_id))
    c.execute('INSERT INTO reviewer_time_events(assignment_item_id,event_type) VALUES(?,?)',(item_id,'ANSWER_REVEALED'))
    c.commit()


def _risk_flags(active_seconds: int, answer_revealed: bool, decision: str, comments: str) -> list[str]:
    flags=[]
    if active_seconds<10: flags.append('VERY_FAST_REVIEW')
    elif active_seconds<20: flags.append('FAST_REVIEW')
    if not answer_revealed: flags.append('ANSWER_NOT_REVEALED')
    if decision!='ACCEPT_UNCHANGED' and not _meaningful_comment(comments): flags.append('THIN_REQUIRED_COMMENT')
    return flags


def submit_decision(c, *, item_id: int, reviewer_user_id: int, decision: str, mastery_suitability: str,
                    comments: str='', independent_answer: str='') -> dict:
    decision=_text(decision).upper(); mastery_suitability=_text(mastery_suitability).upper(); comments=_text(comments)
    if decision not in DECISIONS: raise ValueError('Choose a valid review decision.')
    if mastery_suitability not in MASTERY_SUITABILITY: raise ValueError('Record mastery-level suitability.')
    if decision!='ACCEPT_UNCHANGED' and not _meaningful_comment(comments): raise ValueError('Add a meaningful comment with at least two descriptive words.')
    row=c.execute('''SELECT rai.*,ra.round_no,ra.id assignment_id,ra.reviewer_user_id,rq.id question_id,
      rq.calibration_expected_decision,rq.reviewer2_required,rq.review_requirement
      FROM reviewer_assignment_items rai JOIN reviewer_assignments ra ON ra.id=rai.assignment_id
      JOIN reviewer_questions rq ON rq.id=rai.question_id WHERE rai.id=? AND ra.reviewer_user_id=?''',(item_id,reviewer_user_id)).fetchone()
    if not row: raise PermissionError('Review item unavailable.')
    outcome=c.execute('SELECT * FROM reviewer_question_outcomes WHERE question_id=?',(row['question_id'],)).fetchone()
    if row['status']=='COMPLETED':
        if int(row['round_no'])==2 or (outcome and (outcome['second_assignment_item_id'] or outcome['status'] in ('SECOND_REVIEW_AGREED','ADJUDICATION_REQUIRED'))):
            raise ValueError('This decision is locked because independent review or adjudication has started.')
    prior=_text(row['decision']); now=datetime.now().isoformat(timespec='seconds')
    flags=_risk_flags(int(row['active_seconds'] or 0),bool(row['answer_revealed_at']),decision,comments)
    edit_inc=1 if row['status']=='COMPLETED' and (prior!=decision or _text(row['comments'])!=comments) else 0
    c.execute('''UPDATE reviewer_assignment_items SET status='COMPLETED',previous_decision=?,decision=?,mastery_suitability=?,comments=?,
      independent_answer=?,submitted_at=?,edit_count=edit_count+?,risk_flags_json=? WHERE id=?''',
      (prior,decision,mastery_suitability,comments,_text(independent_answer),now,edit_inc,json.dumps(flags),item_id))
    if int(row['round_no'])==1:
        governed_dual=bool(int(row['reviewer2_required'] or 0)) or 'DUAL_REVIEW_REQUIRED' in _text(row['review_requirement']).upper()
        status='FIRST_REVIEW_ACCEPTED' if decision=='ACCEPT_UNCHANGED' and not governed_dual else 'SECOND_REVIEW_REQUIRED'
        c.execute('''UPDATE reviewer_question_outcomes SET first_assignment_item_id=?,first_decision=?,status=?,updated_at=CURRENT_TIMESTAMP
          WHERE question_id=?''',(item_id,decision,status,row['question_id']))
    else:
        first=_text(outcome['first_decision']) if outcome else ''
        status='SECOND_REVIEW_AGREED' if first==decision else 'ADJUDICATION_REQUIRED'
        c.execute('''UPDATE reviewer_question_outcomes SET second_assignment_item_id=?,second_decision=?,status=?,
          adjudication_status=?,updated_at=CURRENT_TIMESTAMP WHERE question_id=?''',
          (item_id,decision,status,'OPEN' if status=='ADJUDICATION_REQUIRED' else '',row['question_id']))
    c.execute('INSERT INTO reviewer_audit_events(actor_user_id,event_type,assignment_id,assignment_item_id,metadata_json) VALUES(?,?,?,?,?)',
              (reviewer_user_id,'REVIEW_DECISION_SUBMITTED',row['assignment_id'],item_id,json.dumps({'decision':decision,'risk_flags':flags})))
    remaining=c.execute("SELECT COUNT(*) n FROM reviewer_assignment_items WHERE assignment_id=? AND status<>'COMPLETED'",(row['assignment_id'],)).fetchone()['n']
    if not remaining:
        c.execute("UPDATE reviewer_assignments SET status='COMPLETED',completed_at=? WHERE id=?",(now,row['assignment_id']))
    c.commit(); return {'remaining':remaining,'risk_flags':flags,'outcome_status':status}


def assignment_progress(c, assignment_id: int) -> dict:
    row=c.execute('''SELECT COUNT(*) total,SUM(CASE WHEN status='COMPLETED' THEN 1 ELSE 0 END) completed,
      COALESCE(SUM(active_seconds),0) active_seconds,COALESCE(AVG(CASE WHEN status='COMPLETED' THEN active_seconds END),0) avg_seconds
      FROM reviewer_assignment_items WHERE assignment_id=?''',(assignment_id,)).fetchone()
    values=[r['active_seconds'] for r in c.execute("SELECT active_seconds FROM reviewer_assignment_items WHERE assignment_id=? AND status='COMPLETED' ORDER BY active_seconds",(assignment_id,)).fetchall()]
    median=0
    if values:
        n=len(values); median=values[n//2] if n%2 else round((values[n//2-1]+values[n//2])/2,1)
    return {'total':int(row['total'] or 0),'completed':int(row['completed'] or 0),'active_seconds':int(row['active_seconds'] or 0),
            'avg_seconds':round(float(row['avg_seconds'] or 0),1),'median_seconds':median}


def assignment_quality(c, assignment_id: int) -> dict:
    items=c.execute('''SELECT rai.*,rq.calibration_expected_decision FROM reviewer_assignment_items rai
      JOIN reviewer_questions rq ON rq.id=rai.question_id WHERE rai.assignment_id=? ORDER BY rai.display_order''',(assignment_id,)).fetchall()
    fast=sum(1 for x in items if x['status']=='COMPLETED' and int(x['active_seconds'] or 0)<10)
    no_reveal=sum(1 for x in items if x['status']=='COMPLETED' and not x['answer_revealed_at'])
    calibration=[x for x in items if _text(x['calibration_expected_decision'])]
    calibration_correct=sum(1 for x in calibration if _text(x['decision'])==_text(x['calibration_expected_decision']))
    longest_run=0; current=0; previous=None
    for x in items:
        d=_text(x['decision'])
        if not d: continue
        if d==previous: current+=1
        else: previous=d; current=1
        longest_run=max(longest_run,current)
    flags=[]
    if items and fast/max(1,sum(1 for x in items if x['status']=='COMPLETED'))>=.25: flags.append('HIGH_FAST_REVIEW_RATE')
    if longest_run>=20: flags.append('LONG_IDENTICAL_DECISION_RUN')
    if calibration and calibration_correct/len(calibration)<.8: flags.append('CALIBRATION_CONCERN')
    if no_reveal: flags.append('ANSWER_REVEAL_GAPS')
    return {'fast_items':fast,'answer_not_revealed':no_reveal,'longest_identical_run':longest_run,
            'calibration_total':len(calibration),'calibration_correct':calibration_correct,'flags':flags}
