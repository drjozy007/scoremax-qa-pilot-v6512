"""ScoreMax V6.2.6 Pre-Pilot Assurance & Mastery Laboratory.

This module is deliberately isolated from the live question bank, attempts and mastery
records. It accepts governed *candidate* material for technical QA only, scores a broad
set of question families, creates a provisional evidence ledger, replays synthetic
student histories and explains every laboratory mastery decision.

Nothing in this module can publish a question or create real student mastery.
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import re
from dataclasses import dataclass
from datetime import datetime, timedelta
from typing import Any, Iterable, Mapping, Sequence

LAB_RELEASE_FLAGS = {
    "content_environment": "QA_SANDBOX_ONLY",
    "student_release_status": "NOT_STUDENT_RELEASED",
    "bank_approval_status": "NOT_BANK_APPROVED",
    "mastery_validity": "NOT_VALID_FOR_REAL_MASTERY",
}

MASTERY_LEVELS = ["Foundation", "Exam Ready", "Advanced", "Distinction"]
MASTERY_STATES = [
    "UNASSESSED",
    "PROVISIONAL_FOUNDATION",
    "VERIFIED_FOUNDATION",
    "PROVISIONAL_EXAM_READY",
    "VERIFIED_EXAM_READY",
    "PROVISIONAL_ADVANCED",
    "VERIFIED_ADVANCED",
    "PROVISIONAL_DISTINCTION",
    "VERIFIED_DISTINCTION",
    "VERIFICATION_DUE",
    "RECOVERY_REQUIRED",
    "RECOVERY_IN_PROGRESS",
    "RECONFIRMED",
]

SUPPORTED_FAMILIES = {
    "standard_mcq",
    "four_statement_selection",
    "true_false",
    "cloze",
    "diagram_data_stimulus",
    "matching",
    "ordering",
    "multiple_response",
    "numerical_interpretation",
    "constructed_response",
    "misconception_probe",
    "adaptive_recovery",
}

RELATION_TYPES = {
    "independent_seed",
    "true_variant",
    "scaffold",
    "shared_stimulus_pair",
    "integrated_question",
    "recovery_item",
    "reconfirmation_item",
}

RELATION_WEIGHTS = {
    "independent_seed": 1.0,
    "true_variant": 0.35,
    "scaffold": 0.20,
    "shared_stimulus_pair": 0.50,
    "integrated_question": 1.0,
    "recovery_item": 0.50,
    "reconfirmation_item": 1.0,
}

INDEPENDENT_RELATIONS = {"independent_seed", "integrated_question", "reconfirmation_item"}
PROGRESSION_RELATIONS = {"independent_seed", "true_variant", "shared_stimulus_pair", "integrated_question"}

DEFAULT_POLICIES = {
    "Foundation": {
        "min_weighted_accuracy": 0.65,
        "min_independent_units": 2.0,
        "min_concept_coverage": 0.50,
        "min_lo_coverage": 0.50,
        "min_independence_ratio": 0.50,
        "min_application_ratio": 0.00,
        "min_integrated_units": 0.0,
        "min_independent_forms": 2,
        "verification_days": 90,
    },
    "Exam Ready": {
        "min_weighted_accuracy": 0.70,
        "min_independent_units": 3.0,
        "min_concept_coverage": 0.70,
        "min_lo_coverage": 0.65,
        "min_independence_ratio": 0.60,
        "min_application_ratio": 0.20,
        "min_integrated_units": 0.0,
        "min_independent_forms": 2,
        "verification_days": 75,
    },
    "Advanced": {
        "min_weighted_accuracy": 0.75,
        "min_independent_units": 4.0,
        "min_concept_coverage": 0.80,
        "min_lo_coverage": 0.75,
        "min_independence_ratio": 0.70,
        "min_application_ratio": 0.35,
        "min_integrated_units": 0.0,
        "min_independent_forms": 2,
        "verification_days": 60,
    },
    "Distinction": {
        "min_weighted_accuracy": 0.82,
        "min_independent_units": 5.0,
        "min_concept_coverage": 0.90,
        "min_lo_coverage": 0.85,
        "min_independence_ratio": 0.75,
        "min_application_ratio": 0.45,
        "min_integrated_units": 1.0,
        "min_independent_forms": 2,
        "verification_days": 45,
    },
}

SYNTHETIC_PROFILES = {
    "HIGH_RECALL_WEAK_APPLICATION": "High recall and recognition, but weak application and integration.",
    "REPEATED_VARIANTS_ONLY": "Strong-looking performance created mainly from true variants of one seed.",
    "BROAD_ONE_MISSING_CONCEPT": "Broad success with one concept left materially uncovered.",
    "SCAFFOLD_SUCCESS_FAILED_INDEPENDENT": "Scaffold success followed by failure on an independent retest.",
    "DELAYED_FORGETTING": "Initial apparent mastery followed by failed delayed reconfirmation.",
    "GENUINE_DISTINCTION": "Broad, independent, integrated and durable Distinction evidence.",
    "INCONSISTENT_GUESSED": "Inconsistent answers with very short response times and low confidence.",
}


def canonical_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def checksum(value: Any) -> str:
    return hashlib.sha256(canonical_json(value).encode("utf-8")).hexdigest()


def safe_json(value: Any, default: Any) -> Any:
    if value is None or value == "":
        return default
    if isinstance(value, (dict, list, tuple, int, float, bool)):
        return value
    try:
        return json.loads(str(value))
    except (TypeError, ValueError, json.JSONDecodeError):
        return default


def _text(value: Any) -> str:
    return str(value or "").strip()


def _slug(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "_", _text(value).lower()).strip("_")


def canonical_family_type(value: Any) -> str:
    raw = _slug(value)
    mapping = {
        "mcq": "standard_mcq",
        "standard_mcq": "standard_mcq",
        "single_choice": "standard_mcq",
        "single_answer_mcq": "standard_mcq",
        "four_statement": "four_statement_selection",
        "four_statement_selection": "four_statement_selection",
        "statement_selection": "four_statement_selection",
        "true_false": "true_false",
        "truefalse": "true_false",
        "fill_blank": "cloze",
        "fill_in_the_blank": "cloze",
        "cloze": "cloze",
        "diagram": "diagram_data_stimulus",
        "data_stimulus": "diagram_data_stimulus",
        "diagram_data_stimulus": "diagram_data_stimulus",
        "matching": "matching",
        "ordering": "ordering",
        "sequence": "ordering",
        "multiple_select": "multiple_response",
        "multiple_response": "multiple_response",
        "numerical": "numerical_interpretation",
        "numeric": "numerical_interpretation",
        "numerical_interpretation": "numerical_interpretation",
        "constructed": "constructed_response",
        "short_response": "constructed_response",
        "extended_response": "constructed_response",
        "constructed_response": "constructed_response",
        "misconception": "misconception_probe",
        "misconception_probe": "misconception_probe",
        "adaptive_recovery": "adaptive_recovery",
        "recovery_pathway": "adaptive_recovery",
    }
    return mapping.get(raw, raw)


def canonical_relation_type(value: Any) -> str:
    raw = _slug(value)
    mapping = {
        "seed": "independent_seed",
        "independent": "independent_seed",
        "independent_seed": "independent_seed",
        "variant": "true_variant",
        "true_variant": "true_variant",
        "scaffold": "scaffold",
        "shared_stimulus": "shared_stimulus_pair",
        "shared_stimulus_pair": "shared_stimulus_pair",
        "integrated": "integrated_question",
        "integrated_question": "integrated_question",
        "recovery": "recovery_item",
        "recovery_item": "recovery_item",
        "reconfirmation": "reconfirmation_item",
        "reconfirmation_item": "reconfirmation_item",
    }
    return mapping.get(raw, raw or "independent_seed")


def canonical_level(value: Any) -> str:
    raw = _text(value).casefold()
    for level in MASTERY_LEVELS:
        if raw == level.casefold():
            return level
    aliases = {"exam_ready": "Exam Ready", "examready": "Exam Ready"}
    return aliases.get(_slug(value), "Foundation")


def _list(value: Any) -> list[str]:
    if value is None or value == "":
        return []
    if isinstance(value, list):
        return [_text(x) for x in value if _text(x)]
    if isinstance(value, tuple):
        return [_text(x) for x in value if _text(x)]
    parsed = safe_json(value, None)
    if isinstance(parsed, list):
        return [_text(x) for x in parsed if _text(x)]
    return [_text(x) for x in re.split(r"[|;,]", _text(value)) if _text(x)]


def init_mastery_lab_schema(c) -> None:
    """Create the QA-only laboratory schema. Idempotent and separate from live tables."""
    c.executescript(
        """
        CREATE TABLE IF NOT EXISTS mastery_lab_feature_controls(
          feature_code TEXT PRIMARY KEY,
          state TEXT NOT NULL DEFAULT 'QA_ONLY',
          configuration_json TEXT DEFAULT '{}',
          updated_by INTEGER,
          updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

        CREATE TABLE IF NOT EXISTS mastery_lab_batches(
          id INTEGER PRIMARY KEY,
          batch_code TEXT UNIQUE NOT NULL,
          source_system TEXT DEFAULT 'POWER_HOUSE_CANDIDATE_CORPUS',
          source_reference TEXT DEFAULT '',
          filename TEXT DEFAULT '',
          file_type TEXT DEFAULT 'json',
          programme TEXT DEFAULT '',
          subject TEXT DEFAULT '',
          chapter TEXT DEFAULT '',
          curriculum_version TEXT DEFAULT '',
          payload_checksum TEXT UNIQUE NOT NULL,
          row_count INTEGER DEFAULT 0,
          imported_count INTEGER DEFAULT 0,
          error_count INTEGER DEFAULT 0,
          warning_count INTEGER DEFAULT 0,
          unresolved_warning_count INTEGER DEFAULT 0,
          status TEXT NOT NULL DEFAULT 'IMPORTED_QA_ONLY',
          release_flags_json TEXT NOT NULL,
          validation_report_json TEXT DEFAULT '{}',
          imported_by INTEGER,
          created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

        CREATE TABLE IF NOT EXISTS mastery_lab_questions(
          id INTEGER PRIMARY KEY,
          batch_id INTEGER NOT NULL,
          external_question_id TEXT NOT NULL,
          external_version TEXT DEFAULT '1',
          content_checksum TEXT NOT NULL,
          family_type TEXT NOT NULL,
          response_mode TEXT DEFAULT '',
          family_id TEXT DEFAULT '',
          seed_key TEXT DEFAULT '',
          relation_type TEXT NOT NULL DEFAULT 'independent_seed',
          parent_external_question_id TEXT DEFAULT '',
          relation_group_key TEXT DEFAULT '',
          stimulus_group_key TEXT DEFAULT '',
          programme TEXT NOT NULL,
          subject TEXT NOT NULL,
          chapter TEXT NOT NULL,
          topic TEXT DEFAULT '',
          subtopic TEXT DEFAULT '',
          learning_outcome_ids_json TEXT DEFAULT '[]',
          concept_ids_json TEXT DEFAULT '[]',
          mastery_level TEXT NOT NULL DEFAULT 'Foundation',
          mastery_ceiling TEXT NOT NULL DEFAULT 'Foundation',
          cognitive_demand TEXT DEFAULT '',
          command_verb TEXT DEFAULT '',
          assessment_purpose TEXT DEFAULT '',
          question_text TEXT NOT NULL,
          stimulus_json TEXT DEFAULT '{}',
          options_json TEXT DEFAULT '[]',
          answer_config_json TEXT DEFAULT '{}',
          marking_config_json TEXT DEFAULT '{}',
          misconception_tags_json TEXT DEFAULT '[]',
          source_lineage_json TEXT DEFAULT '{}',
          warnings_json TEXT DEFAULT '[]',
          content_environment TEXT NOT NULL DEFAULT 'QA_SANDBOX_ONLY',
          student_release_status TEXT NOT NULL DEFAULT 'NOT_STUDENT_RELEASED',
          bank_approval_status TEXT NOT NULL DEFAULT 'NOT_BANK_APPROVED',
          mastery_validity TEXT NOT NULL DEFAULT 'NOT_VALID_FOR_REAL_MASTERY',
          active INTEGER DEFAULT 1,
          created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
          UNIQUE(batch_id,external_question_id,external_version));

        CREATE TABLE IF NOT EXISTS mastery_lab_question_relations(
          id INTEGER PRIMARY KEY,
          batch_id INTEGER NOT NULL,
          question_id INTEGER NOT NULL,
          related_question_id INTEGER,
          related_external_question_id TEXT DEFAULT '',
          relation_type TEXT NOT NULL,
          relation_group_key TEXT DEFAULT '',
          metadata_json TEXT DEFAULT '{}',
          created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
          UNIQUE(question_id,related_external_question_id,relation_type));

        CREATE TABLE IF NOT EXISTS mastery_lab_policies(
          mastery_level TEXT PRIMARY KEY,
          level_rank INTEGER NOT NULL,
          policy_json TEXT NOT NULL,
          active INTEGER DEFAULT 1,
          updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

        CREATE TABLE IF NOT EXISTS mastery_lab_synthetic_profiles(
          profile_code TEXT PRIMARY KEY,
          title TEXT NOT NULL,
          description TEXT NOT NULL,
          configuration_json TEXT DEFAULT '{}',
          active INTEGER DEFAULT 1);

        CREATE TABLE IF NOT EXISTS mastery_lab_runs(
          id INTEGER PRIMARY KEY,
          run_code TEXT UNIQUE NOT NULL,
          batch_id INTEGER NOT NULL,
          profile_code TEXT NOT NULL,
          programme TEXT NOT NULL,
          subject TEXT NOT NULL,
          chapter TEXT NOT NULL,
          status TEXT NOT NULL DEFAULT 'RUNNING',
          previous_state TEXT DEFAULT 'UNASSESSED',
          final_state TEXT DEFAULT 'UNASSESSED',
          provisional_level TEXT DEFAULT '',
          decision_summary TEXT DEFAULT '',
          next_action TEXT DEFAULT '',
          metrics_json TEXT DEFAULT '{}',
          rationale_json TEXT DEFAULT '{}',
          release_flags_json TEXT NOT NULL,
          started_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
          completed_at TEXT DEFAULT '',
          created_by INTEGER);

        CREATE TABLE IF NOT EXISTS mastery_lab_responses(
          id INTEGER PRIMARY KEY,
          run_id INTEGER NOT NULL,
          question_id INTEGER NOT NULL,
          phase_no INTEGER DEFAULT 1,
          days_from_start INTEGER DEFAULT 0,
          response_json TEXT DEFAULT '{}',
          awarded_marks REAL DEFAULT 0,
          max_marks REAL DEFAULT 1,
          is_correct INTEGER DEFAULT 0,
          score_fraction REAL DEFAULT 0,
          confidence TEXT DEFAULT '',
          response_time_seconds INTEGER DEFAULT 0,
          manual_review_required INTEGER DEFAULT 0,
          scoring_explanation TEXT DEFAULT '',
          diagnostic_tags_json TEXT DEFAULT '[]',
          identity_cluster_key TEXT DEFAULT '',
          identity_weight REAL DEFAULT 0,
          progression_eligible INTEGER DEFAULT 0,
          created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
          UNIQUE(run_id,question_id,phase_no));

        CREATE TABLE IF NOT EXISTS mastery_lab_evidence(
          id INTEGER PRIMARY KEY,
          run_id INTEGER NOT NULL,
          phase_no INTEGER DEFAULT 1,
          evidence_type TEXT NOT NULL,
          evidence_key TEXT NOT NULL,
          mastery_level TEXT NOT NULL,
          raw_weight REAL DEFAULT 0,
          effective_weight REAL DEFAULT 0,
          correct_weight REAL DEFAULT 0,
          accuracy REAL DEFAULT 0,
          independent_weight REAL DEFAULT 0,
          scaffold_weight REAL DEFAULT 0,
          variant_weight REAL DEFAULT 0,
          application_weight REAL DEFAULT 0,
          integrated_weight REAL DEFAULT 0,
          misconception_count INTEGER DEFAULT 0,
          validity TEXT NOT NULL DEFAULT 'SANDBOX_PROVISIONAL_ONLY',
          explanation TEXT DEFAULT '',
          metadata_json TEXT DEFAULT '{}',
          created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
          UNIQUE(run_id,phase_no,evidence_type,evidence_key,mastery_level));

        CREATE TABLE IF NOT EXISTS mastery_lab_state_history(
          id INTEGER PRIMARY KEY,
          run_id INTEGER NOT NULL,
          phase_no INTEGER DEFAULT 1,
          days_from_start INTEGER DEFAULT 0,
          previous_state TEXT NOT NULL,
          new_state TEXT NOT NULL,
          provisional_level TEXT DEFAULT '',
          event_type TEXT NOT NULL,
          rationale_json TEXT DEFAULT '{}',
          next_action TEXT DEFAULT '',
          created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

        CREATE TABLE IF NOT EXISTS mastery_lab_recovery_needs(
          id INTEGER PRIMARY KEY,
          run_id INTEGER NOT NULL,
          concept_id TEXT DEFAULT '',
          learning_outcome_id TEXT DEFAULT '',
          priority TEXT DEFAULT 'MEDIUM',
          reason_code TEXT NOT NULL,
          explanation TEXT NOT NULL,
          recommended_action TEXT NOT NULL,
          status TEXT DEFAULT 'OPEN',
          created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

        CREATE TABLE IF NOT EXISTS mastery_lab_gate_results(
          id INTEGER PRIMARY KEY,
          batch_id INTEGER,
          run_id INTEGER,
          gate_code TEXT NOT NULL,
          gate_name TEXT NOT NULL,
          status TEXT NOT NULL,
          summary TEXT NOT NULL,
          evidence_json TEXT DEFAULT '{}',
          evaluated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

        CREATE TABLE IF NOT EXISTS mastery_lab_blockers(
          id INTEGER PRIMARY KEY,
          blocker_code TEXT UNIQUE NOT NULL,
          batch_id INTEGER,
          run_id INTEGER,
          gate_code TEXT DEFAULT '',
          severity TEXT NOT NULL DEFAULT 'HIGH',
          title TEXT NOT NULL,
          description TEXT NOT NULL,
          status TEXT NOT NULL DEFAULT 'OPEN',
          resolution TEXT DEFAULT '',
          created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
          resolved_at TEXT DEFAULT '');

        CREATE TABLE IF NOT EXISTS mastery_lab_audit_events(
          id INTEGER PRIMARY KEY,
          actor_user_id INTEGER,
          event_type TEXT NOT NULL,
          subject_type TEXT DEFAULT '',
          subject_id TEXT DEFAULT '',
          metadata_json TEXT DEFAULT '{}',
          created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);

        CREATE INDEX IF NOT EXISTS idx_lab_questions_batch_scope
          ON mastery_lab_questions(batch_id,programme,subject,chapter,mastery_level,family_type);
        CREATE INDEX IF NOT EXISTS idx_lab_questions_identity
          ON mastery_lab_questions(seed_key,relation_type,relation_group_key,stimulus_group_key);
        CREATE INDEX IF NOT EXISTS idx_lab_responses_run_phase
          ON mastery_lab_responses(run_id,phase_no,question_id);
        CREATE INDEX IF NOT EXISTS idx_lab_evidence_run
          ON mastery_lab_evidence(run_id,phase_no,evidence_type,evidence_key);
        CREATE INDEX IF NOT EXISTS idx_lab_blockers_status
          ON mastery_lab_blockers(status,severity,gate_code);
        """
    )
    c.execute(
        """INSERT OR IGNORE INTO mastery_lab_feature_controls(feature_code,state,configuration_json)
        VALUES('mastery_laboratory','QA_ONLY',?)""",
        (
            canonical_json(
                {
                    "student_routes": False,
                    "live_bank_writes": False,
                    "real_mastery_writes": False,
                    "require_all_release_flags": True,
                    "allowed_profiles": list(SYNTHETIC_PROFILES),
                }
            ),
        ),
    )
    for rank, level in enumerate(MASTERY_LEVELS, start=1):
        c.execute(
            """INSERT OR IGNORE INTO mastery_lab_policies(mastery_level,level_rank,policy_json)
            VALUES(?,?,?)""",
            (level, rank, canonical_json(DEFAULT_POLICIES[level])),
        )
    for code, description in SYNTHETIC_PROFILES.items():
        title = code.replace("_", " ").title()
        c.execute(
            """INSERT OR IGNORE INTO mastery_lab_synthetic_profiles(profile_code,title,description,configuration_json)
            VALUES(?,?,?,?)""",
            (code, title, description, "{}"),
        )


def parse_candidate_payload(raw: bytes, filename: str = "candidate.json") -> tuple[list[dict[str, Any]], str]:
    """Parse JSON, CSV or XLSX candidate payload into dictionaries.

    XLSX is imported lazily so the pure engine remains usable without openpyxl.
    """
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else "json"
    if suffix == "json":
        payload = json.loads(raw.decode("utf-8-sig"))
        if isinstance(payload, dict):
            rows = payload.get("questions") or payload.get("candidates") or payload.get("items") or []
        else:
            rows = payload
        if not isinstance(rows, list):
            raise ValueError("JSON candidate payload must contain a list of questions.")
        return [dict(x) for x in rows if isinstance(x, Mapping)], "json"
    if suffix == "csv":
        text = raw.decode("utf-8-sig")
        return [dict(r) for r in csv.DictReader(io.StringIO(text))], "csv"
    if suffix in {"xlsx", "xlsm"}:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        rows: list[dict[str, Any]] = []
        for ws in wb.worksheets:
            values = list(ws.iter_rows(values_only=True))
            if not values:
                continue
            headers = [_text(x) for x in values[0]]
            for values_row in values[1:]:
                row = {headers[i]: values_row[i] for i in range(min(len(headers), len(values_row))) if headers[i]}
                if any(_text(v) for v in row.values()):
                    row["_sheet_name"] = ws.title
                    rows.append(row)
        return rows, "xlsx"
    raise ValueError("Mastery Laboratory accepts JSON, CSV or XLSX candidate files.")


def _pick(row: Mapping[str, Any], *names: str, default: Any = "") -> Any:
    for name in names:
        if name in row and row[name] not in (None, ""):
            return row[name]
    lowered = {str(k).strip().casefold(): v for k, v in row.items()}
    for name in names:
        value = lowered.get(name.strip().casefold())
        if value not in (None, ""):
            return value
    return default


def normalize_candidate(row: Mapping[str, Any], row_number: int) -> tuple[dict[str, Any], list[str], list[str]]:
    errors: list[str] = []
    warnings: list[str] = []
    external_id = _text(_pick(row, "external_question_id", "question_id", "Question ID", "id"))
    if not external_id:
        errors.append("Missing external question ID.")
    family_type = canonical_family_type(_pick(row, "family_type", "question_family", "qtype", "Type", "type"))
    if family_type not in SUPPORTED_FAMILIES:
        errors.append(f"Unsupported question family: {family_type or 'blank'}.")
    relation_type = canonical_relation_type(_pick(row, "relation_type", "evidence_identity", "relationship", default="independent_seed"))
    if relation_type not in RELATION_TYPES:
        errors.append(f"Unsupported evidence relationship: {relation_type}.")
    programme = _text(_pick(row, "programme", "framework", "qualification", "Programme"))
    subject = _text(_pick(row, "subject", "Subject"))
    chapter = _text(_pick(row, "chapter", "Chapter"))
    question_text = _text(_pick(row, "question_text", "question", "Question"))
    for label, value in (("programme/framework", programme), ("subject", subject), ("chapter", chapter), ("question text", question_text)):
        if not value:
            errors.append(f"Missing {label}.")
    concepts = _list(_pick(row, "concept_ids", "concept_id", "concept", "Concept ID", "Concept"))
    los = _list(_pick(row, "learning_outcome_ids", "learning_outcome_id", "learning_outcome", "Learning Outcome"))
    if not concepts:
        warnings.append("No concept identity supplied; concept-level evidence will be incomplete.")
    if not los:
        warnings.append("No learning-outcome identity supplied; LO evidence will be incomplete.")
    family_id = _text(_pick(row, "family_id", "Family ID"))
    seed_key = _text(_pick(row, "seed_key", "seed_id", "independent_seed_id", "Seed Key")) or family_id or external_id
    parent_external = _text(_pick(row, "parent_external_question_id", "parent_question_id", "parent_seed_id", "Parent Question ID"))
    relation_group = _text(_pick(row, "relation_group_key", "variant_group", "evidence_group", "Relationship Group"))
    stimulus_group = _text(_pick(row, "stimulus_group_key", "shared_stimulus_id", "stimulus_id", "Stimulus Group"))
    if relation_type in {"true_variant", "scaffold", "recovery_item", "reconfirmation_item"} and not parent_external and not seed_key:
        warnings.append(f"{relation_type} does not identify a parent or seed.")
    if relation_type == "shared_stimulus_pair" and not stimulus_group:
        warnings.append("Shared-stimulus item has no stimulus group key.")
    level = canonical_level(_pick(row, "mastery_level", "level", "Level", default="Foundation"))
    ceiling = canonical_level(_pick(row, "mastery_ceiling", "maximum_mastery", default=level))
    if MASTERY_LEVELS.index(ceiling) < MASTERY_LEVELS.index(level):
        warnings.append("Mastery ceiling is below the authored level; ceiling has been raised to the authored level.")
        ceiling = level
    response_mode = _slug(_pick(row, "response_mode", "marking_mode", default=""))
    options = safe_json(_pick(row, "options", "options_json", default=[]), [])
    if not options:
        legacy_options = []
        for code in ("A", "B", "C", "D"):
            value = _pick(row, code, f"option_{code.lower()}", f"Option {code}")
            if _text(value):
                legacy_options.append({"id": code, "text": _text(value)})
        options = legacy_options
    answer_config = safe_json(_pick(row, "answer_config", "answer_config_json", default={}), {})
    marking_config = safe_json(_pick(row, "marking_config", "marking_config_json", default={}), {})
    legacy_answer = _pick(row, "answer", "correct_answer", "Answer", default="")
    marks = float(_pick(row, "marks", "Marks", default=1) or 1)
    if not isinstance(answer_config, dict):
        answer_config = {}
    if not isinstance(marking_config, dict):
        marking_config = {}
    marking_config.setdefault("marks", marks)
    if family_type in {"standard_mcq", "four_statement_selection", "true_false"}:
        marking_config.setdefault("correct_option_ids", _list(legacy_answer))
        if not marking_config.get("correct_option_ids"):
            errors.append("Objective selection question has no correct option.")
        response_mode = response_mode or "single_choice"
    elif family_type == "multiple_response":
        marking_config.setdefault("correct_option_ids", _list(legacy_answer))
        if not marking_config.get("correct_option_ids"):
            errors.append("Multiple-response question has no correct options.")
        response_mode = response_mode or "multiple_response"
    elif family_type == "true_false":
        response_mode = "true_false"
    elif family_type == "cloze":
        if "blanks" not in answer_config:
            accepted = _list(legacy_answer)
            answer_config["blanks"] = [{"accepted_answers": accepted}] if accepted else []
        if not answer_config.get("blanks"):
            errors.append("Cloze question has no accepted blank answers.")
        response_mode = response_mode or "cloze"
    elif family_type == "matching":
        if "correct_pairs" not in marking_config:
            parsed = safe_json(legacy_answer, {})
            if isinstance(parsed, dict):
                marking_config["correct_pairs"] = parsed
        if not marking_config.get("correct_pairs"):
            errors.append("Matching question has no correct-pairs map.")
        response_mode = response_mode or "matching"
    elif family_type == "ordering":
        if "correct_order" not in marking_config:
            marking_config["correct_order"] = _list(legacy_answer)
        if not marking_config.get("correct_order"):
            errors.append("Ordering question has no correct order.")
        response_mode = response_mode or "ordering"
    elif family_type == "numerical_interpretation":
        if "correct_value" not in marking_config and _text(legacy_answer):
            marking_config["correct_value"] = legacy_answer
        marking_config.setdefault("tolerance", 0)
        if "correct_value" not in marking_config:
            errors.append("Numerical question has no correct value.")
        response_mode = response_mode or "numerical"
    elif family_type == "constructed_response":
        response_mode = response_mode or "constructed_response"
        if not marking_config.get("rubric_points") and not marking_config.get("manual_scoring_required"):
            marking_config["manual_scoring_required"] = True
            warnings.append("Constructed response requires manual or rubric-point scoring in the sandbox.")
    elif family_type in {"diagram_data_stimulus", "misconception_probe", "adaptive_recovery"}:
        response_mode = response_mode or _slug(marking_config.get("response_mode") or answer_config.get("response_mode") or "single_choice")
        marking_config.setdefault("response_mode", response_mode)
    source_lineage = safe_json(_pick(row, "source_lineage", "lineage", "source_evidence", default={}), {})
    if not isinstance(source_lineage, dict):
        source_lineage = {"raw": source_lineage}
    if not source_lineage:
        warnings.append("No source lineage supplied.")
    normalized = {
        "external_question_id": external_id,
        "external_version": _text(_pick(row, "external_version", "version", "question_version", default="1")) or "1",
        "family_type": family_type,
        "response_mode": response_mode,
        "family_id": family_id,
        "seed_key": seed_key,
        "relation_type": relation_type,
        "parent_external_question_id": parent_external,
        "relation_group_key": relation_group,
        "stimulus_group_key": stimulus_group,
        "programme": programme,
        "subject": subject,
        "chapter": chapter,
        "topic": _text(_pick(row, "topic", "Topic")),
        "subtopic": _text(_pick(row, "subtopic", "sub_topic", "Sub-topic")),
        "learning_outcome_ids": los,
        "concept_ids": concepts,
        "mastery_level": level,
        "mastery_ceiling": ceiling,
        "cognitive_demand": _text(_pick(row, "cognitive_demand", "cognitive_skill", "Cognitive Demand", "Cognitive Skill")),
        "command_verb": _text(_pick(row, "command_verb", "command_word", "Command Verb", "Command Word")),
        "assessment_purpose": _text(_pick(row, "assessment_purpose", "purpose", default="sandbox_qa")),
        "question_text": question_text,
        "stimulus": safe_json(_pick(row, "stimulus", "stimulus_json", "stimulus_data", default={}), {}),
        "options": options if isinstance(options, list) else [],
        "answer_config": answer_config,
        "marking_config": marking_config,
        "misconception_tags": _list(_pick(row, "misconception_tags", "misconceptions", "Misconception Tags")),
        "source_lineage": source_lineage,
        "row_number": row_number,
    }
    normalized["content_checksum"] = checksum(normalized)
    return normalized, errors, warnings


def import_candidate_batch(
    c,
    rows: Sequence[Mapping[str, Any]],
    *,
    filename: str,
    file_type: str,
    imported_by: int | None = None,
    source_system: str = "POWER_HOUSE_CANDIDATE_CORPUS",
    source_reference: str = "",
) -> dict[str, Any]:
    """Validate and atomically import a QA-only candidate corpus.

    Any structural error blocks the entire batch. Warnings are preserved and surfaced;
    they are never silently discarded. The function never writes to `questions`.
    """
    normalized_rows: list[dict[str, Any]] = []
    row_reports: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    all_errors: list[str] = []
    all_warnings: list[str] = []
    for idx, row in enumerate(rows, start=1):
        normalized, errors, warnings = normalize_candidate(row, idx)
        key = (normalized["external_question_id"], normalized["external_version"])
        if key in seen and key[0]:
            errors.append(f"Duplicate external question/version inside file: {key[0]} v{key[1]}.")
        seen.add(key)
        normalized_rows.append(normalized)
        row_reports.append({"row_number": idx, "external_question_id": key[0], "errors": errors, "warnings": warnings})
        all_errors.extend([f"Row {idx}: {x}" for x in errors])
        all_warnings.extend([f"Row {idx}: {x}" for x in warnings])
    if not normalized_rows:
        raise ValueError("Candidate corpus is empty.")
    payload = {"source_system": source_system, "source_reference": source_reference, "questions": normalized_rows}
    payload_hash = checksum(payload)
    existing = c.execute("SELECT * FROM mastery_lab_batches WHERE payload_checksum=?", (payload_hash,)).fetchone()
    if existing:
        raise ValueError(f"This exact candidate corpus was already imported as {existing['batch_code']}.")
    if all_errors:
        return {
            "ok": False,
            "errors": all_errors,
            "warnings": all_warnings,
            "rows": row_reports,
            "payload_checksum": payload_hash,
        }
    scopes = {(r["programme"], r["subject"], r["chapter"]) for r in normalized_rows}
    if len(scopes) > 1:
        all_warnings.append("The imported batch contains more than one programme/subject/chapter scope.")
    programme, subject, chapter = sorted(scopes)[0]
    code = "MLB-" + datetime.now().strftime("%Y%m%d%H%M%S") + "-" + payload_hash[:8].upper()
    c.execute("BEGIN IMMEDIATE")
    try:
        cur = c.execute(
            """INSERT INTO mastery_lab_batches(
              batch_code,source_system,source_reference,filename,file_type,programme,subject,chapter,
              payload_checksum,row_count,imported_count,error_count,warning_count,unresolved_warning_count,
              status,release_flags_json,validation_report_json,imported_by)
              VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                code,
                source_system,
                source_reference,
                filename,
                file_type,
                programme,
                subject,
                chapter,
                payload_hash,
                len(normalized_rows),
                len(normalized_rows),
                0,
                len(all_warnings),
                len(all_warnings),
                "IMPORTED_QA_ONLY",
                canonical_json(LAB_RELEASE_FLAGS),
                canonical_json({"rows": row_reports, "errors": [], "warnings": all_warnings}),
                imported_by,
            ),
        )
        batch_id = cur.lastrowid
        inserted: dict[str, int] = {}
        for r in normalized_rows:
            cur = c.execute(
                """INSERT INTO mastery_lab_questions(
                  batch_id,external_question_id,external_version,content_checksum,family_type,response_mode,family_id,seed_key,
                  relation_type,parent_external_question_id,relation_group_key,stimulus_group_key,programme,subject,chapter,
                  topic,subtopic,learning_outcome_ids_json,concept_ids_json,mastery_level,mastery_ceiling,cognitive_demand,
                  command_verb,assessment_purpose,question_text,stimulus_json,options_json,answer_config_json,
                  marking_config_json,misconception_tags_json,source_lineage_json,warnings_json,content_environment,
                  student_release_status,bank_approval_status,mastery_validity,active)
                  VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,1)""",
                (
                    batch_id,
                    r["external_question_id"],
                    r["external_version"],
                    r["content_checksum"],
                    r["family_type"],
                    r["response_mode"],
                    r["family_id"],
                    r["seed_key"],
                    r["relation_type"],
                    r["parent_external_question_id"],
                    r["relation_group_key"],
                    r["stimulus_group_key"],
                    r["programme"],
                    r["subject"],
                    r["chapter"],
                    r["topic"],
                    r["subtopic"],
                    canonical_json(r["learning_outcome_ids"]),
                    canonical_json(r["concept_ids"]),
                    r["mastery_level"],
                    r["mastery_ceiling"],
                    r["cognitive_demand"],
                    r["command_verb"],
                    r["assessment_purpose"],
                    r["question_text"],
                    canonical_json(r["stimulus"]),
                    canonical_json(r["options"]),
                    canonical_json(r["answer_config"]),
                    canonical_json(r["marking_config"]),
                    canonical_json(r["misconception_tags"]),
                    canonical_json(r["source_lineage"]),
                    canonical_json(row_reports[r["row_number"] - 1]["warnings"]),
                    LAB_RELEASE_FLAGS["content_environment"],
                    LAB_RELEASE_FLAGS["student_release_status"],
                    LAB_RELEASE_FLAGS["bank_approval_status"],
                    LAB_RELEASE_FLAGS["mastery_validity"],
                ),
            )
            inserted[r["external_question_id"]] = cur.lastrowid
        for r in normalized_rows:
            qid = inserted[r["external_question_id"]]
            if r["relation_type"] in {"true_variant", "scaffold", "recovery_item", "reconfirmation_item"}:
                related_ext = r["parent_external_question_id"]
            else:
                related_ext = ""
            related_id = inserted.get(related_ext)
            c.execute(
                """INSERT OR IGNORE INTO mastery_lab_question_relations(
                  batch_id,question_id,related_question_id,related_external_question_id,relation_type,relation_group_key,metadata_json)
                  VALUES(?,?,?,?,?,?,?)""",
                (
                    batch_id,
                    qid,
                    related_id,
                    related_ext,
                    r["relation_type"],
                    r["relation_group_key"] or r["stimulus_group_key"] or r["seed_key"],
                    canonical_json({"stimulus_group_key": r["stimulus_group_key"], "seed_key": r["seed_key"]}),
                ),
            )
        c.execute(
            """INSERT INTO mastery_lab_audit_events(actor_user_id,event_type,subject_type,subject_id,metadata_json)
            VALUES(?,?,?,?,?)""",
            (imported_by, "CANDIDATE_BATCH_IMPORTED", "mastery_lab_batch", str(batch_id), canonical_json({"batch_code": code, "rows": len(normalized_rows)})),
        )
        c.commit()
    except Exception:
        c.rollback()
        raise
    evaluate_content_admission_gate(c, batch_id)
    c.commit()
    return {
        "ok": True,
        "batch_id": batch_id,
        "batch_code": code,
        "imported_count": len(normalized_rows),
        "warning_count": len(all_warnings),
        "warnings": all_warnings,
        "payload_checksum": payload_hash,
    }


def _row_dict(q: Any) -> dict[str, Any]:
    if isinstance(q, dict):
        return q
    if hasattr(q, "keys"):
        return {k: q[k] for k in q.keys()}
    raise TypeError("Question must be a mapping or sqlite row.")


def _response_payload(response: Any) -> Any:
    if isinstance(response, (dict, list, int, float, bool)):
        return response
    parsed = safe_json(response, None)
    return parsed if parsed is not None else _text(response)


def _normalise_scalar(value: Any, case_sensitive: bool = False) -> str:
    result = re.sub(r"\s+", " ", _text(value)).strip()
    return result if case_sensitive else result.casefold()


def _score_selection(correct: Iterable[Any], response: Any, max_marks: float, partial: bool = False) -> tuple[float, bool, str]:
    correct_set = {_normalise_scalar(x) for x in correct if _text(x)}
    if isinstance(response, str):
        selected = {_normalise_scalar(x) for x in re.split(r"[,|;]", response) if _text(x)}
    elif isinstance(response, Sequence) and not isinstance(response, (bytes, bytearray)):
        selected = {_normalise_scalar(x) for x in response if _text(x)}
    else:
        selected = {_normalise_scalar(response)} if _text(response) else set()
    exact = bool(correct_set) and selected == correct_set
    if exact:
        return max_marks, True, "Selected response exactly matches the governed correct option set."
    if partial and correct_set:
        good = len(selected & correct_set)
        wrong = len(selected - correct_set)
        fraction = max(0.0, min(1.0, good / len(correct_set) - wrong / len(correct_set)))
        return round(max_marks * fraction, 4), False, f"Partial-credit set score: {good} correct selection(s), {wrong} incorrect selection(s)."
    return 0.0, False, "Selected response does not match the governed correct option set."


def score_lab_response(question: Mapping[str, Any], response: Any, *, manual_score: float | None = None) -> dict[str, Any]:
    """Score one QA-sandbox response across supported question families."""
    q = _row_dict(question)
    family = canonical_family_type(q.get("family_type") or q.get("qtype"))
    response_mode = _slug(q.get("response_mode") or "")
    answer_cfg = safe_json(q.get("answer_config_json") or q.get("answer_config"), {})
    marking = safe_json(q.get("marking_config_json") or q.get("marking_config"), {})
    misconceptions = safe_json(q.get("misconception_tags_json") or q.get("misconception_tags"), [])
    if not isinstance(answer_cfg, dict):
        answer_cfg = {}
    if not isinstance(marking, dict):
        marking = {}
    max_marks = float(marking.get("marks") or 1)
    payload = _response_payload(response)
    awarded = 0.0
    correct = False
    explanation = ""
    manual_required = False
    diagnostic_tags: list[str] = []

    effective_mode = response_mode
    if family in {"standard_mcq", "four_statement_selection", "true_false"}:
        effective_mode = "single_choice"
    elif family == "multiple_response":
        effective_mode = "multiple_response"
    elif family == "cloze":
        effective_mode = "cloze"
    elif family == "matching":
        effective_mode = "matching"
    elif family == "ordering":
        effective_mode = "ordering"
    elif family == "numerical_interpretation":
        effective_mode = "numerical"
    elif family == "constructed_response":
        effective_mode = "constructed_response"
    elif family in {"diagram_data_stimulus", "misconception_probe", "adaptive_recovery"}:
        effective_mode = _slug(marking.get("response_mode") or answer_cfg.get("response_mode") or effective_mode or "single_choice")

    if effective_mode in {"single_choice", "true_false"}:
        correct_ids = marking.get("correct_option_ids") or answer_cfg.get("correct_option_ids") or []
        awarded, correct, explanation = _score_selection(correct_ids, [payload], max_marks, False)
    elif effective_mode in {"multiple_response", "multiple_select"}:
        correct_ids = marking.get("correct_option_ids") or []
        awarded, correct, explanation = _score_selection(correct_ids, payload, max_marks, bool(marking.get("partial_credit")))
    elif effective_mode == "cloze":
        blanks = answer_cfg.get("blanks") or []
        responses = payload if isinstance(payload, list) else [payload]
        points = []
        for idx, blank in enumerate(blanks):
            blank = blank if isinstance(blank, dict) else {"accepted_answers": [blank]}
            accepted = blank.get("accepted_answers") or []
            candidate = responses[idx] if idx < len(responses) else ""
            norm = _normalise_scalar(candidate, bool(blank.get("case_sensitive")))
            accepted_norm = {_normalise_scalar(x, bool(blank.get("case_sensitive"))) for x in accepted}
            points.append(1.0 if norm in accepted_norm and accepted_norm else 0.0)
        fraction = sum(points) / len(points) if points else 0.0
        awarded = max_marks * (fraction if marking.get("partial_credit", True) else (1.0 if fraction == 1 else 0.0))
        correct = bool(points) and fraction == 1.0
        explanation = f"Cloze score: {int(sum(points))}/{len(points)} governed blank(s) correct."
    elif effective_mode == "matching":
        correct_pairs = marking.get("correct_pairs") or {}
        response_pairs = payload if isinstance(payload, dict) else {}
        total = len(correct_pairs)
        good = sum(1 for key, value in correct_pairs.items() if _normalise_scalar(response_pairs.get(key)) == _normalise_scalar(value))
        fraction = good / total if total else 0.0
        awarded = max_marks * (fraction if marking.get("partial_credit", True) else (1.0 if fraction == 1 else 0.0))
        correct = total > 0 and good == total
        explanation = f"Matching score: {good}/{total} governed pair(s) correct."
    elif effective_mode == "ordering":
        correct_order = [_normalise_scalar(x) for x in marking.get("correct_order") or []]
        response_order = [_normalise_scalar(x) for x in (payload if isinstance(payload, list) else _list(payload))]
        total = len(correct_order)
        position_good = sum(1 for idx, value in enumerate(correct_order) if idx < len(response_order) and response_order[idx] == value)
        exact = total > 0 and response_order == correct_order
        fraction = position_good / total if total else 0.0
        awarded = max_marks * (fraction if marking.get("partial_credit") else (1.0 if exact else 0.0))
        correct = exact
        explanation = f"Ordering score: {position_good}/{total} item(s) in the governed position."
    elif effective_mode in {"numerical", "numeric"}:
        try:
            value = float(payload)
            target = float(marking.get("correct_value"))
            tolerance = float(marking.get("tolerance") or 0)
            relative_tolerance = float(marking.get("relative_tolerance") or 0)
            allowed = max(tolerance, abs(target) * relative_tolerance)
            correct = abs(value - target) <= allowed
            awarded = max_marks if correct else 0.0
            explanation = f"Numerical response compared with {target:g} using allowed tolerance {allowed:g}."
        except (TypeError, ValueError):
            explanation = "Response could not be parsed as a number."
    elif effective_mode in {"constructed_response", "short_response", "extended_response"}:
        if manual_score is not None:
            awarded = max(0.0, min(max_marks, float(manual_score)))
            correct = awarded >= max_marks
            explanation = "Manual sandbox score supplied against the governed rubric."
        else:
            rubric_points = marking.get("rubric_points") or []
            text = _normalise_scalar(payload)
            if rubric_points:
                earned = 0.0
                total_available = 0.0
                for point in rubric_points:
                    point = point if isinstance(point, dict) else {"keywords": [point], "marks": 1}
                    point_marks = float(point.get("marks") or 1)
                    total_available += point_marks
                    keywords = [_normalise_scalar(x) for x in point.get("keywords") or point.get("accepted_phrases") or []]
                    if keywords and any(keyword in text for keyword in keywords):
                        earned += point_marks
                scale = max_marks / total_available if total_available else 0.0
                awarded = min(max_marks, earned * scale)
                correct = math.isclose(awarded, max_marks)
                explanation = f"Constructed-response sandbox rubric awarded {awarded:g}/{max_marks:g} marks."
            else:
                manual_required = True
                explanation = "Constructed response requires manual or external governed marking; no automatic mastery evidence was awarded."
    else:
        raise ValueError(f"Unsupported laboratory response mode: {effective_mode or family}.")

    incorrect_map = marking.get("misconception_by_response") or answer_cfg.get("misconception_by_response") or {}
    if not correct:
        key = _normalise_scalar(payload)
        mapped = incorrect_map.get(str(payload)) or incorrect_map.get(key)
        if mapped:
            diagnostic_tags.extend(_list(mapped))
        elif len(misconceptions) == 1:
            diagnostic_tags.append(_text(misconceptions[0]))
    fraction = 0.0 if max_marks <= 0 else max(0.0, min(1.0, awarded / max_marks))
    return {
        "family_type": family,
        "response_mode": effective_mode,
        "awarded_marks": round(awarded, 4),
        "max_marks": round(max_marks, 4),
        "is_correct": bool(correct),
        "score_fraction": round(fraction, 4),
        "manual_review_required": bool(manual_required),
        "scoring_explanation": explanation,
        "diagnostic_tags": sorted(set(x for x in diagnostic_tags if x)),
    }


def identity_cluster(question: Mapping[str, Any]) -> str:
    q = _row_dict(question)
    relation = canonical_relation_type(q.get("relation_type"))
    if relation == "shared_stimulus_pair":
        return _text(q.get("stimulus_group_key") or q.get("relation_group_key") or q.get("seed_key") or q.get("external_question_id"))
    return _text(q.get("seed_key") or q.get("family_id") or q.get("external_question_id"))


def identity_weight(question: Mapping[str, Any]) -> float:
    relation = canonical_relation_type(_row_dict(question).get("relation_type"))
    return float(RELATION_WEIGHTS.get(relation, 0.0))


def progression_eligible(question: Mapping[str, Any]) -> bool:
    relation = canonical_relation_type(_row_dict(question).get("relation_type"))
    return relation in PROGRESSION_RELATIONS


def _demand_is_application(value: Any) -> bool:
    raw = _slug(value)
    return any(token in raw for token in ("apply", "application", "analyse", "analysis", "evaluate", "integrat", "unfamiliar", "reason"))


def _level_rank(level: str) -> int:
    try:
        return MASTERY_LEVELS.index(canonical_level(level))
    except ValueError:
        return 0


def _policy_rows(c) -> dict[str, dict[str, Any]]:
    rows = c.execute("SELECT * FROM mastery_lab_policies WHERE active=1 ORDER BY level_rank").fetchall()
    return {r["mastery_level"]: safe_json(r["policy_json"], DEFAULT_POLICIES[r["mastery_level"]]) for r in rows}


def _effective_clustered_rows(question_rows: Sequence[Mapping[str, Any]], response_rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    by_q = {int(q["id"]): _row_dict(q) for q in question_rows}
    prepared: list[dict[str, Any]] = []
    for response in response_rows:
        r = _row_dict(response)
        q = by_q.get(int(r["question_id"]))
        if not q:
            continue
        relation = canonical_relation_type(q.get("relation_type"))
        raw_weight = float(r.get("identity_weight") or identity_weight(q))
        prepared.append(
            {
                "question": q,
                "response": r,
                "cluster": r.get("identity_cluster_key") or identity_cluster(q),
                "relation": relation,
                "raw_weight": raw_weight,
                "score_fraction": float(r.get("score_fraction") or 0),
            }
        )
    # Cap closely-related evidence. A seed/variant family cannot manufacture breadth.
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for item in prepared:
        level = canonical_level(item["question"].get("mastery_level"))
        grouped.setdefault((item["cluster"], level), []).append(item)
    output: list[dict[str, Any]] = []
    for _, items in grouped.items():
        remaining = 1.0
        # Independent evidence gets first claim on the cluster cap; variants/scaffolds fill only unused capacity.
        items.sort(key=lambda x: (0 if x["relation"] in INDEPENDENT_RELATIONS else 1, -x["raw_weight"]))
        for item in items:
            effective = min(remaining, item["raw_weight"])
            remaining = max(0.0, remaining - effective)
            item = dict(item)
            item["effective_weight"] = effective
            output.append(item)
    return output


def _aggregate_clustered_metrics(clustered: Sequence[Mapping[str, Any]], eligible_questions: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    eligible_ids={int(_row_dict(q)["id"]) for q in eligible_questions}
    all_concepts=sorted({x for q in eligible_questions for x in _list(_row_dict(q).get("concept_ids_json"))})
    all_los=sorted({x for q in eligible_questions for x in _list(_row_dict(q).get("learning_outcome_ids_json"))})
    concept_seen:set[str]=set(); lo_seen:set[str]=set()
    raw_weight=effective_weight=correct_weight=independent_weight=scaffold_weight=variant_weight=0.0
    application_weight=integrated_weight=progression_weight=0.0
    form_clusters:set[str]=set(); manual_review_count=fast_low_confidence=0
    relation_counts:dict[str,int]={}
    level_metrics:dict[str,dict[str,float]]={level:{"effective":0.0,"correct":0.0,"independent":0.0,"application":0.0,"integrated":0.0} for level in MASTERY_LEVELS}
    for item in clustered:
        q=item["question"]
        if int(q["id"]) not in eligible_ids:
            continue
        r=item["response"]; relation=item["relation"]; weight=float(item["effective_weight"])
        raw_weight+=float(item["raw_weight"]); effective_weight+=weight; correct_weight+=weight*float(item["score_fraction"])
        level=canonical_level(q.get("mastery_level")); level_metrics[level]["effective"]+=weight; level_metrics[level]["correct"]+=weight*float(item["score_fraction"])
        relation_counts[relation]=relation_counts.get(relation,0)+1
        if relation in INDEPENDENT_RELATIONS:
            independent_weight+=weight; level_metrics[level]["independent"]+=weight; form_clusters.add(item["cluster"])
        if relation=="scaffold": scaffold_weight+=weight
        if relation=="true_variant": variant_weight+=weight
        if _demand_is_application(q.get("cognitive_demand")):
            application_weight+=weight; level_metrics[level]["application"]+=weight
        if relation=="integrated_question":
            integrated_weight+=weight; level_metrics[level]["integrated"]+=weight
        if progression_eligible(q): progression_weight+=weight
        if float(item["score_fraction"])>0:
            concept_seen.update(_list(q.get("concept_ids_json"))); lo_seen.update(_list(q.get("learning_outcome_ids_json")))
        if int(r.get("manual_review_required") or 0): manual_review_count+=1
        if int(r.get("response_time_seconds") or 0)<=5 and _text(r.get("confidence")).casefold() in {"low","unsure","guess","guessed"}: fast_low_confidence+=1
    weighted_accuracy=correct_weight/effective_weight if effective_weight else 0.0
    independence_ratio=independent_weight/effective_weight if effective_weight else 0.0
    application_ratio=application_weight/effective_weight if effective_weight else 0.0
    return {
      "raw_weight":round(raw_weight,4),"effective_weight":round(effective_weight,4),"weighted_accuracy":round(weighted_accuracy,4),
      "independent_units":round(independent_weight,4),"independence_ratio":round(independence_ratio,4),
      "scaffold_weight":round(scaffold_weight,4),"variant_weight":round(variant_weight,4),"application_weight":round(application_weight,4),
      "application_ratio":round(application_ratio,4),"integrated_units":round(integrated_weight,4),"progression_weight":round(progression_weight,4),
      "concepts_total":len(all_concepts),"concepts_covered":len(concept_seen),"concept_coverage":round(len(concept_seen)/len(all_concepts),4) if all_concepts else 0.0,
      "los_total":len(all_los),"los_covered":len(lo_seen),"lo_coverage":round(len(lo_seen)/len(all_los),4) if all_los else 0.0,
      "independent_form_clusters":len(form_clusters),"manual_review_count":manual_review_count,"fast_low_confidence_count":fast_low_confidence,
      "relation_counts":relation_counts,"level_metrics":level_metrics,"concepts_seen":sorted(concept_seen),"los_seen":sorted(lo_seen)}


def calculate_phase_metrics(c, batch_id: int, response_rows: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    question_rows=c.execute("SELECT * FROM mastery_lab_questions WHERE batch_id=? AND active=1",(batch_id,)).fetchall()
    clustered=_effective_clustered_rows(question_rows,response_rows)
    overall=_aggregate_clustered_metrics(clustered,question_rows)
    target_levels={}
    for target in MASTERY_LEVELS:
        target_rank=_level_rank(target)
        eligible=[q for q in question_rows if _level_rank(q["mastery_ceiling"])>=target_rank]
        target_levels[target]=_aggregate_clustered_metrics(clustered,eligible)
    overall["target_levels"]=target_levels
    return overall

def evaluate_level(metrics: Mapping[str, Any], policy: Mapping[str, Any]) -> tuple[bool, list[dict[str, Any]]]:
    checks = [
        ("weighted_accuracy", float(metrics.get("weighted_accuracy") or 0), float(policy.get("min_weighted_accuracy") or 0), ">="),
        ("independent_units", float(metrics.get("independent_units") or 0), float(policy.get("min_independent_units") or 0), ">="),
        ("concept_coverage", float(metrics.get("concept_coverage") or 0), float(policy.get("min_concept_coverage") or 0), ">="),
        ("lo_coverage", float(metrics.get("lo_coverage") or 0), float(policy.get("min_lo_coverage") or 0), ">="),
        ("independence_ratio", float(metrics.get("independence_ratio") or 0), float(policy.get("min_independence_ratio") or 0), ">="),
        ("application_ratio", float(metrics.get("application_ratio") or 0), float(policy.get("min_application_ratio") or 0), ">="),
        ("integrated_units", float(metrics.get("integrated_units") or 0), float(policy.get("min_integrated_units") or 0), ">="),
    ]
    details = [
        {"metric": name, "actual": round(actual, 4), "required": round(required, 4), "passed": actual >= required, "operator": op}
        for name, actual, required, op in checks
    ]
    return all(x["passed"] for x in details), details


def highest_provisional_level(c, metrics: Mapping[str, Any]) -> tuple[str, dict[str, Any]]:
    policies = _policy_rows(c)
    evaluations: dict[str, Any] = {}
    achieved = ""
    for level in MASTERY_LEVELS:
        level_metrics=(metrics.get("target_levels") or {}).get(level) or metrics
        passed, details = evaluate_level(level_metrics, policies[level])
        evaluations[level] = {"passed": passed, "checks": details, "metrics": level_metrics}
        if passed:
            achieved = level
        else:
            # Levels are cumulative; a failed lower level blocks higher progression.
            break
    return achieved, evaluations


def provisional_state(level: str) -> str:
    return "UNASSESSED" if not level else "PROVISIONAL_" + level.upper().replace(" ", "_")


def verified_state(level: str) -> str:
    return "UNASSESSED" if not level else "VERIFIED_" + level.upper().replace(" ", "_")


def _critical_recovery(metrics: Mapping[str, Any]) -> bool:
    return bool(
        metrics.get("effective_weight")
        and (
            float(metrics.get("weighted_accuracy") or 0) < 0.50
            or (int(metrics.get("concepts_total") or 0) > 0 and float(metrics.get("concept_coverage") or 0) < 0.40)
        )
    )


def transition_mastery_state(
    previous_state: str,
    level: str,
    metrics: Mapping[str, Any],
    *,
    event_type: str,
    independent_confirmation: bool = False,
    verification_due: bool = False,
    recovery_item_only: bool = False,
) -> tuple[str, str, list[str]]:
    """Apply the explicit laboratory state machine and return state, action and reasons."""
    previous_state = previous_state if previous_state in MASTERY_STATES else "UNASSESSED"
    reasons: list[str] = []
    if verification_due and previous_state.startswith("VERIFIED_"):
        reasons.append("The configured reconfirmation interval has elapsed.")
        return "VERIFICATION_DUE", "Schedule an independent reconfirmation form.", reasons
    if event_type == "RECONFIRMATION":
        if level and not _critical_recovery(metrics):
            reasons.append("Independent delayed reconfirmation met the provisional evidence thresholds.")
            return "RECONFIRMED", "Retain the current level and schedule the next adaptive verification window.", reasons
        reasons.append("Delayed reconfirmation did not sustain the held level.")
        return "RECOVERY_REQUIRED", "Create a targeted recovery sequence, then require an unseen independent retest.", reasons
    if recovery_item_only:
        reasons.append("Evidence came from recovery/scaffold items and cannot independently verify mastery.")
        return "RECOVERY_IN_PROGRESS", "Continue recovery, then complete an unseen independent retest.", reasons
    if _critical_recovery(metrics):
        reasons.append("Accuracy or curriculum coverage is below the laboratory recovery threshold.")
        return "RECOVERY_REQUIRED", "Target the missing concepts and reassess with independent evidence.", reasons
    if not level:
        reasons.append("No mastery level met the configured provisional evidence requirements.")
        return "UNASSESSED", "Collect broader independent evidence before making a mastery claim.", reasons
    if independent_confirmation:
        reasons.append("A second independent evidence form confirmed the same or higher provisional level.")
        return verified_state(level), "Continue progression while scheduling later reconfirmation.", reasons
    reasons.append("The evidence meets provisional thresholds but has not yet been independently confirmed.")
    return provisional_state(level), "Collect a second unseen independent form before verification.", reasons


def _correct_response_for(q: Mapping[str, Any]) -> Any:
    q = _row_dict(q)
    family = canonical_family_type(q.get("family_type"))
    answer_cfg = safe_json(q.get("answer_config_json"), {})
    marking = safe_json(q.get("marking_config_json"), {})
    mode = _slug(q.get("response_mode") or marking.get("response_mode") or answer_cfg.get("response_mode"))
    if family in {"standard_mcq", "four_statement_selection", "true_false"} or mode in {"single_choice", "true_false"}:
        ids = marking.get("correct_option_ids") or []
        return ids[0] if ids else ""
    if family == "multiple_response" or mode in {"multiple_response", "multiple_select"}:
        return marking.get("correct_option_ids") or []
    if family == "cloze" or mode == "cloze":
        return [(b.get("accepted_answers") or [""])[0] for b in answer_cfg.get("blanks") or []]
    if family == "matching" or mode == "matching":
        return marking.get("correct_pairs") or {}
    if family == "ordering" or mode == "ordering":
        return marking.get("correct_order") or []
    if family == "numerical_interpretation" or mode in {"numerical", "numeric"}:
        return marking.get("correct_value")
    if family == "constructed_response" or mode == "constructed_response":
        points = marking.get("rubric_points") or []
        return " ".join((p.get("keywords") or [""])[0] for p in points if isinstance(p, dict))
    return ""


def _incorrect_response_for(q: Mapping[str, Any]) -> Any:
    q = _row_dict(q)
    correct = _correct_response_for(q)
    family = canonical_family_type(q.get("family_type"))
    options = safe_json(q.get("options_json"), [])
    if family in {"standard_mcq", "four_statement_selection", "true_false", "misconception_probe", "diagram_data_stimulus", "adaptive_recovery"}:
        correct_set = {_normalise_scalar(x) for x in (correct if isinstance(correct, list) else [correct])}
        for option in options:
            oid = option.get("id") if isinstance(option, dict) else option
            if _normalise_scalar(oid) not in correct_set:
                return oid
        return "__WRONG__"
    if family == "multiple_response":
        return []
    if family == "cloze":
        return ["__wrong__" for _ in (correct if isinstance(correct, list) else [correct])]
    if family == "matching":
        return {k: "__wrong__" for k in (correct or {})}
    if family == "ordering":
        return list(reversed(correct or []))
    if family == "numerical_interpretation":
        try:
            return float(correct) + max(10, abs(float(correct)) * 0.5)
        except (TypeError, ValueError):
            return 999999
    return "irrelevant response"


def _profile_response(profile: str, q: Mapping[str, Any], index: int, phase_no: int, missing_concept: str = "") -> tuple[Any, str, int, float | None]:
    q = _row_dict(q)
    relation = canonical_relation_type(q.get("relation_type"))
    demand_application = _demand_is_application(q.get("cognitive_demand")) or _level_rank(q.get("mastery_level")) >= 2
    concepts = _list(q.get("concept_ids_json"))
    correct = True
    confidence = "high"
    response_time = 35
    manual_score = None
    if profile == "HIGH_RECALL_WEAK_APPLICATION":
        correct = not demand_application
        confidence = "high" if correct else "medium"
    elif profile == "REPEATED_VARIANTS_ONLY":
        correct = relation == "true_variant" or (relation == "independent_seed" and index == 0)
        confidence = "high"
    elif profile == "BROAD_ONE_MISSING_CONCEPT":
        correct = not (missing_concept and missing_concept in concepts)
    elif profile == "SCAFFOLD_SUCCESS_FAILED_INDEPENDENT":
        correct = relation == "scaffold" if phase_no == 1 else relation not in INDEPENDENT_RELATIONS
        confidence = "medium"
    elif profile == "DELAYED_FORGETTING":
        correct = phase_no == 1
        confidence = "high" if phase_no == 1 else "low"
    elif profile == "GENUINE_DISTINCTION":
        correct = True
        confidence = "high"
        response_time = 45
    elif profile == "INCONSISTENT_GUESSED":
        correct = (int(hashlib.sha256(f"{q.get('external_question_id')}|{phase_no}".encode()).hexdigest()[:2], 16) % 3) == 0
        confidence = "guess"
        response_time = 3
    response = _correct_response_for(q) if correct else _incorrect_response_for(q)
    if canonical_family_type(q.get("family_type")) == "constructed_response":
        max_marks = float(safe_json(q.get("marking_config_json"), {}).get("marks") or 1)
        manual_score = max_marks if correct else 0.0
    return response, confidence, response_time, manual_score


def _profile_phases(profile: str) -> list[dict[str, Any]]:
    if profile == "SCAFFOLD_SUCCESS_FAILED_INDEPENDENT":
        return [{"phase_no": 1, "days": 0, "event": "RECOVERY"}, {"phase_no": 2, "days": 7, "event": "INDEPENDENT_RETEST"}]
    if profile == "DELAYED_FORGETTING":
        return [{"phase_no": 1, "days": 0, "event": "INITIAL"}, {"phase_no": 2, "days": 120, "event": "RECONFIRMATION"}]
    if profile == "GENUINE_DISTINCTION":
        return [{"phase_no": 1, "days": 0, "event": "INITIAL"}, {"phase_no": 2, "days": 21, "event": "INDEPENDENT_CONFIRMATION"}, {"phase_no": 3, "days": 80, "event": "RECONFIRMATION"}]
    return [{"phase_no": 1, "days": 0, "event": "INITIAL"}]


def _questions_for_phase(questions: Sequence[Mapping[str, Any]], profile: str, phase_no: int) -> list[Mapping[str, Any]]:
    if profile == "REPEATED_VARIANTS_ONLY":
        variants = [q for q in questions if canonical_relation_type(q["relation_type"]) == "true_variant"]
        if variants:
            seed = variants[0]["seed_key"]
            return [q for q in questions if q["seed_key"] == seed][:12]
    if profile == "SCAFFOLD_SUCCESS_FAILED_INDEPENDENT":
        if phase_no == 1:
            return [q for q in questions if canonical_relation_type(q["relation_type"]) in {"scaffold", "recovery_item"}]
        return [q for q in questions if canonical_relation_type(q["relation_type"]) in INDEPENDENT_RELATIONS]
    if profile == "GENUINE_DISTINCTION" and phase_no == 2:
        return [q for q in questions if canonical_relation_type(q["relation_type"]) in INDEPENDENT_RELATIONS]
    if profile in {"DELAYED_FORGETTING", "GENUINE_DISTINCTION"} and phase_no > 1:
        reconfirm = [q for q in questions if canonical_relation_type(q["relation_type"]) == "reconfirmation_item"]
        independent = [q for q in questions if canonical_relation_type(q["relation_type"]) in INDEPENDENT_RELATIONS]
        seen=set(); combined=[]
        for q in reconfirm+independent:
            if q["id"] not in seen:
                seen.add(q["id"]); combined.append(q)
        return combined
    return list(questions)


def _write_evidence_ledger(c, run_id: int, phase_no: int, batch_id: int, responses: Sequence[Mapping[str, Any]], metrics: Mapping[str, Any]) -> None:
    questions=c.execute("SELECT * FROM mastery_lab_questions WHERE batch_id=? AND active=1",(batch_id,)).fetchall()
    clustered=_effective_clustered_rows(questions,responses)
    buckets:dict[tuple[str,str,str],dict[str,Any]]={}
    for item in clustered:
        q=item["question"]; r=item["response"]; relation=item["relation"]
        for evidence_type,ids in (("CONCEPT",_list(q.get("concept_ids_json"))),("LEARNING_OUTCOME",_list(q.get("learning_outcome_ids_json")))):
            split=max(1,len(ids))
            for key in ids:
                bucket=buckets.setdefault((evidence_type,key,canonical_level(q.get("mastery_level"))),
                  {"raw":0.0,"effective":0.0,"correct":0.0,"independent":0.0,"scaffold":0.0,"variant":0.0,"application":0.0,"integrated":0.0,"misconceptions":0})
                raw=float(item["raw_weight"])/split; effective=float(item["effective_weight"])/split
                bucket["raw"]+=raw; bucket["effective"]+=effective; bucket["correct"]+=effective*float(item["score_fraction"])
                if relation in INDEPENDENT_RELATIONS: bucket["independent"]+=effective
                if relation=="scaffold": bucket["scaffold"]+=effective
                if relation=="true_variant": bucket["variant"]+=effective
                if _demand_is_application(q.get("cognitive_demand")): bucket["application"]+=effective
                if relation=="integrated_question": bucket["integrated"]+=effective
                if safe_json(r.get("diagnostic_tags_json"),[]): bucket["misconceptions"]+=1
    for (etype,key,level),b in buckets.items():
        accuracy=b["correct"]/b["effective"] if b["effective"] else 0.0
        explanation=f"{etype.title().replace('_',' ')} {key}: {b['effective']:.2f} identity-capped unit(s), {accuracy*100:.1f}% weighted accuracy; sandbox evidence only."
        c.execute("""INSERT OR REPLACE INTO mastery_lab_evidence(
          run_id,phase_no,evidence_type,evidence_key,mastery_level,raw_weight,effective_weight,correct_weight,accuracy,
          independent_weight,scaffold_weight,variant_weight,application_weight,integrated_weight,misconception_count,
          validity,explanation,metadata_json) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
          (run_id,phase_no,etype,key,level,round(b["raw"],4),round(b["effective"],4),round(b["correct"],4),round(accuracy,4),
           round(b["independent"],4),round(b["scaffold"],4),round(b["variant"],4),round(b["application"],4),round(b["integrated"],4),
           b["misconceptions"],"SANDBOX_PROVISIONAL_ONLY",explanation,canonical_json({"batch_id":batch_id,"identity_cap_applied":True})))

def _identify_recovery_needs(c, run_id: int, phase_no: int) -> list[dict[str, Any]]:
    rows = c.execute(
        """SELECT * FROM mastery_lab_evidence WHERE run_id=? AND phase_no=? AND evidence_type='CONCEPT'
        ORDER BY accuracy ASC,effective_weight DESC""",
        (run_id, phase_no),
    ).fetchall()
    needs: list[dict[str, Any]] = []
    for row in rows:
        if float(row["accuracy"] or 0) < 0.60 or float(row["independent_weight"] or 0) < 0.50:
            priority = "HIGH" if float(row["accuracy"] or 0) < 0.40 else "MEDIUM"
            reason = "LOW_ACCURACY" if float(row["accuracy"] or 0) < 0.60 else "INSUFFICIENT_INDEPENDENT_EVIDENCE"
            action = "Assign targeted explanation and scaffold, then require an unseen independent retest."
            explanation = f"{row['evidence_key']} has {float(row['accuracy'] or 0)*100:.1f}% weighted accuracy and {float(row['independent_weight'] or 0):.2f} independent unit(s)."
            c.execute(
                """INSERT INTO mastery_lab_recovery_needs(run_id,concept_id,priority,reason_code,explanation,recommended_action)
                VALUES(?,?,?,?,?,?)""",
                (run_id, row["evidence_key"], priority, reason, explanation, action),
            )
            needs.append({"concept_id": row["evidence_key"], "priority": priority, "reason": reason, "action": action})
    return needs


def simulate_profile(c, batch_id: int, profile_code: str, *, created_by: int | None = None) -> dict[str, Any]:
    profile_code = _text(profile_code).upper()
    if profile_code not in SYNTHETIC_PROFILES:
        raise ValueError("Unknown synthetic learner profile.")
    batch = c.execute("SELECT * FROM mastery_lab_batches WHERE id=?", (batch_id,)).fetchone()
    if not batch:
        raise ValueError("Mastery Laboratory batch not found.")
    questions = c.execute("SELECT * FROM mastery_lab_questions WHERE batch_id=? AND active=1 ORDER BY id", (batch_id,)).fetchall()
    if not questions:
        raise ValueError("The batch contains no active sandbox candidates.")
    run_code = "MLR-" + datetime.now().strftime("%Y%m%d%H%M%S%f") + "-" + profile_code[:4]
    cur = c.execute(
        """INSERT INTO mastery_lab_runs(
          run_code,batch_id,profile_code,programme,subject,chapter,status,previous_state,release_flags_json,created_by)
          VALUES(?,?,?,?,?,?,'RUNNING','UNASSESSED',?,?)""",
        (run_code, batch_id, profile_code, batch["programme"], batch["subject"], batch["chapter"], canonical_json(LAB_RELEASE_FLAGS), created_by),
    )
    run_id = cur.lastrowid
    concepts = sorted({x for q in questions for x in _list(q["concept_ids_json"])})
    missing_concept = concepts[-1] if concepts else ""
    previous_state = "UNASSESSED"
    highest_level = ""
    combined_metrics: dict[str, Any] = {}
    final_action = ""
    final_reasons: list[str] = []
    phase_summaries: list[dict[str, Any]] = []
    first_level = ""
    for phase in _profile_phases(profile_code):
        phase_no = phase["phase_no"]
        phase_questions = _questions_for_phase(questions, profile_code, phase_no)
        response_rows = []
        for index, q in enumerate(phase_questions):
            response, confidence, response_time, manual_score = _profile_response(profile_code, q, index, phase_no, missing_concept)
            result = score_lab_response(q, response, manual_score=manual_score)
            cluster = identity_cluster(q)
            weight = identity_weight(q)
            cur = c.execute(
                """INSERT INTO mastery_lab_responses(
                  run_id,question_id,phase_no,days_from_start,response_json,awarded_marks,max_marks,is_correct,score_fraction,
                  confidence,response_time_seconds,manual_review_required,scoring_explanation,diagnostic_tags_json,
                  identity_cluster_key,identity_weight,progression_eligible)
                  VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    run_id,
                    q["id"],
                    phase_no,
                    phase["days"],
                    canonical_json(response),
                    result["awarded_marks"],
                    result["max_marks"],
                    1 if result["is_correct"] else 0,
                    result["score_fraction"],
                    confidence,
                    response_time,
                    1 if result["manual_review_required"] else 0,
                    result["scoring_explanation"],
                    canonical_json(result["diagnostic_tags"]),
                    cluster,
                    weight,
                    1 if progression_eligible(q) else 0,
                ),
            )
            response_rows.append(c.execute("SELECT * FROM mastery_lab_responses WHERE id=?", (cur.lastrowid,)).fetchone())
        metrics = calculate_phase_metrics(c, batch_id, response_rows)
        level, evaluations = highest_provisional_level(c, metrics)
        if phase_no == 1:
            first_level = level
        independent_confirmation = phase["event"] == "INDEPENDENT_CONFIRMATION" and bool(level) and _level_rank(level) >= _level_rank(first_level)
        recovery_item_only = bool(response_rows) and all(
            canonical_relation_type(c.execute("SELECT relation_type FROM mastery_lab_questions WHERE id=?", (r["question_id"],)).fetchone()["relation_type"])
            in {"scaffold", "recovery_item"}
            for r in response_rows
        )
        event_type = phase["event"]
        verification_due = event_type == "RECONFIRMATION" and phase["days"] > 60 and previous_state.startswith("VERIFIED_")
        if verification_due:
            due_state, due_action, due_reasons = transition_mastery_state(previous_state, highest_level or level, metrics, event_type="TIME_ELAPSED", verification_due=True)
            c.execute(
                """INSERT INTO mastery_lab_state_history(run_id,phase_no,days_from_start,previous_state,new_state,provisional_level,event_type,rationale_json,next_action)
                VALUES(?,?,?,?,?,?,?,?,?)""",
                (run_id, phase_no, phase["days"], previous_state, due_state, highest_level or level, "VERIFICATION_DUE", canonical_json({"reasons": due_reasons}), due_action),
            )
            previous_state = due_state
        new_state, next_action, reasons = transition_mastery_state(
            previous_state,
            level,
            metrics,
            event_type="RECONFIRMATION" if event_type == "RECONFIRMATION" else event_type,
            independent_confirmation=independent_confirmation,
            recovery_item_only=recovery_item_only,
        )
        c.execute(
            """INSERT INTO mastery_lab_state_history(run_id,phase_no,days_from_start,previous_state,new_state,provisional_level,event_type,rationale_json,next_action)
            VALUES(?,?,?,?,?,?,?,?,?)""",
            (
                run_id,
                phase_no,
                phase["days"],
                previous_state,
                new_state,
                level,
                event_type,
                canonical_json({"reasons": reasons, "evaluations": evaluations, "metrics": metrics}),
                next_action,
            ),
        )
        _write_evidence_ledger(c, run_id, phase_no, batch_id, response_rows, metrics)
        needs = _identify_recovery_needs(c, run_id, phase_no)
        phase_summaries.append(
            {
                "phase_no": phase_no,
                "event": event_type,
                "days_from_start": phase["days"],
                "state": new_state,
                "level": level,
                "metrics": metrics,
                "evaluations": evaluations,
                "recovery_needs": needs,
                "reasons": reasons,
                "next_action": next_action,
            }
        )
        previous_state = new_state
        highest_level = level or highest_level
        combined_metrics = metrics
        final_action = next_action
        final_reasons = reasons
    summary = f"{profile_code.replace('_', ' ').title()} finished in {previous_state}."
    rationale = {
        "profile": profile_code,
        "profile_description": SYNTHETIC_PROFILES[profile_code],
        "phases": phase_summaries,
        "final_reasons": final_reasons,
        "evidence_boundary": "Synthetic QA evidence only; not valid for real student mastery.",
    }
    c.execute(
        """UPDATE mastery_lab_runs SET status='COMPLETED',final_state=?,provisional_level=?,decision_summary=?,next_action=?,
          metrics_json=?,rationale_json=?,completed_at=? WHERE id=?""",
        (
            previous_state,
            highest_level,
            summary,
            final_action,
            canonical_json(combined_metrics),
            canonical_json(rationale),
            datetime.now().isoformat(timespec="seconds"),
            run_id,
        ),
    )
    c.execute(
        """INSERT INTO mastery_lab_audit_events(actor_user_id,event_type,subject_type,subject_id,metadata_json)
        VALUES(?,?,?,?,?)""",
        (created_by, "SYNTHETIC_PROFILE_REPLAYED", "mastery_lab_run", str(run_id), canonical_json({"profile": profile_code, "final_state": previous_state})),
    )
    evaluate_mastery_gate(c, batch_id, run_id)
    c.commit()
    return {
        "run_id": run_id,
        "run_code": run_code,
        "profile_code": profile_code,
        "final_state": previous_state,
        "provisional_level": highest_level,
        "decision_summary": summary,
        "next_action": final_action,
        "rationale": rationale,
    }


def evaluate_content_admission_gate(c, batch_id: int) -> dict[str, Any]:
    batch = c.execute("SELECT * FROM mastery_lab_batches WHERE id=?", (batch_id,)).fetchone()
    questions = c.execute("SELECT * FROM mastery_lab_questions WHERE batch_id=?", (batch_id,)).fetchall()
    failures: list[str] = []
    warnings: list[str] = []
    if not batch or not questions:
        failures.append("No imported candidate records are available.")
    for q in questions:
        for field, expected in LAB_RELEASE_FLAGS.items():
            if q[field] != expected:
                failures.append(f"{q['external_question_id']} has invalid {field}={q[field]!r}.")
        if q["family_type"] not in SUPPORTED_FAMILIES:
            failures.append(f"{q['external_question_id']} uses unsupported family {q['family_type']}.")
        if q["relation_type"] not in RELATION_TYPES:
            failures.append(f"{q['external_question_id']} uses unsupported relationship {q['relation_type']}.")
        warnings.extend(_list(q["warnings_json"]))
    unresolved_relations = c.execute(
        """SELECT COUNT(*) n FROM mastery_lab_question_relations
        WHERE batch_id=? AND COALESCE(related_external_question_id,'')<>'' AND related_question_id IS NULL""",
        (batch_id,),
    ).fetchone()["n"]
    if unresolved_relations:
        warnings.append(f"{unresolved_relations} relationship(s) reference candidates outside this batch.")
    status = "BLOCKED" if failures else ("PASS_WITH_WARNINGS" if warnings else "PASS")
    summary = "Content Admission blocked." if failures else "Content Admission preserved QA-only status and evidence identity."
    c.execute(
        """INSERT INTO mastery_lab_gate_results(batch_id,gate_code,gate_name,status,summary,evidence_json)
        VALUES(?,?,?,?,?,?)""",
        (batch_id, "GATE_1", "Content Admission", status, summary, canonical_json({"failures": failures, "warnings": warnings, "question_count": len(questions)})),
    )
    if failures:
        _upsert_blocker(c, batch_id=batch_id, gate_code="GATE_1", severity="CRITICAL", title="Mastery Laboratory content admission failed", description="; ".join(failures[:8]))
    return {"status": status, "summary": summary, "failures": failures, "warnings": warnings}


def evaluate_mastery_gate(c, batch_id: int, run_id: int) -> dict[str, Any]:
    run = c.execute("SELECT * FROM mastery_lab_runs WHERE id=?", (run_id,)).fetchone()
    failures: list[str] = []
    if not run or run["status"] != "COMPLETED":
        failures.append("Synthetic run did not complete.")
    if run and run["final_state"] not in MASTERY_STATES:
        failures.append("Run produced an unknown mastery state.")
    if run and not run["rationale_json"]:
        failures.append("Run has no explainability rationale.")
    # Hard boundary: laboratory execution must not create live attempts/mastery.
    live_counts = {
        "attempts": c.execute("SELECT COUNT(*) n FROM attempts WHERE assessment_kind='mastery_lab' OR scope='mastery_lab'").fetchone()["n"],
        "mastery_records": c.execute("SELECT COUNT(*) n FROM mastery_records WHERE source='Mastery Laboratory'").fetchone()["n"],
    }
    if any(live_counts.values()):
        failures.append("Laboratory evidence leaked into live attempts or mastery records.")
    status = "BLOCKED" if failures else "PASS"
    summary = "Mastery and Study Plan gate blocked." if failures else "Synthetic mastery remained provisional, explainable and isolated from live records."
    c.execute(
        """INSERT INTO mastery_lab_gate_results(batch_id,run_id,gate_code,gate_name,status,summary,evidence_json)
        VALUES(?,?,?,?,?,?,?)""",
        (batch_id, run_id, "GATE_3", "Mastery and Study Plan", status, summary, canonical_json({"failures": failures, "live_counts": live_counts, "final_state": run["final_state"] if run else ""})),
    )
    if failures:
        _upsert_blocker(c, batch_id=batch_id, run_id=run_id, gate_code="GATE_3", severity="CRITICAL", title="Laboratory mastery isolation failed", description="; ".join(failures))
    return {"status": status, "summary": summary, "failures": failures}


def evaluate_assessment_execution_gate(c, batch_id: int) -> dict[str, Any]:
    questions = c.execute("SELECT * FROM mastery_lab_questions WHERE batch_id=?", (batch_id,)).fetchall()
    failures: list[str] = []
    families_tested: dict[str, int] = {}
    for q in questions:
        try:
            correct = score_lab_response(q, _correct_response_for(q), manual_score=float(safe_json(q["marking_config_json"], {}).get("marks") or 1) if q["family_type"] == "constructed_response" else None)
            incorrect = score_lab_response(q, _incorrect_response_for(q), manual_score=0.0 if q["family_type"] == "constructed_response" else None)
            if not correct["is_correct"] and not correct["manual_review_required"]:
                failures.append(f"{q['external_question_id']} did not recognise its governed correct response.")
            if incorrect["is_correct"] and q["family_type"] != "constructed_response":
                failures.append(f"{q['external_question_id']} accepted the generated incorrect response.")
            families_tested[q["family_type"]] = families_tested.get(q["family_type"], 0) + 1
        except Exception as exc:
            failures.append(f"{q['external_question_id']} scoring error: {exc}")
    status = "BLOCKED" if failures else "PASS"
    summary = "Assessment Execution blocked." if failures else "All imported candidate families passed deterministic correct/incorrect scoring challenges."
    c.execute(
        """INSERT INTO mastery_lab_gate_results(batch_id,gate_code,gate_name,status,summary,evidence_json)
        VALUES(?,?,?,?,?,?)""",
        (batch_id, "GATE_2", "Assessment Execution", status, summary, canonical_json({"failures": failures, "families_tested": families_tested})),
    )
    if failures:
        _upsert_blocker(c, batch_id=batch_id, gate_code="GATE_2", severity="CRITICAL", title="Candidate-family scoring challenge failed", description="; ".join(failures[:8]))
    c.commit()
    return {"status": status, "summary": summary, "failures": failures, "families_tested": families_tested}


def evaluate_security_isolation_gate(c, batch_id: int) -> dict[str, Any]:
    failures: list[str] = []
    live_question_matches = c.execute(
        """SELECT COUNT(*) n FROM questions q JOIN mastery_lab_questions mlq
        ON q.question_id=mlq.external_question_id WHERE mlq.batch_id=?""",
        (batch_id,),
    ).fetchone()["n"]
    if live_question_matches:
        failures.append("One or more sandbox external IDs already exist in the live question bank; this requires explicit admission reconciliation.")
    invalid_flags = c.execute(
        """SELECT COUNT(*) n FROM mastery_lab_questions WHERE batch_id=? AND (
          content_environment<>'QA_SANDBOX_ONLY' OR student_release_status<>'NOT_STUDENT_RELEASED' OR
          bank_approval_status<>'NOT_BANK_APPROVED' OR mastery_validity<>'NOT_VALID_FOR_REAL_MASTERY')""",
        (batch_id,),
    ).fetchone()["n"]
    if invalid_flags:
        failures.append(f"{invalid_flags} sandbox record(s) have invalid release flags.")
    status = "BLOCKED" if failures else "PASS"
    summary = "Security, Privacy and Isolation blocked." if failures else "The corpus remains in separate QA-only tables with no real-student or live-bank eligibility."
    c.execute(
        """INSERT INTO mastery_lab_gate_results(batch_id,gate_code,gate_name,status,summary,evidence_json)
        VALUES(?,?,?,?,?,?)""",
        (batch_id, "GATE_4", "Security, Privacy and Isolation", status, summary, canonical_json({"failures": failures, "live_question_matches": live_question_matches, "invalid_flags": invalid_flags})),
    )
    if failures:
        _upsert_blocker(c, batch_id=batch_id, gate_code="GATE_4", severity="HIGH", title="Mastery Laboratory isolation warning", description="; ".join(failures))
    c.commit()
    return {"status": status, "summary": summary, "failures": failures}


def evaluate_release_acceptance_gate(c, batch_id: int) -> dict[str, Any]:
    required = {"GATE_1", "GATE_2", "GATE_3", "GATE_4"}
    latest: dict[str, str] = {}
    for row in c.execute(
        """SELECT * FROM mastery_lab_gate_results WHERE batch_id=? ORDER BY id""", (batch_id,)
    ).fetchall():
        latest[row["gate_code"]] = row["status"]
    missing = sorted(required - set(latest))
    blocked = sorted(code for code, status in latest.items() if code in required and status == "BLOCKED")
    open_blockers = c.execute("SELECT COUNT(*) n FROM mastery_lab_blockers WHERE batch_id=? AND status='OPEN'", (batch_id,)).fetchone()["n"]
    failures = []
    if missing:
        failures.append("Missing gate evidence: " + ", ".join(missing))
    if blocked:
        failures.append("Blocked gates: " + ", ".join(blocked))
    if open_blockers:
        failures.append(f"{open_blockers} open blocker(s) remain.")
    status = "BLOCKED" if failures else "PASS"
    summary = "Release Acceptance blocked." if failures else "The QA sandbox milestone is technically accepted; candidates remain barred from student release and real mastery."
    c.execute(
        """INSERT INTO mastery_lab_gate_results(batch_id,gate_code,gate_name,status,summary,evidence_json)
        VALUES(?,?,?,?,?,?)""",
        (batch_id, "GATE_5", "Release Acceptance", status, summary, canonical_json({"failures": failures, "latest_gates": latest, "open_blockers": open_blockers})),
    )
    c.commit()
    return {"status": status, "summary": summary, "failures": failures, "latest_gates": latest}


def _upsert_blocker(c, *, batch_id: int | None = None, run_id: int | None = None, gate_code: str, severity: str, title: str, description: str) -> None:
    code = "MLB-" + checksum({"batch_id": batch_id, "run_id": run_id, "gate_code": gate_code, "title": title})[:16].upper()
    c.execute(
        """INSERT INTO mastery_lab_blockers(blocker_code,batch_id,run_id,gate_code,severity,title,description,status)
        VALUES(?,?,?,?,?,?,?,'OPEN')
        ON CONFLICT(blocker_code) DO UPDATE SET description=excluded.description,severity=excluded.severity,
          status=CASE WHEN mastery_lab_blockers.status='RESOLVED' THEN mastery_lab_blockers.status ELSE 'OPEN' END""",
        (code, batch_id, run_id, gate_code, severity, title, description),
    )


def evaluate_all_gates(c, batch_id: int, run_id: int | None = None) -> dict[str, Any]:
    results = {
        "GATE_1": evaluate_content_admission_gate(c, batch_id),
        "GATE_2": evaluate_assessment_execution_gate(c, batch_id),
        "GATE_4": evaluate_security_isolation_gate(c, batch_id),
    }
    if run_id:
        results["GATE_3"] = evaluate_mastery_gate(c, batch_id, run_id)
    results["GATE_5"] = evaluate_release_acceptance_gate(c, batch_id)
    return results


def batch_summary(c, batch_id: int) -> dict[str, Any] | None:
    batch = c.execute("SELECT * FROM mastery_lab_batches WHERE id=?", (batch_id,)).fetchone()
    if not batch:
        return None
    questions = c.execute("SELECT * FROM mastery_lab_questions WHERE batch_id=? ORDER BY id", (batch_id,)).fetchall()
    family_counts = c.execute(
        "SELECT family_type,COUNT(*) n FROM mastery_lab_questions WHERE batch_id=? GROUP BY family_type ORDER BY family_type",
        (batch_id,),
    ).fetchall()
    relation_counts = c.execute(
        "SELECT relation_type,COUNT(*) n FROM mastery_lab_questions WHERE batch_id=? GROUP BY relation_type ORDER BY relation_type",
        (batch_id,),
    ).fetchall()
    gates = c.execute(
        """SELECT g.* FROM mastery_lab_gate_results g JOIN (
          SELECT gate_code,MAX(id) id FROM mastery_lab_gate_results WHERE batch_id=? GROUP BY gate_code
        ) x ON x.id=g.id ORDER BY g.gate_code""",
        (batch_id,),
    ).fetchall()
    runs = c.execute("SELECT * FROM mastery_lab_runs WHERE batch_id=? ORDER BY id DESC", (batch_id,)).fetchall()
    blockers = c.execute("SELECT * FROM mastery_lab_blockers WHERE batch_id=? ORDER BY CASE severity WHEN 'CRITICAL' THEN 1 WHEN 'HIGH' THEN 2 ELSE 3 END,id DESC", (batch_id,)).fetchall()
    return {
        "batch": batch,
        "questions": questions,
        "family_counts": family_counts,
        "relation_counts": relation_counts,
        "gates": gates,
        "runs": runs,
        "blockers": blockers,
    }


def run_summary(c, run_id: int) -> dict[str, Any] | None:
    run = c.execute("SELECT * FROM mastery_lab_runs WHERE id=?", (run_id,)).fetchone()
    if not run:
        return None
    states = c.execute("SELECT * FROM mastery_lab_state_history WHERE run_id=? ORDER BY id", (run_id,)).fetchall()
    evidence = c.execute("SELECT * FROM mastery_lab_evidence WHERE run_id=? ORDER BY phase_no,evidence_type,accuracy,evidence_key", (run_id,)).fetchall()
    recovery = c.execute("SELECT * FROM mastery_lab_recovery_needs WHERE run_id=? ORDER BY CASE priority WHEN 'HIGH' THEN 1 ELSE 2 END,id", (run_id,)).fetchall()
    responses = c.execute(
        """SELECT r.*,q.external_question_id,q.family_type,q.relation_type,q.mastery_level,q.question_text
        FROM mastery_lab_responses r JOIN mastery_lab_questions q ON q.id=r.question_id
        WHERE r.run_id=? ORDER BY r.phase_no,r.id""",
        (run_id,),
    ).fetchall()
    return {"run": run, "states": states, "evidence": evidence, "recovery": recovery, "responses": responses}


def sample_candidate_corpus() -> dict[str, Any]:
    """Return a small, non-live technical corpus spanning the laboratory families."""
    base = {
        "programme": "FSc Part 1",
        "subject": "Biology",
        "chapter": "Chapter 1 QA Laboratory",
        "learning_outcome_ids": ["LO-QA-1"],
        "concept_ids": ["CONCEPT-QA-A"],
        "source_lineage": {"source": "TECHNICAL_SAMPLE_ONLY", "approved_for_students": False},
    }
    questions = [
        {**base, "question_id": "LAB-MCQ-1", "family_type": "standard_mcq", "relation_type": "independent_seed", "seed_key": "SEED-A", "mastery_level": "Foundation", "cognitive_demand": "Recall", "question": "Technical sample single-choice item.", "options": [{"id": "A", "text": "Correct"}, {"id": "B", "text": "Incorrect"}], "marking_config": {"marks": 1, "correct_option_ids": ["A"]}},
        {**base, "question_id": "LAB-4S-1", "family_type": "four_statement_selection", "relation_type": "true_variant", "seed_key": "SEED-A", "parent_question_id": "LAB-MCQ-1", "mastery_level": "Foundation", "cognitive_demand": "Understanding", "question": "Technical sample four-statement selection.", "options": [{"id": "A", "text": "I only"}, {"id": "B", "text": "II only"}], "marking_config": {"marks": 1, "correct_option_ids": ["A"]}},
        {**base, "question_id": "LAB-TF-1", "family_type": "true_false", "relation_type": "scaffold", "seed_key": "SEED-A", "parent_question_id": "LAB-MCQ-1", "mastery_level": "Foundation", "question": "Technical sample true/false.", "options": [{"id": "A", "text": "True"}, {"id": "B", "text": "False"}], "marking_config": {"marks": 1, "correct_option_ids": ["A"]}},
        {**base, "question_id": "LAB-CLOZE-1", "family_type": "cloze", "relation_type": "independent_seed", "seed_key": "SEED-B", "mastery_level": "Exam Ready", "cognitive_demand": "Application", "question": "Technical sample cloze.", "answer_config": {"blanks": [{"accepted_answers": ["alpha"]}, {"accepted_answers": ["beta"]}]}, "marking_config": {"marks": 2, "partial_credit": True}},
        {**base, "question_id": "LAB-MATCH-1", "family_type": "matching", "relation_type": "shared_stimulus_pair", "seed_key": "SEED-C", "stimulus_group_key": "STIM-1", "mastery_level": "Exam Ready", "question": "Technical sample matching.", "marking_config": {"marks": 2, "correct_pairs": {"1": "A", "2": "B"}, "partial_credit": True}},
        {**base, "question_id": "LAB-ORDER-1", "family_type": "ordering", "relation_type": "shared_stimulus_pair", "seed_key": "SEED-D", "stimulus_group_key": "STIM-1", "mastery_level": "Exam Ready", "question": "Technical sample ordering.", "marking_config": {"marks": 2, "correct_order": ["A", "B", "C"], "partial_credit": True}},
        {**base, "question_id": "LAB-MR-1", "family_type": "multiple_response", "relation_type": "independent_seed", "seed_key": "SEED-E", "mastery_level": "Advanced", "cognitive_demand": "Analysis", "question": "Technical sample multiple response.", "options": [{"id": "A", "text": "Correct one"}, {"id": "B", "text": "Correct two"}, {"id": "C", "text": "Incorrect"}], "marking_config": {"marks": 2, "correct_option_ids": ["A", "B"], "partial_credit": True}},
        {**base, "question_id": "LAB-NUM-1", "family_type": "numerical_interpretation", "relation_type": "independent_seed", "seed_key": "SEED-F", "mastery_level": "Advanced", "cognitive_demand": "Application", "question": "Technical sample numerical interpretation.", "marking_config": {"marks": 2, "correct_value": 10, "tolerance": 0.1}},
        {**base, "question_id": "LAB-DATA-1", "family_type": "diagram_data_stimulus", "response_mode": "single_choice", "relation_type": "integrated_question", "seed_key": "SEED-G", "mastery_level": "Distinction", "mastery_ceiling": "Distinction", "cognitive_demand": "Integrated unfamiliar evidence", "question": "Technical sample data-stimulus interpretation.", "stimulus": {"type": "table", "data": [[1, 2], [3, 4]]}, "options": [{"id": "A", "text": "Correct"}, {"id": "B", "text": "Incorrect"}], "marking_config": {"marks": 2, "response_mode": "single_choice", "correct_option_ids": ["A"]}},
        {**base, "question_id": "LAB-CR-1", "family_type": "constructed_response", "relation_type": "integrated_question", "seed_key": "SEED-H", "mastery_level": "Distinction", "cognitive_demand": "Evaluation", "question": "Technical sample constructed response.", "marking_config": {"marks": 3, "rubric_points": [{"marks": 1, "keywords": ["alpha"]}, {"marks": 1, "keywords": ["beta"]}, {"marks": 1, "keywords": ["gamma"]}]}},
        {**base, "question_id": "LAB-MIS-1", "family_type": "misconception_probe", "response_mode": "single_choice", "relation_type": "recovery_item", "seed_key": "SEED-A", "parent_question_id": "LAB-MCQ-1", "mastery_level": "Foundation", "question": "Technical sample misconception probe.", "options": [{"id": "A", "text": "Correct"}, {"id": "B", "text": "Misconception"}], "misconception_tags": ["MIS-QA-1"], "marking_config": {"marks": 1, "response_mode": "single_choice", "correct_option_ids": ["A"], "misconception_by_response": {"B": "MIS-QA-1"}}},
        {**base, "question_id": "LAB-REC-1", "family_type": "adaptive_recovery", "response_mode": "single_choice", "relation_type": "reconfirmation_item", "seed_key": "SEED-I", "mastery_level": "Advanced", "cognitive_demand": "Application", "question": "Technical sample reconfirmation item.", "options": [{"id": "A", "text": "Correct"}, {"id": "B", "text": "Incorrect"}], "marking_config": {"marks": 1, "response_mode": "single_choice", "correct_option_ids": ["A"]}},
        {**base, "question_id": "LAB-DIST-2", "family_type": "standard_mcq", "relation_type": "independent_seed", "seed_key": "SEED-J", "learning_outcome_ids": ["LO-QA-2"], "concept_ids": ["CONCEPT-QA-B"], "mastery_level": "Distinction", "mastery_ceiling": "Distinction", "cognitive_demand": "Analysis of unfamiliar evidence", "question": "Technical sample Distinction item two.", "options": [{"id": "A", "text": "Correct"}, {"id": "B", "text": "Incorrect"}], "marking_config": {"marks": 1, "correct_option_ids": ["A"]}},
        {**base, "question_id": "LAB-DIST-3", "family_type": "numerical_interpretation", "relation_type": "independent_seed", "seed_key": "SEED-K", "learning_outcome_ids": ["LO-QA-2"], "concept_ids": ["CONCEPT-QA-B"], "mastery_level": "Distinction", "mastery_ceiling": "Distinction", "cognitive_demand": "Integrated application", "question": "Technical sample Distinction numerical item.", "marking_config": {"marks": 2, "correct_value": 20, "tolerance": 0.1}},
        {**base, "question_id": "LAB-DIST-4", "family_type": "ordering", "relation_type": "independent_seed", "seed_key": "SEED-L", "learning_outcome_ids": ["LO-QA-3"], "concept_ids": ["CONCEPT-QA-C"], "mastery_level": "Distinction", "mastery_ceiling": "Distinction", "cognitive_demand": "Evaluation and integration", "question": "Technical sample Distinction ordering item.", "marking_config": {"marks": 2, "correct_order": ["A", "B", "C"], "partial_credit": False}},
    ]
    return {"metadata": {"technical_sample_only": True}, "questions": questions}
