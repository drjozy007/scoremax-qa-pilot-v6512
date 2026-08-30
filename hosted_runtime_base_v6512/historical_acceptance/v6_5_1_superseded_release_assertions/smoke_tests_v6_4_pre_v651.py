"""ScoreMax V6.4.0 Live Pilot UX & Operations deterministic checks.

Uses the established lightweight framework stubs so core schema/logic tests can run in
restricted build environments. Real-browser and hosted-domain acceptance remain separate gates.
"""
from __future__ import annotations
import io, json, os, tempfile, time
from pathlib import Path
from openpyxl import Workbook, load_workbook
from smoke_tests_v5_5 import install_framework_stubs
flask_stub, request_stub = install_framework_stubs()
ROOT=Path(__file__).resolve().parent
TMP=Path(tempfile.mkdtemp(prefix='scoremax_v640_smoke_'))
os.environ['SCOREMAX_DB']=str(TMP/'scoremax.db')
os.environ['SCOREMAX_BACKUP_DIR']=str(TMP/'backups')
os.environ['SCOREMAX_CONTENT_INTAKE_DIR']=str(TMP/'intake')
os.environ['SCOREMAX_UNIVERSAL_MASTERY']='1'
import app

n=0
def ok(name,condition):
    global n
    if not condition: raise AssertionError(name)
    n+=1; print('PASS:',name)

app.init(); c=app.db()
# Current release/migration identity.
health=app.healthz()[0]
ok('health keeps immutable parent compatibility marker and exposes V6.4.0 release identity',health['version']=='6.2.8.1' and health['release_version'] in {'6.4.0','6.5.0'} and health['build_name'] in {'Live Pilot UX & Operations','Three-System Integration Adapter V1'})
cols={r['name'] for r in c.execute('PRAGMA table_info(users)').fetchall()}
ok('V6.4 migration adds persistent active programme context','active_programme' in cols)
ok('V6.4 migration adds referral rule versioning and one-level override fields',all(x in {r['name'] for r in c.execute('PRAGMA table_info(referral_rewards)').fetchall()} for x in ['rule_version','override_referrer_user_id','override_reward_rate','override_reward_amount_minor','override_rule_version']))
ok('V6.4 migration adds governed emergency-release state',all(x in {r['name'] for r in c.execute('PRAGMA table_info(content_import_batches)').fetchall()} for x in ['intake_mode','release_status','release_attested_at','released_count']))

# Persistent programme selector: visible even when a programme has no released questions.
c.execute("""INSERT INTO users(system_user_id,role,full_name,email,username,academic_level,active_programme,subjects,profile_completed)
 VALUES('STU-V640-1','student','V640 Student','v640@example.test','v640-student','FSc Part 1','FSc Part 1','Biology,Chemistry,Physics',1)""")
sid=c.execute("SELECT id FROM users WHERE system_user_id='STU-V640-1'").fetchone()['id']; c.commit()
opts=app.student_programme_options(c,sid)
ok('student programme selector always exposes FSc 1 FSc 2 and MDCAT',[x['label'] for x in opts]==['FSc 1','FSc 2','MDCAT'])
ok('student programme selector starts in the learner selected programme',next(x for x in opts if x['active'])['value']=='FSc Part 1')
app.session.clear(); app.session.update({'user_id':sid,'role':'student'})
request_stub.form={'programme_code':'fsc2','return_to':'/student/subjects'}; request_stub.referrer=''
res=app.student_programme_switch()
ok('programme switch persists FSc 2 rather than being a decorative tab',app.student_programme(c,sid)=='FSc Part 2' and res=='/student/subjects')
request_stub.form={'programme_code':'mdcat','return_to':'/student'}; app.student_programme_switch()
ok('programme switch persists MDCAT independently of original registration programme',app.student_programme(c,sid)=='MDCAT')

# New-student mastery hero never fabricates Foundation mastery.
hero=app.student_mastery_hero(c,sid,'MDCAT')
ok('new student mastery hero says starting point instead of fabricating Foundation',hero['established'] is False and hero['current_level']=='Not established' and hero['status']=='Starting point')

# Teacher referral architecture: code at creation, immutable attribution, direct + one upstream teacher only.
def add_user(system_id,role,name,email):
    cur=c.execute("INSERT INTO users(system_user_id,role,full_name,email,username,password_hash,account_status) VALUES(?,?,?,?,?,'x','active')",
                  (system_id,role,name,email,system_id.lower()))
    return cur.lastrowid
teacher_a=add_user('TCH-A','teacher','Teacher A','a@teacher.test')
code_a=app.ensure_referral_code(c,teacher_a)
teacher_b=add_user('TCH-B','teacher','Teacher B','b@teacher.test')
app.apply_referral_attribution(c,teacher_b,code_a,'teacher_referral'); code_b=app.ensure_referral_code(c,teacher_b)
student_b=add_user('STU-B','student','Paying Student B','studentb@test')
owner=app.apply_referral_attribution(c,student_b,code_b,'teacher_referral')
other=add_user('TCH-X','teacher','Teacher X','x@teacher.test'); code_x=app.ensure_referral_code(c,other)
owner2=app.apply_referral_attribution(c,student_b,code_x,'teacher_referral')
ok('first valid teacher attribution is immutable',owner==teacher_b and owner2==teacher_b and c.execute('SELECT referrer_id FROM referral_attributions WHERE user_id=?',(student_b,)).fetchone()['referrer_id']==teacher_b)
ok('teacher recruited by teacher is explicitly marked as teacher recruitment',c.execute('SELECT attribution_kind FROM referral_attributions WHERE user_id=?',(teacher_b,)).fetchone()['attribution_kind']=='TEACHER_RECRUITMENT')
# Founder-configurable rates; no hard-coded commercial promise.
c.execute("UPDATE referral_programs SET reward_rate=.10,programme_version=2 WHERE role_group='teacher_direct'")
c.execute("UPDATE referral_programs SET reward_rate=.02,programme_version=3 WHERE role_group='teacher_override'")
plan=c.execute("SELECT id FROM plans WHERE code NOT LIKE 'free%' ORDER BY id LIMIT 1").fetchone()
if not plan: plan=c.execute('SELECT id FROM plans ORDER BY id LIMIT 1').fetchone()
tx=app.record_payment(c,student_b,plan['id'],100000,currency='PKR',status='successful',provider='smoke')
r=c.execute('SELECT * FROM referral_rewards WHERE payment_transaction_id=?',(tx,)).fetchone()
ok('paying student creates direct teacher commission linked to the exact payment',r and r['referrer_user_id']==teacher_b and r['reward_amount_minor']==10000 and abs(float(r['reward_rate'])-.10)<1e-9)
ok('teacher A receives only the one-level teacher-recruitment override',r['override_referrer_user_id']==teacher_a and r['override_reward_amount_minor']==2000 and abs(float(r['override_reward_rate'])-.02)<1e-9)
ok('direct and override rewards retain the rule versions used at transaction time',r['rule_version']==2 and r['override_rule_version']==3)
# Build A->B->C->student and prove A is not a third-level beneficiary.
teacher_c=add_user('TCH-C','teacher','Teacher C','c@teacher.test'); app.apply_referral_attribution(c,teacher_c,code_b,'teacher_referral'); code_c=app.ensure_referral_code(c,teacher_c)
student_c=add_user('STU-C','student','Paying Student C','studentc@test'); app.apply_referral_attribution(c,student_c,code_c,'teacher_referral')
tx2=app.record_payment(c,student_c,plan['id'],50000,currency='PKR',status='successful',provider='smoke')
r2=c.execute('SELECT * FROM referral_rewards WHERE payment_transaction_id=?',(tx2,)).fetchone()
ok('teacher referral chain is capped at one upstream teacher level',r2['referrer_user_id']==teacher_c and r2['override_referrer_user_id']==teacher_b and r2['override_referrer_user_id']!=teacher_a)

# Paid referral attribution must never be lost merely because founder commission rates are not configured yet.
c.execute("UPDATE referral_programs SET reward_rate=0,programme_version=4 WHERE role_group='teacher_direct'")
c.execute("UPDATE referral_programs SET reward_rate=0,programme_version=5 WHERE role_group='teacher_override'")
teacher_d=add_user('TCH-D','teacher','Teacher D','d@teacher.test'); app.apply_referral_attribution(c,teacher_d,code_a,'teacher_referral'); code_d=app.ensure_referral_code(c,teacher_d)
student_d=add_user('STU-D','student','Paying Student D','studentd@test'); app.apply_referral_attribution(c,student_d,code_d,'teacher_referral')
tx3=app.record_payment(c,student_d,plan['id'],75000,currency='PKR',status='successful',provider='smoke')
r3=c.execute('SELECT * FROM referral_rewards WHERE payment_transaction_id=?',(tx3,)).fetchone()
ok('paid teacher referral remains in the reward ledger when direct rate is not configured',r3 and r3['referrer_user_id']==teacher_d and r3['status']=='rate_not_configured' and int(r3['reward_amount_minor'] or 0)==0 and r3['rule_version']==4)
ok('one-level teacher-network attribution is retained even when override rate is not configured',r3['override_referrer_user_id']==teacher_a and r3['override_status']=='rate_not_configured' and int(r3['override_reward_amount_minor'] or 0)==0 and r3['override_rule_version']==5)

# ScoreMax records commercial facts; Growth Engine receives asynchronous outbox events only.
events=c.execute("SELECT event_type,payload_json,status FROM universal_growth_event_outbox WHERE event_type IN ('PAYMENT_RECORDED','TEACHER_REFERRAL_CONVERSION') ORDER BY occurred_at,event_id").fetchall()
types=[e['event_type'] for e in events]
ok('successful payments emit PAYMENT_RECORDED to the governed Growth Engine outbox','PAYMENT_RECORDED' in types and all(e['status']=='PENDING' for e in events))
ref_event=None; ref_payload={}
for e in events:
    if e['event_type']!='TEACHER_REFERRAL_CONVERSION': continue
    payload_e=json.loads(e['payload_json'])
    if payload_e.get('payment_transaction_id')==tx3:
        ref_event=e; ref_payload=payload_e; break
ok('teacher referral conversion event preserves payment direct teacher upstream teacher and eligible amount facts',ref_event is not None and ref_payload.get('direct_referrer_user_id')==teacher_d and ref_payload.get('override_referrer_user_id')==teacher_a and ref_payload.get('qualifying_amount_minor')==75000)

# Restore configured rates for the monthly workbook evidence below.
c.execute("UPDATE referral_programs SET reward_rate=.10,programme_version=6 WHERE role_group='teacher_direct'")
c.execute("UPDATE referral_programs SET reward_rate=.02,programme_version=7 WHERE role_group='teacher_override'"); c.commit()

# Monthly workbook: 3 required sheets and spreadsheet-formula hardening.
c.execute("UPDATE users SET full_name='=HYPERLINK(\"https://evil.test\",\"x\")' WHERE id=?",(teacher_b,)); c.commit()
app.session.clear(); admin=c.execute("SELECT id FROM users WHERE role='admin' ORDER BY id LIMIT 1").fetchone(); app.session.update({'user_id':admin['id'],'role':'admin'})
paid_at=c.execute('SELECT paid_at FROM payment_transactions WHERE id=?',(tx,)).fetchone()['paid_at']; request_stub.args={'month':str(paid_at)[:7]}
export_result=app.admin_referrals_export(); out=export_result[1][0] if isinstance(export_result,tuple) and len(export_result)>1 else export_result[0]
# stub send_file returns ('file', args, kwargs)
if isinstance(export_result,tuple) and export_result and export_result[0]=='file': out=export_result[1][0]
out.seek(0); wb=load_workbook(out,data_only=False)
ok('monthly teacher payout workbook contains summary detail and teacher-network sheets',wb.sheetnames==['Teacher Monthly Summary','Student Referral Detail','Teacher-to-Teacher Rewards'])
summary_values=[cell.value for row in wb['Teacher Monthly Summary'].iter_rows() for cell in row]
ok('teacher payout workbook neutralizes formula-leading user text',any(isinstance(v,str) and v.startswith("'=HYPERLINK") for v in summary_values))
summary_headers=[x.value for x in wb['Teacher Monthly Summary'][1]]
ok('teacher monthly summary separates direct downstream and network sales and rewards',all(x in summary_headers for x in ['Direct Paying Students','Downstream Paying Students','Total Paying Students','Direct Gross Sales','Direct Eligible Sales','Network Eligible Sales','Direct Commission','Teacher Referral Override','Total Due','Balance']))
headers=[x.value for x in wb['Student Referral Detail'][1]]
ok('student referral detail includes package programme payment amount direct and override reward evidence',all(x in headers for x in ['Teacher','Student ID','Package','Programme','Payment Date','Gross Amount','Eligible Amount','Direct Rate','Direct Reward','Direct Status','Original Teacher Override','Override Rate','Override Reward','Override Status','Refund']))

# Emergency Direct Intake: source-preserving, Draft/inactive first, conservative eligible-release fence.
def make_small_csv():
    headers=['Question ID','Family ID','Programme','Subject','Chapter','Type','Question','A','B','C','D','Answer','Explanation','Level','Difficulty','Rights Status','ScoreMax Ready','Status','R2 Status','Reviewer 2 Required','Assessment Purpose']
    rows=[
      ['ED-1','EDF-1','FSc Part 1','Biology','Cell Biology','MCQ','Safe approved question?','Yes','No','Maybe','Never','A','Safe explanation','Foundation','Moderate','ScoreMax Original','Yes','Approved','','No','practice|test|mastery'],
      ['ED-2','EDF-2','FSc Part 1','Biology','Cell Biology','MCQ','Held question?','Yes','No','Maybe','Never','A','Held explanation','Foundation','Moderate','ScoreMax Original','Yes','HOLD','','No','practice|test|mastery'],
      ['ED-3','EDF-3','FSc Part 1','Biology','Cell Biology','MCQ','R2 question?','Yes','No','Maybe','Never','A','R2 explanation','Foundation','Moderate','ScoreMax Original','Yes','Approved','','Yes','practice|test|mastery'],
    ]
    import csv as _csv
    b=io.StringIO(); w=_csv.writer(b); w.writerow(headers); w.writerows(rows); return b.getvalue().encode()
class Upload:
    def __init__(self,name,payload): self.filename=name; self._payload=payload
    def read(self): return self._payload
payload=make_small_csv(); request_stub.method='POST'; request_stub.files={'file':Upload('approved_emergency.csv',payload)}
request_stub.form={'intake_mode':'EMERGENCY_DIRECT','source_system':'EMERGENCY_APPROVED_WORKBOOK','source_prompt_pack_id':'PH-REL-SMOKE','source_prompt_pack_version':'1'}; request_stub.args={'mode':'emergency'}
app.admin_import(); batch=c.execute("SELECT * FROM content_import_batches WHERE source_prompt_pack_id='PH-REL-SMOKE' ORDER BY id DESC LIMIT 1").fetchone()
ok('Emergency Direct Intake preview preserves intake mode original checksum and source file',batch and batch['intake_mode']=='EMERGENCY_DIRECT' and batch['row_count']==3 and Path(batch['source_file_path']).exists() and batch['payload_checksum']==__import__('hashlib').sha256(payload).hexdigest())
source_rows=c.execute('SELECT * FROM content_import_batch_rows WHERE batch_id=? ORDER BY row_number',(batch['id'],)).fetchall()
ok('Emergency Direct Intake preserves source row lineage before import',[x['row_number'] for x in source_rows]==[2,3,4])
request_stub.form={'batch_id':str(batch['id'])}; app.admin_import_confirm();
imported=c.execute('SELECT id,status,review_status,active,feedback_config FROM questions WHERE source_import_batch_id=? ORDER BY question_id',(batch['id'],)).fetchall()
ok('Emergency Direct Intake imports all valid records as Draft and inactive before release',len(imported)==3 and all(x['status']=='Draft' and x['review_status']=='Draft' and x['active']==0 for x in imported))
lineage=json.loads(imported[0]['feedback_config'])
ok('imported question retains source worksheet and source row lineage',lineage['source_sheet']=='CSV' and int(lineage['source_row'])==2)
request_stub.form={'attestation':'I CONFIRM THIS IS A FROZEN ACADEMICALLY APPROVED RELEASE','release_note':'smoke controlled release'}; app.admin_import_release_eligible(batch['id'])
qrows=c.execute("SELECT question_id,status,review_status,active,family_key FROM questions WHERE source_import_batch_id=? ORDER BY question_id",(batch['id'],)).fetchall(); by={x['question_id']:x for x in qrows}
ok('emergency release activates only explicitly ready rights-cleared non-held non-R2 rows',by['ED-1']['active']==1 and by['ED-2']['active']==0 and by['ED-3']['active']==0)
ok('emergency release also opens the governed family gate for the eligible question',c.execute("SELECT review_status,active FROM question_families WHERE family_key=?",(by['ED-1']['family_key'],)).fetchone()['active']==1)
ok('held and Reviewer-2-required content remains excluded from learner inventory',c.execute(f"SELECT COUNT(*) n FROM questions q WHERE q.source_import_batch_id=? AND {app.live_question_clause('q')}",(batch['id'],)).fetchone()['n']==1)

# 3,000-row heterogeneous XLSX preflight: question-bearing sheet detection, instruction sheet ignored, canonical headers recognized.
def make_3000_xlsx():
    wb=Workbook(); intro=wb.active; intro.title='Instructions'; intro.append(['Read me']); intro.append(['Do not import this sheet'])
    ws=wb.create_sheet('Question Bank')
    ws.append(['Question ID','Family ID','Programme','Subject','Chapter','Question / Task','Statements / Options','Key Answer','Explanation / Marking Rubric','Mastery','Difficulty','Rights Status','ScoreMax Ready'])
    for i in range(3000):
        ws.append([f'Q3000-{i+1}',f'F3000-{i+1}','FSc Part 1','Chemistry','Chapter 1',f'Question {i+1}?','A. One\nB. Two\nC. Three\nD. Four','A','Because A is correct.','Foundation','Moderate','ScoreMax Original','Yes'])
    b=io.BytesIO(); wb.save(b); return b.getvalue()
xlsx=make_3000_xlsx(); t=time.perf_counter(); fn,ft,raw,rows=app._read_import_upload(Upload('powerhouse_3000.xlsx',xlsx)); elapsed=time.perf_counter()-t
ok('3,000-row emergency workbook parser reads exactly the question-bearing sheet',ft=='XLSX' and len(rows)==3000 and all(r.get('_sheet')=='Question Bank' for r in rows[:5]))
ok('Power House-style V3 headers map to canonical question answer rubric and mastery fields',rows[0]['Question']=='Question 1?' and rows[0]['Answer']=='A' and rows[0]['Explanation']=='Because A is correct.' and rows[0]['Level']=='Foundation')
ok('combined Statements / Options are deterministically split into A-D options',all(rows[0].get(x) for x in 'ABCD'))
print(f'INFO: 3000-row XLSX parse seconds: {elapsed:.3f}')

# Learner/public UX source checks: these are intentional requirements, not screenshot acceptance.
base=(ROOT/'templates/base.html').read_text(encoding='utf-8'); home=(ROOT/'templates/student.html').read_text(encoding='utf-8'); landing=(ROOT/'templates/index.html').read_text(encoding='utf-8'); css=(ROOT/'static/styles.css').read_text(encoding='utf-8'); login=(ROOT/'templates/login.html').read_text(encoding='utf-8'); register=(ROOT/'templates/register.html').read_text(encoding='utf-8'); intake=(ROOT/'templates/import.html').read_text(encoding='utf-8')
ok('logged-in student shell renders the persistent programme selector from the FSc1/FSc2/MDCAT programme model',all(x in base for x in ['programme-context-strip','student_programme_switch','programme_nav_global','programme_code']))
ok('subject chips no longer expose operational Open language','>Open<' not in base and ' Open</' not in base)
ok('Home hero uses mastery identity and one starting-point action without Full Access badge','YOUR MASTERY' in home and 'Find my starting point' in home and 'Full Access' not in home.split('home-identity-hero',1)[1].split('</header>',1)[0])
ok('Home removes verified-mastery-record production jargon','verified mastery records' not in home.casefold())
ok('landing page includes programme context cinematic intelligence and motivational progression',all(x in landing for x in ['public-programme-strip','landing-v640','Discipline','Consistency','Progress','Mastery','Results']) and 'scoremax_intelligence_hero.png' in css)
ok('account entry wording clearly says Email or ScoreMax ID','Email or ScoreMax ID' in login and 'ScoreMax ID' in register)
ok('admin UI exposes Emergency Direct Intake as an explicit fallback route','Emergency Direct Intake' in intake and 'NOT AN ACADEMIC REVIEW BYPASS' in intake)
ok('teacher referral UI is available from teacher shell and referral page supports student and teacher invite links','Referrals' in base and 'teacher_referral_link' in (ROOT/'templates/referrals.html').read_text())

c.close(); wb.close()
print(f'\nV6.4.0 LIVE PILOT UX & OPERATIONS CHECKS PASSED: {n}')
